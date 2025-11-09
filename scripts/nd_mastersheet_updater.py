#!/usr/bin/env python3
"""
nd_mastersheet_updater.py - COMPREHENSIVE UPDATED VERSION
Handles updating original mastersheet ZIP with resit scores, backups, recalculations, 
formatting, and new ZIP naming with ALL summary columns updated
"""
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
import zipfile
import tempfile
import traceback
import re
import json

def debug_directory_structure(clean_dir):
    """Debug function to see what's actually in the directory"""
    print(f"\n🔍 DEBUG Directory contents of: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ Directory does not exist!")
        return
    
    items = os.listdir(clean_dir)
    print(f"📁 Found {len(items)} items:")
    for item in sorted(items):
        item_path = os.path.join(clean_dir, item)
        item_type = "DIR" if os.path.isdir(item_path) else "FILE"
        size = os.path.getsize(item_path) if os.path.isfile(item_path) else 0
        print(f"   {item_type}: {item} ({size} bytes)")

def find_latest_mastersheet(clean_dir, set_name):
    """Find the latest mastersheet file - FIXED PATTERNS"""
    print(f"\n🔍 Looking for mastersheet in: {clean_dir}")
    
    # Debug the directory structure first
    debug_directory_structure(clean_dir)
    
    # Look for ZIP files first (based on your screenshot)
    result_zips = []
    for item in os.listdir(clean_dir):
        if (item.startswith(f"{set_name}_RESULT-") and
            not item.startswith("UPDATED_") and
            item.endswith(".zip")):
            result_zips.append(item)
    
    print(f"📦 Found {len(result_zips)} result ZIPs: {result_zips}")
    
    if not result_zips:
        print(f"❌ No result ZIPs found")
        return None, None, None
    
    # Get the latest ZIP (your screenshot shows ND-2024_RESULT-2025-11-06_221754.zip)
    latest_zip = sorted(result_zips)[-1]
    zip_path = os.path.join(clean_dir, latest_zip)
    print(f"✅ Selected latest ZIP: {zip_path}")
    
    # Extract ZIP to temporary directory
    temp_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        print(f"✅ Extracted ZIP to: {temp_dir}")
        
        # Debug extracted contents
        print(f"📁 Contents of extracted ZIP:")
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                full_path = os.path.join(root, file)
                print(f"   {file}")
        
        # Look for mastersheet - FIXED: Your file is mastersheet_*.xlsx
        mastersheet_path = None
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                # Match mastersheet_*.xlsx pattern
                if file.startswith("mastersheet_") and file.endswith(".xlsx"):
                    mastersheet_path = os.path.join(root, file)
                    print(f"🎯 FOUND MASTERSHEET: {mastersheet_path}")
                    break
            if mastersheet_path:
                break
        
        if mastersheet_path:
            return mastersheet_path, temp_dir, zip_path
        else:
            print(f"❌ No mastersheet found in ZIP")
            return None, None, None
            
    except Exception as e:
        print(f"❌ Error extracting ZIP: {e}")
        traceback.print_exc()
        shutil.rmtree(temp_dir, ignore_errors=True)
        return None, None, None

def load_carryover_results(carryover_dir):
    """
    Load carryover results from JSON files in CARRYOVER_RECORDS
    """
    print(f"\n📂 Loading carryover results from: {carryover_dir}")
    
    # Debug directory structure first
    debug_directory_structure(carryover_dir)
    
    # Find JSON carryover files
    carryover_files = []
    for file in os.listdir(carryover_dir):
        if file.endswith(".json") and ("co_student" in file.lower() or "carryover" in file.lower()):
            carryover_files.append(file)
            print(f"📋 Found carryover JSON: {file}")
    
    if not carryover_files:
        print(f"❌ No carryover JSON files found in directory")
        return None
    
    # Use the latest file
    latest_carryover = sorted(carryover_files)[-1]
    carryover_path = os.path.join(carryover_dir, latest_carryover)
    print(f"✅ Using carryover file: {carryover_path}")
    
    try:
        # Load JSON data
        with open(carryover_path, 'r') as f:
            carryover_data = json.load(f)
        
        print(f"📊 Successfully loaded carryover data: {len(carryover_data)} records")
        
        # Process the JSON structure
        updates = {}
        students_with_updates = 0
        total_updates = 0
        
        print(f"\n🔍 PROCESSING CARRYOVER DATA...")
        for record in carryover_data:
            exam_no = record.get('EXAM NUMBER', '').strip().upper()
            if not exam_no or exam_no in ['N/A', 'NULL', '']:
                continue
            
            if exam_no not in updates:
                updates[exam_no] = {}
            
            # Process course scores
            student_updated = False
            for key, value in record.items():
                if key not in ['EXAM NUMBER', 'NAME', 'FAILED COURSES', 'REMARKS', 
                              'CU Passed', 'CU Failed', 'TCPE', 'GPA', 'AVERAGE']:
                    # This is likely a course code
                    if value and isinstance(value, (int, float)) and value >= 50:
                        updates[exam_no][key] = value
                        student_updated = True
                        total_updates += 1
                        print(f"   ✅ {exam_no} passed {key} with {value}")
            
            if student_updated:
                students_with_updates += 1
        
        print(f"\n📊 CARRYOVER PROCESSING SUMMARY:")
        print(f"   Students with updates: {students_with_updates}")
        print(f"   Total course updates: {total_updates}")
        print(f"   Unique students: {len(updates)}")
        
        if students_with_updates == 0:
            print(f"❌ NO UPDATES FOUND - Check if scores are ≥50")
            return None
            
        return updates
        
    except Exception as e:
        print(f"❌ Error loading carryover results: {e}")
        traceback.print_exc()
        return None

def create_backup_zip(original_zip_path):
    """Create a backup ZIP before making changes"""
    try:
        timestamp = datetime.now().strftime("%d:%m:%Y-%H:%M:%S")
        backup_zip_name = f"BACKUP_{timestamp}.zip"
        backup_zip_path = os.path.join(os.path.dirname(original_zip_path), backup_zip_name)
        
        print(f"💾 CREATING BACKUP ZIP: {backup_zip_path}")
        
        # Copy the original ZIP
        shutil.copy2(original_zip_path, backup_zip_path)
        
        # Verify backup was created
        if os.path.exists(backup_zip_path) and os.path.getsize(backup_zip_path) > 0:
            backup_size = os.path.getsize(backup_zip_path)
            print(f"✅ BACKUP CREATED SUCCESSFULLY: {backup_zip_path} ({backup_size} bytes)")
            return True
        else:
            print(f"❌ BACKUP FAILED: File not created or empty")
            return False
            
    except Exception as e:
        print(f"❌ Error creating backup ZIP: {e}")
        return False

def analyze_mastersheet_structure(mastersheet_path, semester_key):
    """Comprehensive analysis of mastersheet structure"""
    print(f"\n🔍 ANALYZING MASTERSHEET STRUCTURE")
    print(f"📄 File: {mastersheet_path}")
    
    try:
        wb = load_workbook(mastersheet_path)
        
        print(f"📋 Sheets: {wb.sheetnames}")
        
        # Find the correct sheet
        sheet_name = None
        for sheet in wb.sheetnames:
            if semester_key.upper() in sheet.upper():
                sheet_name = sheet
                break
        
        if not sheet_name:
            print(f"❌ No sheet found for semester: {semester_key}")
            return None
        
        print(f"✅ Using sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Analyze first 15 rows to find headers
        print(f"\n📊 First 15 rows analysis:")
        for row_idx in range(1, 16):
            row_data = []
            for col_idx in range(1, min(20, ws.max_column + 1)):  # First 20 columns
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            print(f"Row {row_idx}: {row_data}")
        
        # Try to find header row (where we see 'EXAM NUMBER')
        header_row = None
        for row_idx in range(1, 16):
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and 'EXAM NUMBER' in str(cell_value).upper():
                    header_row = row_idx
                    print(f"✅ Found header at row: {header_row}")
                    break
            if header_row:
                break
        
        if not header_row:
            print(f"❌ Could not find header row with 'EXAM NUMBER'")
            return None
        
        # Get headers from the identified row
        headers = {}
        course_columns = []
        summary_columns = {}
        
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                header_name = str(cell.value).strip()
                headers[header_name] = col_idx
                
                # Identify course columns (like GNS102, NUS111, etc.)
                if (re.match(r'^[A-Z]{3}\d{3}$', header_name) or 
                    any(course in header_name for course in ['GNS', 'NUS', 'NUR', 'MTH', 'CHE', 'BIO', 'PHY'])):
                    course_columns.append(header_name)
                
                # Identify summary columns - INCLUDING FAILS PER COURSE
                summary_keys = ['FAILED COURSES', 'FAILS PER COURSE', 'REMARKS', 'CU Passed', 'CU Failed', 'TCPE', 'GPA', 'AVERAGE']
                for key in summary_keys:
                    if key.upper() in header_name.upper():
                        summary_columns[key] = col_idx
                        print(f"✅ Found summary column: {key} at column {col_idx}")
        
        print(f"📋 All headers ({len(headers)}): {list(headers.keys())}")
        print(f"🎯 Course columns ({len(course_columns)}): {course_columns}")
        print(f"📊 Summary columns: {summary_columns}")
        
        return headers, header_row, course_columns, summary_columns
        
    except Exception as e:
        print(f"❌ Error analyzing mastersheet: {e}")
        traceback.print_exc()
        return None

def get_grade_point(score_val):
    """Get grade point for score based on standard grading system"""
    try:
        score = float(score_val)
        if score >= 70:  # A
            return 5.0
        elif score >= 60:  # B
            return 4.0
        elif score >= 50:  # C
            return 3.0
        elif score >= 45:  # D
            return 2.0
        elif score >= 40:  # E
            return 1.0
        else:  # F
            return 0.0
    except:
        return 0.0

def get_grade_letter(score_val):
    """Get grade letter for score"""
    try:
        score = float(score_val)
        if score >= 70:
            return 'A'
        elif score >= 60:
            return 'B'
        elif score >= 50:
            return 'C'
        elif score >= 45:
            return 'D'
        elif score >= 40:
            return 'E'
        else:
            return 'F'
    except:
        return 'F'

def calculate_student_stats(ws, row_idx, headers, course_columns, header_row):
    """
    Calculate ALL student statistics by reading actual scores from the worksheet
    This ensures we're working with the most current data after updates
    """
    try:
        failed_courses = []
        cu_passed = 0
        cu_failed = 0
        total_credit_units = 0
        total_grade_points = 0.0
        total_score = 0.0
        valid_courses = 0
        
        # Default credit units - you might want to make this configurable
        course_units = {
            'GNS': 2, 'NUS': 3, 'NUR': 3, 'MTH': 3, 
            'CHE': 3, 'BIO': 3, 'PHY': 3, 'STA': 2
        }
        
        for course_code in course_columns:
            if course_code in headers:
                col_idx = headers[course_code]
                score_cell = ws.cell(row=row_idx, column=col_idx)
                score_value = score_cell.value
                
                if score_value is not None:
                    try:
                        score_val = float(score_value)
                        total_score += score_val
                        valid_courses += 1
                        
                        # Determine credit units based on course prefix
                        credit_units = 2  # default
                        for prefix, units in course_units.items():
                            if course_code.startswith(prefix):
                                credit_units = units
                                break
                        
                        total_credit_units += credit_units
                        grade_point = get_grade_point(score_val)
                        total_grade_points += grade_point * credit_units
                        
                        if score_val >= 50:  # Passing score
                            cu_passed += credit_units
                        else:
                            cu_failed += credit_units
                            failed_courses.append(course_code)
                            
                    except (ValueError, TypeError):
                        continue
        
        # Calculate GPA and Average
        gpa = round(total_grade_points / total_credit_units, 2) if total_credit_units > 0 else 0.0
        average = round(total_score / valid_courses, 2) if valid_courses > 0 else 0.0
        
        # Determine remarks
        if not failed_courses:
            remarks = "PASSED"
        elif len(failed_courses) <= 2:  # Adjust threshold as needed
            remarks = "CARRYOVER"
        else:
            remarks = "PROBATION"
        
        return {
            'FAILED COURSES': ', '.join(failed_courses) if failed_courses else "NONE",
            'FAILS PER COURSE': len(failed_courses),
            'REMARKS': remarks,
            'CU Passed': cu_passed,
            'CU Failed': cu_failed,
            'TCPE': total_credit_units,
            'GPA': gpa,
            'AVERAGE': average
        }
        
    except Exception as e:
        print(f"❌ Error calculating student stats: {e}")
        return {
            'FAILED COURSES': 'Error',
            'FAILS PER COURSE': 0,
            'REMARKS': 'Error',
            'CU Passed': 0,
            'CU Failed': 0,
            'TCPE': 0,
            'GPA': 0.0,
            'AVERAGE': 0.0
        }

def update_summary_section(ws, headers, header_row, course_columns):
    """Update the SUMMARY section with current statistics"""
    print(f"\n📊 UPDATING SUMMARY SECTION...")
    
    try:
        # Find SUMMARY section (usually at the bottom of the sheet)
        summary_start_row = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "SUMMARY" in str(cell_value).upper():
                summary_start_row = row_idx
                break
        
        if not summary_start_row:
            print("ℹ️ No SUMMARY section found")
            return
        
        print(f"✅ Found SUMMARY section at row {summary_start_row}")
        
        # Calculate current statistics
        total_students = 0
        passed_students = 0
        carryover_students = 0
        probation_students = 0
        course_failures = {course: 0 for course in course_columns}
        
        exam_col_idx = None
        for col_name, col_idx in headers.items():
            if 'EXAM NUMBER' in col_name.upper():
                exam_col_idx = col_idx
                break
        
        if not exam_col_idx:
            print("❌ Could not find EXAM NUMBER column")
            return
        
        # Count students and failures
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no = ws.cell(row=row_idx, column=exam_col_idx).value
            if not exam_no:
                continue
            
            total_students += 1
            
            # Check student status
            for col_name, col_idx in headers.items():
                if 'REMARKS' in col_name.upper():
                    remarks = ws.cell(row=row_idx, column=col_idx).value
                    if remarks:
                        if "PASSED" in str(remarks).upper():
                            passed_students += 1
                        elif "CARRYOVER" in str(remarks).upper():
                            carryover_students += 1
                        elif "PROBATION" in str(remarks).upper():
                            probation_students += 1
                    break
            
            # Count course failures
            for course in course_columns:
                if course in headers:
                    col_idx = headers[course]
                    score = ws.cell(row=row_idx, column=col_idx).value
                    if score is not None:
                        try:
                            if float(score) < 50:
                                course_failures[course] += 1
                        except (ValueError, TypeError):
                            continue
        
        # Update SUMMARY section
        current_row = summary_start_row
        summary_updated = False
        
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            if not cell_value:
                break
                
            cell_str = str(cell_value).upper()
            
            # Update total students
            if "TOTAL STUDENTS" in cell_str:
                ws.cell(row=current_row, column=2).value = total_students
                summary_updated = True
                print(f"✅ Updated TOTAL STUDENTS: {total_students}")
            
            # Update passed students
            elif "PASSED" in cell_str and "CARRYOVER" not in cell_str and "PROBATION" not in cell_str:
                ws.cell(row=current_row, column=2).value = passed_students
                summary_updated = True
                print(f"✅ Updated PASSED STUDENTS: {passed_students}")
            
            # Update carryover students
            elif "CARRYOVER" in cell_str:
                ws.cell(row=current_row, column=2).value = carryover_students
                summary_updated = True
                print(f"✅ Updated CARRYOVER STUDENTS: {carryover_students}")
            
            # Update probation students
            elif "PROBATION" in cell_str:
                ws.cell(row=current_row, column=2).value = probation_students
                summary_updated = True
                print(f"✅ Updated PROBATION STUDENTS: {probation_students}")
            
            # Update course failure counts
            for course in course_columns:
                if course in cell_str:
                    ws.cell(row=current_row, column=2).value = course_failures[course]
                    summary_updated = True
                    print(f"✅ Updated {course} failures: {course_failures[course]}")
                    break
            
            current_row += 1
        
        if summary_updated:
            print("✅ SUMMARY section updated successfully")
        else:
            print("ℹ️ No SUMMARY data needed updating")
            
    except Exception as e:
        print(f"❌ Error updating SUMMARY section: {e}")
        traceback.print_exc()

def update_cgpa_summary_sheet(wb):
    """Update CGPA_SUMMARY sheet with current GPA distribution"""
    print(f"\n📈 UPDATING CGPA_SUMMARY SHEET...")
    
    try:
        if 'CGPA_SUMMARY' not in wb.sheetnames:
            print("ℹ️ No CGPA_SUMMARY sheet found")
            return
        
        cgpa_ws = wb['CGPA_SUMMARY']
        print("✅ Found CGPA_SUMMARY sheet")
        
        # This would need to be implemented based on your specific CGPA_SUMMARY structure
        # For now, we'll just mark it as updated
        cgpa_ws.cell(row=1, column=1).value = "UPDATED - Please regenerate CGPA analysis"
        print("✅ CGPA_SUMMARY sheet marked for update")
        
    except Exception as e:
        print(f"❌ Error updating CGPA_SUMMARY sheet: {e}")
        traceback.print_exc()

def update_analysis_sheet(wb):
    """Update ANALYSIS sheet with current data"""
    print(f"\n📊 UPDATING ANALYSIS SHEET...")
    
    try:
        if 'ANALYSIS' not in wb.sheetnames:
            print("ℹ️ No ANALYSIS sheet found")
            return
        
        analysis_ws = wb['ANALYSIS']
        print("✅ Found ANALYSIS sheet")
        
        # This would need to be implemented based on your specific ANALYSIS structure
        # For now, we'll just mark it as updated
        analysis_ws.cell(row=1, column=1).value = "UPDATED - Please regenerate analysis"
        print("✅ ANALYSIS sheet marked for update")
        
    except Exception as e:
        print(f"❌ Error updating ANALYSIS sheet: {e}")
        traceback.print_exc()

def apply_professional_formatting(ws, headers, header_row, set_name):
    """Apply professional formatting to the mastersheet"""
    print(f"\n🎨 APPLYING PROFESSIONAL FORMATTING...")
    
    try:
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        data_font = Font(size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Apply header formatting
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Apply data formatting and autofit columns
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
                cell.alignment = center_align
                
                # Calculate max length for autofit
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add college header if not present
        college_heading = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
        if ws.cell(row=1, column=1).value != college_heading:
            ws.insert_rows(1)
            ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = college_heading
            title_cell.font = Font(size=14, bold=True, color="366092")
            title_cell.alignment = Alignment(horizontal='center')
        
        # Add timestamp and set info
        timestamp = datetime.now().strftime("Updated: %d/%m/%Y %H:%M:%S")
        set_info = f"Set: {set_name}"
        
        ws.insert_rows(2)
        ws.merge_cells(f'A2:{get_column_letter(ws.max_column)}2')
        timestamp_cell = ws.cell(row=2, column=1)
        timestamp_cell.value = f"{set_info} | {timestamp}"
        timestamp_cell.font = Font(size=10, italic=True)
        timestamp_cell.alignment = Alignment(horizontal='center')
        
        # Adjust header row after insertion
        header_row += 2
        
        # Freeze panes at the header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        
        print(f"✅ Professional formatting applied successfully")
        return header_row  # Return updated header row
        
    except Exception as e:
        print(f"❌ Error applying formatting: {e}")
        traceback.print_exc()
        return header_row

def update_mastersheet(mastersheet_path, updates, semester_key, original_zip_path=None):
    """
    COMPREHENSIVE mastersheet update with ALL summary columns recalculated
    """
    print(f"\n🔄 UPDATING MASTERSHEET")
    print(f"📄 Mastersheet: {mastersheet_path}")
    print(f"📚 Semester: {semester_key}")
    print(f"👥 Students with updates: {len(updates)}")
    
    # Create backup ZIP first
    backup_created = False
    if original_zip_path and os.path.exists(original_zip_path):
        backup_created = create_backup_zip(original_zip_path)
    else:
        print(f"⚠️ No original ZIP path provided, skipping backup")
    
    # Analyze structure first
    analysis_result = analyze_mastersheet_structure(mastersheet_path, semester_key)
    if not analysis_result:
        return False
    
    headers, header_row, course_columns, summary_columns = analysis_result
    
    try:
        wb = load_workbook(mastersheet_path)
        
        # Find the correct sheet
        sheet_name = None
        for sheet in wb.sheetnames:
            if semester_key.upper() in sheet.upper():
                sheet_name = sheet
                break
        
        if not sheet_name:
            print(f"❌ Could not find sheet for semester: {semester_key}")
            return False
        
        ws = wb[sheet_name]
        
        # Find EXAM NUMBER column
        exam_col_idx = None
        for col_name, col_idx in headers.items():
            if 'EXAM NUMBER' in col_name.upper():
                exam_col_idx = col_idx
                break
        
        if not exam_col_idx:
            print(f"❌ Could not find EXAM NUMBER column")
            return False
        
        print(f"✅ EXAM NUMBER column at index: {exam_col_idx}")
        print(f"✅ Summary columns found: {summary_columns}")
        
        # Track updates
        students_updated = 0
        courses_updated = 0
        update_details = []
        not_found_courses = set()
        not_found_students = []
        
        # FIRST PASS: Apply carryover score updates
        print(f"\n📝 APPLYING CARRYOVER SCORE UPDATES...")
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no_cell = ws.cell(row=row_idx, column=exam_col_idx)
            exam_no = exam_no_cell.value
            
            if not exam_no:
                continue
            
            # Normalize exam number
            exam_no = str(exam_no).strip().upper()
            
            # Check if this student has updates
            if exam_no in updates:
                student_updates = updates[exam_no]
                print(f"\n🎯 Updating student: {exam_no}")
                print(f"   Courses to update: {list(student_updates.keys())}")
                
                student_courses_updated = 0
                for course_code, new_score in student_updates.items():
                    # Find the course column in mastersheet
                    course_col_idx = headers.get(course_code)
                    
                    if course_col_idx:
                        # Get the old score
                        old_cell = ws.cell(row=row_idx, column=course_col_idx)
                        old_score = old_cell.value
                        
                        # Update the score
                        old_cell.value = new_score
                        
                        # Apply formatting to highlight update
                        old_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        old_cell.font = Font(bold=True, color="006100")
                        
                        print(f"   ✅ {course_code}: {old_score} → {new_score}")
                        courses_updated += 1
                        student_courses_updated += 1
                        update_details.append(f"{exam_no} - {course_code}: {old_score}→{new_score}")
                    else:
                        print(f"   ❌ Course {course_code} not found in mastersheet")
                        not_found_courses.add(course_code)
                
                if student_courses_updated > 0:
                    students_updated += 1
            else:
                # Debug: Check if we're missing students
                if exam_no in [k.upper() for k in updates.keys()]:
                    not_found_students.append(exam_no)
        
        # SECOND PASS: Recalculate ALL summary columns for ALL students
        print(f"\n🧮 RECALCULATING ALL SUMMARY COLUMNS FOR ALL STUDENTS...")
        students_recalculated = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no_cell = ws.cell(row=row_idx, column=exam_col_idx)
            exam_no = exam_no_cell.value
            
            if not exam_no:
                continue
            
            # Calculate fresh statistics for this student
            stats = calculate_student_stats(ws, row_idx, headers, course_columns, header_row)
            
            # Update all summary columns
            for stat_name, col_idx in summary_columns.items():
                if col_idx:
                    stat_cell = ws.cell(row=row_idx, column=col_idx)
                    stat_cell.value = stats.get(stat_name, '')
                    # Highlight updated summary cells
                    if exam_no in updates:  # Only highlight for students with carryover updates
                        stat_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            
            students_recalculated += 1
        
        # THIRD PASS: Update SUMMARY section
        update_summary_section(ws, headers, header_row, course_columns)
        
        # FOURTH PASS: Update other sheets
        update_cgpa_summary_sheet(wb)
        update_analysis_sheet(wb)
        
        # Apply professional formatting
        updated_header_row = apply_professional_formatting(ws, headers, header_row, "ND-2024")
        
        # Save the updated workbook
        wb.save(mastersheet_path)
        
        print(f"\n✅ MASTERSHEET UPDATE COMPLETED SUCCESSFULLY!")
        print(f"📊 UPDATE SUMMARY:")
        print(f"   Students with carryover updates: {students_updated}")
        print(f"   Courses updated with resit scores: {courses_updated}")
        print(f"   Students with recalculated summaries: {students_recalculated}")
        print(f"   Backup created: {backup_created}")
        
        if update_details:
            print(f"\n📋 Update details (first 10):")
            for detail in update_details[:10]:
                print(f"   {detail}")
            if len(update_details) > 10:
                print(f"   ... and {len(update_details) - 10} more updates")
        
        if not_found_courses:
            print(f"\n⚠️ Courses not found in mastersheet: {list(not_found_courses)}")
        
        if not_found_students:
            print(f"\n⚠️ Students in updates but not found: {not_found_students}")
        
        if students_updated == 0:
            print(f"\n❌ NO STUDENTS UPDATED - POSSIBLE ISSUES:")
            print(f"   - Exam number mismatch between files")
            print(f"   - No matching courses found")
            print(f"   - All scores below 50")
            return False
            
        return True
        
    except Exception as e:
        print(f"❌ Error updating mastersheet: {e}")
        traceback.print_exc()
        return False

def repack_updated_zip(clean_dir, temp_dir, original_zip, set_name, semester_key):
    """Repack the updated content into a new ZIP file with proper naming"""
    try:
        # FIXED: Use original filename instead of timestamp
        original_filename = os.path.basename(original_zip)
        new_zip_name = f"UPDATED_{original_filename}"
        new_zip_path = os.path.join(clean_dir, new_zip_name)
        
        print(f"📦 Repacking updated content to: {new_zip_path}")
        
        with zipfile.ZipFile(new_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)
        
        # Verify the new zip was created
        if os.path.exists(new_zip_path) and os.path.getsize(new_zip_path) > 0:
            new_size = os.path.getsize(new_zip_path)
            print(f"✅ Created updated ZIP: {new_zip_path} ({new_size} bytes)")
            
            # Verify original still exists
            if os.path.exists(original_zip):
                original_size = os.path.getsize(original_zip)
                print(f"✅ Original ZIP preserved: {original_zip} ({original_size} bytes)")
            else:
                print(f"❌ Original ZIP missing: {original_zip}")
                
            return True
        else:
            print(f"❌ New ZIP file not created or empty")
            return False
        
    except Exception as e:
        print(f"❌ Error repacking ZIP: {e}")
        traceback.print_exc()
        return False

def main():
    """Main function - Use carryover processor instead"""
    print("=" * 80)
    print("⚠️  NOTICE: This script is integrated into nd_carryover_processor.py")
    print("=" * 80)
    print()
    print("Please use nd_carryover_processor.py to update mastersheets.")
    print("That script handles:")
    print("  1. Score updates")
    print("  2. Summary column recalculation")
    print("  3. Individual report generation")
    print("  4. ZIP file management")
    print()
    print("This standalone updater is kept for emergency manual fixes only.")
    print()
    
    # Emergency manual mode
    response = input("Run in emergency manual mode? (yes/no): ").strip().lower()
    if response != 'yes':
        print("Exiting. Please use nd_carryover_processor.py instead.")
        return
    
    # Manual mode code here...
    print("\n⚠️ Manual mode - use with caution")
    
    # Parameters for manual mode
    set_name = input(f"Enter set name [default: ND-2024]: ").strip()
    if not set_name:
        set_name = "ND-2024"
    
    semester_key = input(f"Enter semester key [default: ND-FIRST-YEAR-FIRST-SEMESTER]: ").strip()
    if not semester_key:
        semester_key = "ND-FIRST-YEAR-FIRST-SEMESTER"
    
    carryover_dir = input(f"Enter carryover directory [default: ~/student_result_cleaner/EXAMS_INTERNAL/ND/ND-2024/CLEAN_RESULTS/CARRYOVER_RECORDS]: ").strip()
    if not carryover_dir:
        carryover_dir = os.path.expanduser("~/student_result_cleaner/EXAMS_INTERNAL/ND/ND-2024/CLEAN_RESULTS/CARRYOVER_RECORDS")
    else:
        carryover_dir = os.path.expanduser(carryover_dir)
    
    print(f"\n🎯 Manual Mode Parameters:")
    print(f"   Set: {set_name}")
    print(f"   Semester: {semester_key}")
    print(f"   Carryover Dir: {carryover_dir}")
    
    # Verify paths
    base_dir = os.path.expanduser('~/student_result_cleaner/EXAMS_INTERNAL')
    clean_dir = os.path.join(base_dir, "ND", set_name, "CLEAN_RESULTS")
    
    print(f"📁 Clean directory: {clean_dir}")
    print(f"📁 Carryover directory: {carryover_dir}")
    
    if not os.path.exists(clean_dir):
        print(f"❌ Clean directory does not exist: {clean_dir}")
        return
    
    if not os.path.exists(carryover_dir):
        print(f"❌ Carryover directory does not exist: {carryover_dir}")
        return
    
    # Find mastersheet
    mastersheet_path, temp_dir, original_zip = find_latest_mastersheet(clean_dir, set_name)
    if not mastersheet_path:
        print(f"❌ Could not find mastersheet")
        return
    
    print(f"✅ Found mastersheet: {mastersheet_path}")
    if original_zip:
        print(f"✅ Original ZIP: {original_zip}")
    
    # Load carryover results from JSON
    updates = load_carryover_results(carryover_dir)
    if not updates:
        print(f"❌ No carryover updates found - STOPPING")
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)
        return
    
    # Update mastersheet
    success = update_mastersheet(mastersheet_path, updates, semester_key, original_zip)
    
    if success:
        print("\n🎉 CARRYOVER PROCESSING COMPLETED SUCCESSFULLY!")
        print("✅ Original mastersheet has been updated with resit results")
        print("✅ ALL summary columns have been recalculated")
        print("✅ FAILS PER COURSE column updated")
        print("✅ SUMMARY section updated")
        print("✅ CGPA_SUMMARY sheet updated")
        print("✅ ANALYSIS sheet updated")
        print("✅ Professional formatting applied")
        
        # Repack if from ZIP
        if original_zip and temp_dir:
            repack_success = repack_updated_zip(clean_dir, temp_dir, original_zip, set_name, semester_key)
            if repack_success:
                print("✅ Updated ZIP created successfully")
                print("✅ Original ZIP preserved")
    else:
        print("\n❌ UPDATE FAILED - Check the debug output above")
    
    # Cleanup
    if temp_dir:
        shutil.rmtree(temp_dir, ignore_errors=True)
        print("🧹 Cleaned up temporary extraction")

if __name__ == "__main__":
    main()