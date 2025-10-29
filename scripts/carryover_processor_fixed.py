#!/usr/bin/env python3
"""
carryover_processor_fixed.py - UPDATED WORKING VERSION
Fixed CGPA calculation and enhanced presentation with course titles, credit units, and proper formatting.
"""

import os
import sys
import re
import pandas as pd
from datetime import datetime
import glob
import json
import traceback
import shutil
import zipfile
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------
# Configuration
# ----------------------------

def get_base_directory():
    """Get base directory."""
    if os.getenv('BASE_DIR'):
        base_dir = os.getenv('BASE_DIR')
        if os.path.exists(base_dir):
            return base_dir
    
    return os.path.join(os.path.expanduser('~'), 'student_result_cleaner')

BASE_DIR = get_base_directory()
TIMESTAMP_FMT = "%d-%m-%Y_%H%M%S"
DEFAULT_PASS_THRESHOLD = 50.0
DEFAULT_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")

def sanitize_filename(filename):
    """Remove or replace characters that are not safe for filenames."""
    return re.sub(r'[^\w\-_.]', '_', filename)

def find_exam_number_column(df):
    """Find the exam number column in a DataFrame."""
    possible_names = ['EXAM NUMBER', 'REG. No', 'REG NO', 'REGISTRATION NUMBER', 'MAT NO', 'STUDENT ID']
    for col in df.columns:
        col_upper = str(col).upper()
        for possible_name in possible_names:
            if possible_name in col_upper:
                return col
    return None

def load_course_data():
    """Load course data from course-code-creditUnit.xlsx with robust matching."""
    course_file = os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", "ND-COURSES", "course-code-creditUnit.xlsx")
    print(f"📚 Loading course data from: {course_file}")
    
    if not os.path.exists(course_file):
        print(f"❌ Course file not found: {course_file}")
        return {}, {}, {}
    
    try:
        xl = pd.ExcelFile(course_file)
        semester_course_titles = {}
        semester_credit_units = {}
        course_code_to_title = {}
        
        for sheet in xl.sheet_names:
            print(f"📖 Reading sheet: {sheet}")
            df = pd.read_excel(course_file, sheet_name=sheet, engine='openpyxl', header=0)
            df.columns = [str(c).strip() for c in df.columns]
            
            expected = ['COURSE CODE', 'COURSE TITLE', 'CU']
            if not all(col in df.columns for col in expected):
                print(f"⚠️ Sheet '{sheet}' missing expected columns {expected} — skipped")
                print(f"   Available columns: {list(df.columns)}")
                continue
            
            # More robust cleaning
            dfx = df.dropna(subset=['COURSE CODE', 'COURSE TITLE'])
            dfx = dfx[~dfx['COURSE CODE'].astype(str).str.contains('TOTAL', case=False, na=False)]
            
            valid_mask = dfx['CU'].astype(str).str.replace('.', '', regex=False).str.isdigit()
            dfx = dfx[valid_mask]
            
            if dfx.empty:
                print(f"⚠️ Sheet '{sheet}' has no valid rows after cleaning — skipped")
                continue
            
            codes = dfx['COURSE CODE'].astype(str).str.strip().tolist()
            titles = dfx['COURSE TITLE'].astype(str).str.strip().tolist()
            cus = dfx['CU'].astype(float).tolist()

            print(f"📋 Found {len(codes)} courses in {sheet}:")
            for code, title in zip(codes[:5], titles[:5]):  # Show first 5 for debugging
                print(f"   - {code}: {title}")

            # Create multiple normalization variants for robust matching
            normalized_codes_variants = []
            for code in codes:
                # Variant 1: Remove all spaces and uppercase
                variant1 = re.sub(r'\s+', '', code.upper())
                # Variant 2: Remove spaces but keep original case
                variant2 = re.sub(r'\s+', '', code)
                # Variant 3: Just uppercase
                variant3 = code.upper()
                # Variant 4: Just strip spaces
                variant4 = code.strip()
                
                normalized_codes_variants.extend([variant1, variant2, variant3, variant4])
            
            # Create mapping dictionaries with all variants
            sheet_titles = {}
            sheet_units = {}
            
            for code, title, cu in zip(codes, titles, cus):
                # Create all normalization variants for this code
                variants = [
                    re.sub(r'\s+', '', code.upper()),  # No spaces, uppercase
                    re.sub(r'\s+', '', code),           # No spaces, original case  
                    code.upper(),                       # Uppercase, with spaces
                    code.strip(),                       # Just strip spaces
                    code.strip().upper(),               # Strip and uppercase
                    code.strip().replace(' ', ''),      # Strip and remove spaces
                ]
                
                # Add all variants to the mapping
                for variant in variants:
                    sheet_titles[variant] = title
                    sheet_units[variant] = cu
                    course_code_to_title[variant] = title
            
            semester_course_titles[sheet] = sheet_titles
            semester_credit_units[sheet] = sheet_units
        
        print(f"✅ Loaded course data for sheets: {list(semester_course_titles.keys())}")
        print(f"📊 Total course mappings: {len(course_code_to_title)}")
        
        # Debug: Show some course mappings
        print("🔍 Sample course mappings:")
        for i, (code, title) in enumerate(list(course_code_to_title.items())[:10]):
            print(f"   {code} -> {title}")
            
        return semester_course_titles, semester_credit_units, course_code_to_title
        
    except Exception as e:
        print(f"❌ Error loading course data: {e}")
        traceback.print_exc()
        return {}, {}, {}

def find_course_title(course_code, course_titles_dict, course_code_to_title):
    """Robust function to find course title with multiple matching strategies."""
    if not course_code or str(course_code).upper() in ['NAN', 'NONE', '']:
        return course_code
    
    original_code = str(course_code).strip()
    
    # Try multiple matching strategies
    matching_strategies = [
        # Strategy 1: Exact match with various normalizations
        re.sub(r'\s+', '', original_code.upper()),
        re.sub(r'\s+', '', original_code),
        original_code.upper(),
        original_code.strip(),
        original_code.strip().upper(),
        original_code.strip().replace(' ', ''),
        
        # Strategy 2: Remove special characters and normalize
        re.sub(r'[^a-zA-Z0-9]', '', original_code.upper()),
        re.sub(r'[^a-zA-Z0-9]', '', original_code),
        
        # Strategy 3: Try partial matches (for codes like "EEd216" vs "EED216")
        original_code.upper().replace(' ', ''),
        original_code.replace(' ', ''),
    ]
    
    # Remove duplicates
    matching_strategies = list(dict.fromkeys(matching_strategies))
    
    # Try each strategy
    for strategy in matching_strategies:
        # Try course_titles_dict first (semester-specific)
        if strategy in course_titles_dict:
            print(f"✅ Found course title for '{original_code}' -> '{strategy}': {course_titles_dict[strategy]}")
            return course_titles_dict[strategy]
        
        # Try global course_code_to_title
        if strategy in course_code_to_title:
            print(f"✅ Found course title for '{original_code}' -> '{strategy}': {course_code_to_title[strategy]}")
            return course_code_to_title[strategy]
    
    # If no match found, log and return original code
    print(f"⚠️ Could not find course title for: '{original_code}'")
    print(f"   Tried strategies: {matching_strategies}")
    return original_code

def extract_mastersheet_from_zip(zip_path, semester_key):
    """Extract mastersheet from ZIP file and return temporary file path."""
    try:
        print(f"📦 Looking for mastersheet in ZIP: {zip_path}")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # List all files in ZIP for debugging
            all_files = zip_ref.namelist()
            print(f"📁 Files in ZIP: {all_files}")
            
            # Look for mastersheet files in the ZIP
            mastersheet_files = [f for f in all_files if 'mastersheet' in f.lower() and f.endswith('.xlsx')]
            
            if not mastersheet_files:
                print(f"❌ No mastersheet found in ZIP")
                return None, None
            
            # Use the first mastersheet found
            mastersheet_name = mastersheet_files[0]
            print(f"✅ Found mastersheet: {mastersheet_name}")
            
            # Extract to temporary file
            temp_dir = tempfile.mkdtemp()
            temp_mastersheet_path = os.path.join(temp_dir, f"mastersheet_{semester_key}.xlsx")
            
            with open(temp_mastersheet_path, 'wb') as f:
                f.write(zip_ref.read(mastersheet_name))
            
            print(f"✅ Extracted mastersheet to: {temp_mastersheet_path}")
            return temp_mastersheet_path, temp_dir
            
    except Exception as e:
        print(f"❌ Error extracting mastersheet from ZIP: {e}")
        traceback.print_exc()
        return None, None

def find_latest_zip_file(clean_dir):
    """Find the latest ZIP file in clean results directory."""
    print(f"🔍 Looking for ZIP files in: {clean_dir}")
    
    if not os.path.exists(clean_dir):
        print(f"❌ Clean directory doesn't exist: {clean_dir}")
        return None
    
    # List all files in directory
    all_files = os.listdir(clean_dir)
    print(f"📁 Files in clean directory: {all_files}")
    
    # Look for ZIP files
    zip_files = [f for f in all_files if f.lower().endswith('.zip')]
    
    if not zip_files:
        print(f"❌ No ZIP files found")
        return None
    
    print(f"✅ Found ZIP files: {zip_files}")
    
    # Sort by modification time and return the latest
    zip_files_with_path = [os.path.join(clean_dir, f) for f in zip_files]
    latest_zip = sorted(zip_files_with_path, key=os.path.getmtime, reverse=True)[0]
    
    print(f"🎯 Using latest ZIP: {latest_zip}")
    return latest_zip

def get_semester_display_info(semester_key):
    """Get display information for a semester key."""
    semester_lower = semester_key.lower()
    if 'first-year-first-semester' in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI", "Semester 1"
    elif 'first-year-second-semester' in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "NDI", "Semester 2"
    elif 'second-year-first-semester' in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "NDII", "Semester 3"
    elif 'second-year-second-semester' in semester_lower:
        return 2, 2, "YEAR TWO", "SECOND SEMESTER", "NDII", "Semester 4"
    else:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI", "Semester 1"

def get_grade_point(score):
    """Determine grade point based on score - NIGERIAN 5.0 SCALE."""
    try:
        score = float(score)
        if score >= 70: return 5.0
        elif score >= 60: return 4.0
        elif score >= 50: return 3.0
        elif score >= 45: return 2.0
        elif score >= 40: return 1.0
        else: return 0.0
    except (ValueError, TypeError):
        return 0.0

def load_previous_gpas(mastersheet_path, current_semester_key):
    """Load previous GPA data from mastersheet for CGPA calculation."""
    all_student_data = {}
    current_year, current_semester_num, _, _, _, current_semester_name = get_semester_display_info(current_semester_key)

    # Define all previous semesters based on current semester
    if current_semester_num == 1 and current_year == 1:
        semesters_to_load = []
    elif current_semester_num == 2 and current_year == 1:
        semesters_to_load = ["ND-FIRST-YEAR-FIRST-SEMESTER"]
    elif current_semester_num == 1 and current_year == 2:
        semesters_to_load = ["ND-FIRST-YEAR-FIRST-SEMESTER", "ND-FIRST-YEAR-SECOND-SEMESTER"]
    elif current_semester_num == 2 and current_year == 2:
        semesters_to_load = ["ND-FIRST-YEAR-FIRST-SEMESTER", "ND-FIRST-YEAR-SECOND-SEMESTER", 
                            "ND-SECOND-YEAR-FIRST-SEMESTER"]

    if not os.path.exists(mastersheet_path):
        print(f"❌ Mastersheet not found: {mastersheet_path}")
        return {}

    for semester in semesters_to_load:
        try:
            df = pd.read_excel(mastersheet_path, sheet_name=semester, header=5)
            exam_col = find_exam_number_column(df)
            gpa_col = None
            credit_col = None
            
            for col in df.columns:
                col_str = str(col).upper()
                if 'GPA' in col_str and 'CGPA' not in col_str:
                    gpa_col = col
                if 'CU PASSED' in col_str:
                    credit_col = col
            
            if exam_col and gpa_col:
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    gpa = row[gpa_col]
                    credits = int(row[credit_col]) if credit_col and pd.notna(row[credit_col]) else 30
                    
                    if pd.notna(gpa) and pd.notna(exam_no) and exam_no != 'nan' and str(exam_no) != 'nan':
                        try:
                            if exam_no not in all_student_data:
                                all_student_data[exam_no] = {'gpas': [], 'credits': [], 'semesters': []}
                            all_student_data[exam_no]['gpas'].append(float(gpa))
                            all_student_data[exam_no]['credits'].append(credits)
                            all_student_data[exam_no]['semesters'].append(semester)
                            print(f"📊 Loaded GPA for {exam_no}: {gpa} with {credits} credits from {semester}")
                        except (ValueError, TypeError) as e:
                            print(f"⚠️ Error processing GPA for {exam_no}: {e}")
                            continue
        except Exception as e:
            print(f"⚠️ Could not load data from {semester}: {e}")
    
    print(f"📊 Loaded cumulative data for {len(all_student_data)} students")
    return all_student_data

def calculate_cgpa(student_data, current_gpa, current_credits):
    """Calculate Cumulative GPA - FIXED VERSION."""
    if not student_data or not student_data.get('gpas'):
        print(f"⚠️ No previous GPA data, using current GPA: {current_gpa}")
        return current_gpa

    total_grade_points = 0.0
    total_credits = 0

    print(f"🔢 Calculating CGPA from {len(student_data['gpas'])} previous semesters")
    
    for prev_gpa, prev_credits in zip(student_data['gpas'], student_data['credits']):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
        print(f"   - GPA: {prev_gpa}, Credits: {prev_credits}, Running Total: {total_grade_points}/{total_credits}")

    # Add current semester
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits

    print(f"📊 Final calculation: {total_grade_points} / {total_credits}")

    if total_credits > 0:
        cgpa = round(total_grade_points / total_credits, 2)
        print(f"✅ Calculated CGPA: {cgpa}")
        return cgpa
    else:
        print(f"⚠️ No credits, returning current GPA: {current_gpa}")
        return current_gpa

def process_carryover_results(resit_file_path, zip_file_path, semester_key, set_name, pass_threshold, output_dir):
    """
    Process carryover results and generate CARRYOVER_mastersheet.
    """
    print(f"\n🔄 PROCESSING CARRYOVER RESULTS FOR {semester_key}")
    print("=" * 60)
    
    # Load course data
    semester_course_titles, semester_credit_units, course_code_to_title = load_course_data()
    
    # Create output directory
    timestamp = datetime.now().strftime(TIMESTAMP_FMT)
    carryover_output_dir = os.path.join(output_dir, f"CARRYOVER_{set_name}_{semester_key}_{timestamp}")
    os.makedirs(carryover_output_dir, exist_ok=True)
    
    if not os.path.exists(resit_file_path):
        print(f"❌ Resit file not found: {resit_file_path}")
        return False
    
    if not os.path.exists(zip_file_path):
        print(f"❌ ZIP file not found: {zip_file_path}")
        return False
    
    temp_mastersheet_path = None
    temp_dir = None
    
    try:
        # Extract mastersheet from ZIP
        print(f"📦 Extracting mastersheet from ZIP: {zip_file_path}")
        temp_mastersheet_path, temp_dir = extract_mastersheet_from_zip(zip_file_path, semester_key)
        
        if not temp_mastersheet_path:
            print("❌ Failed to extract mastersheet from ZIP")
            return False
        
        # Read files
        print("📖 Reading files...")
        resit_df = pd.read_excel(resit_file_path, header=0)
        mastersheet_df = pd.read_excel(temp_mastersheet_path, sheet_name=semester_key, header=5)
        
        print(f"✅ Files loaded - Resit: {len(resit_df)} rows, Mastersheet: {len(mastersheet_df)} students")
        
        # Find exam number columns
        resit_exam_col = find_exam_number_column(resit_df)
        mastersheet_exam_col = find_exam_number_column(mastersheet_df) or 'EXAM NUMBER'
        
        if not resit_exam_col:
            print("❌ Cannot find exam number column in resit file")
            return False
        
        print(f"📝 Exam columns - Resit: '{resit_exam_col}', Mastersheet: '{mastersheet_exam_col}'")
        
        # Load previous GPAs for CGPA calculation
        cgpa_data = load_previous_gpas(temp_mastersheet_path, semester_key)
        
        # Create carryover mastersheet data structure
        carryover_data = []
        updated_students = set()
        
        print(f"\n🎯 PROCESSING RESIT SCORES...")
        
        for idx, resit_row in resit_df.iterrows():
            exam_no = str(resit_row[resit_exam_col]).strip().upper()
            if not exam_no or exam_no in ['NAN', 'NONE', '']:
                continue
            
            # Find student in mastersheet
            student_mask = mastersheet_df[mastersheet_exam_col].astype(str).str.strip().str.upper() == exam_no
            if not student_mask.any():
                print(f"⚠️ Student {exam_no} not found in mastersheet - skipping")
                continue
            
            student_data = mastersheet_df[student_mask].iloc[0]
            student_name = student_data.get('NAME', 'Unknown')
            
            # Get current credits passed for CGPA calculation
            current_credits = 0
            for col in mastersheet_df.columns:
                if 'CU PASSED' in str(col).upper():
                    current_credits = student_data.get(col, 0)
                    break
            
            # Initialize student record for carryover mastersheet
            student_record = {
                'EXAM NUMBER': exam_no,
                'NAME': student_name,
                'RESIT_COURSES': {},
                'CURRENT_GPA': student_data.get('GPA', 0),
                'CURRENT_CREDITS': current_credits
            }
            
            # Calculate CGPA properly
            if exam_no in cgpa_data:
                student_record['CURRENT_CGPA'] = calculate_cgpa(
                    cgpa_data[exam_no], 
                    student_record['CURRENT_GPA'], 
                    current_credits
                )
            else:
                student_record['CURRENT_CGPA'] = student_record['CURRENT_GPA']
            
            # Add previous GPAs (Semester 1, 2, 3 for Semester 4)
            if exam_no in cgpa_data:
                student_gpa_data = cgpa_data[exam_no]
                for i, prev_semester in enumerate(student_gpa_data['semesters']):
                    sem_num = get_semester_display_info(prev_semester)[5]  # Get "Semester X"
                    student_record[f'GPA_{sem_num}'] = student_gpa_data['gpas'][i]
            
            # Process each course in resit file
            for col in resit_df.columns:
                if col == resit_exam_col or col == 'NAME' or 'Unnamed' in str(col):
                    continue
                    
                resit_score = resit_row.get(col)
                if pd.isna(resit_score) or resit_score == '':
                    continue
                
                try:
                    resit_score_val = float(resit_score)
                except (ValueError, TypeError):
                    continue
                
                # Check if this course column exists in mastersheet
                if col in mastersheet_df.columns:
                    original_score = student_data.get(col)
                    if pd.isna(original_score):
                        continue
                    
                    try:
                        original_score_val = float(original_score) if not pd.isna(original_score) else 0.0
                    except (ValueError, TypeError):
                        original_score_val = 0.0
                    
                    # Only include courses that were re-sat (failed originally and now resat)
                    if original_score_val < pass_threshold:
                        # Get course title using robust matching
                        course_title = find_course_title(col, semester_course_titles.get(semester_key, {}), course_code_to_title)
                        # Get credit unit
                        credit_unit = semester_credit_units.get(semester_key, {}).get(col, 0)
                        
                        student_record['RESIT_COURSES'][col] = {
                            'original_score': original_score_val,
                            'resit_score': resit_score_val,
                            'updated': resit_score_val >= pass_threshold,
                            'course_title': course_title,
                            'credit_unit': credit_unit
                        }
            
            # Only add student to carryover mastersheet if they have resit courses
            if student_record['RESIT_COURSES']:
                carryover_data.append(student_record)
                updated_students.add(exam_no)
                print(f"✅ {exam_no}: {len(student_record['RESIT_COURSES'])} resit courses, CGPA: {student_record['CURRENT_CGPA']}")
        
        # Generate CARRYOVER_mastersheet
        if carryover_data:
            print(f"\n📊 GENERATING CARRYOVER MASTERSHEET...")
            carryover_mastersheet_path = generate_carryover_mastersheet(
                carryover_data, carryover_output_dir, semester_key, set_name, timestamp, 
                cgpa_data, semester_course_titles.get(semester_key, {}), semester_credit_units.get(semester_key, {})
            )
            
            # Generate individual student reports
            print(f"\n📄 GENERATING INDIVIDUAL STUDENT REPORTS...")
            generate_individual_reports(
                carryover_data, carryover_output_dir, semester_key, set_name, timestamp, cgpa_data
            )
            
            # Create final ZIP
            zip_path = os.path.join(output_dir, f"CARRYOVER_{set_name}_{semester_key}_{timestamp}.zip")
            if create_carryover_zip(carryover_output_dir, zip_path):
                print(f"✅ Final carryover ZIP created: {zip_path}")
            
            print(f"\n🎉 CARRYOVER PROCESSING COMPLETED!")
            print(f"📁 Output directory: {carryover_output_dir}")
            print(f"📦 ZIP file: {zip_path}")
            print(f"👨‍🎓 Students processed: {len(carryover_data)}")
            
            return True
        else:
            print("❌ No carryover data found to process")
            return False
            
    except Exception as e:
        print(f"❌ Error processing carryover results: {e}")
        traceback.print_exc()
        return False
    finally:
        # Clean up temporary files
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print("✅ Cleaned up temporary files")

def generate_carryover_mastersheet(carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data, course_titles, course_units):
    """Generate the CARRYOVER_mastersheet with enhanced formatting, course titles, and credit units."""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CARRYOVER_RESULTS"
    
    # Add logo if available
    if os.path.exists(DEFAULT_LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image
            img = Image(DEFAULT_LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"⚠️ Could not add logo: {e}")
    
    # Title and headers - UPDATED with RESIT and 2025/2026
    current_year = 2025  # Hardcoded as requested
    next_year = 2026
    year, sem_num, level, sem_display, set_code, sem_name = get_semester_display_info(semester_key)
    
    # Calculate total columns needed for merging
    all_courses = set()
    for student in carryover_data:
        all_courses.update(student['RESIT_COURSES'].keys())
    
    # Build headers structure to calculate total columns
    headers = ['S/N', 'EXAM NUMBER', 'NAME']
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        headers.extend(['GPA Semester 1', 'GPA Semester 2', 'GPA Semester 3'])
    
    course_headers = []
    for course in sorted(all_courses):
        course_headers.extend([f'{course}', f'{course}_RESIT'])
    
    headers.extend(course_headers)
    headers.extend(['GPA Semester 4', 'CGPA', 'REMARKS'])
    
    total_columns = len(headers)
    last_column = get_column_letter(total_columns)
    
    # CENTRALIZED TITLE ROWS - merged across all columns
    ws.merge_cells(f'A3:{last_column}3')  # Merge across all columns
    title_cell = ws['A3']
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(f'A4:{last_column}4')  # Merge across all columns
    subtitle_cell = ws['A4']
    subtitle_cell.value = f"RESIT - {current_year}/{next_year} SESSION NATIONAL DIPLOMA {level} {sem_display} EXAMINATIONS RESULT — October 28, 2025"
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"🔍 Courses found in resit data: {sorted(all_courses)}")
    
    # Build headers structure with course titles, codes, and credit units
    headers = ['S/N', 'EXAM NUMBER', 'NAME']
    
    # Add previous GPA columns (Semester 1, 2, 3 for Semester 4)
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        headers.extend(['GPA Semester 1', 'GPA Semester 2', 'GPA Semester 3'])
    
    # Add course columns with titles, codes, and credit units
    course_headers = []
    course_title_mapping = {}  # Store the actual titles we find
    course_unit_mapping = {}   # Store the credit units
    
    for course in sorted(all_courses):
        # Use robust title lookup
        course_title = find_course_title(course, course_titles, {})
        course_title_mapping[course] = course_title
        
        # Get credit unit
        credit_unit = course_units.get(course, 0)
        course_unit_mapping[course] = credit_unit
        
        # Truncate long course titles for display
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        course_headers.extend([f'{course}', f'{course}_RESIT'])
    
    headers.extend(course_headers)
    headers.extend(['GPA Semester 4', 'CGPA', 'REMARKS'])
    
    # Write course titles row (row 5) with counterclockwise orientation
    title_row = [''] * 3  # S/N, EXAM NUMBER, NAME
    
    # Add previous GPA placeholders
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        title_row.extend(['', '', ''])  # GPA Semester 1, 2, 3
    
    # Add course titles with counterclockwise orientation
    for course in sorted(all_courses):
        course_title = course_title_mapping[course]
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        title_row.extend([course_title, course_title])  # Use title for both original and resit columns
    
    title_row.extend(['', '', ''])  # GPA Semester 4, CGPA, REMARKS
    
    ws.append(title_row)  # This is row 5
    
    # Write credit units row (row 6)
    credit_row = [''] * 3  # S/N, EXAM NUMBER, NAME
    
    # Add previous GPA placeholders
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        credit_row.extend(['', '', ''])  # GPA Semester 1, 2, 3
    
    # Add credit units for each course
    for course in sorted(all_courses):
        credit_unit = course_unit_mapping[course]
        credit_row.extend([f'CU: {credit_unit}', f'CU: {credit_unit}'])  # Credit unit for both original and resit columns
    
    credit_row.extend(['', '', ''])  # GPA Semester 4, CGPA, REMARKS
    
    ws.append(credit_row)  # This is row 6
    
    # Write course codes row (row 7)
    code_row = ['S/N', 'EXAM NUMBER', 'NAME']
    
    # Add previous GPA headers
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        code_row.extend(['GPA Semester 1', 'GPA Semester 2', 'GPA Semester 3'])
    
    # Add course codes
    for course in sorted(all_courses):
        code_row.extend([f'{course}', f'{course}_RESIT'])
    
    code_row.extend(['GPA Semester 4', 'CGPA', 'REMARKS'])
    
    ws.append(code_row)  # This is row 7
    
    # Define print-friendly colors for course title columns (light pastel colors)
    course_colors = [
        "E6F3FF",  # Light blue
        "FFF0E6",  # Light orange
        "E6FFE6",  # Light green
        "FFF6E6",  # Light peach
        "F0E6FF",  # Light purple
        "E6FFFF",  # Light cyan
        "FFE6F2",  # Light pink
        "F5F5DC",  # Light beige
        "E6F7FF",  # Light sky blue
        "FFF5E6",  # Light apricot
    ]
    
    # Apply colors to course columns in all header rows (5, 6, 7)
    start_col = 4  # Start after S/N, EXAM NUMBER, NAME
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        start_col += 3  # Skip GPA columns
    
    color_index = 0
    for course in sorted(all_courses):
        # Apply colors to all three header rows for this course pair
        for row in [5, 6, 7]:  # CHANGED: Now rows 5, 6, 7
            for offset in [0, 1]:  # Original and resit columns
                cell = ws.cell(row=row, column=start_col + offset)
                cell.fill = PatternFill(start_color=course_colors[color_index % len(course_colors)], 
                                      end_color=course_colors[color_index % len(course_colors)], 
                                      fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        # Apply rotation only to course titles row (row 5)
        for offset in [0, 1]:
            cell = ws.cell(row=5, column=start_col + offset)  # CHANGED: Row 5 for course titles
            cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=9)
        
        color_index += 1
        start_col += 2  # Move to next course pair
    
    # Style the non-course header columns (S/N, EXAM NUMBER, NAME, GPA columns)
    for row in [5, 6, 7]:  # CHANGED: Rows 5, 6, 7
        for col in range(1, 4):  # S/N, EXAM NUMBER, NAME
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Style GPA columns if they exist
        if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
            for col in range(4, 7):  # GPA Semester 1, 2, 3
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        # Style final GPA columns
        for col in range(len(headers)-2, len(headers)+1):  # GPA Semester 4, CGPA, REMARKS
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Write data starting from row 8 (after headers)
    row_idx = 8  # CHANGED: Data starts at row 8 now
    failed_counts = {course: 0 for course in all_courses}
    
    # Apply colors to data rows for course columns
    start_col = 4  # Reset start column
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        start_col += 3  # Skip GPA columns
    
    for student in carryover_data:
        exam_no = student['EXAM NUMBER']
        
        # Basic info
        ws.cell(row=row_idx, column=1, value=row_idx-7)  # S/N (adjusted for new row)
        ws.cell(row=row_idx, column=2, value=student['EXAM NUMBER'])
        ws.cell(row=row_idx, column=3, value=student['NAME'])
        
        # Previous GPAs
        gpa_col = 4
        if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
            for sem_num in ['Semester 1', 'Semester 2', 'Semester 3']:
                gpa_value = student.get(f'GPA_{sem_num}', '')
                ws.cell(row=row_idx, column=gpa_col, value=gpa_value)
                gpa_col += 1
        
        # Course scores - APPLY COLORS TO DATA ROWS
        course_col = gpa_col
        color_index = 0
        for course in sorted(all_courses):
            # Apply the same alternating colors to data cells
            for offset in [0, 1]:
                cell = ws.cell(row=row_idx, column=course_col + offset)
                cell.fill = PatternFill(start_color=course_colors[color_index % len(course_colors)], 
                                      end_color=course_colors[color_index % len(course_colors)], 
                                      fill_type="solid")
            
            if course in student['RESIT_COURSES']:
                course_data = student['RESIT_COURSES'][course]
                
                # Original score (color red if failed)
                orig_cell = ws.cell(row=row_idx, column=course_col, value=course_data['original_score'])
                if course_data['original_score'] < DEFAULT_PASS_THRESHOLD:
                    orig_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                # Resit score (color green if passed, red if failed)
                resit_cell = ws.cell(row=row_idx, column=course_col+1, value=course_data['resit_score'])
                if course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD:
                    resit_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    resit_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    failed_counts[course] += 1
            else:
                ws.cell(row=row_idx, column=course_col, value='')
                ws.cell(row=row_idx, column=course_col+1, value='')
            
            color_index += 1
            course_col += 2
        
        # Current GPA and CGPA
        ws.cell(row=row_idx, column=course_col, value=student['CURRENT_GPA'])
        ws.cell(row=row_idx, column=course_col+1, value=student['CURRENT_CGPA'])
        
        # Remarks
        remarks = generate_remarks(student['RESIT_COURSES'])
        ws.cell(row=row_idx, column=course_col+2, value=remarks)
        
        row_idx += 1
    
    # Add failed count summary - MOVED TO THE EMPTY ROW IMMEDIATELY AFTER DATA
    failed_row_idx = row_idx  # Use the current row (empty row after data)
    ws.cell(row=failed_row_idx, column=1, value="FAILED COUNT BY COURSE:").font = Font(bold=True)
    
    # Apply color to failed count row - LIGHT YELLOW BACKGROUND
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=failed_row_idx, column=col)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Write failed counts under each course's RESIT column
    course_col = gpa_col  # Start at first course column
    for course in sorted(all_courses):
        # Write the failed count in the RESIT column (course_col + 1)
        count_cell = ws.cell(row=failed_row_idx, column=course_col+1, value=failed_counts[course])
        count_cell.font = Font(bold=True)
        count_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        course_col += 2
    
    # Add main summary section - LEFT ALIGNED
    summary_start_row = failed_row_idx + 2  # One empty row after failed count
    
    # Calculate summary statistics
    total_students = len(carryover_data)
    passed_all = sum(1 for student in carryover_data 
                    if all(course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD 
                          for course_data in student['RESIT_COURSES'].values()))
    
    carryover_count = total_students - passed_all
    total_failed_attempts = sum(failed_counts.values())
    
    # LEFT-ALIGNED SUMMARY DATA
    summary_data = [
        ["SUMMARY"],
        [f"A total of {total_students} students registered and sat for the Carryover Examination"],
        [f"A total of {passed_all} students passed all carryover courses"],
        [f"A total of {carryover_count} students failed one or more carryover courses and must repeat them"],
        [f"Total failed resit attempts: {total_failed_attempts} across all courses"],
        [f"Carryover processing completed on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"],
        [""],  # Empty row for spacing
        [""],  # Another empty row
        ["", ""],  # Signatories will be placed in separate columns
        ["________________________", "________________________"],
        ["Mrs. Abini Hauwa", "Mrs. Olukemi Ogunleye"],
        ["Head of Exams", "Chairman, ND/HND Program C'tee"]
    ]
    
    for i, row_data in enumerate(summary_data):
        row_num = summary_start_row + i
        if len(row_data) == 1:
            if row_data[0]:  # Only merge if there's actual content
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
                cell = ws.cell(row=row_num, column=1, value=row_data[0])
                if i == 0:  # "SUMMARY" header
                    cell.font = Font(bold=True, size=12, underline='single')
                else:
                    cell.font = Font(bold=False, size=11)
                # LEFT ALIGNMENT for summary text
                cell.alignment = Alignment(horizontal='left', vertical='center')
        elif len(row_data) == 2:
            # MOVED SIGNATORIES FURTHER LEFT - aligned with summary
            left_cell = ws.cell(row=row_num, column=1, value=row_data[0])
            right_cell = ws.cell(row=row_num, column=4, value=row_data[1])
            
            # Style signatory rows
            if i >= len(summary_data) - 3:  # Last 3 rows are signatories
                left_cell.alignment = Alignment(horizontal='left')
                right_cell.alignment = Alignment(horizontal='left')
                left_cell.font = Font(bold=True, size=11)
                right_cell.font = Font(bold=True, size=11)
    
    # Apply borders to data area
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=7, max_row=row_idx-1, min_col=1, max_col=len(headers)):  # CHANGED: min_row=7
        for cell in row:
            cell.border = thin_border
    
    # Professional formatting
    ws.freeze_panes = 'D8'  # CHANGED: Freeze at row 8 (data start)
    
    # Set professional font for entire worksheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or not cell.font.bold:
                cell.font = Font(name='Calibri', size=11)
    
    # IMPROVED AUTO-ADJUST COLUMN WIDTHS - PROPERLY FITS LONGEST TEXT
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in the column to find the longest content
        for cell in column:
            try:
                if cell.value is not None:
                    # Convert to string and calculate length
                    cell_value = str(cell.value)
                    cell_length = len(cell_value)
                    
                    # For rotated text in row 5, we need to handle differently
                    if cell.row == 5 and cell.alignment.text_rotation == 90:
                        # For rotated text, we want wider columns to accommodate the text
                        cell_length = max(cell_length, 10)  # Minimum width for rotated text
                    
                    # Adjust for numeric values (scores, GPAs)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell_length = max(cell_length, 8)  # Ensure enough space for numbers
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width based on content with reasonable limits
        adjusted_width = min(max_length + 2, 50)  # Add padding, cap at 50
        
        # Apply specific adjustments for different column types
        if col_idx == 1:  # S/N
            adjusted_width = 8
        elif col_idx == 2:  # EXAM NUMBER
            adjusted_width = 18
        elif col_idx == 3:  # NAME
            adjusted_width = 35  # Generous space for full names
        elif col_idx >= 4 and col_idx <= (4 + (3 if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER" else 0)):
            # GPA columns
            adjusted_width = 15
        elif col_idx >= len(headers) - 2:  # GPA Semester 4, CGPA, REMARKS
            adjusted_width = 15
        else:
            # Course columns - ensure they're wide enough for content
            adjusted_width = min(max(adjusted_width, 12), 25)  # Course columns between 12-25 width
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply alternating row colors for better readability
    for row_idx in range(8, row_idx):  # Data rows only (starting from row 8)
        if row_idx % 2 == 0:  # Even rows
            for cell in ws[row_idx]:
                # Only apply if no special fill (like course colors or pass/fail colors)
                if (cell.fill.start_color.index == '00000000' or 
                    cell.fill.start_color.index == '00FFFFFF'):
                    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    
    # Color code GPA columns in data area for better distinction
    gpa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light purple
    if semester_key == "ND-SECOND-YEAR-SECOND-SEMESTER":
        for row in range(8, row_idx):  # CHANGED: Starting from row 8
            for col in range(4, 7):  # GPA Semester 1, 2, 3 columns
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.index == '00000000':
                    cell.fill = gpa_fill
    
    # Color code final GPA columns in data area
    final_gpa_fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light cyan
    for row in range(8, row_idx):  # CHANGED: Starting from row 8
        for col in range(len(headers)-2, len(headers)+1):  # GPA Semester 4, CGPA, REMARKS
            cell = ws.cell(row=row, column=col)
            if cell.fill.start_color.index == '00000000':
                cell.fill = final_gpa_fill
    
    # Save file
    filename = f"CARRYOVER_mastersheet_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    print(f"✅ CARRYOVER mastersheet generated: {filepath}")
    print(f"📊 Course title mapping used: {course_title_mapping}")
    print(f"📊 Credit units used: {course_unit_mapping}")
    print(f"🎨 Applied color coding: Course title row (row 5) with pastel colors")
    return filepath

def generate_remarks(resit_courses):
    """Generate remarks based on resit performance."""
    passed_count = sum(1 for course_data in resit_courses.values() 
                      if course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD)
    total_count = len(resit_courses)
    
    if passed_count == total_count:
        return "All courses passed in resit"
    elif passed_count > 0:
        return f"{passed_count}/{total_count} courses passed in resit"
    else:
        return "No improvement in resit"

def generate_individual_reports(carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data):
    """Generate individual student reports in CSV format."""
    reports_dir = os.path.join(output_dir, "INDIVIDUAL_REPORTS")
    os.makedirs(reports_dir, exist_ok=True)
    
    for student in carryover_data:
        exam_no = student['EXAM NUMBER']
        # Sanitize the exam number for filename safety
        safe_exam_no = sanitize_filename(exam_no)
        filename = f"carryover_report_{safe_exam_no}_{timestamp}.csv"
        filepath = os.path.join(reports_dir, filename)
        
        report_data = []
        report_data.append(["CARRYOVER RESULT REPORT"])
        report_data.append(["FCT COLLEGE OF NURSING SCIENCES"])
        report_data.append([f"Set: {set_name}"])
        report_data.append([f"Semester: {semester_key}"])
        report_data.append([])
        report_data.append(["STUDENT INFORMATION"])
        report_data.append(["Exam Number:", student['EXAM NUMBER']])
        report_data.append(["Name:", student['NAME']])
        report_data.append([])
        
        # Previous GPAs
        report_data.append(["PREVIOUS GPAs"])
        for key in sorted([k for k in student.keys() if k.startswith('GPA_')]):
            semester = key.replace('GPA_', '')
            report_data.append([f"{semester}:", student[key]])
        report_data.append([])
        
        # Current GPA and CGPA
        report_data.append(["CURRENT ACADEMIC RECORD"])
        report_data.append(["Current GPA:", student['CURRENT_GPA']])
        report_data.append(["Current CGPA:", student['CURRENT_CGPA']])
        report_data.append([])
        
        # Resit courses
        report_data.append(["RESIT COURSES"])
        report_data.append(["Course Code", "Course Title", "Credit Unit", "Original Score", "Resit Score", "Status"])
        
        for course_code, course_data in student['RESIT_COURSES'].items():
            status = "PASSED" if course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD else "FAILED"
            course_title = course_data.get('course_title', course_code)
            credit_unit = course_data.get('credit_unit', 0)
            report_data.append([
                course_code, 
                course_title,
                credit_unit,
                course_data['original_score'], 
                course_data['resit_score'], 
                status
            ])
        
        # Save CSV
        try:
            df = pd.DataFrame(report_data)
            df.to_csv(filepath, index=False, header=False)
            print(f"✅ Generated report for: {exam_no}")
        except Exception as e:
            print(f"❌ Error generating report for {exam_no}: {e}")
    
    print(f"✅ Generated {len(carryover_data)} individual student reports in {reports_dir}")

def create_carryover_zip(source_dir, zip_path):
    """Create ZIP file of carryover results."""
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        print(f"✅ ZIP file created: {zip_path}")
        return True
    except Exception as e:
        print(f"❌ Error creating ZIP: {e}")
        return False

def main():
    """Main function to process carryover results."""
    print("🎯 CARRYOVER RESULT PROCESSOR - UPDATED WORKING VERSION")
    print("=" * 50)
    
    # Configuration
    set_name = "ND-2024"
    semester_key = "ND-SECOND-YEAR-SECOND-SEMESTER"
    pass_threshold = DEFAULT_PASS_THRESHOLD
    
    # Path setup
    raw_dir = os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", set_name, "RAW_RESULTS")
    clean_dir = os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", set_name, "CLEAN_RESULTS")
    output_dir = clean_dir
    
    print(f"📁 Base directory: {BASE_DIR}")
    print(f"📁 Raw directory: {raw_dir}")
    print(f"📁 Clean directory: {clean_dir}")
    
    # Check if directories exist
    if not os.path.exists(raw_dir):
        print(f"❌ Raw directory doesn't exist: {raw_dir}")
        print("💡 Please check if the set name and directory structure are correct")
        return
    
    if not os.path.exists(clean_dir):
        print(f"❌ Clean directory doesn't exist: {clean_dir}")
        print("💡 Please run the regular result processor first to generate clean results")
        return
    
    # Find resit file
    resit_subdir = os.path.join(raw_dir, "CARRYOVER")
    if not os.path.exists(resit_subdir):
        print(f"❌ CARRYOVER subdirectory doesn't exist: {resit_subdir}")
        print("💡 Creating CARRYOVER directory...")
        os.makedirs(resit_subdir, exist_ok=True)
        print(f"📁 Please place your carryover resit file in: {resit_subdir}")
        return
    
    resit_files = [f for f in os.listdir(resit_subdir) 
                   if f.lower().endswith(('.xlsx', '.xls')) and 'carryover' in f.lower()]
    
    if not resit_files:
        print(f"❌ No resit files found in {resit_subdir}")
        print(f"📁 Available files: {os.listdir(resit_subdir)}")
        print("💡 Please ensure your resit file has 'carryover' in the filename")
        return
    
    resit_file_path = os.path.join(resit_subdir, resit_files[0])
    print(f"📄 Using resit file: {resit_file_path}")
    
    # Find latest ZIP file
    zip_file_path = find_latest_zip_file(clean_dir)
    if not zip_file_path:
        print(f"❌ No ZIP files found in {clean_dir}")
        print("💡 Please run the regular result processor first to generate clean results ZIP")
        return
    
    # Process carryover results
    success = process_carryover_results(
        resit_file_path, zip_file_path, semester_key, set_name, pass_threshold, output_dir
    )
    
    if success:
        print(f"\n✅ Carryover processing completed successfully!")
        print(f"📁 Check the CLEAN_RESULTS directory for the CARRYOVER output")
    else:
        print(f"\n❌ Carryover processing failed!")

if __name__ == "__main__":
    main()