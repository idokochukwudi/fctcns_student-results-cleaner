#!/usr/bin/env python3
"""
BM CARRYOVER PROCESSOR - Complete Basic Midwifery Carryover Processor
FIXED VERSION for Web Interface Integration
Based on BN Carryover Processor with BM-specific adaptations
Handles Basic Midwifery (BM) carryover/resit result processing with:
- 6 semesters (3 years × 2 semesters)
- M- prefix for semester keys
- Proper GPA/CGPA recalculation
- Mastersheet updates with all enhancements
- SINGLE workbook session to prevent corruption
- WEB INTERFACE COMPATIBLE

CRITICAL FIX APPLIED:
1. Fixed GPA calculation functions with correct Nigerian 5.0 scale
2. Fixed withdrawn status override - now properly resets withdrawn status based on CURRENT carryover results
3. Removed duplicate exception block
4. Added robust ZIP creation function
5. Optimized course variant generation
6. Safe dictionary access in CGPA summary
7. Early variable initialization
8. Progress indicators for long operations
9. Fixed improper finally block placement
10. ADDED: Complete ZIP creation at end of processing
"""

import os
import sys
import re
import pandas as pd
import numpy as np
from datetime import datetime
import glob
import json
import traceback
import shutil
import zipfile
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CRITICAL FIX 1: Configuration with Early Initialization
# ============================================================
def get_base_directory():
    """Get base directory - FIXED for web interface"""
    # Priority 1: Environment variable
    if os.getenv("BASE_DIR"):
        base_dir = os.getenv("BASE_DIR")
        # CRITICAL FIX: Adjust if BASE_DIR ends with EXAMS_INTERNAL
        if os.path.basename(base_dir) == "EXAMS_INTERNAL":
            base_dir = os.path.dirname(base_dir)
            print(f"✅ Adjusted BASE_DIR to parent: {base_dir}")
        if os.path.exists(base_dir):
            print(f"✅ Using BASE_DIR from environment: {base_dir}")
            return base_dir
    # Priority 2: Standard locations
    possible_dirs = [
        os.path.join(os.path.expanduser("~"), "student_result_cleaner"),
        os.path.join(
            os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"
        ),
        os.getcwd(),
        os.path.dirname(os.path.abspath(__file__)),
    ]
    for dir_path in possible_dirs:
        if os.path.exists(os.path.join(dir_path, "EXAMS_INTERNAL")):
            print(f"✅ Found EXAMS_INTERNAL in: {dir_path}")
            return dir_path
    # Fallback
    default_dir = os.path.join(os.path.expanduser("~"), "student_result_cleaner")
    print(f"⚠️ Using default directory: {default_dir}")
    return default_dir


BASE_DIR = get_base_directory()
TIMESTAMP_FMT = "%d-%m-%Y_%H%M%S"
DEFAULT_PASS_THRESHOLD = 50.0
DEFAULT_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")
print(f"🔧 BASE_DIR set to: {BASE_DIR}")


# ============================================================
# CRITICAL FIX 2: New ZIP Creation Function
# ============================================================
def create_updated_zip_from_directory(temp_extract_dir, updated_zip_path):
    """
    FIXED VERSION: Create a ZIP file from a directory with proper error handling.
    Ensures only ONE updated ZIP is created.
    """
    try:
        # Remove existing ZIP if it exists to avoid duplicates
        if os.path.exists(updated_zip_path):
            os.remove(updated_zip_path)
            print(f"🧹 Removed existing ZIP: {updated_zip_path}")

        with zipfile.ZipFile(updated_zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_extract_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_extract_dir)
                    zipf.write(file_path, arcname)
                    print(f"📁 Added to updated ZIP: {arcname}")

        # Verify ZIP was created
        if os.path.exists(updated_zip_path) and os.path.getsize(updated_zip_path) > 0:
            print(
                f"✅ SUCCESS: Created SINGLE updated ZIP ({os.path.getsize(updated_zip_path):,} bytes)"
            )

            # Verify mastersheet is in ZIP
            try:
                with zipfile.ZipFile(updated_zip_path, "r") as test_zip:
                    zip_files = test_zip.namelist()
                    mastersheet_in_zip = any(
                        "mastersheet" in f.lower() for f in zip_files
                    )

                    if mastersheet_in_zip:
                        print(f"✅ Verified: Mastersheet is in the updated ZIP")
                    else:
                        print(f"⚠️ Warning: No mastersheet found in updated ZIP")

            except Exception as e:
                print(f"⚠️ Could not verify ZIP contents: {e}")

            return True
        else:
            print(f"❌ ERROR: ZIP was not created properly")
            return False

    except Exception as zip_error:
        print(f"❌ Error creating ZIP: {zip_error}")
        traceback.print_exc()
        return False

# ============================================================
# CRITICAL FIX 3: Optimized Course Variant Generation
# ============================================================
def generate_course_variants(course_code, max_variants=15):
    """
    Generate normalized course code variants with limit.

    Args:
        course_code: Original course code
        max_variants: Maximum number of variants to generate

    Returns:
        list: Unique course code variants
    """
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return []

    original_code = str(course_code).strip()
    variants = []

    # Priority variants (most likely to match)
    priority_variants = [
        original_code.upper().strip(),
        re.sub(r"[^A-Z0-9]", "", original_code.upper()),
        original_code.upper().replace(" ", ""),
        f"MID{original_code.upper()}",
        f"MWF{original_code.upper()}",
    ]
    variants.extend(priority_variants)

    # Additional variants only if needed
    if len(variants) < max_variants:
        additional_variants = [
            original_code.strip(),
            original_code.upper(),
            re.sub(r"\s+", "", original_code.upper()),
            original_code.upper().replace("-", ""),
            original_code.upper().replace("_", ""),
            f"MID{re.sub(r'[^A-Z0-9]', '', original_code.upper())}",
            f"MWF{re.sub(r'[^A-Z0-9]', '', original_code.upper())}",
            re.sub(r"^(MID|MWF)", "", original_code.upper()).strip(),
            original_code.upper().replace(".", ""),
        ]
        variants.extend(additional_variants)

    # Remove duplicates while preserving order
    seen = set()
    unique_variants = []
    for variant in variants[:max_variants]:
        if variant and variant not in seen and variant not in ["NAN", "NONE", ""]:
            seen.add(variant)
            unique_variants.append(variant)

    return unique_variants


# ============================================================
# CRITICAL FIX 4: Safe Dictionary Access in CGPA Summary
# ============================================================
def write_cgpa_summary_data_bm_safe(cgpa_ws, semester_data, all_withdrawn_students):
    """
    Safely write CGPA summary data with proper error handling.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Collect unique students
    all_exam_no = set()
    for semester_dict in semester_data.values():
        all_exam_no.update(semester_dict.keys())

    students = []
    for exam_no in all_exam_no:
        total_gp = 0.0
        total_cr = 0.0
        gpas = {}
        name = None
        withdrawn = False

        for key, semester_dict in semester_data.items():
            # SAFE ACCESS: Use .get() instead of direct dictionary access
            student_data = semester_dict.get(exam_no)

            if student_data:
                gpas[key] = student_data.get("gpa", 0.0)
                total_gp += student_data.get("gpa", 0.0) * student_data.get(
                    "credits", 0.0
                )
                total_cr += student_data.get("credits", 0.0)

                if student_data.get("withdrawn", False):
                    withdrawn = True

                if not name:
                    name = student_data.get("name", "Unknown")

        cgpa = round(total_gp / total_cr, 2) if total_cr > 0 else 0.0

        students.append(
            {
                "exam_no": exam_no,
                "name": name,
                "gpas": gpas,
                "cgpa": cgpa,
                "withdrawn": withdrawn,
            }
        )

    # Sort students (non-withdrawn first)
    non_withdrawn = [s for s in students if not s["withdrawn"]]
    withdrawn_list = [s for s in students if s["withdrawn"]]

    non_withdrawn.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    withdrawn_list.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))

    sorted_students = non_withdrawn + withdrawn_list

    # Write data starting from row 7
    start_row = 7
    for idx, student in enumerate(sorted_students, start_row):
        # Serial number
        cgpa_ws.cell(row=idx, column=1, value=idx - 6)
        # Exam number
        cgpa_ws.cell(row=idx, column=2, value=student["exam_no"])
        # Name
        cgpa_ws.cell(row=idx, column=3, value=student["name"])

        # GPA for each semester
        col = 4
        for semester_key in [
            "M-FIRST-YEAR-FIRST-SEMESTER",
            "M-FIRST-YEAR-SECOND-SEMESTER",
            "M-SECOND-YEAR-FIRST-SEMESTER",
            "M-SECOND-YEAR-SECOND-SEMESTER",
            "M-THIRD-YEAR-FIRST-SEMESTER",
            "M-THIRD-YEAR-SECOND-SEMESTER",
        ]:
            gpa_value = student["gpas"].get(semester_key, "")
            cgpa_ws.cell(row=idx, column=col, value=gpa_value)
            col += 1

        # CGPA
        cgpa_ws.cell(row=idx, column=10, value=student["cgpa"])

        # Withdrawn status
        withdrawn_status = "YES" if student["withdrawn"] else "NO"
        cgpa_ws.cell(row=idx, column=11, value=withdrawn_status)

    return len(sorted_students)


# ============================================================
# FIXED GPA CALCULATION FUNCTIONS
# ============================================================
def get_grade_point(score):
    """Determine grade point based on score - NIGERIAN 5.0 SCALE - FIXED VERSION."""
    try:
        score = float(score)
        if score >= 70:
            return 5.0  # A
        elif score >= 60:
            return 4.0  # B
        elif score >= 50:
            return 3.0  # C
        elif score >= 45:
            return 2.0  # D
        elif score >= 40:
            return 1.0  # E
        else:
            return 0.0  # F
    except (ValueError, TypeError):
        return 0.0


def calculate_gpa_correctly(scores, credit_units_dict, course_code_to_unit):
    """
    CORRECT GPA CALCULATION:
    GPA = (Sum of (Grade Point * Credit Units)) / (Total Credit Units)
    
    Args:
        scores: Dictionary of course_code -> score
        credit_units_dict: Semester-specific credit units
        course_code_to_unit: Global course code to credit units mapping
    
    Returns:
        tuple: (gpa, total_grade_points, total_credits, cu_passed, cu_failed)
    """
    total_grade_points = 0.0
    total_credits = 0
    cu_passed = 0
    cu_failed = 0
    
    for course_code, score in scores.items():
        if score is None or score == "":
            continue
            
        try:
            score_val = float(score)
            # Get credit unit for this course
            credit_unit = find_credit_unit(course_code, credit_units_dict, course_code_to_unit)
            
            # Get grade point for this score
            grade_point = get_grade_point(score_val)
            
            # Calculate quality points for this course
            quality_points = grade_point * credit_unit
            
            # Add to totals
            total_grade_points += quality_points
            total_credits += credit_unit
            
            # Track passed/failed credits
            if score_val >= 50:  # Passing score
                cu_passed += credit_unit
            else:
                cu_failed += credit_unit
                
        except (ValueError, TypeError):
            continue
    
    # Calculate GPA
    gpa = round(total_grade_points / total_credits, 2) if total_credits > 0 else 0.0
    
    return gpa, total_grade_points, total_credits, cu_passed, cu_failed


def recalculate_student_gpa_fixed(student_row, headers, course_columns, credit_units_dict, course_code_to_unit):
    """
    FIXED: Recalculate student GPA using correct method.
    
    Args:
        student_row: The row from mastersheet DataFrame
        headers: Dictionary of header names to column indices
        course_columns: Dictionary of course codes to column indices  
        credit_units_dict: Semester-specific credit units
        course_code_to_unit: Global course code to credit units mapping
    
    Returns:
        dict: Updated student metrics
    """
    # Extract scores for all courses
    scores = {}
    for course_code, col_idx in course_columns.items():
        score = student_row.iloc[col_idx - 1] if col_idx - 1 < len(student_row) else None
        scores[course_code] = score
    
    # Calculate GPA correctly
    gpa, total_grade_points, total_credits, cu_passed, cu_failed = calculate_gpa_correctly(
        scores, credit_units_dict, course_code_to_unit
    )
    
    # Calculate average score (for reference, not for GPA)
    valid_scores = [float(score) for score in scores.values() if score is not None and score != ""]
    average = round(sum(valid_scores) / len(valid_scores), 2) if valid_scores else 0.0
    
    # Identify failed courses
    failed_courses = []
    for course_code, score in scores.items():
        if score is not None and score != "":
            try:
                if float(score) < 50:
                    failed_courses.append(course_code)
            except (ValueError, TypeError):
                continue
    
    return {
        'gpa': gpa,
        'average': average,
        'total_credits': total_credits,
        'cu_passed': cu_passed,
        'cu_failed': cu_failed,
        'failed_courses': failed_courses,
        'total_grade_points': total_grade_points
    }


# ============================================================
# UTILITY FUNCTIONS (BM-Compatible)
# ============================================================
def sanitize_filename(filename):
    """Remove or replace characters that are not safe for filenames."""
    return re.sub(r"[^\w\-_.]", "_", filename)


def find_exam_number_column(df):
    """Find the exam number column in a DataFrame - FIXED for EXAMS NUMBER"""
    possible_names = [
        "EXAMS NUMBER",  # PRIMARY - plural form (this is what's actually used)
        "EXAM NUMBER",  # Secondary - singular form
        "REG. No",
        "REG NO",
        "REGISTRATION NUMBER",
        "MAT NO",
        "STUDENT ID",
    ]
    for col in df.columns:
        col_upper = str(col).upper()
        for possible_name in possible_names:
            if possible_name in col_upper:
                print(f"✅ Found exam column: '{col}' matches '{possible_name}'")
                return col
    print(f"❌ No exam number column found in: {list(df.columns)}")
    return None


def generate_remarks(resit_courses):
    """Generate remarks for resit performance."""
    passed_count = sum(
        1
        for course_data in resit_courses.values()
        if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
    )
    total_count = len(resit_courses)
    if passed_count == total_count:
        return "All courses passed in resit"
    elif passed_count > 0:
        return f"{passed_count}/{total_count} courses passed in resit"
    else:
        return "No improvement in resit"


def extract_class_from_set_name(set_name):
    """Extract class name from set_name (e.g., 'SET47' from 'SET47')"""
    return set_name


# ============================================================
# BM-SPECIFIC SEMESTER FUNCTIONS - FIXED VERSION
# ============================================================
def is_bm_semester(semester_key):
    """Check if semester key belongs to BM program"""
    if not semester_key:
        return False
    key_upper = semester_key.upper()
    # BM indicators
    bm_indicators = [
        key_upper.startswith("M-"),
        " M " in key_upper,
        " BM " in key_upper,
        key_upper.startswith("BM-"),
        "BASIC MIDWIFERY" in key_upper,
        "B.M." in key_upper,
    ]
    return any(bm_indicators)


def standardize_semester_key(semester_key):
    """Standardize semester key to canonical format for BM - FIXED VERSION"""
    if not semester_key:
        return None
    key_upper = semester_key.upper()
    print(f"🔍 Standardizing BM semester key: '{semester_key}' -> '{key_upper}'")
    # BM canonical mappings (M- prefix) - ENHANCED WITH BETTER PATTERN MATCHING
    canonical_mappings = {
        # Year 1 - BM patterns (M- prefix)
        ("M", "FIRST", "YEAR", "FIRST", "SEMESTER"): "M-FIRST-YEAR-FIRST-SEMESTER",
        ("M", "FIRST", "YEAR", "1ST", "SEMESTER"): "M-FIRST-YEAR-FIRST-SEMESTER",
        ("M", "1ST", "YEAR", "FIRST", "SEMESTER"): "M-FIRST-YEAR-FIRST-SEMESTER",
        ("M", "YEAR", "1", "SEMESTER", "1"): "M-FIRST-YEAR-FIRST-SEMESTER",
        (
            "FIRST",
            "YEAR",
            "FIRST",
            "SEMESTER",
        ): "M-FIRST-YEAR-FIRST-SEMESTER",  # Fallback
        ("1ST", "YEAR", "1ST", "SEMESTER"): "M-FIRST-YEAR-FIRST-SEMESTER",  # Fallback
        ("M", "FIRST", "YEAR", "SECOND", "SEMESTER"): "M-FIRST-YEAR-SECOND-SEMESTER",
        ("M", "FIRST", "YEAR", "2ND", "SEMESTER"): "M-FIRST-YEAR-SECOND-SEMESTER",
        ("M", "1ST", "YEAR", "SECOND", "SEMESTER"): "M-FIRST-YEAR-SECOND-SEMESTER",
        ("M", "YEAR", "1", "SEMESTER", "2"): "M-FIRST-YEAR-SECOND-SEMESTER",
        (
            "FIRST",
            "YEAR",
            "SECOND",
            "SEMESTER",
        ): "M-FIRST-YEAR-SECOND-SEMESTER",  # Fallback
        ("1ST", "YEAR", "2ND", "SEMESTER"): "M-FIRST-YEAR-SECOND-SEMESTER",  # Fallback
        # Year 2 - BM patterns
        ("M", "SECOND", "YEAR", "FIRST", "SEMESTER"): "M-SECOND-YEAR-FIRST-SEMESTER",
        ("M", "SECOND", "YEAR", "1ST", "SEMESTER"): "M-SECOND-YEAR-FIRST-SEMESTER",
        ("M", "2ND", "YEAR", "FIRST", "SEMESTER"): "M-SECOND-YEAR-FIRST-SEMESTER",
        ("M", "YEAR", "2", "SEMESTER", "1"): "M-SECOND-YEAR-FIRST-SEMESTER",
        (
            "SECOND",
            "YEAR",
            "FIRST",
            "SEMESTER",
        ): "M-SECOND-YEAR-FIRST-SEMESTER",  # Fallback
        ("2ND", "YEAR", "1ST", "SEMESTER"): "M-SECOND-YEAR-FIRST-SEMESTER",  # Fallback
        ("M", "SECOND", "YEAR", "SECOND", "SEMESTER"): "M-SECOND-YEAR-SECOND-SEMESTER",
        ("M", "SECOND", "YEAR", "2ND", "SEMESTER"): "M-SECOND-YEAR-SECOND-SEMESTER",
        ("M", "2ND", "YEAR", "SECOND", "SEMESTER"): "M-SECOND-YEAR-SECOND-SEMESTER",
        ("M", "YEAR", "2", "SEMESTER", "2"): "M-SECOND-YEAR-SECOND-SEMESTER",
        (
            "SECOND",
            "YEAR",
            "SECOND",
            "SEMESTER",
        ): "M-SECOND-YEAR-SECOND-SEMESTER",  # Fallback
        ("2ND", "YEAR", "2ND", "SEMESTER"): "M-SECOND-YEAR-SECOND-SEMESTER",  # Fallback
        # Year 3 - BM patterns
        ("M", "THIRD", "YEAR", "FIRST", "SEMESTER"): "M-THIRD-YEAR-FIRST-SEMESTER",
        ("M", "THIRD", "YEAR", "1ST", "SEMESTER"): "M-THIRD-YEAR-FIRST-SEMESTER",
        ("M", "3RD", "YEAR", "FIRST", "SEMESTER"): "M-THIRD-YEAR-FIRST-SEMESTER",
        ("M", "YEAR", "3", "SEMESTER", "1"): "M-THIRD-YEAR-FIRST-SEMESTER",
        (
            "THIRD",
            "YEAR",
            "FIRST",
            "SEMESTER",
        ): "M-THIRD-YEAR-FIRST-SEMESTER",  # Fallback
        ("3RD", "YEAR", "1ST", "SEMESTER"): "M-THIRD-YEAR-FIRST-SEMESTER",  # Fallback
        ("M", "THIRD", "YEAR", "SECOND", "SEMESTER"): "M-THIRD-YEAR-SECOND-SEMESTER",
        ("M", "THIRD", "YEAR", "2ND", "SEMESTER"): "M-THIRD-YEAR-SECOND-SEMESTER",
        ("M", "3RD", "YEAR", "SECOND", "SEMESTER"): "M-THIRD-YEAR-SECOND-SEMESTER",
        ("M", "YEAR", "3", "SEMESTER", "2"): "M-THIRD-YEAR-SECOND-SEMESTER",
        (
            "THIRD",
            "YEAR",
            "SECOND",
            "SEMESTER",
        ): "M-THIRD-YEAR-SECOND-SEMESTER",  # Fallback
        ("3RD", "YEAR", "2ND", "SEMESTER"): "M-THIRD-YEAR-SECOND-SEMESTER",  # Fallback
    }
    # First, try exact matching with BM prefixes
    for key_parts, canonical in canonical_mappings.items():
        if all(part in key_upper for part in key_parts):
            print(f"✅ Exact BM match: '{semester_key}' -> '{canonical}'")
            return canonical
    # BM patterns with M- prefix
    bm_patterns = [
        (
            r"(M-?)\s*(FIRST|1ST|YEAR\s*1).*?(FIRST|1ST|SEMESTER\s*1)",
            "M-FIRST-YEAR-FIRST-SEMESTER",
        ),
        (
            r"(M-?)\s*(FIRST|1ST|YEAR\s*1).*?(SECOND|2ND|SEMESTER\s*2)",
            "M-FIRST-YEAR-SECOND-SEMESTER",
        ),
        (
            r"(M-?)\s*(SECOND|2ND|YEAR\s*2).*?(FIRST|1ST|SEMESTER\s*1)",
            "M-SECOND-YEAR-FIRST-SEMESTER",
        ),
        (
            r"(M-?)\s*(SECOND|2ND|YEAR\s*2).*?(SECOND|2ND|SEMESTER\s*2)",
            "M-SECOND-YEAR-SECOND-SEMESTER",
        ),
        (
            r"(M-?)\s*(THIRD|3RD|YEAR\s*3).*?(FIRST|1ST|SEMESTER\s*1)",
            "M-THIRD-YEAR-FIRST-SEMESTER",
        ),
        (
            r"(M-?)\s*(THIRD|3RD|YEAR\s*3).*?(SECOND|2ND|SEMESTER\s*2)",
            "M-THIRD-YEAR-SECOND-SEMESTER",
        ),
    ]
    for pattern, canonical in bm_patterns:
        if re.search(pattern, key_upper, re.IGNORECASE):
            print(f"✅ BM pattern match: '{semester_key}' -> '{canonical}'")
            return canonical
    # If we get here and the key starts with M-, preserve it as BM
    if key_upper.startswith("M-") or " M " in key_upper:
        # Clean up the key but keep it as BM
        clean_key = key_upper.replace(" ", "-").replace("_", "-").replace("--", "-")
        if clean_key.startswith("M-"):
            print(f"✅ Preserving BM key: '{semester_key}' -> '{clean_key}'")
            return clean_key
        else:
            bm_key = f"M-{clean_key}"
            print(f"✅ Converted to BM key: '{semester_key}' -> '{bm_key}'")
            return bm_key
    print(f"⚠️ Could not standardize BM semester key: {semester_key}, using as-is")
    return semester_key


def standardize_semester_name(semester_name):
    """Standardize semester name - alias for standardize_semester_key for compatibility."""
    return standardize_semester_key(semester_name)


def extract_semester_from_filename(filename):
    """Extract semester from carryover filename - BM VERSION FIXED"""
    try:
        # Handle both .json and .xlsx files
        if filename.endswith(".json") or filename.endswith(".xlsx"):
            # Pattern for BM files: co_student_BM-SET2023_M-SECOND-YEAR-FIRST-SEMESTER_20251107_100522.json
            # Also handle: co_student_BM-M-FIRST-YEAR-FIRST-SEMESTER_20251112_192740.json
            patterns = [
                r"co_student_BM-SET\d+_(M-.*?)_\d+_\d+\.(json|xlsx)",
                r"co_student_BM-(M-.*?)_\d+_\d+\.(json|xlsx)",
                r"BM_.*?_(M-.*?)_\d+\.(json|xlsx)",
                r"BM-CARRYOVER.*?_(M-.*?)_\d+\.(json|xlsx)",
            ]
            for pattern in patterns:
                match = re.search(pattern, filename)
                if match:
                    semester = match.group(1)
                    print(f"✅ Extracted BM semester from '{filename}': '{semester}'")
                    return semester.upper().replace("-", " ").replace("_", " ")
        # Fallback: try to extract any M- pattern (BM specific)
        match = re.search(
            r"(M-[A-Za-z-]+(?:YEAR|SEMESTER)[A-Za-z-]*)", filename, re.IGNORECASE
        )
        if match:
            semester = match.group(1)
            print(f"✅ Fallback extracted BM semester: '{semester}'")
            return semester.upper().replace("-", " ").replace("_", " ")
        print(f"❌ Could not extract semester from BM filename: {filename}")
        return None
    except Exception as e:
        print(f"Error extracting semester from {filename}: {e}")
        return None


def get_semester_display_info(semester_key):
    """Get display information for BM semester key."""
    semester_lower = semester_key.lower()
    semester_info = {
        "first-year-first-semester": (
            1,
            1,
            "YEAR ONE",
            "FIRST SEMESTER",
            "BMI",
            "Semester 1",
        ),
        "first-year-second-semester": (
            1,
            2,
            "YEAR ONE",
            "SECOND SEMESTER",
            "BMI",
            "Semester 2",
        ),
        "second-year-first-semester": (
            2,
            1,
            "YEAR TWO",
            "FIRST SEMESTER",
            "BMII",
            "Semester 3",
        ),
        "second-year-second-semester": (
            2,
            2,
            "YEAR TWO",
            "SECOND SEMESTER",
            "BMII",
            "Semester 4",
        ),
        "third-year-first-semester": (
            3,
            1,
            "YEAR THREE",
            "FIRST SEMESTER",
            "BMIII",
            "Semester 5",
        ),
        "third-year-second-semester": (
            3,
            2,
            "YEAR THREE",
            "SECOND SEMESTER",
            "BMIII",
            "Semester 6",
        ),
    }
    for key, info in semester_info.items():
        if key in semester_lower:
            return info
    return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BMI", "Semester 1"


def get_previous_semester(semester_key):
    """Get the previous semester key for BM carryover."""
    standardized = standardize_semester_key(semester_key)
    # BM semesters chain
    semester_chain = {
        "M-FIRST-YEAR-SECOND-SEMESTER": "M-FIRST-YEAR-FIRST-SEMESTER",
        "M-SECOND-YEAR-FIRST-SEMESTER": "M-FIRST-YEAR-SECOND-SEMESTER",
        "M-SECOND-YEAR-SECOND-SEMESTER": "M-SECOND-YEAR-FIRST-SEMESTER",
        "M-THIRD-YEAR-FIRST-SEMESTER": "M-SECOND-YEAR-SECOND-SEMESTER",
        "M-THIRD-YEAR-SECOND-SEMESTER": "M-THIRD-YEAR-FIRST-SEMESTER",
    }
    return semester_chain.get(standardized)


def get_previous_semesters_for_display(current_semester_key):
    """Get list of previous semesters for BM GPA display in mastersheet."""
    current_standard = standardize_semester_key(current_semester_key)
    semester_mapping = {
        "M-FIRST-YEAR-FIRST-SEMESTER": [],
        "M-FIRST-YEAR-SECOND-SEMESTER": ["Semester 1"],
        "M-SECOND-YEAR-FIRST-SEMESTER": ["Semester 1", "Semester 2"],
        "M-SECOND-YEAR-SECOND-SEMESTER": ["Semester 1", "Semester 2", "Semester 3"],
        "M-THIRD-YEAR-FIRST-SEMESTER": [
            "Semester 1",
            "Semester 2",
            "Semester 3",
            "Semester 4",
        ],
        "M-THIRD-YEAR-SECOND-SEMESTER": [
            "Semester 1",
            "Semester 2",
            "Semester 3",
            "Semester 4",
            "Semester 5",
        ],
    }
    return semester_mapping.get(current_standard, [])


# ============================================================
# BM COURSE DATA MANAGEMENT WITH OPTIMIZED VARIANT GENERATION
# ============================================================
def load_course_data():
    """Load BM course data ONLY."""
    return load_bm_course_data()


def load_bm_course_data():
    """Load BM course data from M-course-code-creditUnit.xlsx."""
    possible_course_files = [
        os.path.join(
            BASE_DIR,
            "EXAMS_INTERNAL",
            "BM",
            "BM-COURSES",
            "M-course-code-creditUnit.xlsx",
        ),
        os.path.join(BASE_DIR, "BM", "BM-COURSES", "M-course-code-creditUnit.xlsx"),
        os.path.join(
            BASE_DIR, "EXAMS_INTERNAL", "BM-COURSES", "M-course-code-creditUnit.xlsx"
        ),
        os.path.join(BASE_DIR, "M-course-code-creditUnit.xlsx"),
    ]
    course_file = None
    for possible_file in possible_course_files:
        if os.path.exists(possible_file):
            course_file = possible_file
            print(f"✅ Found BM course file: {possible_file}")
            break
    if not course_file:
        print(f"❌ Main BM course file not found in standard locations")
        alternative_files = find_alternative_bm_course_files()
        if alternative_files:
            course_file = alternative_files[0]
            print(f"🔄 Using alternative BM course file: {course_file}")
        else:
            print(f"❌ No BM course files found anywhere!")
            return {}, {}, {}, {}
    print(f"📚 Loading BM course data from: {course_file}")
    return _load_course_data_from_file_bm(course_file)


def _load_course_data_from_file_bm(course_file):
    """Generic function to load BM course data from Excel file."""
    try:
        xl = pd.ExcelFile(course_file)
        semester_course_titles = {}
        semester_credit_units = {}
        course_code_to_title = {}
        course_code_to_unit = {}
        print(f"📖 Available sheets: {xl.sheet_names}")
        for sheet in xl.sheet_names:
            sheet_standard = standardize_semester_key(sheet)
            print(f"📖 Reading sheet: {sheet} (standardized: {sheet_standard})")
            try:
                df = pd.read_excel(
                    course_file, sheet_name=sheet, engine="openpyxl", header=0
                )
                # Convert columns to string and clean
                df.columns = [str(c).strip().upper() for c in df.columns]
                # Look for course code, title, and credit unit columns with flexible matching
                code_col = None
                title_col = None
                unit_col = None
                for col in df.columns:
                    col_clean = str(col).upper()
                    if any(
                        keyword in col_clean
                        for keyword in ["COURSE CODE", "CODE", "COURSECODE"]
                    ):
                        code_col = col
                    elif any(
                        keyword in col_clean
                        for keyword in ["COURSE TITLE", "TITLE", "COURSENAME"]
                    ):
                        title_col = col
                    elif any(
                        keyword in col_clean
                        for keyword in ["CU", "CREDIT", "UNIT", "CREDIT UNIT"]
                    ):
                        unit_col = col
                print(
                    f"🔍 Detected columns - Code: {code_col}, Title: {title_col}, Unit: {unit_col}"
                )
                if not all([code_col, title_col, unit_col]):
                    print(
                        f"⚠️ Sheet '{sheet}' missing required columns - found: code={code_col}, title={title_col}, unit={unit_col}"
                    )
                    # Try to use first three columns as fallback
                    if len(df.columns) >= 3:
                        code_col, title_col, unit_col = (
                            df.columns[0],
                            df.columns[1],
                            df.columns[2],
                        )
                        print(
                            f"🔄 Using fallback columns: {code_col}, {title_col}, {unit_col}"
                        )
                    else:
                        print(
                            f"❌ Sheet '{sheet}' doesn't have enough columns - skipped"
                        )
                        continue
                # Clean the data
                df_clean = df.dropna(subset=[code_col]).copy()
                if df_clean.empty:
                    print(f"⚠️ Sheet '{sheet}' has no data after cleaning - skipped")
                    continue
                # Convert credit units to numeric, handling errors
                df_clean[unit_col] = pd.to_numeric(df_clean[unit_col], errors="coerce")
                df_clean = df_clean.dropna(subset=[unit_col])
                # Remove rows with "TOTAL" in course code
                df_clean = df_clean[
                    ~df_clean[code_col]
                    .astype(str)
                    .str.contains("TOTAL", case=False, na=False)
                ]
                if df_clean.empty:
                    print(
                        f"⚠️ Sheet '{sheet}' has no valid rows after cleaning - skipped"
                        )
                    continue
                codes = df_clean[code_col].astype(str).str.strip().tolist()
                titles = df_clean[title_col].astype(str).str.strip().tolist()
                units = df_clean[unit_col].astype(float).tolist()
                print(f"📋 Found {len(codes)} courses in {sheet}:")
                for i, (code, title, unit) in enumerate(
                    zip(codes[:5], titles[:5], units[:5])
                ):
                    print(f" - '{code}': '{title}' (CU: {unit})")
                # Create mapping dictionaries with OPTIMIZED variant generation
                sheet_titles = {}
                sheet_units = {}
                for code, title, unit in zip(codes, titles, units):
                    if not code or code.upper() in ["NAN", "NONE", ""]:
                        continue

                    # USE OPTIMIZED VARIANT GENERATION
                    variants = generate_course_variants(code)

                    # Add all variants to mappings
                    for variant in variants:
                        sheet_titles[variant] = title
                        sheet_units[variant] = unit
                        course_code_to_title[variant] = title
                        course_code_to_unit[variant] = unit
                semester_course_titles[sheet_standard] = sheet_titles
                semester_credit_units[sheet_standard] = sheet_units
            except Exception as e:
                print(f"❌ Error processing sheet '{sheet}': {e}")
                traceback.print_exc()
                continue
        print(
            f"✅ Loaded course data for sheets: {list(semester_course_titles.keys())}"
        )
        print(f"📊 Total course mappings: {len(course_code_to_title)}")
        # Debug: Show some course mappings
        print("🔍 Sample course mappings:")
        sample_items = list(course_code_to_title.items())[:15]
        for code, title in sample_items:
            unit = course_code_to_unit.get(code, 0)
            print(f" '{code}' -> '{title}' (CU: {unit})")
        return (
            semester_course_titles,
            semester_credit_units,
            course_code_to_title,
            course_code_to_unit,
        )
    except Exception as e:
        print(f"❌ Error loading BM course data: {e}")
        traceback.print_exc()
        return {}, {}, {}, {}


def find_alternative_bm_course_files():
    """Look for alternative BM course files."""
    base_dirs = [
        os.path.join(BASE_DIR, "EXAMS_INTERNAL", "BM", "BM-COURSES"),
        os.path.join(BASE_DIR, "BM", "BM-COURSES"),
        os.path.join(BASE_DIR, "COURSES"),
        os.path.join(BASE_DIR, "EXAMS_INTERNAL"),
    ]
    course_files = []
    for base_dir in base_dirs:
        if os.path.exists(base_dir):
            for file in os.listdir(base_dir):
                if "course" in file.lower() and file.endswith((".xlsx", ".xls")):
                    full_path = os.path.join(base_dir, file)
                    course_files.append(full_path)
    return course_files


def debug_course_matching_bm(
    resit_file_path, course_code_to_title, course_code_to_unit
):
    """Debug function to check why BM course codes aren't matching."""
    print(f"\n🔍 DEBUGGING BM COURSE MATCHING")
    print("=" * 50)
    # Read resit file to see what course codes we have
    resit_df = pd.read_excel(resit_file_path, header=0)
    resit_exam_col = find_exam_number_column(resit_df)
    # Get all course codes from resit file
    resit_courses = []
    for col in resit_df.columns:
        if col != resit_exam_col and col != "NAME" and not "Unnamed" in str(col):
            resit_courses.append(col)
    print(f"📋 BM Course codes from resit file: {resit_courses}")
    print(f"📊 Total courses in BM resit file: {len(resit_courses)}")
    # Check each resit course against course file
    for course in resit_courses:
        print(f"\n🔍 Checking BM course: '{course}'")
        original_code = str(course).strip()

        # USE OPTIMIZED VARIANT GENERATION
        variants = generate_course_variants(original_code)

        print(f" Generated {len(variants)} variants to try")
        found = False
        for variant in variants:
            if variant in course_code_to_title:
                title = course_code_to_title[variant]
                unit = course_code_to_unit.get(variant, 0)
                print(f" ✅ FOUND: '{variant}' -> '{title}' (CU: {unit})")
                found = True
                break
        if not found:
            print(f" ❌ NOT FOUND: No match for '{course}'")
            # Show some similar keys from course file
            similar_keys = []
            # Check for partial matches
            for key in list(course_code_to_title.keys())[:50]:  # Check first 50 keys
                if any(
                    part in key.upper()
                    for part in original_code.upper().split()
                    if len(part) > 2
                ):
                    similar_keys.append(key)
            if similar_keys:
                print(f" 💡 Similar keys found: {similar_keys[:5]}")
            else:
                print(
                    f" 💡 Sample available keys: {list(course_code_to_title.keys())[:10]}"
                )


def find_course_title(course_code, course_titles_dict, course_code_to_title):
    """Robust function to find course title with OPTIMIZED variant generation."""
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return str(course_code) if course_code else "Unknown Course"
    original_code = str(course_code).strip()

    # USE OPTIMIZED VARIANT GENERATION
    variants = generate_course_variants(original_code)

    # Try each strategy in order
    for variant in variants:
        # Try course_titles_dict first (semester-specific)
        if variant in course_titles_dict:
            title = course_titles_dict[variant]
            print(
                f"✅ Found title for '{original_code}' using variant '{variant}': '{title}'"
            )
            return title
        # Try global course_code_to_title
        if variant in course_code_to_title:
            title = course_code_to_title[variant]
            print(
                f"✅ Found title for '{original_code}' using global variant '{variant}': '{title}'"
            )
            return title
    # If no match found, log and return descriptive original code
    print(f"⚠️ Could not find course title for: '{original_code}'")
    print(f" Tried {len(variants)} variants without success")
    return f"{original_code} (Title Not Found)"


def find_credit_unit(course_code, credit_units_dict, course_code_to_unit):
    """Robust function to find credit unit with OPTIMIZED variant generation."""
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return 0
    original_code = str(course_code).strip()

    # USE OPTIMIZED VARIANT GENERATION
    variants = generate_course_variants(original_code)

    # Try each strategy
    for variant in variants:
        if variant in credit_units_dict:
            unit = credit_units_dict[variant]
            return unit
        if variant in course_code_to_unit:
            unit = course_code_to_unit[variant]
            return unit
    print(f"⚠️ Could not find credit unit for: '{original_code}', defaulting to 2")
    return 2  # Default credit unit


def find_credit_unit_simple(course_code, credit_units_dict):
    """Simplified version of find_credit_unit for internal use."""
    return find_credit_unit(course_code, credit_units_dict, {})


# ============================================================
# CRITICAL FIXES: Mastersheet Reading Functions
# ============================================================
def read_mastersheet_with_flexible_headers(mastersheet_path, sheet_name):
    """FIXED VERSION: Read mastersheet with flexible header detection for BM - FIXED for EXAMS NUMBER"""
    print(f"🔍 FIXED: Reading BM mastersheet with flexible headers...")
    # First, let's examine the actual structure
    xl = pd.ExcelFile(mastersheet_path)
    df_raw = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=None)
    print(f"📊 Raw BM mastersheet shape: {df_raw.shape}")
    print(f"📊 First 10 rows sample:")
    for i in range(min(10, len(df_raw))):
        print(f" Row {i}: {df_raw.iloc[i].dropna().tolist()}")
    # Look for the header row that contains "EXAMS NUMBER" or similar
    header_row_idx = None
    for idx in range(len(df_raw)):
        row_values = df_raw.iloc[idx].dropna().astype(str).str.upper().tolist()
        row_combined = " ".join(row_values)
        # Check for exam number indicators - FIXED for EXAMS NUMBER
        if any(
            keyword in row_combined
            for keyword in ["EXAMS NUMBER", "EXAM NUMBER", "REG NO", "REGISTRATION"]
        ):
            header_row_idx = idx
            print(f"✅ FOUND header row at index {idx}: {row_values}")
            break
    if header_row_idx is None:
        print(f"❌ No header row found with exam number indicators")
        # Try common header row positions
        for idx in [5, 4, 3, 2, 1, 0]:
            try:
                df_test = pd.read_excel(
                    mastersheet_path, sheet_name=sheet_name, header=idx
                )
                if len(df_test.columns) > 3:  # Reasonable number of columns
                    header_row_idx = idx
                    print(f"🔄 Using fallback header row: {idx}")
                    break
            except:
                continue
    if header_row_idx is None:
        print(f"❌ Could not determine header row")
        return None, None
    # Read with the found header row
    try:
        df = pd.read_excel(
            mastersheet_path, sheet_name=sheet_name, header=header_row_idx
        )
        print(f"✅ Successfully read BM mastersheet with header row {header_row_idx}")
        print(f"📊 Columns: {df.columns.tolist()}")
        # Find exam number column - FIXED for EXAMS NUMBER
        exam_col = None
        for col in df.columns:
            col_str = str(col).upper()
            if any(
                keyword in col_str
                for keyword in ["EXAMS NUMBER", "EXAM NUMBER", "REG NO", "REGISTRATION"]
            ):
                exam_col = col
                break
        if not exam_col:
            print(f"❌ No exam number column found in: {df.columns.tolist()}")
            return None, None
        print(f"✅ Exam column found: '{exam_col}'")
        return df, exam_col
    except Exception as e:
        print(f"❌ Error reading BM mastersheet: {e}")
        return None, None


def find_student_in_mastersheet_fixed(exam_no, mastersheet_df, exam_col):
    """FIXED VERSION: Robust student matching in BM mastersheet"""
    if mastersheet_df is None or exam_col not in mastersheet_df.columns:
        return None
    # Clean the exam number for matching
    exam_no_clean = str(exam_no).strip().upper()
    exam_no_clean = re.sub(r"[^A-Z0-9]", "", exam_no_clean)
    # Method 1: Exact match after cleaning
    for idx, row in mastersheet_df.iterrows():
        current_exam = (
            str(row[exam_col]).strip().upper() if pd.notna(row[exam_col]) else ""
        )
        current_exam_clean = re.sub(r"[^A-Z0-9]", "", current_exam)
        if current_exam_clean == exam_no_clean:
            return row
    # Method 2: Partial match
    for idx, row in mastersheet_df.iterrows():
        current_exam = (
            str(row[exam_col]).strip().upper() if pd.notna(row[exam_col]) else ""
        )
        current_exam_clean = re.sub(r"[^A-Z0-9]", "", current_exam)
        if exam_no_clean in current_exam_clean or current_exam_clean in exam_no_clean:
            return row
    # Method 3: Try with different cleaning approaches
    for idx, row in mastersheet_df.iterrows():
        current_exam = (
            str(row[exam_col]).strip().upper() if pd.notna(row[exam_col]) else ""
        )
        # Remove common prefixes/suffixes
        current_clean = re.sub(r"^(BM|MID|MWF)", "", current_exam)
        exam_clean = re.sub(r"^(BM|MID|MWF)", "", exam_no_clean)
        if current_clean == exam_clean:
            return row
    return None


def quick_fix_read_mastersheet(mastersheet_path, sheet_name):
    """QUICK FIX: Read BM mastersheet using header row 5 (which we know works)"""
    try:
        # Force header row 5 which we know contains the correct columns
        df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)
        # Manually find exam column
        exam_col = None
        for col in df.columns:
            if "EXAM" in str(col).upper() or "REG" in str(col).upper():
                exam_col = col
                break
        if exam_col:
            print(f"✅ QUICK FIX: Using forced header row 5, exam column: '{exam_col}'")
            return df, exam_col
        else:
            print(f"❌ QUICK FIX: No exam column found even with header row 5")
            return None, None
    except Exception as e:
        print(f"❌ QUICK FIX Error: {e}")
        return None, None


# ============================================================
# File and ZIP Handling (BM-Compatible)
# ============================================================
def extract_mastersheet_from_zip(zip_path, semester_key):
    """Extract mastersheet from ZIP file and return temporary file path."""
    try:
        print(f"📦 Looking for mastersheet in ZIP: {zip_path}")
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            all_files = zip_ref.namelist()
            print(f"📁 Files in ZIP: {all_files}")
            mastersheet_files = [
                f
                for f in all_files
                if "mastersheet" in f.lower() and f.endswith(".xlsx")
            ]
            if not mastersheet_files:
                print(f"❌ No mastersheet found in ZIP")
                return None, None
            mastersheet_name = mastersheet_files[0]
            print(f"✅ Found mastersheet: {mastersheet_name}")
            temp_dir = tempfile.mkdtemp()
            temp_mastersheet_path = os.path.join(
                temp_dir, f"mastersheet_{semester_key}.xlsx"
            )
            with open(temp_mastersheet_path, "wb") as f:
                f.write(zip_ref.read(mastersheet_name))
            print(f"✅ Extracted mastersheet to: {temp_mastersheet_path}")
            return temp_mastersheet_path, temp_dir
    except Exception as e:
        print(f"❌ Error extracting mastersheet from ZIP: {e}")
        traceback.print_exc()
        return None, None


def find_latest_zip_file(clean_dir):
    """Find the latest ZIP file in clean results directory."""
    print(f"🔍 Looking for BM ZIP files in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ BM clean directory doesn't exist: {clean_dir}")
        return None
    all_files = os.listdir(clean_dir)
    zip_files = []
    for f in all_files:
        if f.lower().endswith(".zip"):
            if "carryover" in f.lower():
                print(f"⚠️ Skipping BM carryover ZIP: {f}")
                continue
            if any(pattern in f for pattern in ["_RESULT-", "RESULT_", "RESULT-"]):
                zip_files.append(f)
                print(f"✅ Found BM regular results ZIP: {f}")
            else:
                print(f"ℹ️ Found other BM ZIP (not a result file): {f}")
    if not zip_files:
        print(f"❌ No BM regular results ZIP files found (excluding carryover files)")
        fallback_zips = [
            f
            for f in all_files
            if f.lower().endswith(".zip") and "carryover" not in f.lower()
        ]
        if fallback_zips:
            print(f"⚠️ Using fallback BM ZIP files: {fallback_zips}")
            zip_files = fallback_zips
        else:
            print(f"❌ No BM ZIP files found at all in {clean_dir}")
            return None
    print(f"✅ Final BM ZIP files to consider: {zip_files}")
    zip_files_with_path = [os.path.join(clean_dir, f) for f in zip_files]
    latest_zip = sorted(zip_files_with_path, key=os.path.getmtime, reverse=True)[0]
    print(f"🎯 Using latest BM ZIP: {latest_zip}")
    return latest_zip


def find_latest_result_folder(clean_dir, set_name):
    """Find the latest result folder in clean results directory."""
    print(f"🔍 Looking for BM result folders in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ BM clean directory doesn't exist: {clean_dir}")
        return None
    all_items = os.listdir(clean_dir)
    result_folders = [
        f
        for f in all_items
        if os.path.isdir(os.path.join(clean_dir, f))
        and f.startswith(f"{set_name}_RESULT-")
    ]
    if not result_folders:
        print(f"❌ No BM result folders found")
        return None
    print(f"✅ Found BM result folders: {result_folders}")
    folders_with_path = [os.path.join(clean_dir, f) for f in result_folders]
    latest_folder = sorted(folders_with_path, key=os.path.getmtime, reverse=True)[0]
    print(f"🎯 Using latest BM result folder: {latest_folder}")
    return latest_folder


def find_latest_mastersheet_source(clean_dir, set_name):
    """Find the latest source for mastersheet: prefer ZIP, fallback to folder."""
    print(f"🔍 Looking for BM mastersheet source in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ BM clean directory doesn't exist: {clean_dir}")
        return None, None
    zip_path = find_latest_zip_file(clean_dir)
    if zip_path:
        print(f"✅ Using BM ZIP source: {zip_path}")
        try:
            with zipfile.ZipFile(zip_path, "r") as zip_ref:
                zip_files = zip_ref.namelist()
                mastersheet_files = [
                    f
                    for f in zip_files
                    if "mastersheet" in f.lower() and f.endswith(".xlsx")
                ]
                if mastersheet_files:
                    print(f"✅ BM ZIP contains mastersheet files: {mastersheet_files}")
                    return zip_path, "zip"
                else:
                    print(f"⚠️ BM ZIP found but no mastersheet inside: {zip_path}")
        except Exception as e:
            print(f"⚠️ Error checking BM ZIP contents: {e}")
    folder_path = find_latest_result_folder(clean_dir, set_name)
    if folder_path:
        print(f"✅ Using BM folder source: {folder_path}")
        return folder_path, "folder"
    print(f"❌ No valid BM ZIP files or result folders found in {clean_dir}")
    return None, None


def get_mastersheet_path(source_path, source_type, semester_key):
    """Get mastersheet path based on source type (zip or folder)."""
    temp_dir = None
    if source_type == "zip":
        temp_mastersheet_path, temp_dir = extract_mastersheet_from_zip(
            source_path, semester_key
        )
        if not temp_mastersheet_path:
            print("❌ Failed to extract mastersheet from ZIP")
            return None, None
    elif source_type == "folder":
        all_files = os.listdir(source_path)
        mastersheet_files = [
            f for f in all_files if "mastersheet" in f.lower() and f.endswith(".xlsx")
        ]
        if not mastersheet_files:
            print(f"❌ No mastersheet found in folder {source_path}")
            return None, None
        mastersheet_name = mastersheet_files[0]
        temp_mastersheet_path = os.path.join(source_path, mastersheet_name)
        print(f"✅ Found mastersheet in folder: {temp_mastersheet_path}")
    else:
        return None, None
    return temp_mastersheet_path, temp_dir


def find_matching_sheet(sheet_names, target_key):
    target_upper = (
        target_key.upper().replace("-", " ").replace("_", " ").replace(".", " ")
    )
    target_upper = " ".join(target_upper.split())
    possible_keys = [
        target_key,
        target_key.upper(),
        target_key.lower(),
        target_key.title(),
        target_key.replace("-", " ").upper(),
        target_key.replace("-", " ").lower(),
        target_key.replace("-", " ").title(),
        target_key.replace("-", "_").upper(),
        target_key.replace("-", "_").lower(),
        target_key.replace("-", "_").title(),
        target_key.replace("First", "1st"),
        target_key.replace("Second", "2nd"),
        target_key.replace("Third", "3rd"),
        target_key.replace("YEAR", "YR"),
        target_key.replace("SEMESTER", "SEM"),
        target_upper,
        target_upper.replace("FIRST", "1st"),
        target_upper.replace("SECOND", "2ND"),
        target_upper.replace("THIRD", "3RD"),
        target_upper.replace("YEAR", "YR"),
        target_upper.replace("SEMESTER", "SEM"),
    ]
    possible_keys = list(set(possible_keys))
    print(f"🔍 Trying sheet variants for '{target_key}': {possible_keys}")
    for sheet in sheet_names:
        sheet_normalized = (
            sheet.upper().replace("-", " ").replace("_", " ").replace(".", " ")
        )
        sheet_normalized = " ".join(sheet_normalized.split())
        if any(
            p == sheet or p in sheet or p == sheet_normalized or p in sheet_normalized
            for p in possible_keys
        ):
            print(f"✅ Found matching sheet: '{sheet}' for '{target_key}'")
            return sheet
    print(f"❌ No matching sheet found for '{target_key}'")
    print(f"📖 Available sheets: {sheet_names}")
    return None


def create_carryover_zip(source_dir, zip_path):
    """Create ZIP file of carryover results."""
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        print(f"✅ ZIP file created: {zip_path}")
        return True
    except Exception as e:
        print(f"❌ Error creating ZIP: {e}")
        return False


# ============================================================
# GPA/CGPA Management (BM-Compatible)
# ============================================================
def load_previous_gpas(mastersheet_path, current_semester_key):
    """Load previous GPA data from mastersheet for BM CGPA calculation - FIXED with flexible headers."""
    all_student_data = {}
    current_standard = standardize_semester_key(current_semester_key)
    # BM semesters
    all_semesters = {
        "M-FIRST-YEAR-FIRST-SEMESTER": [],
        "M-FIRST-YEAR-SECOND-SEMESTER": ["M-FIRST-YEAR-FIRST-SEMESTER"],
        "M-SECOND-YEAR-FIRST-SEMESTER": [
            "M-FIRST-YEAR-FIRST-SEMESTER",
            "M-FIRST-YEAR-SECOND-SEMESTER",
        ],
        "M-SECOND-YEAR-SECOND-SEMESTER": [
            "M-FIRST-YEAR-FIRST-SEMESTER",
            "M-FIRST-YEAR-SECOND-SEMESTER",
            "M-SECOND-YEAR-FIRST-SEMESTER",
        ],
        "M-THIRD-YEAR-FIRST-SEMESTER": [
            "M-FIRST-YEAR-FIRST-SEMESTER",
            "M-FIRST-YEAR-SECOND-SEMESTER",
            "M-SECOND-YEAR-FIRST-SEMESTER",
            "M-SECOND-YEAR-SECOND-SEMESTER",
        ],
        "M-THIRD-YEAR-SECOND-SEMESTER": [
            "M-FIRST-YEAR-FIRST-SEMESTER",
            "M-FIRST-YEAR-SECOND-SEMESTER",
            "M-SECOND-YEAR-FIRST-SEMESTER",
            "M-SECOND-YEAR-SECOND-SEMESTER",
            "M-THIRD-YEAR-FIRST-SEMESTER",
        ],
    }
    semesters_to_load = all_semesters.get(current_standard, [])
    print(f"📊 Loading previous BM GPAs for {current_standard}: {semesters_to_load}")
    if not os.path.exists(mastersheet_path):
        print(f"❌ BM Mastersheet not found: {mastersheet_path}")
        return {}
    try:
        xl = pd.ExcelFile(mastersheet_path)
        print(f"📖 Available sheets in BM mastersheet: {xl.sheet_names}")
    except Exception as e:
        print(f"❌ Error opening BM mastersheet: {e}")
        return {}
    for semester in semesters_to_load:
        try:
            sheet_name = find_matching_sheet(xl.sheet_names, semester)
            if not sheet_name:
                print(f"⚠️ Skipping BM semester {semester} - no matching sheet found")
                continue
            print(f"📖 Reading BM sheet '{sheet_name}' for semester {semester}")
            # FIXED: Use flexible header reading for each previous semester sheet
            df, exam_col = read_mastersheet_with_flexible_headers(
                mastersheet_path, sheet_name
            )
            if df is None or exam_col is None:
                print(f"⚠️ Could not read BM sheet '{sheet_name}' with flexible headers")
                continue
            gpa_col = None
            credit_col = None
            # ENHANCED: Prioritize total attempted credits for CGPA accuracy
            for col in df.columns:
                col_str = str(col).upper()
                if "GPA" in col_str and "CGPA" not in col_str:
                    gpa_col = col
                if (
                    "TCPE" in col_str
                    or "TOTAL CREDIT" in col_str
                    or "TOTAL UNIT" in col_str
                ):
                    credit_col = col  # Prioritize total attempted credits
                elif "CU PASSED" in col_str or "CREDIT" in col_str or "UNIT" in col_str:
                    credit_col = col  # Fallback
            print(
                f"🔍 Columns found - Exam: {exam_col}, GPA: {gpa_col}, Credits: {credit_col}"
            )
            if exam_col and gpa_col:
                for idx, row in df.iterrows():
                    try:
                        exam_no = str(row[exam_col]).strip()
                        if pd.isna(exam_no) or exam_no in ["", "NAN", "NONE"]:
                            continue
                        gpa_value = row[gpa_col]
                        if pd.isna(gpa_value):
                            continue
                        credits = 30
                        if (
                            credit_col
                            and credit_col in row
                            and pd.notna(row[credit_col])
                        ):
                            try:
                                credits = int(float(row[credit_col]))
                            except (ValueError, TypeError):
                                credits = 30
                        if exam_no not in all_student_data:
                            all_student_data[exam_no] = {"gpas": [], "credits": []}
                        all_student_data[exam_no]["gpas"].append(float(gpa_value))
                        all_student_data[exam_no]["credits"].append(credits)
                        if idx < 3:
                            print(
                                f"📊 Loaded BM GPA for {exam_no}: {gpa_value} with {credits} credits"
                            )
                    except (ValueError, TypeError) as e:
                        print(f"⚠️ Error processing row {idx} for BM {semester}: {e}")
                        continue
            else:
                print(
                    f"⚠️ Missing required columns in BM {sheet_name}: exam_col={exam_col}, gpa_col={gpa_col}"
                )
        except Exception as e:
            print(f"⚠️ Could not load data from BM {semester}: {e}")
            traceback.print_exc()
    print(f"📊 Loaded cumulative BM data for {len(all_student_data)} students")
    return all_student_data


def calculate_cgpa(student_data, current_gpa, current_credits):
    """Calculate Cumulative GPA for BM."""
    if not student_data or not student_data.get("gpas"):
        print(f"⚠️ No previous BM GPA data, using current GPA: {current_gpa}")
        return current_gpa
    total_grade_points = 0.0
    total_credits = 0
    print(f"🔢 Calculating BM CGPA from {len(student_data['gpas'])} previous semesters")
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
        print(
            f" - GPA: {prev_gpa}, Credits: {prev_credits}, Running Total: {total_grade_points}/{total_credits}"
        )
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits
    print(f"📊 Final BM calculation: {total_grade_points} / {total_credits}")
    if total_credits > 0:
        cgpa = round(total_grade_points / total_credits, 2)
        print(f"✅ Calculated BM CGPA: {cgpa}")
        return cgpa
    else:
        print(f"⚠️ No BM credits, returning current GPA: {current_gpa}")
        return current_gpa


# ============================================================
# CRITICAL FIX: Enhanced Remarks Calculation
# ============================================================
def calculate_student_remarks(cu_passed, cu_failed, total_credits, gpa, student_had_carryover_update=False):
    """
    CRITICAL FIX: Calculate student remarks based on CURRENT performance.
    Override withdrawn status for students who passed all courses in resit.
    INCLUDES PROBATION status as per NBTE standards.
    
    Args:
        cu_passed: Credits passed
        cu_failed: Credits failed  
        total_credits: Total credits
        gpa: Current GPA
        student_had_carryover_update: Whether student had carryover updates
    
    Returns:
        str: Remarks (PASSED, RESIT, PROBATION, or WITHDRAW)
    """
    # Calculate passed percentage
    passed_percent = cu_passed / total_credits if total_credits > 0 else 0
    
    # CRITICAL FIX: If student has NO failed courses (cu_failed == 0), they PASSED
    # This overrides any previous withdrawn status
    if cu_failed == 0:
        if student_had_carryover_update:
            print(f"  ✅ CARRYOVER STUDENT NOW PASSED: 0 failures, GPA: {gpa}")
        return "PASSED"
    
    # CRITICAL FIX: If student passed less than 45% of courses, WITHDRAW
    # But ONLY if they still have failures after carryover processing
    elif passed_percent < 0.45:
        if student_had_carryover_update:
            print(f"  ⚠️ CARRYOVER STUDENT STILL WITHDRAWN: passed only {passed_percent*100:.1f}%")
        return "WITHDRAW"
    
    # Has failures but passed ≥45%, assign RESIT or PROBATION based on GPA
    else:
        # CRITICAL FIX: Include PROBATION status for students with GPA < 2.0
        remarks = "RESIT" if gpa >= 2.0 else "PROBATION"
        if student_had_carryover_update:
            print(f"  📝 CARRYOVER STUDENT {remarks}: GPA: {gpa}, {cu_failed} CU failed")
        return remarks

# ============================================================
# Mastersheet Update Functions (CRITICAL FIXES - BM-Compatible)
# ============================================================
def find_sheet_structure(ws):
    """FIXED: Find header row - BM compatible with EXAMS NUMBER"""
    header_row = None
    headers = {}

    # BM-specific keywords (EXAMS NUMBER is plural in BM)
    exam_keywords = [
        "EXAMS NUMBER",  # ← CRITICAL: BM uses plural
        "EXAM NUMBER",
        "REG NO",
        "REG. NO",
        "REGISTRATION",
    ]

    # Search first 20 rows for headers
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_text = " ".join(
            [
                str(ws.cell(row_idx, col).value or "").upper()
                for col in range(1, min(15, ws.max_column + 1))
            ]
        )

        # Check if this row contains exam number indicators
        if any(keyword in row_text for keyword in exam_keywords):
            header_row = row_idx
            print(f"✅ FOUND BM header row at: {row_idx}")

            # Build headers dictionary with both original and uppercase keys
            for col_idx in range(1, ws.max_column + 1):
                header_val = ws.cell(row=header_row, column=col_idx).value
                if header_val:
                    header_clean = str(header_val).strip()
                    headers[header_clean] = col_idx
                    headers[header_clean.upper()] = col_idx

            print(f"📋 Found {len(set(headers.values()))} unique column headers")
            return header_row, headers

    print(f"❌ No BM header row found")
    return None, {}


def emergency_find_header_row(ws, semester_key):
    """EMERGENCY FALLBACK: Force-find header row for BM sheets"""
    print(f"🚨 EMERGENCY: Force-searching BM header for {semester_key}")
    # BM sheets typically have headers at row 5, 6, or 7
    for forced_row in [5, 6, 7, 4, 8]:
        if forced_row > ws.max_row:
            continue
        headers = {}
        # Build headers from this forced row
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=forced_row, column=col_idx).value
            if header_val:
                header_clean = str(header_val).strip()
                headers[header_clean] = col_idx
                headers[header_clean.upper()] = col_idx
        if headers:
            print(f"🚨 USING EMERGENCY HEADER ROW: {forced_row}")
            print(f"🚨 Headers: {list(headers.keys())[:8]}...")
            return forced_row, headers
    print(f"❌ EMERGENCY: No suitable header row found")
    return None, {}


def apply_student_sorting(ws, header_row, headers_dict):
    """Apply sorting to students - compatibility function for ANALYSIS sheet"""
    apply_student_sorting_with_serial_numbers(ws, header_row, headers_dict)


def apply_student_sorting_with_serial_numbers(ws, header_row, headers_dict):
    """Apply sorting to students with PROPER serial numbers - BM VERSION"""
    from openpyxl.styles import PatternFill, Font, Border

    exam_col = headers_dict.get("EXAMS NUMBER") or headers_dict.get("EXAM NUMBER")
    remarks_col = headers_dict.get("REMARKS")
    gpa_col = headers_dict.get("GPA")
    serial_col = 1  # Serial number is always column 1
    if not all([exam_col, remarks_col, gpa_col]):
        return
    # Collect all student rows
    student_rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        exam_no = ws.cell(row, exam_col).value
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break
        # Get basic row data (values only, not styles)
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            row_data.append({"value": cell.value, "number_format": cell.number_format})
        remarks = ws.cell(row, remarks_col).value or ""
        remarks_upper = str(remarks).upper()
        # Assign priority for sorting
        if "WITHDRAW" in remarks_upper:
            priority = 3  # Withdrawn - last
        elif (
            "RESIT" in remarks_upper
            or "CARRYOVER" in remarks_upper
            or "PROBATION" in remarks_upper
        ):
            priority = 2  # Carryover - middle
        else:
            priority = 1  # Passed - first
        student_rows.append(
            {
                "row_num": row,
                "priority": priority,
                "gpa": ws.cell(row, gpa_col).value or 0,
                "row_data": row_data,
                "remarks": remarks_upper,
                "exam_no": exam_no,  # Store exam number for secondary sorting
            }
        )
    # Sort by priority (Passed first, then Carryover, then Withdrawn)
    # Within each group, sort by GPA (descending), then by exam number
    student_rows.sort(
        key=lambda x: (x["priority"], -float(x["gpa"] if x["gpa"] else 0), x["exam_no"])
    )
    # Write sorted data back to worksheet
    for idx, student in enumerate(student_rows):
        target_row = header_row + 1 + idx
        row_data = student["row_data"]
        for col_idx, cell_data in enumerate(row_data, 1):
            cell = ws.cell(target_row, col_idx)
            cell.value = cell_data["value"]
            # Apply number format if present
            if cell_data.get("number_format"):
                cell.number_format = cell_data["number_format"]
            # CRITICAL FIX: Update serial numbers to be consecutive
            if col_idx == serial_col:  # Serial number column
                cell.value = idx + 1
            # Re-apply styling based on content and student status
            if student["priority"] == 3:  # Withdrawn
                cell.fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )
                if "WITHDRAW" in student["remarks"]:
                    cell.font = Font(bold=True, color="FF0000")
            elif idx % 2 == 0:  # Alternate row coloring
                cell.fill = PatternFill(
                    start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
                )
            # Apply GPA styling for GPA column
            if col_idx == headers_dict.get("GPA"):
                try:
                    gpa_val = float(cell.value) if cell.value else 0
                    if gpa_val >= 3.5:
                        cell.font = Font(
                            bold=True, color="006400"
                        )  # Dark green for high GPA
                    elif gpa_val < 2.0:
                        cell.font = Font(bold=True, color="FF0000")  # Red for low GPA
                except (ValueError, TypeError):
                    pass
    print(
        f" ✅ Applied student sorting with proper serial numbers (1 to {len(student_rows)})"
    )


def identify_course_columns_properly(headers):
    """Identify BM course columns with comprehensive pattern matching"""
    import re

    course_columns = {}

    # BM course patterns: MID101, MWF201, GNS111, etc.
    patterns = [
        r"^[A-Z]{3}\d{3}$",  # Standard: 3 letters + 3 digits
        r"^[A-Z]{2,4}\s*\d{3}$",  # With optional space
    ]

    for header, col_idx in headers.items():
        header_clean = str(header).strip().upper()

        for pattern in patterns:
            if re.match(pattern, header_clean):
                # Store multiple variants for robust matching
                normalized = re.sub(r"[^A-Z0-9]", "", header_clean)

                course_columns[header] = col_idx
                course_columns[header_clean] = col_idx
                course_columns[normalized] = col_idx

                print(f"✅ BM Course: '{header}' -> column {col_idx}")
                break

    print(f"📊 Total BM course columns: {len(set(course_columns.values()))}")
    return course_columns


def update_summary_section_fixed(ws, headers, header_row, course_columns):
    """BM VERSION - Counts failures from CURRENT worksheet scores"""

    print(f" 📊 Updating BM summary section...")

    try:
        # Find SUMMARY section
        summary_start_row = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "SUMMARY" in str(cell_value).upper():
                summary_start_row = row_idx
                break

        if not summary_start_row:
            print(" ℹ️ No SUMMARY section found")
            return

        # Find exam column
        exam_col_idx = None
        for col_name, col_idx in headers.items():
            if "EXAM" in col_name.upper() and "NUMBER" in col_name.upper():
                exam_col_idx = col_idx
                break

        if not exam_col_idx:
            print(" ❌ No exam column found")
            return

        # ═══════════════════════════════════════════════════════
        # CRITICAL: Count from CURRENT worksheet values
        # ═══════════════════════════════════════════════════════

        # Re-identify course columns from CURRENT headers
        fresh_headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                fresh_headers[str(header_val).strip()] = col_idx

        fresh_course_columns = identify_course_columns_properly(fresh_headers)

        # Initialize counters
        total_students = 0
        passed_students = 0
        resit_students = 0
        probation_students = 0
        withdrawn_students = 0
        course_failures = {course: 0 for course in fresh_course_columns}

        # Count from CURRENT data
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no = ws.cell(row=row_idx, column=exam_col_idx).value

            if not exam_no or "SUMMARY" in str(exam_no).upper():
                break

            if str(exam_no).strip() in ["", "NAN", "NONE"]:
                continue

            total_students += 1

            # Count remarks from CURRENT values
            for col_name, col_idx in headers.items():
                if "REMARKS" in col_name.upper():
                    remarks = ws.cell(row=row_idx, column=col_idx).value
                    if remarks:
                        remarks_upper = str(remarks).upper()
                        if "PASSED" in remarks_upper:
                            passed_students += 1
                        elif "RESIT" in remarks_upper or "CARRYOVER" in remarks_upper:
                            resit_students += 1
                        elif "PROBATION" in remarks_upper:
                            probation_students += 1
                        elif "WITHDRAW" in remarks_upper:
                            withdrawn_students += 1
                    break

            # ═══════════════════════════════════════════════════════
            # CRITICAL: Count failures from CURRENT scores (including updated resit scores)
            # ═══════════════════════════════════════════════════════
            for course in fresh_course_columns:
                if course in fresh_headers:
                    col_idx = fresh_headers[course]
                    score = ws.cell(row=row_idx, column=col_idx).value
                    if score is not None and score != "":
                        try:
                            if float(score) < 50:
                                course_failures[course] += 1
                        except (ValueError, TypeError):
                            continue

        # ═══════════════════════════════════════════════════════
        # Update "Fails Per Course" row (aligned to course columns)
        # ═══════════════════════════════════════════════════════
        fails_row = find_fails_per_course_row(ws, header_row)

        if fails_row:
            print(f"✅ Found fails per course row at: {fails_row}")
            sorted_courses = sorted(
                fresh_course_columns, key=lambda k: fresh_course_columns[k]
            )

            for course in sorted_courses:
                col = fresh_course_columns[course]
                ws.cell(row=fails_row, column=col).value = course_failures[course]
                print(f" ✅ {course}: {course_failures[course]} failures")

        # ═══════════════════════════════════════════════════════
        # Update summary text with embedded numbers
        # ═══════════════════════════════════════════════════════
        update_summary_text_with_counts(
            ws,
            summary_start_row,
            total_students,
            passed_students,
            resit_students,
            probation_students,
            withdrawn_students,
        )

        print(f" ✅ BM summary section updated with current data")

    except Exception as e:
        print(f" ❌ Error updating BM summary: {e}")
        traceback.print_exc()


def ensure_required_sheets_exist(wb):
    """Ensure CGPA_SUMMARY and ANALYSIS sheets exist in workbook"""
    from openpyxl.styles import Font, Alignment, PatternFill

    print(f"\n🔍 CHECKING FOR REQUIRED SHEETS...")
    print(f" Current sheets: {wb.sheetnames}")
    # Create CGPA_SUMMARY if it doesn't exist
    if "CGPA_SUMMARY" not in wb.sheetnames:
        print(f"⚠️ CGPA_SUMMARY sheet not found - creating it...")
        cgpa_ws = wb.create_sheet("CGPA_SUMMARY")
        # Add basic structure
        cgpa_ws["A1"] = "CGPA SUMMARY"
        cgpa_ws["A1"].font = Font(bold=True, size=14)
        cgpa_ws["A1"].alignment = Alignment(horizontal="center")
        print(f"✅ Created CGPA_SUMMARY sheet")
    else:
        print(f"✅ CGPA_SUMMARY sheet exists")
    # Create ANALYSIS if it doesn't exist
    if "ANALYSIS" not in wb.sheetnames:
        print(f"⚠️ ANALYSIS sheet not found - creating it...")
        analysis_ws = wb.create_sheet("ANALYSIS")
        # Add basic structure
        analysis_ws["A1"] = "PERFORMANCE ANALYSIS"
        analysis_ws["A1"].font = Font(bold=True, size=14)
        analysis_ws["A1"].alignment = Alignment(horizontal="center")
        print(f"✅ Created ANALYSIS sheet")
    else:
        print(f"✅ ANALYSIS sheet exists")
    print(f"✅ All required sheets verified")
    return True


def clear_sheet_completely(ws):
    """Clear all data and formatting from sheet"""
    from openpyxl.styles import PatternFill, Font, Border

    # Unmerge all merged cells
    try:
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
    except:
        pass

    # Clear all cells
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()


def find_summary_columns(headers):
    """Find all summary-related columns"""
    summary_columns = {}
    summary_mapping = {
        "FAILED COURSES": ["FAILED COURSES", "FAILED COURSE", "COURSES FAILED"],
        "REMARKS": ["REMARKS", "REMARK", "STATUS"],
        "CU Passed": ["CU PASSED", "CREDIT PASSED", "UNITS PASSED", "CUP"],
        "CU Failed": ["CU FAILED", "CREDIT FAILED", "UNITS FAILED", "CUF"],
        "TCPE": ["TCPE", "TOTAL CREDIT", "TOTAL CREDITS", "TCP"],
        "GPA": ["GPA", "GRADE POINT"],
        "AVERAGE": ["AVERAGE", "AVG", "MEAN"],
        "CGPA": ["CGPA", "CUMULATIVE GPA"],
    }

    for key, keywords in summary_mapping.items():
        for header, col_idx in headers.items():
            header_upper = header.upper()
            if any(kw in header_upper for kw in keywords):
                if key == "GPA" and "CGPA" in header_upper:
                    continue
                summary_columns[key] = col_idx
                break

    return summary_columns


def find_fails_per_course_row(ws, header_row):
    """Find the 'Fails Per Course' row"""
    for row_idx in range(header_row + 1, ws.max_row + 1):
        for col_idx in [1, 2, 3]:
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and "FAILS PER COURSE:" in str(cell_value).upper():
                return row_idx
    return None


def update_summary_text_with_counts(
    ws,
    summary_start_row,
    total_students,
    passed_students,
    resit_students,
    probation_students,
    withdrawn_students,
):
    """Update summary text with current counts"""
    current_row = summary_start_row + 1
    while current_row <= ws.max_row:
        cell_value = ws.cell(row=current_row, column=1).value
        if not cell_value:
            break
        cell_str = str(cell_value).upper()
        if "REGISTERED AND SAT" in cell_str:
            new_value = re.sub(
                r"A TOTAL OF \d+ STUDENTS",
                f"A total of {total_students} students",
                cell_value,
                flags=re.I,
            )
            ws.cell(row=current_row, column=1).value = new_value
            print(f" ✅ Updated total students: {total_students}")
        elif (
            "PASSED IN ALL COURSES REGISTERED" in cell_str and "FAILED" not in cell_str
        ):
            new_value = re.sub(
                r"A TOTAL OF \d+ STUDENTS",
                f"A total of {passed_students} students",
                cell_value,
                flags=re.I,
            )
            ws.cell(row=current_row, column=1).value = new_value
            print(f" ✅ Updated passed: {passed_students}")
        elif "GRADE POINT AVERAGE (GPA) OF 2.00 AND ABOVE FAILED" in cell_str:
            new_value = re.sub(
                r"A TOTAL OF \d+ STUDENTS",
                f"A total of {resit_students} students",
                cell_value,
                flags=re.I,
            )
            ws.cell(row=current_row, column=1).value = new_value
            print(f" ✅ Updated resit (GPA >=2.00): {resit_students}")
        elif "GRADE POINT AVERAGE (GPA) BELOW 2.00 FAILED" in cell_str:
            new_value = re.sub(
                r"A TOTAL OF \d+ STUDENTS",
                f"A total of {probation_students} students",
                cell_value,
                flags=re.I,
            )
            ws.cell(row=current_row, column=1).value = new_value
            print(f" ✅ Updated probation: {probation_students}")
        elif "FAILED IN MORE THAN 45%" in cell_str:
            new_value = re.sub(
                r"A TOTAL OF \d+ STUDENTS",
                f"A total of {withdrawn_students} students",
                cell_value,
                flags=re.I,
            )
            ws.cell(row=current_row, column=1).value = new_value
            print(f" ✅ Updated withdrawn: {withdrawn_students}")
        current_row += 1


def update_cgpa_summary_sheet_fixed(wb, semester_key, header_row, set_name):
    """BM VERSION - Reads CURRENT data from ALL updated sheets with professional formatting"""
    print(f" 📈 Updating BM CGPA_SUMMARY with professional formatting...")

    if "CGPA_SUMMARY" not in wb.sheetnames:
        print(" ❌ CGPA_SUMMARY sheet missing")
        return

    cgpa_ws = wb["CGPA_SUMMARY"]

    # Clear old data completely
    clear_sheet_completely(cgpa_ws)

    # Create professional headers
    create_professional_headers_bm(cgpa_ws, set_name, semester_key)

    # Collect data from all semester sheets
    semester_keys = [
        "M-FIRST-YEAR-FIRST-SEMESTER",
        "M-FIRST-YEAR-SECOND-SEMESTER",
        "M-SECOND-YEAR-FIRST-SEMESTER",
        "M-SECOND-YEAR-SECOND-SEMESTER",
        "M-THIRD-YEAR-FIRST-SEMESTER",
        "M-THIRD-YEAR-SECOND-SEMESTER",
    ]

    # Track withdrawn students across ALL semesters
    all_withdrawn_students = set()

    # FIRST PASS: Identify withdrawn students from CURRENT data
    for key in semester_keys:
        sheet_name = find_matching_sheet(wb.sheetnames, key)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        header_row_found, headers_dict = find_sheet_structure(ws)
        if not header_row_found:
            continue

        exam_col = headers_dict.get("EXAM NUMBER") or headers_dict.get("EXAMS NUMBER")
        remarks_col = headers_dict.get("REMARKS")

        if not all([exam_col, remarks_col]):
            continue

        # Read CURRENT remarks from worksheet
        for row in range(header_row_found + 1, ws.max_row + 1):
            exam_no = ws.cell(row, exam_col).value
            if not exam_no or "SUMMARY" in str(exam_no).upper():
                break

            # Read CURRENT remarks (includes any updates)
            remarks = ws.cell(row, remarks_col).value or ""
            if "WITHDRAW" in str(remarks).upper():
                all_withdrawn_students.add(str(exam_no).strip().upper())

    print(f" ✅ Found {len(all_withdrawn_students)} withdrawn BM students")

    # SECOND PASS: Collect semester data with CURRENT values
    semester_data = {}
    for key in semester_keys:
        sheet_name = find_matching_sheet(wb.sheetnames, key)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        header_row_found, headers_dict = find_sheet_structure(ws)
        if not header_row_found:
            continue

        exam_col = headers_dict.get("EXAM NUMBER") or headers_dict.get("EXAMS NUMBER")
        name_col = headers_dict.get("NAME")
        gpa_col = headers_dict.get("GPA")
        credits_col = headers_dict.get("TCPE")

        if not all([exam_col, name_col, gpa_col]):
            continue

        data = {}
        for row in range(header_row_found + 1, ws.max_row + 1):
            exam_no_cell = ws.cell(row, exam_col).value
            if not exam_no_cell or "SUMMARY" in str(exam_no_cell).upper():
                break

            exam_no = str(exam_no_cell).strip().upper()

            # Read CURRENT values from worksheet
            name = ws.cell(row, name_col).value
            gpa_val = ws.cell(row, gpa_col).value
            credits_val = ws.cell(row, credits_col).value if credits_col else 0

            is_withdrawn = exam_no in all_withdrawn_students

            try:
                data[exam_no] = {
                    "name": name,
                    "gpa": float(gpa_val) if gpa_val else 0,
                    "credits": float(credits_val) if credits_val else 0,
                    "withdrawn": is_withdrawn,
                }
            except (ValueError, TypeError):
                continue

        semester_data[key] = data
        print(f" ✅ Collected data for {len(data)} students in {key}")

    # Compile and write student data with professional formatting
    write_cgpa_summary_data_with_formatting(
        cgpa_ws, semester_data, all_withdrawn_students
    )

    print(f" ✅ BM CGPA_SUMMARY professionally formatted with current data")


def write_cgpa_summary_data_with_formatting(cgpa_ws, semester_data, all_withdrawn_students):
    """Write CGPA summary data with professional formatting and abbreviated semester columns"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Collect unique students
    all_exam_no = set()
    for semester_dict in semester_data.values():
        all_exam_no.update(semester_dict.keys())

    students = []
    for exam_no in all_exam_no:
        total_gp = 0.0
        total_cr = 0.0
        gpas = {}
        name = None
        withdrawn = exam_no in all_withdrawn_students

        for key, semester_dict in semester_data.items():
            student_data = semester_dict.get(exam_no)

            if student_data:
                gpas[key] = student_data.get("gpa", 0.0)
                total_gp += student_data.get("gpa", 0.0) * student_data.get("credits", 0.0)
                total_cr += student_data.get("credits", 0.0)

                if not name:
                    name = student_data.get("name", "Unknown")

        cgpa = round(total_gp / total_cr, 2) if total_cr > 0 else 0.0

        students.append(
            {
                "exam_no": exam_no,
                "name": name,
                "gpas": gpas,
                "cgpa": cgpa,
                "withdrawn": withdrawn,
            }
        )

    # Sort students (non-withdrawn first, then by CGPA descending)
    non_withdrawn = [s for s in students if not s["withdrawn"]]
    withdrawn_list = [s for s in students if s["withdrawn"]]

    non_withdrawn.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    withdrawn_list.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))

    sorted_students = non_withdrawn + withdrawn_list

    # Define styles
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    even_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    withdrawn_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    excellent_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    good_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

    # Write data starting from row 7
    start_row = 7
    for idx, student in enumerate(sorted_students, start_row):
        # Determine row styling
        is_even_row = (idx - start_row) % 2 == 0
        base_fill = even_row_fill if is_even_row else PatternFill()

        if student["withdrawn"]:
            row_fill = withdrawn_fill
            status_text = "WITHDRAWN"
            status_color = "FF0000"
        else:
            row_fill = base_fill
            status_text = "ACTIVE"
            status_color = "006400"

        # Apply special fills for high performers
        if student["cgpa"] >= 4.0 and not student["withdrawn"]:
            row_fill = excellent_fill
        elif student["cgpa"] >= 3.5 and not student["withdrawn"]:
            row_fill = good_fill

        # Serial number
        cell = cgpa_ws.cell(row=idx, column=1, value=idx - 6)
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Exam number
        cell = cgpa_ws.cell(row=idx, column=2, value=student["exam_no"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Name
        cell = cgpa_ws.cell(row=idx, column=3, value=student["name"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # GPA for each semester - USING ABBREVIATED SEMESTER KEYS
        col = 4
        semester_mapping = {
            "M-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
            "M-FIRST-YEAR-SECOND-SEMESTER": "Y1S2", 
            "M-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
            "M-SECOND-YEAR-SECOND-SEMESTER": "Y2S2",
            "M-THIRD-YEAR-FIRST-SEMESTER": "Y3S1",
            "M-THIRD-YEAR-SECOND-SEMESTER": "Y3S2"
        }
        
        semester_keys_ordered = list(semester_mapping.keys())

        for semester_key in semester_keys_ordered:
            gpa_value = student["gpas"].get(semester_key, "")
            cell = cgpa_ws.cell(row=idx, column=col, value=gpa_value)
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"

            # Color code GPA values
            if gpa_value and isinstance(gpa_value, (int, float)):
                if gpa_value >= 4.0:
                    cell.font = Font(bold=True, color="006100")  # Excellent
                elif gpa_value >= 3.5:
                    cell.font = Font(bold=True, color="00B050")  # Very Good
                elif gpa_value >= 3.0:
                    cell.font = Font(bold=True, color="92D050")  # Good
                elif gpa_value >= 2.5:
                    cell.font = Font(bold=True, color="FFC000")  # Average
                elif gpa_value >= 2.0:
                    cell.font = Font(bold=True, color="FF6600")  # Below Average
                else:
                    cell.font = Font(bold=True, color="FF0000")  # Poor

            col += 1

        # CGPA
        cell = cgpa_ws.cell(row=idx, column=10, value=student["cgpa"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"

        # Color code CGPA
        if student["cgpa"] >= 4.0:
            cell.font = Font(bold=True, color="006100", size=11)  # First Class
        elif student["cgpa"] >= 3.5:
            cell.font = Font(bold=True, color="00B050", size=11)  # Second Class Upper
        elif student["cgpa"] >= 3.0:
            cell.font = Font(bold=True, color="92D050", size=11)  # Second Class Lower
        elif student["cgpa"] >= 2.5:
            cell.font = Font(bold=True, color="FFC000", size=11)  # Third Class
        elif student["cgpa"] >= 2.0:
            cell.font = Font(bold=True, color="FF6600", size=11)  # Pass
        else:
            cell.font = Font(bold=True, color="FF0000", size=11)  # Fail

        # Status
        cell = cgpa_ws.cell(row=idx, column=11, value=status_text)
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color=status_color)

    # Add summary statistics
    summary_row = start_row + len(sorted_students) + 2

    if sorted_students:
        active_students = [s for s in sorted_students if not s["withdrawn"]]
        if active_students:
            avg_cgpa = sum(s["cgpa"] for s in active_students) / len(active_students)
            max_cgpa = max(s["cgpa"] for s in active_students)
            min_cgpa = min(s["cgpa"] for s in active_students)

            summary_data = [
                ["SUMMARY STATISTICS:", ""],
                [f"Total Students: {len(sorted_students)}", f"Active: {len(active_students)}"],
                [f"Withdrawn: {len(withdrawn_list)}", f"Withdrawn Rate: {len(withdrawn_list)/len(sorted_students)*100:.1f}%"],
                [f"Average CGPA: {avg_cgpa:.2f}", f"Maximum CGPA: {max_cgpa:.2f}"],
                [f"Minimum CGPA: {min_cgpa:.2f}", ""],
            ]

            for i, row_data in enumerate(summary_data):
                for j, value in enumerate(row_data):
                    cell = cgpa_ws.cell(row=summary_row + i, column=1 + j, value=value)
                    if i == 0:
                        cell.font = Font(bold=True, size=12, color="1F4E78")
                    else:
                        cell.font = Font(bold=True, size=10)

    print(f"✅ CGPA summary data written for {len(sorted_students)} students with abbreviated semester columns")
    return len(sorted_students)

def create_professional_headers_bm(cgpa_ws, set_name, semester_key):
    """Create professional headers for BM CGPA summary with enhanced formatting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Clear any existing content
    clear_sheet_completely(cgpa_ws)

    # Get display information
    class_name = f"BM {set_name}"
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )

    # Calculate total columns needed (11 columns for BM: 6 semesters + basic info)
    total_columns = 11
    last_column = get_column_letter(total_columns)

    # Define professional color scheme
    header_purple = "7D3C98"  # Purple for main headers (BM specific)
    accent_purple = "E8DAEF"  # Light purple for accents
    white = "FFFFFF"
    light_gray = "F8F8F8"

    # Define borders
    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Row 1: Institution Name (Merged and Centered)
    cgpa_ws.merge_cells(f"A1:{last_column}1")
    title_cell = cgpa_ws["A1"]
    title_cell.value = "COLLEGE OF NURSING SCIENCES, GWAGWALADA"
    title_cell.font = Font(bold=True, size=16, name="Calibri", color=white)
    title_cell.fill = PatternFill(
        start_color=header_purple, end_color=header_purple, fill_type="solid"
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thick_border
    cgpa_ws.row_dimensions[1].height = 30

    # Row 2: Department (Merged and Centered)
    cgpa_ws.merge_cells(f"A2:{last_column}2")
    dept_cell = cgpa_ws["A2"]
    dept_cell.value = "DEPARTMENT OF MIDWIFERY"
    dept_cell.font = Font(bold=True, size=14, name="Calibri", color=white)
    dept_cell.fill = PatternFill(
        start_color=header_purple, end_color=header_purple, fill_type="solid"
    )
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")
    dept_cell.border = thick_border
    cgpa_ws.row_dimensions[2].height = 25

    # Row 3: Class and Sheet Title (Merged and Centered)
    cgpa_ws.merge_cells(f"A3:{last_column}3")
    class_cell = cgpa_ws["A3"]
    class_cell.value = f"{set_name} CLASS - CUMULATIVE GPA SUMMARY"
    class_cell.font = Font(bold=True, size=14, name="Calibri", color=white)
    class_cell.fill = PatternFill(
        start_color=accent_purple, end_color=accent_purple, fill_type="solid"
    )
    class_cell.alignment = Alignment(horizontal="center", vertical="center")
    class_cell.border = thick_border
    cgpa_ws.row_dimensions[3].height = 25

    # Row 4: Date and Academic Session (Merged and Centered)
    cgpa_ws.merge_cells(f"A4:{last_column}4")
    date_cell = cgpa_ws["A4"]
    current_year = datetime.now().year
    date_cell.value = f"{current_year}/{current_year + 1} Academic Session - Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"
    date_cell.font = Font(size=11, name="Calibri", italic=True, color="4A235A")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    cgpa_ws.row_dimensions[4].height = 20

    # Row 5: Empty spacer
    cgpa_ws.row_dimensions[5].height = 5

    # Row 6: Column Headers - USING ABBREVIATED SEMESTER NAMES
    headers = [
        "S/N",
        "EXAM NUMBER",
        "NAME",
        "Y1S1",  # M-FIRST-YEAR-FIRST-SEMESTER
        "Y1S2",  # M-FIRST-YEAR-SECOND-SEMESTER  
        "Y2S1",  # M-SECOND-YEAR-FIRST-SEMESTER
        "Y2S2",  # M-SECOND-YEAR-SECOND-SEMESTER
        "Y3S1",  # M-THIRD-YEAR-FIRST-SEMESTER
        "Y3S2",  # M-THIRD-YEAR-SECOND-SEMESTER
        "CGPA",
        "STATUS",
    ]

    header_fill = PatternFill(
        start_color=header_purple, end_color=header_purple, fill_type="solid"
    )
    header_font = Font(color=white, bold=True, size=11, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = cgpa_ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        # Set column widths
        if col_idx == 1:  # S/N
            cgpa_ws.column_dimensions[get_column_letter(col_idx)].width = 8
        elif col_idx == 2:  # EXAM NUMBER
            cgpa_ws.column_dimensions[get_column_letter(col_idx)].width = 20
        elif col_idx == 3:  # NAME
            cgpa_ws.column_dimensions[get_column_letter(col_idx)].width = 35
        elif col_idx in [4, 5, 6, 7, 8, 9]:  # Semester GPAs (abbreviated)
            cgpa_ws.column_dimensions[get_column_letter(col_idx)].width = 10
        elif col_idx == 10:  # CGPA
            cgpa_ws.column_dimensions[get_column_letter(col_idx)].width = 10
        elif col_idx == 11:  # STATUS
            cgpa_ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Set row height for header row
    cgpa_ws.row_dimensions[6].height = 25

    print(f"✅ Created professional BM CGPA headers for {class_name} with abbreviated semester columns")


def create_analysis_headers_bm(analysis_ws, set_name, semester_key):
    """Create professional headers for BM Analysis sheet with abbreviated columns"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Clear any existing content
    clear_sheet_completely(analysis_ws)

    # Get display information
    class_name = f"BM {set_name}"
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )

    # Total columns for analysis
    total_columns = 7
    last_column = get_column_letter(total_columns)

    # Define professional color scheme
    header_purple = "7D3C98"
    accent_purple = "E8DAEF"
    white = "FFFFFF"

    # Define borders
    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Row 1: Institution Name
    analysis_ws.merge_cells(f"A1:{last_column}1")
    title_cell = analysis_ws["A1"]
    title_cell.value = "COLLEGE OF NURSING SCIENCES, GWAGWALADA"
    title_cell.font = Font(bold=True, size=16, name="Calibri", color=white)
    title_cell.fill = PatternFill(
        start_color=header_purple, end_color=header_purple, fill_type="solid"
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thick_border
    analysis_ws.row_dimensions[1].height = 30

    # Row 2: Department
    analysis_ws.merge_cells(f"A2:{last_column}2")
    dept_cell = analysis_ws["A2"]
    dept_cell.value = "DEPARTMENT OF MIDWIFERY"
    dept_cell.font = Font(bold=True, size=14, name="Calibri", color=white)
    dept_cell.fill = PatternFill(
        start_color=header_purple, end_color=header_purple, fill_type="solid"
    )
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")
    dept_cell.border = thick_border
    analysis_ws.row_dimensions[2].height = 25

    # Row 3: Analysis Title
    analysis_ws.merge_cells(f"A3:{last_column}3")
    analysis_cell = analysis_ws["A3"]
    analysis_cell.value = f"{set_name} CLASS - ACADEMIC PERFORMANCE ANALYSIS"
    analysis_cell.font = Font(bold=True, size=14, name="Calibri", color=white)
    analysis_cell.fill = PatternFill(
        start_color=accent_purple, end_color=accent_purple, fill_type="solid"
    )
    analysis_cell.alignment = Alignment(horizontal="center", vertical="center")
    analysis_cell.border = thick_border
    analysis_ws.row_dimensions[3].height = 25

    # Row 4: Date and Session
    analysis_ws.merge_cells(f"A4:{last_column}4")
    date_cell = analysis_ws["A4"]
    current_year = datetime.now().year
    date_cell.value = f"{current_year}/{current_year + 1} Academic Session - Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"
    date_cell.font = Font(size=11, name="Calibri", italic=True, color="4A235A")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    analysis_ws.row_dimensions[4].height = 20

    # Row 5: Empty spacer
    analysis_ws.row_dimensions[5].height = 5

    # Row 6: Column Headers - USING ABBREVIATED COLUMN NAMES
    headers = [
        "SEMESTER",
        "TOTAL",      # Was "TOTAL STUDENTS"
        "PASSED",     # Was "PASSED" 
        "RESIT",      # Was "CARRYOVER"
        "W/D",        # Was "WITHDRAWN" - abbreviated
        "AVG GPA",    # Was "AVG GPA"
        "PASS %",     # Was "PASS RATE (%)" - abbreviated
    ]

    header_fill = PatternFill(
        start_color=header_purple, end_color=header_purple, fill_type="solid"
    )
    header_font = Font(color=white, bold=True, size=11, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = analysis_ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        # Set column widths
        if col_idx == 1:  # SEMESTER
            analysis_ws.column_dimensions[get_column_letter(col_idx)].width = 20
        elif col_idx == 2:  # TOTAL (abbreviated)
            analysis_ws.column_dimensions[get_column_letter(col_idx)].width = 12
        elif col_idx in [3, 4, 5]:  # PASSED, RESIT, W/D (abbreviated)
            analysis_ws.column_dimensions[get_column_letter(col_idx)].width = 10
        elif col_idx == 6:  # AVG GPA
            analysis_ws.column_dimensions[get_column_letter(col_idx)].width = 10
        elif col_idx == 7:  # PASS % (abbreviated)
            analysis_ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Set row height for header row
    analysis_ws.row_dimensions[6].height = 25

    print(f"✅ Created professional BM Analysis headers for {class_name} with abbreviated columns")


def write_analysis_data_bm(analysis_ws, semester_stats, overall_stats):
    """Write BM analysis data with professional formatting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Define styles
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Color schemes
    even_row_fill = PatternFill(
        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
    )
    passed_purple = PatternFill(
        start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"
    )
    carryover_yellow = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    withdrawn_red = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    overall_fill = PatternFill(
        start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"
    )
    overall_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # BM semester mapping with display names
    semester_display_names = {
        "M-FIRST-YEAR-FIRST-SEMESTER": "Year 1 - Semester 1",
        "M-FIRST-YEAR-SECOND-SEMESTER": "Year 1 - Semester 2",
        "M-SECOND-YEAR-FIRST-SEMESTER": "Year 2 - Semester 1",
        "M-SECOND-YEAR-SECOND-SEMESTER": "Year 2 - Semester 2",
        "M-THIRD-YEAR-FIRST-SEMESTER": "Year 3 - Semester 1",
        "M-THIRD-YEAR-SECOND-SEMESTER": "Year 3 - Semester 2",
    }

    semester_keys = list(semester_display_names.keys())

    # Start writing data from row 7
    current_row = 7

    for idx, key in enumerate(semester_keys):
        if key in semester_stats:
            stats = semester_stats[key]

            # Determine row fill for alternating colors
            row_fill = even_row_fill if idx % 2 == 0 else PatternFill()

            # Semester name
            cell = analysis_ws.cell(current_row, 1, value=semester_display_names[key])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(bold=True)

            # Total students
            cell = analysis_ws.cell(current_row, 2, value=stats["total"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Passed (purple background for BM)
            cell = analysis_ws.cell(current_row, 3, value=stats["passed"])
            cell.border = data_border
            cell.fill = passed_purple if stats["passed"] > 0 else row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if stats["passed"] > 0:
                cell.font = Font(bold=True, color="4A235A")

            # Carryover (yellow background for warning)
            cell = analysis_ws.cell(current_row, 4, value=stats["carryover"])
            cell.border = data_border
            cell.fill = carryover_yellow if stats["carryover"] > 0 else row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if stats["carryover"] > 0:
                cell.font = Font(bold=True, color="9C5700")

            # Withdrawn (red background for critical)
            cell = analysis_ws.cell(current_row, 5, value=stats["withdrawn"])
            cell.border = data_border
            cell.fill = withdrawn_red if stats["withdrawn"] > 0 else row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if stats["withdrawn"] > 0:
                cell.font = Font(bold=True, color="9C0006")

            # Average GPA (formatted to 2 decimal places)
            cell = analysis_ws.cell(current_row, 6, value=stats["avg_gpa"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"

            # Color code GPA based on performance
            if stats["avg_gpa"] >= 3.5:
                cell.font = Font(bold=True, color="4A235A")  # Excellent - dark purple
            elif stats["avg_gpa"] >= 3.0:
                cell.font = Font(bold=True, color="7D3C98")  # Good - purple
            elif stats["avg_gpa"] >= 2.5:
                cell.font = Font(bold=True, color="BB8FCE")  # Average - light purple
            else:
                cell.font = Font(bold=True, color="FF0000")  # Poor - red

            # Pass Rate (formatted to 2 decimal places with %)
            cell = analysis_ws.cell(current_row, 7, value=stats["pass_rate"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"

            # Color code pass rate
            if stats["pass_rate"] >= 80:
                cell.font = Font(bold=True, color="4A235A")  # Excellent
            elif stats["pass_rate"] >= 70:
                cell.font = Font(bold=True, color="7D3C98")  # Good
            elif stats["pass_rate"] >= 60:
                cell.font = Font(bold=True, color="BB8FCE")  # Average
            else:
                cell.font = Font(bold=True, color="FF0000")  # Poor

            current_row += 1

    # Add empty row before overall summary
    current_row += 1

    # OVERALL SUMMARY ROW
    # Calculate overall statistics
    overall_avg_gpa = (
        round(overall_stats["gpa_sum"] / overall_stats["total"], 2)
        if overall_stats["total"] > 0
        else 0
    )
    overall_pass_rate = (
        round(overall_stats["passed"] / overall_stats["total"] * 100, 2)
        if overall_stats["total"] > 0
        else 0
    )

    # OVERALL label
    cell = analysis_ws.cell(current_row, 1, value="OVERALL SUMMARY")
    cell.border = overall_border
    cell.fill = overall_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(bold=True, size=12, color="4A235A")

    # Overall stats
    overall_data = [
        overall_stats["total"],
        overall_stats["passed"],
        overall_stats["carryover"],
        overall_stats["withdrawn"],
        overall_avg_gpa,
        overall_pass_rate,
    ]

    for col_idx, value in enumerate(overall_data, 2):
        cell = analysis_ws.cell(current_row, col_idx, value=value)
        cell.border = overall_border
        cell.fill = overall_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=11, color="4A235A")

        # Apply number formatting for GPA and Pass Rate
        if col_idx >= 6:  # GPA and Pass Rate columns
            cell.number_format = "0.00"

        # Color code the overall values
        if col_idx == 3 and value > 0:  # Passed
            cell.fill = passed_purple
        elif col_idx == 4 and value > 0:  # Carryover
            cell.fill = carryover_yellow
        elif col_idx == 5 and value > 0:  # Withdrawn
            cell.fill = withdrawn_red
        elif col_idx == 6:  # Overall GPA
            if value >= 3.5:
                cell.font = Font(bold=True, color="4A235A")
            elif value >= 3.0:
                cell.font = Font(bold=True, color="7D3C98")
            elif value >= 2.5:
                cell.font = Font(bold=True, color="BB8FCE")
            else:
                cell.font = Font(bold=True, color="FF0000")
        elif col_idx == 7:  # Overall Pass Rate
            if value >= 80:
                cell.font = Font(bold=True, color="4A235A")
            elif value >= 70:
                cell.font = Font(bold=True, color="7D3C98")
            elif value >= 60:
                cell.font = Font(bold=True, color="BB8FCE")
            else:
                cell.font = Font(bold=True, color="FF0000")

    # Add summary notes
    current_row += 2
    notes = [
        "PERFORMANCE INDICATORS:",
        "• Excellent: GPA ≥ 3.50 / Pass Rate ≥ 80%",
        "• Good: GPA 3.00-3.49 / Pass Rate 70-79%",
        "• Average: GPA 2.50-2.99 / Pass Rate 60-69%",
        "• Needs Improvement: GPA < 2.50 / Pass Rate < 60%",
        f"• Total Students Analyzed: {overall_stats['total']}",
        f"• Overall Pass Rate: {overall_pass_rate}%",
        f"• Overall Average GPA: {overall_avg_gpa}",
    ]

    for note in notes:
        cell = analysis_ws.cell(current_row, 1, value=note)
        if note.startswith("•"):
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.font = Font(bold=True, size=11, color="4A235A")
        current_row += 1

    print(f"✅ BM Analysis data written with {len(semester_stats)} semesters analyzed")


def update_analysis_sheet_fixed(wb, semester_key, set_name):
    """BM VERSION - Reads CURRENT data with persistent withdrawn tracking"""

    print(f" 📊 Updating BM ANALYSIS...")

    if "ANALYSIS" not in wb.sheetnames:
        print(" ❌ ANALYSIS sheet missing")
        return

    analysis_ws = wb["ANALYSIS"]

    # Clear and create headers
    clear_sheet_completely(analysis_ws)
    create_analysis_headers_bm(analysis_ws, set_name, semester_key)

    semester_keys = [
        "M-FIRST-YEAR-FIRST-SEMESTER",
        "M-FIRST-YEAR-SECOND-SEMESTER",
        "M-SECOND-YEAR-FIRST-SEMESTER",
        "M-SECOND-YEAR-SECOND-SEMESTER",
        "M-THIRD-YEAR-FIRST-SEMESTER",
        "M-THIRD-YEAR-SECOND-SEMESTER",
    ]

    # Track withdrawn students across ALL semesters
    all_withdrawn_students = set()

    # FIRST PASS: Identify withdrawn
    for key in semester_keys:
        sheet_name = find_matching_sheet(wb.sheetnames, key)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        header_row_found, headers_dict = find_sheet_structure(ws)
        if not header_row_found:
            continue

        exam_col = headers_dict.get("EXAM NUMBER") or headers_dict.get("EXAMS NUMBER")
        remarks_col = headers_dict.get("REMARKS")

        if not all([exam_col, remarks_col]):
            continue

        # ═══════════════════════════════════════════════════════
        # CRITICAL: Read CURRENT remarks from worksheet
        # ═══════════════════════════════════════════════════════
        for row in range(header_row_found + 1, ws.max_row + 1):
            exam_no = ws.cell(row, exam_col).value
            if not exam_no or "SUMMARY" in str(exam_no).upper():
                break

            remarks = ws.cell(row, remarks_col).value or ""
            if "WITHDRAW" in str(remarks).upper():
                all_withdrawn_students.add(str(exam_no).strip().upper())

    # SECOND PASS: Collect statistics with CURRENT data
    semester_stats = {}
    overall_stats = {
        "total": 0,
        "passed": 0,
        "carryover": 0,
        "withdrawn": 0,
        "gpa_sum": 0,
    }

    for key in semester_keys:
        sheet_name = find_matching_sheet(wb.sheetnames, key)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        header_row_found, headers_dict = find_sheet_structure(ws)
        if not header_row_found:
            continue

        exam_col = headers_dict.get("EXAM NUMBER") or headers_dict.get("EXAMS NUMBER")
        gpa_col = headers_dict.get("GPA")
        remarks_col = headers_dict.get("REMARKS")

        if not all([exam_col, gpa_col, remarks_col]):
            continue

        stats = {"total": 0, "passed": 0, "carryover": 0, "withdrawn": 0, "gpa_sum": 0}

        # ═══════════════════════════════════════════════════════
        # CRITICAL: Process CURRENT worksheet state
        # ═══════════════════════════════════════════════════════
        for row in range(header_row_found + 1, ws.max_row + 1):
            exam_no = ws.cell(row, exam_col).value
            if not exam_no or "SUMMARY" in str(exam_no).upper():
                break

            exam_no_clean = str(exam_no).strip().upper()
            stats["total"] += 1

            # Read CURRENT remarks and GPA
            remarks = ws.cell(row, remarks_col).value or ""
            gpa_val = ws.cell(row, gpa_col).value

            # Check persistent withdrawn status
            is_withdrawn = exam_no_clean in all_withdrawn_students

            if is_withdrawn:
                stats["withdrawn"] += 1
                # Ensure withdrawn status is maintained
                if "WITHDRAW" not in str(remarks).upper():
                    ws.cell(row, remarks_col).value = "WITHDRAWN"
            elif "PASSED" in str(remarks).upper():
                stats["passed"] += 1
            elif "RESIT" in str(remarks).upper() or "CARRYOVER" in str(remarks).upper():
                stats["carryover"] += 1
            elif "PROBATION" in str(remarks).upper():
                stats["carryover"] += 1

            try:
                stats["gpa_sum"] += float(gpa_val) if gpa_val else 0
            except:
                pass

        # Calculate averages
        stats["avg_gpa"] = (
            round(stats["gpa_sum"] / stats["total"], 2) if stats["total"] > 0 else 0
        )
        stats["pass_rate"] = (
            round(stats["passed"] / stats["total"] * 100, 2)
            if stats["total"] > 0
            else 0
        )

        semester_stats[key] = stats

        # Accumulate overall stats
        for k in ["total", "passed", "carryover", "withdrawn", "gpa_sum"]:
            overall_stats[k] += stats[k]

        # Apply sorting to maintain order
        apply_student_sorting_with_serial_numbers(ws, header_row_found, headers_dict)

    # Write statistics to ANALYSIS sheet
    write_analysis_data_bm(analysis_ws, semester_stats, overall_stats)

    print(f" ✅ BM ANALYSIS populated with current data")


def apply_complete_professional_formatting(wb, semester_key, header_row, set_name):
    """Apply complete professional formatting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    print(f" 🎨 Applying formatting...")
    # Find sheet
    current_sheet = None
    for sheet in wb.sheetnames:
        if semester_key.upper() in sheet.upper():
            current_sheet = sheet
            break
    if not current_sheet:
        return
    ws = wb[current_sheet]
    # Define styles
    header_fill = PatternFill(
        start_color="7D3C98", end_color="7D3C98", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    # Format header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    # Set specific widths
    if header_row > 0:
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                header_str = str(header_val).upper()
                if "EXAM NUMBER" in header_str:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20
                elif "NAME" in header_str:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 35
    print(f" ✅ Formatting applied")


# ============================================================
# CRITICAL FIX 6: Add Progress Indicators for Long Operations
# ============================================================
def process_students_with_progress(students, total_students):
    """
    Process students with progress indicators.
    """
    print(f"\n🎯 Processing {total_students} students...")

    progress_points = [10, 25, 50, 75, 90, 100]

    for idx, student in enumerate(students, 1):
        # Calculate progress percentage
        progress = (idx / total_students) * 100

        # Show progress at key points
        if any(abs(progress - point) < 1 for point in progress_points):
            print(f"📊 Progress: {progress:.0f}% ({idx}/{total_students})")

        # Process student
        # ... processing logic

    print(f"✅ Completed processing all {total_students} students")


def update_mastersheet_with_recalculation_COMPLETE_FIX(
    mastersheet_path,
    updates,
    semester_key,
    original_zip_path,
    course_titles_dict,
    course_units_dict,
    set_name,
):
    """
    COMPLETELY REWRITTEN VERSION - Enhanced matching and robust score updates
    WITH ALL CRITICAL FIXES APPLIED INCLUDING CONSISTENT COLORING
    """
    print(f"\n{'='*80}")
    print(f"🔄 COMPLETELY REWRITTEN: UPDATING BM MASTERSHEET")
    print(f"{'='*80}")

    # Constants
    DEFAULT_PASS_THRESHOLD = 50.0

    # Create backup first
    backup_path = mastersheet_path.replace(".xlsx", "_BACKUP.xlsx")
    try:
        shutil.copy2(mastersheet_path, backup_path)
        print(f"💾 Created backup: {backup_path}")
    except Exception as e:
        print(f"⚠️ Could not create backup: {e}")

    wb = None
    try:
        # ================================================================
        # PHASE 1: LOAD WORKBOOK AND FIND STRUCTURES
        # ================================================================
        print(f"\n📖 PHASE 1: Loading workbook and finding structures...")
        wb = load_workbook(mastersheet_path)

        # Ensure required sheets exist
        from datetime import datetime

        if "CGPA_SUMMARY" not in wb.sheetnames:
            wb.create_sheet("CGPA_SUMMARY")
        if "ANALYSIS" not in wb.sheetnames:
            wb.create_sheet("ANALYSIS")

        # Find semester sheet
        sheet_name = None
        for sheet in wb.sheetnames:
            if semester_key.upper() in sheet.upper():
                sheet_name = sheet
                break
        if not sheet_name:
            print(f"❌ No sheet found for semester: {semester_key}")
            return False
        ws = wb[sheet_name]
        print(f"✅ Using sheet: {sheet_name}")

        # ================================================================
        # PHASE 2: FIND HEADER ROW WITH ENHANCED DETECTION
        # ================================================================
        print(f"\n🔍 PHASE 2: Finding header row...")

        header_row = None
        headers = {}

        # Search for header row with exam number indicators
        exam_keywords = [
            "EXAMS NUMBER",
            "EXAM NUMBER",
            "REG NO",
            "REG. NO",
            "REGISTRATION",
        ]

        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_text = " ".join(
                [
                    str(ws.cell(row_idx, col).value or "").upper()
                    for col in range(1, min(15, ws.max_column + 1))
                ]
            )

            if any(keyword in row_text for keyword in exam_keywords):
                header_row = row_idx
                print(f"✅ Found header row at: {row_idx}")

                # Build headers dictionary
                for col_idx in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=header_row, column=col_idx).value
                    if header_val:
                        header_clean = str(header_val).strip()
                        headers[header_clean] = col_idx
                        headers[header_clean.upper()] = col_idx
                break

        if not header_row:
            print(f"❌ Could not find header row")
            return False
        print(f"📋 Found {len(set(headers.values()))} unique column headers")

        # ================================================================
        # PHASE 3: IDENTIFY CRITICAL COLUMNS
        # ================================================================
        print(f"\n🎯 PHASE 3: Identifying critical columns...")

        # Find exam number column (prioritize EXAMS NUMBER)
        exam_col = None
        exam_col_name = None

        for keyword in exam_keywords:
            for header, col_idx in headers.items():
                if keyword in header.upper():
                    exam_col = col_idx
                    exam_col_name = header
                    print(
                        f"✅ Exam column: '{header}' (matched '{keyword}') at column {col_idx}"
                    )
                    break
            if exam_col:
                break
        if not exam_col:
            print(f"❌ CRITICAL: No exam column found!")
            print(f"📋 Available headers: {list(set(headers.keys()))[:20]}")
            return False

        # ================================================================
        # PHASE 4: BUILD COURSE COLUMN MAPPING
        # ================================================================
        print(f"\n📚 PHASE 4: Building comprehensive course column mapping...")

        course_columns = {}  # All variants -> column index
        course_column_map = {}  # Normalized course code -> column index
        course_code_to_header = {}  # Normalized -> original header name

        for header, col_idx in headers.items():
            header_clean = str(header).strip().upper()

            # Match course code patterns (MID101, MWF201, etc.)
            patterns = [
                r"^[A-Z]{2,4}\s*\d{3}$",  # 2-4 letters, optional space, 3 digits
            ]

            for pattern in patterns:
                if re.match(pattern, header_clean):
                    # Normalize: remove all non-alphanumeric
                    normalized = re.sub(r"[^A-Z0-9]", "", header_clean)

                    # Store all variants
                    course_columns[header] = col_idx
                    course_columns[header_clean] = col_idx
                    course_columns[normalized] = col_idx

                    # Store primary mapping
                    course_column_map[normalized] = col_idx
                    course_code_to_header[normalized] = header

                    # Also store version with space (MID 101)
                    with_space = re.sub(r"([A-Z]+)(\d+)", r"\1 \2", header_clean)
                    course_columns[with_space] = col_idx

                    print(f" 📖 Course: '{header}' -> normalized: '{normalized}'")
                    break
        print(f"✅ Mapped {len(course_column_map)} unique course columns")
        print(f"📊 Total course variants: {len(course_columns)}")
        if not course_column_map:
            print(f"❌ No course columns found!")
            return False

        # ================================================================
        # PHASE 5: FIND SUMMARY COLUMNS
        # ================================================================
        print(f"\n📊 PHASE 5: Finding summary columns...")

        summary_columns = {}
        summary_mapping = {
            "FAILED COURSES": ["FAILED COURSES", "FAILED COURSE", "COURSES FAILED"],
            "REMARKS": ["REMARKS", "REMARK", "STATUS"],
            "CU Passed": ["CU PASSED", "CREDIT PASSED", "UNITS PASSED", "CUP"],
            "CU Failed": ["CU FAILED", "CREDIT FAILED", "UNITS FAILED", "CUF"],
            "TCPE": ["TCPE", "TOTAL CREDIT", "TOTAL CREDITS", "TCP"],
            "GPA": ["GPA", "GRADE POINT"],
            "AVERAGE": ["AVERAGE", "AVG", "MEAN"],
            "CGPA": ["CGPA", "CUMULATIVE GPA"],
        }

        for key, keywords in summary_mapping.items():
            for header, col_idx in headers.items():
                header_upper = header.upper()
                if any(kw in header_upper for kw in keywords):
                    # Skip CGPA when looking for GPA
                    if key == "GPA" and "CGPA" in header_upper:
                        continue
                    summary_columns[key] = col_idx
                    print(f" ✅ {key}: column {col_idx} ('{header}')")
                    break
        print(f"✅ Found {len(summary_columns)} summary columns")

        # ================================================================
        # PHASE 6: NORMALIZE ALL UPDATES FOR MATCHING
        # ================================================================
        print(f"\n🔄 PHASE 6: Normalizing updates for fast matching...")

        normalized_updates = {}
        for exam_no, courses in updates.items():
            # Normalize exam number
            exam_normalized = re.sub(r"[^A-Z0-9]", "", str(exam_no).strip().upper())

            normalized_updates[exam_normalized] = {
                "original_key": exam_no,
                "courses": {},
            }

            # Normalize each course code
            for course_code, score in courses.items():
                course_normalized = re.sub(
                    r"[^A-Z0-9]", "", str(course_code).strip().upper()
                )
                normalized_updates[exam_normalized]["courses"][course_normalized] = {
                    "original_code": course_code,
                    "score": score,
                }

        print(f"✅ Normalized {len(normalized_updates)} student updates")

        # ================================================================
        # PHASE 7: BUILD STUDENT ROW INDEX
        # ================================================================
        print(f"\n📇 PHASE 7: Building student index for fast lookup...")

        student_rows = {}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no = ws.cell(row_idx, exam_col).value

            if not exam_no or str(exam_no).upper() in ["", "NAN", "NONE"]:
                continue
            if "SUMMARY" in str(exam_no).upper():
                break

            exam_normalized = re.sub(r"[^A-Z0-9]", "", str(exam_no).strip().upper())
            student_rows[exam_normalized] = {
                "row": row_idx,
                "original_exam": str(exam_no).strip(),
            }

        print(f"✅ Indexed {len(student_rows)} students in mastersheet")

        # ================================================================
        # PHASE 8: APPLY SCORE UPDATES WITH CONSISTENT COLORING
        # ================================================================
        print(f"\n{'='*80}")
        print(f"📝 PHASE 8: APPLYING SCORE UPDATES WITH CONSISTENT COLORING")
        print(f"{'='*80}")

        students_updated = 0
        courses_updated = 0
        update_log = []

        # Progress indicator
        total_updates = len(normalized_updates)
        print(f"🎯 Processing {total_updates} students with progress indicators...")

        for idx, (exam_normalized, update_data) in enumerate(
            normalized_updates.items(), 1
        ):
            # Progress indicator
            progress = (idx / total_updates) * 100
            if progress % 10 == 0:
                print(f"📊 Progress: {progress:.0f}% ({idx}/{total_updates})")

            original_exam_key = update_data["original_key"]

            # Check if student exists in mastersheet
            if exam_normalized not in student_rows:
                msg = f"NOT FOUND: {original_exam_key}"
                print(f"⚠️ {msg}")
                update_log.append(msg)
                continue

            row_idx = student_rows[exam_normalized]["row"]
            original_exam = student_rows[exam_normalized]["original_exam"]

            print(f"\n🎯 Updating: {original_exam} (row {row_idx})")
            student_courses_updated = 0

            for course_normalized, course_data in update_data["courses"].items():
                original_code = course_data["original_code"]
                new_score = course_data["score"]

                # Find course column using normalized lookup
                if course_normalized not in course_column_map:
                    msg = f"COURSE NOT FOUND: {original_exam} - {original_code} (normalized: {course_normalized})"
                    print(f" ⚠️ {msg}")
                    update_log.append(msg)
                    continue

                course_col = course_column_map[course_normalized]
                old_score = ws.cell(row=row_idx, column=course_col).value

                # Update the score
                ws.cell(row=row_idx, column=course_col).value = float(new_score)

                # ═══════════════════════════════════════════════════════
                # CRITICAL FIX: Apply consistent coloring based on pass/fail
                # ═══════════════════════════════════════════════════════
                cell = ws.cell(row=row_idx, column=course_col)
                if new_score >= DEFAULT_PASS_THRESHOLD:
                    # GREEN for passed courses (consistent coloring)
                    cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="006100")  # Dark green text
                else:
                    # ORANGE for failed courses (consistent coloring)
                    cell.fill = PatternFill(
                        start_color="FFD580", end_color="FFD580", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="8B4500")  # Dark orange text

                status = "PASSED" if new_score >= DEFAULT_PASS_THRESHOLD else "FAILED"
                msg = f"UPDATED: {original_exam} - {original_code}: {old_score} → {new_score} ({status})"
                print(f" ✅ {original_code}: {old_score} → {new_score} ({status})")
                update_log.append(msg)

                student_courses_updated += 1
                courses_updated += 1

            if student_courses_updated > 0:
                students_updated += 1
                print(f" 📊 Updated {student_courses_updated} courses")

        print(f"\n{'='*80}")
        print(f"✅ PHASE 8 COMPLETE:")
        print(f" Students processed: {len(normalized_updates)}")
        print(f" Students updated: {students_updated}")
        print(f" Courses updated: {courses_updated}")
        print(f"{'='*80}")

        # Save update log
        log_path = mastersheet_path.replace(".xlsx", "_update_log.txt")
        try:
            with open(log_path, "w", encoding="utf-8") as f:
                f.write("BM CARRYOVER UPDATE LOG\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"Total students to update: {len(normalized_updates)}\n")
                f.write(f"Students successfully updated: {students_updated}\n")
                f.write(f"Total courses updated: {courses_updated}\n\n")
                f.write("=" * 80 + "\n\n")
                f.write("\n".join(update_log))
            print(f"📝 Update log saved: {log_path}")
        except Exception as e:
            print(f"⚠️ Could not save log: {e}")

        # ================================================================
        # PHASE 9: RECALCULATE ALL STUDENT RECORDS
        # ================================================================
        print(f"\n🧮 PHASE 9: Recalculating student records with CORRECT GPA...")

        # Load previous GPAs for CGPA calculation
        cgpa_data = load_previous_gpas(mastersheet_path, semester_key)

        recalc_count = 0
        total_students = len(
            [
                row
                for row in range(header_row + 1, ws.max_row + 1)
                if ws.cell(row, exam_col).value
                and "SUMMARY" not in str(ws.cell(row, exam_col).value).upper()
            ]
        )

        print(f"🎯 Recalculating {total_students} student records...")

        # Track students who got updates from carryover processing
        updated_students_set = set(normalized_updates.keys())

        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                exam_no = ws.cell(row_idx, exam_col).value
                if not exam_no or str(exam_no).upper() in ["", "NAN", "NONE"]:
                    continue
                if "SUMMARY" in str(exam_no).upper():
                    break

                # Progress indicator
                if recalc_count % 10 == 0:
                    progress = (recalc_count / total_students) * 100
                    print(f"📊 Recalculation progress: {progress:.0f}% ({recalc_count}/{total_students})")

                exam_normalized = re.sub(r"[^A-Z0-9]", "", str(exam_no).strip().upper())
                student_had_carryover_update = exam_normalized in updated_students_set

                # Collect all course scores for GPA calculation
                scores = {}
                for course_normalized, course_col in course_column_map.items():
                    score_value = ws.cell(row_idx, course_col).value
                    if score_value is not None and score_value != "":
                        try:
                            scores[course_normalized] = float(score_value)
                        except (ValueError, TypeError):
                            scores[course_normalized] = 0.0

                # Calculate GPA using CORRECT method
                gpa, total_grade_points, total_credits, cu_passed, cu_failed = calculate_gpa_correctly(
                    scores, course_units_dict, course_units_dict
                )

                # Calculate average
                valid_scores = [score for score in scores.values() if score is not None]
                average = round(sum(valid_scores) / len(valid_scores), 2) if valid_scores else 0.0

                # Identify failed courses
                failed_courses = []
                for course_normalized, score in scores.items():
                    if score < 50:
                        original_course = course_code_to_header.get(course_normalized, course_normalized)
                        failed_courses.append(original_course)

                # Calculate remarks with enhanced logic
                remarks = calculate_student_remarks(
                    cu_passed, 
                    cu_failed, 
                    total_credits, 
                    gpa, 
                    student_had_carryover_update
                )

                # Calculate CGPA
                if exam_no in cgpa_data:
                    cgpa = calculate_cgpa(cgpa_data[exam_no], gpa, total_credits)
                else:
                    cgpa = gpa

                # Update summary columns
                if "FAILED COURSES" in summary_columns:
                    ws.cell(row_idx, summary_columns["FAILED COURSES"]).value = (
                        ", ".join(failed_courses) if failed_courses else "NONE"
                    )
                if "REMARKS" in summary_columns:
                    ws.cell(row_idx, summary_columns["REMARKS"]).value = remarks
                if "CU Passed" in summary_columns:
                    ws.cell(row_idx, summary_columns["CU Passed"]).value = cu_passed
                if "CU Failed" in summary_columns:
                    ws.cell(row_idx, summary_columns["CU Failed"]).value = cu_failed
                if "TCPE" in summary_columns:
                    ws.cell(row_idx, summary_columns["TCPE"]).value = total_credits
                if "GPA" in summary_columns:
                    ws.cell(row_idx, summary_columns["GPA"]).value = gpa
                if "AVERAGE" in summary_columns:
                    ws.cell(row_idx, summary_columns["AVERAGE"]).value = average
                if "CGPA" in summary_columns:
                    ws.cell(row_idx, summary_columns["CGPA"]).value = cgpa

                recalc_count += 1

            except Exception as e:
                print(f"⚠️ Error recalculating row {row_idx}: {e}")
                continue

        print(f"✅ Recalculated {recalc_count} student records")

        # ================================================================
        # PHASE 10-13: UPDATE OTHER SHEETS
        # ================================================================
        print(f"\n📊 PHASE 10-13: Updating summary, CGPA, and analysis sheets...")

        try:
            update_summary_section_fixed(ws, headers, header_row, course_columns)
            print(f"✅ Summary section updated")
        except Exception as e:
            print(f"⚠️ Error updating summary: {e}")

        try:
            update_cgpa_summary_sheet_fixed(wb, semester_key, header_row, set_name)
            print(f"✅ CGPA_SUMMARY updated")
        except Exception as e:
            print(f"⚠️ Error updating CGPA_SUMMARY: {e}")

        try:
            update_analysis_sheet_fixed(wb, semester_key, set_name)
            print(f"✅ ANALYSIS updated")
        except Exception as e:
            print(f"⚠️ Error updating ANALYSIS: {e}")

        try:
            apply_complete_professional_formatting(wb, semester_key, header_row, set_name)
            apply_student_sorting_with_serial_numbers(ws, header_row, headers)
            print(f"✅ Formatting and sorting applied")
        except Exception as e:
            print(f"⚠️ Error applying formatting: {e}")

        # ================================================================
        # FINAL: SAVE WORKBOOK
        # ================================================================
        print(f"\n{'='*80}")
        print(f"💾 SAVING WORKBOOK")
        print(f"{'='*80}")

        try:
            wb.save(mastersheet_path)
            file_size = os.path.getsize(mastersheet_path)
            print(f"✅ Saved successfully")
            print(f"📁 File size: {file_size:,} bytes")

            # Verify file integrity
            test_wb = load_workbook(mastersheet_path)
            test_wb.close()
            print(f"✅ File integrity verified")

            print(f"\n{'='*80}")
            print(f"🎉 MASTERSHEET UPDATE COMPLETE!")
            print(f"{'='*80}")
            print(f"📊 Summary:")
            print(f" Students updated: {students_updated}/{len(normalized_updates)}")
            print(f" Courses updated: {courses_updated}")
            print(f"{'='*80}")

            return True

        except Exception as save_error:
            print(f"❌ Save error: {save_error}")
            import traceback
            traceback.print_exc()

            # Restore from backup
            try:
                shutil.copy2(backup_path, mastersheet_path)
                print(f"🔄 Restored from backup")
            except:
                pass
            return False

    except Exception as e:
        print(f"❌ Critical error: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # Always close workbook
        if wb:
            try:
                wb.close()
                print("✅ Workbook closed")
            except:
                pass


# ============================================================
# Carryover Processing Functions (BM-Compatible)
# ============================================================
def load_carryover_files(carryover_dir, semester_key=None):
    """Load carryover files - BM VERSION FIXED"""
    carryover_files = []
    for file in os.listdir(carryover_dir):
        # Look for both JSON and Excel carryover files
        if file.startswith("co_student_") and (
            file.endswith(".json") or file.endswith(".xlsx")
        ):
            file_semester = extract_semester_from_filename(file)
            file_semester_standardized = (
                standardize_semester_name(file_semester) if file_semester else None
            )
            # Skip if semester doesn't match filter
            if semester_key and file_semester_standardized != semester_key:
                continue
            file_path = os.path.join(carryover_dir, file)
            try:
                if file.endswith(".json"):
                    with open(file_path, "r") as f:
                        data = json.load(f)
                else:  # .xlsx
                    df = pd.read_excel(file_path)
                    data = df.to_dict("records")
                carryover_files.append(
                    {
                        "filename": file,
                        "semester": file_semester_standardized,
                        "data": data,
                        "count": len(data),
                        "file_path": file_path,
                    }
                )
                print(f"✅ Loaded {len(data)} records from {file}")
            except Exception as e:
                print(f"❌ Error loading {file}: {e}")
    if not carryover_files:
        print(f"❌ No carryover files found in {carryover_dir}")
        return []
    print(f"📚 Total BM carryover files loaded: {len(carryover_files)}")
    return carryover_files


def save_carryover_json_records(carryover_data, carryover_output_dir, semester_key):
    """
    Save BM carryover records as JSON files
    """
    json_dir = os.path.join(carryover_output_dir, "CARRYOVER_RECORDS")
    os.makedirs(json_dir, exist_ok=True)
    print(f"\n💾 SAVING BM CARRYOVER JSON RECORDS")
    print(f"📁 JSON directory: {json_dir}")
    # Convert carryover_data to JSON-friendly format
    json_records = {}
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        # Build the JSON record structure
        record = {
            "exam_number": exam_no,
            "name": student["NAME"],
            "carryover_courses": {},
            "passed_resit_courses": {},
            "failed_resit_courses": {},
        }
        # Process each resit course
        for course_code, course_data in student["RESIT_COURSES"].items():
            course_info = {
                "course_code": course_code,
                "course_title": course_data.get("course_title", course_code),
                "credit_unit": course_data.get("credit_unit", 2),
                "original_score": course_data["original_score"],
                "resit_score": course_data["resit_score"],
            }
            # Add to carryover courses (all courses that were taken as resit)
            record["carryover_courses"][course_code] = course_info
            # Separate into passed/failed based on resit score
            if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD:
                record["passed_resit_courses"][course_code] = course_info
            else:
                record["failed_resit_courses"][course_code] = course_info
        json_records[exam_no] = record
    # Save as JSON file with timestamp and semester
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f"co_student_BM-{semester_key}_{timestamp}.json"
    json_filepath = os.path.join(json_dir, json_filename)
    try:
        with open(json_filepath, "w", encoding="utf-8") as f:
            json.dump(json_records, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved BM JSON carryover records: {json_filepath}")
        print(f"📊 Records saved: {len(json_records)} students")
        return json_filepath
    except Exception as e:
        print(f"❌ Error saving BM JSON records: {e}")
        traceback.print_exc()
        return None


def copy_json_to_centralized_location(json_filepath, set_name, semester_key):
    """
    Copy JSON file to centralized CARRYOVER_RECORDS location for BM
    """
    try:
        # Determine the centralized location
        base_dir = get_base_directory()
        centralized_dir = os.path.join(
            base_dir,
            "EXAMS_INTERNAL",
            "BM",
            set_name,
            "CLEAN_RESULTS",
            "CARRYOVER_RECORDS",
        )
        os.makedirs(centralized_dir, exist_ok=True)
        # Copy the JSON file
        filename = os.path.basename(json_filepath)
        dest_path = os.path.join(centralized_dir, filename)
        shutil.copy2(json_filepath, dest_path)
        print(f"\n📋 COPIED TO BM CENTRALIZED LOCATION")
        print(f"✅ From: {json_filepath}")
        print(f"✅ To: {dest_path}")
        return dest_path
    except Exception as e:
        print(f"❌ Error copying to BM centralized location: {e}")
        traceback.print_exc()
        return None


def generate_carryover_mastersheet(
    carryover_data,
    output_dir,
    semester_key,
    set_name,
    timestamp,
    cgpa_data,
    course_titles,
    course_units,
    course_code_to_title,
    course_code_to_unit,
):
    """Generate BM CARRYOVER_mastersheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BM_CARRYOVER_RESULTS"
    program_name = "BASIC MIDWIFERY"
    program_abbr = "BM"
    if os.path.exists(DEFAULT_LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image

            img = Image(DEFAULT_LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"⚠️ Could not add logo: {e}")
    current_year = 2025
    next_year = 2026
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )
    all_courses = set()
    for student in carryover_data:
        all_courses.update(student["RESIT_COURSES"].keys())
    previous_semesters = get_previous_semesters_for_display(semester_key)
    headers = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        headers.append(f"GPA {prev_sem}")
    course_headers = []
    for course in sorted(all_courses):
        course_headers.extend([f"{course}", f"{course}_RESIT"])
    headers.extend(course_headers)
    headers.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    total_columns = len(headers)
    last_column = get_column_letter(total_columns)
    # Header section
    ws.merge_cells(f"A3:{last_column}3")
    title_cell = ws["A3"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A4:{last_column}4")
    subtitle_cell = ws["A4"]
    subtitle_cell.value = f"RESIT - {current_year}/{next_year} SESSION {program_name} {level} {sem_display} EXAMINATIONS RESULT — {datetime.now().strftime('%B %d, %Y')}"
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    print(f"🔍 BM Courses found in resit data: {sorted(all_courses)}")
    print(
        f"📊 BM GPA columns for {semester_key}: Previous={previous_semesters}, Current={current_semester_name}"
    )
    # Course title row
    title_row = [""] * 3
    for prev_sem in previous_semesters:
        title_row.extend([""])
    for course in sorted(all_courses):
        course_title = find_course_title(course, course_titles, course_code_to_title)
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        title_row.extend([course_title, course_title])
    title_row.extend(["", "", ""])
    ws.append(title_row)
    # Credit unit row
    credit_row = [""] * 3
    for prev_sem in previous_semesters:
        credit_row.extend([""])
    for course in sorted(all_courses):
        credit_unit = find_credit_unit(course, course_units, course_code_to_unit)
        credit_row.extend([f"CU: {credit_unit}", f"CU: {credit_unit}"])
    credit_row.extend(["", "", ""])
    ws.append(credit_row)
    # Code row
    code_row = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        code_row.append(f"GPA {prev_sem}")
    for course in sorted(all_courses):
        code_row.append(f"{course}")
        code_row.append(f"{course}_RESIT")
    code_row.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    ws.append(code_row)
    # Apply formatting to header rows
    course_colors = [
        "E8DAEF",
        "F4ECF7",
        "D6EAF8",
        "D1F2EB",
        "FDEBD0",
        "FADBD8",
        "EBDEF0",
        "D5F5E3",
        "FCF3CF",
        "F6DDCC",
    ]
    start_col = 4
    if previous_semesters:
        start_col += len(previous_semesters)
    color_index = 0
    for course in sorted(all_courses):
        for row in [5, 6, 7]:
            for offset in [0, 1]:
                cell = ws.cell(row=row, column=start_col + offset)
                cell.fill = PatternFill(
                    start_color=course_colors[color_index % len(course_colors)],
                    end_color=course_colors[color_index % len(course_colors)],
                    fill_type="solid",
                )
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        for offset in [0, 1]:
            cell = ws.cell(row=5, column=start_col + offset)
            cell.alignment = Alignment(
                text_rotation=90, horizontal="center", vertical="center"
            )
            cell.font = Font(bold=True, size=9)
        color_index += 1
        start_col += 2
    # Format basic info columns
    for row in [5, 6, 7]:
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color="7D3C98", end_color="7D3C98", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
    # Add student data
    row_idx = 8
    failed_counts = {course: 0 for course in all_courses}
    # CRITICAL FIX: Proper serial numbers from 1 to n
    serial_number = 1
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        # Serial Number (PROPERLY SORTED from 1 to n)
        ws.cell(row=row_idx, column=1, value=serial_number)
        serial_number += 1
        ws.cell(row=row_idx, column=2, value=student["EXAM NUMBER"])
        ws.cell(row=row_idx, column=3, value=student["NAME"])
        # Previous GPAs
        gpa_col = 4
        for prev_sem in previous_semesters:
            gpa_value = student.get(f"GPA_{prev_sem}", "")
            ws.cell(row=row_idx, column=gpa_col, value=gpa_value)
            gpa_col += 1
        # Course scores
        course_col = gpa_col
        color_index = 0
        for course in sorted(all_courses):
            for offset in [0, 1]:
                cell = ws.cell(row=row_idx, column=course_col + offset)
                cell.fill = PatternFill(
                    start_color=course_colors[color_index % len(course_colors)],
                    end_color=course_colors[color_index % len(course_colors)],
                    fill_type="solid",
                )
            if course in student["RESIT_COURSES"]:
                course_data = student["RESIT_COURSES"][course]
                orig_cell = ws.cell(
                    row=row_idx, column=course_col, value=course_data["original_score"]
                )
                if course_data["original_score"] < DEFAULT_PASS_THRESHOLD:
                    orig_cell.fill = PatternFill(
                        start_color="FADBD8", end_color="FADBD8", fill_type="solid"
                    )
                resit_cell = ws.cell(
                    row=row_idx, column=course_col + 1, value=course_data["resit_score"]
                )
                if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD:
                    resit_cell.fill = PatternFill(
                        start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"
                    )
                else:
                    resit_cell.fill = PatternFill(
                        start_color="FFD580", end_color="FFD580", fill_type="solid"
                    )
                    failed_counts[course] += 1
            else:
                ws.cell(row=row_idx, column=course_col, value="")
                ws.cell(row=row_idx, column=course_col + 1, value="")
            color_index += 1
            course_col += 2
        # Current GPA and CGPA
        ws.cell(row=row_idx, column=course_col, value=student["CURRENT_GPA"])
        ws.cell(row=row_idx, column=course_col + 1, value=student["CURRENT_CGPA"])
        # Remarks
        remarks = generate_remarks(student["RESIT_COURSES"])
        ws.cell(row=row_idx, column=course_col + 2, value=remarks)
        row_idx += 1
    # Add failed counts row
    failed_row_idx = row_idx
    ws.cell(row=failed_row_idx, column=1, value="FAILED COUNT BY COURSE:").font = Font(
        bold=True
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=failed_row_idx, column=col)
        cell.fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
    course_col = gpa_col
    for course in sorted(all_courses):
        count_cell = ws.cell(
            row=failed_row_idx, column=course_col + 1, value=failed_counts[course]
        )
        count_cell.font = Font(bold=True)
        count_cell.fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
        course_col += 2
    # Add summary section
    summary_start_row = failed_row_idx + 2
    total_students = len(carryover_data)
    passed_all = sum(
        1
        for student in carryover_data
        if all(
            course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
            for course_data in student["RESIT_COURSES"].values()
        )
    )
    carryover_count = total_students - passed_all
    total_failed_attempts = sum(failed_counts.values())
    summary_data = [
        ["BM CARRYOVER SUMMARY"],
        [
            f"A total of {total_students} students registered and sat for the Carryover Examination"
        ],
        [f"A total of {passed_all} students passed all carryover courses"],
        [
            f"A total of {carryover_count} students failed one or more carryover courses and must repeat them"
        ],
        [f"Total failed resit attempts: {total_failed_attempts} across all courses"],
        [
            f"BM Carryover processing completed on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"
        ],
    ]
    for i, row_data in enumerate(summary_data):
        row_num = summary_start_row + i
        if row_data[0]:
            ws.merge_cells(
                start_row=row_num, start_column=1, end_row=row_num, end_column=10
            )
            cell = ws.cell(row=row_num, column=1, value=row_data[0])
            if i == 0:
                cell.font = Font(bold=True, size=12, underline="single")
            else:
                cell.font = Font(bold=False, size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to data area
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=7, max_row=row_idx - 1, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border = thin_border
    # Freeze panes and adjust column widths
    ws.freeze_panes = "D8"
    # Set column widths
    ws.column_dimensions["A"].width = 8  # S/N
    ws.column_dimensions["B"].width = 18  # EXAM NUMBER
    ws.column_dimensions["C"].width = 35  # NAME
    # Set GPA column widths
    for col in range(4, 4 + len(previous_semesters)):
        ws.column_dimensions[get_column_letter(col)].width = 15
    # Set course column widths
    for col in range(4 + len(previous_semesters), len(headers) - 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    # Set summary column widths
    for col in range(len(headers) - 2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    # Save the file
    filename = f"BM_CARRYOVER_mastersheet_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    print(f"✅ BM CARRYOVER mastersheet generated: {filepath}")
    return filepath


def generate_individual_reports(
    carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data
):
    """Generate individual BM student reports."""
    reports_dir = os.path.join(output_dir, "INDIVUAL_REPORTS")
    os.makedirs(reports_dir, exist_ok=True)
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        safe_exam_no = sanitize_filename(exam_no)
        filename = f"bm_carryover_report_{safe_exam_no}_{timestamp}.csv"
        filepath = os.path.join(reports_dir, filename)
        report_data = []
        report_data.append(["BM CARRYOVER RESULT REPORT"])
        report_data.append(["FCT COLLEGE OF NURSING SCIENCES"])
        report_data.append([f"BM Set: {set_name}"])
        report_data.append([f"BM Semester: {semester_key}"])
        report_data.append([])
        report_data.append(["BM STUDENT INFORMATION"])
        report_data.append(["Exam Number:", student["EXAM NUMBER"]])
        report_data.append(["Name:", student["NAME"]])
        report_data.append([])
        report_data.append(["BM PREVIOUS GPAs"])
        for key in sorted([k for k in student.keys() if k.startswith("GPA_")]):
            semester = key.replace("GPA_", "")
            report_data.append([f"{semester}:", student[key]])
        report_data.append([])
        report_data.append(["BM CURRENT ACADEMIC RECORD"])
        report_data.append(["Current GPA:", student["CURRENT_GPA"]])
        report_data.append(["Current CGPA:", student["CURRENT_CGPA"]])
        report_data.append([])
        report_data.append(["BM RESIT COURSES"])
        report_data.append(
            [
                "Course Code",
                "Course Title",
                "Credit Unit",
                "Original Score",
                "Resit Score",
                "Status",
            ]
        )
        for course_code, course_data in student["RESIT_COURSES"].items():
            status = (
                "PASSED"
                if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
                else "FAILED"
            )
            course_title = course_data.get("course_title", course_code)
            credit_unit = course_data.get("credit_unit", 0)
            report_data.append(
                [
                    course_code,
                    course_title,
                    credit_unit,
                    course_data["original_score"],
                    course_data["resit_score"],
                    status,
                ]
            )
        try:
            df = pd.DataFrame(report_data)
            df.to_csv(filepath, index=False, header=False)
            print(f"✅ Generated BM report for: {exam_no}")
        except Exception as e:
            print(f"❌ Error generating BM report for {exam_no}: {e}")
    print(
        f"✅ Generated {len(carryover_data)} individual BM student reports in {reports_dir}"
    )


# ============================================================
# CRITICAL FIX: Output Directory Management
# ============================================================
def get_output_directory(set_name):
    """Get the correct output directory for carryover results - FIXED"""
    # For web interface: Save directly to CLEAN_RESULTS
    clean_dir = os.path.join(
        BASE_DIR, "EXAMS_INTERNAL", "BM", set_name, "CLEAN_RESULTS"
    )
    # Alternative paths
    if not os.path.exists(clean_dir):
        clean_dir = os.path.join(BASE_DIR, "BM", set_name, "CLEAN_RESULTS")
    if not os.path.exists(clean_dir):
        print(f"⚠️ CLEAN_RESULTS directory doesn't exist, will create: {clean_dir}")
        os.makedirs(clean_dir, exist_ok=True)
    print(f"📁 Output directory: {clean_dir}")
    return clean_dir


# ============================================================
# CRITICAL FIX: Main Processing Function with All Fixes
# ============================================================
def process_carryover_results(
    resit_file_path,
    source_path,
    source_type,
    semester_key,
    set_name,
    pass_threshold,
    output_dir,
):
    """
    FIXED VERSION: Process BM carryover results with robust mastersheet reading
    ALL CRITICAL FIXES APPLIED - Consistent output, proper coloring, and correct ZIP management
    """
    print(f"\n🔄 FIXED BM VERSION: PROCESSING CARRYOVER FOR {semester_key}")
    print("=" * 60)

    # Initialize variables
    temp_mastersheet_path = None
    temp_dir = None
    updated_zip_path = None
    update_success = False

    try:
        # Validate this is a BM semester
        if not is_bm_semester(semester_key):
            print(f"❌ ERROR: Semester '{semester_key}' is not a valid BM semester!")
            return False

        # Load BM course data
        (
            semester_course_titles,
            semester_credit_units,
            course_code_to_title,
            course_code_to_unit,
        ) = load_course_data()

        # Debug course matching
        debug_course_matching_bm(
            resit_file_path, course_code_to_title, course_code_to_unit
        )

        # Get semester info
        year, sem_num, level, sem_display, set_code, sem_name = (
            get_semester_display_info(semester_key)
        )

        # Find course titles for the specific semester
        possible_sheet_keys = [
            f"{set_code} {sem_display}",
            f"{set_code} {sem_name}",
            semester_key,
            semester_key.replace("-", " ").upper(),
            f"{level} {sem_display}",
        ]
        course_titles_dict = {}
        credit_units_dict = {}
        for sheet_key in possible_sheet_keys:
            sheet_standard = standardize_semester_key(sheet_key)
            if sheet_standard in semester_course_titles:
                course_titles_dict = semester_course_titles[sheet_standard]
                credit_units_dict = semester_credit_units[sheet_standard]
                print(
                    f"✅ Using BM sheet key: '{sheet_key}' with {len(course_titles_dict)} courses"
                )
                break
            else:
                print(f"❌ BM sheet key not found: '{sheet_key}'")
        if not course_titles_dict:
            print(
                f"⚠️ No BM semester-specific course data found, using global course mappings"
            )
            course_titles_dict = course_code_to_title
            credit_units_dict = course_code_to_unit
        print(
            f"📊 Final BM course mappings: {len(course_titles_dict)} titles, {len(credit_units_dict)} units"
        )

        timestamp = datetime.now().strftime(TIMESTAMP_FMT)
        carryover_output_dir = os.path.join(
            output_dir, f"BM_CARRYOVER_{set_name}_{semester_key}_{timestamp}"
        )
        os.makedirs(carryover_output_dir, exist_ok=True)
        print(f"📁 BM Output directory: {carryover_output_dir}")

        if not os.path.exists(resit_file_path):
            print(f"❌ BM resit file not found: {resit_file_path}")
            return False

        temp_mastersheet_path, temp_dir = get_mastersheet_path(
            source_path, source_type, semester_key
        )
        if not temp_mastersheet_path:
            print(f"❌ Failed to get BM mastersheet")
            return False

        print(f"📖 Reading BM files...")
        resit_df = pd.read_excel(resit_file_path, header=0)
        print(f"📊 BM Resit file rows: {len(resit_df)}")
        print(f"📊 BM Resit file columns: {resit_df.columns.tolist()}")
        resit_exam_col = find_exam_number_column(resit_df)
        print(f"📊 BM Resit exam column: '{resit_exam_col}'")
        if resit_exam_col:
            print(
                f"📊 Sample BM resit exam numbers: {resit_df[resit_exam_col].head().tolist()}"
            )

        xl = pd.ExcelFile(temp_mastersheet_path)
        sheet_name = find_matching_sheet(xl.sheet_names, semester_key)
        if not sheet_name:
            print(f"❌ No matching BM sheet found for {semester_key}")
            return False

        print(f"📖 Using BM sheet '{sheet_name}' for current semester {semester_key}")
        # Use the enhanced mastersheet reading function
        mastersheet_df, mastersheet_exam_col = read_mastersheet_with_flexible_headers(
            temp_mastersheet_path, sheet_name
        )
        if mastersheet_df is None or mastersheet_exam_col is None:
            print(f"❌ Could not read BM mastersheet with flexible headers")
            # Fallback to quick fix
            print(f"🔄 Trying quick fix...")
            mastersheet_df, mastersheet_exam_col = quick_fix_read_mastersheet(
                temp_mastersheet_path, sheet_name
            )
        if mastersheet_df is None or mastersheet_exam_col is None:
            print(f"❌ Could not read BM mastersheet with any method")
            return False

        # DEBUG: Print mastersheet info
        print(f"📊 BM Mastersheet rows: {len(mastersheet_df)}")
        print(f"📊 BM Mastersheet columns: {mastersheet_df.columns.tolist()}")
        if mastersheet_exam_col in mastersheet_df.columns:
            print(
                f"📊 Sample BM mastersheet exam numbers: {mastersheet_df[mastersheet_exam_col].head().tolist()}"
            )
        print(
            f"✅ BM files loaded - Resit: {len(resit_df)} rows, Mastersheet: {len(mastersheet_df)} students"
        )

        resit_exam_col = find_exam_number_column(resit_df)
        if not resit_exam_col:
            print(f"❌ Cannot find exam number column in BM resit file")
            return None

        print(
            f"📝 BM Exam columns - Resit: '{resit_exam_col}', Mastersheet: '{mastersheet_exam_col}'"
        )

        # Load previous GPAs for CGPA calculation
        cgpa_data = load_previous_gpas(temp_mastersheet_path, semester_key)
        carryover_data = []
        updated_students = set()

        print(f"\n🎯 PROCESSING BM RESIT SCORES...")

        # CRITICAL FIX: Add progress indicators
        total_students = len(resit_df)
        print(f"🎯 Processing {total_students} students with progress indicators...")

        for idx, resit_row in enumerate(resit_df.iterrows(), 1):
            # PROGRESS INDICATOR
            progress = (idx / total_students) * 100
            if progress % 10 == 0:  # Show progress every 10%
                print(f"📊 Progress: {progress:.0f}% ({idx}/{total_students})")

            try:
                exam_no = str(resit_row[1][resit_exam_col]).strip().upper()
                if not exam_no or exam_no in ["NAN", "NONE", ""]:
                    continue

                # Use enhanced student matching
                student_data = find_student_in_mastersheet_fixed(
                    exam_no, mastersheet_df, mastersheet_exam_col
                )
                if student_data is None:
                    print(f"⚠️ BM Student {exam_no} not found in mastersheet - skipping")
                    continue

                student_name = student_data.get("NAME", "Unknown")
                current_credits = 0
                # Find credits column
                for col in mastersheet_df.columns:
                    if "TCPE" in str(col).upper():
                        current_credits = student_data.get(col, 0)
                        break

                student_record = {
                    "EXAM NUMBER": exam_no,
                    "NAME": student_name,
                    "RESIT_COURSES": {},
                    "CURRENT_GPA": student_data.get("GPA", 0),
                    "CURRENT_CREDITS": current_credits,
                }

                # Process resit courses
                for col in resit_df.columns:
                    if col == resit_exam_col or col == "NAME" or "Unnamed" in str(col):
                        continue
                    resit_score = resit_row[1].get(col)
                    if pd.isna(resit_score) or resit_score == "":
                        continue
                    try:
                        resit_score_val = float(resit_score)
                    except (ValueError, TypeError):
                        continue
                    # Check if course exists in mastersheet
                    if col in mastersheet_df.columns:
                        original_score = student_data.get(col)
                        if pd.isna(original_score):
                            continue
                    else:
                        # Try to find course with similar name
                        course_found = False
                        for ms_col in mastersheet_df.columns:
                            if col.upper() == ms_col.upper() or col.replace(
                                " ", ""
                            ) == ms_col.replace(" ", ""):
                                original_score = student_data.get(ms_col)
                                course_found = True
                                break
                        if not course_found:
                            continue
                    try:
                        original_score_val = (
                            float(original_score)
                            if not pd.isna(original_score)
                            else 0.0
                        )
                    except (ValueError, TypeError):
                        original_score_val = 0.0
                    if original_score_val < pass_threshold:
                        course_title = find_course_title(
                            col, course_titles_dict, course_code_to_title
                        )
                        credit_unit = find_credit_unit(
                            col, credit_units_dict, course_code_to_unit
                        )
                        student_record["RESIT_COURSES"][col] = {
                            "original_score": original_score_val,
                            "resit_score": resit_score_val,
                            "updated": resit_score_val >= pass_threshold,
                            "course_title": course_title,
                            "credit_unit": credit_unit,
                        }
                # Process previous GPAs
                previous_semesters = get_previous_semesters_for_display(semester_key)
                for prev_sem in previous_semesters:
                    student_record[f"GPA_{prev_sem}"] = student_data.get(
                        f"GPA_{prev_sem}", ""
                    )
                # Enhanced GPA/CGPA calculation with resit scores
                if student_record["RESIT_COURSES"]:
                    # Identify course columns in mastersheet
                    import re

                    course_columns = [
                        col
                        for col in mastersheet_df.columns
                        if re.match(r"^[A-Z]{3}\d{3}$", str(col).upper())
                    ]
                    # Recalculate updated current GPA with resit overrides
                    total_grade_points = 0.0
                    total_credits = 0
                    for col in course_columns:
                        if col in mastersheet_df.columns:
                            original_score = student_data.get(col, 0)
                            score = original_score
                            # Apply resit scores if available
                            if col in student_record["RESIT_COURSES"]:
                                score = student_record["RESIT_COURSES"][col][
                                    "resit_score"
                                ]
                            try:
                                score_val = float(score)
                                credit_unit = find_credit_unit(
                                    col, credit_units_dict, course_code_to_unit
                                )
                                grade_point = get_grade_point(score_val)
                                total_grade_points += grade_point * credit_unit
                                total_credits += credit_unit
                            except (ValueError, TypeError):
                                continue
                    updated_gpa = (
                        round(total_grade_points / total_credits, 2)
                        if total_credits > 0
                        else 0.0
                    )
                    student_record["CURRENT_GPA"] = updated_gpa
                    student_record["CURRENT_CREDITS"] = total_credits
                    # Recalculate CGPA
                    if exam_no in cgpa_data:
                        student_record["CURRENT_CGPA"] = calculate_cgpa(
                            cgpa_data[exam_no], updated_gpa, total_credits
                        )
                    else:
                        student_record["CURRENT_CGPA"] = updated_gpa
                    carryover_data.append(student_record)
                    updated_students.add(exam_no)
                    print(
                        f"✅ BM {exam_no}: {len(student_record['RESIT_COURSES'])} resit courses, Updated GPA: {student_record['CURRENT_GPA']}, CGPA: {student_record['CURRENT_CGPA']}"
                    )
            except Exception as e:
                print(
                    f"❌ Error processing BM student {exam_no if 'exam_no' in locals() else 'unknown'}: {e}"
                )
                continue

        # DEBUG: Print final stats
        print(f"\n📊 BM FINAL STATS:")
        print(f" Total carryover students processed: {len(carryover_data)}")
        print(f" Students with updates: {len(updated_students)}")

        if carryover_data:
            print(f"\n📊 GENERATING BM OUTPUTS...")
            # 1. Generate the Excel carryover mastersheet
            carryover_mastersheet_path = generate_carryover_mastersheet(
                carryover_data,
                carryover_output_dir,
                semester_key,
                set_name,
                timestamp,
                cgpa_data,
                course_titles_dict,
                credit_units_dict,
                course_code_to_title,
                course_code_to_unit,
            )
            # 2. Generate individual reports
            generate_individual_reports(
                carryover_data,
                carryover_output_dir,
                semester_key,
                set_name,
                timestamp,
                cgpa_data,
            )
            # 3. Generate JSON records
            json_filepath = save_carryover_json_records(
                carryover_data, carryover_output_dir, semester_key
            )
            if json_filepath:
                # 4. Copy to centralized location
                centralized_json = copy_json_to_centralized_location(
                    json_filepath, set_name, semester_key
                )

            # ============================================
            # STEP 6: UPDATE ORIGINAL MASTERSHEET WITH ALL ENHANCEMENTS
            # ============================================
            if carryover_data:
                print(f"\n{'='*60}")
                print(
                    f"🔄 STEP 6: UPDATING ORIGINAL BM MASTERSHEET WITH ALL ENHANCEMENTS"
                )
                print(f"{'='*60}")
                try:
                    # Find the original result ZIP
                    clean_dir_parent = output_dir
                    all_result_zips = [
                            f
                            for f in os.listdir(clean_dir_parent)
                            if f.lower().endswith(".zip") and "carryover" not in f.lower()
                        ]

                    if not all_result_zips:
                        print(f"❌ No result ZIP found in {clean_dir_parent}")
                        return False

                    # Use the most recently modified zip (this guarantees chaining/persistence)
                    latest_zip_name = max(
                        all_result_zips,
                        key=lambda f: os.path.getmtime(os.path.join(clean_dir_parent, f))
                    )
                    original_zip_path = os.path.join(clean_dir_parent, latest_zip_name)
                    print(f"✅ Using latest result ZIP (persistent): {original_zip_path}")
                    
                    # Create updated ZIP name
                    match = re.search(r"UPDATED_(\d+)_", latest_zip_name)
                    current_count = int(match.group(1)) if match else 0
                    new_count = current_count + 1
                    if current_count == 0:
                        updated_zip_name = f"UPDATED_{new_count}_{latest_zip_name}"
                    else:
                        updated_zip_name = re.sub(
                            r"UPDATED_\d+", f"UPDATED_{new_count}", latest_zip_name
                        )
                    updated_zip_path = os.path.join(clean_dir_parent, updated_zip_name)
                    
                    print(f"✅ Found latest BM ZIP: {original_zip_path}")
                    # Extract the ZIP to temporary directory
                    temp_extract_dir = tempfile.mkdtemp()
                    try:
                        with zipfile.ZipFile(original_zip_path, "r") as zip_ref:
                            zip_ref.extractall(temp_extract_dir)
                        print(f"✅ Extracted to temp directory: {temp_extract_dir}")
                        # Find the mastersheet in the extracted files
                        mastersheet_path = None
                        for root, dirs, files in os.walk(temp_extract_dir):
                            for file in files:
                                if "mastersheet" in file.lower() and file.endswith(
                                    ".xlsx"
                                ):
                                    mastersheet_path = os.path.join(root, file)
                                    print(
                                        f"✅ Found BM mastersheet: {mastersheet_path}"
                                    )
                                    break
                            if mastersheet_path:
                                break
                        if not mastersheet_path:
                            print(f"❌ No BM mastersheet found in ZIP")
                        else:
                            print(f"✅ Found BM mastersheet: {mastersheet_path}")
                            # Build updates dictionary from carryover data
                            updates = {}
                            for student in carryover_data:
                                exam_no = student["EXAM NUMBER"]
                                updates[exam_no] = {}
                                for course_code, course_data in student[
                                    "RESIT_COURSES"
                                ].items():
                                    # CRITICAL FIX: Update ALL resit scores (both passed and failed)
                                    # This ensures failed resit values are also recorded in the mastersheet
                                    updates[exam_no][course_code] = course_data[
                                        "resit_score"
                                    ]
                            print(f"📊 Prepared BM updates for {len(updates)} students")
                            # Update the mastersheet with full recalculation and ALL enhancements
                            # CRITICAL FIX: Use the new COMPLETE_FIX function with SINGLE session
                            update_success = (
                                update_mastersheet_with_recalculation_COMPLETE_FIX(
                                    mastersheet_path=mastersheet_path,
                                    updates=updates,
                                    semester_key=semester_key,
                                    original_zip_path=original_zip_path,
                                    course_titles_dict=course_titles_dict,
                                    course_units_dict=credit_units_dict,
                                    set_name=set_name,
                                )
                            )
                            if update_success:
                                # Create backup of original (only if doesn't exist)
                                backup_zip = original_zip_path.replace(
                                    ".zip", "_BACKUP.zip"
                                )
                                if not os.path.exists(backup_zip):
                                    shutil.copy2(original_zip_path, backup_zip)
                                    print(f"💾 Created BM backup: {backup_zip}")
                                print(f"📦 Creating updated BM ZIP: {updated_zip_name}")

                                # CRITICAL FIX: Use new robust ZIP creation function
                                zip_success = create_updated_zip_from_directory(
                                    temp_extract_dir, updated_zip_path
                                )

                                if zip_success:
                                    print(f"✅ SUCCESS: Created {updated_zip_name}")
                                else:
                                    print(
                                        f"❌ ERROR: Updated BM ZIP was not created properly"
                                    )

                                print(
                                    f"✅ BM Original preserved: {os.path.basename(original_zip_path)}"
                                )
                            else:
                                print(
                                    f"❌ BM Mastersheet update had some errors, but continuing"
                                )
                                # Even if update had errors, try to create the ZIP anyway
                                zip_success = create_updated_zip_from_directory(
                                    temp_extract_dir, updated_zip_path
                                )

                    except Exception as e:
                        print(f"❌ Error during BM ZIP processing: {e}")
                        traceback.print_exc()
                    finally:
                        # Clean up temp directory
                        if os.path.exists(temp_extract_dir):
                            shutil.rmtree(temp_extract_dir)
                            print(f"🧹 Cleaned up BM temp extraction directory")
                except Exception as e:
                    print(f"❌ Error updating BM mastersheet: {e}")
                    traceback.print_exc()
                
                # ============================================================
                # CRITICAL FIX: Create FINAL CARRYOVER ZIP (only this one)
                # ============================================================

                # Create SINGLE carryover ZIP file from carryover results
                if 'carryover_output_dir' in locals() and carryover_output_dir and os.path.exists(carryover_output_dir):
                    carryover_zip_filename = f"BM_CARRYOVER_{set_name}_{semester_key}_{timestamp}.zip"
                    carryover_zip_path = os.path.join(output_dir, carryover_zip_filename)
                    
                    print(f"📦 Creating FINAL CARRYOVER ZIP file: {carryover_zip_filename}")
                    
                    with zipfile.ZipFile(carryover_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for root, dirs, files in os.walk(carryover_output_dir):
                            for file in files:
                                file_path = os.path.join(root, file)
                                arcname = os.path.relpath(file_path, carryover_output_dir)
                                zipf.write(file_path, arcname)
                                print(f"  ✅ Added to CARRYOVER ZIP: {arcname}")
                    
                    # Verify ZIP was created
                    if os.path.exists(carryover_zip_path) and os.path.getsize(carryover_zip_path) > 100:
                        print(f"✅ Successfully created CARRYOVER ZIP: {carryover_zip_path} ({os.path.getsize(carryover_zip_path):,} bytes)")
                        
                        # Clean up the carryover directory after zipping
                        try:
                            shutil.rmtree(carryover_output_dir)
                            print(f"🗑️ Cleaned up temporary directory: {carryover_output_dir}")
                        except Exception as e:
                            print(f"⚠️ Could not remove temporary directory: {e}")
                    else:
                        print(f"❌ Failed to create valid CARRYOVER ZIP file")
                
                print(f"\n{'='*60}")
                print(f"🎉 BM CARRYOVER PROCESSING COMPLETE WITH ALL ENHANCEMENTS!")
                print(f"{'='*60}")
                print(f"📊 Summary:")
                print(f" Original ZIP: {os.path.basename(original_zip_path)}")
                print(f" Carryover ZIP: {carryover_zip_filename}")
                print(f" Updated ZIP: {updated_zip_name}")
                print(f"{'='*60}")

                return True
        else:
            print(f"❌ No BM carryover data found to process")
            return False
    except Exception as e:
        print(f"❌ Error processing BM carryover results: {e}")
        traceback.print_exc()
        return False
    finally:
        # Safe cleanup
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"✅ Cleaned up BM temporary files")
        # Report final status
        if updated_zip_path and os.path.exists(updated_zip_path):
            print(f"✅ BM UPDATED ZIP successfully created: {updated_zip_path}")
        else:
            print(f"⚠️ BM UPDATED ZIP was not created - check logs above")

# ============================================================
# Main Function
# ============================================================
def main():
    """Main function - FIXED for web interface"""
    print("=" * 60)
    print("🎯 BM CARRYOVER RESULT PROCESSOR")
    print("🌐 WEB INTERFACE COMPATIBLE VERSION")
    print("🔧 WITH ALL CRITICAL FIXES APPLIED")
    print("=" * 60)
    # Get environment variables
    set_name = os.getenv("SELECTED_SET", "")
    semester_key = os.getenv("SELECTED_SEMESTERS", "")
    resit_file_path = os.getenv("RESIT_FILE_PATH", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", str(DEFAULT_PASS_THRESHOLD)))
    print(f"\n📋 PARAMETERS:")
    print(f" Set: {set_name}")
    print(f" Semester: {semester_key}")
    print(f" Resit File: {resit_file_path}")
    print(f" Pass Threshold: {pass_threshold}")
    print(f" Base Dir: {BASE_DIR}")
    # Validate inputs
    if not set_name:
        print("❌ ERROR: SELECTED_SET not provided")
        sys.exit(1)
    if not semester_key:
        print("❌ ERROR: SELECTED_SEMESTERS not provided")
        sys.exit(1)
    if not resit_file_path or not os.path.exists(resit_file_path):
        print(f"❌ ERROR: Resit file not found: {resit_file_path}")
        sys.exit(1)
    # Validate BM set
    BM_SETS = ["SET2023", "SET2024", "SET2025"]
    if set_name not in BM_SETS:
        print(f"❌ ERROR: Invalid BM set: {set_name}")
        print(f"💡 Valid BM sets: {BM_SETS}")
        sys.exit(1)
    print(f"\n✅ Processing BM Set: {set_name}")
    print(f"✅ Processing Semester: {semester_key}")
    # Find clean directory
    clean_dir = get_output_directory(set_name)
    output_dir = clean_dir  # FIXED: Output to CLEAN_RESULTS
    print(f"📁 Output Directory: {output_dir}")
    # Find mastersheet source
    print(f"\n🔍 Looking for mastersheet source...")
    source_path, source_type = find_latest_mastersheet_source(clean_dir, set_name)
    if not source_path:
        print(f"❌ ERROR: No ZIP files or result folders found in {clean_dir}")
        print(f"💡 Run BM regular processor first")
        sys.exit(1)
    print(f"✅ Using source: {source_path}")
    print(f"✅ Source type: {source_type}")
    # Process carryover results
    print(f"\n🚀 Starting carryover processing...")
    success = process_carryover_results(
        resit_file_path=resit_file_path,
        source_path=source_path,
        source_type=source_type,
        semester_key=semester_key,
        set_name=set_name,
        pass_threshold=pass_threshold,
        output_dir=output_dir,
    )
    if success:
        print("\n" + "=" * 60)
        print("✅ BM CARRYOVER PROCESSING COMPLETED")
        print("=" * 60)
        print(f"📂 Check CLEAN_RESULTS for the ZIP file")
        print(f"💡 The file should now appear in Download Center")
        sys.exit(0)
    else:
        print("\n" + "=" * 60)
        print("❌ BM CARRYOVER PROCESSING FAILED")
        print("=" * 60)
        sys.exit(1)


# ============================================================
# Script Execution
# ============================================================
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Fatal error: {e}")
        traceback.print_exc()
        sys.exit(1)