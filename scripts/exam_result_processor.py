#!/usr/bin/env python3

"""
exam_result_processor.py

Complete script with ENFORCED probation/withdrawal rule, integrated carryover student management, and NOT REG detection.

WITH FIXED PREVIOUS CGPA LOADING LOGIC:
- Previous CGPA now loads from actual previous semester worksheet (not CGPA_SUMMARY)
- Previous CGPA correctly mapped by semester progression
- Proper CGPA calculation with 4.0 scale
"""

from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import re
import pandas as pd
from datetime import datetime
import platform
import difflib
import math
import glob
import tempfile
import shutil
import json
import subprocess
import numpy as np

# ----------------------------
# Configuration
# ----------------------------

def is_running_on_railway():
    """Check if we're running on Railway"""
    return any(
        key in os.environ
        for key in [
            "RAILWAY_ENVIRONMENT",
            "RAILWAY_STATIC_URL",
            "RAILWAY_PROJECT_ID",
            "RAILWAY_SERVICE_NAME",
        ]
    )

def get_base_directory():
    """Get base directory - compatible with both local and Railway environments"""
    # Check if BASE_DIR is explicitly set in environment (highest priority)
    base_dir_env = os.getenv("BASE_DIR")
    if base_dir_env:
        if os.path.exists(base_dir_env):
            print(f"✅ Using BASE_DIR from environment: {base_dir_env}")
            return base_dir_env
        else:
            print(
                f"⚠️ BASE_DIR from environment doesn't exist: {base_dir_env}, trying alternatives..."
            )
    # Check if we're running on Railway
    if is_running_on_railway():
        # Create the directory structure on Railway
        railway_base = "/app/EXAMS_INTERNAL"
        os.makedirs(railway_base, exist_ok=True)
        os.makedirs(os.path.join(railway_base, "ND", "ND-COURSES"), exist_ok=True)
        print(f"✅ Using Railway base directory: {railway_base}")
        return railway_base
    # Local development fallbacks - check multiple possible locations
    local_paths = [
        # Your specific structure
        os.path.join(
            os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"
        ),
        # Common development locations
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "EXAMS_INTERNAL"),
        os.path.join(os.getcwd(), "EXAMS_INTERNAL"),
        # Relative to script location
        os.path.join(os.path.dirname(__file__), "EXAMS_INTERNAL"),
    ]
    for local_path in local_paths:
        if os.path.exists(local_path):
            print(f"✅ Using local base directory: {local_path}")
            return local_path
    # Final fallback - create in current working directory
    fallback_path = os.path.join(os.getcwd(), "EXAMS_INTERNAL")
    print(f"⚠️ No existing directory found, creating fallback: {fallback_path}")
    os.makedirs(fallback_path, exist_ok=True)
    os.makedirs(os.path.join(fallback_path, "ND", "ND-COURSES"), exist_ok=True)
    return fallback_path

BASE_DIR = get_base_directory()
# UPDATED: ND directories now under ND folder
ND_BASE_DIR = os.path.join(BASE_DIR, "ND")
ND_COURSES_DIR = os.path.join(ND_BASE_DIR, "ND-COURSES")
# Ensure directories exist
os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(ND_BASE_DIR, exist_ok=True)
os.makedirs(ND_COURSES_DIR, exist_ok=True)
print(f"📁 Base directory: {BASE_DIR}")
print(f"📁 ND base directory: {ND_BASE_DIR}")
print(f"📁 ND courses directory: {ND_COURSES_DIR}")

# Global variables for threshold upgrade
THRESHOLD_UPGRADED = False
ORIGINAL_THRESHOLD = 50.0
UPGRADE_MIN = None
UPGRADE_MAX = 49

# Global carryover tracker
CARRYOVER_STUDENTS = {}

# Global inactive students tracker (kept for internal tracking but not displayed)
INACTIVE_STUDENTS = {}

def is_web_mode():
    """Check if running in web mode (file upload)"""
    return os.getenv("WEB_MODE") == "true"

def get_uploaded_file_path():
    """Get path of uploaded file in web mode"""
    return os.getenv("UPLOADED_FILE_PATH")

def should_use_interactive_mode():
    """Check if we should use interactive mode (CLI) or non-interactive mode (web)."""
    # If specific environment variables are set by web form, use non-interactive
    if os.getenv("SELECTED_SET") or os.getenv("PROCESSING_MODE") or is_web_mode():
        return False
    # If we're running in a terminal with stdin available, use interactive mode
    if sys.stdin.isatty():
        return True
    # Default to interactive for backward compatibility
    return True

def get_upgrade_threshold_from_env():
    """Get upgrade threshold from environment variables"""
    upgrade_threshold_str = os.getenv("UPGRADE_THRESHOLD", "0").strip()
    if upgrade_threshold_str and upgrade_threshold_str.isdigit():
        upgrade_value = int(upgrade_threshold_str)
        if 0 <= upgrade_value <= 49:
            return upgrade_value if upgrade_value > 0 else None
    return None

def get_form_parameters():
    """Get parameters from environment variables set by the web form."""
    selected_set = os.getenv("SELECTED_SET", "all")
    processing_mode = os.getenv("PROCESSING_MODE", "auto")
    selected_semesters_str = os.getenv("SELECTED_SEMESTERS", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", "50.0"))
    generate_pdf = os.getenv("GENERATE_PDF", "True").lower() == "true"
    track_withdrawn = os.getenv("TRACK_WITHDRAWN", "True").lower() == "true"
    
    # NEW: Check for carryover processing mode
    process_carryover = os.getenv("PROCESS_CARRYOVER", "False").lower() == "true"
    carryover_file_path = os.getenv("CARRYOVER_FILE_PATH", "")
    
    # Convert semester string to list - handle both comma-separated and single values
    selected_semesters = []
    if selected_semesters_str:
        if "," in selected_semesters_str:
            selected_semesters = [
                sem.strip() for sem in selected_semesters_str.split(",") if sem.strip()
            ]
        else:
            selected_semesters = [selected_semesters_str.strip()]
    
    # If no semesters selected or 'all' in selected, use all semesters
    if not selected_semesters or "all" in selected_semesters:
        selected_semesters = SEMESTER_ORDER.copy()
    
    # FIX: Convert selected semesters to UPPERCASE to match course data
    selected_semesters = [sem.upper() for sem in selected_semesters]
    
    print(f"🎯 FORM PARAMETERS:")
    print(f" Selected Set: {selected_set}")
    print(f" Processing Mode: {processing_mode}")
    print(f" Selected Semesters: {selected_semesters}")
    print(f" Pass Threshold: {pass_threshold}")
    print(f" Generate PDF: {generate_pdf}")
    print(f" Track Withdrawn: {track_withdrawn}")
    print(f" Process Carryover: {process_carryover}")
    print(f" Carryover File Path: {carryover_file_path}")
    
    return {
        "selected_set": selected_set,
        "processing_mode": processing_mode,
        "selected_semesters": selected_semesters,
        "pass_threshold": pass_threshold,
        "generate_pdf": generate_pdf,
        "track_withdrawn": track_withdrawn,
        "process_carryover": process_carryover,
        "carryover_file_path": carryover_file_path,
    }

def get_pass_threshold():
    """Get pass threshold - now handles upgrade logic interactively."""
    threshold_str = os.getenv("PASS_THRESHOLD", "50.0")
    try:
        threshold = float(threshold_str)
    except ValueError:
        threshold = 50.0
    return threshold

DEFAULT_PASS_THRESHOLD = get_pass_threshold()
TIMESTAMP_FMT = "%Y-%m-%d_%H%M%S"

# FIXED: More robust logo path detection
def get_logo_path():
    """Get the logo path with multiple fallback options."""
    possible_paths = [
        # Relative to script
        os.path.normpath(
            os.path.join(
                os.path.dirname(__file__), "..", "launcher", "static", "logo.png"
            )
        ),
        # Common locations
        os.path.join(os.path.dirname(__file__), "logo.png"),
        os.path.join(os.getcwd(), "logo.png"),
        # Absolute path fallback
        "/app/launcher/static/logo.png", # For Railway deployment
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            print(f"✅ Found logo at: {path}")
            return path
    
    print("⚠️ Logo not found, PDF generation will proceed without logo")
    return None

DEFAULT_LOGO_PATH = get_logo_path()
NAME_WIDTH_CAP = 40

# Define semester processing order - FIXED: Use consistent uppercase
SEMESTER_ORDER = [
    "ND-FIRST-YEAR-FIRST-SEMESTER",
    "ND-FIRST-YEAR-SECOND-SEMESTER",
    "ND-SECOND-YEAR-FIRST-SEMESTER",
    "ND-SECOND-YEAR-SECOND-SEMESTER",
]

# Global student tracker
STUDENT_TRACKER = {}
WITHDRAWN_STUDENTS = {}

# ----------------------------
# NEW: Inactive Student Detection Functions - FIXED VERSION
# ----------------------------

def initialize_inactive_students_tracker():
    """Initialize the global inactive students tracker."""
    global INACTIVE_STUDENTS
    INACTIVE_STUDENTS = {}

def identify_inactive_students():
    """
    Identify students who were not withdrawn but didn't sit for subsequent semester examinations.
    These students appear in earlier semesters but are missing from later ones, yet have status in CGPA summary.
    FIXED: Properly track names and all missing semesters
    """
    global INACTIVE_STUDENTS, STUDENT_TRACKER, WITHDRAWN_STUDENTS
    
    print(f"\n🔍 IDENTIFYING INACTIVE STUDENTS (Not Withdrawn but Missing from Subsequent Semesters)")
    print("=" * 80)
    
    inactive_students = {}
    
    # Track all semesters we've processed
    all_processed_semesters = set()
    for student_data in STUDENT_TRACKER.values():
        all_processed_semesters.update(student_data["semesters_present"])
    
    # Sort semesters in processing order
    processed_semesters_ordered = [sem for sem in SEMESTER_ORDER if sem in all_processed_semesters]
    
    if len(processed_semesters_ordered) < 2:
        print("ℹ️ Need at least 2 semesters to identify inactive students")
        return {}
    
    print(f"📊 Processed semesters in order: {processed_semesters_ordered}")
    
    for exam_no, student_data in STUDENT_TRACKER.items():
        # Skip withdrawn students
        if exam_no in WITHDRAWN_STUDENTS:
            continue
            
        semesters_present = student_data["semesters_present"]
        first_semester = student_data["first_seen"]
        last_semester = student_data["last_seen"]
        
        # Check if student is missing any semesters between first and last appearance
        first_index = processed_semesters_ordered.index(first_semester)
        last_index = processed_semesters_ordered.index(last_semester)
        
        expected_semesters = processed_semesters_ordered[first_index:last_index + 1]
        missing_semesters = [sem for sem in expected_semesters if sem not in semesters_present]
        
        if missing_semesters:
            # FIX 1: Ensure name is properly captured
            student_name = student_data.get("name", "Unknown")
            if student_name == "Unknown" and "name" in student_data:
                student_name = student_data["name"]
                
            inactive_students[exam_no] = {
                "name": student_name,
                "first_semester": first_semester,
                "last_semester": last_semester,
                "semesters_present": semesters_present,
                "missing_semesters": missing_semesters,  # FIX 2: Store all missing semesters
                "status": student_data.get("status", "Active"),
                "current_gpa": student_data.get("current_gpa", "N/A"),
                "cgpa_status": student_data.get("cgpa_status", "N/A"),
                "identified_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            print(f"⚠️ INACTIVE: {exam_no} - {student_name} - Present in {len(semesters_present)} semesters, missing {len(missing_semesters)}: {missing_semesters}")
    
    # Also check for students who have CGPA status but are missing from recent semesters
    for exam_no, student_data in STUDENT_TRACKER.items():
        if exam_no in inactive_students or exam_no in WITHDRAWN_STUDENTS:
            continue
            
        # Check if student has CGPA data but is missing from the most recent semester
        if student_data.get("has_cgpa_data", False):
            most_recent_semester = processed_semesters_ordered[-1]
            if most_recent_semester not in student_data["semesters_present"]:
                # FIX 3: Check if student already in inactive list and append missing semester
                if exam_no in inactive_students:
                    # Add to existing missing semesters if not already there
                    if most_recent_semester not in inactive_students[exam_no]["missing_semesters"]:
                        inactive_students[exam_no]["missing_semesters"].append(most_recent_semester)
                else:
                    # Create new entry
                    student_name = student_data.get("name", "Unknown")
                    if student_name == "Unknown" and "name" in student_data:
                        student_name = student_data["name"]
                        
                    inactive_students[exam_no] = {
                        "name": student_name,
                        "first_semester": student_data["first_seen"],
                        "last_semester": student_data["last_seen"],
                        "semesters_present": student_data["semesters_present"],
                        "missing_semesters": [most_recent_semester],
                        "status": student_data.get("status", "Active"),
                        "current_gpa": student_data.get("current_gpa", "N/A"),
                        "cgpa_status": student_data.get("cgpa_status", "Active in CGPA"),
                        "identified_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                print(f"🎓 CGPA ACTIVE BUT MISSING: {exam_no} - {student_name} - Has CGPA but missing from {most_recent_semester}")
    
    INACTIVE_STUDENTS = inactive_students
    print(f"\n📊 INACTIVE STUDENTS SUMMARY:")
    print(f" Total inactive students identified: {len(INACTIVE_STUDENTS)}")
    
    # Breakdown by type
    cgpa_active_count = sum(1 for s in inactive_students.values() if s.get("cgpa_status") == "Active in CGPA")
    regular_inactive_count = len(inactive_students) - cgpa_active_count
    
    print(f" - Regular inactive (missing intermediate semesters): {regular_inactive_count}")
    print(f" - CGPA active but missing from recent semester: {cgpa_active_count}")
    
    return inactive_students

def abbreviate_semester_name(semester_name):
    """Abbreviate semester names to Y1S1, Y1S2, Y2S1, Y2S2 format"""
    abbreviations = {
        "ND-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
        "ND-FIRST-YEAR-SECOND-SEMESTER": "Y1S2",
        "ND-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
        "ND-SECOND-YEAR-SECOND-SEMESTER": "Y2S2"
    }
    return abbreviations.get(semester_name, semester_name)

# ----------------------------
# NOT REG Detection and Handling Functions
# ----------------------------

def detect_not_registered_content(cell_value):
    """Detect if a cell contains NOT REG or similar content indicating non-registration."""
    if pd.isna(cell_value) or cell_value == "":
        return False
    
    cell_str = str(cell_value).strip().upper()
    not_reg_patterns = [
        "NOT REG", "NOT REGISTERED", "NOT-REG", "NOT_REG",
        "NOT REGISTERED FOR COURSE", "NO REG", "NOT ENROLLED",
        "NOT TAKING", "NOT OFFERED", "NOT ATTEMPTED"
    ]
    
    for pattern in not_reg_patterns:
        if pattern in cell_str:
            return True
    return False

def process_not_registered_scores(df, course_columns):
    """
    Process NOT REG content in the dataframe.
    Returns: (processed_df, not_reg_counts_per_course)
    """
    not_reg_counts = {course: 0 for course in course_columns}
    
    for course in course_columns:
        if course in df.columns:
            # Convert column to object type first to allow mixed types
            if df[course].dtype != 'object':
                df[course] = df[course].astype('object')
                
            for idx in df.index:
                cell_value = df.at[idx, course]
                if detect_not_registered_content(cell_value):
                    # Replace NOT REG content with a special marker
                    df.at[idx, course] = "NOT REG"
                    not_reg_counts[course] += 1
                    
    return df, not_reg_counts

def calculate_course_statistics(mastersheet, ordered_codes, pass_threshold):
    """
    Calculate course statistics excluding NOT REG students.
    Returns: (fails_per_course, not_reg_per_course, registered_students_per_course)
    """
    fails_per_course = {}
    not_reg_per_course = {}
    registered_students_per_course = {}
    
    for code in ordered_codes:
        if code in mastersheet.columns:
            # Count registered students (those with actual scores, not "NOT REG")
            registered_mask = mastersheet[code].apply(
                lambda x: not detect_not_registered_content(x) and pd.notna(x) and x != ""
            )
            registered_students = registered_mask.sum()
            registered_students_per_course[code] = registered_students
            
            # Count NOT REG students
            not_reg_mask = mastersheet[code].apply(detect_not_registered_content)
            not_reg_count = not_reg_mask.sum()
            not_reg_per_course[code] = not_reg_count
            
            # Count failures only among registered students
            if registered_students > 0:
                # Convert to numeric, excluding NOT REG students
                scores = pd.to_numeric(mastersheet[code][registered_mask], errors='coerce')
                fail_count = (scores < pass_threshold).sum()
                fails_per_course[code] = int(fail_count)
            else:
                fails_per_course[code] = 0
                
    return fails_per_course, not_reg_per_course, registered_students_per_course

def calculate_gpa_with_not_reg(row, ordered_codes, filtered_credit_units):
    """
    Calculate GPA excluding NOT REG courses.
    """
    tcpe = 0.0
    total_registered_cu = 0
    
    for code in ordered_codes:
        score = row.get(code)
        
        # Skip NOT REG courses
        if detect_not_registered_content(score):
            continue
            
        try:
            score_val = float(score) if pd.notna(score) and score != "" else 0
            cu = filtered_credit_units.get(code, 0)
            gp = get_grade_point(score_val)
            tcpe += gp * cu
            total_registered_cu += cu
        except (ValueError, TypeError):
            continue
            
    return round((tcpe / total_registered_cu), 2) if total_registered_cu > 0 else 0.0, total_registered_cu

# ----------------------------
# DATA TRANSFORMATION FUNCTIONS - NEW ADDITION
# ----------------------------

def transform_transposed_data(df, sheet_type):
    """
    Transform transposed data format to wide format.
    Input: Each student appears multiple times with different courses
    Output: Each student appears once with all courses as columns
    """
    print(f"🔄 Transforming {sheet_type} sheet from transposed to wide format...")
    
    # Find the registration and name columns
    reg_col = find_column_by_names(
        df, ["REG. No", "Reg No", "Registration Number", "EXAM NUMBER"]
    )
    name_col = find_column_by_names(df, ["NAME", "Full Name", "Candidate Name"])
    
    if not reg_col:
        print("❌ Could not find registration column for transformation")
        return df
    # Get all course columns (columns that contain course codes)
    course_columns = [
        col
        for col in df.columns
        if col not in [reg_col, name_col] and col not in ["", None]
    ]
    
    print(f"📊 Found {len(course_columns)} course columns: {course_columns}")
    # Create a new dataframe to store transformed data
    transformed_data = []
    student_dict = {}
    # Process each row
    for idx, row in df.iterrows():
        exam_no = str(row[reg_col]).strip()
        student_name = (
            str(row[name_col]).strip()
            if name_col and pd.notna(row.get(name_col))
            else ""
        )
        
        if exam_no not in student_dict:
            student_dict[exam_no] = {"REG. No": exam_no, "NAME": student_name}
            
        # Add course scores for this student
        for course_col in course_columns:
            score = row.get(course_col)
            if pd.notna(score) and score != "" and score != " ":
                # Create column name with sheet type suffix
                column_name = f"{course_col}_{sheet_type}"
                student_dict[exam_no][column_name] = score
    # Convert dictionary to list
    transformed_data = list(student_dict.values())
    
    # Create new DataFrame
    if transformed_data:
        transformed_df = pd.DataFrame(transformed_data)
        print(
            f"✅ Transformed data: {len(transformed_df)} students, {len(transformed_df.columns)} columns"
        )
        return transformed_df
    else:
        print("❌ No data after transformation")
        return df

def detect_data_format(df, sheet_type):
    """
    Detect if data is in transposed format (students appear multiple times)
    Returns True if transposed format is detected
    """
    reg_col = find_column_by_names(
        df, ["REG. No", "Reg No", "Registration Number", "EXAM NUMBER"]
    )
    if not reg_col:
        return False
    # Count occurrences of each student
    student_counts = df[reg_col].value_counts()
    max_occurrences = student_counts.max()
    
    # If any student appears more than one, it's likely transposed format
    if max_occurrences > 1:
        print(f"📊 Data format detection for {sheet_type}:")
        print(f" Total students: {len(student_counts)}")
        print(f" Max occurrences per student: {max_occurrences}")
        print(f" Students with multiple entries: {(student_counts > 1).sum()}")
        return True
    
    return False

# ----------------------------
# Enhanced Course Name Matching Functions
# ----------------------------

def normalize_course_name(name):
    """Enhanced normalization for course title matching with better variations handling."""
    if not isinstance(name, str):
        return ""
    
    # Convert to lowercase and remove extra spaces
    normalized = name.lower().strip()
    # Replace multiple spaces with single space
    normalized = re.sub(r"\s+", " ", normalized)
    # Remove special characters and extra words
    normalized = re.sub(r"[^\w\s]", "", normalized)
    
    # Enhanced substitutions for variations
    substitutions = {
        "coomunication": "communication",
        "nsg": "nursing",
        "foundation": "foundations",
        "of of": "of", # handle double "of"
        "emergency care": "emergency",
        "nursing/ emergency": "nursing emergency",
        "care i": "care",
        "foundations of nursing": "foundations nursing",
        "foundation of nsg": "foundations nursing",
        "foundation of nursing": "foundations nursing",
    }
    
    for old, new in substitutions.items():
        normalized = normalized.replace(old, new)
    
    return normalized.strip()

def find_best_course_match(column_name, course_map):
    """Find the best matching course using enhanced matching algorithm.

    UPDATED: First try matching as code (handles headers like 'NUR221' or 'NUR221_CA').
    """
    if not isinstance(column_name, str):
        return None
    normalized_column = normalize_course_name(column_name)
    
    # NEW: Extract potential code (strip suffix like '_ca' if present)
    if ' ' in normalized_column:
        potential_code = normalized_column.split(' ')[0]
    elif '_' in normalized_column:
        potential_code = normalized_column.split('_')[0]
    else:
        potential_code = normalized_column
        
    potential_code = re.sub(r"[^a-z0-9]", "", potential_code) # Clean to alphanumeric
   
    # NEW: Create reverse map for code lookup (case-insensitive)
    code_to_info = {course_info["code"].lower(): course_info for course_info in course_map.values()}
   
    # NEW: Check if potential_code matches a known course code
    if potential_code in code_to_info:
        return code_to_info[potential_code]
   
    # Existing: exact match on normalized title
    if normalized_column in course_map:
        return course_map[normalized_column]
   
    # Existing: partial/contained matches
    for course_norm, course_info in course_map.items():
        if course_norm in normalized_column or normalized_column in course_norm:
            return course_info
   
    # Existing: word-based matching
    column_words = set(normalized_column.split())
    best_match = None
    best_score = 0
    
    for course_norm, course_info in course_map.items():
        course_words = set(course_norm.split())
        common_words = column_words.intersection(course_words)
        if common_words:
            score = len(common_words)
            key_words = ["foundation", "nursing", "emergency", "care", "communication", "anatomy", "physiology"]
            for word in key_words:
                if word in column_words and word in course_words:
                    score += 2
            if score > best_score:
                best_score = score
                best_match = course_info
                
    if best_match and best_score >= 2:
        return best_match
   
    # Existing: fuzzy matching fallback
    best_match = None
    best_ratio = 0
    
    for course_norm, course_info in course_map.items():
        ratio = difflib.SequenceMatcher(None, normalized_column, course_norm).ratio()
        if ratio > best_ratio and ratio > 0.6:
            best_ratio = ratio
            best_match = course_info
            
    return best_match

# ----------------------------
# Carryover Management Functions - UPDATED: No enhanced formatting
# ----------------------------

def initialize_carryover_tracker():
    """Initialize the global carryover tracker."""
    global CARRYOVER_STUDENTS
    CARRYOVER_STUDENTS = {}

def identify_carryover_students(
    mastersheet_df, semester_key, set_name, pass_threshold=50.0
):
    """
    Identify students with carryover courses from current semester processing.
    UPDATED: Includes both Resit and Probation students as carryover students
    """
    carryover_students = []
    
    # Get course columns (excluding student info columns)
    course_columns = [
        col
        for col in mastersheet_df.columns
        if col
        not in [
            "S/N",
            "EXAM NUMBER",
            "NAME",
            "FAILED COURSES",
            "REMARKS",
            "CU Passed",
            "CU Failed",
            "TCPE",
            "GPA",
            "AVERAGE",
        ]
    ]
    
    for idx, student in mastersheet_df.iterrows():
        failed_courses = []
        exam_no = str(student["EXAM NUMBER"])
        student_name = student["NAME"]
        remarks = str(student["REMARKS"])
        
        # Include both Resit and Probation students in carryover
        if remarks in ["Resit", "Probation"]:
            for course in course_columns:
                score = student.get(course, 0)
                try:
                    score_val = float(score) if pd.notna(score) else 0
                    if score_val < pass_threshold:
                        failed_courses.append(
                            {
                                "course_code": course,
                                "original_score": score_val,
                                "semester": semester_key,
                                "set": set_name,
                                "resit_attempts": 0,
                                "best_score": score_val,
                                "status": "Failed",
                            }
                        )
                except (ValueError, TypeError):
                    continue
                    
            if failed_courses:
                carryover_data = {
                    "exam_number": exam_no,
                    "name": student_name,
                    "failed_courses": failed_courses,
                    "semester": semester_key,
                    "set": set_name,
                    "identified_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_resit_attempts": 0,
                    "status": "Active",
                    "probation_status": remarks == "Probation", # Track if on probation
                }
                carryover_students.append(carryover_data)
                # Update global tracker
                student_key = f"{exam_no}_{semester_key}"
                CARRYOVER_STUDENTS[student_key] = carryover_data
                
    print(
        f"📊 Identified {len(carryover_students)} carryover students ({len([s for s in carryover_students if s['probation_status']])} on probation)"
    )
    return carryover_students

def save_carryover_records(carryover_students, output_dir, set_name, semester_key):
    """
    Save carryover student records to the clean results folder.
    UPDATED: SIMPLE Excel structure WITHOUT enhanced formatting
    """
    if not carryover_students:
        print("ℹ️ No carryover students to save")
        return None
    # Create carryover subdirectory in clean results
    carryover_dir = os.path.join(output_dir, "CARRYOVER_RECORDS")
    os.makedirs(carryover_dir, exist_ok=True)
    # Generate filename with set and semester tags
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"co_student_{set_name}_{semester_key}_{timestamp}"
    # Save as Excel
    excel_file = os.path.join(carryover_dir, f"{filename}.xlsx")
    
    # Prepare data for Excel - simple structure without enhanced formatting
    records_data = []
    for student in carryover_students:
        for course in student["failed_courses"]:
            records_data.append(
                {
                    "EXAM NUMBER": student["exam_number"],
                    "NAME": student["name"],
                    "COURSE CODE": course["course_code"],
                    "ORIGINAL SCORE": course["original_score"],
                    "SEMESTER": student["semester"],
                    "SET": student["set"],
                    "RESIT ATTEMPTS": course["resit_attempts"],
                    "BEST SCORE": course["best_score"],
                    "STATUS": course["status"],
                    "IDENTIFIED DATE": student["identified_date"],
                }
            )
    if records_data:
        df = pd.DataFrame(records_data)
        df.to_excel(excel_file, index=False)
        print(f"✅ Carryover records saved: {excel_file}")
        
        # UPDATED: NO enhanced formatting - keep it simple for regular processing
        try:
            # Just add basic header formatting without logo or title
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Simple header formatting only
            header_row = ws[1]
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                
            # Auto-adjust column widths for readability
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            wb.save(excel_file)
            print("✅ Added basic formatting to carryover Excel file")
        except Exception as e:
            print(f"⚠️ Could not add basic formatting to carryover file: {e}")
    # Save as JSON for easy processing
    json_file = os.path.join(carryover_dir, f"{filename}.json")
    with open(json_file, "w") as f:
        json.dump(carryover_students, f, indent=2)
        
    print(f"📁 Regular carryover records saved in: {carryover_dir}")
    return carryover_dir

def check_existing_carryover_files(raw_dir, set_name, semester_key):
    """
    Check if carryover files already exist for a given set and semester.
    Returns list of existing carryover file paths.
    """
    carryover_dir = os.path.join(raw_dir, "CARRYOVER")
    if not os.path.exists(carryover_dir):
        return []
    # Look for carryover files matching the pattern
    pattern = f"CARRYOVER-{semester_key}-{set_name}*.xlsx"
    existing_files = []
    for file in os.listdir(carryover_dir):
        if file.startswith(f"CARRYOVER-{semester_key}-{set_name}") and file.endswith(
            ".xlsx"
        ):
            existing_files.append(os.path.join(carryover_dir, file))
            
    print(
        f"🔍 Found {len(existing_files)} existing carryover files for {set_name}/{semester_key}"
    )
    return existing_files

# ----------------------------
# CGPA Tracking Functions - FIXED: Proper GPA vs CGPA terminology
# UPDATED: Recalculated for 4.0 scale and proper sorting
# ----------------------------

def get_grade_point(score):
    """Convert score to grade point for GPA calculation - ND 4.0 SCALE."""
    try:
        score = float(score)
        if score >= 70:
            return 4.0  # A
        elif score >= 60:
            return 3.0  # B
        elif score >= 50:
            return 2.0  # C
        elif score >= 45:
            return 1.0  # D
        else:
            return 0.0  # F
    except BaseException:
        return 0.0

def calculate_cumulative_cgpa(student_data, current_gpa, current_credits):
    """
    Calculate Cumulative CGPA based on all previous semesters and current semester.
    UPDATED: Proper calculation with weighted credits
    """
    if not student_data or "gpas" not in student_data or "credits" not in student_data:
        return current_gpa  # No previous data, return current GPA
        
    if len(student_data["gpas"]) == 0 or len(student_data["credits"]) == 0:
        return current_gpa
        
    total_grade_points = 0.0
    total_credits = 0
    
    # Add previous semesters
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
    
    # Add current semester
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits
    
    if total_credits > 0:
        cumulative_cgpa = total_grade_points / total_credits
        return round(cumulative_cgpa, 2)
    else:
        return current_gpa

def create_cgpa_summary_sheet(mastersheet_path, timestamp):
    """
    Create a CGPA summary sheet that aggregates GPA across all semesters.
    FIXED: Proper header row detection based on actual worksheet structure
    UPDATED: Added S/N column and fixed creation logic
    UPDATED: Professional column headings (Y1S1, Y1S2, Y2S1, Y2S2) and color-coded withdrawn status
    UPDATED TITLE: Changed to match the requested format
    UPDATED: Added CLASS OF AWARD column with the specified criteria
    FIXED: Dynamic GPA column detection and proper data extraction
    UPDATED: Auto-fit column width for all columns and subtle color coding for CLASS OF AWARD
    UPDATED: INACTIVE class of award for students with less than 4 valid semester GPAs
    UPDATED: Recalculated all GPAs for 4.0 scale and proper sorting
    FIXED: Freeze headings (rows 1-5), proper serial numbering, and auto-fit columns
    UPDATED: NAME and PROBATION HISTORY left aligned, abbreviated probation history
    """
    try:
        print("📊 Creating CGPA Summary Sheet...")
        
        # Load the mastersheet workbook
        wb = load_workbook(mastersheet_path)
        
        # Collect CGPA data from all semesters
        cgpa_data = {}
        
        # Map full semester names to abbreviated professional headings
        semester_abbreviation_map = {
            "ND-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
            "ND-FIRST-YEAR-SECOND-SEMESTER": "Y1S2",
            "ND-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
            "ND-SECOND-YEAR-SECOND-SEMESTER": "Y2S2"
        }
        # Track which semesters we actually found data for
        semesters_with_data = set()
        for sheet_name in wb.sheetnames:
            if sheet_name in SEMESTER_ORDER:
                try:
                    print(f"🔍 Processing sheet: {sheet_name}")
                    
                    # DYNAMIC APPROACH: Try multiple header rows to find the actual data
                    best_header_row = None
                    best_gpa_col = None
                    best_exam_col = None
                    best_df = None
                    
                    # Try header rows from 0 to 20 (covering all possible positions)
                    for header_row in range(0, 21):
                        try:
                            df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=header_row)
                            
                            # Skip if dataframe is empty or has very few columns
                            if df.empty or len(df.columns) < 3:
                                continue
                                
                            # Look for exam number column
                            exam_col = find_exam_number_column(df)
                            if not exam_col:
                                continue
                                
                            # Look for GPA column with multiple possible patterns
                            gpa_col = None
                            for col in df.columns:
                                col_str = str(col).upper().strip()
                                # More flexible GPA column detection
                                if any(pattern in col_str for pattern in ["GPA", "GRADE POINT", "POINT"]):
                                    gpa_col = col
                                    break
                            
                            if not gpa_col:
                                continue
                            
                            # Validate that we have actual data in these columns
                            valid_students = 0
                            sample_gpas = []
                            
                            for idx, row in df.iterrows():
                                exam_no = str(row[exam_col]).strip()
                                gpa_val = row[gpa_col]
                                
                                # Check for valid exam number and GPA
                                if (exam_no and exam_no != "nan" and exam_no != "" and 
                                    not exam_no.lower().startswith(("exam", "reg", "registration")) and
                                    len(exam_no) >= 5):  # Reasonable exam number length
                                    
                                    if pd.notna(gpa_val) and str(gpa_val).strip() != "":
                                        try:
                                            gpa_float = float(gpa_val)
                                            if 0 <= gpa_float <= 5.0:  # Valid GPA range
                                                valid_students += 1
                                                sample_gpas.append((exam_no, gpa_float))
                                                if valid_students >= 5:  # Found enough valid data
                                                    break
                                        except (ValueError, TypeError):
                                            continue
                            
                            if valid_students >= 3:  # Require at least 3 valid students
                                best_header_row = header_row
                                best_gpa_col = gpa_col
                                best_exam_col = exam_col
                                best_df = df
                                print(f"✅ Found valid data at header row {header_row}: {valid_students} students with GPA")
                                print(f"   Sample: {sample_gpas[:3]}")  # Show first 3 samples
                                break
                                    
                        except Exception as e:
                            # Continue to next header row if this one fails
                            continue
                    
                    # Process the data if we found a valid configuration
                    if best_header_row is not None and best_df is not None:
                        df = best_df
                        exam_col = best_exam_col
                        gpa_col = best_gpa_col
                        
                        print(f"📊 Processing {sheet_name} with header row {best_header_row}")
                        print(f"📝 Using columns - Exam: '{exam_col}', GPA: '{gpa_col}'")
                        
                        # Also look for name and remarks columns
                        name_col = None
                        remarks_col = None
                        for col in df.columns:
                            col_str = str(col).upper().strip()
                            if "NAME" in col_str and "COURSE" not in col_str:
                                name_col = col
                            elif "REMARKS" in col_str:
                                remarks_col = col
                        
                        students_found = 0
                        for idx, row in df.iterrows():
                            exam_no = str(row[exam_col]).strip()
                            
                            # Validate exam number format
                            if (exam_no and exam_no != "nan" and exam_no != "" and 
                                not exam_no.lower().startswith(("exam", "reg", "registration")) and
                                len(exam_no) >= 5):
                                
                                if exam_no not in cgpa_data:
                                    cgpa_data[exam_no] = {
                                        "name": (
                                            str(row[name_col]) if name_col and pd.notna(row.get(name_col))
                                            else ""
                                        ),
                                        "gpas": {},
                                        "status": "Active",
                                        "probation_semesters": [],
                                    }
                                
                                # Extract GPA value and convert to 4.0 scale if needed
                                gpa_value = row[gpa_col]
                                if pd.notna(gpa_value) and str(gpa_value).strip() != "":
                                    try:
                                        gpa_float = float(gpa_value)
                                        # Convert from 5.0 scale to 4.0 scale if necessary
                                        if gpa_float > 4.0:  # Assume it's in 5.0 scale
                                            gpa_float = (gpa_float / 5.0) * 4.0
                                        if 0 <= gpa_float <= 4.0:  # Valid GPA range for 4.0 scale
                                            cgpa_data[exam_no]["gpas"][sheet_name] = round(gpa_float, 2)
                                            students_found += 1
                                            semesters_with_data.add(sheet_name)
                                            
                                            # Update student tracker with CGPA data
                                            if exam_no in STUDENT_TRACKER:
                                                STUDENT_TRACKER[exam_no]["has_cgpa_data"] = True
                                                STUDENT_TRACKER[exam_no]["current_gpa"] = gpa_float
                                                STUDENT_TRACKER[exam_no]["cgpa_status"] = "Active in CGPA"
                                    except (ValueError, TypeError):
                                        continue
                                
                                # Track probation status - ABBREVIATE SEMESTER NAMES
                                if remarks_col and pd.notna(row.get(remarks_col)):
                                    remarks = str(row[remarks_col])
                                    if (remarks == "Probation" and 
                                        sheet_name not in cgpa_data[exam_no]["probation_semesters"]):
                                        # Use abbreviated semester name for probation history
                                        abbrev_semester = semester_abbreviation_map.get(sheet_name, sheet_name)
                                        cgpa_data[exam_no]["probation_semesters"].append(abbrev_semester)
                        
                        print(f"📊 Extracted GPA data for {students_found} students in {sheet_name}")
                        
                    else:
                        print(f"❌ Could not find valid GPA data in {sheet_name} after trying all header rows")
                        # Try to debug by showing available columns from first few rows
                        try:
                            for test_row in [0, 1, 2, 3, 4, 5]:
                                test_df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=test_row)
                                if not test_df.empty:
                                    print(f"   Row {test_row} columns: {[str(col) for col in test_df.columns[:8]]}")
                                    # Show sample data
                                    if len(test_df) > 0:
                                        sample = {}
                                        for col in test_df.columns[:5]:
                                            sample[col] = test_df.iloc[0][col]
                                        print(f"   Sample data: {sample}")
                                    break
                        except Exception as e:
                            print(f"   Debug failed: {e}")
                        
                except Exception as e:
                    print(f"⚠️ Warning: Could not process sheet {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
        print(f"📊 Total students with CGPA data: {len(cgpa_data)}")
        print(f"📊 Semesters with data: {semesters_with_data}")
        # Create CGPA summary dataframe with probation tracking
        summary_data = []
        for exam_no, data in cgpa_data.items():
            row = {
                "EXAM NUMBER": exam_no,
                "NAME": data["name"],
                "PROBATION HISTORY": (
                    ", ".join(data["probation_semesters"])
                    if data["probation_semesters"]
                    else "None"
                ),
            }
            
            # Add GPA for each semester using abbreviated headings and calculate cumulative
            total_gpa = 0
            semester_count = 0
            for semester in SEMESTER_ORDER:
                if semester in data["gpas"]:
                    # Use abbreviated column name
                    abbrev_semester = semester_abbreviation_map.get(semester, semester)
                    row[abbrev_semester] = data["gpas"][semester]
                    if pd.notna(data["gpas"][semester]):
                        total_gpa += data["gpas"][semester]
                        semester_count += 1
                else:
                    abbrev_semester = semester_abbreviation_map.get(semester, semester)
                    row[abbrev_semester] = None
                    
            # Calculate Cumulative CGPA
            cumulative_cgpa = (
                round(total_gpa / semester_count, 2) if semester_count > 0 else 0.0
            )
            row["CUMULATIVE CGPA"] = cumulative_cgpa
            
            # FIX 1: Calculate CLASS OF AWARD with WITHDRAWN check FIRST
            is_withdrawn = is_student_withdrawn(exam_no)
            valid_semester_count = semester_count  # Count of semesters with valid GPA data
            # CRITICAL: Check withdrawal status FIRST
            if is_withdrawn:
                class_of_award = "WITHDRAWN"  # Withdrawn students should show WITHDRAWN, not Pass/Fail
            elif not is_withdrawn and valid_semester_count < 4:
                class_of_award = "INACTIVE"
            else:
                # Normal award calculation based on CGPA (4.0 scale)
                if cumulative_cgpa >= 3.5:
                    class_of_award = "Distinction"
                elif cumulative_cgpa >= 3.0:
                    class_of_award = "Upper Credit"
                elif cumulative_cgpa >= 2.0:
                    class_of_award = "Lower Credit"
                elif cumulative_cgpa >= 1.0:
                    class_of_award = "Pass"
                else:
                    class_of_award = "Fail"
                    
            row["CLASS OF AWARD"] = class_of_award
            row["WITHDRAWN"] = "Yes" if is_withdrawn else "No"
            summary_data.append(row)
        # Create summary dataframe
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            
            # UPDATED: Sort the data as requested
            def get_sort_key(row):
                award = row["CLASS OF AWARD"]
                if award in ["Distinction", "Upper Credit", "Lower Credit"]:
                    return (0, -row["CUMULATIVE CGPA"])  # Active students first
                elif award == "INACTIVE":
                    return (1, 0)
                elif award == "WITHDRAWN":
                    return (3, 0)  # Withdrawn students at the end
                elif award == "Fail":
                    return (2, 0)
                else:  # Pass and others
                    return (0, -row["CUMULATIVE CGPA"])
            
            # Create a temporary sort key column
            sort_keys = summary_df.apply(get_sort_key, axis=1)
            summary_df = summary_df.iloc[sort_keys.argsort()].reset_index(drop=True)
            
            # FIX 2: Add proper serial numbering after sorting
            summary_df.insert(0, "S/N", range(1, len(summary_df) + 1))
            
            print(f"✅ Successfully created CGPA summary with {len(summary_df)} students")
            # Show sample of the summary data
            print("📋 Sample of CGPA summary data:")
            for i in range(min(5, len(summary_df))):
                student = summary_df.iloc[i]
                print(f"  {i+1}. {student['S/N']}. {student['EXAM NUMBER']}: CGPA={student.get('CUMULATIVE CGPA', 'N/A')}, Award={student.get('CLASS OF AWARD', 'N/A')}, Semesters={sum(1 for sem in SEMESTER_ORDER if pd.notna(student.get(semester_abbreviation_map.get(sem, sem))))}")
        else:
            print("⚠️ No CGPA data found for any students")
            # Create empty dataframe with correct columns
            headers = (
                ["S/N", "EXAM NUMBER", "NAME", "PROBATION HISTORY"]
                + [semester_abbreviation_map.get(sem, sem) for sem in SEMESTER_ORDER]
                + ["CUMULATIVE CGPA", "WITHDRAWN", "CLASS OF AWARD"]
            )
            summary_df = pd.DataFrame(columns=headers)
        # Add the summary sheet to the workbook
        if "CGPA_SUMMARY" in wb.sheetnames:
            del wb["CGPA_SUMMARY"]
        ws = wb.create_sheet("CGPA_SUMMARY")
        # Define headers for column calculation
        headers = (
            ["S/N", "EXAM NUMBER", "NAME", "PROBATION HISTORY"]
            + [semester_abbreviation_map.get(sem, sem) for sem in SEMESTER_ORDER]
            + ["CUMULATIVE CGPA", "WITHDRAWN", "CLASS OF AWARD"]
        )
        num_columns = len(headers)
        end_col_letter = get_column_letter(num_columns)
        # ADD PROFESSIONAL HEADER WITH SCHOOL NAME - FIXED: Use exact requested format
        ws.merge_cells(f"A1:{end_col_letter}1")
        title_cell = ws["A1"]
        title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )
        ws.merge_cells(f"A2:{end_col_letter}2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = "DEPARTMENT OF NURSING"
        subtitle_cell.font = Font(bold=True, size=14, color="000000")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        subtitle_cell.fill = PatternFill(
            start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
        )
        # ADD DYNAMIC TITLE ROW - FIXED: Use exact requested format
        ws.merge_cells(f"A3:{end_col_letter}3")
        exam_title_cell = ws["A3"]
        exam_title_cell.value = "NDII CGPA SUMMARY REPORT"
        exam_title_cell.font = Font(bold=True, size=12, color="000000")
        exam_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        exam_title_cell.fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
        )
        ws.merge_cells(f"A4:{end_col_letter}4")
        date_cell = ws["A4"]
        date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Add empty row for spacing
        ws.row_dimensions[5].height = 10
        # Write header starting from row 6
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        # Write data starting from row 7
        if not summary_df.empty:
            for row_idx, row_data in enumerate(summary_df.to_dict("records"), 7):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    if pd.isna(value):
                        value = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply basic styling to data cells
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
                    
                    # Alternate row coloring
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # FIX 1: LEFT ALIGN NAME AND PROBATION HISTORY COLUMNS
                    if col_idx == 1:  # S/N - center
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 2:  # EXAM NUMBER - center
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx in [3, 4]:  # NAME and PROBATION HISTORY - LEFT ALIGNED with wrap text
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:  # All other columns (numeric) - center
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                    # Apply subtle color coding for CLASS OF AWARD column
                    if header == "CLASS OF AWARD" and value:
                        if value == "Distinction":
                            cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Very light green
                        elif value == "Upper Credit":
                            cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")  # Very light blue
                        elif value == "Lower Credit":
                            cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # Very light yellow
                        elif value == "Pass":
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Very light gray
                        elif value == "Fail":
                            cell.fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")  # Very light red
                        elif value == "INACTIVE":  # FIX 1: Add color for INACTIVE
                            cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")  # Very light red/pink
                        elif value == "WITHDRAWN":  # FIX 1: Add color for WITHDRAWN
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light gray
            
            print(f"✅ Written {len(summary_df)} students to CGPA summary sheet")
        else:
            print("⚠️ No data to write to CGPA summary sheet")
            ws.cell(row=7, column=1, value="No CGPA data available")
        # FIX 1: FREEZE PANES AT ROW 7 (so rows 1-6 are frozen)
        ws.freeze_panes = "A7"
        print("✅ Frozen headings (rows 1-6)")
        # FIX 2: AUTO-FIT COLUMN WIDTHS FOR ALL COLUMNS WITH BETTER LOGIC
        print("📏 Auto-fitting column widths to fit content...")
        
        # Define minimum and maximum widths for better control
        min_width = 8
        max_width = 35
        
        # Enhanced auto-fit: Calculate maximum content length for each column
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Check header length first
            header_length = len(str(header))
            max_length = max(max_length, header_length)
            
            # Check data content lengths
            for row_idx in range(7, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        # For text columns, allow more width
                        if header in ["NAME", "PROBATION HISTORY"]:
                            cell_length = min(cell_length, 50)  # Cap very long text
                        max_length = max(max_length, cell_length)
                    except:
                        pass
            
            # Add padding and apply limits
            adjusted_width = min(max_length + 3, max_width)  # Increased padding to 3
            adjusted_width = max(adjusted_width, min_width)
            
            # Special handling for specific columns
            if header == "S/N":
                adjusted_width = 6
            elif header == "EXAM NUMBER":
                adjusted_width = min(max(adjusted_width, 15), 20)
            elif header == "NAME":
                adjusted_width = min(max(adjusted_width, 20), 35)
            elif header == "PROBATION HISTORY":
                adjusted_width = min(max(adjusted_width, 15), 25)  # Reduced width since we abbreviated
            elif header in ["Y1S1", "Y1S2", "Y2S1", "Y2S2"]:
                adjusted_width = min(max(adjusted_width, 8), 12)
            elif header == "CUMULATIVE CGPA":
                adjusted_width = min(max(adjusted_width, 12), 18)
            elif header == "WITHDRAWN":
                adjusted_width = min(max(adjusted_width, 10), 12)
            elif header == "CLASS OF AWARD":
                adjusted_width = min(max(adjusted_width, 12), 20)
            
            ws.column_dimensions[column_letter].width = adjusted_width
            print(f"   📐 Column {column_letter} ({header}): width {adjusted_width}")
        # Adjust row heights for better visibility
        for row in range(6, ws.max_row + 1):
            if row == 6:  # Header row
                ws.row_dimensions[row].height = 25
            else:
                ws.row_dimensions[row].height = 20
        # Add summary statistics if we have data
        if not summary_df.empty:
            stats_row = len(summary_df) + 8
            stats_cell = ws.cell(row=stats_row, column=1, value="SUMMARY STATISTICS")
            stats_cell.font = Font(bold=True, size=12)
            stats_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            if "CUMULATIVE CGPA" in summary_df.columns:
                # FIX 1: Count INACTIVE and WITHDRAWN students
                inactive_count = (summary_df['CLASS OF AWARD'] == 'INACTIVE').sum()
                withdrawn_count = (summary_df['CLASS OF AWARD'] == 'WITHDRAWN').sum()
                
                stats_data = [
                    f"Total Students: {len(summary_df)}",
                    f"Average Cumulative CGPA: {summary_df['CUMULATIVE CGPA'].mean():.2f}",
                    f"Highest Cumulative CGPA: {summary_df['CUMULATIVE CGPA'].max():.2f}",
                    f"Lowest Cumulative CGPA: {summary_df['CUMULATIVE CGPA'].min():.2f}",
                    f"Withdrawn Students: {withdrawn_count}",
                    f"INACTIVE Students (less than 4 semesters): {inactive_count}",
                    f"Students with Probation History: {(summary_df['PROBATION HISTORY'] != 'None').sum()}",
                ]
                
                # Add award distribution
                if "CLASS OF AWARD" in summary_df.columns:
                    award_counts = summary_df['CLASS OF AWARD'].value_counts()
                    stats_data.append("Award Distribution:")
                    for award, count in award_counts.items():
                        stats_data.append(f"  - {award}: {count} students")
                
                for i, stat in enumerate(stats_data, start=stats_row + 1):
                    cell = ws.cell(row=i, column=1, value=stat)
                    if stat.startswith("Award Distribution:"):
                        cell.font = Font(bold=True)
                    elif stat.startswith("  -"):
                        cell.font = Font(italic=True)
        # Save the workbook
        wb.save(mastersheet_path)
        print("✅ CGPA Summary sheet created successfully with:")
        print("   - Frozen headings (rows 1-6)")
        print("   - Proper serial numbering")
        print("   - Auto-fit column widths")
        print("   - NAME and PROBATION HISTORY left aligned")
        print("   - Abbreviated probation history (Y1S1, Y1S2, Y2S1, Y2S2)")
        return summary_df
        
    except Exception as e:
        print(f"❌ Error creating CGPA summary sheet: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_analysis_sheet(mastersheet_path, timestamp):
    """
    Create an analysis sheet with comprehensive statistics.
    FIXED: Dynamic header detection, robust column finding, accurate student counting, and proper serial numbering
    ENHANCED: Professional formatting with proper cell fitting and visual appeal
    UPDATED: REMOVED INACTIVE STUDENTS column as requested
    """
    try:
        print("📈 Creating Analysis Sheet...")
        wb = load_workbook(mastersheet_path)
        
        # Collect data from all semesters
        analysis_data = {
            "semester": [],
            "total_students": [],
            "passed_all": [],
            "resit_students": [],
            "probation_students": [],
            "withdrawn_students": [],
            "average_gpa": [],
            "pass_rate": [],
        }
        # Semester short name mapping
        semester_short_names = {
            "ND-FIRST-YEAR-FIRST-SEMESTER": "Y1-1STS",
            "ND-FIRST-YEAR-SECOND-SEMESTER": "Y1-2NDS", 
            "ND-SECOND-YEAR-FIRST-SEMESTER": "Y2-1STS",
            "ND-SECOND-YEAR-SECOND-SEMESTER": "Y2-2NDS"
        }
        for sheet_name in wb.sheetnames:
            if sheet_name not in SEMESTER_ORDER:
                continue
                
            try:
                print(f"\n🔍 Processing sheet: {sheet_name}")
                
                # ========================================
                # DYNAMIC HEADER ROW DETECTION
                # ========================================
                best_header_row = None
                best_df = None
                
                # Try different header rows (0 to 20)
                for header_row in range(0, 21):
                    try:
                        df = pd.read_excel(
                            mastersheet_path, 
                            sheet_name=sheet_name, 
                            header=header_row,
                            dtype=str
                        )
                        
                        if df.empty or len(df.columns) < 3:
                            continue
                        
                        # Check if we have the required columns
                        columns_upper = [str(col).upper().strip() for col in df.columns]
                        
                        # Look for exam number column
                        exam_col_found = any(
                            keyword in col_upper 
                            for col_upper in columns_upper 
                            for keyword in ["EXAM", "REG", "MATRIC", "STUDENT"]
                        )
                        
                        # Look for remarks column
                        remarks_col_found = any(
                            "REMARK" in col_upper 
                            for col_upper in columns_upper
                        )
                        
                        # Look for GPA column
                        gpa_col_found = any(
                            "GPA" in col_upper or "GRADE POINT" in col_upper
                            for col_upper in columns_upper
                        )
                        
                        if not (exam_col_found and remarks_col_found and gpa_col_found):
                            continue
                        
                        # Validate we have actual student data
                        exam_col = None
                        for col in df.columns:
                            col_str = str(col).upper().strip()
                            if any(keyword in col_str for keyword in ["EXAM", "REG", "MATRIC"]):
                                exam_col = col
                                break
                        
                        if not exam_col:
                            continue
                        
                        # Count valid exam numbers
                        valid_count = 0
                        for idx, row in df.iterrows():
                            exam_no = str(row[exam_col]).strip()
                            # Check if it looks like a real exam number
                            if (exam_no and 
                                exam_no != "nan" and 
                                exam_no != "" and
                                not exam_no.upper().startswith(("EXAM", "REG", "REGISTRATION", "STUDENT")) and
                                len(exam_no) >= 5):
                                valid_count += 1
                                if valid_count >= 3:  # Need at least 3 valid students
                                    break
                        
                        if valid_count >= 3:
                            best_header_row = header_row
                            best_df = df
                            print(f"✅ Found valid header at row {header_row} with {valid_count}+ students")
                            break
                            
                    except Exception as e:
                        continue
                
                if best_header_row is None or best_df is None:
                    print(f"❌ Could not find valid data structure in {sheet_name}")
                    continue
                
                df = best_df
                print(f"📊 Using header row {best_header_row}")
                
                # ========================================
                # DYNAMIC COLUMN DETECTION
                # ========================================
                
                # Find EXAM NUMBER column
                exam_col = None
                for col in df.columns:
                    col_upper = str(col).upper().strip()
                    if any(keyword in col_upper for keyword in [
                        "EXAM NUMBER", "EXAM NO", "REG NO", "REG. NO", 
                        "REGISTRATION", "MATRIC", "STUDENT ID"
                    ]):
                        exam_col = col
                        print(f"✅ Found EXAM column: '{col}'")
                        break
                
                # Find REMARKS column
                remarks_col = None
                for col in df.columns:
                    col_upper = str(col).upper().strip()
                    if "REMARK" in col_upper:
                        remarks_col = col
                        print(f"✅ Found REMARKS column: '{col}'")
                        break
                
                # Find GPA column
                gpa_col = None
                for col in df.columns:
                    col_upper = str(col).upper().strip()
                    if "GPA" in col_upper or "GRADE POINT" in col_upper:
                        gpa_col = col
                        print(f"✅ Found GPA column: '{col}'")
                        break
                
                # Validate we found all required columns
                if not exam_col:
                    print(f"❌ Could not find EXAM NUMBER column in {sheet_name}")
                    continue
                if not remarks_col:
                    print(f"❌ Could not find REMARKS column in {sheet_name}")
                    continue
                if not gpa_col:
                    print(f"❌ Could not find GPA column in {sheet_name}")
                    continue
                
                # ========================================
                # COUNT STUDENTS BY STATUS
                # ========================================
                
                # Use sets to track unique students
                all_students = set()
                passed_students = set()
                resit_students = set()
                probation_students = set()
                withdrawn_students_this_sem = set()
                gpa_values = []
                
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    
                    # Validate exam number
                    if not exam_no or exam_no == "nan" or exam_no == "":
                        continue
                    
                    # Skip header-like values
                    exam_no_upper = exam_no.upper()
                    if any(keyword in exam_no_upper for keyword in [
                        "EXAM", "REG", "REGISTRATION", "STUDENT", "MATRIC", "NUMBER"
                    ]):
                        continue
                    
                    # Must be reasonable length
                    if len(exam_no) < 5:
                        continue
                    
                    # Valid student found
                    all_students.add(exam_no)
                    
                    # Get remarks
                    remarks = str(row[remarks_col]).strip() if pd.notna(row.get(remarks_col)) else ""
                    
                    # Categorize by status
                    if remarks == "Passed":
                        passed_students.add(exam_no)
                    elif remarks == "Resit":
                        resit_students.add(exam_no)
                    elif remarks == "Probation":
                        probation_students.add(exam_no)
                    elif remarks == "Withdrawn":
                        withdrawn_students_this_sem.add(exam_no)
                    
                    # Collect GPA values
                    gpa_val = row[gpa_col]
                    if pd.notna(gpa_val):
                        try:
                            gpa_float = float(gpa_val)
                            if 0 <= gpa_float <= 5.0:
                                gpa_values.append(gpa_float)
                        except (ValueError, TypeError):
                            pass
                
                # Calculate statistics
                total_students = len(all_students)
                passed_all = len(passed_students)
                resit_count = len(resit_students)
                probation_count = len(probation_students)
                withdrawn_count = len(withdrawn_students_this_sem)
                
                avg_gpa = round(sum(gpa_values) / len(gpa_values), 2) if gpa_values else 0.0
                pass_rate = round((passed_all / total_students * 100), 2) if total_students > 0 else 0.0
                
                # Log results
                print(f"📊 {sheet_name} Statistics:")
                print(f"   Total Students: {total_students}")
                print(f"   Passed: {passed_all}")
                print(f"   Resit: {resit_count}")
                print(f"   Probation: {probation_count}")
                print(f"   Withdrawn: {withdrawn_count}")
                print(f"   Average GPA: {avg_gpa}")
                print(f"   Pass Rate: {pass_rate}%")
                
                # Add to analysis data
                short_semester = semester_short_names.get(sheet_name, sheet_name)
                analysis_data["semester"].append(short_semester)
                analysis_data["total_students"].append(total_students)
                analysis_data["passed_all"].append(passed_all)
                analysis_data["resit_students"].append(resit_count)
                analysis_data["probation_students"].append(probation_count)
                analysis_data["withdrawn_students"].append(withdrawn_count)
                analysis_data["average_gpa"].append(avg_gpa)
                analysis_data["pass_rate"].append(pass_rate)
                
            except Exception as e:
                print(f"⚠️ Error processing sheet {sheet_name}: {e}")
                import traceback
                traceback.print_exc()
                continue
        # ========================================
        # CREATE ANALYSIS DATAFRAME
        # ========================================
        
        if not analysis_data["semester"]:
            print("❌ No analysis data collected from any semester")
            # Create empty dataframe
            analysis_df = pd.DataFrame({
                "SEMESTER": [],
                "TOTAL STUDENTS": [],
                "PASSED ALL": [],
                "RESIT STUDENTS": [],
                "PROBATION STUDENTS": [],
                "WITHDRAWN STUDENTS": [],
                "AVERAGE GPA": [],
                "PASS RATE (%)": []
            })
        else:
            analysis_df = pd.DataFrame(analysis_data)
            
            # Rename columns to match expected format
            analysis_df.columns = [
                "SEMESTER",
                "TOTAL STUDENTS",
                "PASSED ALL",
                "RESIT STUDENTS",
                "PROBATION STUDENTS",
                "WITHDRAWN STUDENTS",
                "AVERAGE GPA",
                "PASS RATE (%)"
            ]
            
            # Add overall statistics
            overall_stats = {
                "SEMESTER": "OVERALL",
                "TOTAL STUDENTS": sum(analysis_data["total_students"]),
                "PASSED ALL": sum(analysis_data["passed_all"]),
                "RESIT STUDENTS": sum(analysis_data["resit_students"]),
                "PROBATION STUDENTS": sum(analysis_data["probation_students"]),
                "WITHDRAWN STUDENTS": sum(analysis_data["withdrawn_students"]),
                "AVERAGE GPA": round(sum(analysis_data["average_gpa"]) / len(analysis_data["average_gpa"]), 2) if analysis_data["average_gpa"] else 0.0,
                "PASS RATE (%)": round(sum(analysis_data["pass_rate"]) / len(analysis_data["pass_rate"]), 2) if analysis_data["pass_rate"] else 0.0,
            }
            analysis_df = pd.concat([analysis_df, pd.DataFrame([overall_stats])], ignore_index=True)
            
            print(f"\n✅ Analysis data collected for {len(analysis_data['semester'])} semesters")
        # Add serial number AFTER creating the dataframe
        analysis_df.insert(0, "S/N", range(1, len(analysis_df) + 1))
        # ========================================
        # CREATE ANALYSIS WORKSHEET
        # ========================================
        
        if "ANALYSIS" in wb.sheetnames:
            del wb["ANALYSIS"]
        ws = wb.create_sheet("ANALYSIS")
        headers = list(analysis_df.columns)  # Use actual dataframe columns
        num_columns = len(headers)
        end_col_letter = get_column_letter(num_columns)
        # Professional header with proper row heights
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A1:{end_col_letter}1")
        title_cell = ws["A1"]
        title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        ws.row_dimensions[2].height = 25
        ws.merge_cells(f"A2:{end_col_letter}2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = "DEPARTMENT OF NURSING"
        subtitle_cell.font = Font(bold=True, size=14, color="000000")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        subtitle_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.row_dimensions[3].height = 22
        ws.merge_cells(f"A3:{end_col_letter}3")
        exam_title_cell = ws["A3"]
        exam_title_cell.value = "NDII SEMESTER ANALYSIS REPORT"
        exam_title_cell.font = Font(bold=True, size=12, color="000000")
        exam_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        exam_title_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        ws.row_dimensions[4].height = 20
        ws.merge_cells(f"A4:{end_col_letter}4")
        date_cell = ws["A4"]
        date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Add spacing row
        ws.row_dimensions[5].height = 10
        # Write headers (row 6) with proper height
        ws.row_dimensions[6].height = 35
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        # Write data (starting row 7)
        for row_idx, (_, row_data) in enumerate(analysis_df.iterrows(), start=7):
            ws.row_dimensions[row_idx].height = 25
            
            for col_idx, header in enumerate(headers, 1):
                value = row_data[header]
                if pd.isna(value):
                    value = ""
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Styling
                is_overall = (row_data["SEMESTER"] == "OVERALL")
                
                if is_overall:
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    cell.font = Font(bold=True, size=10)
                else:
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.font = Font(size=10)
                
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Professional column widths
        column_widths = {
            'A': 8,   # S/N
            'B': 12,  # SEMESTER
            'C': 16,  # TOTAL STUDENTS
            'D': 14,  # PASSED ALL
            'E': 16,  # RESIT STUDENTS
            'F': 18,  # PROBATION STUDENTS
            'G': 18,  # WITHDRAWN STUDENTS
            'H': 14,  # AVERAGE GPA
            'I': 16,  # PASS RATE (%)
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        # Analysis notes with proper spacing
        notes_row = ws.max_row + 2
        ws.row_dimensions[notes_row].height = 25
        ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=num_columns)
        notes_cell = ws.cell(row=notes_row, column=1, value="ANALYSIS NOTES:")
        notes_cell.font = Font(bold=True, size=12, color="1E90FF")
        notes_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        notes_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        notes = [
            "• Resit Students: Students with GPA ≥ 2.0 who failed ≤45% of credits",
            "• Probation Students: Students with GPA < 2.0 OR failed >45% with GPA ≥ 2.0",
            "• Withdrawn Students: Students who failed >45% of credits with GPA < 2.0",
            "• Pass Rate: Percentage of students who passed all courses in the semester",
        ]
        
        for i, note in enumerate(notes):
            note_row = notes_row + i + 1
            ws.row_dimensions[note_row].height = 22
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=num_columns)
            note_cell = ws.cell(row=note_row, column=1, value=note)
            note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            note_cell.font = Font(size=10, italic=True)
        wb.save(mastersheet_path)
        print("✅ Analysis sheet created successfully with proper serial numbering and professional formatting")
        return analysis_df
        
    except Exception as e:
        print(f"❌ Error creating analysis sheet: {e}")
        import traceback
        traceback.print_exc()
        return None

# ----------------------------
# ENFORCED STUDENT STATUS DETERMINATION - ENFORCED PROBATION/WITHDRAWAL RULE
# ----------------------------

def determine_student_status(row, total_cu, pass_threshold):
    """
    ENFORCE the rule based on:
    | Category | GPA Condition | Credit Units Passed | Status |
    | -------- | ------------- | ------------------- | ------------------------------------------------------ |
    | 1 | GPA ≥ 2.00 | ≥ 45% | To **resit** failed courses next session |
    | 2 | GPA < 2.00 | ≥ 45% | **Placed on Probation**, to resit courses next session |
    | 3 | Any GPA | < 45% | **Advised to withdraw** |
    Note: "Credit Units Passed" refers to the percentage of total credit units passed.
    """
    exam_no = row.get("EXAM NUMBER", "Unknown")
    gpa = float(row.get("GPA", 0))
    cu_passed = int(row.get("CU Passed", 0))
    cu_failed = int(row.get("CU Failed", 0))
    
    # Calculate percentage of credit units passed
    passed_percentage = (cu_passed / total_cu * 100) if total_cu > 0 else 0
    # ENFORCED DECISION LOGIC BASED ON THE RULE
    # Rule 1: No failures → PASSED
    if cu_failed == 0:
        status = "Passed"
        reason = "No failed courses"
    # Rule 3: Passed < 45% of credits → WITHDRAWN (regardless of GPA)
    elif passed_percentage < 45:
        status = "Withdrawn"
        reason = f"Passed only {passed_percentage:.1f}% of credits (<45%) - Advised to withdraw"
    # Rule 1 & 2: Passed ≥ 45% of credits
    else:
        if gpa >= 2.00:
            status = "Resit"
            reason = f"GPA {gpa:.2f} ≥ 2.00, passed {passed_percentage:.1f}% of credits (≥45%) - To resit failed courses"
        else:
            status = "Probation"
            reason = f"GPA {gpa:.2f} < 2.00, passed {passed_percentage:.1f}% of credits (≥45%) - Placed on probation, to resit courses"
    # Debug output for specific students
    if hasattr(determine_student_status, "debug_students"):
        if exam_no in determine_student_status.debug_students or not hasattr(
            determine_student_status, "count"
        ):
            if not hasattr(determine_student_status, "count"):
                determine_student_status.count = 0
            determine_student_status.count += 1
            if determine_student_status.count <= 10:
                print(f"\n Student {exam_no}:")
                print(
                    f" CU Passed: {cu_passed} ({passed_percentage:.1f}%), CU Failed: {cu_failed}"
                )
                print(f" GPA: {gpa:.2f}")
                print(f" → Status: {status}")
                print(f" → Reason: {reason}")
                
    return status

# Initialize debug list for specific students
determine_student_status.debug_students = ["FCTCONS/ND24/109"]
determine_student_status.count = 0

def validate_probation_withdrawal_logic(mastersheet, total_cu):
    """
    Validate that probation and withdrawal statuses are correctly assigned.
    Specifically check edge cases around the 45% threshold.
    """
    print("\n" + "=" * 70)
    print("🔍 VALIDATING PROBATION/WITHDRAWAL LOGIC - ENFORCED RULE")
    print("=" * 70)
    
    # Check students who passed < 45% (should be Withdrawn regardless of GPA)
    low_pass_students = mastersheet[
        (mastersheet["CU Passed"] / total_cu < 0.45)
        & (mastersheet["CU Failed"] > 0) # Exclude students with no failures
    ]
    print(f"\n📊 Students with <45% credits passed:")
    print(f" Total: {len(low_pass_students)}")
    if len(low_pass_students) > 0:
        print(f"\n Should ALL be 'Withdrawn' (regardless of GPA):")
        for idx, row in low_pass_students.head(10).iterrows():
            exam_no = row["EXAM NUMBER"]
            gpa = row["GPA"]
            cu_passed = row["CU Passed"]
            cu_failed = row["CU Failed"]
            passed_pct = cu_passed / total_cu * 100
            status = row["REMARKS"]
            correct = "✅" if status == "Withdrawn" else f"❌ (got {status})"
            print(
                f" {exam_no}: GPA={gpa:.2f}, Passed={passed_pct:.1f}%, Failed={cu_failed}, Status={status} {correct}"
            )
    # Check students who passed ≥ 45% with GPA >= 2.00 (should be Resit)
    high_gpa_adequate_pass = mastersheet[
        (mastersheet["CU Passed"] / total_cu >= 0.45)
        & (mastersheet["GPA"] >= 2.00)
        & (mastersheet["CU Failed"] > 0) # Must have some failures
    ]
    print(f"\n📊 Students with ≥45% credits passed AND GPA ≥ 2.00:")
    print(f" Total: {len(high_gpa_adequate_pass)}")
    if len(high_gpa_adequate_pass) > 0:
        print(f"\n Should ALL be 'Resit':")
        for idx, row in high_gpa_adequate_pass.head(10).iterrows():
            exam_no = row["EXAM NUMBER"]
            gpa = row["GPA"]
            cu_passed = row["CU Passed"]
            passed_pct = cu_passed / total_cu * 100
            status = row["REMARKS"]
            correct = "✅" if status == "Resit" else f"❌ (got {status})"
            print(
                f" {exam_no}: GPA={gpa:.2f}, Passed={passed_pct:.1f}%, Status={status} {correct}"
            )
    # Check students who passed ≥ 45% with GPA < 2.00 (should be Probation)
    low_gpa_adequate_pass = mastersheet[
        (mastersheet["CU Passed"] / total_cu >= 0.45)
        & (mastersheet["GPA"] < 2.00)
        & (mastersheet["CU Failed"] > 0) # Must have some failures
    ]
    print(f"\n📊 Students with ≥45% credits passed AND GPA < 2.00:")
    print(f" Total: {len(low_gpa_adequate_pass)}")
    if len(low_gpa_adequate_pass) > 0:
        print(f"\n Should ALL be 'Probation':")
        for idx, row in low_gpa_adequate_pass.head(10).iterrows():
            exam_no = row["EXAM NUMBER"]
            gpa = row["GPA"]
            cu_passed = row["CU Passed"]
            passed_pct = cu_passed / total_cu * 100
            status = row["REMARKS"]
            correct = "✅" if status == "Probation" else f"❌ (got {status})"
            print(
                f" {exam_no}: GPA={gpa:.2f}, Passed={passed_pct:.1f}%, Status={status} {correct}"
            )
    # Status distribution
    print(f"\n📊 Overall Status Distribution:")
    status_counts = mastersheet["REMARKS"].value_counts()
    for status in ["Passed", "Resit", "Probation", "Withdrawn"]:
        count = status_counts.get(status, 0)
        pct = (count / len(mastersheet) * 100) if len(mastersheet) > 0 else 0
        print(f" {status:12s}: {count:3d} ({pct:5.1f}%)")
        
    print("=" * 70)

# ----------------------------
# Upgrade Rule Functions - FIXED: Added detailed logging
# ----------------------------

def get_upgrade_threshold_from_user(semester_key, set_name):
    """
    Prompt user to choose upgrade threshold for ND results.
    Returns: (min_threshold, upgraded_count) or (None, 0) if skipped
    """
    print(f"\n🎯 MANAGEMENT THRESHOLD UPGRADE RULE DETECTED")
    print(f"📚 Semester: {semester_key}")
    print(f"📁 Set: {set_name}")
    print(
        "\nSelect minimum score to upgrade (45-49). All scores >= selected value up to 49 will be upgraded to 50."
    )
    print("Enter 0 to skip upgrade.")
    
    while True:
        try:
            choice = input("\nEnter your choice (0, 45, 46, 47, 48, 49): ").strip()
            if not choice:
                print("❌ Please enter a value.")
                continue
            if choice == "0":
                print("⏭️ Skipping upgrade for this semester.")
                return None, 0
            if choice in ["45", "46", "47", "48", "49"]:
                min_threshold = int(choice)
                print(f"✅ Upgrade rule selected: {min_threshold}–49 → 50")
                return min_threshold, 0
            else:
                print("❌ Invalid choice. Please enter 0, 45, 46, 47, 48, or 49.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print(f"❌ Error: {e}. Please try again.")

def apply_upgrade_rule(mastersheet, ordered_codes, min_threshold):
    """
    Apply upgrade rule to mastersheet scores.
    Returns: (updated_mastersheet, upgraded_count)
    """
    if min_threshold is None:
        return mastersheet, 0
        
    upgraded_count = 0
    upgraded_students = set()
    upgrade_details = []  # Track details for verification
    
    print(f"🔄 Applying upgrade rule: {min_threshold}–49 → 50")
    
    for code in ordered_codes:
        for idx in mastersheet.index:
            score = mastersheet.at[idx, code]
            
            # Skip NOT REG
            if detect_not_registered_content(score):
                continue
                
            try:
                score_val = float(score)
                if min_threshold <= score_val <= 49:
                    exam_no = mastersheet.at[idx, "EXAM NUMBER"]
                    original_score = score_val
                    mastersheet.at[idx, code] = 50
                    upgraded_count += 1
                    upgraded_students.add(exam_no)
                    
                    # Store upgrade details
                    upgrade_details.append({
                        'exam_no': exam_no,
                        'course': code,
                        'original': original_score,
                        'upgraded': 50
                    })
                    
                    # Log first 10 upgrades for verification
                    if upgraded_count <= 10:
                        print(f"🔼 {exam_no} - {code}: {original_score} → 50")
            except (ValueError, TypeError):
                continue
                    
    if upgraded_count > 0:
        print(f"✅ Upgraded {upgraded_count} scores from {min_threshold}–49 to 50")
        print(f"📊 Affected {len(upgraded_students)} students")
        
        # Show sample of upgrades for verification
        if upgrade_details:
            print("\n📋 Sample upgrades:")
            for detail in upgrade_details[:5]:
                print(f"   {detail['exam_no']} - {detail['course']}: {detail['original']} → {detail['upgraded']}")
    else:
        print(f"ℹ️ No scores found in range {min_threshold}–49 to upgrade")
        print("🔍 Checking for scores in the upgrade range...")
        
        # Debug: Check what scores exist
        for code in ordered_codes:
            scores_in_range = []
            for idx in mastersheet.index:
                try:
                    score = float(mastersheet.at[idx, code])
                    if 40 <= score <= 49:
                        scores_in_range.append(score)
                except:
                    continue
            if scores_in_range:
                print(f"   {code}: Found {len(scores_in_range)} scores in 40-49 range: {sorted(set(scores_in_range))}")
        
    return mastersheet, upgraded_count

# ----------------------------
# Utilities
# ----------------------------

def normalize_path(path: str) -> str:
    """Normalize user paths for Windows/WSL/Linux."""
    path = os.path.expanduser(path)
    path = os.path.normpath(path)
    if platform.system().lower() == "linux" and path.startswith("C:\\"):
        path = "/mnt/" + path[0].lower() + path[2:].replace("\\", "/")
    if platform.system().lower() == "linux" and path.startswith("c:\\"):
        path = "/mnt/" + path[0].lower() + path[2:].replace("\\", "/")
    if path.startswith("/c/") and os.path.exists("/mnt/c"):
        path = path.replace("/c/", "/mnt/c/", 1)
    return os.path.abspath(path)

def create_zip_folder(source_dir, zip_path):
    """
    Create a ZIP file from a directory.
    Returns True if successful, False otherwise.
    """
    try:
        import zipfile
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Create relative path for ZIP
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        print(f"✅ Successfully created ZIP: {zip_path}")
        return True
    except Exception as e:
        print(f"❌ Failed to create ZIP: {e}")
        return False

def normalize_for_matching(s):
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"\b1st\b", "first", s)
    s = re.sub(r"\b2nd\b", "second", s)
    s = re.sub(r"\b3rd\b", "third", s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# ----------------------------
# Student Tracking Functions - FIXED: Now properly tracks names
# ----------------------------

def initialize_student_tracker():
    """Initialize the global student tracker."""
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    STUDENT_TRACKER = {}
    WITHDRAWN_STUDENTS = {}

def update_student_tracker(
    semester_key, exam_numbers, withdrawn_students=None, probation_students=None, exam_number_to_name_map=None
):
    """
    Update the student tracker with current semester's students.
    UPDATED: Tracks probation status separately and FIXED name tracking
    """
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    
    print(f"📊 Updating student tracker for {semester_key}")
    print(f"📝 Current students in this semester: {len(exam_numbers)}")
    
    # Track withdrawn students
    if withdrawn_students:
        for exam_no in withdrawn_students:
            if exam_no not in WITHDRAWN_STUDENTS:
                WITHDRAWN_STUDENTS[exam_no] = {
                    "withdrawn_semester": semester_key,
                    "withdrawn_date": datetime.now().strftime(TIMESTAMP_FMT),
                    "reappeared_semesters": [],
                }
                print(f"🚫 Marked as withdrawn: {exam_no} in {semester_key}")
    # Track probation students
    probation_count = 0
    
    # FIX 1: Use the exam_number_to_name_map to get proper names
    for exam_no in exam_numbers:
        # Get the name from the mapping if available
        student_name = "Unknown"
        if exam_number_to_name_map and exam_no in exam_number_to_name_map:
            student_name = exam_number_to_name_map[exam_no]
        
        if exam_no not in STUDENT_TRACKER:
            STUDENT_TRACKER[exam_no] = {
                "first_seen": semester_key,
                "last_seen": semester_key,
                "semesters_present": [semester_key],
                "status": "Active",
                "withdrawn": False,
                "withdrawn_semester": None,
                "probation_history": [],
                "current_probation": False,
                "has_cgpa_data": False,
                "current_gpa": 0.0,
                "cgpa_status": "Not in CGPA",
                "name": student_name,  # FIX 1: Store the name properly
            }
        else:
            STUDENT_TRACKER[exam_no]["last_seen"] = semester_key
            if semester_key not in STUDENT_TRACKER[exam_no]["semesters_present"]:
                STUDENT_TRACKER[exam_no]["semesters_present"].append(semester_key)
            
            # FIX 1: Update the name if we have a new one
            if exam_number_to_name_map and exam_no in exam_number_to_name_map:
                STUDENT_TRACKER[exam_no]["name"] = exam_number_to_name_map[exam_no]
            
            # Check if student was previously withdrawn and has reappeared
            if STUDENT_TRACKER[exam_no]["withdrawn"]:
                print(f"⚠️ PREVIOUSLY WITHDRAWN STUDENT REAPPEARED: {exam_no}")
                if exam_no in WITHDRAWN_STUDENTS:
                    if (
                        semester_key
                        not in WITHDRAWN_STUDENTS[exam_no]["reappeared_semesters"]
                    ):
                        WITHDRAWN_STUDENTS[exam_no]["reappeared_semesters"].append(
                            semester_key
                        )
        # Update probation status if this student is on probation
        if probation_students and exam_no in probation_students:
            if semester_key not in STUDENT_TRACKER[exam_no]["probation_history"]:
                STUDENT_TRACKER[exam_no]["probation_history"].append(semester_key)
            STUDENT_TRACKER[exam_no]["current_probation"] = True
            probation_count += 1
    print(f"📈 Total unique students tracked: {len(STUDENT_TRACKER)}")
    print(f"🚫 Total withdrawn students: {len(WITHDRAWN_STUDENTS)}")
    print(f"⚠️ Total probation students: {probation_count}")

def mark_student_withdrawn(exam_no, semester_key):
    """Mark a student as withdrawn in a specific semester."""
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    if exam_no in STUDENT_TRACKER:
        STUDENT_TRACKER[exam_no]["withdrawn"] = True
        STUDENT_TRACKER[exam_no]["withdrawn_semester"] = semester_key
        STUDENT_TRACKER[exam_no]["status"] = "Withdrawn"
        
    if exam_no not in WITHDRAWN_STUDENTS:
        WITHDRAWN_STUDENTS[exam_no] = {
            "withdrawn_semester": semester_key,
            "withdrawn_date": datetime.now().strftime(TIMESTAMP_FMT),
            "reappeared_semesters": [],
        }

def is_student_withdrawn(exam_no):
    """Check if a student has been withdrawn in any previous semester."""
    return exam_no in WITHDRAWN_STUDENTS

def get_withdrawal_history(exam_no):
    """Get withdrawal history for a student."""
    if exam_no in WITHDRAWN_STUDENTS:
        return WITHDRAWN_STUDENTS[exam_no]
    return None

def filter_out_withdrawn_students(mastersheet, semester_key):
    """
    Filter out students who were withdrawn in previous semesters.
    Returns filtered mastersheet and list of removed students.
    """
    removed_students = []
    filtered_mastersheet = mastersheet.copy()
    
    exam_col = find_exam_number_column(mastersheet)
    if not exam_col:
        print("❌ Could not find exam number column for filtering withdrawn students")
        return mastersheet, []
        
    for idx, row in mastersheet.iterrows():
        exam_no = str(row[exam_col]).strip()
        if is_student_withdrawn(exam_no):
            withdrawal_history = get_withdrawal_history(exam_no)
            # Only remove if student was withdrawn in a PREVIOUS semester
            if (
                withdrawal_history
                and withdrawal_history["withdrawn_semester"] != semester_key
            ):
                removed_students.append(exam_no)
                filtered_mastersheet = filtered_mastersheet[
                    filtered_mastersheet[exam_col].astype(str) != exam_no
                ]
                
    if removed_students:
        print(
            f"🚫 Removed {len(removed_students)} previously withdrawn students from {semester_key}:"
        )
        for exam_no in removed_students:
            withdrawal_history = get_withdrawal_history(exam_no)
            print(
                f" - {exam_no} (withdrawn in {withdrawal_history['withdrawn_semester']})"
            )
            
    return filtered_mastersheet, removed_students

# ----------------------------
# Set Selection Functions
# ----------------------------

def get_available_sets(base_dir):
    """Get all available ND sets (ND-2024, ND-2025, etc.) from the ND folder"""
    # UPDATED: Look in the ND subdirectory
    nd_dir = os.path.join(base_dir, "ND")
    if not os.path.exists(nd_dir):
        print(f"❌ ND directory not found: {nd_dir}")
        return []
        
    sets = []
    for item in os.listdir(nd_dir):
        item_path = os.path.join(nd_dir, item)
        if os.path.isdir(item_path) and item.upper().startswith("ND-"):
            sets.append(item)
            
    return sorted(sets)

def get_user_set_choice(available_sets):
    """
    Prompt user to choose which set to process.
    Returns the selected set directory name.
    """
    print("\n🎯 AVAILABLE SETS:")
    for i, set_name in enumerate(available_sets, 1):
        print(f"{i}. {set_name}")
    print(f"{len(available_sets) + 1}. Process ALL sets")
    
    while True:
        try:
            choice = input(
                f"\nEnter your choice (1-{len(available_sets) + 1}): "
            ).strip()
            if not choice:
                print("❌ Please enter a choice.")
                continue
                
            if choice.isdigit():
                choice_num = int(choice)
                if 1 <= choice_num <= len(available_sets):
                    selected_set = available_sets[choice_num - 1]
                    print(f"✅ Selected set: {selected_set}")
                    return [selected_set]
                elif choice_num == len(available_sets) + 1:
                    print("✅ Selected: ALL sets")
                    return available_sets
                else:
                    print(
                        f"❌ Invalid choice. Please enter a number between 1-{len(available_sets) + 1}."
                    )
            else:
                print("❌ Please enter a valid number.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print(f"❌ Error: {e}. Please try again.")

# ----------------------------
# Grade and GPA calculation - UPDATED for 4.0 scale
# ----------------------------

def get_grade(score):
    """Convert numeric score to letter grade - single letter only."""
    try:
        score = float(score)
        if score >= 70:
            return "A"
        elif score >= 60:
            return "B"
        elif score >= 50:
            return "C"
        elif score >= 45:
            return "D"
        elif score >= 40:
            return "E"
        else:
            return "F"
    except BaseException:
        return "F"

# UPDATED: get_grade_point function is now at the top with the CGPA functions
# ----------------------------
# Load Course Data
# ----------------------------

def load_course_data():
    """
    Reads course-code-creditUnit.xlsx and returns:
      (semester_course_maps, semester_credit_units,
       semester_lookup, semester_course_titles)
    """
    course_file = os.path.join(ND_COURSES_DIR, "course-code-creditUnit.xlsx")
    print(f"Loading course data from: {course_file}")
    if not os.path.exists(course_file):
        raise FileNotFoundError(f"Course file not found: {course_file}")
    xl = pd.ExcelFile(course_file)
    semester_course_maps = {}
    semester_credit_units = {}
    semester_lookup = {}
    semester_course_titles = {} # code -> title mapping
    for sheet in xl.sheet_names:
        df = pd.read_excel(course_file, sheet_name=sheet, engine="openpyxl", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        expected = ["COURSE CODE", "COURSE TITLE", "CU"]
        if not all(col in df.columns for col in expected):
            print(
                f"Warning: sheet '{sheet}' missing expected columns {expected} — skipped"
            )
            continue
        dfx = df.dropna(subset=["COURSE CODE", "COURSE TITLE"])
        dfx = dfx[
            ~dfx["COURSE CODE"].astype(str).str.contains("TOTAL", case=False, na=False)
        ]
        valid_mask = (
            dfx["CU"].astype(str).str.replace(".", "", regex=False).str.isdigit()
        )
        dfx = dfx[valid_mask]
        if dfx.empty:
            print(
                f"Warning: sheet '{sheet}' has no valid rows after cleaning — skipped"
            )
            continue
        codes = dfx["COURSE CODE"].astype(str).str.strip().tolist()
        titles = dfx["COURSE TITLE"].astype(str).str.strip().tolist()
        cus = dfx["CU"].astype(float).astype(int).tolist()
        # Create enhanced course mapping with normalized titles
        enhanced_course_map = {}
        for title, code in zip(titles, codes):
            normalized_title = normalize_course_name(title)
            enhanced_course_map[normalized_title] = {
                "original_name": title,
                "code": code,
                "normalized": normalized_title,
            }
        semester_course_maps[sheet] = enhanced_course_map
        semester_credit_units[sheet] = dict(zip(codes, cus))
        semester_course_titles[sheet] = dict(zip(codes, titles))
        # Create multiple lookup variations for flexible matching
        norm = normalize_for_matching(sheet)
        semester_lookup[norm] = sheet
        # Add variations without "ND-" prefix
        norm_no_nd = norm.replace("nd-", "").replace("nd ", "")
        semester_lookup[norm_no_nd] = sheet
        # Add variations with different separators
        norm_hyphen = norm.replace("-", " ")
        semester_lookup[norm_hyphen] = sheet
        norm_space = norm.replace(" ", "-")
        semester_lookup[norm_space] = sheet
    if not semester_course_maps:
        raise ValueError("No course data loaded from course workbook")
        
    print(f"Loaded course sheets: {list(semester_course_maps.keys())}")
    return (
        semester_course_maps,
        semester_credit_units,
        semester_lookup,
        semester_course_titles,
    )

# ----------------------------
# Helper functions
# ----------------------------

def detect_semester_from_filename(filename):
    """
    Detect semester from filename.
    Returns: (
    semester_key,
    year,
    semester_num,
    level_display,
    semester_display,
     set_code)
    """
    filename_upper = filename.upper()
    # Map filename patterns to actual course sheet names
    if (
        "FIRST-YEAR-FIRST-SEMESTER" in filename_upper
        or "FIRST_YEAR_FIRST_SEMESTER" in filename_upper
        or "FIRST SEMESTER" in filename_upper
    ):
        return "ND-FIRST-YEAR-FIRST-SEMESTER", 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI"
    elif (
        "FIRST-YEAR-SECOND-SEMESTER" in filename_upper
        or "FIRST_YEAR_SECOND_SEMESTER" in filename_upper
        or "SECOND SEMESTER" in filename_upper
    ):
        return (
            "ND-FIRST-YEAR-SECOND-SEMESTER",
            1,
            2,
            "YEAR ONE",
            "SECOND SEMESTER",
            "NDI",
        )
    elif (
        "SECOND-YEAR-FIRST-SEMESTER" in filename_upper
        or "SECOND_YEAR_FIRST_SEMESTER" in filename_upper
    ):
        return (
            "ND-SECOND-YEAR-FIRST-SEMESTER",
            2,
            1,
            "YEAR TWO",
            "FIRST SEMESTER",
            "NDII",
        )
    elif (
        "SECOND-YEAR-SECOND-SEMESTER" in filename_upper
        or "SECOND_YEAR_SECOND_SEMESTER" in filename_upper
    ):
        return (
            "ND-SECOND-YEAR-SECOND-SEMESTER",
            2,
            2,
            "YEAR TWO",
            "SECOND SEMESTER",
            "NDII",
        )
    elif "FIRST" in filename_upper and "SECOND" not in filename_upper:
        return "ND-FIRST-YEAR-FIRST-SEMESTER", 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI"
    elif "SECOND" in filename_upper:
        return (
            "ND-FIRST-YEAR-SECOND-SEMESTER",
            1,
            2,
            "YEAR ONE",
            "SECOND SEMESTER",
            "NDI",
        )
    else:
        # Default fallback
        print(
            f"⚠️ Could not detect semester from filename: {filename}, defaulting to ND-FIRST-YEAR-FIRST-SEMESTER"
        )
        return "ND-FIRST-YEAR-FIRST-SEMESTER", 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI"

def get_semester_display_info(semester_key):
    """
    Get display information for a given semester key.
    Returns: (year, semester_num, level_display, semester_display, set_code)
    """
    semester_lower = semester_key.lower()
    if "first-year-first-semester" in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI"
    elif "first-year-second-semester" in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "NDI"
    elif "second-year-first-semester" in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "NDII"
    elif "second-year-second-semester" in semester_lower:
        return 2, 2, "YEAR TWO", "SECOND SEMESTER", "NDII"
    elif "first" in semester_lower and "second" not in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI"
    elif "second" in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "NDI"
    else:
        # Default to first semester, first year
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI"

def match_semester_from_filename(fname, semester_lookup):
    """Match semester using the lookup table with flexible matching."""
    fn = normalize_for_matching(fname)
    # Try exact matches first
    for norm, sheet in semester_lookup.items():
        if norm in fn:
            return sheet
    # Try close matches
    keys = list(semester_lookup.keys())
    best = difflib.get_close_matches(fn, keys, n=1, cutoff=0.55)
    if best:
        return semester_lookup[best[0]]
    # Fallback to filename-based detection
    sem, _, _, _, _, _ = detect_semester_from_filename(fname)
    return sem

def find_column_by_names(df, candidate_names):
    norm_map = {
        col: re.sub(r"\s+", " ", str(col).strip().lower()) for col in df.columns
    }
    candidates = [re.sub(r"\s+", " ", c.strip().lower()) for c in candidate_names]
    for cand in candidates:
        for col, ncol in norm_map.items():
            if ncol == cand:
                return col
    return None

def find_exam_number_column(df):
    """Find the exam number column in a DataFrame with enhanced pattern matching."""
    # Primary patterns to look for
    primary_patterns = [
        "EXAM NUMBER",
        "EXAM NO",
        "EXAM_NO",
        "EXAMNUMBER",
        "REG NO",
        "REG NO.",
        "REGNO",
        "REGISTRATION NUMBER",
        "MAT NO",
        "MATRIC NO",
        "MATRICULATION NUMBER",
        "STUDENT ID",
        "STUDENTID",
        "STUDENT NUMBER",
    ]
    # Secondary patterns (less common but possible)
    secondary_patterns = ["EXAM", "REG", "MATRIC", "STUDENT", "ID", "NUMBER"]
    # First pass: exact matches for primary patterns
    for col in df.columns:
        col_upper = str(col).upper().strip()
        for pattern in primary_patterns:
            if pattern == col_upper:
                print(f"✅ Found exact exam number column: '{col}'")
                return col
    # Second pass: partial matches for primary patterns
    for col in df.columns:
        col_upper = str(col).upper().strip()
        for pattern in primary_patterns:
            if pattern in col_upper:
                print(f"✅ Found partial match exam number column: '{col}'")
                return col
    # Third pass: check for columns that might contain exam numbers by sampling data
    for col in df.columns:
        if df[col].notna().sum() > 0: # Column has some data
            sample_values = df[col].dropna().head(10).astype(str)
            exam_number_like = 0
            total_samples = len(sample_values)
            if total_samples > 0:
                for value in sample_values:
                    value_str = str(value).strip()
                    # Check if value looks like an exam number (alphanumeric, reasonable length)
                    if (
                        len(value_str) >= 5
                        and len(value_str) <= 20
                        and any(c.isalpha() for c in value_str)
                        and any(c.isdigit() for c in value_str)
                    ):
                        exam_number_like += 1
                # If most samples look like exam numbers, use this column
                if exam_number_like / total_samples >= 0.7:
                    print(f"✅ Detected exam number pattern in column: '{col}'")
                    return col
    # Final fallback: try common column positions
    common_positions = [0, 1] # Often first or second column
    for pos in common_positions:
        if pos < len(df.columns):
            col = df.columns[pos]
            print(f"⚠️ Using fallback exam number column (position {pos}): '{col}'")
            return col
    print("❌ Could not find exam number column")
    return None

# ----------------------------
# REFACTORED: PREVIOUS CGPA LOADING FROM ACTUAL PREVIOUS SEMESTER WORKSHEET
# ----------------------------

def load_previous_cgpas_from_processed_files(output_dir, current_semester_key, timestamp):
    """
    Load previous CGPA data from the ACTUAL previous semester's worksheet (not CGPA_SUMMARY).
    
    Returns dict: {exam_number: previous_cgpa}
    
    Semester mapping:
    - ND-FIRST-YEAR-FIRST-SEMESTER      → No previous CGPA
    - ND-FIRST-YEAR-SECOND-SEMESTER     → ND-FIRST-YEAR-FIRST-SEMESTER
    - ND-SECOND-YEAR-FIRST-SEMESTER     → ND-FIRST-YEAR-SECOND-SEMESTER  
    - ND-SECOND-YEAR-SECOND-SEMESTER    → ND-SECOND-YEAR-FIRST-SEMESTER
    """
    previous_cgpas = {}
    
    # Map current semester to its previous semester
    semester_to_previous_map = {
        "ND-FIRST-YEAR-FIRST-SEMESTER": None,
        "ND-FIRST-YEAR-SECOND-SEMESTER": "ND-FIRST-YEAR-FIRST-SEMESTER",
        "ND-SECOND-YEAR-FIRST-SEMESTER": "ND-FIRST-YEAR-SECOND-SEMESTER",
        "ND-SECOND-YEAR-SECOND-SEMESTER": "ND-SECOND-YEAR-FIRST-SEMESTER",
    }
    
    previous_semester_key = semester_to_previous_map.get(current_semester_key)
    
    if not previous_semester_key:
        print(f"📊 First semester ({current_semester_key}) - no previous CGPA available")
        return previous_cgpas
    
    print(f"\n🔍 LOADING PREVIOUS CGPA for: {current_semester_key}")
    print(f"📚 Looking for previous semester data from: {previous_semester_key}")
    
    # Look for the mastersheet file
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")
    print(f"🔍 Checking for mastersheet: {mastersheet_path}")
    
    if not os.path.exists(mastersheet_path):
        print(f"❌ Mastersheet not found: {mastersheet_path}")
        return previous_cgpas
    
    print(f"✅ Found mastersheet: {mastersheet_path}")
    
    try:
        # Read the Excel file to check sheets
        excel_file = pd.ExcelFile(mastersheet_path)
        print(f"📋 Available sheets: {excel_file.sheet_names}")
        
        if previous_semester_key not in excel_file.sheet_names:
            print(f"❌ Previous semester sheet '{previous_semester_key}' not found in mastersheet")
            print(f"📋 Available semester sheets: {[s for s in excel_file.sheet_names if s in SEMESTER_ORDER]}")
            return previous_cgpas
        
        print(f"✅ Found previous semester sheet: {previous_semester_key}")
        
        # Try multiple header rows to find the actual data (5-10 rows as requested)
        best_header_row = None
        best_df = None
        best_exam_col = None
        best_gpa_col = None
        
        for header_row in range(5, 11):  # Try rows 5-10 as requested
            try:
                df = pd.read_excel(mastersheet_path, sheet_name=previous_semester_key, header=header_row)
                
                if df.empty or len(df.columns) < 3:
                    continue
                
                # Find exam number column
                exam_col = None
                for col in df.columns:
                    col_str = str(col).upper().strip()
                    if any(keyword in col_str for keyword in ["EXAM", "REG", "NUMBER", "NO"]):
                        exam_col = col
                        break
                
                if not exam_col:
                    continue
                
                # Find GPA column
                gpa_col = None
                for col in df.columns:
                    col_str = str(col).upper().strip()
                    if any(pattern in col_str for pattern in ["GPA", "GRADE POINT", "POINT"]):
                        gpa_col = col
                        break
                
                if not gpa_col:
                    continue
                
                # Validate we have actual student data
                valid_students = 0
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    
                    # Skip invalid exam numbers
                    if (not exam_no or exam_no == "nan" or exam_no == "" or
                        any(keyword in exam_no.upper() for keyword in ["EXAM", "REG", "REGISTRATION", "STUDENT", "MATRIC", "NUMBER"]) or
                        len(exam_no) < 5):
                        continue
                    
                    # Check if GPA exists
                    if pd.notna(row[gpa_col]) and str(row[gpa_col]).strip() != "":
                        valid_students += 1
                        if valid_students >= 3:  # Found enough valid data
                            break
                
                if valid_students >= 3:
                    best_header_row = header_row
                    best_df = df
                    best_exam_col = exam_col
                    best_gpa_col = gpa_col
                    print(f"✅ Found valid data at header row {header_row} with {valid_students}+ students")
                    break
                    
            except Exception as e:
                continue
        
        if best_header_row is None or best_df is None:
            print(f"❌ Could not find valid data structure in {previous_semester_key}")
            return previous_cgpas
        
        print(f"📊 Processing {previous_semester_key} with header row {best_header_row}")
        print(f"📝 Using columns - Exam: '{best_exam_col}', GPA: '{best_gpa_col}'")
        
        # Load the data
        cgpas_loaded = 0
        for idx, row in best_df.iterrows():
            exam_no = str(row[best_exam_col]).strip()
            
            # Validate exam number
            if not exam_no or exam_no == "nan" or exam_no == "":
                continue
                
            # Skip header-like values
            exam_no_upper = exam_no.upper()
            if any(keyword in exam_no_upper for keyword in ["EXAM", "REG", "REGISTRATION", "STUDENT", "MATRIC", "NUMBER"]):
                continue
            
            # Must be reasonable length
            if len(exam_no) < 5:
                continue
            
            # Get GPA value
            gpa_val = row[best_gpa_col]
            if pd.notna(gpa_val) and str(gpa_val).strip() != "":
                try:
                    gpa_float = float(gpa_val)
                    # Convert from 5.0 scale to 4.0 scale if needed
                    if gpa_float > 4.0:  # Assume it's in 5.0 scale
                        gpa_float = (gpa_float / 5.0) * 4.0
                    
                    if 0 <= gpa_float <= 4.0:  # Valid GPA range for 4.0 scale
                        previous_cgpas[exam_no] = round(gpa_float, 2)
                        cgpas_loaded += 1
                        
                        if cgpas_loaded <= 3:  # Show first 3 for debugging
                            print(f"📝 Loaded previous CGPA: {exam_no} → {gpa_float:.2f}")
                except (ValueError, TypeError):
                    continue
        
        print(f"✅ Loaded previous CGPAs for {cgpas_loaded} students from {previous_semester_key}")
        
        if cgpas_loaded > 0:
            # Show summary
            sample_items = list(previous_cgpas.items())[:5]
            print(f"📊 Sample loaded CGPAs: {sample_items}")
        else:
            print(f"⚠️ No valid previous CGPA data found in {previous_semester_key}")
            
    except Exception as e:
        print(f"⚠️ Could not read previous semester sheet: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print(f"📊 FINAL: Loaded {len(previous_cgpas)} previous CGPAs from {previous_semester_key}")
    return previous_cgpas

def get_cumulative_cgpa(current_gpa, previous_cgpa, current_credits, previous_credits):
    """
    Calculate cumulative CGPA based on current and previous semester performance.
    """
    if previous_cgpa is None:
        return current_gpa
        
    # If credits not provided, assume equal weight
    if current_credits is None or previous_credits is None:
        return round((current_gpa + previous_cgpa) / 2, 2)
        
    total_points = (current_gpa * current_credits) + (previous_cgpa * previous_credits)
    total_credits = current_credits + previous_credits
    return round(total_points / total_credits, 2) if total_credits > 0 else 0.0

def load_all_previous_cgpas_for_cumulative(output_dir, current_semester_key, timestamp):
    """
    Load ALL previous CGPAs from all completed semesters for Cumulative CGPA calculation.
    Returns dict: {exam_number: {'gpas': [gpa1, gpa2, ...], 'credits': [credits1, credits2, ...]}}
    """
    print(f"\n🔍 LOADING ALL PREVIOUS CGPAs for Cumulative CGPA calculation: {current_semester_key}")
    
    current_year, current_semester_num, _, _, _ = get_semester_display_info(current_semester_key)
    
    # Determine which semesters to load based on current semester
    semesters_to_load = []
    if current_semester_num == 1 and current_year == 1:
        # First semester - no previous data
        print("📊 First semester of first year - no previous CGPA data")
        return {}
    elif current_semester_num == 2 and current_year == 1:
        # Second semester of first year - load first semester
        semesters_to_load = ["ND-FIRST-YEAR-FIRST-SEMESTER"]
    elif current_semester_num == 1 and current_year == 2:
        # First semester of second year - load both first year semesters
        semesters_to_load = [
            "ND-FIRST-YEAR-FIRST-SEMESTER",
            "ND-FIRST-YEAR-SECOND-SEMESTER",
        ]
    elif current_semester_num == 2 and current_year == 2:
        # Second semester of second year - load all previous semesters
        semesters_to_load = [
            "ND-FIRST-YEAR-FIRST-SEMESTER",
            "ND-FIRST-YEAR-SECOND-SEMESTER",
            "ND-SECOND-YEAR-FIRST-SEMESTER",
        ]
    
    print(f"📚 Semesters to load for Cumulative CGPA: {semesters_to_load}")
    
    all_student_data = {}
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")
    
    if not os.path.exists(mastersheet_path):
        print(f"❌ Mastersheet not found: {mastersheet_path}")
        return {}
    
    for semester in semesters_to_load:
        print(f"📖 Loading data from: {semester}")
        
        # Use the refactored function to load from the actual semester sheet
        semester_cgpas = load_previous_cgpas_from_processed_files(output_dir, semester, timestamp)
        
        # Convert format for cumulative calculation
        for exam_no, gpa in semester_cgpas.items():
            if exam_no not in all_student_data:
                all_student_data[exam_no] = {"gpas": [], "credits": []}
            
            all_student_data[exam_no]["gpas"].append(gpa)
            # Estimate credits for the semester
            if "FIRST" in semester.upper():
                all_student_data[exam_no]["credits"].append(30)  # Estimated first semester credits
            else:
                all_student_data[exam_no]["credits"].append(30)  # Estimated other semester credits
    
    print(f"📊 Loaded cumulative data for {len(all_student_data)} students")
    
    # Debug: Show sample data
    if all_student_data:
        sample_exam = list(all_student_data.keys())[0]
        print(f"📋 Sample cumulative data for {sample_exam}: {all_student_data[sample_exam]}")
    
    return all_student_data

def format_failed_courses_remark(failed_courses, max_line_length=60):
    """
    Format failed courses remark with line breaks for long lists.
    Returns list of formatted lines.
    """
    if not failed_courses:
        return [""]
        
    failed_str = ", ".join(sorted(failed_courses))
    # If the string is short enough, return as single line
    if len(failed_str) <= max_line_length:
        return [failed_str]
        
    # Split into multiple lines
    lines = []
    current_line = ""
    for course in sorted(failed_courses):
        if not current_line:
            current_line = course
        elif len(current_line) + len(course) + 2 <= max_line_length: # +2 for ", "
            current_line += ", " + course
        else:
            lines.append(current_line)
            current_line = course
    if current_line:
        lines.append(current_line)
        
    return lines

def get_user_semester_choice():
    """
    Prompt user to choose which semesters to process.
    Returns list of semester keys to process.
    """
    print("\n🎯 SEMESTER PROCESSING OPTIONS:")
    print("1. Process ALL semesters in order")
    print("2. Process FIRST YEAR - FIRST SEMESTER only")
    print("3. Process FIRST YEAR - SECOND SEMESTER only")
    print("4. Process SECOND YEAR - FIRST SEMESTER only")
    print("5. Process SECOND YEAR - SECOND SEMESTER only")
    print("6. Custom selection")
    
    while True:
        try:
            choice = input("\nEnter your choice (1-6): ").strip()
            if choice == "1":
                return SEMESTER_ORDER.copy()
            elif choice == "2":
                return ["ND-FIRST-YEAR-FIRST-SEMESTER"]
            elif choice == "3":
                return ["ND-FIRST-YEAR-SECOND-SEMESTER"]
            elif choice == "4":
                return ["ND-SECOND-YEAR-FIRST-SEMESTER"]
            elif choice == "5":
                return ["ND-SECOND-YEAR-SECOND-SEMESTER"]
            elif choice == "6":
                return get_custom_semester_selection()
            else:
                print("❌ Invalid choice. Please enter a number between 1-6.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print(f"❌ Error: {e}. Please try again.")

def get_custom_semester_selection():
    """
    Allow user to select multiple semesters for processing.
    """
    print("\n📚 AVAILABLE SEMESTERS:")
    for i, semester in enumerate(SEMESTER_ORDER, 1):
        year, sem_num, level, sem_display, set_code = get_semester_display_info(
            semester
        )
        print(f"{i}. {level} - {sem_display}")
    print(f"{len(SEMESTER_ORDER) + 1}. Select all")
    
    selected = []
    while True:
        try:
            choices = input(
                f"\nEnter semester numbers separated by commas (1-{len(SEMESTER_ORDER) + 1}): "
            ).strip()
            if not choices:
                print("❌ Please enter at least one semester number.")
                continue
                
            choice_list = [c.strip() for c in choices.split(",")]
            # Check for "select all" option
            if str(len(SEMESTER_ORDER) + 1) in choice_list:
                return SEMESTER_ORDER.copy()
                
            # Validate and convert choices
            valid_choices = []
            for choice in choice_list:
                if not choice.isdigit():
                    print(f"❌ '{choice}' is not a valid number.")
                    continue
                choice_num = int(choice)
                if 1 <= choice_num <= len(SEMESTER_ORDER):
                    valid_choices.append(choice_num)
                else:
                    print(f"❌ '{choice}' is not a valid semester number.")
                    
            if valid_choices:
                selected_semesters = [SEMESTER_ORDER[i - 1] for i in valid_choices]
                print(
                    f"✅ Selected semesters: {[get_semester_display_info(sem)[3] for sem in selected_semesters]}"
                )
                return selected_semesters
            else:
                print("❌ No valid semesters selected. Please try again.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print(f"❌ Error: {e}. Please try again.")

# ----------------------------
# PDF Generation - Individual Student Report (FIXED: Proper GPA vs CGPA terminology)
# UPDATED: Now includes both previous CGPA and current CGPA
# FIXED: Added proper error handling for missing reportlab installation
# UPDATED: Dynamic date based on current processing date
# ----------------------------

def generate_individual_student_pdf(
    mastersheet_df,
    out_pdf_path,
    semester_key,
    logo_path=None,
    prev_mastersheet_df=None,
    filtered_credit_units=None,
    ordered_codes=None,
    course_titles_map=None,
    previous_cgpas=None,
    cumulative_cgpa_data=None,
    total_cu=None,
    pass_threshold=None,
    upgrade_min_threshold=None,
):
    """
    FIXED VERSION: Properly separates Current GPA, Previous CGPA, and Current CGPA
    """
    try:
        # Try to import reportlab modules
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
            PageBreak,
            Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError as e:
        print(f"❌ ERROR: ReportLab is not installed. Cannot generate PDF.")
        print(f"💡 Please install it with: pip install reportlab")
        print(f"⚠️ Skipping PDF generation, but Excel processing will continue")
        return None  # Return None instead of False to indicate skipping
    
    # FIX: Validate inputs before proceeding
    if mastersheet_df is None or mastersheet_df.empty:
        print("⚠️ No student data to generate PDF")
        return None
        
    if ordered_codes is None:
        ordered_codes = []
        print("⚠️ No course codes provided for PDF generation")
    
    doc = SimpleDocTemplate(
        out_pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=20,
        bottomMargin=20,
    )
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        "CustomHeader",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=2,
    )
    main_header_style = ParagraphStyle(
        "MainHeader",
        parent=styles["Normal"],
        fontSize=16,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceAfter=6,
        textColor=colors.HexColor("#800080"),
    )
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Normal"],
        fontSize=12,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=10,
        textColor=colors.red,
    )
    left_align_style = ParagraphStyle(
        "LeftAlign",
        parent=styles["Normal"],
        fontSize=9,
        alignment=TA_LEFT,
        leftIndent=4,
    )
    center_align_style = ParagraphStyle(
        "CenterAlign", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER
    )
    remarks_style = ParagraphStyle(
        "RemarksStyle", parent=styles["Normal"], fontSize=8, alignment=TA_LEFT
    )
    elems = []
    
    for idx, r in mastersheet_df.iterrows():
        # Logo and header
        logo_img = None
        if logo_path and os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path, width=0.8 * inch, height=0.8 * inch)
            except Exception as e:
                print(f"Warning: Could not load logo: {e}")
        
        # Header table with logo and title
        if logo_img:
            header_data = [
                [
                    logo_img,
                    Paragraph("FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA", main_header_style),
                ]
            ]
            header_table = Table(header_data, colWidths=[1.0 * inch, 5.0 * inch])
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ]
                )
            )
            elems.append(header_table)
        else:
            elems.append(
                Paragraph("FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA", main_header_style)
            )
        
        # Address and contact info
        elems.append(Paragraph("P.O.Box 507, Gwagwalada-Abuja, Nigeria", header_style))
        elems.append(Paragraph("<b>DEPARTMENT OF NURSING</b>", header_style))
        elems.append(Paragraph("fctsonexamsoffice@gmail.com", header_style))
        elems.append(Spacer(1, 8))
        
        # Dynamic title based on semester
        year, semester_num, level_display, semester_display, set_code = get_semester_display_info(semester_key)
        
        # Get current date
        current_date = datetime.now().strftime("%B %d, %Y")
        if "SECOND-YEAR-SECOND-SEMESTER" in semester_key:
            exam_title = f"NATIONAL DIPLOMA YEAR TWO SECOND SEMESTER EXAMINATIONS RESULT — {current_date}"
        else:
            exam_title = f"NATIONAL DIPLOMA {level_display} {semester_display} EXAMINATIONS RESULT — {current_date}"
        
        elems.append(Paragraph(exam_title, title_style))
        elems.append(Paragraph("(THIS IS NOT A TRANSCRIPT)", subtitle_style))
        elems.append(Spacer(1, 8))
        
        # Student particulars
        exam_no = str(r.get("EXAM NUMBER", r.get("REG. No", "")))
        student_name = str(r.get("NAME", ""))
        
        # Particulars table
        particulars_data = [
            [Paragraph("<b>STUDENT'S PARTICULARS</b>", styles["Normal"])],
            [Paragraph("<b>NAME:</b>", styles["Normal"]), student_name],
            [
                Paragraph("<b>LEVEL OF<br/>STUDY:</b>", styles["Normal"]),
                level_display,
                Paragraph("<b>SEMESTER:</b>", styles["Normal"]),
                semester_display,
            ],
            [
                Paragraph("<b>REG NO.</b>", styles["Normal"]),
                exam_no,
                Paragraph("<b>SET:</b>", styles["Normal"]),
                set_code,
            ],
        ]
        particulars_table = Table(
            particulars_data, colWidths=[1.2 * inch, 2.3 * inch, 0.8 * inch, 1.5 * inch]
        )
        particulars_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (3, 0)),
                    ("SPAN", (1, 1), (3, 1)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        
        # Passport photo table
        passport_data = [
            [Paragraph("Affix Recent<br/>Passport<br/>Photograph", styles["Normal"])]
        ]
        passport_table = Table(
            passport_data, colWidths=[1.5 * inch], rowHeights=[1.2 * inch]
        )
        passport_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        
        # Combined table
        combined_data = [[particulars_table, passport_table]]
        combined_table = Table(combined_data, colWidths=[5.8 * inch, 1.5 * inch])
        combined_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        elems.append(combined_table)
        elems.append(Spacer(1, 12))
        
        # Semester result header
        elems.append(Paragraph("<b>SEMESTER RESULT</b>", title_style))
        elems.append(Spacer(1, 6))
        
        # Course results table
        course_data = [
            [
                Paragraph("<b>S/N</b>", styles["Normal"]),
                Paragraph("<b>CODE</b>", styles["Normal"]),
                Paragraph("<b>COURSE TITLE</b>", styles["Normal"]),
                Paragraph("<b>UNITS</b>", styles["Normal"]),
                Paragraph("<b>SCORE</b>", styles["Normal"]),
                Paragraph("<b>GRADE</b>", styles["Normal"]),
            ]
        ]
        sn = 1
        current_semester_grade_points = 0.0
        current_semester_units = 0
        current_semester_units_passed = 0
        current_semester_units_failed = 0
        failed_courses_list = []
        
        for code in ordered_codes:
            score = r.get(code)
            if pd.isna(score) or score == "":
                continue
                
            try:
                # Handle NOT REG content
                if detect_not_registered_content(score):
                    score_val = 0
                    score_display = "NOT REG"
                    grade = "N/A"
                    grade_point = 0.0
                else:
                    score_val = float(score)
                    score_display = str(int(round(score_val)))
                    grade = get_grade(score_val)
                    grade_point = get_grade_point(score_val)
            except Exception:
                score_display = str(score)
                grade = "F"
                grade_point = 0.0
                score_val = 0
            
            # Handle missing data
            cu = 0
            if filtered_credit_units and code in filtered_credit_units:
                cu = filtered_credit_units.get(code, 0)
            
            course_title = code
            if course_titles_map and code in course_titles_map:
                course_title = course_titles_map.get(code, code)
            
            # Calculate only for registered students
            if not detect_not_registered_content(score):
                current_semester_grade_points += grade_point * cu
                current_semester_units += cu
                if score_val >= (pass_threshold or 50):
                    current_semester_units_passed += cu
                else:
                    current_semester_units_failed += cu
                    failed_courses_list.append(code)
            
            course_data.append(
                [
                    Paragraph(str(sn), center_align_style),
                    Paragraph(code, left_align_style),
                    Paragraph(course_title, left_align_style),
                    Paragraph(str(cu), center_align_style),
                    Paragraph(score_display, center_align_style),
                    Paragraph(grade, center_align_style),
                ]
            )
            sn += 1
        
        # Check if we have courses
        if len(course_data) == 1:
            course_data.append([
                Paragraph("1", center_align_style),
                Paragraph("N/A", left_align_style),
                Paragraph("No courses available", left_align_style),
                Paragraph("0", center_align_style),
                Paragraph("0", center_align_style),
                Paragraph("N/A", center_align_style),
            ])
        
        course_table = Table(
            course_data,
            colWidths=[
                0.4 * inch,
                0.7 * inch,
                2.8 * inch,
                0.6 * inch,
                0.6 * inch,
                0.6 * inch,
            ],
        )
        course_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("ALIGN", (1, 1), (2, -1), "LEFT"),
                ]
            )
        )
        elems.append(course_table)
        elems.append(Spacer(1, 14))
        
        # ========================================================================
        # CORRECTED GPA AND CGPA CALCULATION LOGIC
        # ========================================================================
        
        # Step 1: Calculate CURRENT GPA (current semester only)
        current_gpa = (
            round(current_semester_grade_points / current_semester_units, 2) 
            if current_semester_units > 0 
            else 0.0
        )
        
        print(f"✅ Student {exam_no} - Current Semester GPA: {current_gpa:.2f}")
        print(f"   Current Semester: {current_semester_grade_points:.1f} points / {current_semester_units} units")
        
        # Step 2: Get PREVIOUS CGPA (from all completed semesters BEFORE current)
        previous_cgpa_value = None
        
        # Try to load from previous_cgpas dict
        if previous_cgpas and exam_no in previous_cgpas:
            previous_cgpa_value = previous_cgpas[exam_no]
            print(f"✅ Found Previous CGPA for {exam_no}: {previous_cgpa_value:.2f}")
        else:
            # Try to calculate from cumulative data (exclude current semester)
            if cumulative_cgpa_data and exam_no in cumulative_cgpa_data:
                student_history = cumulative_cgpa_data[exam_no]
                if student_history.get("gpas") and len(student_history["gpas"]) > 0:
                    # Calculate PREVIOUS CGPA from historical data only
                    prev_total_points = sum(
                        gpa * credits 
                        for gpa, credits in zip(student_history["gpas"], student_history["credits"])
                    )
                    prev_total_credits = sum(student_history["credits"])
                    
                    if prev_total_credits > 0:
                        previous_cgpa_value = round(prev_total_points / prev_total_credits, 2)
                        print(f"✅ Calculated Previous CGPA for {exam_no}: {previous_cgpa_value:.2f}")
                        print(f"   From {len(student_history['gpas'])} previous semester(s)")
        
        # Set display value
        display_previous_cgpa = previous_cgpa_value if previous_cgpa_value is not None else "N/A"
        
        if display_previous_cgpa == "N/A":
            print(f"⚠️ No previous CGPA data for {exam_no} (first semester or data missing)")
        
        # Step 3: Calculate CURRENT CGPA (cumulative including current semester)
        current_cgpa = current_gpa  # Initialize with current semester GPA
        
        if cumulative_cgpa_data and exam_no in cumulative_cgpa_data:
            student_history = cumulative_cgpa_data[exam_no]
            
            # Sum all previous semesters
            cumulative_grade_points = sum(
                gpa * credits 
                for gpa, credits in zip(
                    student_history.get("gpas", []), 
                    student_history.get("credits", [])
                )
            )
            cumulative_credits = sum(student_history.get("credits", []))
            
            # Add current semester
            cumulative_grade_points += current_semester_grade_points
            cumulative_credits += current_semester_units
            
            # Calculate CGPA
            if cumulative_credits > 0:
                current_cgpa = round(cumulative_grade_points / cumulative_credits, 2)
                print(f"✅ Calculated Current CGPA for {exam_no}: {current_cgpa:.2f}")
                print(f"   Total: {cumulative_grade_points:.1f} points / {cumulative_credits} units")
            else:
                current_cgpa = current_gpa
                print(f"⚠️ No cumulative credits, using current semester GPA: {current_cgpa:.2f}")
        
        elif previous_cgpa_value is not None:
            # Fallback: Use weighted average if we have previous CGPA but no detailed history
            estimated_prev_credits = 30  # Adjust based on your institution's typical load
            
            cumulative_grade_points = (previous_cgpa_value * estimated_prev_credits) + current_semester_grade_points
            cumulative_credits = estimated_prev_credits + current_semester_units
            
            if cumulative_credits > 0:
                current_cgpa = round(cumulative_grade_points / cumulative_credits, 2)
                print(f"✅ Estimated Current CGPA for {exam_no}: {current_cgpa:.2f}")
                print(f"   (Using estimated previous credits: {estimated_prev_credits})")
        else:
            # No previous data: Current CGPA = Current GPA
            current_cgpa = current_gpa
            print(f"ℹ️ No previous data for {exam_no}, Current CGPA = Current GPA: {current_cgpa:.2f}")
        
        # Validation checks
        print(f"\n📊 FINAL VALUES for {exam_no}:")
        print(f"   Current GPA (this semester only):        {current_gpa:.2f}")
        print(f"   Previous CGPA (before this semester):    {display_previous_cgpa}")
        print(f"   Current CGPA (cumulative with current):  {current_cgpa:.2f}")
        
        if previous_cgpa_value is not None:
            if abs(current_gpa - current_cgpa) < 0.01:
                print(f"⚠️ WARNING: Current GPA and Current CGPA are identical!")
                print(f"   This should only happen if student has no prior academic history.")
                print(f"   But Previous CGPA exists ({previous_cgpa_value:.2f}), so this is likely an error.")
        
        # ========================================================================
        # Prepare display values for PDF
        # ========================================================================
        
        display_current_gpa = current_gpa
        display_current_cgpa = current_cgpa
        
        # Summary values for PDF
        tcpe = round(current_semester_grade_points, 1)
        tcup = current_semester_units_passed
        tcuf = current_semester_units_failed
        
        # Student status
        student_status = r.get("REMARKS", "Passed")
        withdrawal_history = get_withdrawal_history(exam_no)
        previously_withdrawn = withdrawal_history is not None
        
        # Failed courses
        failed_courses_str = str(r.get("FAILED COURSES", ""))
        failed_courses_list = (
            [c.strip() for c in failed_courses_str.split(",") if c.strip()]
            if failed_courses_str
            else []
        )
        failed_courses_formatted = format_failed_courses_remark(failed_courses_list)
        
        # Remarks
        final_remarks_lines = []
        if previously_withdrawn and withdrawal_history["withdrawn_semester"] == semester_key:
            if failed_courses_formatted:
                final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                if len(failed_courses_formatted) > 1:
                    final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("Advised to Withdraw")
            else:
                final_remarks_lines.append("Advised to Withdraw")
        elif previously_withdrawn:
            withdrawn_semester = withdrawal_history["withdrawn_semester"]
            year, sem_num, level, sem_display, set_code = get_semester_display_info(withdrawn_semester)
            final_remarks_lines.append(f"STUDENT WAS WITHDRAWN FROM {level} - {sem_display}")
        else:
            if student_status == "Passed":
                final_remarks_lines.append("Passed")
            elif student_status == "Withdrawn":
                if failed_courses_formatted:
                    final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                    if len(failed_courses_formatted) > 1:
                        final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("Advised to Withdraw")
            elif student_status == "Resit":
                if failed_courses_formatted:
                    final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                    if len(failed_courses_formatted) > 1:
                        final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("To Resit Courses")
            elif student_status == "Probation":
                if failed_courses_formatted:
                    final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                    if len(failed_courses_formatted) > 1:
                        final_remarks_lines.extend(failed_courses_formatted[1:])
                passed_percentage = (tcup / (total_cu or 1) * 100)
                if passed_percentage >= 45 and current_gpa < 2.00:
                    final_remarks_lines.append("Placed on Probation (Passed ≥45% but GPA < 2.00)")
                final_remarks_lines.append("To Resit Failed Courses")
        
        final_remarks = "<br/>".join(final_remarks_lines)
        
        # ========================================================================
        # PDF SUMMARY TABLE WITH CORRECT VALUES
        # ========================================================================
        
        summary_data = [
            [Paragraph("<b>SUMMARY</b>", styles["Normal"]), "", "", ""],
            [
                Paragraph("<b>TCPE:</b>", styles["Normal"]),
                f"{tcpe:.1f}",
                Paragraph("<b>CURRENT GPA:</b>", styles["Normal"]),
                f"{display_current_gpa:.2f}",  # ✅ Current semester only
            ],
            [
                Paragraph("<b>TCUP:</b>", styles["Normal"]),
                str(tcup),
                Paragraph("<b>PREVIOUS CGPA:</b>", styles["Normal"]),
                f"{display_previous_cgpa:.2f}" if display_previous_cgpa != "N/A" else "N/A",  # ✅ Historical cumulative
            ],
            [
                Paragraph("<b>TCUF:</b>", styles["Normal"]),
                str(tcuf),
                Paragraph("<b>CURRENT CGPA:</b>", styles["Normal"]),
                f"{display_current_cgpa:.2f}",  # ✅ Updated cumulative
            ],
        ]
        
        remarks_paragraph = Paragraph(final_remarks, remarks_style)
        summary_data.append([
            Paragraph("<b>REMARKS:</b>", styles["Normal"]), 
            remarks_paragraph, "", ""
        ])
        
        # Row heights
        row_heights = [0.3 * inch] * len(summary_data)
        total_remark_lines = len(final_remarks_lines)
        if total_remark_lines > 1:
            row_heights[-1] = max(0.4 * inch, 0.2 * inch * (total_remark_lines + 1))
        
        summary_table = Table(
            summary_data,
            colWidths=[1.5 * inch, 1.0 * inch, 1.5 * inch, 1.0 * inch],
            rowHeights=row_heights,
        )
        summary_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (3, 0)),  # "SUMMARY" spans all columns
                    ("SPAN", (1, len(summary_data) - 1), (3, len(summary_data) - 1)),  # REMARKS spans columns 2-4
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#E0E0E0")),
                    ("ALIGN", (0, 0), (3, 0), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        elems.append(summary_table)
        elems.append(Spacer(1, 25))
        
        # Signature section
        sig_data = [
            ["", ""],
            ["____________________", "____________________"],
            [
                Paragraph("<b>Head of Exams</b>", ParagraphStyle("SigStyle", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER)),
                Paragraph("<b>HOD Nursing</b>", ParagraphStyle("SigStyle", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER)),
            ],
        ]
        sig_table = Table(sig_data, colWidths=[3.0 * inch, 3.0 * inch])
        sig_table.setStyle(
            TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ])
        )
        elems.append(sig_table)
        
        # Page break
        if idx < len(mastersheet_df) - 1:
            elems.append(PageBreak())
    
    try:
        doc.build(elems)
        print(f"✅ Individual student PDF written: {out_pdf_path}")
        return True
    except Exception as e:
        print(f"❌ Failed to generate PDF: {e}")
        import traceback
        traceback.print_exc()
        return False

# ----------------------------
# Main file processing (Enhanced with Data Transformation and NOT REG handling)
# ----------------------------

def process_single_file(
    path,
    output_dir,
    ts,
    pass_threshold,
    semester_course_maps,
    semester_credit_units,
    semester_lookup,
    semester_course_titles,
    logo_path,
    semester_key,
    set_name,
    previous_cgpas,
    cumulative_cgpa_data=None,
    upgrade_min_threshold=None,
):
    """
    Process a single raw file and produce mastersheet Excel and PDFs.
    Enhanced with data transformation for transposed formats and NOT REG handling.
    UPDATED TITLE: Dynamic title based on semester being processed
    UPDATED: Applied row height and column width fixes
    UPDATED: FIXED name tracking in student tracker
    UPDATED: Recalculated GPAs for 4.0 scale
    
    FIXED: Added Previous CGPA and Current CGPA calculation to Excel mastersheet
    """
    fname = os.path.basename(path)
    print(f"🔍 Processing file: {fname} for semester: {semester_key}")
    
    try:
        xl = pd.ExcelFile(path)
        print(f"✅ Successfully opened Excel file: {fname}")
        print(f"📋 Sheets found: {xl.sheet_names}")
    except Exception as e:
        print(f"❌ Error opening excel {path}: {e}")
        return None
    expected_sheets = ["CA", "OBJ", "EXAM"]
    dfs = {}
    for s in expected_sheets:
        if s in xl.sheet_names:
            try:
                # Try reading with different parameters to handle various Excel formats
                dfs[s] = pd.read_excel(path, sheet_name=s, dtype=str, header=0)
                print(f"✅ Loaded sheet {s} with shape: {dfs[s].shape}")
                print(f"📊 Sheet {s} columns: {dfs[s].columns.tolist()}")
                
                # NEW: Check if data is in transposed format and transform if needed
                if detect_data_format(dfs[s], s):
                    print(
                        f"🔄 Data in {s} sheet is in transposed format, transforming..."
                    )
                    dfs[s] = transform_transposed_data(dfs[s], s)
                    print(f"✅ Transformed {s} sheet to wide format")
                    print(f"📊 Transformed shape: {dfs[s].shape}")
                    print(f"📋 Transformed columns: {dfs[s].columns.tolist()}")
                    
                # Debug: Show first few rows of data
                if not dfs[s].empty:
                    print(f"🔍 First 3 rows of {s} sheet:")
                    for i in range(min(3, len(dfs[s]))):
                        row_data = {}
                        for col in dfs[s].columns[:5]: # Show first 5 columns
                            row_data[col] = dfs[s].iloc[i][col]
                        print(f" Row {i}: {row_data}")
                else:
                    print(f"⚠️ Sheet {s} is empty!")
            except Exception as e:
                print(f"❌ Error reading sheet {s}: {e}")
                # Try alternative reading method
                try:
                    dfs[s] = pd.read_excel(path, sheet_name=s, header=0)
                    print(f"✅ Alternative load successful for sheet {s}")
                except Exception as e2:
                    print(f"❌ Alternative load also failed for sheet {s}: {e2}")
                    dfs[s] = pd.DataFrame()
        else:
            print(f"⚠️ Sheet {s} not found in {fname}")
            dfs[s] = pd.DataFrame()
    if not dfs:
        print("❌ No CA/OBJ/EXAM sheets detected — skipping file.")
        return None
    # Use the provided semester key
    sem = semester_key
    year, semester_num, level_display, semester_display, set_code = (
        get_semester_display_info(sem)
    )
    print(f"📁 Processing: {level_display} - {semester_display} - Set: {set_code}")
    print(f"📊 Using course sheet: {sem}")
    print(f"📊 Previous CGPAs provided: {len(previous_cgpas)} students")
    print(
        f"📊 Cumulative CGPA data available for: {len(cumulative_cgpa_data) if cumulative_cgpa_data else 0} students"
    )
    # Check if semester exists in course maps
    if sem not in semester_course_maps:
        print(
            f"❌ Semester '{sem}' not found in course data. Available semesters: {list(semester_course_maps.keys())}"
        )
        return None
    course_map = semester_course_maps[sem]
    credit_units = semester_credit_units[sem]
    course_titles = semester_course_titles[sem]
    ordered_titles = list(course_map.keys())
    ordered_codes = [course_map[t]["code"] for t in ordered_titles if course_map.get(t)]
    ordered_codes = [c for c in ordered_codes if credit_units.get(c, 0) > 0]
    filtered_credit_units = {c: credit_units[c] for c in ordered_codes}
    total_cu = sum(filtered_credit_units.values())
    print(f"📚 Course codes to process: {ordered_codes}")
    print(f"📊 Total credit units: {total_cu}")
    reg_no_cols = {
        s: find_column_by_names(
            df,
            [
                "REG. No",
                "Reg No",
                "Registration Number",
                "Mat No",
                "EXAM NUMBER",
                "Student ID",
            ],
        )
        for s, df in dfs.items()
    }
    name_cols = {
        s: find_column_by_names(df, ["NAME", "Full Name", "Candidate Name"])
        for s, df in dfs.items()
    }
    print(f"🔍 Registration columns found: {reg_no_cols}")
    print(f"🔍 Name columns found: {name_cols}")
    merged = None
    for s, df in dfs.items():
        if df.empty:
            print(f"⚠️ Skipping empty sheet: {s}")
            continue
            
        df = df.copy()
        regcol = reg_no_cols.get(s)
        namecol = name_cols.get(s)
        if not regcol:
            regcol = df.columns[0] if len(df.columns) > 0 else None
        if not namecol and len(df.columns) > 1:
            namecol = df.columns[1]
        if regcol is None:
            print(f"❌ Skipping sheet {s}: no reg column found")
            continue
            
        print(
            f"📝 Processing sheet {s} with reg column: {regcol}, name column: {namecol}"
        )
        df["REG. No"] = df[regcol].astype(str).str.strip()
        if namecol:
            df["NAME"] = df[namecol].astype(str).str.strip()
        else:
            df["NAME"] = pd.NA
        to_drop = [c for c in [regcol, namecol] if c and c not in ["REG. No", "NAME"]]
        df.drop(columns=to_drop, errors="ignore", inplace=True)
        # Debug: Show available columns for matching
        print(f"🔍 Available columns in {s} sheet: {df.columns.tolist()}")
        # ENHANCED COURSE MATCHING - Use the new matching algorithm
        for col in [c for c in df.columns if c not in ["REG. No", "NAME"]]:
            # Skip empty or whitespace-only columns
            if not col or str(col).strip() == '':
                print(f"⚠️ Skipping empty/whitespace column")
                df.drop(columns=[col], inplace=True)
                continue
                
            matched_course = find_best_course_match(col, course_map)
            if matched_course:
                matched_code = matched_course["code"]
                newcol = f"{matched_code}_{s.upper()}"
                
                # Prevent duplicate column names
                if newcol in df.columns:
                    print(f"⚠️ Column '{newcol}' already exists, skipping duplicate match for '{col}'")
                    df.drop(columns=[col], inplace=True)
                    continue
                    
                df.rename(columns={col: newcol}, inplace=True)
                print(
                    f"✅ Matched column '{col}' to course code '{matched_code}' (original: {matched_course['original_name']})"
                )
            else:
                print(f"❌ No match found for column: '{col}'")
        cur_cols = ["REG. No", "NAME"] + [
            c for c in df.columns if c.endswith(f"_{s.upper()}")
        ]
        cur = df[cur_cols].copy()
        # Debug: Show data before merging
        print(f"📊 Data in {s} sheet - Shape: {cur.shape}")
        if not cur.empty:
            print(f"🔍 First 3 rows of {s} data:")
            for i in range(min(2, len(cur))):
                print(
                    f" Row {i}: REG. No='{cur.iloc[i]['REG. No']}', NAME='{cur.iloc[i]['NAME']}'"
                )
        if merged is None:
            merged = cur
            print(f"✅ Initialized merged dataframe with {s} sheet")
        else:
            print(f"🔗 Merging {s} sheet with existing data")
            before_merge = len(merged)
            merged = merged.merge(cur, on="REG. No", how="outer", suffixes=("", "_dup"))
            after_merge = len(merged)
            print(f"📊 Merge result: {before_merge} -> {after_merge} rows")
            if "NAME_dup" in merged.columns:
                merged["NAME"] = merged["NAME"].combine_first(merged["NAME_dup"])
                merged.drop(columns=["NAME_dup"], inplace=True)
    if merged is None or merged.empty:
        print("❌ No data merged from sheets — skipping file.")
        return None
    print(f"✅ Final merged dataframe shape: {merged.shape}")
    print(f"📋 Final merged columns: {merged.columns.tolist()}")
    # NEW: NOT REG DETECTION AND HANDLING
    print("🔍 Checking for NOT REGISTERED candidates...")
    course_columns_to_check = []
    for code in ordered_codes:
        for sheet_type in ["CA", "OBJ", "EXAM"]:
            course_columns_to_check.extend([
                f"{code}_{sheet_type}"
            ])
    # Process NOT REG content
    merged, not_reg_counts = process_not_registered_scores(merged, course_columns_to_check)
    # Print NOT REG summary
    total_not_reg = sum(not_reg_counts.values())
    if total_not_reg > 0:
        print(f"📊 Found {total_not_reg} NOT REGISTERED entries across courses:")
        for course, count in not_reg_counts.items():
            if count > 0:
                print(f"   - {course}: {count} NOT REG entries")
    # CRITICAL FIX: Check if we have actual score data before proceeding
    has_score_data = False
    score_columns = [
        col for col in merged.columns if any(code in col for code in ordered_codes)
    ]
    print(f"🔍 Checking score columns: {score_columns}")
    
    for col in score_columns:
        if col in merged.columns:
            # Check if column has any non-null, non-zero values
            try:
                col_data = merged[col]
                # Handle both Series and DataFrame cases (in case of duplicate columns)
                if isinstance(col_data, pd.DataFrame):
                    # If it's a DataFrame, take the first column
                    col_data = col_data.iloc[:, 0]
                
                non_null_count = int(col_data.notna().sum())
                if non_null_count > 0:
                    # Try to convert to numeric and check for non-zero values
                    numeric_values = pd.to_numeric(col_data, errors="coerce")
                    non_zero_count = int((numeric_values > 0).sum())
                    if non_zero_count > 0:
                        has_score_data = True
                        print(
                            f"✅ Found score data in column {col}: {non_zero_count} non-zero values"
                        )
                        break
            except Exception as e:
                print(f"⚠️ Error checking column {col}: {e}")
    if not has_score_data:
        print(f"❌ CRITICAL: No valid score data found in file {fname}!")
        print(f"🔍 Sample of merged data:")
        print(merged.head(3))
        return None
    mastersheet = merged[["REG. No", "NAME"]].copy()
    mastersheet.rename(columns={"REG. No": "EXAM NUMBER"}, inplace=True)
    print("🎯 Calculating scores for each course...")
    
    # FIX: Ensure course columns can handle mixed data types before NOT REG assignment
    for code in ordered_codes:
        if code in mastersheet.columns and mastersheet[code].dtype != 'object':
            mastersheet[code] = mastersheet[code].astype('object')
    
    for code in ordered_codes:
        ca_col = f"{code}_CA"
        obj_col = f"{code}_OBJ"
        exam_col = f"{code}_EXAM"
        print(f"📊 Processing course {code}:")
        print(f" CA column: {ca_col} - exists: {ca_col in merged.columns}")
        print(f" OBJ column: {obj_col} - exists: {obj_col in merged.columns}")
        print(f" EXAM column: {exam_col} - exists: {exam_col in merged.columns}")
        # NEW: NOT REG handling in score calculation
        mastersheet[code] = 0  # Default to 0
        
        for idx in merged.index:
            exam_no = str(merged.at[idx, "REG. No"]).strip()
            
            # Check if any component has NOT REG
            ca_has_not_reg = ca_col in merged.columns and detect_not_registered_content(merged.at[idx, ca_col])
            obj_has_not_reg = obj_col in merged.columns and detect_not_registered_content(merged.at[idx, obj_col])
            exam_has_not_reg = exam_col in merged.columns and detect_not_registered_content(merged.at[idx, exam_col])
            
            # If ANY component has NOT REG, mark the entire course as NOT REG
            if ca_has_not_reg or obj_has_not_reg or exam_has_not_reg:
                # Convert column to object type first to allow mixed types
                if mastersheet[code].dtype != 'object':
                    mastersheet[code] = mastersheet[code].astype('object')
                mastersheet.at[idx, code] = "NOT REG"
                continue
                
            # Normal score calculation for registered students
            ca_series = (
                pd.to_numeric(merged.at[idx, ca_col], errors="coerce")
                if ca_col in merged.columns and pd.notna(merged.at[idx, ca_col])
                else 0
            )
            obj_series = (
                pd.to_numeric(merged.at[idx, obj_col], errors="coerce") 
                if obj_col in merged.columns and pd.notna(merged.at[idx, obj_col])
                else 0
            )
            exam_series = (
                pd.to_numeric(merged.at[idx, exam_col], errors="coerce")
                if exam_col in merged.columns and pd.notna(merged.at[idx, exam_col])
                else 0
            )
            # Handle NaN values from coercion
            ca_norm = (float(ca_series) / 20) * 100 if not pd.isna(ca_series) else 0
            obj_norm = (float(obj_series) / 20) * 100 if not pd.isna(obj_series) else 0
            exam_norm = (float(exam_series) / 80) * 100 if not pd.isna(exam_series) else 0
            ca_norm = min(ca_norm, 100)
            obj_norm = min(obj_norm, 100) 
            exam_norm = min(exam_norm, 100)
            total = (ca_norm * 0.2) + (((obj_norm + exam_norm) / 2) * 0.8)
            mastersheet.at[idx, code] = round(total, 0)
    # NEW: APPLY FLEXIBLE UPGRADE RULE - Ask user for threshold per semester
    # Only ask in interactive mode
    if should_use_interactive_mode():
        upgrade_min_threshold, upgraded_scores_count = get_upgrade_threshold_from_user(
            semester_key, set_name
        )
    else:
        # In non-interactive mode, use the provided threshold or None
        upgraded_scores_count = 0
        if upgrade_min_threshold is not None:
            print(
                f"🔄 Applying upgrade upgrade from parameters: {upgrade_min_threshold}–49 → 50"
            )
            
    if upgrade_min_threshold is not None:
        mastersheet, upgraded_scores_count = apply_upgrade_rule(
            mastersheet, ordered_codes, upgrade_min_threshold
        )
    for c in ordered_codes:
        if c not in mastersheet.columns:
            mastersheet[c] = 0
    # UPDATED: Compute FAILED COURSES with corrected logic (excluding NOT REG)
    def compute_failed_courses(row):
        """Compute list of failed courses (excluding NOT REG courses)."""
        fails = []
        for c in ordered_codes:
            score = row.get(c)
            # Skip NOT REG courses when counting failures
            if detect_not_registered_content(score):
                continue
            try:
                if float(score or 0) < pass_threshold:
                    fails.append(c)
            except (ValueError, TypeError):
                continue
        return ", ".join(sorted(fails)) if fails else ""

    mastersheet["FAILED COURSES"] = mastersheet.apply(compute_failed_courses, axis=1)
    # NEW: Calculate TCPE, TCUP, TCUF with NOT REG handling
    def calc_tcpe_tcup_tcuf(row):
        tcpe = 0.0
        tcup = 0
        tcuf = 0
        total_registered_cu = 0
        
        for code in ordered_codes:
            score = row.get(code)
            
            # Skip NOT REG courses
            if detect_not_registered_content(score):
                continue
                
            try:
                score_val = float(score) if pd.notna(score) and score != "" else 0
                cu = filtered_credit_units.get(code, 0)
                gp = get_grade_point(score_val)
                tcpe += gp * cu
                total_registered_cu += cu
                
                if score_val >= pass_threshold:
                    tcup += cu
                else:
                    tcuf += cu
            except (ValueError, TypeError):
                continue
                
        return tcpe, tcup, tcuf, total_registered_cu

    results = mastersheet.apply(calc_tcpe_tcup_tcuf, axis=1, result_type="expand")
    mastersheet["TCPE"] = results[0].round(1)
    mastersheet["CU Passed"] = results[1]
    mastersheet["CU Failed"] = results[2]
    mastersheet["Total Registered CU"] = results[3]  # Add this column
    # NEW: Calculate GPA with NOT REG handling
    def calculate_gpa(row):
        tcpe = row["TCPE"]
        total_registered_cu = row["Total Registered CU"]
        return round((tcpe / total_registered_cu), 2) if total_registered_cu > 0 else 0.0

    mastersheet["GPA"] = mastersheet.apply(calculate_gpa, axis=1)
    
    # ========================================================================
    # CRITICAL FIX: ADDED PREVIOUS CGPA AND CURRENT CGPA CALCULATION TO EXCEL
    # ========================================================================
    print("🎯 Calculating CGPA values for Excel mastersheet...")
    
    # Calculate Previous CGPA (from all completed semesters before current)
    def calculate_previous_cgpa(row):
        """Calculate Previous CGPA from cumulative data (excluding current semester)."""
        exam_no = str(row["EXAM NUMBER"]).strip()
        
        if cumulative_cgpa_data and exam_no in cumulative_cgpa_data:
            student_history = cumulative_cgpa_data[exam_no]
            
            # Use only previous semesters (exclude current)
            prev_gpas = student_history.get("gpas", [])
            prev_credits = student_history.get("credits", [])
            
            if prev_gpas and prev_credits and len(prev_gpas) == len(prev_credits):
                total_points = sum(gpa * credit for gpa, credit in zip(prev_gpas, prev_credits))
                total_credits = sum(prev_credits)
                
                if total_credits > 0:
                    previous_cgpa = round(total_points / total_credits, 2)
                    print(f"📊 {exam_no}: Previous CGPA calculated = {previous_cgpa:.2f} from {len(prev_gpas)} semesters")
                    return previous_cgpa
        
        # Return N/A if no previous data
        return "N/A"
    
    # Calculate Current CGPA (cumulative including current semester)
    def calculate_current_cgpa(row):
        """Calculate Current CGPA (cumulative including current semester)."""
        exam_no = str(row["EXAM NUMBER"]).strip()
        current_gpa = row["GPA"]
        
        if cumulative_cgpa_data and exam_no in cumulative_cgpa_data:
            student_history = cumulative_cgpa_data[exam_no]
            
            # Get previous semester data
            prev_gpas = student_history.get("gpas", [])
            prev_credits = student_history.get("credits", [])
            
            # Add current semester
            current_grade_points = row["TCPE"]
            current_credits = row["Total Registered CU"]
            
            if prev_gpas and prev_credits:
                total_points = sum(gpa * credit for gpa, credit in zip(prev_gpas, prev_credits))
                total_points += current_grade_points
                
                total_credits = sum(prev_credits) + current_credits
                
                if total_credits > 0:
                    current_cgpa = round(total_points / total_credits, 2)
                    print(f"📊 {exam_no}: Current CGPA calculated = {current_cgpa:.2f} (includes current semester)")
                    return current_cgpa
        
        # If no previous data, use current GPA as CGPA
        print(f"📊 {exam_no}: No previous data, Current CGPA = Current GPA = {current_gpa:.2f}")
        return current_gpa
    
    # Apply CGPA calculations
    print("🔍 Calculating Previous CGPA values...")
    mastersheet["PREVIOUS CGPA"] = mastersheet.apply(calculate_previous_cgpa, axis=1)
    
    print("🔍 Calculating Current CGPA values...")
    mastersheet["CURRENT CGPA"] = mastersheet.apply(calculate_current_cgpa, axis=1)
    
    print(f"✅ CGPA calculations completed. Sample values:")
    for idx in range(min(3, len(mastersheet))):
        exam_no = mastersheet.iloc[idx]["EXAM NUMBER"]
        current_gpa = mastersheet.iloc[idx]["GPA"]
        prev_cgpa = mastersheet.iloc[idx]["PREVIOUS CGPA"]
        curr_cgpa = mastersheet.iloc[idx]["CURRENT CGPA"]
        print(f"  {exam_no}: GPA={current_gpa:.2f}, Prev CGPA={prev_cgpa}, Curr CGPA={curr_cgpa:.2f}")
    
    # ========================================================================
    # FIX: Replace the problematic AVERAGE calculation with safe mean function
    # ========================================================================
    def safe_mean(row, ordered_codes):
        """Calculate average excluding NOT REG values and handling mixed types."""
        numeric_values = []
        for code in ordered_codes:
            value = row[code]
            if not detect_not_registered_content(value):
                try:
                    numeric_values.append(float(value))
                except (ValueError, TypeError):
                    continue
        return np.mean(numeric_values) if numeric_values else 0

    mastersheet["AVERAGE"] = mastersheet.apply(
        lambda row: safe_mean(row, ordered_codes), axis=1
    ).round(0)
    # ENFORCED: Compute REMARKS with ENFORCED rule logic
    print(
        "\n🎯 Determining student statuses with ENFORCED probation/withdrawal rule..."
    )
    determine_student_status.debug_students = [
        "FCTCONS/ND24/109"
    ] # Add specific students to debug
    determine_student_status.count = 0
    mastersheet["REMARKS"] = mastersheet.apply(
        lambda row: determine_student_status(row, total_cu, pass_threshold), axis=1
    )
    # Validate the probation/withdrawal logic
    validate_probation_withdrawal_logic(mastersheet, total_cu)
    # FILTER OUT PREVIOUSLY WITHDRAWN STUDENTS
    mastersheet, removed_students = filter_out_withdrawn_students(
        mastersheet, semester_key
    )
    # Identify withdrawn students in this semester (after filtering)
    withdrawn_students = []
    for idx, row in mastersheet.iterrows():
        if row["REMARKS"] == "Withdrawn":
            exam_no = str(row["EXAM NUMBER"]).strip()
            withdrawn_students.append(exam_no)
            mark_student_withdrawn(exam_no, semester_key)
            print(f"🚫 Student {exam_no} marked as withdrawn in {semester_key}")
    # UPDATED: Identify probation students for tracking
    probation_students = []
    for idx, row in mastersheet.iterrows():
        if row["REMARKS"] == "Probation":
            exam_no = str(row["EXAM NUMBER"]).strip()
            probation_students.append(exam_no)
    # FIX 2: Create exam number to name mapping for proper name tracking
    exam_number_to_name_map = dict(zip(mastersheet["EXAM NUMBER"], mastersheet["NAME"]))
    # Update student tracker with current semester's students (after filtering)
    exam_numbers = mastersheet["EXAM NUMBER"].astype(str).str.strip().tolist()
    update_student_tracker(
        semester_key, exam_numbers, withdrawn_students, probation_students, exam_number_to_name_map
    )
    # Identify and save carryover students after processing
    carryover_students = identify_carryover_students(
        mastersheet, semester_key, set_name, pass_threshold
    )
    if carryover_students:
        carryover_dir = save_carryover_records(
            carryover_students, output_dir, set_name, semester_key
        )
        print(
            f"✅ Saved {len(carryover_students)} carryover records to: {carryover_dir}"
        )
        # ADD: Log the carryover record file path for debugging
        carryover_file = os.path.join(
            carryover_dir, f"co_student_{set_name}_{semester_key}_*.json"
        )
        print(f"📁 Carryover file pattern: {carryover_file}")
        # Print carryover summary
        total_failed_courses = sum(len(s["failed_courses"]) for s in carryover_students)
        print(
            f"📊 Carryover Summary: {total_failed_courses} failed courses across all students"
        )
        # Show most frequently failed courses
        course_fail_count = {}
        for student in carryover_students:
            for course in student["failed_courses"]:
                course_code = course["course_code"]
                course_fail_count[course_code] = (
                    course_fail_count.get(course_code, 0) + 1
                )
        if course_fail_count:
            top_failed = sorted(
                course_fail_count.items(), key=lambda x: x[1], reverse=True
            )[:5]
            print(f"📚 Most failed courses: {top_failed}")
    else:
        print("✅ No carryover students identified")
    # NEW: Sorting by REMARKS with custom order and secondary by GPA descending
    def status_key(s):
        return {"Passed": 0, "Resit": 1, "Probation": 2, "Withdrawn": 3}.get(s, 4)

    mastersheet["status_key"] = mastersheet["REMARKS"].apply(status_key)
    mastersheet = (
        mastersheet.sort_values(by=["status_key", "GPA"], ascending=[True, False])
        .drop(columns=["status_key"])
        .reset_index(drop=True)
    )
    if "S/N" not in mastersheet.columns:
        mastersheet.insert(0, "S/N", range(1, len(mastersheet) + 1))
    else:
        mastersheet["S/N"] = range(1, len(mastersheet) + 1)
        cols = list(mastersheet.columns)
        if cols[0] != "S/N":
            cols.remove("S/N")
            mastersheet = mastersheet[["S/N"] + cols]
    course_cols = ordered_codes
    # UPDATED: Include Total Registered CU, PREVIOUS CGPA, and CURRENT CGPA in output columns
    out_cols = (
        ["S/N", "EXAM NUMBER", "NAME"]
        + course_cols
        + [
            "FAILED COURSES",
            "REMARKS", 
            "CU Passed",
            "CU Failed",
            "Total Registered CU",  # Add this
            "TCPE",
            "GPA",
            "PREVIOUS CGPA",  # ADDED: Previous CGPA column
            "CURRENT CGPA",   # ADDED: Current CGPA column
            "AVERAGE",
        ]
    )
    for c in out_cols:
        if c not in mastersheet.columns:
            mastersheet[c] = pd.NA
    mastersheet = mastersheet[out_cols]
    # FIXED: Create proper output directory structure - all files go directly to the set output directory
    out_xlsx = os.path.join(output_dir, f"mastersheet_{ts}.xlsx")
    if not os.path.exists(out_xlsx):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
    else:
        wb = load_workbook(out_xlsx)
    if sem not in wb.sheetnames:
        ws = wb.create_sheet(title=sem)
    else:
        ws = wb[sem]
    try:
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
    except Exception:
        pass
    ws.insert_rows(1, 2)
    logo_path_norm = os.path.normpath(logo_path) if logo_path else None
    if logo_path_norm and os.path.exists(logo_path_norm):
        try:
            img = XLImage(logo_path_norm)
            img.width, img.height = 110, 110
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"⚠ Could not place logo: {e}")
    # =========================================================================
    # APPLIED FIX: UPDATED HEADER SECTION WITH SPACING AND ROW HEIGHT FIXES
    # =========================================================================
    # UPDATED HEADER: Dynamic title based on semester being processed
    ws.merge_cells("C1:Q1")
    title_cell = ws["C1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="1E90FF", end_color="1E90FF", fill_type="solid"
    )
    border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    title_cell.border = border
    # UPDATED: Use new header format with DEPARTMENT OF NURSING
    ws.merge_cells("C2:Q2")
    dept_cell = ws["C2"]
    dept_cell.value = "DEPARTMENT OF NURSING"
    dept_cell.font = Font(bold=True, size=14, color="000000")
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")
    dept_cell.fill = PatternFill(
        start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
    )
    # UPDATED: Dynamic title based on semester being processed
    ws.merge_cells("C3:Q3")
    subtitle_cell = ws["C3"]
    
    # Create dynamic title based on semester
    # FIXED: Use current date dynamically
    current_date = datetime.now().strftime("%B %d, %Y")
    if "SECOND-YEAR-SECOND-SEMESTER" in semester_key:
        # Use the specific format for NDII Second Semester
        exam_title = f"NATIONAL DIPLOMA YEAR TWO SECOND SEMESTER EXAMINATIONS RESULT — {current_date}"
    else:
        # Dynamic title for other semesters
        year, semester_num, level_display, semester_display, set_code = get_semester_display_info(semester_key)
        exam_title = f"NATIONAL DIPLOMA {level_display} {semester_display} EXAMINATIONS RESULT — {current_date}"
    
    subtitle_cell.value = exam_title
    subtitle_cell.font = Font(bold=True, size=12, color="000000")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    # ADDED: Create space between header and course titles by adding empty rows
    ws.append([])  # Empty row 4 - creates space
    ws.append([])  # Empty row 5 - creates more space
    # FIX 5: Add visual indicator in header for upgraded scores
    start_row = 6  # CHANGED: Now starting at row 6 due to added empty rows
    if upgrade_min_threshold is not None:
        # Add upgrade notice row
        ws.append(
            ["", "", f"UPGRADED SCORES: {upgrade_min_threshold}–49 → 50"]
            + [""] * len(ordered_codes)
            + [""] * 6
        )
        # Merge cells for the notice
        ws.merge_cells(f"C{start_row}:E{start_row}")
        notice_cell = ws.cell(row=start_row, column=3)
        notice_cell.font = Font(bold=True, size=10, color="FFFFFF")
        notice_cell.fill = PatternFill(
            start_color="FF6B35", end_color="FF6B35", fill_type="solid"
        )  # Orange background
        notice_cell.alignment = Alignment(horizontal="center", vertical="center")
        start_row += 1  # Increment start_row for subsequent rows
    # ADDED: Additional space after upgrade notice (if present) or after header (if no upgrade)
    ws.append([])  # Empty row for spacing
    start_row += 1  # Adjust start_row for the empty row we just added
    # FIX 1: EXPAND ROW HEIGHT FOR COURSE TITLES (ROW 8)
    display_course_titles = []
    for t in ordered_titles:
        course_info = course_map.get(t)
        if course_info and course_info["code"] in ordered_codes:
            display_course_titles.append(course_info["original_name"])
    # Calculate the maximum title length to determine appropriate row height
    max_title_length = max([len(title) for title in display_course_titles]) if display_course_titles else 0
    print(f"📏 Longest course title length: {max_title_length} characters")
    
    # Set appropriate row height based on title length - FIXED ROW HEIGHT
    course_title_row_height = 60  # Increased height to accommodate wrapped text
    ws.row_dimensions[start_row].height = course_title_row_height
    ws.append([""] * 3 + display_course_titles + [""] * 6)
    
    # Apply text wrapping and rotation to course titles
    for i, cell in enumerate(
        ws[start_row][3 : 3 + len(display_course_titles)], start=3
    ):
        cell.alignment = Alignment(
            horizontal="center", 
            vertical="center", 
            text_rotation=45,
            wrap_text=True  # ADDED: Enable text wrapping
        )
        cell.font = Font(bold=True, size=9)
    
    # FIX 2: AUTO-FIT COLUMN WIDTHS FOR ALL COLUMNS
    cu_list = [filtered_credit_units.get(c, "") for c in ordered_codes]
    ws.append([""] * 3 + cu_list + [""] * 6)
    for cell in ws[start_row + 1][3 : 3 + len(cu_list)]:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=135
        )
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
    headers = out_cols
    ws.append(headers)
    for cell in ws[start_row + 2]:
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="4A90E2", end_color="4A90E2", fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    for _, r in mastersheet.iterrows():
        rowvals = [r[col] for col in headers]
        ws.append(rowvals)
    # =========================================================================
    # APPLIED FIX: UPDATED FREEZE PANES TO ACCOUNT FOR NEW ROW POSITIONS
    # =========================================================================
    # Adjust freeze panes based on new structure
    freeze_row = start_row + 4  # CHANGED: Added +1 to account for extra spacing
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=start_row + 3, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
    # FIX 3: Fix the Excel colorization to properly identify upgraded scores
    not_reg_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray for NOT REG
    upgraded_fill = PatternFill(
        start_color="E6FFCC", end_color="E6FFCC", fill_type="solid"
    ) # Light green for upgraded scores
    passed_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    ) # Normal green for passed
    failed_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    ) # White for failed
    # CRITICAL FIX: Track which scores were upgraded by comparing against original merged data
    upgraded_scores_tracker = {}
    if upgrade_min_threshold is not None:
        # Build tracker of which student/course combinations were upgraded
        for idx in mastersheet.index:
            exam_no = str(mastersheet.at[idx, "EXAM NUMBER"]).strip()
            for code in ordered_codes:
                # Check if this score is exactly 50 (potentially upgraded)
                if mastersheet.at[idx, code] == 50:
                    # Mark as potentially upgraded
                    if exam_no not in upgraded_scores_tracker:
                        upgraded_scores_tracker[exam_no] = set()
                    upgraded_scores_tracker[exam_no].add(code)
    for idx, code in enumerate(ordered_codes, start=4):
        col_letter = get_column_letter(idx)
        for r_idx in range(start_row + 3, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=idx)
            cell_value = cell.value
            
            # Get exam number for this row
            exam_no_cell = ws.cell(row=r_idx, column=2)  # EXAM NUMBER is column B (index 2)
            exam_no = str(exam_no_cell.value).strip() if exam_no_cell.value else ""
            
            # Check for NOT REG content
            if detect_not_registered_content(cell_value):
                cell.fill = not_reg_fill
                cell.font = Font(color="666666", italic=True)
                continue
                
            try:
                val = float(cell_value) if cell_value not in (None, "") else 0
                
                # CRITICAL FIX: Check if this score was upgraded
                is_upgraded = (
                    upgrade_min_threshold is not None 
                    and exam_no in upgraded_scores_tracker 
                    and code in upgraded_scores_tracker.get(exam_no, set())
                    and val == 50
                )
                
                if is_upgraded:
                    # This score was upgraded - use special color
                    cell.fill = upgraded_fill
                    cell.font = Font(color="006600", bold=True)
                elif val >= pass_threshold:
                    cell.fill = passed_fill
                    cell.font = Font(color="006100")
                else:
                    cell.fill = failed_fill
                    cell.font = Font(color="FF0000", bold=True)
            except (ValueError, TypeError):
                continue
    # Apply specific column alignments
    left_align_columns = [
        "CU Passed",
        "CU Failed",
        "Total Registered CU",
        "TCPE",
        "GPA",
        "PREVIOUS CGPA",  # ADDED
        "CURRENT CGPA",   # ADDED
        "AVERAGE",
        "FAILED COURSES",
        "REMARKS",
    ]
    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in left_align_columns:
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row + 3, ws.max_row + 1): # Start from data rows
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        # Center align S/N column
        elif col_name == "S/N":
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row + 3, ws.max_row + 1): # Start from data rows
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    # NEW: Wrap text for FAILED COURSES and REMARKS
    failed_col_idx = (
        headers.index("FAILED COURSES") + 1 if "FAILED COURSES" in headers else None
    )
    remarks_col_idx = headers.index("REMARKS") + 1 if "REMARKS" in headers else None
    for row_idx in range(start_row + 3, ws.max_row + 1):
        for col in [failed_col_idx, remarks_col_idx]:
            if col:
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
    # UPDATED: Color coding for REMARKS column - ADDED PROBATION COLOR
    passed_remarks_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    ) # green
    resit_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    ) # yellow
    probation_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    ) # orange for probation
    withdrawn_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ) # red
    for r_idx in range(start_row + 3, ws.max_row + 1):
        cell = ws.cell(row=r_idx, column=remarks_col_idx)
        if cell.value == "Passed":
            cell.fill = passed_remarks_fill
        elif cell.value == "Resit":
            cell.fill = resit_fill
        elif cell.value == "Probation": # NEW: Color for probation
            cell.fill = probation_fill
        elif cell.value == "Withdrawn":
            cell.fill = withdrawn_fill
    # FIX 2: AUTO-FIT COLUMN WIDTHS FOR ALL COLUMNS PROFESSIONALLY
    print("🔄 Auto-fitting column widths for professional appearance...")
    
    # Calculate optimal column widths with special handling for all columns
    longest_name_len = (
        max([len(str(x)) for x in mastersheet["NAME"].fillna("")])
        if "NAME" in mastersheet.columns
        else 10
    )
    name_col_width = min(max(longest_name_len + 2, 10), NAME_WIDTH_CAP)
    # Enhanced FAILED COURSES column width calculation
    longest_failed_len = max(
        [len(str(x)) for x in mastersheet["FAILED COURSES"].fillna("")]
    )
    failed_col_width = min(max(longest_failed_len + 4, 40), 80)
    # REMARKS column width
    longest_remark_len = max([len(str(x)) for x in mastersheet["REMARKS"].fillna("")])
    remarks_col_width = min(max(longest_remark_len + 4, 15), 30)
    # PREVIOUS CGPA and CURRENT CGPA column widths
    prev_cgpa_width = min(max(max([len(str(x)) for x in mastersheet["PREVIOUS CGPA"].fillna("")]) + 2, 12), 15)
    curr_cgpa_width = min(max(max([len(str(x)) for x in mastersheet["CURRENT CGPA"].fillna("")]) + 2, 12), 15)
    
    # Apply column widths with professional auto-fitting
    for col_idx, col in enumerate(ws.columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Calculate maximum content length for this column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Apply professional width constraints based on column type
        if col_idx == 1: # S/N
            ws.column_dimensions[column_letter].width = 6
        elif column_letter == "B" or headers[col_idx - 1] in ["EXAM NUMBER", "EXAM NO"]:
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 15), 20)
        elif headers[col_idx - 1] == "NAME":
            ws.column_dimensions[column_letter].width = name_col_width
        elif 4 <= col_idx < 4 + len(ordered_codes): # course columns
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 8), 12)
        elif headers[col_idx - 1] == "FAILED COURSES":
            ws.column_dimensions[column_letter].width = failed_col_width
        elif headers[col_idx - 1] == "REMARKS":
            ws.column_dimensions[column_letter].width = remarks_col_width
        elif headers[col_idx - 1] == "PREVIOUS CGPA":
            ws.column_dimensions[column_letter].width = prev_cgpa_width
        elif headers[col_idx - 1] == "CURRENT CGPA":
            ws.column_dimensions[column_letter].width = curr_cgpa_width
        elif headers[col_idx - 1] in ["CU Passed", "CU Failed", "Total Registered CU", "TCPE", "GPA", "AVERAGE"]:
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 15)
        else:
            # Default auto-fit for other columns
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 8), 20)
    # NEW: Enhanced course statistics with NOT REG information
    fails_per_course, not_reg_per_course, registered_per_course = calculate_course_statistics(
        mastersheet, ordered_codes, pass_threshold
    )
    # Add footer with enhanced statistics
    footer_vals1 = [""] * 2 + ["FAILS PER COURSE:"] + [fails_per_course.get(c, 0) for c in ordered_codes] + [""] * (len(headers) - 3 - len(ordered_codes))
    ws.append(footer_vals1)
    footer_vals2 = [""] * 2 + ["NOT REG PER COURSE:"] + [not_reg_per_course.get(c, 0) for c in ordered_codes] + [""] * (len(headers) - 3 - len(ordered_codes))
    ws.append(footer_vals2)
    footer_vals3 = [""] * 2 + ["REGISTERED STUDENTS:"] + [registered_per_course.get(c, 0) for c in ordered_codes] + [""] * (len(headers) - 3 - len(ordered_codes))
    ws.append(footer_vals3)
    # Style the footer rows
    for row_num in [ws.max_row - 2, ws.max_row - 1, ws.max_row]:
        for cell in ws[row_num]:
            if 4 <= cell.column < 4 + len(ordered_codes):
                if row_num == ws.max_row - 2:  # Fails row
                    cell.fill = PatternFill(start_color="F0E68C", end_color="F0E68C", fill_type="solid")  # Light yellow
                elif row_num == ws.max_row - 1:  # NOT REG row  
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light purple
                else:  # Registered students row
                    cell.fill = PatternFill(start_color="E6FFCC", end_color="E6FFCC", fill_type="solid")  # Light green
                    
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif cell.column == 3:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
    # UPDATED: COMPREHENSIVE SUMMARY BLOCK - ENFORCED RULE WITH NOT REG INFORMATION
    total_students = len(mastersheet)
    passed_all = len(mastersheet[mastersheet["REMARKS"] == "Passed"])
    
    # Count students by status with ENFORCED rule
    resit_students = len(mastersheet[mastersheet["REMARKS"] == "Resit"])
    probation_students = len(mastersheet[mastersheet["REMARKS"] == "Probation"])
    withdrawn_students = len(mastersheet[mastersheet["REMARKS"] == "Withdrawn"])
    # ENFORCED RULE: Break down students by the new criteria
    resit_rule_students = len(
        mastersheet[
            (mastersheet["REMARKS"] == "Resit")
            & (mastersheet["CU Passed"] / total_cu >= 0.45)
            & (mastersheet["GPA"] >= 2.00)
        ]
    )
    probation_rule_students = len(
        mastersheet[
            (mastersheet["REMARKS"] == "Probation")
            & (mastersheet["CU Passed"] / total_cu >= 0.45)
            & (mastersheet["GPA"] < 2.00)
        ]
    )
    withdrawn_rule_students = len(
        mastersheet[
            (mastersheet["REMARKS"] == "Withdrawn")
            & (mastersheet["CU Passed"] / total_cu < 0.45)
        ]
    )
    # NEW: Calculate NOT REG statistics
    total_not_reg_students = 0
    not_reg_courses_count = {}
    for code in ordered_codes:
        if code in mastersheet.columns:
            not_reg_count = not_reg_per_course.get(code, 0)
            if not_reg_count > 0:
                total_not_reg_students = max(total_not_reg_students, not_reg_count)  # Approximate count
                not_reg_courses_count[code] = not_reg_count
    # Add summary rows
    ws.append([])
    ws.append(["SUMMARY"])
    # FIX 4: Ensure upgrade summary is displayed properly
    # Add upgrade notice FIRST if applicable
    if upgrade_min_threshold is not None and upgraded_scores_count > 0:
        ws.append(
            [
                f"✅ MANAGEMENT DECISION: All scores between {upgrade_min_threshold}–49 were upgraded to 50 ({upgraded_scores_count} scores upgraded)"
            ]
        )
        ws.append([])  # Add blank line for separation
    # Then add other summary items
    ws.append(
        [f"A total of {total_students} students registered and sat for the Examination"]
    )
    ws.append(
        [
            f"A total of {passed_all} students passed in all courses registered and are to proceed to the next semester"
        ]
    )
    ws.append(
        [
            f"A total of {resit_rule_students} students with Grade Point Average (GPA) of 2.00 and above who passed ≥45% of credit units failed various courses, and are to resit these courses in the next session."
        ]
    )
    ws.append(
        [
            f"A total of {probation_rule_students} students with Grade Point Average (GPA) below 2.00 who passed ≥45% of credit units failed various courses, and are placed on Probation, to resit these courses in the next session."
        ]
    )
    ws.append(
        [
            f"A total of {withdrawn_rule_students} students who passed less than 45% of their registered credit units have been advised to withdraw"
        ]
    )
    
    # NEW: Add NOT REG information to summary
    if total_not_reg_students > 0:
        ws.append([
            f"A total of {total_not_reg_students} candidates did not register for certain courses and were excluded from assessment in those courses."
        ])
        ws.append([
            "NOT REGISTERED candidates are not included in pass/fail statistics for the courses they did not register for."
        ])
        
        # Show courses with NOT REG candidates
        if not_reg_courses_count:
            not_reg_courses_str = ", ".join([f"{code}({count})" for code, count in not_reg_courses_count.items()])
            ws.append([
                f"Courses with NOT REGISTERED candidates: {not_reg_courses_str}"
            ])
            
    if removed_students:
        ws.append(
            [
                f"NOTE: {len(removed_students)} previously withdrawn students were removed from this semester's results as they should not be processed."
            ]
        )
    ws.append(
        [
            "The above decisions are in line with the provisions of the General Information Section of the NMCN/NBTE Examinations Regulations (Pg 4) adopted by the College."
        ]
    )
    ws.append([])
    ws.append(
        [
            "________________________",
            "",
            "",
            "________________________",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Mrs. Abini Hauwa",
            "",
            "",
            "Dr. Kigbu Job Yaro",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Head of Exams",
            "",
            "",
            "HOD Nursing",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    wb.save(out_xlsx)
    print(f"✅ Mastersheet saved: {out_xlsx}")
    print(f"📊 CGPA columns added to Excel: PREVIOUS CGPA and CURRENT CGPA")
    
    # Generate individual student PDF with previous CGPAs and Cumulative CGPA
    safe_sem = re.sub(r"[^\w\-]", "_", sem)
    student_pdf_path = os.path.join(
        output_dir, f"mastersheet_students_{ts}_{safe_sem}.pdf"
    )
    
    print(f"📊 FINAL CHECK before PDF generation:")
    print(f" Previous CGPAs loaded: {len(previous_cgpas)}")
    print(
        f" Cumulative CGPA data available for: {len(cumulative_cgpa_data) if cumulative_cgpa_data else 0} students"
    )
    if previous_cgpas:
        sample = list(previous_cgpas.items())[:3]
        print(f" Sample CGPAs: {sample}")
    try:
        # FIX: Check if ordered_codes is valid before passing to PDF generation
        if ordered_codes and len(ordered_codes) > 0:
            # FIX: Call generate_individual_student_pdf without resit_count parameter
            pdf_success = generate_individual_student_pdf(
                mastersheet,
                student_pdf_path,
                sem,
                logo_path=logo_path_norm,
                prev_mastersheet_df=None,
                filtered_credit_units=filtered_credit_units,
                ordered_codes=ordered_codes,
                course_titles_map=course_titles,
                previous_cgpas=previous_cgpas,
                cumulative_cgpa_data=cumulative_cgpa_data,
                total_cu=total_cu,
                pass_threshold=pass_threshold,
                upgrade_min_threshold=upgrade_min_threshold,
            ) # PASS THE UPGRADE THRESHOLD TO PDF
            if pdf_success:
                print(f"✅ PDF generated successfully for {sem}")
            else:
                print(f"⚠️ PDF generation failed for {sem}")
        else:
            print(f"⚠️ No course codes found for {sem}, skipping PDF generation")
    except Exception as e:
        print(f"❌ Failed to generate student PDF for {sem}: {e}")
        import traceback
        traceback.print_exc()
    return mastersheet

def process_semester_files(
    semester_key,
    raw_files,
    raw_dir,
    output_dir,
    ts,
    pass_threshold,
    semester_course_maps,
    semester_credit_units,
    semester_lookup,
    semester_course_titles,
    logo_path,
    set_name,
    previous_cgpas=None,
    upgrade_min_threshold=None,
):
    """
    Process all files for a specific semester with carryover integration.
    """
    print(f"\n{'='*60}")
    print(f"PROCESSING SEMESTER: {semester_key}")
    print(f"{'='*60}")
    
    # Filter files for this semester
    normalized_key = semester_key.replace("ND-", "").upper()
    semester_files = [
        f for f in raw_files if normalized_key in f.upper().replace("ND-", "")
    ]
    
    if not semester_files:
        print(f"⚠️ No files found for semester {semester_key}")
        print(f"🔍 Available files: {raw_files}")
        return None
        
    print(f"📁 Found {len(semester_files)} files for {semester_key}: {semester_files}")
    # Check for existing carryover files
    existing_carryover_files = check_existing_carryover_files(
        raw_dir, set_name, semester_key
    )
    if existing_carryover_files:
        print(f"📋 Found existing carryover files: {existing_carryover_files}")
        print("ℹ️ Carryover processing will be available after regular processing")
    # Process each file for this semester
    mastersheet_result = None
    for rf in semester_files:
        raw_path = os.path.join(raw_dir, rf)
        print(f"\n📄 Processing: {rf}")
        try:
            # Load previous CGPAs for this specific semester
            current_previous_cgpas = (
                load_previous_cgpas_from_processed_files(output_dir, semester_key, ts)
                if previous_cgpas is None
                else previous_cgpas
            )
            
            # Load Cumulative CGPA data (all previous semesters)
            cumulative_cgpa_data = load_all_previous_cgpas_for_cumulative(
                output_dir, semester_key, ts
            )
            
            # Process the file with both previous and cumulative CGPA data
            result = process_single_file(
                raw_path,
                output_dir,
                ts,
                pass_threshold,
                semester_course_maps,
                semester_credit_units,
                semester_lookup,
                semester_course_titles,
                logo_path,
                semester_key,
                set_name,
                current_previous_cgpas,  # This should have the previous CGPA data
                cumulative_cgpa_data,
                upgrade_min_threshold,
            )
            if result is not None:
                print(f"✅ Successfully processed {rf}")
                mastersheet_result = result
            else:
                print(f"❌ Failed to process {rf}")
        except Exception as e:
            print(f"❌ Error processing {rf}: {e}")
            import traceback
            traceback.print_exc()
    # ADD: Verify carryover records were created
    carryover_records_dir = os.path.join(output_dir, "CARRYOVER_RECORDS")
    if os.path.exists(carryover_records_dir):
        json_files = glob.glob(
            os.path.join(
                carryover_records_dir, f"co_student_{set_name}_{semester_key}_*.json"
            )
        )
        if json_files:
            print(f"✅ Carryover records created: {len(json_files)} file(s)")
            print(f"📝 Latest: {sorted(json_files)[-1]}")
        else:
            print(f"⚠️ No carryover records found for {semester_key}")
            
    return mastersheet_result

# ----------------------------
# Main runner (Enhanced)
# ----------------------------

def main():
    print("Starting ND Examination Results Processing with Data Transformation and NOT REG handling...")
    ts = datetime.now().strftime(TIMESTAMP_FMT)
    
    # Initialize trackers
    initialize_student_tracker()
    initialize_carryover_tracker()
    initialize_inactive_students_tracker()
    # Check if running in web mode
    if is_web_mode():
        uploaded_file_path = get_uploaded_file_path()
        if uploaded_file_path and os.path.exists(uploaded_file_path):
            print("🔧 Running in WEB MODE with uploaded file")
            # This would need to be adapted for your specific uploaded file processing
            print(
                "⚠️ Uploaded file processing for individual files not fully implemented in this version"
            )
            return
    # Get parameters from form
    params = get_form_parameters()
    
    # Use the parameters
    global DEFAULT_PASS_THRESHOLD
    DEFAULT_PASS_THRESHOLD = params["pass_threshold"]
    base_dir_norm = normalize_path(BASE_DIR)
    print(f"Using base directory: {base_dir_norm}")
    # Check if we should use interactive or non-interactive mode
    if should_use_interactive_mode():
        print("🔧 Running in INTERACTIVE mode (CLI)")
        try:
            (
                semester_course_maps,
                semester_credit_units,
                semester_lookup,
                semester_course_titles,
            ) = load_course_data()
        except Exception as e:
            print(f"❌ Could not load course data: {e}")
            return
        # Get available sets and let user choose
        available_sets = get_available_sets(base_dir_norm)
        if not available_sets:
            print(f"No ND-* directories found in {base_dir_norm}. Nothing to process.")
            print(f"Available directories: {os.listdir(base_dir_norm)}")
            return
        print(f"📚 Found {len(available_sets)} available sets: {available_sets}")
        
        # Let user choose which set(s) to process
        sets_to_process = get_user_set_choice(available_sets)
        print(f"\n🎯 PROCESSING SELECTED SETS: {sets_to_process}")
        for nd_set in sets_to_process:
            print(f"\n{'='*60}")
            print(f"PROCESSING SET: {nd_set}")
            print(f"{'='*60}")
            
            # Generate a single timestamp for this set processing
            ts = datetime.now().strftime(TIMESTAMP_FMT)
            
            # UPDATED: Raw and clean directories now under ND folder
            raw_dir = normalize_path(
                os.path.join(base_dir_norm, "ND", nd_set, "RAW_RESULTS")
            )
            clean_dir = normalize_path(
                os.path.join(base_dir_norm, "ND", nd_set, "CLEAN_RESULTS")
            )
            
            # Create directories if they don't exist
            os.makedirs(raw_dir, exist_ok=True)
            os.makedirs(clean_dir, exist_ok=True)
            # Check if raw directory exists and has files
            if not os.path.exists(raw_dir):
                print(f"⚠️ RAW_RESULTS directory not found: {raw_dir}")
                continue
                
            raw_files = [
                f
                for f in os.listdir(raw_dir)
                if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
            ]
            if not raw_files:
                print(f"⚠️ No raw files in {raw_dir}; skipping {nd_set}")
                print(f" Available files: {os.listdir(raw_dir)}")
                continue
            print(f"📁 Found {len(raw_files)} raw files in {nd_set}: {raw_files}")
            # Create a single timestamped folder for this set
            set_output_dir = os.path.join(clean_dir, f"{nd_set}_RESULT-{ts}")
            os.makedirs(set_output_dir, exist_ok=True)
            print(f"📁 Created set output directory: {set_output_dir}")
            # Get user choice for which semesters to process
            semesters_to_process = get_user_semester_choice()
            print(
                f"\n🎯 PROCESSING SELECTED SEMESTERS for {nd_set}: {[get_semester_display_info(sem)[3] for sem in semesters_to_process]}"
            )
            # Process selected semesters in the correct order
            for semester_key in semesters_to_process:
                if semester_key not in SEMESTER_ORDER:
                    print(f"⚠️ Skipping unknown semester: {semester_key}")
                    continue
                    
                # Check if there are files for this semester
                semester_files_exist = False
                for rf in raw_files:
                    detected_sem, _, _, _, _, _ = detect_semester_from_filename(rf)
                    if detected_sem == semester_key:
                        semester_files_exist = True
                        break
                        
                if semester_files_exist:
                    print(f"\n🎯 Processing {semester_key} in {nd_set}...")
                    process_semester_files(
                        semester_key,
                        raw_files,
                        raw_dir,
                        set_output_dir,
                        ts,
                        DEFAULT_PASS_THRESHOLD,
                        semester_course_maps,
                        semester_credit_units,
                        semester_lookup,
                        semester_course_titles,
                        DEFAULT_LOGO_PATH,
                        nd_set,
                    )
                else:
                    print(
                        f"⚠️ No files found for {semester_key} in {nd_set}, skipping..."
                    )
            # Create CGPA_SUMMARY and ANALYSIS worksheets
            mastersheet_path = os.path.join(set_output_dir, f"mastersheet_{ts}.xlsx")
            if os.path.exists(mastersheet_path):
                print(f"📊 Creating CGPA_SUMMARY and ANALYSIS worksheets...")
                
                # FIX 2: Identify inactive students FIRST before creating sheets
                print(f"🔍 Identifying inactive students...")
                identify_inactive_students()
                
                # THEN create the sheets that depend on this data
                create_cgpa_summary_sheet(mastersheet_path, ts)
                create_analysis_sheet(mastersheet_path, ts)  # Now INACTIVE_STUDENTS will be populated
                
                print(f"✅ Successfully added all worksheets (CGPA_SUMMARY, ANALYSIS)")
            # Create ZIP of the entire set results
            try:
                zip_path = os.path.join(clean_dir, f"{nd_set}_RESULT-{ts}.zip")
                zip_success = create_zip_folder(set_output_dir, zip_path)
                if zip_success:
                    print(f"✅ ZIP file created: {zip_path}")
                    # Verify file size
                    if os.path.exists(zip_path):
                        zip_size = os.path.getsize(zip_path)
                        zip_size_mb = zip_size / (1024 * 1024)
                        print(f"📦 ZIP file size: {zip_size_mb:.2f} MB")
                else:
                    print(f"❌ Failed to create ZIP file for {nd_set}")
            except Exception as e:
                print(f"⚠️ Failed to create ZIP for {nd_set}: {e}")
        # Print student tracking summary
        print(f"\n📊 STUDENT TRACKING SUMMARY:")
        print(f"Total unique students tracked: {len(STUDENT_TRACKER)}")
        print(f"Total withdrawn students: {len(WITHDRAWN_STUDENTS)}")
        print(f"Total inactive students: {len(INACTIVE_STUDENTS)}")
        
        # Print carryover summary
        if CARRYOVER_STUDENTS:
            print(f"\n📋 CARRYOVER STUDENT SUMMARY:")
            print(f"Total carryover students: {len(CARRYOVER_STUDENTS)}")
            # Count by semester
            semester_counts = {}
            for student_key, data in CARRYOVER_STUDENTS.items():
                semester = data["semester"]
                semester_counts[semester] = semester_counts.get(semester, 0) + 1
            for semester, count in semester_counts.items():
                print(f" {semester}: {count} students")
        # Print inactive students summary
        if INACTIVE_STUDENTS:
            print(f"\n📋 INACTIVE STUDENTS SUMMARY:")
            cgpa_active_count = sum(1 for s in INACTIVE_STUDENTS.values() if s.get("cgpa_status") == "Active in CGPA")
            regular_inactive_count = len(INACTIVE_STUDENTS) - cgpa_active_count
            print(f" Regular inactive (missing intermediate semesters): {regular_inactive_count}")
            print(f" CGPA active but missing from recent semester: {cgpa_active_count}")
            
            # Show sample of inactive students
            print(f" Sample inactive students:")
            for i, (exam_no, data) in enumerate(list(INACTIVE_STUDENTS.items())[:5]):
                print(f"  {i+1}. {exam_no}: {data['name']} - Present in {len(data['semesters_present'])} semesters, missing {len(data['missing_semesters'])}: {', '.join(data['missing_semesters'])}")
        # Print withdrawn students who reappeared
        reappeared_count = 0
        for exam_no, data in WITHDRAWN_STUDENTS.items():
            if data["reappeared_semesters"]:
                reappeared_count += 1
                print(
                    f"🚨 {exam_no}: Withdrawn in {data['withdrawn_semester']}, reappeared in {data['reappeared_semesters']}"
                )
        if reappeared_count > 0:
            print(
                f"🚨 ALERT: {reappeared_count} previously withdrawn students have reappeared in later semesters!"
            )
        # Analyze student progression
        sem_counts = {}
        for student_data in STUDENT_TRACKER.values():
            sem_count = len(student_data["semesters_present"])
            if sem_count not in sem_counts:
                sem_counts[sem_count] = 0
            sem_counts[sem_count] += 1
        for sem_count, student_count in sorted(sem_counts.items()):
            print(f"Students present in {sem_count} semester(s): {student_count}")
        print("\n✅ ND Examination Results Processing completed successfully.")
    else:
        print("🔧 Running in NON-INTERACTIVE mode (Web)")
        # NEW: Check if this is carryover processing mode
        if params.get("process_carryover", False):
            print(
                "🎯 Detected CARRYOVER processing mode - redirecting to integrated_carryover_processor.py"
            )
            # Set environment variables for the carryover processor
            os.environ["CARRYOVER_FILE_PATH"] = params["carryover_file_path"]
            os.environ["SET_NAME"] = params["selected_set"]
            os.environ["SEMESTER_KEY"] = (
                params["selected_semesters"][0] if params["selected_semesters"] else ""
            )
            os.environ["BASE_DIR"] = BASE_DIR
            
            # Path to the integrated_carryover_processor.py script
            carryover_script_path = os.path.join(
                os.path.dirname(__file__), "integrated_carryover_processor.py"
            )
            if not os.path.exists(carryover_script_path):
                print(
                    f"❌ Carryover processor script not found: {carryover_script_path}"
                )
                return False
                
            print(f"🚀 Running carryover processor: {carryover_script_path}")
            # Run the carryover processor
            result = subprocess.run(
                [sys.executable, carryover_script_path], capture_output=True, text=True
            )
            # Print the output and return
            print(result.stdout)
            if result.stderr:
                print(result.stderr)
            return result.returncode == 0
        else:
            # Regular processing mode
            success = process_in_non_interactive_mode(params, base_dir_norm)
            if success:
                print("✅ ND Examination Results Processing completed successfully")
            else:
                print("❌ ND Examination Results Processing failed")
        return

def process_in_non_interactive_mode(params, base_dir_norm):
    """Process exams in non-interactive mode for web interface."""
    print("🔧 Running in NON-INTERACTIVE mode (web interface)")
    
    # Use parameters from environment variables
    selected_set = params["selected_set"]
    selected_semesters = params["selected_semesters"]
    
    # FIX: Normalize semester names to uppercase for consistent matching
    selected_semesters = [sem.upper() for sem in selected_semesters]
    print(f"🎯 Processing semesters (normalized): {selected_semesters}")
    
    # Get upgrade threshold from environment variable if provided
    upgrade_min_threshold = get_upgrade_threshold_from_env()
    # Get available sets
    available_sets = get_available_sets(base_dir_norm)
    if not available_sets:
        print("❌ No ND sets found")
        return False
        
    # Remove ND-COURSES from available sets if present
    available_sets = [s for s in available_sets if s != "ND-COURSES"]
    if not available_sets:
        print("❌ No valid ND sets found (only ND-COURSES present)")
        return False
    # Determine which sets to process
    if selected_set == "all":
        sets_to_process = available_sets
        print(f"🎯 Processing ALL sets: {sets_to_process}")
    else:
        if selected_set in available_sets:
            sets_to_process = [selected_set]
            print(f"🎯 Processing selected set: {selected_set}")
        else:
            print(f"⚠️ Selected set '{selected_set}' not found, processing all sets")
            sets_to_process = available_sets
    # Load course data once
    try:
        (
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
        ) = load_course_data()
        print(
            f"✅ Loaded course data for semesters: {list(semester_course_maps.keys())}"
        )
    except Exception as e:
        print(f"❌ Could not load course data: {e}")
        return False
    # Initialize carryover tracker
    initialize_carryover_tracker()
    initialize_inactive_students_tracker()
    # Process each set and semester
    total_processed = 0
    for nd_set in sets_to_process:
        print(f"\n{'='*60}")
        print(f"PROCESSING SET: {nd_set}")
        print(f"{'='*60}")
        
        # Generate a single timestamp for this set processing
        ts = datetime.now().strftime(TIMESTAMP_FMT)
        
        # UPDATED: Raw and clean directories now under ND folder
        raw_dir = normalize_path(
            os.path.join(base_dir_norm, "ND", nd_set, "RAW_RESULTS")
        )
        clean_dir = normalize_path(
            os.path.join(base_dir_norm, "ND", nd_set, "CLEAN_RESULTS")
        )
        
        # Create directories if they don't exist
        os.makedirs(raw_dir, exist_ok=True)
        os.makedirs(clean_dir, exist_ok=True)
        if not os.path.exists(raw_dir):
            print(f"⚠️ RAW_RESULTS directory not found: {raw_dir}")
            continue
            
        raw_files = [
            f
            for f in os.listdir(raw_dir)
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
        ]
        if not raw_files:
            print(f"⚠️ No raw files in {raw_dir}; skipping {nd_set}")
            continue
            
        print(f"📁 Found {len(raw_files)} raw files in {nd_set}: {raw_files}")
        # Create a single timestamped folder for this set
        set_output_dir = os.path.join(clean_dir, f"{nd_set}_RESULT-{ts}")
        os.makedirs(set_output_dir, exist_ok=True)
        print(f"📁 Created set output directory: {set_output_dir}")
        # Process selected semesters - FIXED: Use normalized (uppercase) semester names
        for semester_key in selected_semesters:
            # FIX: Check if semester exists in course data (case-sensitive)
            if semester_key not in semester_course_maps:
                print(
                    f"⚠️ Semester '{semester_key}' not found in course data. Available: {list(semester_course_maps.keys())}"
                )
                continue
                
            # Check if there are files for this semester
            semester_files_exist = False
            for rf in raw_files:
                detected_sem, _, _, _, _, _ = detect_semester_from_filename(rf)
                # FIX: Compare in uppercase for case-insensitive matching
                if detected_sem.upper() == semester_key.upper():
                    semester_files_exist = True
                    break
                    
            if semester_files_exist:
                print(f"\n🎯 Processing {semester_key} in {nd_set}...")
                try:
                    # Process the semester with the upgrade threshold
                    result = process_semester_files(
                        semester_key,
                        raw_files,
                        raw_dir,
                        set_output_dir,
                        ts,
                        params["pass_threshold"],
                        semester_course_maps,
                        semester_credit_units,
                        semester_lookup,
                        semester_course_titles,
                        DEFAULT_LOGO_PATH,
                        nd_set,
                        previous_cgpas=None,
                        upgrade_min_threshold=upgrade_min_threshold,
                    )
                    if result is not None:
                        print(f"✅ Successfully processed {semester_key}")
                        total_processed += 1
                    else:
                        print(f"❌ Failed to process {semester_key}")
                except Exception as e:
                    print(f"❌ Error processing {semester_key}: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"⚠️ No files found for {semester_key} in {nd_set}, skipping...")
        # Create CGPA_SUMMARY and ANALYSIS worksheets
        mastersheet_path = os.path.join(set_output_dir, f"mastersheet_{ts}.xlsx")
        if os.path.exists(mastersheet_path):
            print(f"📊 Creating CGPA_SUMMARY and ANALYSIS worksheets...")
            
            # FIX 2: Identify inactive students FIRST before creating sheets
            print(f"🔍 Identifying inactive students...")
            identify_inactive_students()
            
            # THEN create the sheets that depend on this data
            create_cgpa_summary_sheet(mastersheet_path, ts)
            create_analysis_sheet(mastersheet_path, ts)  # Now INACTIVE_STUDENTS will be populated
            
            print(f"✅ Successfully added all worksheets")
        # Create ZIP of the entire set results
        try:
            zip_path = os.path.join(clean_dir, f"{nd_set}_RESULT-{ts}.zip")
            zip_success = create_zip_folder(set_output_dir, zip_path)
            if zip_success:
                # Verify the ZIP file was created and has content
                if os.path.exists(zip_path):
                    zip_size = os.path.getsize(zip_path)
                    print(f"✅ ZIP file created: {zip_path} ({zip_size} bytes)")
                    # Convert bytes to MB for readability
                    zip_size_mb = zip_size / (1024 * 1024)
                    print(f"📦 ZIP file size: {zip_size_mb:.2f} MB")
                else:
                    print(f"❌ ZIP file was not created: {zip_path}")
            else:
                print(f"❌ Failed to create ZIP file for {nd_set}")
        except Exception as e:
            print(f"⚠️ Failed to create ZIP for {nd_set}: {e}")
            import traceback
            traceback.print_exc()
    print(f"\n📊 PROCESSING SUMMARY: {total_processed} semester(s) processed")
    
    # Print carryover summary
    if CARRYOVER_STUDENTS:
        print(f"\n📋 CARRYOVER SUMMARY:")
        print(f" Total carryover students: {len(CARRYOVER_STUDENTS)}")
        # Count by semester
        semester_counts = {}
        for student_key, data in CARRYOVER_STUDENTS.items():
            semester = data["semester"]
            semester_counts[semester] = semester_counts.get(semester, 0) + 1
        for semester, count in semester_counts.items():
            print(f" {semester}: {count} students")
            
    # Print inactive students summary
    if INACTIVE_STUDENTS:
        print(f"\n📋 INACTIVE STUDENTS SUMMARY:")
        print(f" Total inactive students: {len(INACTIVE_STUDENTS)}")
        cgpa_active_count = sum(1 for s in INACTIVE_STUDENTS.values() if s.get("cgpa_status") == "Active in CGPA")
        regular_inactive_count = len(INACTIVE_STUDENTS) - cgpa_active_count
        print(f" - Regular inactive (missing intermediate semesters): {regular_inactive_count}")
        print(f" - CGPA active but missing from recent semester: {cgpa_active_count}")
            
    return total_processed > 0

if __name__ == "__main__":
    try:
        main()
        print("✅ ND Examination Results Processing completed successfully")
    except Exception as e:
        print(f"❌ Error during processing: {e}")
        import traceback
        traceback.print_exc()