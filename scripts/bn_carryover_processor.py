#!/usr/bin/env python3

"""
BN CARRYOVER PROCESSOR - Complete Basic Nursing Carryover Processor
FIXED VERSION for Web Interface Integration

Based on ND Carryover Processor with BN-specific adaptations
Handles Basic Nursing (BN) carryover/resit result processing with:
- 6 semesters (3 years × 2 semesters)
- N- prefix for semester keys
- Proper GPA/CGPA recalculation
- Mastersheet updates with all enhancements
- SINGLE workbook session to prevent corruption
- WEB INTERFACE COMPATIBLE
"""

import os
import sys
import re
import pandas as pd
import numpy as np
from datetime import datetime
import glob
import json
import traceback
import shutil
import zipfile
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------
# CRITICAL FIX: Configuration
# ----------------------------

def get_base_directory():
    """Get base directory - FIXED for web interface"""
    # Priority 1: Environment variable
    if os.getenv("BASE_DIR"):
        base_dir = os.getenv("BASE_DIR")
        if os.path.exists(base_dir):
            print(f"✅ Using BASE_DIR from environment: {base_dir}")
            return base_dir
    
    # Priority 2: Standard locations
    possible_dirs = [
        os.path.join(os.path.expanduser("~"), "student_result_cleaner"),
        os.path.join(os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"),
        os.getcwd(),
        os.path.dirname(os.path.abspath(__file__)),
    ]
    
    for dir_path in possible_dirs:
        if os.path.exists(os.path.join(dir_path, "EXAMS_INTERNAL")):
            print(f"✅ Found EXAMS_INTERNAL in: {dir_path}")
            return dir_path
    
    # Fallback
    default_dir = os.path.join(os.path.expanduser("~"), "student_result_cleaner")
    print(f"⚠️ Using default directory: {default_dir}")
    return default_dir

BASE_DIR = get_base_directory()
TIMESTAMP_FMT = "%d-%m-%Y_%H%M%S"
DEFAULT_PASS_THRESHOLD = 50.0
DEFAULT_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")
print(f"🔧 BASE_DIR set to: {BASE_DIR}")

# ----------------------------
# Utility Functions (BN-Compatible)
# ----------------------------

def sanitize_filename(filename):
    """Remove or replace characters that are not safe for filenames."""
    return re.sub(r"[^\w\-_.]", "_", filename)

def find_exam_number_column(df):
    """Find the exam number column in a DataFrame."""
    possible_names = [
        "EXAM NUMBER",
        "REG. No",
        "REG NO",
        "REGISTRATION NUMBER",
        "MAT NO",
        "STUDENT ID",
    ]
    for col in df.columns:
        col_upper = str(col).upper()
        for possible_name in possible_names:
            if possible_name in col_upper:
                return col
    return None

def get_grade_point(score):
    """Determine grade point based on score - NIGERIAN 5.0 SCALE."""
    try:
        score = float(score)
        if score >= 70:
            return 5.0
        elif score >= 60:
            return 4.0
        elif score >= 50:
            return 3.0
        elif score >= 45:
            return 2.0
        elif score >= 40:
            return 1.0
        else:
            return 0.0
    except (ValueError, TypeError):
        return 0.0

def generate_remarks(resit_courses):
    """Generate remarks for resit performance."""
    passed_count = sum(
        1
        for course_data in resit_courses.values()
        if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
    )
    total_count = len(resit_courses)
    if passed_count == total_count:
        return "All courses passed in resit"
    elif passed_count > 0:
        return f"{passed_count}/{total_count} courses passed in resit"
    else:
        return "No improvement in resit"

def extract_class_from_set_name(set_name):
    """Extract class name from set_name (e.g., 'SET47' from 'SET47')"""
    return set_name

# ----------------------------
# BN-Specific Semester Functions
# ----------------------------

def standardize_semester_key(semester_key):
    """Standardize semester key to canonical format for BN."""
    if not semester_key:
        return None
    
    key_upper = semester_key.upper()
    
    # BN canonical mappings (N- prefix)
    canonical_mappings = {
        # Year 1
        ("FIRST", "YEAR", "FIRST", "SEMESTER"): "N-FIRST-YEAR-FIRST-SEMESTER",
        ("1ST", "YEAR", "1ST", "SEMESTER"): "N-FIRST-YEAR-FIRST-SEMESTER",
        ("YEAR", "1", "SEMESTER", "1"): "N-FIRST-YEAR-FIRST-SEMESTER",
        
        ("FIRST", "YEAR", "SECOND", "SEMESTER"): "N-FIRST-YEAR-SECOND-SEMESTER",
        ("1ST", "YEAR", "2ND", "SEMESTER"): "N-FIRST-YEAR-SECOND-SEMESTER",
        ("YEAR", "1", "SEMESTER", "2"): "N-FIRST-YEAR-SECOND-SEMESTER",
        
        # Year 2
        ("SECOND", "YEAR", "FIRST", "SEMESTER"): "N-SECOND-YEAR-FIRST-SEMESTER",
        ("2ND", "YEAR", "1ST", "SEMESTER"): "N-SECOND-YEAR-FIRST-SEMESTER",
        ("YEAR", "2", "SEMESTER", "1"): "N-SECOND-YEAR-FIRST-SEMESTER",
        
        ("SECOND", "YEAR", "SECOND", "SEMESTER"): "N-SECOND-YEAR-SECOND-SEMESTER",
        ("2ND", "YEAR", "2ND", "SEMESTER"): "N-SECOND-YEAR-SECOND-SEMESTER",
        ("YEAR", "2", "SEMESTER", "2"): "N-SECOND-YEAR-SECOND-SEMESTER",
        
        # Year 3
        ("THIRD", "YEAR", "FIRST", "SEMESTER"): "N-THIRD-YEAR-FIRST-SEMESTER",
        ("3RD", "YEAR", "1ST", "SEMESTER"): "N-THIRD-YEAR-FIRST-SEMESTER",
        ("YEAR", "3", "SEMESTER", "1"): "N-THIRD-YEAR-FIRST-SEMESTER",
        
        ("THIRD", "YEAR", "SECOND", "SEMESTER"): "N-THIRD-YEAR-SECOND-SEMESTER",
        ("3RD", "YEAR", "2ND", "SEMESTER"): "N-THIRD-YEAR-SECOND-SEMESTER",
        ("YEAR", "3", "SEMESTER", "2"): "N-THIRD-YEAR-SECOND-SEMESTER",
    }
    
    # BN patterns
    patterns = [
        r"(FIRST|1ST|YEAR.?1).*?(FIRST|1ST|SEMESTER.?1)",
        r"(FIRST|1ST|YEAR.?1).*?(SECOND|2ND|SEMESTER.?2)",
        r"(SECOND|2ND|YEAR.?2).*?(FIRST|1ST|SEMESTER.?1)",
        r"(SECOND|2ND|YEAR.?2).*?(SECOND|2ND|SEMESTER.?2)",
        r"(THIRD|3RD|YEAR.?3).*?(FIRST|1ST|SEMESTER.?1)",
        r"(THIRD|3RD|YEAR.?3).*?(SECOND|2ND|SEMESTER.?2)",
    ]
    
    for pattern_idx, pattern in enumerate(patterns):
        if re.search(pattern, key_upper):
            semester_map = [
                "N-FIRST-YEAR-FIRST-SEMESTER",
                "N-FIRST-YEAR-SECOND-SEMESTER",
                "N-SECOND-YEAR-FIRST-SEMESTER",
                "N-SECOND-YEAR-SECOND-SEMESTER",
                "N-THIRD-YEAR-FIRST-SEMESTER",
                "N-THIRD-YEAR-SECOND-SEMESTER",
            ]
            if pattern_idx < len(semester_map):
                return semester_map[pattern_idx]
    
    print(f"Could not standardize BN semester key: {semester_key}")
    return semester_key

def standardize_semester_name(semester_name):
    """Standardize semester name - alias for standardize_semester_key for compatibility."""
    return standardize_semester_key(semester_name)

def extract_semester_from_filename(filename):
    """Extract semester from carryover filename - BN VERSION"""
    try:
        # Handle both .json and .xlsx files
        if filename.endswith(".json") or filename.endswith(".xlsx"):
            # Pattern: co_student_BN-SET47_N-SECOND-YEAR-FIRST-SEMESTER_20251107_100522.json
            # Extract the part between the second 'N-' and the timestamp
            pattern = r"co_student_BN-SET\d+_(N-.*?)_\d+_\d+\.(json|xlsx)"
            match = re.search(pattern, filename)
            if match:
                semester = match.group(1)
                return semester.upper().replace("-", " ").replace("_", " ")
        # Fallback: try to extract any N- pattern
        match = re.search(
            r"(N-[A-Za-z-]+(?:YEAR|SEMESTER)[A-Za-z-]*)", filename, re.IGNORECASE
        )
        if match:
            return match.group(1).upper().replace("-", " ").replace("_", " ")
        return None
    except Exception as e:
        print(f"Error extracting semester from {filename}: {e}")
        return None

def get_semester_display_info(semester_key):
    """Get display information for BN semester key."""
    semester_lower = semester_key.lower()
    
    semester_info = {
        "first-year-first-semester": (1, 1, "YEAR ONE", "FIRST SEMESTER", "BNI", "Semester 1"),
        "first-year-second-semester": (1, 2, "YEAR ONE", "SECOND SEMESTER", "BNI", "Semester 2"),
        "second-year-first-semester": (2, 1, "YEAR TWO", "FIRST SEMESTER", "BNII", "Semester 3"),
        "second-year-second-semester": (2, 2, "YEAR TWO", "SECOND SEMESTER", "BNII", "Semester 4"),
        "third-year-first-semester": (3, 1, "YEAR THREE", "FIRST SEMESTER", "BNIII", "Semester 5"),
        "third-year-second-semester": (3, 2, "YEAR THREE", "SECOND SEMESTER", "BNIII", "Semester 6"),
    }
    
    for key, info in semester_info.items():
        if key in semester_lower:
            return info
    
    return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BNI", "Semester 1"

def get_previous_semester(semester_key):
    """Get the previous semester key for BN carryover."""
    standardized = standardize_semester_key(semester_key)
    
    # BN semesters chain
    semester_chain = {
        "N-FIRST-YEAR-SECOND-SEMESTER": "N-FIRST-YEAR-FIRST-SEMESTER",
        "N-SECOND-YEAR-FIRST-SEMESTER": "N-FIRST-YEAR-SECOND-SEMESTER",
        "N-SECOND-YEAR-SECOND-SEMESTER": "N-SECOND-YEAR-FIRST-SEMESTER",
        "N-THIRD-YEAR-FIRST-SEMESTER": "N-SECOND-YEAR-SECOND-SEMESTER",
        "N-THIRD-YEAR-SECOND-SEMESTER": "N-THIRD-YEAR-FIRST-SEMESTER",
    }
    
    return semester_chain.get(standardized)

def get_previous_semesters_for_display(current_semester_key):
    """Get list of previous semesters for BN GPA display in mastersheet."""
    current_standard = standardize_semester_key(current_semester_key)
    
    semester_mapping = {
        "N-FIRST-YEAR-FIRST-SEMESTER": [],
        "N-FIRST-YEAR-SECOND-SEMESTER": ["Semester 1"],
        "N-SECOND-YEAR-FIRST-SEMESTER": ["Semester 1", "Semester 2"],
        "N-SECOND-YEAR-SECOND-SEMESTER": ["Semester 1", "Semester 2", "Semester 3"],
        "N-THIRD-YEAR-FIRST-SEMESTER": ["Semester 1", "Semester 2", "Semester 3", "Semester 4"],
        "N-THIRD-YEAR-SECOND-SEMESTER": ["Semester 1", "Semester 2", "Semester 3", "Semester 4", "Semester 5"],
    }
    
    return semester_mapping.get(current_standard, [])

# ----------------------------
# BN Course Data Management
# ----------------------------

def load_course_data():
    """Load BN course data ONLY."""
    return load_bn_course_data()

def load_bn_course_data():
    """Load BN course data from course-code-creditUnit.xlsx."""
    possible_course_files = [
        os.path.join(
            BASE_DIR,
            "EXAMS_INTERNAL",
            "BN",
            "BN-COURSES",
            "course-code-creditUnit.xlsx",
        ),
        os.path.join(BASE_DIR, "BN", "BN-COURSES", "course-code-creditUnit.xlsx"),
        os.path.join(
            BASE_DIR, "EXAMS_INTERNAL", "BN-COURSES", "course-code-creditUnit.xlsx"
        ),
        os.path.join(BASE_DIR, "course-code-creditUnit.xlsx"),
    ]
    course_file = None
    for possible_file in possible_course_files:
        if os.path.exists(possible_file):
            course_file = possible_file
            print(f"✅ Found BN course file: {possible_file}")
            break
    if not course_file:
        print(f"❌ Main BN course file not found in standard locations")
        alternative_files = find_alternative_bn_course_files()
        if alternative_files:
            course_file = alternative_files[0]
            print(f"🔄 Using alternative BN course file: {course_file}")
        else:
            print(f"❌ No BN course files found anywhere!")
            return {}, {}, {}, {}
    print(f"📚 Loading BN course data from: {course_file}")
    return _load_course_data_from_file_bn(course_file)

def _load_course_data_from_file_bn(course_file):
    """Generic function to load BN course data from Excel file."""
    try:
        xl = pd.ExcelFile(course_file)
        semester_course_titles = {}
        semester_credit_units = {}
        course_code_to_title = {}
        course_code_to_unit = {}
        print(f"📖 Available sheets: {xl.sheet_names}")
        for sheet in xl.sheet_names:
            sheet_standard = standardize_semester_key(sheet)
            print(f"📖 Reading sheet: {sheet} (standardized: {sheet_standard})")
            try:
                df = pd.read_excel(
                    course_file, sheet_name=sheet, engine="openpyxl", header=0
                )
                # Convert columns to string and clean
                df.columns = [str(c).strip().upper() for c in df.columns]
                # Look for course code, title, and credit unit columns with flexible matching
                code_col = None
                title_col = None
                unit_col = None
                for col in df.columns:
                    col_clean = str(col).upper()
                    if any(
                        keyword in col_clean
                        for keyword in ["COURSE CODE", "CODE", "COURSECODE"]
                    ):
                        code_col = col
                    elif any(
                        keyword in col_clean
                        for keyword in ["COURSE TITLE", "TITLE", "COURSENAME"]
                    ):
                        title_col = col
                    elif any(
                        keyword in col_clean
                        for keyword in ["CU", "CREDIT", "UNIT", "CREDIT UNIT"]
                    ):
                        unit_col = col
                print(
                    f"🔍 Detected columns - Code: {code_col}, Title: {title_col}, Unit: {unit_col}"
                )
                if not all([code_col, title_col, unit_col]):
                    print(
                        f"⚠️ Sheet '{sheet}' missing required columns - found: code={code_col}, title={title_col}, unit={unit_col}"
                    )
                    # Try to use first three columns as fallback
                    if len(df.columns) >= 3:
                        code_col, title_col, unit_col = (
                            df.columns[0],
                            df.columns[1],
                            df.columns[2],
                        )
                        print(
                            f"🔄 Using fallback columns: {code_col}, {title_col}, {unit_col}"
                        )
                    else:
                        print(
                            f"❌ Sheet '{sheet}' doesn't have enough columns - skipped"
                        )
                        continue
                # Clean the data
                df_clean = df.dropna(subset=[code_col]).copy()
                if df_clean.empty:
                    print(f"⚠️ Sheet '{sheet}' has no data after cleaning - skipped")
                    continue
                # Convert credit units to numeric, handling errors
                df_clean[unit_col] = pd.to_numeric(df_clean[unit_col], errors="coerce")
                df_clean = df_clean.dropna(subset=[unit_col])
                # Remove rows with "TOTAL" in course code
                df_clean = df_clean[
                    ~df_clean[code_col]
                    .astype(str)
                    .str.contains("TOTAL", case=False, na=False)
                ]
                if df_clean.empty:
                    print(
                        f"⚠️ Sheet '{sheet}' has no valid rows after cleaning - skipped"
                    )
                    continue
                codes = df_clean[code_col].astype(str).str.strip().tolist()
                titles = df_clean[title_col].astype(str).str.strip().tolist()
                units = df_clean[unit_col].astype(float).tolist()
                print(f"📋 Found {len(codes)} courses in {sheet}:")
                for i, (code, title, unit) in enumerate(
                    zip(codes[:5], titles[:5], units[:5])
                ):
                    print(f" - '{code}': '{title}' (CU: {unit})")
                # Create mapping dictionaries with ENHANCED normalization strategies
                sheet_titles = {}
                sheet_units = {}
                for code, title, unit in zip(codes, titles, units):
                    if not code or code.upper() in ["NAN", "NONE", ""]:
                        continue
                    # ENHANCED: Create comprehensive normalization variants for robust matching
                    variants = [
                        # Basic variants
                        code.upper().strip(),
                        code.strip(),
                        code.upper(),
                        code.lower(),
                        code.title(),
                        # Space removal variants
                        code.upper().replace(" ", ""),
                        code.replace(" ", ""),
                        re.sub(r"\s+", "", code.upper()),
                        re.sub(r"\s+", "", code),
                        # Special character removal (keep only alphanumeric)
                        re.sub(r"[^a-zA-Z0-9]", "", code.upper()),
                        re.sub(r"[^a-zA-Z0-9]", "", code),
                        # Dash and underscore variants
                        code.upper().replace("-", ""),
                        code.upper().replace("_", ""),
                        code.replace("-", "").replace("_", ""),
                        code.upper().replace("-", "").replace("_", "").replace(" ", ""),
                        # WITH common prefixes (for matching with prefix)
                        f"NUR{code.upper()}",
                        f"NUR{code.upper().replace(' ', '')}",
                        f"NUR{re.sub(r'[^a-zA-Z0-9]', '', code.upper())}",
                        f"NSC{code.upper()}",
                        f"NSC{code.upper().replace(' ', '')}",
                        f"NSC{re.sub(r'[^a-zA-Z0-9]', '', code.upper())}",
                        # WITHOUT common prefixes (for matching without prefix)
                        code.upper().replace("NUR", "").strip(),
                        code.upper().replace("NSC", "").strip(),
                        re.sub(r"^(NUR|NSC)", "", code.upper()).strip(),
                        re.sub(r"^(NUR|NSC)", "", code.upper())
                        .replace(" ", "")
                        .strip(),
                        # Number-focused variants (for codes like "101", "201")
                        re.sub(r"[^0-9]", "", code),
                        # Common variations with dots
                        code.upper().replace(".", ""),
                        code.replace(".", ""),
                    ]
                    # Remove duplicates while preserving order
                    variants = list(
                        dict.fromkeys(
                            [v for v in variants if v and v not in ["NAN", "NONE", ""]]
                        )
                    )
                    # Add all variants to mappings
                    for variant in variants:
                        sheet_titles[variant] = title
                        sheet_units[variant] = unit
                        course_code_to_title[variant] = title
                        course_code_to_unit[variant] = unit
                semester_course_titles[sheet_standard] = sheet_titles
                semester_credit_units[sheet_standard] = sheet_units
            except Exception as e:
                print(f"❌ Error processing sheet '{sheet}': {e}")
                traceback.print_exc()
                continue
        print(
            f"✅ Loaded course data for sheets: {list(semester_course_titles.keys())}"
        )
        print(f"📊 Total course mappings: {len(course_code_to_title)}")
        # Debug: Show some course mappings
        print("🔍 Sample course mappings:")
        sample_items = list(course_code_to_title.items())[:15]
        for code, title in sample_items:
            unit = course_code_to_unit.get(code, 0)
            print(f" '{code}' -> '{title}' (CU: {unit})")
        return (
            semester_course_titles,
            semester_credit_units,
            course_code_to_title,
            course_code_to_unit,
        )
    except Exception as e:
        print(f"❌ Error loading BN course data: {e}")
        traceback.print_exc()
        return {}, {}, {}, {}

def find_alternative_bn_course_files():
    """Look for alternative BN course files."""
    base_dirs = [
        os.path.join(BASE_DIR, "EXAMS_INTERNAL", "BN", "BN-COURSES"),
        os.path.join(BASE_DIR, "BN", "BN-COURSES"),
        os.path.join(BASE_DIR, "COURSES"),
        os.path.join(BASE_DIR, "EXAMS_INTERNAL"),
    ]
    course_files = []
    for base_dir in base_dirs:
        if os.path.exists(base_dir):
            for file in os.listdir(base_dir):
                if "course" in file.lower() and file.endswith((".xlsx", ".xls")):
                    full_path = os.path.join(base_dir, file)
                    course_files.append(full_path)
    return course_files

def debug_course_matching_bn(resit_file_path, course_code_to_title, course_code_to_unit):
    """Debug function to check why BN course codes aren't matching."""
    print(f"\n🔍 DEBUGGING BN COURSE MATCHING")
    print("=" * 50)
    # Read resit file to see what course codes we have
    resit_df = pd.read_excel(resit_file_path, header=0)
    resit_exam_col = find_exam_number_column(resit_df)
    # Get all course codes from resit file
    resit_courses = []
    for col in resit_df.columns:
        if col != resit_exam_col and col != "NAME" and not "Unnamed" in str(col):
            resit_courses.append(col)
    print(f"📋 BN Course codes from resit file: {resit_courses}")
    print(f"📊 Total courses in BN resit file: {len(resit_courses)}")
    # Check each resit course against course file
    for course in resit_courses:
        print(f"\n🔍 Checking BN course: '{course}'")
        original_code = str(course).strip()
        # Generate ENHANCED variants for matching
        variants = [
            original_code.upper().strip(),
            original_code.strip(),
            original_code.upper(),
            original_code,
            original_code.lower(),
            original_code.title(),
            original_code.upper().replace(" ", ""),
            original_code.replace(" ", ""),
            re.sub(r"\s+", "", original_code.upper()),
            re.sub(r"\s+", "", original_code),
            re.sub(r"[^a-zA-Z0-9]", "", original_code.upper()),
            re.sub(r"[^a-zA-Z0-9]", "", original_code),
            original_code.upper().replace("-", ""),
            original_code.upper().replace("_", ""),
            original_code.replace("-", "").replace("_", ""),
            original_code.upper().replace("-", "").replace("_", "").replace(" ", ""),
            f"NUR{original_code.upper()}",
            f"NUR{original_code.upper().replace(' ', '')}",
            f"NSC{original_code.upper()}",
            f"NSC{original_code.upper().replace(' ', '')}",
            original_code.upper().replace("NUR", "").strip(),
            original_code.upper().replace("NSC", "").strip(),
            re.sub(r"^(NUR|NSC)", "", original_code.upper()).strip(),
            re.sub(r"^(NUR|NSC)", "", original_code.upper()).replace(" ", "").strip(),
            re.sub(r"[^0-9]", "", original_code),
            original_code.upper().replace(".", ""),
            original_code.replace(".", ""),
        ]
        # Remove duplicates
        variants = list(dict.fromkeys([v for v in variants if v and v != "NAN"]))
        print(f" Generated {len(variants)} variants to try")
        found = False
        for variant in variants:
            if variant in course_code_to_title:
                title = course_code_to_title[variant]
                unit = course_code_to_unit.get(variant, 0)
                print(f" ✅ FOUND: '{variant}' -> '{title}' (CU: {unit})")
                found = True
                break
        if not found:
            print(f" ❌ NOT FOUND: No match for '{course}'")
            # Show some similar keys from course file
            similar_keys = []
            # Check for partial matches
            for key in list(course_code_to_title.keys())[:50]:  # Check first 50 keys
                if any(
                    part in key.upper()
                    for part in original_code.upper().split()
                    if len(part) > 2
                ):
                    similar_keys.append(key)
            if similar_keys:
                print(f" 💡 Similar keys found: {similar_keys[:5]}")
            else:
                print(
                    f" 💡 Sample available keys: {list(course_code_to_title.keys())[:10]}"
                )

def find_course_title(course_code, course_titles_dict, course_code_to_title):
    """Robust function to find course title with comprehensive matching strategies."""
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return str(course_code) if course_code else "Unknown Course"
    original_code = str(course_code).strip()
    # Generate ENHANCED comprehensive matching variants
    variants = [
        # Basic normalizations
        original_code.upper().strip(),
        original_code.strip(),
        original_code.upper(),
        original_code,
        original_code.lower(),
        original_code.title(),
        # Space handling variations
        original_code.upper().replace(" ", ""),
        original_code.replace(" ", ""),
        re.sub(r"\s+", "", original_code.upper()),
        re.sub(r"\s+", "", original_code),
        # Special character handling
        re.sub(r"[^a-zA-Z0-9]", "", original_code.upper()),
        re.sub(r"[^a-zA-Z0-9]", "", original_code),
        # Common formatting issues
        original_code.upper().replace("-", ""),
        original_code.upper().replace("_", ""),
        original_code.replace("-", "").replace("_", ""),
        original_code.upper().replace("-", "").replace("_", "").replace(" ", ""),
        # WITH common prefixes
        f"NUR{original_code.upper()}",
        f"NUR{original_code.upper().replace(' ', '')}",
        f"NUR{re.sub(r'[^a-zA-Z0-9]', '', original_code.upper())}",
        f"NSC{original_code.upper()}",
        f"NSC{original_code.upper().replace(' ', '')}",
        f"NSC{re.sub(r'[^a-zA-Z0-9]', '', original_code.upper())}",
        # WITHOUT common prefixes
        original_code.upper().replace("NUR", "").strip(),
        original_code.upper().replace("NSC", "").strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).replace(" ", "").strip(),
        # Number-focused variants
        re.sub(r"[^0-9]", "", original_code),
        # Dot removal
        original_code.upper().replace(".", ""),
        original_code.replace(".", ""),
    ]
    # Remove duplicates
    variants = list(dict.fromkeys([v for v in variants if v and v != "NAN"]))
    # Try each strategy in order
    for variant in variants:
        # Try course_titles_dict first (semester-specific)
        if variant in course_titles_dict:
            title = course_titles_dict[variant]
            print(
                f"✅ Found title for '{original_code}' using variant '{variant}': '{title}'"
            )
            return title
        # Try global course_code_to_title
        if variant in course_code_to_title:
            title = course_code_to_title[variant]
            print(
                f"✅ Found title for '{original_code}' using global variant '{variant}': '{title}'"
            )
            return title
    # If no match found, log and return descriptive original code
    print(f"⚠️ Could not find course title for: '{original_code}'")
    print(f" Tried {len(variants)} variants without success")
    return f"{original_code} (Title Not Found)"

def find_credit_unit(course_code, credit_units_dict, course_code_to_unit):
    """Robust function to find credit unit with comprehensive matching strategies."""
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return 0
    original_code = str(course_code).strip()
    # Generate the same ENHANCED variants as title matching
    variants = [
        original_code.upper().strip(),
        original_code.strip(),
        original_code.upper(),
        original_code,
        original_code.lower(),
        original_code.title(),
        original_code.upper().replace(" ", ""),
        original_code.replace(" ", ""),
        re.sub(r"\s+", "", original_code.upper()),
        re.sub(r"\s+", "", original_code),
        re.sub(r"[^a-zA-Z0-9]", "", original_code.upper()),
        re.sub(r"[^a-zA-Z0-9]", "", original_code),
        original_code.upper().replace("-", ""),
        original_code.upper().replace("_", ""),
        original_code.replace("-", "").replace("_", ""),
        original_code.upper().replace("-", "").replace("_", "").replace(" ", ""),
        f"NUR{original_code.upper()}",
        f"NUR{original_code.upper().replace(' ', '')}",
        f"NSC{original_code.upper()}",
        f"NSC{original_code.upper().replace(' ', '')}",
        original_code.upper().replace("NUR", "").strip(),
        original_code.upper().replace("NSC", "").strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).replace(" ", "").strip(),
        re.sub(r"[^0-9]", "", original_code),
        original_code.upper().replace(".", ""),
        original_code.replace(".", ""),
    ]
    # Remove duplicates
    variants = list(dict.fromkeys([v for v in variants if v and v != "NAN"]))
    # Try each strategy
    for variant in variants:
        if variant in credit_units_dict:
            unit = credit_units_dict[variant]
            return unit
        if variant in course_code_to_unit:
            unit = course_code_to_unit[variant]
            return unit
    print(f"⚠️ Could not find credit unit for: '{original_code}', defaulting to 2")
    return 2  # Default credit unit

# ----------------------------
# CRITICAL FIXES: Mastersheet Reading Functions
# ----------------------------

def read_mastersheet_with_flexible_headers(mastersheet_path, sheet_name):
    """FIXED VERSION: Read mastersheet with flexible header detection for BN"""
    print(f"🔍 FIXED: Reading BN mastersheet with flexible headers...")
    
    # First, let's examine the actual structure
    xl = pd.ExcelFile(mastersheet_path)
    df_raw = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=None)
    
    print(f"📊 Raw BN mastersheet shape: {df_raw.shape}")
    print(f"📊 First 10 rows sample:")
    for i in range(min(10, len(df_raw))):
        print(f"  Row {i}: {df_raw.iloc[i].dropna().tolist()}")
    
    # Look for the header row that contains "EXAM NUMBER" or similar
    header_row_idx = None
    for idx in range(len(df_raw)):
        row_values = df_raw.iloc[idx].dropna().astype(str).str.upper().tolist()
        row_combined = " ".join(row_values)
        
        # Check for exam number indicators
        if any(keyword in row_combined for keyword in ["EXAM NUMBER", "EXAMS NUMBER", "REG NO", "REGISTRATION"]):
            header_row_idx = idx
            print(f"✅ FOUND header row at index {idx}: {row_values}")
            break
    
    if header_row_idx is None:
        print(f"❌ No header row found with exam number indicators")
        # Try common header row positions
        for idx in [5, 4, 3, 2, 1, 0]:
            try:
                df_test = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=idx)
                if len(df_test.columns) > 3:  # Reasonable number of columns
                    header_row_idx = idx
                    print(f"🔄 Using fallback header row: {idx}")
                    break
            except:
                continue
    
    if header_row_idx is None:
        print(f"❌ Could not determine header row")
        return None, None
    
    # Read with the found header row
    try:
        df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=header_row_idx)
        print(f"✅ Successfully read BN mastersheet with header row {header_row_idx}")
        print(f"📊 Columns: {df.columns.tolist()}")
        
        # Find exam number column
        exam_col = None
        for col in df.columns:
            col_str = str(col).upper()
            if any(keyword in col_str for keyword in ["EXAM NUMBER", "EXAMS NUMBER", "REG NO", "REGISTRATION"]):
                exam_col = col
                break
        
        if not exam_col:
            print(f"❌ No exam number column found in: {df.columns.tolist()}")
            return None, None
            
        print(f"✅ Exam column found: '{exam_col}'")
        return df, exam_col
        
    except Exception as e:
        print(f"❌ Error reading BN mastersheet: {e}")
        return None, None

def find_student_in_mastersheet_fixed(exam_no, mastersheet_df, exam_col):
    """FIXED VERSION: Robust student matching in BN mastersheet"""
    if mastersheet_df is None or exam_col not in mastersheet_df.columns:
        return None
        
    # Clean the exam number for matching
    exam_no_clean = str(exam_no).strip().upper()
    exam_no_clean = re.sub(r'[^A-Z0-9]', '', exam_no_clean)
    
    # Method 1: Exact match after cleaning
    for idx, row in mastersheet_df.iterrows():
        current_exam = str(row[exam_col]).strip().upper() if pd.notna(row[exam_col]) else ""
        current_exam_clean = re.sub(r'[^A-Z0-9]', '', current_exam)
        
        if current_exam_clean == exam_no_clean:
            return row
    
    # Method 2: Partial match
    for idx, row in mastersheet_df.iterrows():
        current_exam = str(row[exam_col]).strip().upper() if pd.notna(row[exam_col]) else ""
        current_exam_clean = re.sub(r'[^A-Z0-9]', '', current_exam)
        
        if exam_no_clean in current_exam_clean or current_exam_clean in exam_no_clean:
            return row
    
    # Method 3: Try with different cleaning approaches
    for idx, row in mastersheet_df.iterrows():
        current_exam = str(row[exam_col]).strip().upper() if pd.notna(row[exam_col]) else ""
        
        # Remove common prefixes/suffixes
        current_clean = re.sub(r'^(BN|NUR|NSC)', '', current_exam)
        exam_clean = re.sub(r'^(BN|NUR|NSC)', '', exam_no_clean)
        
        if current_clean == exam_clean:
            return row
    
    return None

def quick_fix_read_mastersheet(mastersheet_path, sheet_name):
    """QUICK FIX: Read BN mastersheet using header row 5 (which we know works)"""
    try:
        # Force header row 5 which we know contains the correct columns
        df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)
        
        # Manually find exam column
        exam_col = None
        for col in df.columns:
            if "EXAM" in str(col).upper() or "REG" in str(col).upper():
                exam_col = col
                break
                
        if exam_col:
            print(f"✅ QUICK FIX: Using forced header row 5, exam column: '{exam_col}'")
            return df, exam_col
        else:
            print(f"❌ QUICK FIX: No exam column found even with header row 5")
            return None, None
            
    except Exception as e:
        print(f"❌ QUICK FIX Error: {e}")
        return None, None

# ----------------------------
# File and ZIP Handling (BN-Compatible)
# ----------------------------

def extract_mastersheet_from_zip(zip_path, semester_key):
    """Extract mastersheet from ZIP file and return temporary file path."""
    try:
        print(f"📦 Looking for mastersheet in ZIP: {zip_path}")
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            all_files = zip_ref.namelist()
            print(f"📁 Files in ZIP: {all_files}")
            mastersheet_files = [
                f
                for f in all_files
                if "mastersheet" in f.lower() and f.endswith(".xlsx")
            ]
            if not mastersheet_files:
                print(f"❌ No mastersheet found in ZIP")
                return None, None
            mastersheet_name = mastersheet_files[0]
            print(f"✅ Found mastersheet: {mastersheet_name}")
            temp_dir = tempfile.mkdtemp()
            temp_mastersheet_path = os.path.join(
                temp_dir, f"mastersheet_{semester_key}.xlsx"
            )
            with open(temp_mastersheet_path, "wb") as f:
                f.write(zip_ref.read(mastersheet_name))
            print(f"✅ Extracted mastersheet to: {temp_mastersheet_path}")
            return temp_mastersheet_path, temp_dir
    except Exception as e:
        print(f"❌ Error extracting mastersheet from ZIP: {e}")
        traceback.print_exc()
        return None, None

def find_latest_zip_file(clean_dir):
    """Find the latest ZIP file in clean results directory."""
    print(f"🔍 Looking for BN ZIP files in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ BN clean directory doesn't exist: {clean_dir}")
        return None
    all_files = os.listdir(clean_dir)
    zip_files = []
    for f in all_files:
        if f.lower().endswith(".zip"):
            if "carryover" in f.lower():
                print(f"⚠️ Skipping BN carryover ZIP: {f}")
                continue
            if any(pattern in f for pattern in ["_RESULT-", "RESULT_", "RESULT-"]):
                zip_files.append(f)
                print(f"✅ Found BN regular results ZIP: {f}")
            else:
                print(f"ℹ️ Found other BN ZIP (not a result file): {f}")
    if not zip_files:
        print(f"❌ No BN regular results ZIP files found (excluding carryover files)")
        fallback_zips = [
            f
            for f in all_files
            if f.lower().endswith(".zip") and "carryover" not in f.lower()
        ]
        if fallback_zips:
            print(f"⚠️ Using fallback BN ZIP files: {fallback_zips}")
            zip_files = fallback_zips
        else:
            print(f"❌ No BN ZIP files found at all in {clean_dir}")
            return None
    print(f"✅ Final BN ZIP files to consider: {zip_files}")
    zip_files_with_path = [os.path.join(clean_dir, f) for f in zip_files]
    latest_zip = sorted(zip_files_with_path, key=os.path.getmtime, reverse=True)[0]
    print(f"🎯 Using latest BN ZIP: {latest_zip}")
    return latest_zip

def find_latest_result_folder(clean_dir, set_name):
    """Find the latest result folder in clean results directory."""
    print(f"🔍 Looking for BN result folders in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ BN clean directory doesn't exist: {clean_dir}")
        return None
    all_items = os.listdir(clean_dir)
    result_folders = [
        f
        for f in all_items
        if os.path.isdir(os.path.join(clean_dir, f))
        and f.startswith(f"{set_name}_RESULT-")
    ]
    if not result_folders:
        print(f"❌ No BN result folders found")
        return None
    print(f"✅ Found BN result folders: {result_folders}")
    folders_with_path = [os.path.join(clean_dir, f) for f in result_folders]
    latest_folder = sorted(folders_with_path, key=os.path.getmtime, reverse=True)[0]
    print(f"🎯 Using latest BN result folder: {latest_folder}")
    return latest_folder

def find_latest_mastersheet_source(clean_dir, set_name):
    """Find the latest source for mastersheet: prefer ZIP, fallback to folder."""
    print(f"🔍 Looking for BN mastersheet source in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ BN clean directory doesn't exist: {clean_dir}")
        return None, None
    zip_path = find_latest_zip_file(clean_dir)
    if zip_path:
        print(f"✅ Using BN ZIP source: {zip_path}")
        try:
            with zipfile.ZipFile(zip_path, "r") as zip_ref:
                zip_files = zip_ref.namelist()
                mastersheet_files = [
                    f
                    for f in zip_files
                    if "mastersheet" in f.lower() and f.endswith(".xlsx")
                ]
                if mastersheet_files:
                    print(f"✅ BN ZIP contains mastersheet files: {mastersheet_files}")
                    return zip_path, "zip"
                else:
                    print(f"⚠️ BN ZIP found but no mastersheet inside: {zip_path}")
        except Exception as e:
            print(f"⚠️ Error checking BN ZIP contents: {e}")
    folder_path = find_latest_result_folder(clean_dir, set_name)
    if folder_path:
        print(f"✅ Using BN folder source: {folder_path}")
        return folder_path, "folder"
    print(f"❌ No valid BN ZIP files or result folders found in {clean_dir}")
    return None, None

def get_mastersheet_path(source_path, source_type, semester_key):
    """Get mastersheet path based on source type (zip or folder)."""
    temp_dir = None
    if source_type == "zip":
        temp_mastersheet_path, temp_dir = extract_mastersheet_from_zip(
            source_path, semester_key
        )
        if not temp_mastersheet_path:
            print("❌ Failed to extract mastersheet from ZIP")
            return None, None
    elif source_type == "folder":
        all_files = os.listdir(source_path)
        mastersheet_files = [
            f for f in all_files if "mastersheet" in f.lower() and f.endswith(".xlsx")
        ]
        if not mastersheet_files:
            print(f"❌ No mastersheet found in folder {source_path}")
            return None, None
        mastersheet_name = mastersheet_files[0]
        temp_mastersheet_path = os.path.join(source_path, mastersheet_name)
        print(f"✅ Found mastersheet in folder: {temp_mastersheet_path}")
    else:
        return None, None
    return temp_mastersheet_path, temp_dir

def get_matching_sheet(xl, target_key):
    """Find matching sheet name with variants."""
    target_upper = (
        target_key.upper().replace("-", " ").replace("_", " ").replace(".", " ")
    )
    target_upper = " ".join(target_upper.split())
    possible_keys = [
        target_key,
        target_key.upper(),
        target_key.lower(),
        target_key.title(),
        target_key.replace("-", " ").upper(),
        target_key.replace("-", " ").lower(),
        target_key.replace("-", " ").title(),
        target_key.replace("-", "_").upper(),
        target_key.replace("-", "_").lower(),
        target_key.replace("-", "_").title(),
        target_key.replace("First", "1st"),
        target_key.replace("Second", "2nd"),
        target_key.replace("Third", "3rd"),
        target_key.replace("YEAR", "YR"),
        target_key.replace("SEMESTER", "SEM"),
        target_upper,
        target_upper.replace("FIRST", "1ST"),
        target_upper.replace("SECOND", "2ND"),
        target_upper.replace("THIRD", "3RD"),
        target_upper.replace("YEAR", "YR"),
        target_upper.replace("SEMESTER", "SEM"),
    ]
    possible_keys = list(set([k for k in possible_keys if k]))
    print(f"🔍 Trying sheet variants for '{target_key}': {possible_keys}")
    for sheet in xl.sheet_names:
        sheet_normalized = (
            sheet.upper().replace("-", " ").replace("_", " ").replace(".", " ")
        )
        sheet_normalized = " ".join(sheet_normalized.split())
        if any(
            p == sheet or p in sheet or p == sheet_normalized or p in sheet_normalized
            for p in possible_keys
        ):
            print(f"✅ Found matching sheet: '{sheet}' for '{target_key}'")
            return sheet
    print(f"❌ No matching sheet found for '{target_key}'")
    print(f"📖 Available sheets: {xl.sheet_names}")
    return None

def create_carryover_zip(source_dir, zip_path):
    """Create ZIP file of carryover results."""
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        print(f"✅ ZIP file created: {zip_path}")
        return True
    except Exception as e:
        print(f"❌ Error creating ZIP: {e}")
        return False

# ----------------------------
# GPA/CGPA Management (BN-Compatible)
# ----------------------------

def load_previous_gpas(mastersheet_path, current_semester_key):
    """Load previous GPA data from mastersheet for BN CGPA calculation."""
    all_student_data = {}
    current_standard = standardize_semester_key(current_semester_key)
    
    # BN semesters
    all_semesters = {
        "N-FIRST-YEAR-FIRST-SEMESTER": [],
        "N-FIRST-YEAR-SECOND-SEMESTER": ["N-FIRST-YEAR-FIRST-SEMESTER"],
        "N-SECOND-YEAR-FIRST-SEMESTER": [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
        ],
        "N-SECOND-YEAR-SECOND-SEMESTER": [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
        ],
        "N-THIRD-YEAR-FIRST-SEMESTER": [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
            "N-SECOND-YEAR-SECOND-SEMESTER",
        ],
        "N-THIRD-YEAR-SECOND-SEMESTER": [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
            "N-SECOND-YEAR-SECOND-SEMESTER",
            "N-THIRD-YEAR-FIRST-SEMESTER",
        ],
    }
    
    semesters_to_load = all_semesters.get(current_standard, [])
    print(f"📊 Loading previous BN GPAs for {current_standard}: {semesters_to_load}")
    
    if not os.path.exists(mastersheet_path):
        print(f"❌ BN Mastersheet not found: {mastersheet_path}")
        return {}
    
    try:
        xl = pd.ExcelFile(mastersheet_path)
        print(f"📖 Available sheets in BN mastersheet: {xl.sheet_names}")
    except Exception as e:
        print(f"❌ Error opening BN mastersheet: {e}")
        return {}
    
    for semester in semesters_to_load:
        try:
            sheet_name = get_matching_sheet(xl, semester)
            if not sheet_name:
                print(f"⚠️ Skipping BN semester {semester} - no matching sheet found")
                continue
            print(f"📖 Reading BN sheet '{sheet_name}' for semester {semester}")
            df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)
            if df.empty or len(df.columns) < 3:
                df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=0)
                print(f"🔄 Using header row 0 for BN sheet '{sheet_name}'")
            exam_col = find_exam_number_column(df)
            gpa_col = None
            credit_col = None
            # ENHANCED: Prioritize total attempted credits for CGPA accuracy
            for col in df.columns:
                col_str = str(col).upper()
                if "GPA" in col_str and "CGPA" not in col_str:
                    gpa_col = col
                if (
                    "TCPE" in col_str
                    or "TOTAL CREDIT" in col_str
                    or "TOTAL UNIT" in col_str
                ):
                    credit_col = col  # Prioritize total attempted credits
                elif "CU PASSED" in col_str or "CREDIT" in col_str or "UNIT" in col_str:
                    credit_col = col  # Fallback
            print(
                f"🔍 Columns found - Exam: {exam_col}, GPA: {gpa_col}, Credits: {credit_col}"
            )
            if exam_col and gpa_col:
                for idx, row in df.iterrows():
                    try:
                        exam_no = str(row[exam_col]).strip()
                        if pd.isna(exam_no) or exam_no in ["", "NAN", "NONE"]:
                            continue
                        gpa_value = row[gpa_col]
                        if pd.isna(gpa_value):
                            continue
                        credits = 30
                        if (
                            credit_col
                            and credit_col in row
                            and pd.notna(row[credit_col])
                        ):
                            try:
                                credits = int(float(row[credit_col]))
                            except (ValueError, TypeError):
                                credits = 30
                        if exam_no not in all_student_data:
                            all_student_data[exam_no] = {"gpas": [], "credits": []}
                        all_student_data[exam_no]["gpas"].append(float(gpa_value))
                        all_student_data[exam_no]["credits"].append(credits)
                        if idx < 3:
                            print(
                                f"📊 Loaded BN GPA for {exam_no}: {gpa_value} with {credits} credits"
                            )
                    except (ValueError, TypeError) as e:
                        print(f"⚠️ Error processing row {idx} for BN {semester}: {e}")
                        continue
            else:
                print(
                    f"⚠️ Missing required columns in BN {sheet_name}: exam_col={exam_col}, gpa_col={gpa_col}"
                )
        except Exception as e:
            print(f"⚠️ Could not load data from BN {semester}: {e}")
            traceback.print_exc()
    
    print(f"📊 Loaded cumulative BN data for {len(all_student_data)} students")
    return all_student_data

def calculate_cgpa(student_data, current_gpa, current_credits):
    """Calculate Cumulative GPA for BN."""
    if not student_data or not student_data.get("gpas"):
        print(f"⚠️ No previous BN GPA data, using current GPA: {current_gpa}")
        return current_gpa
    
    total_grade_points = 0.0
    total_credits = 0
    print(f"🔢 Calculating BN CGPA from {len(student_data['gpas'])} previous semesters")
    
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
        print(
            f" - GPA: {prev_gpa}, Credits: {prev_credits}, Running Total: {total_grade_points}/{total_credits}"
        )
    
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits
    print(f"📊 Final BN calculation: {total_grade_points} / {total_credits}")
    
    if total_credits > 0:
        cgpa = round(total_grade_points / total_credits, 2)
        print(f"✅ Calculated BN CGPA: {cgpa}")
        return cgpa
    else:
        print(f"⚠️ No BN credits, returning current GPA: {current_gpa}")
        return current_gpa

# ----------------------------
# Mastersheet Update Functions (CRITICAL FIXES - BN-Compatible)
# ----------------------------

def find_sheet_structure(ws):
    """Find the header row and build headers dictionary"""
    header_row = None
    headers = {}
    # Look for header row (contains 'EXAM NUMBER')
    for row_idx in range(1, 30):  # Check first 30 rows
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and "EXAM NUMBER" in str(cell_value).upper():
                header_row = row_idx
                # Build headers dictionary
                for col in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=header_row, column=col).value
                    if header_val:
                        headers[str(header_val).strip()] = col
                print(f"✅ Found header row at: {header_row}")
                return header_row, headers
    print(f"❌ Could not find header row")
    return None, {}

def apply_student_sorting(ws, header_row, headers_dict):
    """Apply sorting to students - compatibility function for ANALYSIS sheet"""
    apply_student_sorting_with_serial_numbers(ws, header_row, headers_dict)

def apply_student_sorting_with_serial_numbers(ws, header_row, headers_dict):
    """Apply sorting to students with PROPER serial numbers - BN VERSION"""
    from openpyxl.styles import Font, PatternFill
    exam_col = headers_dict.get("EXAM NUMBER")
    remarks_col = headers_dict.get("REMARKS")
    gpa_col = headers_dict.get("GPA")
    serial_col = 1  # Serial number is always column 1
    if not all([exam_col, remarks_col, gpa_col]):
        return
    # Collect all student rows
    student_rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        exam_no = ws.cell(row, exam_col).value
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break
        # Get basic row data (values only, not styles)
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            row_data.append({"value": cell.value, "number_format": cell.number_format})
        remarks = ws.cell(row, remarks_col).value or ""
        remarks_upper = str(remarks).upper()
        # Assign priority for sorting
        if "WITHDRAW" in remarks_upper:
            priority = 3  # Withdrawn - last
        elif (
            "RESIT" in remarks_upper
            or "CARRYOVER" in remarks_upper
            or "PROBATION" in remarks_upper
        ):
            priority = 2  # Carryover - middle
        else:
            priority = 1  # Passed - first
        student_rows.append(
            {
                "row_num": row,
                "priority": priority,
                "gpa": ws.cell(row, gpa_col).value or 0,
                "row_data": row_data,
                "remarks": remarks_upper,
                "exam_no": exam_no,  # Store exam number for secondary sorting
            }
        )
    # Sort by priority (Passed first, then Carryover, then Withdrawn)
    # Within each group, sort by GPA (descending), then by exam number
    student_rows.sort(
        key=lambda x: (x["priority"], -float(x["gpa"] if x["gpa"] else 0), x["exam_no"])
    )
    # Write sorted data back to worksheet
    for idx, student in enumerate(student_rows):
        target_row = header_row + 1 + idx
        row_data = student["row_data"]
        for col_idx, cell_data in enumerate(row_data, 1):
            cell = ws.cell(target_row, col_idx)
            cell.value = cell_data["value"]
            # Apply number format if present
            if cell_data.get("number_format"):
                cell.number_format = cell_data["number_format"]
            # CRITICAL FIX: Update serial numbers to be consecutive
            if col_idx == serial_col:  # Serial number column
                cell.value = idx + 1
            # Re-apply styling based on content and student status
            if student["priority"] == 3:  # Withdrawn
                cell.fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )
                if "WITHDRAW" in student["remarks"]:
                    cell.font = Font(bold=True, color="FF0000")
            elif idx % 2 == 0:  # Alternate row coloring
                cell.fill = PatternFill(
                    start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
                )
            # Apply GPA styling for GPA column
            if col_idx == headers_dict.get("GPA"):
                try:
                    gpa_val = float(cell.value) if cell.value else 0
                    if gpa_val >= 3.5:
                        cell.font = Font(
                            bold=True, color="006400"
                        )  # Dark green for high GPA
                    elif gpa_val < 2.0:
                        cell.font = Font(bold=True, color="FF0000")  # Red for low GPA
                except (ValueError, TypeError):
                    pass
    
    print(
        f" ✅ Applied student sorting with proper serial numbers (1 to {len(student_rows)})"
    )

def identify_course_columns_properly(headers):
    """Identify course columns using PROPER pattern matching"""
    import re
    course_columns = {}
    for header, col_idx in headers.items():
        # Match patterns like: NUR101, NSC201, etc. (3 letters + 3 digits)
        # Also match GNS111, etc.
        if re.match(r"^[A-Z]{3}\d{3}$", str(header).strip()):
            course_columns[header] = col_idx
            print(f"✅ Identified course column: '{header}' at index {col_idx}")
    print(f"📊 Total course columns identified: {len(course_columns)}")
    return course_columns

def update_summary_section_fixed(ws, headers, header_row, course_columns):
    """Update SUMMARY section - FIXED VERSION with embedded number updates and aligned fails"""
    print(f" 📊 Updating summary section...")
    try:
        # Find SUMMARY section
        summary_start_row = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "SUMMARY" in str(cell_value).upper():
                summary_start_row = row_idx
                break
        if not summary_start_row:
            print(" ℹ️ No SUMMARY section found")
            return
        # Find exam column
        exam_col_idx = None
        for col_name, col_idx in headers.items():
            if "EXAM NUMBER" in col_name.upper():
                exam_col_idx = col_idx
                break
        if not exam_col_idx:
            print(" ❌ No exam column found")
            return
        # CRITICAL FIX: Re-read headers to ensure we have fresh column mappings
        fresh_headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                fresh_headers[str(header_val).strip()] = col_idx
        # CRITICAL FIX: Re-identify course columns from fresh headers
        import re
        fresh_course_columns = {}
        for header, col_idx in fresh_headers.items():
            if re.match(r"^[A-Z]{3}\d{3}$", str(header).strip()):
                fresh_course_columns[header] = col_idx
        # Count statistics from CURRENT data (reading directly from worksheet)
        total_students = 0
        passed_students = 0
        resit_students = 0
        probation_students = 0
        withdrawn_students = 0
        course_failures = {course: 0 for course in fresh_course_columns}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no = ws.cell(row=row_idx, column=exam_col_idx).value
            # Stop at summary
            if not exam_no or "SUMMARY" in str(exam_no).upper():
                break
            if str(exam_no).strip() in ["", "NAN", "NONE"]:
                continue
            total_students += 1
            # Count remarks
            for col_name, col_idx in headers.items():
                if "REMARKS" in col_name.upper():
                    remarks = ws.cell(row=row_idx, column=col_idx).value
                    if remarks:
                        remarks_upper = str(remarks).upper()
                        if "PASSED" in remarks_upper:
                            passed_students += 1
                        elif "RESIT" in remarks_upper or "CARRYOVER" in remarks_upper:
                            resit_students += 1
                        elif "PROBATION" in remarks_upper:
                            probation_students += 1
                        elif "WITHDRAW" in remarks_upper:
                            withdrawn_students += 1
                    break
            # Count ACTUAL failures from CURRENT scores (INCLUDING updated resit scores)
            for course in fresh_course_columns:
                if course in fresh_headers:
                    col_idx = fresh_headers[course]
                    score = ws.cell(row=row_idx, column=col_idx).value
                    if score is not None and score != "":
                        try:
                            if float(score) < 50:
                                course_failures[course] += 1
                        except (ValueError, TypeError):
                            continue
        # Update fails per course (aligned to course columns)
        fails_row = None
        # Search for the fails per course row in the first three columns
        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col_idx in [1, 2, 3]:
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and "FAILS PER COURSE" in str(cell_value).upper():
                    fails_row = row_idx
                    break
            if fails_row:
                break
        if fails_row:
            print(f"✅ Found fails per course row at row {fails_row}")
            sorted_courses = sorted(
                fresh_course_columns, key=lambda k: fresh_course_columns[k]
            )
            for i, course in enumerate(sorted_courses):
                col = fresh_course_columns[course]
                ws.cell(row=fails_row, column=col).value = course_failures[course]
                print(
                    f" ✅ Updated {course} failures: {course_failures[course]} at column {col}"
                )
        else:
            print("❌ Could not find fails per course row")
        # Update summary rows with embedded numbers
        current_row = summary_start_row + 1
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            if not cell_value:
                break
            cell_str = str(cell_value).upper()
            if "REGISTERED AND SAT" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {total_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated total students: {total_students}")
            elif (
                "PASSED IN ALL COURSES REGISTERED" in cell_str
                and "FAILED" not in cell_str
            ):
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {passed_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated passed: {passed_students}")
            elif "GRADE POINT AVERAGE (GPA) OF 2.00 AND ABOVE FAILED" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {resit_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated resit (GPA >=2.00): {resit_students}")
            elif "GRADE POINT AVERAGE (GPA) BELOW 2.00 FAILED" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {probation_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated probation: {probation_students}")
            elif "FAILED IN MORE THAN 45%" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {withdrawn_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated withdrawn: {withdrawn_students}")
            current_row += 1
        print(f" ✅ Summary section updated")
    except Exception as e:
        print(f" ❌ Error updating summary: {e}")
        traceback.print_exc()

def ensure_required_sheets_exist(wb):
    """Ensure CGPA_SUMMARY and ANALYSIS sheets exist in workbook"""
    from openpyxl.styles import Font, Alignment, PatternFill
    print(f"\n🔍 CHECKING FOR REQUIRED SHEETS...")
    print(f" Current sheets: {wb.sheetnames}")
    # Create CGPA_SUMMARY if it doesn't exist
    if "CGPA_SUMMARY" not in wb.sheetnames:
        print(f"⚠️ CGPA_SUMMARY sheet not found - creating it...")
        cgpa_ws = wb.create_sheet("CGPA_SUMMARY")
        # Add basic structure
        cgpa_ws["A1"] = "CGPA SUMMARY"
        cgpa_ws["A1"].font = Font(bold=True, size=14)
        cgpa_ws["A1"].alignment = Alignment(horizontal="center")
        print(f"✅ Created CGPA_SUMMARY sheet")
    else:
        print(f"✅ CGPA_SUMMARY sheet exists")
    # Create ANALYSIS if it doesn't exist
    if "ANALYSIS" not in wb.sheetnames:
        print(f"⚠️ ANALYSIS sheet not found - creating it...")
        analysis_ws = wb.create_sheet("ANALYSIS")
        # Add basic structure
        analysis_ws["A1"] = "PERFORMANCE ANALYSIS"
        analysis_ws["A1"].font = Font(bold=True, size=14)
        analysis_ws["A1"].alignment = Alignment(horizontal="center")
        print(f"✅ Created ANALYSIS sheet")
    else:
        print(f"✅ ANALYSIS sheet exists")
    print(f"✅ All required sheets verified")
    return True

def update_cgpa_summary_sheet_fixed(wb, semester_key, header_row, set_name):
    """Update CGPA_SUMMARY sheet - BN VERSION with 6 semesters"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    print(f" 📈 Updating BN CGPA_SUMMARY...")
    if "CGPA_SUMMARY" not in wb.sheetnames:
        print(" ℹ️ No CGPA_SUMMARY sheet")
        return
    cgpa_ws = wb["CGPA_SUMMARY"]
    # FIXED: Properly handle merged cells by unmerging first
    try:
        merged_ranges = list(cgpa_ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            cgpa_ws.unmerge_cells(str(merged_range))
        print(f" ✅ Unmerged {len(merged_ranges)} cell ranges")
    except Exception as e:
        print(f" ⚠️ Could not unmerge cells: {e}")
    # Clear ALL old data (including headers)
    for row in range(1, cgpa_ws.max_row + 1):
        for col in range(1, cgpa_ws.max_column + 1):
            cgpa_ws.cell(row, col).value = None
            cgpa_ws.cell(row, col).fill = PatternFill()  # Clear formatting
            cgpa_ws.cell(row, col).font = Font()
            cgpa_ws.cell(row, col).border = Border()
    # ===================================================================
    # STEP 1: CREATE PROFESSIONAL HEADER SECTION
    # ===================================================================
    class_name = f"BN {set_name}"
    # Get semester info
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )
    # Calculate total columns needed (11 columns for BN: 6 semesters + basic info)
    total_columns = 11
    last_column = get_column_letter(total_columns)
    # Row 1: Institution Name (Merged and Centered)
    cgpa_ws.merge_cells(f"A1:{last_column}1")
    title_cell = cgpa_ws["A1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 2: Department (Merged and Centered)
    cgpa_ws.merge_cells(f"A2:{last_column}2")
    dept_cell = cgpa_ws["A2"]
    dept_cell.value = "DEPARTMENT OF NURSING"
    dept_cell.font = Font(bold=True, size=12, name="Calibri")
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 3: Class and Sheet Title (Merged and Centered)
    cgpa_ws.merge_cells(f"A3:{last_column}3")
    class_cell = cgpa_ws["A3"]
    class_cell.value = f"{class_name} CLASS - CUMULATIVE GPA SUMMARY"
    class_cell.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
    class_cell.fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    class_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 4: Date and Academic Session (Merged and Centered)
    cgpa_ws.merge_cells(f"A4:{last_column}4")
    date_cell = cgpa_ws["A4"]
    current_year = datetime.now().year
    date_cell.value = f"{current_year}/{current_year + 1} Academic Session - Generated on {datetime.now().strftime('%B %d, %Y')}"
    date_cell.font = Font(size=11, name="Calibri", italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 5: Empty spacer
    cgpa_ws.row_dimensions[5].height = 5
    # ===================================================================
    # STEP 2: CREATE COLUMN HEADERS FOR BN (6 SEMESTERS)
    # ===================================================================
    # Row 6: Column Headers
    headers = [
        "S/N",
        "EXAM NUMBER",
        "NAME",
        "Semester 1",
        "Semester 2", 
        "Semester 3",
        "Semester 4",
        "Semester 5",
        "Semester 6",
        "CGPA",
        "WITHDRAWN",
    ]
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        cell = cgpa_ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    # Set row height for header
    cgpa_ws.row_dimensions[6].height = 25
    # ===================================================================
    # STEP 3: COLLECT AND POPULATE DATA WITH PROPER WITHDRAWN TRACKING
    # ===================================================================
    # Collect data from all BN semesters using SINGLE workbook session
    semester_keys = [
        "N-FIRST-YEAR-FIRST-SEMESTER",
        "N-FIRST-YEAR-SECOND-SEMESTER",
        "N-SECOND-YEAR-FIRST-SEMESTER",
        "N-SECOND-YEAR-SECOND-SEMESTER",
        "N-THIRD-YEAR-FIRST-SEMESTER",
        "N-THIRD-YEAR-SECOND-SEMESTER",
    ]
    semester_data = {}
    # Track ALL historically withdrawn students across ALL semesters - CRITICAL FIX
    all_withdrawn_students = set()
    # FIRST PASS: Identify ALL historically withdrawn students from ALL semesters
    print(f" 🔍 FIRST PASS: Identifying ALL historically withdrawn students...")
    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue
            exam_col = headers_dict.get("EXAM NUMBER")
            remarks_col = headers_dict.get("REMARKS")
            if not all([exam_col, remarks_col]):
                continue
            # Collect withdrawn students from this semester
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break
                exam_no_clean = str(exam_no).strip().upper()
                remarks = ws.cell(row, remarks_col).value or ""
                remarks_upper = str(remarks).upper()
                # CRITICAL FIX: Track ALL students who were EVER withdrawn
                if "WITHDRAW" in remarks_upper:
                    all_withdrawn_students.add(exam_no_clean)
                    print(
                        f" 📝 Found historically withdrawn: {exam_no_clean} in {sheet_name}"
                    )
    print(
        f" ✅ Identified {len(all_withdrawn_students)} historically withdrawn students"
    )
    # SECOND PASS: Collect semester data with PERSISTENT withdrawn status
    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue
            exam_col = headers_dict.get("EXAM NUMBER")
            name_col = headers_dict.get("NAME")
            gpa_col = headers_dict.get("GPA")
            credits_col = headers_dict.get("TCPE")
            remarks_col = headers_dict.get("REMARKS")
            if not all([exam_col, name_col, gpa_col]):
                continue
            data = {}
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break
                exam_no = str(exam_no).strip().upper()
                name = ws.cell(row, name_col).value
                gpa_val = ws.cell(row, gpa_col).value
                credits_val = ws.cell(row, credits_col).value if credits_col else 0
                remarks = ws.cell(row, remarks_col).value if remarks_col else ""
                # CRITICAL FIX: Check if student is historically withdrawn (PERSISTENT STATUS)
                is_withdrawn = exam_no in all_withdrawn_students
                try:
                    gpa = float(gpa_val) if gpa_val else 0
                    credits = float(credits_val) if credits_val else 0
                    data[exam_no] = {
                        "name": name,
                        "gpa": gpa,
                        "credits": credits,
                        "remarks": remarks,
                        "withdrawn": is_withdrawn,  # Store withdrawn status
                    }
                except (ValueError, TypeError):
                    continue
            semester_data[key] = data
    # Collect unique students
    all_exam_no = set()
    for d in semester_data.values():
        all_exam_no.update(d.keys())
    students = []
    for exam_no in all_exam_no:
        total_gp = 0.0
        total_cr = 0.0
        gpas = {}
        name = None
        # CRITICAL FIX: Check if student is withdrawn in ANY semester
        withdrawn = False
        for key, d in semester_data.items():
            if exam_no in d:
                data = d[exam_no]
                gpas[key] = data["gpa"]
                total_gp += data["gpa"] * data["credits"]
                total_cr += data["credits"]
                # If student is withdrawn in ANY semester, mark as withdrawn
                if data["withdrawn"]:
                    withdrawn = True
                    print(f" 🔒 Student {exam_no} marked as withdrawn (found in {key})")
                if not name:
                    name = data["name"]
        cgpa = round(total_gp / total_cr, 2) if total_cr > 0 else 0.0
        students.append(
            {
                "exam_no": exam_no,
                "name": name,
                "gpas": gpas,
                "cgpa": cgpa,
                "withdrawn": withdrawn,  # PERSISTENT withdrawn status
            }
        )
    # Sort: non-withdrawn by CGPA (descending), then withdrawn by CGPA (descending)
    non_withdrawn = [s for s in students if not s["withdrawn"]]
    withdrawn_list = [s for s in students if s["withdrawn"]]
    # CRITICAL FIX: Sort by CGPA descending, then by exam number for consistency
    non_withdrawn.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    withdrawn_list.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    sorted_students = non_withdrawn + withdrawn_list
    # ===================================================================
    # STEP 4: WRITE STUDENT DATA WITH FORMATTING AND PROPER SERIAL NUMBERS
    # ===================================================================
    row = 7  # Start from row 7 (after header)
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    # Alternate row colors
    even_row_fill = PatternFill(
        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
    )
    withdrawn_fill = PatternFill(
        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
    )
    # CRITICAL FIX: Proper serial numbers from 1 to n
    serial_number = 1
    for idx, s in enumerate(sorted_students):
        # Determine row fill
        if s["withdrawn"]:
            row_fill = withdrawn_fill
        elif idx % 2 == 0:
            row_fill = even_row_fill
        else:
            row_fill = PatternFill()  # White
        # Serial Number (PROPERLY SORTED from 1 to n)
        cell = cgpa_ws.cell(row, 1, value=serial_number)
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        serial_number += 1
        # Exam Number
        cell = cgpa_ws.cell(row, 2, value=s["exam_no"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Name
        cell = cgpa_ws.cell(row, 3, value=s["name"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # Semester GPAs (6 semesters for BN)
        for col_idx, sem_key in enumerate(
            [
                "N-FIRST-YEAR-FIRST-SEMESTER",
                "N-FIRST-YEAR-SECOND-SEMESTER",
                "N-SECOND-YEAR-FIRST-SEMESTER",
                "N-SECOND-YEAR-SECOND-SEMESTER",
                "N-THIRD-YEAR-FIRST-SEMESTER",
                "N-THIRD-YEAR-SECOND-SEMESTER",
            ],
            4,
        ):
            gpa_value = s["gpas"].get(sem_key, "")
            cell = cgpa_ws.cell(row, col_idx, value=gpa_value)
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if gpa_value:
                cell.number_format = "0.00"
        # CGPA
        cell = cgpa_ws.cell(row, 10, value=s["cgpa"])  # Column 10 for BN
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"
        cell.font = Font(bold=True)
        # Withdrawn status - CRITICAL FIX: Show "Yes" for withdrawn students
        withdrawn_status = "Yes" if s["withdrawn"] else "No"
        cell = cgpa_ws.cell(row, 11, value=withdrawn_status)  # Column 11 for BN
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if s["withdrawn"]:
            cell.font = Font(bold=True, color="FF0000")
            print(f" ✅ Marked {s['exam_no']} as WITHDRAWN in CGPA_SUMMARY")
        row += 1
    # ===================================================================
    # STEP 5: ADD SUMMARY STATISTICS
    # ===================================================================
    summary_start_row = row + 2
    total_students = len(students)
    avg_cgpa = (
        round(sum(s["cgpa"] for s in students) / total_students, 2)
        if total_students > 0
        else 0
    )
    highest_cgpa = max(s["cgpa"] for s in students) if students else 0
    lowest_cgpa = min(s["cgpa"] for s in students if s["cgpa"] > 0) if students else 0
    withdrawn_count = len(withdrawn_list)
    # Summary header
    cgpa_ws.merge_cells(f"A{summary_start_row}:{last_column}{summary_start_row}")
    summary_header = cgpa_ws.cell(summary_start_row, 1)
    summary_header.value = "SUMMARY STATISTICS"
    summary_header.font = Font(bold=True, size=12, name="Calibri")
    summary_header.fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    summary_header.alignment = Alignment(horizontal="center", vertical="center")
    summary_header.border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="thin"),
    )
    # Summary data
    summary_data = [
        ("Total Students:", total_students),
        ("Average Cumulative CGPA:", avg_cgpa),
        ("Highest Cumulative CGPA:", highest_cgpa),
        ("Lowest Cumulative CGPA:", lowest_cgpa),
        ("Withdrawn Students:", withdrawn_count),
    ]
    summary_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    for i, (label, value) in enumerate(summary_data):
        current_row = summary_start_row + 1 + i
        # Label (columns A-D merged)
        cgpa_ws.merge_cells(f"A{current_row}:D{current_row}")
        label_cell = cgpa_ws.cell(current_row, 1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=11, name="Calibri")
        label_cell.fill = summary_fill
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.border = Border(
            left=Side(style="medium"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        # Value (columns E-K merged)
        cgpa_ws.merge_cells(f"E{current_row}:{last_column}{current_row}")
        value_cell = cgpa_ws.cell(current_row, 5)
        value_cell.value = value
        value_cell.font = Font(bold=True, size=11, name="Calibri")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="medium"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        if isinstance(value, float):
            value_cell.number_format = "0.00"
    # Bottom border for summary section
    last_summary_row = summary_start_row + len(summary_data)
    for col in range(1, total_columns + 1):
        cell = cgpa_ws.cell(last_summary_row, col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=Side(style="medium"),
        )
    # ===================================================================
    # STEP 6: ADJUST COLUMN WIDTHS FOR BN
    # ===================================================================
    cgpa_ws.column_dimensions["A"].width = 8  # S/N
    cgpa_ws.column_dimensions["B"].width = 18  # EXAM NUMBER
    cgpa_ws.column_dimensions["C"].width = 35  # NAME
    cgpa_ws.column_dimensions["D"].width = 12  # Semester 1
    cgpa_ws.column_dimensions["E"].width = 12  # Semester 2
    cgpa_ws.column_dimensions["F"].width = 12  # Semester 3
    cgpa_ws.column_dimensions["G"].width = 12  # Semester 4
    cgpa_ws.column_dimensions["H"].width = 12  # Semester 5
    cgpa_ws.column_dimensions["I"].width = 12  # Semester 6
    cgpa_ws.column_dimensions["J"].width = 12  # CGPA
    cgpa_ws.column_dimensions["K"].width = 12  # WITHDRAWN
    # Set specific row heights
    cgpa_ws.row_dimensions[1].height = 20  # Title
    cgpa_ws.row_dimensions[2].height = 18  # Department
    cgpa_ws.row_dimensions[3].height = 20  # Class/Title
    cgpa_ws.row_dimensions[4].height = 16  # Date
    print(
        f" ✅ BN CGPA_SUMMARY updated with {len(students)} students, {withdrawn_count} withdrawn, and proper serial numbers"
    )
    print(
        f" ✅ Withdrawn students properly marked: {[s['exam_no'] for s in withdrawn_list]}"
    )

def update_cgpa_summary_with_withdrawn(wb, withdrawn_students):
    """Update CGPA SUMMARY sheet with PERSISTENT withdrawn status and apply sorting"""
    if "CGPA SUMMARY" not in wb.sheetnames:
        return
    cgpa_ws = wb["CGPA SUMMARY"]
    header_row_found, headers_dict = find_sheet_structure(cgpa_ws)
    if not header_row_found:
        return
    exam_col = headers_dict.get("EXAM NUMBER")
    remarks_col = headers_dict.get("REMARKS")
    cgpa_col = headers_dict.get("CGPA")
    withdrawn_col = None
    # Find withdrawn column if it exists
    for header, col in headers_dict.items():
        if "WITHDRAWN" in header.upper():
            withdrawn_col = col
            break
    if not all([exam_col, remarks_col, cgpa_col]):
        return
    # Update remarks AND withdrawn column for withdrawn students
    for row in range(header_row_found + 1, cgpa_ws.max_row + 1):
        exam_no = cgpa_ws.cell(row, exam_col).value
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break
        exam_no_clean = str(exam_no).strip()
        if exam_no_clean in withdrawn_students:
            # Update remarks to ensure withdrawn status
            cgpa_ws.cell(row, remarks_col).value = "WITHDRAWN"
            # Update withdrawn column if it exists
            if withdrawn_col:
                cgpa_ws.cell(row, withdrawn_col).value = "Yes"
            print(f" 🔒 Maintaining withdrawn status in CGPA: {exam_no_clean}")
    # Apply sorting to CGPA SUMMARY
    apply_student_sorting(cgpa_ws, header_row_found, headers_dict)
    print(f" ✅ CGPA SUMMARY updated with PERSISTENT withdrawn status")

def update_analysis_sheet_fixed(
    wb, semester_key, course_columns, headers, header_row, set_name
):
    """Update ANALYSIS sheet - BN VERSION with 6 semesters"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    print(f" 📊 Updating BN ANALYSIS with PERSISTENT withdrawn tracking...")
    if "ANALYSIS" not in wb.sheetnames:
        print(" ℹ️ No ANALYSIS sheet")
        return
    analysis_ws = wb["ANALYSIS"]
    # FIXED: Properly handle merged cells by unmerging first
    try:
        merged_ranges = list(analysis_ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            analysis_ws.unmerge_cells(str(merged_range))
        print(f" ✅ Unmerged {len(merged_ranges)} cell ranges in ANALYSIS sheet")
    except Exception as e:
        print(f" ⚠️ Could not unmerge cells in ANALYSIS sheet: {e}")
    # Clear ALL old data (including headers)
    for row in range(1, analysis_ws.max_row + 1):
        for col in range(1, analysis_ws.max_column + 1):
            try:
                cell = analysis_ws.cell(row, col)
                if cell.value is not None:
                    cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()
                cell.border = Border()
            except AttributeError:
                continue
    # ===================================================================
    # STEP 1: CREATE PROFESSIONAL HEADER SECTION
    # ===================================================================
    class_name = f"BN {set_name}"
    # Get semester info
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )
    # Calculate total columns needed (7 columns for our data)
    total_columns = 7
    last_column = get_column_letter(total_columns)
    # Row 1: Institution Name
    analysis_ws.merge_cells(f"A1:{last_column}1")
    title_cell = analysis_ws["A1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 2: Department
    analysis_ws.merge_cells(f"A2:{last_column}2")
    dept_cell = analysis_ws["A2"]
    dept_cell.value = "DEPARTMENT OF NURSING"
    dept_cell.font = Font(bold=True, size=12, name="Calibri")
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 3: Class and Sheet Title
    analysis_ws.merge_cells(f"A3:{last_column}3")
    class_cell = analysis_ws["A3"]
    class_cell.value = f"{class_name} CLASS - PERFORMANCE ANALYSIS"
    class_cell.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
    class_cell.fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    class_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 4: Date and Academic Session
    analysis_ws.merge_cells(f"A4:{last_column}4")
    date_cell = analysis_ws["A4"]
    current_year = datetime.now().year
    date_cell.value = f"{current_year}/{current_year + 1} Academic Session - Generated on {datetime.now().strftime('%B %d, %Y')}"
    date_cell.font = Font(size=11, name="Calibri", italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 5: Empty spacer
    analysis_ws.row_dimensions[5].height = 5
    # ===================================================================
    # STEP 2: CREATE COLUMN HEADERS
    # ===================================================================
    # Row 6: Column Headers
    headers_list = [
        "SEMESTER",
        "TOTAL",
        "PASSED",
        "CARRYOVER",
        "WITHDRAWN",
        "AVG GPA",
        "PASS RATE (%)",
    ]
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers_list, 1):
        cell = analysis_ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    analysis_ws.row_dimensions[6].height = 25
    # ===================================================================
    # STEP 3: ENHANCED DATA COLLECTION WITH PERSISTENT WITHDRAWN TRACKING
    # ===================================================================
    semester_keys = [
        "N-FIRST-YEAR-FIRST-SEMESTER",
        "N-FIRST-YEAR-SECOND-SEMESTER",
        "N-SECOND-YEAR-FIRST-SEMESTER",
        "N-SECOND-YEAR-SECOND-SEMESTER",
        "N-THIRD-YEAR-FIRST-SEMESTER",
        "N-THIRD-YEAR-SECOND-SEMESTER",
    ]
    semester_display_names = {
        "N-FIRST-YEAR-FIRST-SEMESTER": "Year 1 - Semester 1",
        "N-FIRST-YEAR-SECOND-SEMESTER": "Year 1 - Semester 2",
        "N-SECOND-YEAR-FIRST-SEMESTER": "Year 2 - Semester 1",
        "N-SECOND-YEAR-SECOND-SEMESTER": "Year 2 - Semester 2",
        "N-THIRD-YEAR-FIRST-SEMESTER": "Year 3 - Semester 1",
        "N-THIRD-YEAR-SECOND-SEMESTER": "Year 3 - Semester 2",
    }
    semester_stats = {}
    overall_total = 0
    overall_passed = 0
    overall_carryover = 0
    overall_withdrawn = 0
    overall_gpa_sum = 0
    overall_students_for_gpa = 0
    # Track withdrawn students across ALL semesters - CRITICAL FIX
    all_withdrawn_students = set()
    # FIRST PASS: Identify ALL historically withdrawn students from ALL semesters
    print(f" 🔍 FIRST PASS: Identifying ALL historically withdrawn students...")
    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue
            exam_col = headers_dict.get("EXAM NUMBER")
            remarks_col = headers_dict.get("REMARKS")
            if not all([exam_col, remarks_col]):
                continue
            # Collect withdrawn students from this semester
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break
                exam_no_clean = str(exam_no).strip()
                remarks = ws.cell(row, remarks_col).value or ""
                remarks_upper = str(remarks).upper()
                # CRITICAL FIX: Track ALL students who were EVER withdrawn
                if "WITHDRAW" in remarks_upper:
                    all_withdrawn_students.add(exam_no_clean)
                    print(
                        f" 📝 Found historically withdrawn: {exam_no_clean} in {sheet_name}"
                    )
    print(
        f" ✅ Identified {len(all_withdrawn_students)} historically withdrawn students"
    )
    # SECOND PASS: Process each semester with PERSISTENT withdrawn status
    print(f" 🔍 SECOND PASS: Processing semesters with persistent withdrawn status...")
    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue
            exam_col = headers_dict.get("EXAM NUMBER")
            gpa_col = headers_dict.get("GPA")
            remarks_col = headers_dict.get("REMARKS")
            if not all([exam_col, gpa_col, remarks_col]):
                continue
            total = 0
            passed = 0
            resit = 0
            probation = 0
            withdrawn = 0
            gpa_sum = 0
            # Track students for this semester
            semester_students = set()
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break
                exam_no_clean = str(exam_no).strip()
                semester_students.add(exam_no_clean)
                total += 1
                remarks = ws.cell(row, remarks_col).value or ""
                remarks_upper = str(remarks).upper()
                gpa_val = ws.cell(row, gpa_col).value
                # CRITICAL FIX: Check if student is historically withdrawn (PERSISTENT STATUS)
                is_withdrawn = exam_no_clean in all_withdrawn_students
                if is_withdrawn:
                    withdrawn += 1
                    # ENSURE withdrawn status is maintained in remarks
                    if "WITHDRAW" not in remarks_upper:
                        ws.cell(row, remarks_col).value = "WITHDRAWN"
                        print(
                            f" 🔒 Maintaining withdrawn status for: {exam_no_clean} in {sheet_name}"
                        )
                elif "PASSED" in remarks_upper:
                    passed += 1
                elif "RESIT" in remarks_upper or "CARRYOVER" in remarks_upper:
                    resit += 1
                elif "PROBATION" in remarks_upper:
                    probation += 1
                try:
                    gpa_sum += float(gpa_val) if gpa_val else 0
                except (ValueError, TypeError):
                    pass
            # Apply sorting to semester sheet (Passed -> Carryover -> Withdrawn)
            apply_student_sorting_with_serial_numbers(
                ws, header_row_found, headers_dict
            )
            avg_gpa = round(gpa_sum / total, 2) if total > 0 else 0
            pass_rate = round(passed / total * 100, 2) if total > 0 else 0
            carryover = resit + probation
            semester_stats[key] = {
                "total": total,
                "passed": passed,
                "carryover": carryover,
                "withdrawn": withdrawn,
                "avg_gpa": avg_gpa,
                "pass_rate": pass_rate,
            }
            overall_total += total
            overall_passed += passed
            overall_carryover += carryover
            overall_withdrawn += withdrawn
            overall_gpa_sum += gpa_sum
            overall_students_for_gpa += total
    # Update CGPA SUMMARY with PERSISTENT withdrawn status
    update_cgpa_summary_with_withdrawn(wb, all_withdrawn_students)
    # ===================================================================
    # STEP 4: WRITE SEMESTER DATA WITH FORMATTING
    # ===================================================================
    row = 7  # Start from row 7 (after header)
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    # Alternate row colors
    even_row_fill = PatternFill(
        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
    )
    for idx, key in enumerate(semester_keys):
        if key in semester_stats:
            stats = semester_stats[key]
            # Determine row fill
            row_fill = even_row_fill if idx % 2 == 0 else PatternFill()
            # Semester name
            cell = analysis_ws.cell(row, 1, value=semester_display_names[key])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(bold=True)
            # Total students
            cell = analysis_ws.cell(row, 2, value=stats["total"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Passed
            cell = analysis_ws.cell(row, 3, value=stats["passed"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Carryover
            cell = analysis_ws.cell(row, 4, value=stats["carryover"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Withdrawn
            cell = analysis_ws.cell(row, 5, value=stats["withdrawn"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Average GPA
            cell = analysis_ws.cell(row, 6, value=stats["avg_gpa"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"
            # Pass Rate
            cell = analysis_ws.cell(row, 7, value=stats["pass_rate"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"
            row += 1
    # ===================================================================
    # STEP 5: ADD OVERALL SUMMARY ROW
    # ===================================================================
    row += 1  # Skip a row
    overall_avg_gpa = (
        round(overall_gpa_sum / overall_students_for_gpa, 2)
        if overall_students_for_gpa > 0
        else 0
    )
    overall_pass_rate = (
        round(overall_passed / overall_total * 100, 2) if overall_total > 0 else 0
    )
    overall_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    overall_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    # OVERALL label
    cell = analysis_ws.cell(row, 1, value="OVERALL")
    cell.border = overall_border
    cell.fill = overall_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(bold=True, size=12)
    # Overall stats
    overall_data = [
        overall_total,
        overall_passed,
        overall_carryover,
        overall_withdrawn,
        overall_avg_gpa,
        overall_pass_rate,
    ]
    for col_idx, value in enumerate(overall_data, 2):
        cell = analysis_ws.cell(row, col_idx, value=value)
        cell.border = overall_border
        cell.fill = overall_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=11)
        if col_idx >= 6:  # GPA and Pass Rate
            cell.number_format = "0.00"
    # ===================================================================
    # STEP 6: ADJUST COLUMN WIDTHS
    # ===================================================================
    analysis_ws.column_dimensions["A"].width = 25  # SEMESTER
    analysis_ws.column_dimensions["B"].width = 12  # TOTAL
    analysis_ws.column_dimensions["C"].width = 12  # PASSED
    analysis_ws.column_dimensions["D"].width = 14  # CARRYOVER
    analysis_ws.column_dimensions["E"].width = 14  # WITHDRAWN
    analysis_ws.column_dimensions["F"].width = 12  # AVG GPA
    analysis_ws.column_dimensions["G"].width = 15  # PASS RATE (%)
    # Set specific row heights
    analysis_ws.row_dimensions[1].height = 20  # Title
    analysis_ws.row_dimensions[2].height = 18  # Department
    analysis_ws.row_dimensions[3].height = 20  # Class/Title
    analysis_ws.row_dimensions[4].height = 16  # Date
    print(
        f" ✅ BN ANALYSIS updated with PERSISTENT withdrawn tracking: {overall_withdrawn} withdrawn students"
    )
    print(f" ✅ Student sorting applied across all sheets")

def apply_complete_professional_formatting(wb, semester_key, header_row, set_name):
    """Apply complete professional formatting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    print(f" 🎨 Applying formatting...")
    # Find sheet
    current_sheet = None
    for sheet in wb.sheetnames:
        if semester_key.upper() in sheet.upper():
            current_sheet = sheet
            break
    if not current_sheet:
        return
    ws = wb[current_sheet]
    # Define styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    # Format header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    # Set specific widths
    if header_row > 0:
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                header_str = str(header_val).upper()
                if "EXAM NUMBER" in header_str:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20
                elif "NAME" in header_str:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 35
    print(f" ✅ Formatting applied")

# CRITICAL FIX: SINGLE WORKBOOK SESSION VERSION - NO CORRUPTION
def update_mastersheet_with_recalculation_FINAL(
    mastersheet_path,
    updates,
    semester_key,
    original_zip_path,
    course_titles_dict,
    course_units_dict,
    set_name,
):
    """FINAL FIXED VERSION - SINGLE workbook session to prevent Excel corruption - BN VERSION"""
    print(f"\n{'='*80}")
    print(f"🔄 FINAL FIX: UPDATING BN MASTERSHEET WITH SINGLE SESSION")
    print(f" SET NAME: '{set_name}'")
    print(f"{'='*80}")
    # Create backup first
    backup_path = mastersheet_path.replace(".xlsx", "_BACKUP.xlsx")
    try:
        shutil.copy2(mastersheet_path, backup_path)
        print(f"💾 Created backup: {backup_path}")
    except Exception as e:
        print(f"⚠️ Could not create backup: {e}")
    wb = None
    try:
        # SINGLE WORKBOOK LOAD - No repeated loading/closing
        print(f"📖 Loading workbook ONCE...")
        wb = load_workbook(mastersheet_path)
        # ⭐ ADD THIS CRITICAL LINE ⭐
        ensure_required_sheets_exist(wb)
        # Find semester sheet
        sheet_name = None
        for sheet in wb.sheetnames:
            if semester_key.upper() in sheet.upper():
                sheet_name = sheet
                break
        if not sheet_name:
            print(f"❌ No sheet found for: {semester_key}")
            return False
        ws = wb[sheet_name]
        print(f"✅ Working on sheet: {sheet_name}")
        # Find header row using the fixed function
        header_row, headers = find_sheet_structure(ws)
        if not header_row or not headers:
            print(f"❌ Could not find header row")
            return False
        print(f"✅ Found header row at: {header_row}")
        # Identify course columns properly
        course_columns = identify_course_columns_properly(headers)
        if not course_columns:
            print(f"❌ No course columns found!")
            return False
        print(f"✅ Found {len(course_columns)} course columns")
        # Find summary columns
        summary_columns = {}
        summary_keys = [
            "FAILED COURSES",
            "REMARKS",
            "CU Passed",
            "CU Failed",
            "TCPE",
            "GPA",
            "AVERAGE",
            "CGPA",
        ]
        for key in summary_keys:
            for header, col_idx in headers.items():
                if key.upper() in header.upper():
                    summary_columns[key] = col_idx
                    break
        # Helper functions
        def get_grade_point(score):
            try:
                score = float(score)
                if score >= 70:
                    return 5.0
                elif score >= 60:
                    return 4.0
                elif score >= 50:
                    return 3.0
                elif score >= 45:
                    return 2.0
                elif score >= 40:
                    return 1.0
                else:
                    return 0.0
            except:
                return 0.0

        def find_credit_unit_simple(course_code, units_dict):
            if course_code in units_dict:
                return units_dict[course_code]
            code_no_space = course_code.replace(" ", "")
            if code_no_space in units_dict:
                return units_dict[code_no_space]
            return 2
        # =============================================================
        # STEP 1: Apply ALL score updates (PASSED AND FAILED RESITS)
        # =============================================================
        print(f"\n📝 STEP 1: APPLYING ALL SCORE UPDATES (PASSED AND FAILED RESITS)...")
        students_updated = 0
        courses_updated = 0
        # DEBUG: Verify updates are being applied
        print(f"🔍 DEBUG: Checking if updates are being applied...")
        for exam_no, course_updates in updates.items():
            print(f"📝 Updates for {exam_no}: {len(course_updates)} courses")
            for course_code, new_score in course_updates.items():
                print(f" - {course_code}: {new_score}")
        # Also add debug to verify we're finding the right students
        exam_col = None
        for header, col_idx in headers.items():
            if "EXAM NUMBER" in header.upper():
                exam_col = col_idx
                break
        if not exam_col:
            print("❌ No exam column found")
            return False
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no = ws.cell(row_idx, column=exam_col)
            exam_no = str(exam_no.value).strip().upper() if exam_no.value else None
            if not exam_no or exam_no in ["", "NAN", "NONE"]:
                continue
            if "SUMMARY" in str(exam_no).upper():
                break
            if exam_no in updates:
                print(
                    f"🎯 FOUND student {exam_no} at row {row_idx} - will update {len(updates[exam_no])} courses"
                )
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no_cell = ws.cell(row=row_idx, column=exam_col)
            exam_no = (
                str(exam_no_cell.value).strip().upper() if exam_no_cell.value else None
            )
            if not exam_no or exam_no in ["", "NAN", "NONE"]:
                continue
            if "SUMMARY" in str(exam_no).upper():
                break
            if exam_no in updates:
                for course_code, new_score in updates[exam_no].items():
                    if course_code in course_columns:
                        course_col = course_columns[course_code]
                        old_score = ws.cell(row=row_idx, column=course_col).value
                        # CRITICAL FIX: Update ALL resit scores (both passed and failed)
                        # This ensures failed resit values are also recorded
                        ws.cell(row=row_idx, column=course_col).value = new_score
                        # Apply color coding based on pass/fail status
                        if new_score >= DEFAULT_PASS_THRESHOLD:
                            # GREEN for passed resits
                            ws.cell(row=row_idx, column=course_col).fill = PatternFill(
                                start_color="90EE90",
                                end_color="90EE90",
                                fill_type="solid",
                            )
                        else:
                            # ORANGE for failed resits (to distinguish from original fails)
                            ws.cell(row=row_idx, column=course_col).fill = PatternFill(
                                start_color="FFD580",
                                end_color="FFD580",
                                fill_type="solid",
                            )
                        ws.cell(row=row_idx, column=course_col).font = Font(bold=True)
                        print(
                            f" ✅ {exam_no} - {course_code}: {old_score} → {new_score} ({'PASSED' if new_score >= DEFAULT_PASS_THRESHOLD else 'FAILED'})"
                        )
                        courses_updated += 1
                students_updated += 1
        # VERIFICATION: Check if scores were actually updated
        print(f"🔍 VERIFICATION: Checking if scores were updated...")
        for exam_no in updates:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                exam_no_cell = ws.cell(row_idx, column=exam_col)
                current_exam_no = (
                    str(exam_no_cell.value).strip().upper()
                    if exam_no_cell.value
                    else None
                )
                if current_exam_no == exam_no:
                    for course_code, expected_score in updates[exam_no].items():
                        if course_code in course_columns:
                            course_col = course_columns[course_code]
                            actual_score = ws.cell(row_idx, column=course_col).value
                            print(
                                f" {exam_no} - {course_code}: Expected {expected_score}, Got {actual_score}"
                            )
                    break
        print(
            f"✅ Step 1 Complete: Updated {courses_updated} scores ({students_updated} students)"
        )
        # =============================================================
        # STEP 2: Recalculate student records with CURRENT scores
        # =============================================================
        print(f"\n🧮 STEP 2: RECALCULATING WITH CURRENT SCORES...")
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                exam_no_cell = ws.cell(row=row_idx, column=exam_col)
                exam_no = (
                    str(exam_no_cell.value).strip().upper()
                    if exam_no_cell.value
                    else None
                )
                if not exam_no or exam_no in ["", "NAN", "NONE"]:
                    continue
                if "SUMMARY" in str(exam_no).upper():
                    break
                # Read CURRENT scores from worksheet (including updated resit scores)
                failed_courses = []
                cu_passed = 0
                cu_failed = 0
                total_credits = 0
                total_grade_points = 0.0
                total_score = 0.0
                valid_courses = 0
                for course_code, course_col in course_columns.items():
                    score_cell = ws.cell(row=row_idx, column=course_col)
                    score_value = score_cell.value
                    if score_value is not None and score_value != "":
                        try:
                            score = float(score_value)
                            total_score += score
                            valid_courses += 1
                            credit_unit = find_credit_unit_simple(
                                course_code, course_units_dict
                            )
                            total_credits += credit_unit
                            grade_point = get_grade_point(score)
                            total_grade_points += grade_point * credit_unit
                            if score >= 50:
                                cu_passed += credit_unit
                            else:
                                cu_failed += credit_unit
                                failed_courses.append(course_code)
                        except (ValueError, TypeError):
                            continue
                # Calculate metrics
                gpa = (
                    round(total_grade_points / total_credits, 2)
                    if total_credits > 0
                    else 0.0
                )
                average = (
                    round(total_score / valid_courses, 2) if valid_courses > 0 else 0.0
                )
                cgpa = gpa  # Placeholder, updated later if needed
                # Determine remarks with fixed logic
                passed_percent = cu_passed / total_credits if total_credits > 0 else 0
                if passed_percent < 0.45:
                    remarks = "WITHDRAW"
                elif cu_failed == 0:
                    remarks = "PASSED"
                else:
                    if gpa >= 2.0:
                        remarks = "RESIT"
                    else:
                        remarks = "PROBATION"
                # Update all columns
                if "FAILED COURSES" in summary_columns:
                    ws.cell(
                        row=row_idx, column=summary_columns["FAILED COURSES"]
                    ).value = (", ".join(failed_courses) if failed_courses else "NONE")
                if "REMARKS" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["REMARKS"]).value = (
                        remarks
                    )
                if "CU Passed" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["CU Passed"]).value = (
                        cu_passed
                    )
                if "CU Failed" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["CU Failed"]).value = (
                        cu_failed
                    )
                if "TCPE" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["TCPE"]).value = (
                        total_credits
                    )
                if "GPA" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["GPA"]).value = gpa
                if "AVERAGE" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["AVERAGE"]).value = (
                        average
                    )
                if "CGPA" in summary_columns:
                    ws.cell(row=row_idx, column=summary_columns["CGPA"]).value = cgpa
                # Highlight updated students
                if exam_no in updates:
                    for col_idx in summary_columns.values():
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
                        )
            except Exception as e:
                print(f"⚠️ Error processing row {row_idx}: {e}")
                continue
        print(f"✅ Step 2 Complete: Recalculated all records")
        # =============================================================
        # STEP 3: Update SUMMARY section with CURRENT data
        # =============================================================
        print(f"\n📊 STEP 3: UPDATING SUMMARY SECTION...")
        update_summary_section_fixed(ws, headers, header_row, course_columns)
        print(f"✅ Step 3 Complete: Summary section updated")
        # =============================================================
        # STEP 4: Update CGPA_SUMMARY - FIXED VERSION with SINGLE session
        # =============================================================
        print(f"\n📈 STEP 4: UPDATING CGPA_SUMMARY...")
        print(f" Available sheets: {wb.sheetnames}")
        print(f" set_name parameter: '{set_name}'")
        if "CGPA_SUMMARY" in wb.sheetnames:
            print(f"✅ Calling update_cgpa_summary_sheet_fixed...")
            update_cgpa_summary_sheet_fixed(wb, semester_key, header_row, set_name)
            print(f"✅ Step 4 Complete: CGPA_SUMMARY updated")
        else:
            print(f"❌ CGPA_SUMMARY sheet still not found!")
        # =============================================================
        # STEP 5: Update ANALYSIS
        # =============================================================
        print(f"\n📊 STEP 5: UPDATING ANALYSIS...")
        print(f" Available sheets: {wb.sheetnames}")
        print(f" set_name parameter: '{set_name}'")
        if "ANALYSIS" in wb.sheetnames:
            print(f"✅ Calling update_analysis_sheet_fixed...")
            update_analysis_sheet_fixed(
                wb, semester_key, course_columns, headers, header_row, set_name
            )
            print(f"✅ Step 5 Complete: ANALYSIS updated")
        else:
            print(f"❌ ANALYSIS sheet still not found!")
        # =============================================================
        # STEP 6: Apply formatting and PROPER sorting
        # =============================================================
        print(f"\n🎨 STEP 6: APPLYING FORMATTING AND PROPER SORTING...")
        apply_complete_professional_formatting(wb, semester_key, header_row, set_name)
        # Apply student sorting to maintain proper order with CORRECT serial numbers
        apply_student_sorting_with_serial_numbers(ws, header_row, headers)
        print(f"✅ Step 6 Complete: Formatting and proper sorting applied")
        # =============================================================
        # FINAL SINGLE SAVE - CRITICAL FIX: Only save once at the end
        # =============================================================
        print(f"\n💾 FINAL SAVE: Saving workbook...")
        try:
            wb.save(mastersheet_path)
            print(f"✅ FINAL SAVE COMPLETE")
            # Verify file integrity
            file_size = os.path.getsize(mastersheet_path)
            print(f"📁 File size: {file_size} bytes")
            if file_size > 10000:  # Reasonable minimum size for Excel file
                print(f"✅ File integrity verified - Size looks good")
                # Quick test load
                try:
                    test_wb = load_workbook(mastersheet_path)
                    test_wb.close()
                    print(f"✅ File can be opened successfully - No corruption")
                    return True
                except Exception as test_error:
                    print(f"❌ File may be corrupted - Test open failed: {test_error}")
                    # Restore from backup
                    try:
                        shutil.copy2(backup_path, mastersheet_path)
                        print(f"🔄 Restored from backup due to corruption")
                    except:
                        pass
                    return False
            else:
                print(f"❌ File too small - likely corrupted")
                return False
        except Exception as save_error:
            print(f"❌ Error during final save: {save_error}")
            # Restore from backup
            try:
                shutil.copy2(backup_path, mastersheet_path)
                print(f"🔄 Restored from backup due to save error")
            except:
                pass
            return False
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # CRITICAL: Always close workbook if it exists
        if wb:
            try:
                wb.close()
                print("✅ Workbook properly closed")
            except:
                pass

# ----------------------------
# Carryover Processing Functions (BN-Compatible)
# ----------------------------

def load_carryover_files(carryover_dir, semester_key=None):
    """Load carryover files - BN VERSION"""
    carryover_files = []
    for file in os.listdir(carryover_dir):
        # Look for both JSON and Excel carryover files
        if file.startswith("co_student_") and (
            file.endswith(".json") or file.endswith(".xlsx")
        ):
            file_semester = extract_semester_from_filename(file)
            file_semester_standardized = (
                standardize_semester_name(file_semester) if file_semester else None
            )
            # Skip if semester doesn't match filter
            if semester_key and file_semester_standardized != semester_key:
                continue
            file_path = os.path.join(carryover_dir, file)
            try:
                if file.endswith(".json"):
                    with open(file_path, "r") as f:
                        data = json.load(f)
                else:  # .xlsx
                    df = pd.read_excel(file_path)
                    data = df.to_dict("records")
                carryover_files.append(
                    {
                        "filename": file,
                        "semester": file_semester_standardized,
                        "data": data,
                        "count": len(data),
                        "file_path": file_path,
                    }
                )
                print(f"✅ Loaded {len(data)} records from {file}")
            except Exception as e:
                print(f"❌ Error loading {file}: {e}")
    if not carryover_files:
        print(f"❌ No carryover files found in {carryover_dir}")
        return []
    print(f"📚 Total BN carryover files loaded: {len(carryover_files)}")
    return carryover_files

def save_carryover_json_records(carryover_data, carryover_output_dir, semester_key):
    """
    Save BN carryover records as JSON files
    """
    json_dir = os.path.join(carryover_output_dir, "CARRYOVER_RECORDS")
    os.makedirs(json_dir, exist_ok=True)
    print(f"\n💾 SAVING BN CARRYOVER JSON RECORDS")
    print(f"📁 JSON directory: {json_dir}")
    # Convert carryover_data to JSON-friendly format
    json_records = {}
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        # Build the JSON record structure
        record = {
            "exam_number": exam_no,
            "name": student["NAME"],
            "carryover_courses": {},
            "passed_resit_courses": {},
            "failed_resit_courses": {},
        }
        # Process each resit course
        for course_code, course_data in student["RESIT_COURSES"].items():
            course_info = {
                "course_code": course_code,
                "course_title": course_data.get("course_title", course_code),
                "credit_unit": course_data.get("credit_unit", 2),
                "original_score": course_data["original_score"],
                "resit_score": course_data["resit_score"],
            }
            # Add to carryover courses (all courses that were taken as resit)
            record["carryover_courses"][course_code] = course_info
            # Separate into passed/failed based on resit score
            if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD:
                record["passed_resit_courses"][course_code] = course_info
            else:
                record["failed_resit_courses"][course_code] = course_info
        json_records[exam_no] = record
    # Save as JSON file with timestamp and semester
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f"co_student_BN-{semester_key}_{timestamp}.json"
    json_filepath = os.path.join(json_dir, json_filename)
    try:
        with open(json_filepath, "w", encoding="utf-8") as f:
            json.dump(json_records, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved BN JSON carryover records: {json_filepath}")
        print(f"📊 Records saved: {len(json_records)} students")
        return json_filepath
    except Exception as e:
        print(f"❌ Error saving BN JSON records: {e}")
        traceback.print_exc()
        return None

def copy_json_to_centralized_location(json_filepath, set_name, semester_key):
    """
    Copy JSON file to centralized CARRYOVER_RECORDS location for BN
    """
    try:
        # Determine the centralized location
        base_dir = get_base_directory()
        centralized_dir = os.path.join(
            base_dir,
            "EXAMS_INTERNAL",
            "BN",
            set_name,
            "CLEAN_RESULTS",
            "CARRYOVER_RECORDS",
        )
        os.makedirs(centralized_dir, exist_ok=True)
        # Copy the JSON file
        filename = os.path.basename(json_filepath)
        dest_path = os.path.join(centralized_dir, filename)
        shutil.copy2(json_filepath, dest_path)
        print(f"\n📋 COPIED TO BN CENTRALIZED LOCATION")
        print(f"✅ From: {json_filepath}")
        print(f"✅ To: {dest_path}")
        return dest_path
    except Exception as e:
        print(f"❌ Error copying to BN centralized location: {e}")
        traceback.print_exc()
        return None

def generate_carryover_mastersheet(
    carryover_data,
    output_dir,
    semester_key,
    set_name,
    timestamp,
    cgpa_data,
    course_titles,
    course_units,
    course_code_to_title,
    course_code_to_unit,
):
    """Generate BN CARRYOVER_mastersheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BN_CARRYOVER_RESULTS"
    program_name = "BASIC NURSING"
    program_abbr = "BN"
    
    if os.path.exists(DEFAULT_LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image
            img = Image(DEFAULT_LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"⚠️ Could not add logo: {e}")
    
    current_year = 2025
    next_year = 2026
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )
    
    all_courses = set()
    for student in carryover_data:
        all_courses.update(student["RESIT_COURSES"].keys())
    
    previous_semesters = get_previous_semesters_for_display(semester_key)
    
    headers = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        headers.append(f"GPA {prev_sem}")
    
    course_headers = []
    for course in sorted(all_courses):
        course_headers.extend([f"{course}", f"{course}_RESIT"])
    
    headers.extend(course_headers)
    headers.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    
    total_columns = len(headers)
    last_column = get_column_letter(total_columns)
    # Header section
    ws.merge_cells(f"A3:{last_column}3")
    title_cell = ws["A3"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A4:{last_column}4")
    subtitle_cell = ws["A4"]
    subtitle_cell.value = f"RESIT - {current_year}/{next_year} SESSION {program_name} {level} {sem_display} EXAMINATIONS RESULT — {datetime.now().strftime('%B %d, %Y')}"
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    print(f"🔍 BN Courses found in resit data: {sorted(all_courses)}")
    print(
        f"📊 BN GPA columns for {semester_key}: Previous={previous_semesters}, Current={current_semester_name}"
    )
    # Course title row
    title_row = [""] * 3
    for prev_sem in previous_semesters:
        title_row.extend([""])
    for course in sorted(all_courses):
        course_title = find_course_title(course, course_titles, course_code_to_title)
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        title_row.extend([course_title, course_title])
    title_row.extend(["", "", ""])
    ws.append(title_row)
    # Credit unit row
    credit_row = [""] * 3
    for prev_sem in previous_semesters:
        credit_row.extend([""])
    for course in sorted(all_courses):
        credit_unit = find_credit_unit(course, course_units, course_code_to_unit)
        credit_row.extend([f"CU: {credit_unit}", f"CU: {credit_unit}"])
    credit_row.extend(["", "", ""])
    ws.append(credit_row)
    # Code row
    code_row = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        code_row.append(f"GPA {prev_sem}")
    for course in sorted(all_courses):
        code_row.append(f"{course}")
        code_row.append(f"{course}_RESIT")
    code_row.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    ws.append(code_row)
    # Apply formatting to header rows
    course_colors = [
        "E6F3FF", "FFF0E6", "E6FFE6", "FFF6E6", "F0E6FF", "E6FFFF", "FFE6F2", "F5F5DC", "E6F7FF", "FFF5E6"
    ]
    
    start_col = 4
    if previous_semesters:
        start_col += len(previous_semesters)
    color_index = 0
    
    for course in sorted(all_courses):
        for row in [5, 6, 7]:
            for offset in [0, 1]:
                cell = ws.cell(row=row, column=start_col + offset)
                cell.fill = PatternFill(
                    start_color=course_colors[color_index % len(course_colors)],
                    end_color=course_colors[color_index % len(course_colors)],
                    fill_type="solid"
                )
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        for offset in [0, 1]:
            cell = ws.cell(row=5, column=start_col + offset)
            cell.alignment = Alignment(
                text_rotation=90, horizontal="center", vertical="center"
            )
            cell.font = Font(bold=True, size=9)
        color_index += 1
        start_col += 2
    # Format basic info columns
    for row in [5, 6, 7]:
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
    # Add student data
    row_idx = 8
    failed_counts = {course: 0 for course in all_courses}
    
    # CRITICAL FIX: Proper serial numbers from 1 to n
    serial_number = 1
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        # Serial Number (PROPERLY SORTED from 1 to n)
        ws.cell(row=row_idx, column=1, value=serial_number)
        serial_number += 1
        ws.cell(row=row_idx, column=2, value=student["EXAM NUMBER"])
        ws.cell(row=row_idx, column=3, value=student["NAME"])
        # Previous GPAs
        gpa_col = 4
        for prev_sem in previous_semesters:
            gpa_value = student.get(f"GPA_{prev_sem}", "")
            ws.cell(row=row_idx, column=gpa_col, value=gpa_value)
            gpa_col += 1
        # Course scores
        course_col = gpa_col
        color_index = 0
        for course in sorted(all_courses):
            for offset in [0, 1]:
                cell = ws.cell(row=row_idx, column=course_col + offset)
                cell.fill = PatternFill(
                    start_color=course_colors[color_index % len(course_colors)],
                    end_color=course_colors[color_index % len(course_colors)],
                    fill_type="solid"
                )
            if course in student["RESIT_COURSES"]:
                course_data = student["RESIT_COURSES"][course]
                orig_cell = ws.cell(
                    row=row_idx, column=course_col, value=course_data["original_score"]
                )
                if course_data["original_score"] < DEFAULT_PASS_THRESHOLD:
                    orig_cell.fill = PatternFill(
                        start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                    )
                resit_cell = ws.cell(
                    row=row_idx, column=course_col + 1, value=course_data["resit_score"]
                )
                if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD:
                    resit_cell.fill = PatternFill(
                        start_color="90EE90", end_color="90EE90", fill_type="solid"
                    )
                else:
                    resit_cell.fill = PatternFill(
                        start_color="FFD580", end_color="FFD580", fill_type="solid"
                    )
                    failed_counts[course] += 1
            else:
                ws.cell(row=row_idx, column=course_col, value="")
                ws.cell(row=row_idx, column=course_col + 1, value="")
            color_index += 1
            course_col += 2
        # Current GPA and CGPA
        ws.cell(row=row_idx, column=course_col, value=student["CURRENT_GPA"])
        ws.cell(row=row_idx, column=course_col + 1, value=student["CURRENT_CGPA"])
        # Remarks
        remarks = generate_remarks(student["RESIT_COURSES"])
        ws.cell(row=row_idx, column=course_col + 2, value=remarks)
        row_idx += 1
    # Add failed counts row
    failed_row_idx = row_idx
    ws.cell(row=failed_row_idx, column=1, value="FAILED COUNT BY COURSE:").font = Font(
        bold=True
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=failed_row_idx, column=col)
        cell.fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
    
    course_col = gpa_col
    for course in sorted(all_courses):
        count_cell = ws.cell(
            row=failed_row_idx, column=course_col + 1, value=failed_counts[course]
        )
        count_cell.font = Font(bold=True)
        count_cell.fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
        course_col += 2
    # Add summary section
    summary_start_row = failed_row_idx + 2
    total_students = len(carryover_data)
    passed_all = sum(
        1
        for student in carryover_data
        if all(
            course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
            for course_data in student["RESIT_COURSES"].values()
        )
    )
    carryover_count = total_students - passed_all
    total_failed_attempts = sum(failed_counts.values())
    summary_data = [
        ["BN CARRYOVER SUMMARY"],
        [
            f"A total of {total_students} students registered and sat for the Carryover Examination"
        ],
        [f"A total of {passed_all} students passed all carryover courses"],
        [
            f"A total of {carryover_count} students failed one or more carryover courses and must repeat them"
        ],
        [f"Total failed resit attempts: {total_failed_attempts} across all courses"],
        [
            f"BN Carryover processing completed on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"
        ],
    ]
    for i, row_data in enumerate(summary_data):
        row_num = summary_start_row + i
        if row_data[0]:
            ws.merge_cells(
                start_row=row_num, start_column=1, end_row=row_num, end_column=10
            )
            cell = ws.cell(row=row_num, column=1, value=row_data[0])
            if i == 0:
                cell.font = Font(bold=True, size=12, underline="single")
            else:
                cell.font = Font(bold=False, size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    # Apply borders to data area
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=7, max_row=row_idx - 1, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border = thin_border
    # Freeze panes and adjust column widths
    ws.freeze_panes = "D8"
    
    # Set column widths
    ws.column_dimensions["A"].width = 8  # S/N
    ws.column_dimensions["B"].width = 18  # EXAM NUMBER
    ws.column_dimensions["C"].width = 35  # NAME
    
    # Set GPA column widths
    for col in range(4, 4 + len(previous_semesters)):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Set course column widths
    for col in range(4 + len(previous_semesters), len(headers) - 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Set summary column widths
    for col in range(len(headers) - 2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    # Save the file
    filename = f"BN_CARRYOVER_mastersheet_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()
    print(f"✅ BN CARRYOVER mastersheet generated: {filepath}")
    return filepath

def generate_individual_reports(
    carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data
):
    """Generate individual BN student reports."""
    reports_dir = os.path.join(output_dir, "INDIVIDUAL_REPORTS")
    os.makedirs(reports_dir, exist_ok=True)
    
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        safe_exam_no = sanitize_filename(exam_no)
        filename = f"bn_carryover_report_{safe_exam_no}_{timestamp}.csv"
        filepath = os.path.join(reports_dir, filename)
        report_data = []
        report_data.append(["BN CARRYOVER RESULT REPORT"])
        report_data.append(["FCT COLLEGE OF NURSING SCIENCES"])
        report_data.append([f"BN Set: {set_name}"])
        report_data.append([f"BN Semester: {semester_key}"])
        report_data.append([])
        report_data.append(["BN STUDENT INFORMATION"])
        report_data.append(["Exam Number:", student["EXAM NUMBER"]])
        report_data.append(["Name:", student["NAME"]])
        report_data.append([])
        report_data.append(["BN PREVIOUS GPAs"])
        for key in sorted([k for k in student.keys() if k.startswith("GPA_")]):
            semester = key.replace("GPA_", "")
            report_data.append([f"{semester}:", student[key]])
        report_data.append([])
        report_data.append(["BN CURRENT ACADEMIC RECORD"])
        report_data.append(["Current GPA:", student["CURRENT_GPA"]])
        report_data.append(["Current CGPA:", student["CURRENT_CGPA"]])
        report_data.append([])
        report_data.append(["BN RESIT COURSES"])
        report_data.append(
            [
                "Course Code",
                "Course Title",
                "Credit Unit",
                "Original Score",
                "Resit Score",
                "Status",
            ]
        )
        for course_code, course_data in student["RESIT_COURSES"].items():
            status = (
                "PASSED"
                if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
                else "FAILED"
            )
            course_title = course_data.get("course_title", course_code)
            credit_unit = course_data.get("credit_unit", 0)
            report_data.append(
                [
                    course_code,
                    course_title,
                    credit_unit,
                    course_data["original_score"],
                    course_data["resit_score"],
                    status,
                ]
            )
        try:
            df = pd.DataFrame(report_data)
            df.to_csv(filepath, index=False, header=False)
            print(f"✅ Generated BN report for: {exam_no}")
        except Exception as e:
            print(f"❌ Error generating BN report for {exam_no}: {e}")
    
    print(
        f"✅ Generated {len(carryover_data)} individual BN student reports in {reports_dir}"
    )

# ----------------------------
# CRITICAL FIX: Output Directory Management
# ----------------------------

def get_output_directory(set_name):
    """Get the correct output directory for carryover results - FIXED"""
    # For web interface: Save directly to CLEAN_RESULTS
    clean_dir = os.path.join(BASE_DIR, "EXAMS_INTERNAL", "BN", set_name, "CLEAN_RESULTS")
    
    # Alternative paths
    if not os.path.exists(clean_dir):
        clean_dir = os.path.join(BASE_DIR, "BN", set_name, "CLEAN_RESULTS")
    
    if not os.path.exists(clean_dir):
        print(f"⚠️ CLEAN_RESULTS directory doesn't exist, will create: {clean_dir}")
        os.makedirs(clean_dir, exist_ok=True)
    
    print(f"📁 Output directory: {clean_dir}")
    return clean_dir

# ----------------------------
# CRITICAL FIX: Main Processing Function
# ----------------------------

def debug_mastersheet_structure(mastersheet_path, sheet_name):
    """Debug function to examine mastersheet structure"""
    print(f"\n🔍 DEBUGGING MASTERSHEET STRUCTURE")
    print("=" * 50)
    
    # Try reading with different headers
    for header_row in [0, 1, 2, 3, 4, 5]:
        try:
            df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=header_row)
            print(f"\n📊 Header Row {header_row}:")
            print(f"   Shape: {df.shape}")
            print(f"   Columns: {df.columns.tolist()}")
            
            # Look for exam number column
            exam_col = find_exam_number_column(df)
            print(f"   Exam column found: {exam_col}")
            
            if exam_col and exam_col in df.columns:
                print(f"   Sample exam numbers: {df[exam_col].head(3).tolist()}")
                
        except Exception as e:
            print(f"   Error with header {header_row}: {e}")

def process_carryover_results(
    resit_file_path,
    source_path,
    source_type,
    semester_key,
    set_name,
    pass_threshold,
    output_dir,
):
    """
    FIXED VERSION: Process BN carryover results with robust mastersheet reading
    """
    print(f"\n🔄 FIXED VERSION: PROCESSING BN CARRYOVER RESULT FOR {semester_key}")
    print("=" * 60)
    
    # Load BN course data
    (
        semester_course_titles,
        semester_credit_units,
        course_code_to_title,
        course_code_to_unit,
    ) = load_course_data()
    
    # Debug course matching
    debug_course_matching_bn(resit_file_path, course_code_to_title, course_code_to_unit)
    
    # Get semester info
    year, sem_num, level, sem_display, set_code, sem_name = get_semester_display_info(
        semester_key
    )
    
    # Find course titles for the specific semester
    possible_sheet_keys = [
        f"{set_code} {sem_display}",
        f"{set_code} {sem_name}",
        semester_key,
        semester_key.replace("-", " ").upper(),
        f"{level} {sem_display}",
    ]
    
    course_titles_dict = {}
    credit_units_dict = {}
    
    for sheet_key in possible_sheet_keys:
        sheet_standard = standardize_semester_key(sheet_key)
        if sheet_standard in semester_course_titles:
            course_titles_dict = semester_course_titles[sheet_standard]
            credit_units_dict = semester_credit_units[sheet_standard]
            print(
                f"✅ Using BN sheet key: '{sheet_key}' with {len(course_titles_dict)} courses"
            )
            break
        else:
            print(f"❌ BN sheet key not found: '{sheet_key}'")
    
    if not course_titles_dict:
        print(
            f"⚠️ No BN semester-specific course data found, using global course mappings"
        )
        course_titles_dict = course_code_to_title
        credit_units_dict = course_code_to_unit
    
    print(
        f"📊 Final BN course mappings: {len(course_titles_dict)} titles, {len(credit_units_dict)} units"
    )
    
    timestamp = datetime.now().strftime(TIMESTAMP_FMT)
    carryover_output_dir = os.path.join(
        output_dir, f"BN_CARRYOVER_{set_name}_{semester_key}_{timestamp}"
    )
    os.makedirs(carryover_output_dir, exist_ok=True)
    print(f"📁 BN Output directory: {carryover_output_dir}")
    
    if not os.path.exists(resit_file_path):
        print(f"❌ BN resit file not found: {resit_file_path}")
        return False
    
    temp_mastersheet_path = None
    temp_dir = None
    updated_zip_path = None
    
    try:
        temp_mastersheet_path, temp_dir = get_mastersheet_path(
            source_path, source_type, semester_key
        )
        if not temp_mastersheet_path:
            print(f"❌ Failed to get BN mastersheet")
            return False
        
        print(f"📖 Reading BN files...")
        resit_df = pd.read_excel(resit_file_path, header=0)
        
        # DEBUG: Print resit file info
        print(f"📊 BN Resit file rows: {len(resit_df)}")
        print(f"📊 BN Resit file columns: {resit_df.columns.tolist()}")
        resit_exam_col = find_exam_number_column(resit_df)
        print(f"📊 BN Resit exam column: '{resit_exam_col}'")
        
        if resit_exam_col:
            print(
                f"📊 Sample BN resit exam numbers: {resit_df[resit_exam_col].head().tolist()}"
            )
        
        xl = pd.ExcelFile(temp_mastersheet_path)
        sheet_name = get_matching_sheet(xl, semester_key)
        if not sheet_name:
            print(f"❌ No matching BN sheet found for {semester_key}")
            return False
        
        print(f"📖 Using BN sheet '{sheet_name}' for current semester {semester_key}")
        
        # FIXED: Use the enhanced mastersheet reading function
        mastersheet_df, mastersheet_exam_col = read_mastersheet_with_flexible_headers(
            temp_mastersheet_path, sheet_name
        )
        
        if mastersheet_df is None or mastersheet_exam_col is None:
            print(f"❌ Could not read BN mastersheet with flexible headers")
            # Fallback to quick fix
            print(f"🔄 Trying quick fix...")
            mastersheet_df, mastersheet_exam_col = quick_fix_read_mastersheet(
                temp_mastersheet_path, sheet_name
            )
            
        if mastersheet_df is None or mastersheet_exam_col is None:
            print(f"❌ Could not read BN mastersheet with any method")
            return False
            
        # DEBUG: Print mastersheet info
        print(f"📊 BN Mastersheet rows: {len(mastersheet_df)}")
        print(f"📊 BN Mastersheet columns: {mastersheet_df.columns.tolist()}")
        
        if mastersheet_exam_col in mastersheet_df.columns:
            print(f"📊 Sample BN mastersheet exam numbers: {mastersheet_df[mastersheet_exam_col].head().tolist()}")
        
        print(f"✅ BN files loaded - Resit: {len(resit_df)} rows, Mastersheet: {len(mastersheet_df)} students")
        
        resit_exam_col = find_exam_number_column(resit_df)
        if not resit_exam_col:
            print(f"❌ Cannot find exam number column in BN resit file")
            return None
            
        print(f"📝 BN Exam columns - Resit: '{resit_exam_col}', Mastersheet: '{mastersheet_exam_col}'")
        
        # Load previous GPAs for CGPA calculation
        cgpa_data = load_previous_gpas(temp_mastersheet_path, semester_key)
        carryover_data = []
        updated_students = set()
        
        print(f"\n🎯 PROCESSING BN RESIT SCORES...")
        
        # FIXED: Use enhanced student matching
        for idx, resit_row in resit_df.iterrows():
            try:
                exam_no = str(resit_row[resit_exam_col]).strip().upper()
                if not exam_no or exam_no in ["NAN", "NONE", ""]:
                    continue
                
                # FIXED: Use enhanced student matching
                student_data = find_student_in_mastersheet_fixed(
                    exam_no, mastersheet_df, mastersheet_exam_col
                )
                
                if student_data is None:
                    print(f"⚠️ BN Student {exam_no} not found in mastersheet - skipping")
                    continue
                    
                student_name = student_data.get("NAME", "Unknown")
                current_credits = 0
                
                # Find credits column
                for col in mastersheet_df.columns:
                    if "TCPE" in str(col).upper():
                        current_credits = student_data.get(col, 0)
                        break
                
                student_record = {
                    "EXAM NUMBER": exam_no,
                    "NAME": student_name,
                    "RESIT_COURSES": {},
                    "CURRENT_GPA": student_data.get("GPA", 0),
                    "CURRENT_CREDITS": current_credits,
                }
                
                # Process resit courses
                for col in resit_df.columns:
                    if col == resit_exam_col or col == "NAME" or "Unnamed" in str(col):
                        continue
                        
                    resit_score = resit_row.get(col)
                    if pd.isna(resit_score) or resit_score == "":
                        continue
                        
                    try:
                        resit_score_val = float(resit_score)
                    except (ValueError, TypeError):
                        continue
                    
                    # Check if course exists in mastersheet
                    if col in mastersheet_df.columns:
                        original_score = student_data.get(col)
                        if pd.isna(original_score):
                            continue
                    else:
                        # Try to find course with similar name
                        course_found = False
                        for ms_col in mastersheet_df.columns:
                            if col.upper() == ms_col.upper() or col.replace(" ", "") == ms_col.replace(" ", ""):
                                original_score = student_data.get(ms_col)
                                course_found = True
                                break
                        if not course_found:
                            continue
                    
                    try:
                        original_score_val = (
                            float(original_score) if not pd.isna(original_score) else 0.0
                        )
                    except (ValueError, TypeError):
                        original_score_val = 0.0
                    
                    if original_score_val < pass_threshold:
                        course_title = find_course_title(
                            col, course_titles_dict, course_code_to_title
                        )
                        credit_unit = find_credit_unit(
                            col, credit_units_dict, course_code_to_unit
                        )
                        student_record["RESIT_COURSES"][col] = {
                            "original_score": original_score_val,
                            "resit_score": resit_score_val,
                            "updated": resit_score_val >= pass_threshold,
                            "course_title": course_title,
                            "credit_unit": credit_unit,
                        }
                
                # Enhanced GPA/CGPA calculation with resit scores
                if student_record["RESIT_COURSES"]:
                    # Identify course columns in mastersheet
                    import re
                    course_columns = [
                        col
                        for col in mastersheet_df.columns
                        if re.match(r"^[A-Z]{3}\d{3}$", str(col).upper())
                    ]
                    
                    # Recalculate updated current GPA with resit overrides
                    total_grade_points = 0.0
                    total_credits = 0
                    
                    for col in course_columns:
                        if col in mastersheet_df.columns:
                            original_score = student_data.get(col, 0)
                            score = original_score
                            
                            # Apply resit scores if available
                            if col in student_record["RESIT_COURSES"]:
                                score = student_record["RESIT_COURSES"][col]["resit_score"]
                            
                            try:
                                score_val = float(score)
                                credit_unit = find_credit_unit(
                                    col, credit_units_dict, course_code_to_unit
                                )
                                grade_point = get_grade_point(score_val)
                                total_grade_points += grade_point * credit_unit
                                total_credits += credit_unit
                            except (ValueError, TypeError):
                                continue
                    
                    updated_gpa = (
                        round(total_grade_points / total_credits, 2)
                        if total_credits > 0
                        else 0.0
                    )
                    student_record["CURRENT_GPA"] = updated_gpa
                    student_record["CURRENT_CREDITS"] = total_credits
                    
                    # Recalculate CGPA
                    if exam_no in cgpa_data:
                        student_record["CURRENT_CGPA"] = calculate_cgpa(
                            cgpa_data[exam_no], updated_gpa, total_credits
                        )
                    else:
                        student_record["CURRENT_CGPA"] = updated_gpa
                    
                    carryover_data.append(student_record)
                    updated_students.add(exam_no)
                    print(
                        f"✅ BN {exam_no}: {len(student_record['RESIT_COURSES'])} resit courses, Updated GPA: {student_record['CURRENT_GPA']}, CGPA: {student_record['CURRENT_CGPA']}"
                    )
                    
            except Exception as e:
                print(f"❌ Error processing BN student {exam_no if 'exam_no' in locals() else 'unknown'}: {e}")
                continue

        # DEBUG: Print final stats
        print(f"\n📊 BN FINAL STATS:")
        print(f" Total carryover students processed: {len(carryover_data)}")
        print(f" Students with updates: {len(updated_students)}")
        
        if carryover_data:
            print(f"\n📊 GENERATING BN OUTPUTS...")
            # 1. Generate the Excel carryover mastersheet
            carryover_mastersheet_path = generate_carryover_mastersheet(
                carryover_data,
                carryover_output_dir,
                semester_key,
                set_name,
                timestamp,
                cgpa_data,
                course_titles_dict,
                credit_units_dict,
                course_code_to_title,
                course_code_to_unit,
            )
            
            # 2. Generate individual reports
            generate_individual_reports(
                carryover_data,
                carryover_output_dir,
                semester_key,
                set_name,
                timestamp,
                cgpa_data,
            )
            
            # 3. Generate JSON records
            json_filepath = save_carryover_json_records(
                carryover_data, carryover_output_dir, semester_key
            )
            if json_filepath:
                # 4. Copy to centralized location
                centralized_json = copy_json_to_centralized_location(
                    json_filepath, set_name, semester_key
                )
            
            # 5. Create carryover ZIP file
            zip_path = os.path.join(
                output_dir, f"BN_CARRYOVER_{set_name}_{semester_key}_{timestamp}.zip"
            )
            if create_carryover_zip(carryover_output_dir, zip_path):
                print(f"✅ Final BN carryover ZIP created: {zip_path}")
            
            # ============================================
            # STEP 6: UPDATE ORIGINAL MASTERSHEET WITH ALL ENHANCEMENTS
            # ============================================
            if carryover_data:
                print(f"\n{'='*60}")
                print(f"🔄 STEP 6: UPDATING ORIGINAL BN MASTERSHEET WITH ALL ENHANCEMENTS")
                print(f"{'='*60}")
                try:
                    # Find the original result ZIP
                    clean_dir_parent = output_dir
                    all_result_zips = [
                        f
                        for f in os.listdir(clean_dir_parent)
                        if f.endswith(".zip") and ("RESULT-" in f or "UPDATED_" in f)
                    ]
                    all_result_zips = [
                        f for f in all_result_zips if "CARRYOVER" not in f.upper()
                    ]
                    if not all_result_zips:
                        print(f"❌ No original BN result ZIP found")
                    else:
                        latest_zip_name = max(
                            all_result_zips,
                            key=lambda f: os.path.getmtime(
                                os.path.join(clean_dir_parent, f)
                            ),
                        )
                        latest_result_zip = latest_zip_name
                        original_zip_path = os.path.join(
                            clean_dir_parent, latest_result_zip
                        )
                        print(f"✅ Found latest BN ZIP: {original_zip_path}")
                        
                        # Compute new zip name
                        match = re.search(r"UPDATED_(\d+)_", latest_zip_name)
                        current_count = int(match.group(1)) if match else 0
                        new_count = current_count + 1
                        if current_count == 0:
                            updated_zip_name = f"UPDATED_{new_count}_{latest_zip_name}"
                        else:
                            updated_zip_name = re.sub(
                                r"UPDATED_\d+", f"UPDATED_{new_count}", latest_zip_name
                            )
                        updated_zip_path = os.path.join(
                            clean_dir_parent, updated_zip_name
                        )
                        
                        # Extract the ZIP to temporary directory
                        temp_extract_dir = tempfile.mkdtemp()
                        try:
                            with zipfile.ZipFile(original_zip_path, "r") as zip_ref:
                                zip_ref.extractall(temp_extract_dir)
                            print(f"✅ Extracted to temp directory: {temp_extract_dir}")
                            
                            # Find the mastersheet in the extracted files
                            mastersheet_path = None
                            for root, dirs, files in os.walk(temp_extract_dir):
                                for file in files:
                                    if "mastersheet" in file.lower() and file.endswith(
                                        ".xlsx"
                                    ):
                                        mastersheet_path = os.path.join(root, file)
                                        print(
                                            f"✅ Found BN mastersheet: {mastersheet_path}"
                                        )
                                        break
                                if mastersheet_path:
                                    break
                            
                            if not mastersheet_path:
                                print(f"❌ No BN mastersheet found in ZIP")
                            else:
                                print(f"✅ Found BN mastersheet: {mastersheet_path}")
                                
                                # Build updates dictionary from carryover data
                                updates = {}
                                for student in carryover_data:
                                    exam_no = student["EXAM NUMBER"]
                                    updates[exam_no] = {}
                                    for course_code, course_data in student[
                                        "RESIT_COURSES"
                                    ].items():
                                        # CRITICAL FIX: Update ALL resit scores (both passed and failed)
                                        # This ensures failed resit values are also recorded in the mastersheet
                                        updates[exam_no][course_code] = course_data[
                                            "resit_score"
                                        ]
                                
                                print(
                                    f"📊 Prepared BN updates for {len(updates)} students"
                                )
                                
                                # Update the mastersheet with full recalculation and ALL enhancements
                                # CRITICAL FIX: Use the new FINAL function with SINGLE session
                                update_success = (
                                    update_mastersheet_with_recalculation_FINAL(
                                        mastersheet_path=mastersheet_path,
                                        updates=updates,
                                        semester_key=semester_key,
                                        original_zip_path=original_zip_path,
                                        course_titles_dict=course_titles_dict,
                                        course_units_dict=credit_units_dict,
                                        set_name=set_name,
                                    )
                                )
                                
                                if update_success:
                                    # Create backup of original (only if doesn't exist)
                                    backup_zip = original_zip_path.replace(
                                        ".zip", "_BACKUP.zip"
                                    )
                                    if not os.path.exists(backup_zip):
                                        shutil.copy2(original_zip_path, backup_zip)
                                        print(f"💾 Created BN backup: {backup_zip}")
                                    
                                    print(
                                        f"📦 Creating updated BN ZIP: {updated_zip_name}"
                                    )
                                    
                                    # Create new ZIP with ALL files from temp_extract_dir (including updated mastersheet)
                                    try:
                                        with zipfile.ZipFile(
                                            updated_zip_path, "w", zipfile.ZIP_DEFLATED
                                        ) as zipf:
                                            for root, dirs, files in os.walk(
                                                temp_extract_dir
                                            ):
                                                for file in files:
                                                    file_path = os.path.join(root, file)
                                                    arcname = os.path.relpath(
                                                        file_path, temp_extract_dir
                                                    )
                                                    zipf.write(file_path, arcname)
                                                    print(
                                                        f"📁 Added to updated BN ZIP: {arcname}"
                                                    )
                                        
                                        # Verify the updated ZIP was created
                                        if (
                                            os.path.exists(updated_zip_path)
                                            and os.path.getsize(updated_zip_path) > 0
                                        ):
                                            print(
                                                f"✅ SUCCESS: Created {updated_zip_name} ({os.path.getsize(updated_zip_path)} bytes)"
                                            )
                                            # Also verify the mastersheet is in the new ZIP
                                            try:
                                                with zipfile.ZipFile(
                                                    updated_zip_path, "r"
                                                ) as test_zip:
                                                    zip_files = test_zip.namelist()
                                                    mastersheet_in_zip = any(
                                                        "mastersheet" in f.lower()
                                                        for f in zip_files
                                                    )
                                                    if mastersheet_in_zip:
                                                        print(
                                                            f"✅ Verified: BN Mastersheet is in the updated ZIP"
                                                        )
                                                    else:
                                                        print(
                                                            f"⚠️ Warning: No BN mastersheet found in updated ZIP"
                                                        )
                                            except Exception as e:
                                                print(
                                                    f"⚠️ Could not verify updated BN ZIP contents: {e}"
                                                )
                                        else:
                                            print(
                                                f"❌ ERROR: Updated BN ZIP was not created properly"
                                            )
                                    except Exception as zip_error:
                                        print(
                                            f"❌ Error creating updated BN ZIP: {zip_error}"
                                        )
                                        # Try alternative ZIP creation method
                                        try:
                                            # Fallback: use shutil.make_archive
                                            temp_zip_dir = tempfile.mkdtemp()
                                            shutil.copytree(
                                                temp_extract_dir,
                                                os.path.join(temp_zip_dir, "content"),
                                            )
                                            shutil.make_archive(
                                                updated_zip_path.replace(".zip", ""),
                                                "zip",
                                                temp_zip_dir,
                                            )
                                            shutil.rmtree(temp_zip_dir)
                                            print(
                                                f"✅ Created updated BN ZIP using fallback method"
                                            )
                                        except Exception as fallback_error:
                                            print(
                                                f"❌ Fallback BN ZIP creation also failed: {fallback_error}"
                                            )
                                    
                                    print(
                                        f"✅ BN Original preserved: {os.path.basename(original_zip_path)}"
                                    )
                                else:
                                    print(
                                        f"❌ BN Mastersheet update had some errors, but continuing"
                                    )
                                    # Even if update had errors, try to create the ZIP anyway
                                    try:
                                        with zipfile.ZipFile(
                                            updated_zip_path, "w", zipfile.ZIP_DEFLATED
                                        ) as zipf:
                                            for root, dirs, files in os.walk(
                                                temp_extract_dir
                                            ):
                                                for file in files:
                                                    file_path = os.path.join(root, file)
                                                    arcname = os.path.relpath(
                                                        file_path, temp_extract_dir
                                                    )
                                                    zipf.write(file_path, arcname)
                                                    print(
                                                        f"📁 Added to updated BN ZIP: {arcname}"
                                                    )
                                        # Verify the updated ZIP was created
                                        if (
                                            os.path.exists(updated_zip_path)
                                            and os.path.getsize(updated_zip_path) > 0
                                        ):
                                            print(
                                                f"✅ SUCCESS: Created {updated_zip_name} ({os.path.getsize(updated_zip_path)} bytes)"
                                            )
                                            # Also verify the mastersheet is in the new ZIP
                                            try:
                                                with zipfile.ZipFile(
                                                    updated_zip_path, "r"
                                                ) as test_zip:
                                                    zip_files = test_zip.namelist()
                                                    mastersheet_in_zip = any(
                                                        "mastersheet" in f.lower()
                                                        for f in zip_files
                                                    )
                                                    if mastersheet_in_zip:
                                                        print(
                                                            f"✅ Verified: BN Mastersheet is in the updated ZIP"
                                                        )
                                                    else:
                                                        print(
                                                            f"⚠️ Warning: No BN mastersheet found in updated ZIP"
                                                        )
                                            except Exception as e:
                                                print(
                                                    f"⚠️ Could not verify updated BN ZIP contents: {e}"
                                                )
                                        else:
                                            print(
                                                f"❌ ERROR: Updated BN ZIP was not created properly"
                                            )
                                    except Exception as zip_error:
                                        print(
                                            f"❌ Error creating updated BN ZIP: {zip_error}"
                                        )
                                        # Try alternative ZIP creation method
                                        try:
                                            # Fallback: use shutil.make_archive
                                            temp_zip_dir = tempfile.mkdtemp()
                                            shutil.copytree(
                                                temp_extract_dir,
                                                os.path.join(temp_zip_dir, "content"),
                                            )
                                            shutil.make_archive(
                                                updated_zip_path.replace(".zip", ""),
                                                "zip",
                                                temp_zip_dir,
                                            )
                                            shutil.rmtree(temp_zip_dir)
                                            print(
                                                f"✅ Created updated BN ZIP using fallback method"
                                            )
                                        except Exception as fallback_error:
                                            print(
                                                f"❌ Fallback BN ZIP creation also failed: {fallback_error}"
                                            )
                        except Exception as e:
                            print(f"❌ Error during BN ZIP processing: {e}")
                            traceback.print_exc()
                        finally:
                            # Clean up temp directory
                            if os.path.exists(temp_extract_dir):
                                shutil.rmtree(temp_extract_dir)
                                print(f"🧹 Cleaned up BN temp extraction directory")
                except Exception as e:
                    print(f"❌ Error updating BN mastersheet: {e}")
                    traceback.print_exc()
                
                print(f"\n{'='*60}")
                print(f"🎉 BN CARRYOVER PROCESSING COMPLETE WITH ALL ENHANCEMENTS!")
                print(f"{'='*60}")
                return True
        else:
            print(f"❌ No BN carryover data found to process")
            return False
            
    except Exception as e:
        print(f"❌ Error processing BN carryover results: {e}")
        traceback.print_exc()
        return False
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"✅ Cleaned up BN temporary files")
        
        # Report final status
        if (
            "updated_zip_path" in locals()
            and updated_zip_path
            and os.path.exists(updated_zip_path)
        ):
            print(f"✅ BN UPDATED ZIP successfully created: {updated_zip_path}")
        else:
            print(f"⚠️ BN UPDATED ZIP was not created - check logs above")

# ----------------------------
# CRITICAL FIX: Main Function
# ----------------------------

def main():
    """Main function - FIXED for web interface"""
    print("=" * 60)
    print("🎯 BN CARRYOVER RESULT PROCESSOR")
    print("🌐 WEB INTERFACE COMPATIBLE VERSION")
    print("=" * 60)
    
    # Get environment variables
    set_name = os.getenv("SELECTED_SET", "")
    semester_key = os.getenv("SELECTED_SEMESTERS", "")
    resit_file_path = os.getenv("RESIT_FILE_PATH", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", str(DEFAULT_PASS_THRESHOLD)))
    
    print(f"\n📋 PARAMETERS:")
    print(f"  Set: {set_name}")
    print(f"  Semester: {semester_key}")
    print(f"  Resit File: {resit_file_path}")
    print(f"  Pass Threshold: {pass_threshold}")
    print(f"  Base Dir: {BASE_DIR}")
    
    # Validate inputs
    if not set_name:
        print("❌ ERROR: SELECTED_SET not provided")
        sys.exit(1)
    
    if not semester_key:
        print("❌ ERROR: SELECTED_SEMESTERS not provided")
        sys.exit(1)
    
    if not resit_file_path or not os.path.exists(resit_file_path):
        print(f"❌ ERROR: Resit file not found: {resit_file_path}")
        sys.exit(1)
    
    # Validate BN set
    BN_SETS = ["SET47", "SET48"]
    if set_name not in BN_SETS:
        print(f"❌ ERROR: Invalid BN set: {set_name}")
        print(f"💡 Valid BN sets: {BN_SETS}")
        sys.exit(1)
    
    print(f"\n✅ Processing BN Set: {set_name}")
    print(f"✅ Processing Semester: {semester_key}")
    
    # Find clean directory
    clean_dir = get_output_directory(set_name)
    output_dir = clean_dir  # FIXED: Output to CLEAN_RESULTS
    
    print(f"📁 Output Directory: {output_dir}")
    
    # Find mastersheet source
    print(f"\n🔍 Looking for mastersheet source...")
    source_path, source_type = find_latest_mastersheet_source(clean_dir, set_name)
    
    if not source_path:
        print(f"❌ ERROR: No ZIP files or result folders found in {clean_dir}")
        print(f"💡 Run BN regular processor first")
        sys.exit(1)
    
    print(f"✅ Using source: {source_path}")
    print(f"✅ Source type: {source_type}")
    
    # Process carryover results
    print(f"\n🚀 Starting carryover processing...")
    success = process_carryover_results(
        resit_file_path=resit_file_path,
        source_path=source_path,
        source_type=source_type,
        semester_key=semester_key,
        set_name=set_name,
        pass_threshold=pass_threshold,
        output_dir=output_dir,
    )
    
    if success:
        print("\n" + "=" * 60)
        print("✅ BN CARRYOVER PROCESSING COMPLETED")
        print("=" * 60)
        print(f"📂 Check CLEAN_RESULTS for the ZIP file")
        print(f"💡 The file should now appear in Download Center")
        sys.exit(0)
    else:
        print("\n" + "=" * 60)
        print("❌ BN CARRYOVER PROCESSING FAILED")
        print("=" * 60)
        sys.exit(1)

# ----------------------------
# Script Execution
# ----------------------------

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Fatal error: {e}")
        traceback.print_exc()
        sys.exit(1)