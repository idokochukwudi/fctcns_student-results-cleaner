#!/usr/bin/env python3
"""
COMPLETE ENHANCED SCRIPT WITH SINGLE BACKUP AND CUMULATIVE UPDATES
WITH FIXED CGPA SUMMARY AND ANALYSIS SHEETS FOR CARRYOVER PROCESSOR
"""

import os
import sys
import re
import pandas as pd
import numpy as np
from datetime import datetime
import glob
import json
import traceback
import shutil
import zipfile
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------
# Configuration and Constants
# ----------------------------
def get_base_directory():
    """Get base directory - ENHANCED VERSION."""
    if os.getenv("BASE_DIR"):
        base_dir = os.getenv("BASE_DIR")
        if os.path.exists(base_dir):
            return base_dir
    home_dir = os.path.expanduser("~")
    default_dir = os.path.join(home_dir, "student_result_cleaner")
    if os.path.exists(os.path.join(default_dir, "EXAMS_INTERNAL")):
        return default_dir
    current_script_dir = os.path.dirname(os.path.abspath(__file__))
    if os.path.exists(os.path.join(current_script_dir, "EXAMS_INTERNAL")):
        return current_script_dir
    parent_dir = os.path.dirname(current_script_dir)
    if os.path.exists(os.path.join(parent_dir, "EXAMS_INTERNAL")):
        return parent_dir
    return default_dir


BASE_DIR = get_base_directory()
TIMESTAMP_FMT = "%d-%m-%Y_%H%M%S"
DEFAULT_PASS_THRESHOLD = 50.0
DEFAULT_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")

# Add these constants at the top with other configuration
SEMESTER_ORDER = [
    "ND-FIRST-YEAR-FIRST-SEMESTER",
    "ND-FIRST-YEAR-SECOND-SEMESTER", 
    "ND-SECOND-YEAR-FIRST-SEMESTER",
    "ND-SECOND-YEAR-SECOND-SEMESTER"
]

# Global student tracker for CGPA data
STUDENT_TRACKER = {}

def is_student_withdrawn(exam_no):
    """Check if student is withdrawn based on global tracker or sheet data."""
    # Check global tracker first
    if exam_no in STUDENT_TRACKER and STUDENT_TRACKER[exam_no].get('withdrawn', False):
        return True
    
    # Additional checks can be added here if needed
    return False


# ----------------------------
# ENHANCED DYNAMIC HEADER DETECTION FOR CARRYOVER PROCESSOR - FIXED VERSION
# ----------------------------

def detect_all_headers_robust(df, start_row=0):
    """
    FIXED: Simple header detection for resit files
    """
    detected_headers = {
        'registration_col': None,
        'name_col': None, 
        'course_cols': [],
        'score_cols': [],
        'status_col': None,
        'credit_col': None
    }
    
    # If df is a file path, read it
    if isinstance(df, str):
        df = pd.read_excel(df, header=0)  # FIXED: Always use header=0 for resit files
    
    columns = [str(col).strip() for col in df.columns]
    
    print(f"🔍 Columns found: {columns}")
    
    # 1. Find EXAM NUMBER column - SIMPLIFIED
    for col in df.columns:
        col_upper = str(col).upper().strip()
        if 'EXAM NUMBER' in col_upper or 'EXAM NO' in col_upper:
            detected_headers['registration_col'] = col
            print(f"✅ Found EXAM NUMBER column: '{col}'")
            break
    
    # 2. Find NAME column
    for col in df.columns:
        if str(col).upper().strip() == 'NAME':
            detected_headers['name_col'] = col
            print(f"✅ Found NAME column: '{col}'")
            break
    
    # 3. Find course columns - FIXED PATTERN
    import re
    course_pattern = re.compile(r'^[A-Z]{3}\d{3}$', re.IGNORECASE)
    
    for col in df.columns:
        col_str = str(col).strip()
        
        # Skip EXAM NUMBER and NAME
        if col == detected_headers['registration_col'] or col == detected_headers['name_col']:
            continue
        
        # Check if matches course pattern
        if course_pattern.match(col_str):
            detected_headers['course_cols'].append(col)
            print(f"✅ Found course column: '{col}'")
    
    print(f"✅ Total course columns: {len(detected_headers['course_cols'])}")
    
    # Validation
    if not detected_headers['registration_col']:
        print("❌ ERROR: Could not find EXAM NUMBER column!")
        print(f"Available columns: {columns}")
        return detected_headers, 0
    
    return detected_headers, 0


def validate_headers_detection(detected_headers):
    """
    Validate that all required headers are detected
    """
    validation_result = {
        'is_valid': True,
        'missing_headers': [],
        'warnings': []
    }
    
    required_headers = ['registration_col', 'course_cols']
    
    for header in required_headers:
        if not detected_headers.get(header):
            validation_result['is_valid'] = False
            validation_result['missing_headers'].append(header)
    
    if not detected_headers.get('name_col'):
        validation_result['warnings'].append("Name column not detected - student names may not be available")
    
    if not detected_headers.get('status_col'):
        validation_result['warnings'].append("Status column not detected - carryover status assignment may be limited")
    
    return validation_result


def detect_not_registered_content(score_value):
    """
    Detect if score content indicates 'NOT REGISTERED' or similar
    """
    if pd.isna(score_value):
        return True
        
    score_str = str(score_value).upper().strip()
    not_registered_indicators = [
        "NOT REG", "NOT REGISTERED", "NOT-REG", "NR", "ABSENT", 
        "MISSING", "NA", "N/A", "INCOMPLETE"
    ]
    
    return any(indicator in score_str for indicator in not_registered_indicators)


# ----------------------------
# BACKUP AND VERSION MANAGEMENT
# ----------------------------
def find_latest_updated_file(clean_dir):
    """Find the most recent UPDATED_ file in the directory."""
    print(f"🔍 Looking for latest UPDATED_ file in: {clean_dir}")
    
    if not os.path.exists(clean_dir):
        print(f"❌ Directory doesn't exist: {clean_dir}")
        return None
    
    all_files = os.listdir(clean_dir)
    updated_files = []
    
    for f in all_files:
        if f.startswith("UPDATED_") and f.endswith(".zip"):
            # Extract version number
            match = re.search(r"UPDATED_(\d+)", f)
            if match:
                version = int(match.group(1))
                updated_files.append((version, f))
    
    if not updated_files:
        print("ℹ️ No UPDATED_ files found")
        return None
    
    # Sort by version number (highest first)
    updated_files.sort(key=lambda x: x[0], reverse=True)
    latest_file = updated_files[0][1]
    latest_path = os.path.join(clean_dir, latest_file)
    
    print(f"✅ Found latest updated file: {latest_file} (version {updated_files[0][0]})")
    return latest_path


def get_next_version_number(clean_dir):
    """Determine the next version number for UPDATED_ files."""
    latest_file = find_latest_updated_file(clean_dir)
    
    if latest_file:
        # Extract version from existing UPDATED_ file
        match = re.search(r"UPDATED_(\d+)", os.path.basename(latest_file))
        if match:
            return int(match.group(1)) + 1
    
    # No UPDATED_ files found, check for original files
    all_files = os.listdir(clean_dir) if os.path.exists(clean_dir) else []
    original_files = [f for f in all_files if f.endswith(".zip") and "UPDATED_" not in f and "CARRYOVER" not in f]
    
    if original_files:
        print("ℹ️ Found original files, starting with UPDATED_1")
        return 1
    
    print("⚠️ No existing files found, starting with UPDATED_1")
    return 1


def create_backup_if_not_exists(original_zip_path):
    """Create backup only if it doesn't exist."""
    backup_path = original_zip_path.replace(".zip", "_BACKUP.zip")
    
    if os.path.exists(backup_path):
        print(f"ℹ️ Backup already exists: {backup_path}")
        return backup_path
    
    try:
        shutil.copy2(original_zip_path, backup_path)
        print(f"💾 Created backup: {backup_path}")
        return backup_path
    except Exception as e:
        print(f"⚠️ Could not create backup: {e}")
        return None


def get_mastersheet_source(clean_dir, set_name):
    """Get the appropriate mastersheet source - prefers UPDATED_ files, falls back to original."""
    print(f"\n🔍 DETERMINING MASTERSHEET SOURCE")
    
    # First, look for UPDATED_ files
    latest_updated = find_latest_updated_file(clean_dir)
    if latest_updated:
        print(f"✅ Using latest updated file: {latest_updated}")
        return latest_updated, "zip"
    
    # Fallback to original files
    print("ℹ️ No UPDATED_ files found, looking for original files...")
    original_source, source_type = find_latest_mastersheet_source(clean_dir, set_name)
    
    if original_source:
        print(f"✅ Using original file: {original_source}")
        return original_source, source_type
    
    print("❌ No suitable mastersheet source found")
    return None, None


# ----------------------------
# WITHDRAWN STUDENT MANAGEMENT
# ----------------------------
def remove_withdrawn_from_semester_sheets(wb, semester_key):
    """Remove withdrawn students only from the processed semester sheet."""
    print(f"\n👥 PROCESSING WITHDRAWN STUDENTS FOR {semester_key}")
    
    # Find the current semester sheet
    current_sheet = None
    for sheet in wb.sheetnames:
        if semester_key.upper() in sheet.upper():
            current_sheet = sheet
            break
    
    if not current_sheet:
        print(f"❌ No sheet found for semester: {semester_key}")
        return
    
    ws = wb[current_sheet]
    header_row, headers = find_sheet_structure(ws)
    
    if not header_row:
        print("❌ Could not find header structure")
        return
    
    # Find remarks and exam number columns
    remarks_col = None
    exam_col = None
    
    for header, col_idx in headers.items():
        if "REMARKS" in header.upper():
            remarks_col = col_idx
        if "EXAM NUMBER" in header.upper():
            exam_col = col_idx
    
    if not remarks_col or not exam_col:
        print("❌ Could not find required columns")
        return
    
    # Identify withdrawn students
    withdrawn_students = []
    rows_to_delete = []
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        exam_no = ws.cell(row=row_idx, column=exam_col).value
        remarks = ws.cell(row=row_idx, column=remarks_col).value
        
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break
        
        if remarks and "WITHDRAW" in str(remarks).upper():
            withdrawn_students.append(str(exam_no).strip())
            rows_to_delete.append(row_idx)
    
    print(f"📊 Found {len(withdrawn_students)} withdrawn students: {withdrawn_students}")
    
    # Delete rows from bottom to top to maintain indices
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
        print(f"🗑️ Removed withdrawn student from row {row_idx}")
    
    # Update summary section after removal
    course_columns = identify_course_columns_properly(headers)
    update_summary_section_fixed(ws, headers, header_row, course_columns)
    
    print(f"✅ Removed {len(withdrawn_students)} withdrawn students from {current_sheet}")


def preserve_withdrawn_in_summary_sheets(wb, withdrawn_students):
    """Ensure withdrawn students remain in CGPA_SUMMARY and ANALYSIS sheets."""
    print(f"\n📊 PRESERVING WITHDRAWN STUDENTS IN SUMMARY SHEETS")
    
    # Update CGPA_SUMMARY if it exists
    if "CGPA_SUMMARY" in wb.sheetnames:
        update_cgpa_summary_with_withdrawn(wb, withdrawn_students)
        print("✅ Updated CGPA_SUMMARY with withdrawn students")
    
    # Update ANALYSIS if it exists  
    if "ANALYSIS" in wb.sheetnames:
        print("✅ ANALYSIS sheet maintains withdrawn student records")
    
    print(f"📈 Withdrawn students preserved in summary sheets: {len(withdrawn_students)} students")


# ----------------------------
# Utility Functions
# ----------------------------
def sanitize_filename(filename):
    """Remove or replace characters that are not safe for filenames."""
    return re.sub(r"[^\w\-_.]", "_", filename)


def find_exam_number_column(df):
    """Find the exam number column in a DataFrame."""
    possible_names = [
        "EXAM NUMBER",
        "REG. No",
        "REG NO",
        "REGISTRATION NUMBER",
        "MAT NO",
        "STUDENT ID",
    ]
    for col in df.columns:
        col_upper = str(col).upper()
        for possible_name in possible_names:
            if possible_name in col_upper:
                return col
    return None


def standardize_semester_key(semester_key):
    """Standardize semester key to canonical format - ND ONLY."""
    if not semester_key:
        return None
    key_upper = semester_key.upper()
    # ND-ONLY canonical mappings
    canonical_mappings = {
        # ND First Year First Semester variants
        ("FIRST", "YEAR", "FIRST", "SEMESTER"): "ND-FIRST-YEAR-FIRST-SEMESTER",
        ("1ST", "YEAR", "1ST", "SEMESTER"): "ND-FIRST-YEAR-FIRST-SEMESTER",
        ("YEAR", "1", "SEMESTER", "1"): "ND-FIRST-YEAR-FIRST-SEMESTER",
        # ND First Year Second Semester variants
        ("FIRST", "YEAR", "SECOND", "SEMESTER"): "ND-FIRST-YEAR-SECOND-SEMESTER",
        ("1ST", "YEAR", "2ND", "SEMESTER"): "ND-FIRST-YEAR-SECOND-SEMESTER",
        ("YEAR", "1", "SEMESTER", "2"): "ND-FIRST-YEAR-SECOND-SEMESTER",
        # ND Second Year First Semester variants
        ("SECOND", "YEAR", "FIRST", "SEMESTER"): "ND-SECOND-YEAR-FIRST-SEMESTER",
        ("2ND", "YEAR", "1ST", "SEMESTER"): "ND-SECOND-YEAR-FIRST-SEMESTER",
        ("YEAR", "2", "SEMESTER", "1"): "ND-SECOND-YEAR-FIRST-SEMESTER",
        # ND Second Year Second Semester variants
        ("SECOND", "YEAR", "SECOND", "SEMESTER"): "ND-SECOND-YEAR-SECOND-SEMESTER",
        ("2ND", "YEAR", "2ND", "SEMESTER"): "ND-SECOND-YEAR-SECOND-SEMESTER",
        ("YEAR", "2", "SEMESTER", "2"): "ND-SECOND-YEAR-SECOND-SEMESTER",
    }
    # ND-only patterns
    patterns = [
        r"(FIRST|1ST|YEAR.?1).*?(FIRST|1ST|SEMESTER.?1)",
        r"(FIRST|1ST|YEAR.?1).*?(SECOND|2ND|SEMESTER.?2)",
        r"(SECOND|2ND|YEAR.?2).*?(FIRST|1ST|SEMESTER.?1)",
        r"(SECOND|2ND|YEAR.?2).*?(SECOND|2ND|SEMESTER.?2)",
    ]
    for pattern_idx, pattern in enumerate(patterns):
        if re.search(pattern, key_upper):
            if pattern_idx == 0:
                return "ND-FIRST-YEAR-FIRST-SEMESTER"
            elif pattern_idx == 1:
                return "ND-FIRST-YEAR-SECOND-SEMESTER"
            elif pattern_idx == 2:
                return "ND-SECOND-YEAR-FIRST-SEMESTER"
            elif pattern_idx == 3:
                return "ND-SECOND-YEAR-SECOND-SEMESTER"
    # If no match, return original
    print(f"Could not standardize semester key: {semester_key}")
    return semester_key


def standardize_semester_name(semester_name):
    """Standardize semester name - alias for standardize_semester_key for compatibility."""
    return standardize_semester_key(semester_name)


def extract_semester_from_filename(filename):
    """Extract semester from carryover filename - FIXED VERSION"""
    try:
        # Handle both .json and .xlsx files
        if filename.endswith(".json") or filename.endswith(".xlsx"):
            # Pattern: co_student_ND-2024_ND-SECOND-YEAR-FIRST-SEMESTER_20251107_100522.json
            # Extract the part between the second 'ND-' and the timestamp
            pattern = r"co_student_ND-\d+_(ND-.*?)_\d+_\d+\.(json|xlsx)"
            match = re.search(pattern, filename)
            if match:
                semester = match.group(1)
                return semester.upper().replace("-", " ").replace("_", " ")

        # Fallback: try to extract any ND- pattern
        match = re.search(
            r"(ND-[A-Za-z-]+(?:YEAR|SEMESTER)[A-Za-z-]*)", filename, re.IGNORECASE
        )
        if match:
            return match.group(1).upper().replace("-", " ").replace("_", " ")

        return None
    except Exception as e:
        print(f"Error extracting semester from {filename}: {e}")
        return None


def get_semester_display_info(semester_key):
    """Get display information for ND semester key ONLY."""
    semester_lower = semester_key.lower()
    # ND ONLY
    if "first-year-first-semester" in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI", "Semester 1"
    elif "first-year-second-semester" in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "NDI", "Semester 2"
    elif "second-year-first-semester" in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "NDII", "Semester 3"
    elif "second-year-second-semester" in semester_lower:
        return 2, 2, "YEAR TWO", "SECOND SEMESTER", "NDII", "Semester 4"
    else:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI", "Semester 1"


def get_grade_point(score):
    """Determine grade point based on score - UPDATED TO 4.0 SCALE (A=4, B=3, C=2, D=1, E=0)"""
    try:
        score = float(score)
        if score >= 70:    # A: 70-100
            return 4.0
        elif score >= 60:  # B: 60-69
            return 3.0
        elif score >= 50:  # C: 50-59
            return 2.0
        elif score >= 45:  # D: 45-49
            return 1.0
        else:              # E/F: 0-44
            return 0.0
    except (ValueError, TypeError):
        return 0.0


def get_previous_semester(semester_key):
    """Get the previous semester key for ND carryover ONLY."""
    standardized = standardize_semester_key(semester_key)
    # ND semesters ONLY
    if standardized == "ND-FIRST-YEAR-SECOND-SEMESTER":
        return "ND-FIRST-YEAR-FIRST-SEMESTER"
    elif standardized == "ND-SECOND-YEAR-FIRST-SEMESTER":
        return "ND-FIRST-YEAR-SECOND-SEMESTER"
    elif standardized == "ND-SECOND-YEAR-SECOND-SEMESTER":
        return "ND-SECOND-YEAR-FIRST-SEMESTER"
    else:
        return None  # No previous for first semester


def get_previous_semesters_for_display(current_semester_key):
    """Get list of previous semesters for ND GPA display in mastersheet."""
    current_standard = standardize_semester_key(current_semester_key)
    # ND ONLY
    semester_mapping = {
        "ND-FIRST-YEAR-FIRST-SEMESTER": [],
        "ND-FIRST-YEAR-SECOND-SEMESTER": ["Semester 1"],
        "ND-SECOND-YEAR-FIRST-SEMESTER": ["Semester 1", "Semester 2"],
        "ND-SECOND-YEAR-SECOND-SEMESTER": ["Semester 1", "Semester 2", "Semester 3"],
    }
    return semester_mapping.get(current_standard, [])


def generate_remarks(resit_courses):
    """Generate remarks for resit performance."""
    passed_count = sum(
        1
        for course_data in resit_courses.values()
        if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
    )
    total_count = len(resit_courses)
    if passed_count == total_count:
        return "All courses passed in resit"
    elif passed_count > 0:
        return f"{passed_count}/{total_count} courses passed in resit"
    else:
        return "No improvement in resit"


def extract_class_from_set_name(set_name):
    """Extract class name from set_name (e.g., 'ND-2024' from 'ND-2024')"""
    # set_name is already in format like "ND-2024"
    return set_name


# ----------------------------
# Course Data Management
# ----------------------------
def load_course_data():
    """Load ND course data ONLY."""
    return load_nd_course_data()


def load_nd_course_data():
    """Load ND course data from course-code-creditUnit.xlsx."""
    possible_course_files = [
        os.path.join(
            BASE_DIR,
            "EXAMS_INTERNAL",
            "ND",
            "ND-COURSES",
            "course-code-creditUnit.xlsx",
        ),
        os.path.join(BASE_DIR, "ND", "ND-COURSES", "course-code-creditUnit.xlsx"),
        os.path.join(
            BASE_DIR, "EXAMS_INTERNAL", "ND-COURSES", "course-code-creditUnit.xlsx"
        ),
        os.path.join(BASE_DIR, "course-code-creditUnit.xlsx"),
    ]
    course_file = None
    for possible_file in possible_course_files:
        if os.path.exists(possible_file):
            course_file = possible_file
            print(f"✅ Found ND course file: {possible_file}")
            break
    if not course_file:
        print(f"❌ Main ND course file not found in standard locations")
        alternative_files = find_alternative_course_files()
        if alternative_files:
            course_file = alternative_files[0]
            print(f"🔄 Using alternative ND course file: {course_file}")
        else:
            print(f"❌ No ND course files found anywhere!")
            return {}, {}, {}, {}
    print(f"📚 Loading ND course data from: {course_file}")
    return _load_course_data_from_file(course_file)


def _load_course_data_from_file(course_file):
    """Generic function to load course data from Excel file - FIXED VERSION."""
    try:
        xl = pd.ExcelFile(course_file)
        semester_course_titles = {}
        semester_credit_units = {}
        course_code_to_title = {}
        course_code_to_unit = {}

        print(f"📖 Available sheets: {xl.sheet_names}")  # FIXED: sheet_names instead of sheetnames

        for sheet in xl.sheet_names:  # FIXED: sheet_names instead of sheetnames
            sheet_standard = standardize_semester_key(sheet)
            print(f"📖 Reading sheet: {sheet} (standardized: {sheet_standard})")
            try:
                df = pd.read_excel(
                    course_file, sheet_name=sheet, engine="openpyxl", header=0
                )

                # Convert columns to string and clean
                df.columns = [str(c).strip().upper() for c in df.columns]

                # Look for course code, title, and credit unit columns with flexible matching
                code_col = None
                title_col = None
                unit_col = None

                for col in df.columns:
                    col_clean = str(col).upper()
                    if any(
                        keyword in col_clean
                        for keyword in ["COURSE CODE", "CODE", "COURSECODE"]
                    ):
                        code_col = col
                    elif any(
                        keyword in col_clean
                        for keyword in ["COURSE TITLE", "TITLE", "COURSENAME"]
                    ):
                        title_col = col
                    elif any(
                        keyword in col_clean
                        for keyword in ["CU", "CREDIT", "UNIT", "CREDIT UNIT"]
                    ):
                        unit_col = col

                print(
                    f"🔍 Detected columns - Code: {code_col}, Title: {title_col}, Unit: {unit_col}"
                )

                if not all([code_col, title_col, unit_col]):
                    print(
                        f"⚠️ Sheet '{sheet}' missing required columns - found: code={code_col}, title={title_col}, unit={unit_col}"
                    )
                    # Try to use first three columns as fallback
                    if len(df.columns) >= 3:
                        code_col, title_col, unit_col = (
                            df.columns[0],
                            df.columns[1],
                            df.columns[2],
                        )
                        print(
                            f"🔄 Using fallback columns: {code_col}, {title_col}, {unit_col}"
                        )
                    else:
                        print(
                            f"❌ Sheet '{sheet}' doesn't have enough columns - skipped"
                        )
                        continue

                # Clean the data
                df_clean = df.dropna(subset=[code_col]).copy()
                if df_clean.empty:
                    print(f"⚠️ Sheet '{sheet}' has no data after cleaning - skipped")
                    continue

                # Convert credit units to numeric, handling errors
                df_clean[unit_col] = pd.to_numeric(df_clean[unit_col], errors="coerce")
                df_clean = df_clean.dropna(subset=[unit_col])

                # FIXED: Remove rows with "TOTAL" in course code
                df_clean = df_clean[
                    ~df_clean[code_col]
                    .astype(str)
                    .str.contains("TOTAL", case=False, na=False)
                ]

                if df_clean.empty:
                    print(f"⚠️ Sheet '{sheet}' has no valid rows after cleaning - skipped")
                    continue

                codes = df_clean[code_col].astype(str).str.strip().tolist()
                titles = df_clean[title_col].astype(str).str.strip().tolist()
                units = df_clean[unit_col].astype(float).tolist()
                print(f"📋 Found {len(codes)} courses in {sheet}:")
                for i, (code, title, unit) in enumerate(
                    zip(codes[:5], titles[:5], units[:5])
                ):
                    print(f" - '{code}': '{title}' (CU: {unit})")
                # Create mapping dictionaries with ENHANCED normalization strategies
                sheet_titles = {}
                sheet_units = {}

                for code, title, unit in zip(codes, titles, units):
                    if not code or code.upper() in ["NAN", "NONE", ""]:
                        continue

                    # ENHANCED: Create comprehensive normalization variants for robust matching
                    variants = [
                        # Basic variants
                        code.upper().strip(),
                        code.strip(),
                        code.upper(),
                        code.lower(),
                        code.title(),
                        # Space removal variants
                        code.upper().replace(" ", ""),
                        code.replace(" ", ""),
                        re.sub(r"\s+", "", code.upper()),
                        re.sub(r"\s+", "", code),
                        # Special character handling
                        re.sub(r"[^a-zA-Z0-9]", "", code.upper()),
                        re.sub(r"[^a-zA-Z0-9]", "", code),
                        # Common formatting issues
                        code.upper().replace("-", ""),
                        code.upper().replace("_", ""),
                        code.replace("-", "").replace("_", ""),
                        code.upper().replace("-", "").replace("_", "").replace(" ", ""),
                        # WITH common prefixes (for matching with prefix)
                        f"NUR{code.upper()}",
                        f"NUR{code.upper().replace(' ', '')}",
                        f"NUR{re.sub(r'[^a-zA-Z0-9]', '', code.upper())}",
                        f"NSC{code.upper()}",
                        f"NSC{code.upper().replace(' ', '')}",
                        f"NSC{re.sub(r'[^a-zA-Z0-9]', '', code.upper())}",
                        # WITHOUT common prefixes (for matching without prefix)
                        code.upper().replace("NUR", "").strip(),
                        code.upper().replace("NSC", "").strip(),
                        re.sub(r"^(NUR|NSC)", "", code.upper()).strip(),
                        re.sub(r"^(NUR|NSC)", "", code.upper())
                        .replace(" ", "")
                        .strip(),
                        # Number-focused variants (for codes like "101", "201")
                        re.sub(r"[^0-9]", "", code),
                        # Common variations with dots
                        code.upper().replace(".", ""),
                        code.replace(".", ""),
                    ]

                    # Remove duplicates while preserving order
                    variants = list(
                        dict.fromkeys(
                            [v for v in variants if v and v not in ["NAN", "NONE", ""]]
                        )
                    )

                    # Add all variants to mappings
                    for variant in variants:
                        sheet_titles[variant] = title
                        sheet_units[variant] = unit
                        course_code_to_title[variant] = title
                        course_code_to_unit[variant] = unit

                semester_course_titles[sheet_standard] = sheet_titles
                semester_credit_units[sheet_standard] = sheet_units

            except Exception as e:
                print(f"❌ Error processing sheet '{sheet}': {e}")
                traceback.print_exc()
                continue

        print(
            f"✅ Loaded course data for sheets: {list(semester_course_titles.keys())}"
        )
        print(f"📊 Total course mappings: {len(course_code_to_title)}")

        # Debug: Show some course mappings
        print("🔍 Sample course mappings:")
        sample_items = list(course_code_to_title.items())[:15]
        for code, title in sample_items:
            unit = course_code_to_unit.get(code, 0)
            print(f" '{code}' -> '{title}' (CU: {unit})")

        return (
            semester_course_titles,
            semester_credit_units,
            course_code_to_title,
            course_code_to_unit,
        )

    except Exception as e:
        print(f"❌ Error loading course data: {e}")
        traceback.print_exc()
        return {}, {}, {}, {}


def find_alternative_course_files():
    """Look for alternative course files for ND."""
    base_dirs = [
        os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", "ND-COURSES"),
        os.path.join(BASE_DIR, "ND", "ND-COURSES"),
        os.path.join(BASE_DIR, "COURSES"),
        os.path.join(BASE_DIR, "EXAMS_INTERNAL"),
    ]
    course_files = []
    for base_dir in base_dirs:
        if os.path.exists(base_dir):
            for file in os.listdir(base_dir):
                if "course" in file.lower() and file.endswith((".xlsx", ".xls")):
                    full_path = os.path.join(base_dir, file)
                    course_files.append(full_path)
    return course_files


def debug_course_matching(resit_file_path, course_code_to_title, course_code_to_unit):
    """Debug function to check why course codes aren't matching - ENHANCED."""
    print(f"\n🔍 DEBUGGING ND COURSE MATCHING")
    print("=" * 50)
    # Read resit file to see what course codes we have
    resit_df = pd.read_excel(resit_file_path, header=0)
    resit_exam_col = find_exam_number_column(resit_df)
    # Get all course codes from resit file
    resit_courses = []
    for col in resit_df.columns:
        if col != resit_exam_col and col != "NAME" and not "Unnamed" in str(col):
            resit_courses.append(col)
    print(f"📋 ND Course codes from resit file: {resit_courses}")
    print(f"📊 Total courses in ND resit file: {len(resit_courses)}")
    # Check each resit course against course file
    for course in resit_courses:
        print(f"\n🔍 Checking ND course: '{course}'")
        original_code = str(course).strip()

        # Generate ENHANCED variants for matching
        variants = [
            original_code.upper().strip(),
            original_code.strip(),
            original_code.upper(),
            original_code,
            original_code.lower(),
            original_code.title(),
            original_code.upper().replace(" ", ""),
            original_code.replace(" ", ""),
            re.sub(r"\s+", "", original_code.upper()),
            re.sub(r"\s+", "", original_code),
            re.sub(r"[^a-zA-Z0-9]", "", original_code.upper()),
            re.sub(r"[^a-zA-Z0-9]", "", original_code),
            original_code.upper().replace("-", ""),
            original_code.upper().replace("_", ""),
            original_code.replace("-", "").replace("_", ""),
            original_code.upper().replace("-", "").replace("_", "").replace(" ", ""),
            f"NUR{original_code.upper()}",
            f"NUR{original_code.upper().replace(' ', '')}",
            f"NSC{original_code.upper()}",
            f"NSC{original_code.upper().replace(' ', '')}",
            original_code.upper().replace("NUR", "").strip(),
            original_code.upper().replace("NSC", "").strip(),
            re.sub(r"^(NUR|NSC)", "", original_code.upper()).strip(),
            re.sub(r"^(NUR|NSC)", "", original_code.upper()).replace(" ", "").strip(),
            re.sub(r"[^0-9]", "", original_code),
            original_code.upper().replace(".", ""),
            original_code.replace(".", ""),
        ]

        # Remove duplicates
        variants = list(dict.fromkeys([v for v in variants if v and v != "NAN"]))

        print(f" Generated {len(variants)} variants to try")

        found = False
        for variant in variants:
            if variant in course_code_to_title:
                title = course_code_to_title[variant]
                unit = course_code_to_unit.get(variant, 0)
                print(f" ✅ FOUND: '{variant}' -> '{title}' (CU: {unit})")
                found = True
                break

        if not found:
            print(f" ❌ NOT FOUND: No match for '{course}'")
            # Show some similar keys from course file
            similar_keys = []
            # Check for partial matches
            for key in list(course_code_to_title.keys())[:50]:  # Check first 50 keys
                if any(
                    part in key.upper()
                    for part in original_code.upper().split()
                    if len(part) > 2
                ):
                    similar_keys.append(key)

            if similar_keys:
                print(f" 💡 Similar keys found: {similar_keys[:5]}")
            else:
                print(
                    f" 💡 Sample available keys: {list(course_code_to_title.keys())[:10]}"
                )


def find_course_title(course_code, course_titles_dict, course_code_to_title):
    """Robust function to find course title with comprehensive matching strategies - ENHANCED."""
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return str(course_code) if course_code else "Unknown Course"
    original_code = str(course_code).strip()
    # Generate ENHANCED comprehensive matching variants
    variants = [
        # Basic normalizations
        original_code.upper().strip(),
        original_code.strip(),
        original_code.upper(),
        original_code,
        original_code.lower(),
        original_code.title(),
        # Space handling variations
        original_code.upper().replace(" ", ""),
        original_code.replace(" ", ""),
        re.sub(r"\s+", "", original_code.upper()),
        re.sub(r"\s+", "", original_code),
        # Special character handling
        re.sub(r"[^a-zA-Z0-9]", "", original_code.upper()),
        re.sub(r"[^a-zA-Z0-9]", "", original_code),
        # Common formatting issues
        original_code.upper().replace("-", ""),
        original_code.upper().replace("_", ""),
        original_code.replace("-", "").replace("_", ""),
        original_code.upper().replace("-", "").replace("_", "").replace(" ", ""),
        # WITH common prefixes
        f"NUR{original_code.upper()}",
        f"NUR{original_code.upper().replace(' ', '')}",
        f"NUR{re.sub(r'[^a-zA-Z0-9]', '', original_code.upper())}",
        f"NSC{original_code.upper()}",
        f"NSC{original_code.upper().replace(' ', '')}",
        f"NSC{re.sub(r'[^a-zA-Z0-9]', '', original_code.upper())}",
        # WITHOUT common prefixes
        original_code.upper().replace("NUR", "").strip(),
        original_code.upper().replace("NSC", "").strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).replace(" ", "").strip(),
        # Number-focused variants
        re.sub(r"[^0-9]", "", original_code),
        # Dot removal
        original_code.upper().replace(".", ""),
        original_code.replace(".", ""),
    ]
    # Remove duplicates
    variants = list(dict.fromkeys([v for v in variants if v and v != "NAN"]))
    # Try each strategy in order
    for variant in variants:
        # Try course_titles_dict first (semester-specific)
        if variant in course_titles_dict:
            title = course_titles_dict[variant]
            print(
                f"✅ Found title for '{original_code}' using variant '{variant}': '{title}'"
            )
            return title

        # Try global course_code_to_title
        if variant in course_code_to_title:
            title = course_code_to_title[variant]
            print(
                f"✅ Found title for '{original_code}' using global variant '{variant}': '{title}'"
            )
            return title
    # If no match found, log and return descriptive original code
    print(f"⚠️ Could not find course title for: '{original_code}'")
    print(f" Tried {len(variants)} variants without success")
    return f"{original_code} (Title Not Found)"


def find_credit_unit(course_code, credit_units_dict, course_code_to_unit):
    """Robust function to find credit unit with comprehensive matching strategies - ENHANCED."""
    if not course_code or str(course_code).upper() in ["NAN", "NONE", ""]:
        return 0
    original_code = str(course_code).strip()
    # Generate the same ENHANCED variants as title matching
    variants = [
        original_code.upper().strip(),
        original_code.strip(),
        original_code.upper(),
        original_code,
        original_code.lower(),
        original_code.title(),
        original_code.upper().replace(" ", ""),
        original_code.replace(" ", ""),
        re.sub(r"\s+", "", original_code.upper()),
        re.sub(r"\s+", "", original_code),
        re.sub(r"[^a-zA-Z0-9]", "", original_code.upper()),
        re.sub(r"[^a-zA-Z0-9]", "", original_code),
        original_code.upper().replace("-", ""),
        original_code.upper().replace("_", ""),
        original_code.replace("-", "").replace("_", ""),
        original_code.upper().replace("-", "").replace("_", "").replace(" ", ""),
        f"NUR{original_code.upper()}",
        f"NUR{original_code.upper().replace(' ', '')}",
        f"NUR{re.sub(r'[^a-zA-Z0-9]', '', original_code.upper())}",
        f"NSC{original_code.upper()}",
        f"NSC{original_code.upper().replace(' ', '')}",
        f"NSC{re.sub(r'[^a-zA-Z0-9]', '', original_code.upper())}",
        original_code.upper().replace("NUR", "").strip(),
        original_code.upper().replace("NSC", "").strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).strip(),
        re.sub(r"^(NUR|NSC)", "", original_code.upper()).replace(" ", "").strip(),
        re.sub(r"[^0-9]", "", original_code),
        original_code.upper().replace(".", ""),
        original_code.replace(".", ""),
    ]
    # Remove duplicates
    variants = list(dict.fromkeys([v for v in variants if v and v != "NAN"]))
    # Try each strategy
    for variant in variants:
        if variant in credit_units_dict:
            unit = credit_units_dict[variant]
            return unit

        if variant in course_code_to_unit:
            unit = course_code_to_unit[variant]
            return unit
    print(f"⚠️ Could not find credit unit for: '{original_code}', defaulting to 2")
    return 2  # Default credit unit


# ----------------------------
# File and ZIP Handling
# ----------------------------
def extract_mastersheet_from_zip(zip_path, semester_key):
    """Extract mastersheet from ZIP file and return temporary file path."""
    try:
        print(f"📦 Looking for mastersheet in ZIP: {zip_path}")
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            all_files = zip_ref.namelist()
            print(f"📁 Files in ZIP: {all_files}")

            mastersheet_files = [
                f
                for f in all_files
                if "mastersheet" in f.lower() and f.endswith(".xlsx")
            ]

            if not mastersheet_files:
                print(f"❌ No mastersheet found in ZIP")
                return None, None

            mastersheet_name = mastersheet_files[0]
            print(f"✅ Found mastersheet: {mastersheet_name}")

            temp_dir = tempfile.mkdtemp()
            temp_mastersheet_path = os.path.join(
                temp_dir, f"mastersheet_{semester_key}.xlsx"
            )

            with open(temp_mastersheet_path, "wb") as f:
                f.write(zip_ref.read(mastersheet_name))

            print(f"✅ Extracted mastersheet to: {temp_mastersheet_path}")
            return temp_mastersheet_path, temp_dir

    except Exception as e:
        print(f"❌ Error extracting mastersheet from ZIP: {e}")
        traceback.print_exc()
        return None, None


def find_latest_zip_file(clean_dir):
    """Find the latest ZIP file in clean results directory."""
    print(f"🔍 Looking for ND ZIP files in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ ND clean directory doesn't exist: {clean_dir}")
        return None
    all_files = os.listdir(clean_dir)
    zip_files = []
    for f in all_files:
        if f.lower().endswith(".zip"):
            if "carryover" in f.lower():
                print(f"⚠️ Skipping ND carryover ZIP: {f}")
                continue

            if any(pattern in f for pattern in ["_RESULT-", "RESULT_", "RESULT-"]):
                zip_files.append(f)
                print(f"✅ Found ND regular results ZIP: {f}")
            else:
                print(f"ℹ️ Found other ND ZIP (not a result file): {f}")
    if not zip_files:
        print(f"❌ No ND regular results ZIP files found (excluding carryover files)")
        fallback_zips = [
            f
            for f in all_files
            if f.lower().endswith(".zip") and "carryover" not in f.lower()
        ]
        if fallback_zips:
            print(f"⚠️ Using fallback ND ZIP files: {fallback_zips}")
            zip_files = fallback_zips
        else:
            print(f"❌ No ND ZIP files found at all in {clean_dir}")
            return None
    print(f"✅ Final ND ZIP files to consider: {zip_files}")
    zip_files_with_path = [os.path.join(clean_dir, f) for f in zip_files]
    latest_zip = sorted(zip_files_with_path, key=os.path.getmtime, reverse=True)[0]
    print(f"🎯 Using latest ND ZIP: {latest_zip}")
    return latest_zip


def find_latest_result_folder(clean_dir, set_name):
    """Find the latest result folder in clean results directory."""
    print(f"🔍 Looking for ND result folders in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ ND clean directory doesn't exist: {clean_dir}")
        return None
    all_items = os.listdir(clean_dir)
    result_folders = [
        f
        for f in all_items
        if os.path.isdir(os.path.join(clean_dir, f))
        and f.startswith(f"{set_name}_RESULT-")
    ]
    if not result_folders:
        print(f"❌ No ND result folders found")
        return None
    print(f"✅ Found ND result folders: {result_folders}")
    folders_with_path = [os.path.join(clean_dir, f) for f in result_folders]
    latest_folder = sorted(folders_with_path, key=os.path.getmtime, reverse=True)[0]
    print(f"🎯 Using latest ND result folder: {latest_folder}")
    return latest_folder


def find_latest_mastersheet_source(clean_dir, set_name):
    """Find the latest source for mastersheet: prefer ZIP, fallback to folder."""
    print(f"🔍 Looking for ND mastersheet source in: {clean_dir}")
    if not os.path.exists(clean_dir):
        print(f"❌ ND clean directory doesn't exist: {clean_dir}")
        return None, None
    zip_path = find_latest_zip_file(clean_dir)
    if zip_path:
        print(f"✅ Using ND ZIP source: {zip_path}")
        try:
            with zipfile.ZipFile(zip_path, "r") as zip_ref:
                zip_files = zip_ref.namelist()
                mastersheet_files = [
                    f
                    for f in zip_files
                    if "mastersheet" in f.lower() and f.endswith(".xlsx")
                ]
                if mastersheet_files:
                    print(f"✅ ND ZIP contains mastersheet files: {mastersheet_files}")
                    return zip_path, "zip"
                else:
                    print(f"⚠️ ND ZIP found but no mastersheet inside: {zip_path}")
        except Exception as e:
            print(f"⚠️ Error checking ND ZIP contents: {e}")
    folder_path = find_latest_result_folder(clean_dir, set_name)
    if folder_path:
        print(f"✅ Using ND folder source: {folder_path}")
        return folder_path, "folder"
    print(f"❌ No valid ND ZIP files or result folders found in {clean_dir}")
    return None, None


def get_mastersheet_path(source_path, source_type, semester_key):
    """Get mastersheet path based on source type (zip or folder)."""
    temp_dir = None
    if source_type == "zip":
        temp_mastersheet_path, temp_dir = extract_mastersheet_from_zip(
            source_path, semester_key
        )
        if not temp_mastersheet_path:
            print("❌ Failed to extract mastersheet from ZIP")
            return None, None
    elif source_type == "folder":
        all_files = os.listdir(source_path)
        mastersheet_files = [
            f for f in all_files if "mastersheet" in f.lower() and f.endswith(".xlsx")
        ]

        if not mastersheet_files:
            print(f"❌ No mastersheet found in folder {source_path}")
            return None, None

        mastersheet_name = mastersheet_files[0]
        temp_mastersheet_path = os.path.join(source_path, mastersheet_name)
        print(f"✅ Found mastersheet in folder: {temp_mastersheet_path}")
    else:
        return None, None
    return temp_mastersheet_path, temp_dir


def get_matching_sheet(xl, target_key):
    """Find matching sheet name with variants - FIXED VERSION."""
    target_upper = (
        target_key.upper().replace("-", " ").replace("_", " ").replace(".", " ")
    )
    target_upper = " ".join(target_upper.split())
    possible_keys = [
        target_key,
        target_key.upper(),
        target_key.lower(),
        target_key.title(),
        target_key.replace("-", " ").upper(),
        target_key.replace("-", " ").lower(),
        target_key.replace("-", " ").title(),
        target_key.replace("-", "_").upper(),
        target_key.replace("-", "_").lower(),
        target_key.replace("-", "_").title(),
        target_key.replace("First", "1st"),
        target_key.replace("Second", "2nd"),
        target_key.replace("Third", "3rd"),
        target_key.replace("YEAR", "YR"),
        target_key.replace("SEMESTER", "SEM"),
        target_upper,
        target_upper.replace("FIRST", "1ST"),
        target_upper.replace("SECOND", "2ND"),
        target_upper.replace("THIRD", "3RD"),
        target_upper.replace("YEAR", "YR"),
        target_upper.replace("SEMESTER", "SEM"),
    ]
    possible_keys = list(set([k for k in possible_keys if k]))
    print(f"🔍 Trying sheet variants for '{target_key}': {possible_keys}")
    
    # FIX: Use sheet_names instead of sheetnames for pandas ExcelFile
    for sheet in xl.sheet_names:  # FIXED: sheet_names instead of sheetnames
        sheet_normalized = (
            sheet.upper().replace("-", " ").replace("_", " ").replace(".", " ")
        )
        sheet_normalized = " ".join(sheet_normalized.split())

        if any(
            p == sheet or p in sheet or p == sheet_normalized or p in sheet_normalized
            for p in possible_keys
        ):
            print(f"✅ Found matching sheet: '{sheet}' for '{target_key}'")
            return sheet
    print(f"❌ No matching sheet found for '{target_key}'")
    print(f"📖 Available sheets: {xl.sheet_names}")  # FIXED: sheet_names instead of sheetnames
    return None


def create_carryover_zip(source_dir, zip_path):
    """Create ZIP file of carryover results."""
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        print(f"✅ ZIP file created: {zip_path}")
        return True
    except Exception as e:
        print(f"❌ Error creating ZIP: {e}")
        return False


# ----------------------------
# GPA/CGPA Management - FIXED VERSION
# ----------------------------
def load_previous_gpas_enhanced(mastersheet_path, current_semester_key):
    """Enhanced function to load previous GPA data with better sheet detection."""
    all_student_data = {}
    current_standard = standardize_semester_key(current_semester_key)
    
    # ND semesters ONLY
    all_semesters = {
        "ND-FIRST-YEAR-FIRST-SEMESTER": [],
        "ND-FIRST-YEAR-SECOND-SEMESTER": ["ND-FIRST-YEAR-FIRST-SEMESTER"],
        "ND-SECOND-YEAR-FIRST-SEMESTER": [
            "ND-FIRST-YEAR-FIRST-SEMESTER",
            "ND-FIRST-YEAR-SECOND-SEMESTER",
        ],
        "ND-SECOND-YEAR-SECOND-SEMESTER": [
            "ND-FIRST-YEAR-FIRST-SEMESTER",
            "ND-FIRST-YEAR-SECOND-SEMESTER",
            "ND-SECOND-YEAR-FIRST-SEMESTER",
        ],
    }
    
    semesters_to_load = all_semesters.get(current_standard, [])
    print(f"📊 Loading previous ND GPAs for {current_standard}: {semesters_to_load}")
    
    if not os.path.exists(mastersheet_path):
        print(f"❌ ND Mastersheet not found: {mastersheet_path}")
        return {}
    
    try:
        xl = pd.ExcelFile(mastersheet_path)
        print(f"📖 Available sheets in ND mastersheet: {xl.sheet_names}")
    except Exception as e:
        print(f"❌ Error opening ND mastersheet: {e}")
        return {}
    
    for semester in semesters_to_load:
        try:
            sheet_name = get_matching_sheet(xl, semester)
            if not sheet_name:
                print(f"⚠️ Skipping ND semester {semester} - no matching sheet found")
                continue

            print(f"📖 Reading ND sheet '{sheet_name}' for semester {semester}")
            
            # Try multiple header rows to find the right structure
            df = None
            header_row_found = None
            
            for header_row in range(0, 10):  # Try first 10 rows
                try:
                    temp_df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=header_row)
                    
                    # Check if this row has EXAM NUMBER and GPA columns
                    has_exam_col = any('EXAM NUMBER' in str(col).upper() for col in temp_df.columns)
                    has_gpa_col = any('GPA' in str(col).upper() and 'CGPA' not in str(col).upper() for col in temp_df.columns)
                    
                    if has_exam_col and has_gpa_col:
                        df = temp_df
                        header_row_found = header_row
                        print(f"✅ Found valid headers at row {header_row_found}")
                        break
                except Exception as e:
                    continue
            
            if df is None or df.empty:
                print(f"⚠️ Could not find valid data structure in sheet '{sheet_name}'")
                continue

            exam_col = find_exam_number_column(df)
            gpa_col = None
            credit_col = None

            # Find GPA and credit columns
            for col in df.columns:
                col_str = str(col).upper()
                if "GPA" in col_str and "CGPA" not in col_str:
                    gpa_col = col
                if "TCPE" in col_str or "TOTAL CREDIT" in col_str or "TOTAL UNIT" in col_str:
                    credit_col = col

            print(f"🔍 Columns found - Exam: {exam_col}, GPA: {gpa_col}, Credits: {credit_col}")

            if exam_col and gpa_col:
                student_count = 0
                for idx, row in df.iterrows():
                    try:
                        exam_no = str(row[exam_col]).strip().upper()
                        if pd.isna(exam_no) or exam_no in ["", "NAN", "NONE", "SUMMARY"]:
                            continue

                        gpa_value = row[gpa_col]
                        if pd.isna(gpa_value):
                            continue

                        credits = 30  # Default
                        if credit_col and credit_col in row and pd.notna(row[credit_col]):
                            try:
                                credits = int(float(row[credit_col]))
                            except (ValueError, TypeError):
                                credits = 30

                        if exam_no not in all_student_data:
                            all_student_data[exam_no] = {"gpas": [], "credits": []}

                        all_student_data[exam_no]["gpas"].append(float(gpa_value))
                        all_student_data[exam_no]["credits"].append(credits)
                        student_count += 1

                        if student_count <= 3:  # Print first 3 for verification
                            print(f"📊 Loaded ND GPA for {exam_no}: {gpa_value} with {credits} credits")

                    except (ValueError, TypeError) as e:
                        continue
                
                print(f"✅ Loaded {student_count} student records from {sheet_name}")
            else:
                print(f"⚠️ Missing required columns in ND {sheet_name}: exam_col={exam_col}, gpa_col={gpa_col}")

        except Exception as e:
            print(f"⚠️ Could not load data from ND {semester}: {e}")
            continue
    
    print(f"📊 Loaded cumulative ND data for {len(all_student_data)} students")
    return all_student_data


def load_previous_gpas(mastersheet_path, current_semester_key):
    """Load previous GPA data from mastersheet for ND CGPA calculation - ENHANCED."""
    # Use the enhanced version
    return load_previous_gpas_enhanced(mastersheet_path, current_semester_key)


def calculate_cgpa(student_data, current_gpa, current_credits):
    """Calculate Cumulative GPA for ND - UPDATED FOR 4.0 SCALE"""
    if not student_data or not student_data.get("gpas"):
        print(f"⚠️ No previous ND GPA data, using current GPA: {current_gpa}")
        return current_gpa
    
    total_grade_points = 0.0
    total_credits = 0
    print(f"🔢 Calculating ND CGPA from {len(student_data['gpas'])} previous semesters (4.0 scale)")
    
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
        print(
            f" - GPA: {prev_gpa}, Credits: {prev_credits}, Running Total: {total_grade_points}/{total_credits}"
        )
    
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits
    print(f"📊 Final ND calculation: {total_grade_points} / {total_credits}")
    
    if total_credits > 0:
        cgpa = round(total_grade_points / total_credits, 2)
        print(f"✅ Calculated ND CGPA (4.0 scale): {cgpa}")
        return cgpa
    else:
        print(f"⚠️ No ND credits, returning current GPA: {current_gpa}")
        return current_gpa


# ----------------------------
# Mastersheet Update Functions (CRITICAL FIXES)
# ----------------------------
def find_sheet_structure(ws):
    """FIXED: Find the header row and build headers dictionary"""
    header_row = None
    headers = {}

    # Look for header row (contains 'EXAM NUMBER')
    for row_idx in range(1, 30):  # Check first 30 rows
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and "EXAM NUMBER" in str(cell_value).upper():
                header_row = row_idx
                # Build headers dictionary
                for col in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=header_row, column=col).value
                    if header_val:
                        headers[str(header_val).strip()] = col
                print(f"✅ Found header row at: {header_row}")
                return header_row, headers

    print(f"❌ Could not find header row")
    return None, {}


def apply_student_sorting(ws, header_row, headers_dict):
    """Apply sorting to students - compatibility function for ANALYSIS sheet"""
    # This is just an alias for the main sorting function
    apply_student_sorting_with_serial_numbers(ws, header_row, headers_dict)


def apply_student_sorting_with_serial_numbers(ws, header_row, headers_dict):
    """Apply sorting to students with PROPER serial numbers - FIXED VERSION"""
    from openpyxl.styles import Font, PatternFill

    exam_col = headers_dict.get("EXAM NUMBER")
    remarks_col = headers_dict.get("REMARKS")
    gpa_col = headers_dict.get("GPA")
    serial_col = 1  # Serial number is always column 1

    if not all([exam_col, remarks_col, gpa_col]):
        return

    # Collect all student rows
    student_rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        exam_no = ws.cell(row, exam_col).value
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break

        # Get basic row data (values only, not styles)
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            row_data.append({"value": cell.value, "number_format": cell.number_format})

        remarks = ws.cell(row, remarks_col).value or ""
        remarks_upper = str(remarks).upper()

        # Assign priority for sorting
        if "WITHDRAW" in remarks_upper:
            priority = 3  # Withdrawn - last
        elif (
            "RESIT" in remarks_upper
            or "CARRYOVER" in remarks_upper
            or "PROBATION" in remarks_upper
        ):
            priority = 2  # Carryover - middle
        else:
            priority = 1  # Passed - first

        student_rows.append(
            {
                "row_num": row,
                "priority": priority,
                "gpa": ws.cell(row, gpa_col).value or 0,
                "row_data": row_data,
                "remarks": remarks_upper,
                "exam_no": exam_no,  # Store exam number for secondary sorting
            }
        )

    # Sort by priority (Passed first, then Carryover, then Withdrawn)
    # Within each group, sort by GPA (descending), then by exam number
    student_rows.sort(
        key=lambda x: (x["priority"], -float(x["gpa"] if x["gpa"] else 0), x["exam_no"])
    )

    # Write sorted data back to worksheet
    for idx, student in enumerate(student_rows):
        target_row = header_row + 1 + idx
        row_data = student["row_data"]

        for col_idx, cell_data in enumerate(row_data, 1):
            cell = ws.cell(target_row, col_idx)
            cell.value = cell_data["value"]

            # Apply number format if present
            if cell_data.get("number_format"):
                cell.number_format = cell_data["number_format"]

            # CRITICAL FIX: Update serial numbers to be consecutive
            if col_idx == serial_col:  # Serial number column
                cell.value = idx + 1

            # Re-apply styling based on content and student status
            if student["priority"] == 3:  # Withdrawn
                cell.fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )
                if "WITHDRAW" in student["remarks"]:
                    cell.font = Font(bold=True, color="FF0000")
            elif idx % 2 == 0:  # Alternate row coloring
                cell.fill = PatternFill(
                    start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
                )

            # Apply GPA styling for GPA column
            if col_idx == headers_dict.get("GPA"):
                try:
                    gpa_val = float(cell.value) if cell.value else 0
                    if gpa_val >= 3.5:
                        cell.font = Font(
                            bold=True, color="006400"
                        )  # Dark green for high GPA
                    elif gpa_val < 2.0:
                        cell.font = Font(bold=True, color="FF0000")  # Red for low GPA
                except (ValueError, TypeError):
                    pass
    print(
        f" ✅ Applied student sorting with proper serial numbers (1 to {len(student_rows)})"
    )


# CRITICAL FIX #1: Proper course column identification
def identify_course_columns_properly(headers):
    """Identify course columns using PROPER pattern matching"""
    import re

    course_columns = {}

    for header, col_idx in headers.items():
        # Match patterns like: NUR101, NSC201, etc. (3 letters + 3 digits)
        # Also match GNS111, etc.
        if re.match(r"^[A-Z]{3}\d{3}$", str(header).strip()):
            course_columns[header] = col_idx
            print(f"✅ Identified course column: '{header}' at index {col_idx}")

    print(f"📊 Total course columns identified: {len(course_columns)}")
    return course_columns


def update_summary_section_fixed(ws, headers, header_row, course_columns):
    """Update SUMMARY section - FIXED VERSION with embedded number updates and aligned fails"""
    print(f" 📊 Updating summary section...")

    try:
        # Find SUMMARY section
        summary_start_row = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "SUMMARY" in str(cell_value).upper():
                summary_start_row = row_idx
                break

        if not summary_start_row:
            print(" ℹ️ No SUMMARY section found")
            return

        # Find exam column
        exam_col_idx = None
        for col_name, col_idx in headers.items():
            if "EXAM NUMBER" in col_name.upper():
                exam_col_idx = col_idx
                break

        if not exam_col_idx:
            print(" ❌ No exam column found")
            return

        # CRITICAL FIX: Re-read headers to ensure we have fresh column mappings
        fresh_headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                fresh_headers[str(header_val).strip()] = col_idx

        # CRITICAL FIX: Re-identify course columns from fresh headers
        import re

        fresh_course_columns = {}
        for header, col_idx in fresh_headers.items():
            if re.match(r"^[A-Z]{3}\d{3}$", str(header).strip()):
                fresh_course_columns[header] = col_idx

        # Count statistics from CURRENT data (reading directly from worksheet)
        total_students = 0
        passed_students = 0
        resit_students = 0
        probation_students = 0
        withdrawn_students = 0
        course_failures = {course: 0 for course in fresh_course_columns}

        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no = ws.cell(row=row_idx, column=exam_col_idx).value

            # Stop at summary
            if not exam_no or "SUMMARY" in str(exam_no).upper():
                break

            if str(exam_no).strip() in ["", "NAN", "NONE"]:
                continue

            total_students += 1

            # Count remarks
            for col_name, col_idx in headers.items():
                if "REMARKS" in col_name.upper():
                    remarks = ws.cell(row=row_idx, column=col_idx).value
                    if remarks:
                        remarks_upper = str(remarks).upper()
                        if "PASSED" in remarks_upper:
                            passed_students += 1
                        elif "RESIT" in remarks_upper or "CARRYOVER" in remarks_upper:
                            resit_students += 1
                        elif "PROBATION" in remarks_upper:
                            probation_students += 1
                        elif "WITHDRAW" in remarks_upper:
                            withdrawn_students += 1
                    break

            # Count ACTUAL failures from CURRENT scores (INCLUDING updated resit scores)
            for course in fresh_course_columns:
                if course in fresh_headers:
                    col_idx = fresh_headers[course]
                    score = ws.cell(row=row_idx, column=col_idx).value
                    if score is not None and score != "":
                        try:
                            if float(score) < 50:
                                course_failures[course] += 1
                        except (ValueError, TypeError):
                            continue

        # Update fails per course (aligned to course columns)
        fails_row = None
        # Search for the fails per course row in the first three columns
        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col_idx in [1, 2, 3]:
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and "FAILS PER COURSE" in str(cell_value).upper():
                    fails_row = row_idx
                    break
            if fails_row:
                break

        if fails_row:
            print(f"✅ Found fails per course row at row {fails_row}")
            sorted_courses = sorted(
                fresh_course_columns, key=lambda k: fresh_course_columns[k]
            )
            for i, course in enumerate(sorted_courses):
                col = fresh_course_columns[course]
                ws.cell(row=fails_row, column=col).value = course_failures[course]
                print(
                    f" ✅ Updated {course} failures: {course_failures[course]} at column {col}"
                )
        else:
            print("❌ Could not find fails per course row")

        # Update summary rows with embedded numbers
        current_row = summary_start_row + 1

        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            if not cell_value:
                break

            cell_str = str(cell_value).upper()

            if "REGISTERED AND SAT" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {total_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated total students: {total_students}")

            elif (
                "PASSED IN ALL COURSES REGISTERED" in cell_str
                and "FAILED" not in cell_str
            ):
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {passed_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated passed: {passed_students}")

            elif "GRADE POINT AVERAGE (GPA) OF 2.00 AND ABOVE FAILED" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {resit_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated resit (GPA >=2.00): {resit_students}")

            elif "GRADE POINT AVERAGE (GPA) BELOW 2.00 FAILED" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {probation_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated probation: {probation_students}")

            elif "FAILED IN MORE THAN 45%" in cell_str:
                new_value = re.sub(
                    r"A TOTAL OF \d+ STUDENTS",
                    f"A total of {withdrawn_students} students",
                    cell_value,
                    flags=re.I,
                )
                ws.cell(row=current_row, column=1).value = new_value
                print(f" ✅ Updated withdrawn: {withdrawn_students}")

            current_row += 1

        print(f" ✅ Summary section updated")

    except Exception as e:
        print(f" ❌ Error updating summary: {e}")
        traceback.print_exc()


def ensure_required_sheets_exist(wb):
    """Ensure CGPA_SUMMARY and ANALYSIS sheets exist in workbook"""
    from openpyxl.styles import Font, Alignment, PatternFill

    print(f"\n🔍 CHECKING FOR REQUIRED SHEETS...")
    print(f" Current sheets: {wb.sheetnames}")

    # Create CGPA_SUMMARY if it doesn't exist
    if "CGPA_SUMMARY" not in wb.sheetnames:
        print(f"⚠️ CGPA_SUMMARY sheet not found - creating it...")
        cgpa_ws = wb.create_sheet("CGPA_SUMMARY")

        # Add basic structure
        cgpa_ws["A1"] = "CGPA SUMMARY"
        cgpa_ws["A1"].font = Font(bold=True, size=14)
        cgpa_ws["A1"].alignment = Alignment(horizontal="center")

        print(f"✅ Created CGPA_SUMMARY sheet")
    else:
        print(f"✅ CGPA_SUMMARY sheet exists")

    # Create ANALYSIS if it doesn't exist
    if "ANALYSIS" not in wb.sheetnames:
        print(f"⚠️ ANALYSIS sheet not found - creating it...")
        analysis_ws = wb.create_sheet("ANALYSIS")

        # Add basic structure
        analysis_ws["A1"] = "PERFORMANCE ANALYSIS"
        analysis_ws["A1"].font = Font(bold=True, size=14)
        analysis_ws["A1"].alignment = Alignment(horizontal="center")

        print(f"✅ Created ANALYSIS sheet")
    else:
        print(f"✅ ANALYSIS sheet exists")

    print(f"✅ All required sheets verified")
    return True


def update_cgpa_summary_sheet_fixed(wb, semester_key, header_row, set_name):
    """Update CGPA_SUMMARY sheet - COMPLETELY FIXED VERSION with professional headers, color coding, and CLASS OF AWARD"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    print(f" 📈 Updating CGPA_SUMMARY...")

    if "CGPA_SUMMARY" not in wb.sheetnames:
        print(" ℹ️ No CGPA_SUMMARY sheet")
        return

    cgpa_ws = wb["CGPA_SUMMARY"]

    # FIXED: Properly handle merged cells by unmerging first
    try:
        merged_ranges = list(cgpa_ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            cgpa_ws.unmerge_cells(str(merged_range))
        print(f" ✅ Unmerged {len(merged_ranges)} cell ranges")
    except Exception as e:
        print(f" ⚠️ Could not unmerge cells: {e}")

    # Clear ALL old data (including headers)
    for row in range(1, cgpa_ws.max_row + 1):
        for col in range(1, cgpa_ws.max_column + 1):
            cgpa_ws.cell(row, col).value = None
            cgpa_ws.cell(row, col).fill = PatternFill()  # Clear formatting
            cgpa_ws.cell(row, col).font = Font()
            cgpa_ws.cell(row, col).border = Border()

    # ===================================================================
    # STEP 1: CREATE PROFESSIONAL HEADER SECTION
    # ===================================================================

    class_name = set_name

    # Get semester info
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )

    # Calculate total columns needed (11 columns for our data - added PROBATION HISTORY and CLASS OF AWARD)
    total_columns = 11
    last_column = get_column_letter(total_columns)

    # Row 1: Institution Name (Merged and Centered)
    cgpa_ws.merge_cells(f"A1:{last_column}1")
    title_cell = cgpa_ws["A1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Department (Merged and Centered)
    cgpa_ws.merge_cells(f"A2:{last_column}2")
    dept_cell = cgpa_ws["A2"]
    dept_cell.value = "DEPARTMENT OF NURSING"
    dept_cell.font = Font(bold=True, size=12, name="Calibri")
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Class and Sheet Title (Merged and Centered)
    cgpa_ws.merge_cells(f"A3:{last_column}3")
    class_cell = cgpa_ws["A3"]
    class_cell.value = f"{class_name} CLASS - CUMULATIVE GPA SUMMARY"
    class_cell.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
    class_cell.fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    class_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 4: Date and Academic Session (Merged and Centered)
    cgpa_ws.merge_cells(f"A4:{last_column}4")
    date_cell = cgpa_ws["A4"]
    current_year = datetime.now().year
    date_cell.value = f"{current_year}/{current_year + 1} Academic Session - Generated on {datetime.now().strftime('%B %d, %Y')}"
    date_cell.font = Font(size=11, name="Calibri", italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5: Empty spacer
    cgpa_ws.row_dimensions[5].height = 5

    # ===================================================================
    # STEP 2: CREATE COLUMN HEADERS WITH PROFESSIONAL NAMES
    # ===================================================================

    # Row 6: Column Headers - UPDATED WITH PROBATION HISTORY AND CLASS OF AWARD
    headers = [
        "S/N",
        "EXAM NUMBER",
        "NAME",
        "PROBATION HISTORY",
        "Y1S1",
        "Y1S2",
        "Y2S1",
        "Y2S2",
        "CGPA",
        "CLASS OF AWARD",
        "WITHDRAWN",
    ]

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = cgpa_ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Set row height for header
    cgpa_ws.row_dimensions[6].height = 25

    # ===================================================================
    # STEP 3: COLLECT AND POPULATE DATA WITH PROPER WITHDRAWN TRACKING
    # ===================================================================

    # Collect data from all semesters using SINGLE workbook session
    semester_keys = [
        "ND-FIRST-YEAR-FIRST-SEMESTER",
        "ND-FIRST-YEAR-SECOND-SEMESTER",
        "ND-SECOND-YEAR-FIRST-SEMESTER",
        "ND-SECOND-YEAR-SECOND-SEMESTER",
    ]
    
    # Map full semester names to abbreviated headings
    semester_abbreviation_map = {
        "ND-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
        "ND-FIRST-YEAR-SECOND-SEMESTER": "Y1S2",
        "ND-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
        "ND-SECOND-YEAR-SECOND-SEMESTER": "Y2S2"
    }
    
    semester_data = {}

    # Track ALL historically withdrawn students across ALL semesters - CRITICAL FIX
    all_withdrawn_students = set()

    # FIRST PASS: Identify ALL historically withdrawn students from ALL semesters
    print(f" 🔍 FIRST PASS: Identifying ALL historically withdrawn students...")
    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue

            exam_col = headers_dict.get("EXAM NUMBER")
            remarks_col = headers_dict.get("REMARKS")

            if not all([exam_col, remarks_col]):
                continue

            # Collect withdrawn students from this semester
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break

                exam_no_clean = str(exam_no).strip().upper()
                remarks = ws.cell(row, remarks_col).value or ""
                remarks_upper = str(remarks).upper()

                # CRITICAL FIX: Track ALL students who were EVER withdrawn
                if "WITHDRAW" in remarks_upper:
                    all_withdrawn_students.add(exam_no_clean)
                    print(
                        f" 📝 Found historically withdrawn: {exam_no_clean} in {sheet_name}"
                    )

    print(
        f" ✅ Identified {len(all_withdrawn_students)} historically withdrawn students"
    )

    # SECOND PASS: Collect semester data with PERSISTENT withdrawn status and probation tracking
    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue

            exam_col = headers_dict.get("EXAM NUMBER")
            name_col = headers_dict.get("NAME")
            gpa_col = headers_dict.get("GPA")
            credits_col = headers_dict.get("TCPE")
            remarks_col = headers_dict.get("REMARKS")

            if not all([exam_col, name_col, gpa_col]):
                continue

            data = {}
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no_cell = ws.cell(row, exam_col).value
                if not exam_no_cell or "SUMMARY" in str(exam_no_cell).upper():
                    break

                exam_no = str(exam_no_cell).strip().upper()
                name = ws.cell(row, name_col).value
                gpa_val = ws.cell(row, gpa_col).value
                credits_val = ws.cell(row, credits_col).value if credits_col else 0
                remarks = ws.cell(row, remarks_col).value if remarks_col else ""

                # CRITICAL FIX: Check if student is historically withdrawn (PERSISTENT STATUS)
                is_withdrawn = exam_no in all_withdrawn_students
                
                # Track probation status
                is_probation = "PROBATION" in str(remarks).upper()

                try:
                    gpa = float(gpa_val) if gpa_val else 0
                    credits = float(credits_val) if credits_val else 0
                    data[exam_no] = {
                        "name": name,
                        "gpa": gpa,
                        "credits": credits,
                        "remarks": remarks,
                        "withdrawn": is_withdrawn,
                        "probation": is_probation,
                    }
                except (ValueError, TypeError):
                    continue

            semester_data[key] = data

    # Collect unique students
    all_exam_no = set()
    for d in semester_data.values():
        all_exam_no.update(d.keys())

    students = []
    for exam_no in all_exam_no:
        total_gp = 0.0
        total_cr = 0.0
        gpas = {}
        name = None
        probation_semesters = []

        # CRITICAL FIX: Check if student is withdrawn in ANY semester
        withdrawn = False
        for key, d in semester_data.items():
            if exam_no in d:
                data = d[exam_no]
                gpas[key] = data["gpa"]
                total_gp += data["gpa"] * data["credits"]
                total_cr += data["credits"]
                
                # Track probation semesters with abbreviated names
                if data.get("probation", False):
                    abbrev_semester = semester_abbreviation_map.get(key, key)
                    probation_semesters.append(abbrev_semester)
                
                # If student is withdrawn in ANY semester, mark as withdrawn
                if data["withdrawn"]:
                    withdrawn = True
                    print(f" 🔒 Student {exam_no} marked as withdrawn (found in {key})")
                if not name:
                    name = data["name"]

        cgpa = round(total_gp / total_cr, 2) if total_cr > 0 else 0.0
        
        # Calculate CLASS OF AWARD - FIXED: INACTIVE below pass before withdrawn
        valid_semester_count = len(gpas)
        
        if withdrawn:
            class_of_award = "WITHDRAWN"
        elif valid_semester_count < 4:
            # INACTIVE students (haven't completed all semesters)
            class_of_award = "INACTIVE"
        else:
            # Normal award calculation based on CGPA (4.0 scale)
            # FIXED: INACTIVE is now handled above, so only active students reach here
            if cgpa >= 3.5:
                class_of_award = "Distinction"
            elif cgpa >= 3.0:
                class_of_award = "Upper Credit"
            elif cgpa >= 2.0:
                class_of_award = "Lower Credit"
            elif cgpa >= 1.0:
                class_of_award = "Pass"
            else:
                class_of_award = "Fail"
        
        students.append(
            {
                "exam_no": exam_no,
                "name": name,
                "probation_history": ", ".join(probation_semesters) if probation_semesters else "None",
                "gpas": gpas,
                "cgpa": cgpa,
                "class_of_award": class_of_award,
                "withdrawn": withdrawn,
            }
        )

    # ===================================================================
    # STEP 4: SORT STUDENTS WITH INACTIVE BELOW PASS AND BEFORE WITHDRAWN
    # ===================================================================
    
    # FIXED: Create separate groups for proper sorting
    distinction_students = [s for s in students if s["class_of_award"] == "Distinction" and not s["withdrawn"]]
    upper_credit_students = [s for s in students if s["class_of_award"] == "Upper Credit" and not s["withdrawn"]]
    lower_credit_students = [s for s in students if s["class_of_award"] == "Lower Credit" and not s["withdrawn"]]
    pass_students = [s for s in students if s["class_of_award"] == "Pass" and not s["withdrawn"]]
    fail_students = [s for s in students if s["class_of_award"] == "Fail" and not s["withdrawn"]]
    inactive_students = [s for s in students if s["class_of_award"] == "INACTIVE" and not s["withdrawn"]]
    withdrawn_students = [s for s in students if s["withdrawn"]]
    
    # Sort each group by CGPA descending, then by exam number
    distinction_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    upper_credit_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    lower_credit_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    pass_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    fail_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    inactive_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    withdrawn_students.sort(key=lambda s: (-s["cgpa"], s["exam_no"]))
    
    # FIXED: Combine in correct order - INACTIVE below pass and before withdrawn
    sorted_students = (distinction_students + upper_credit_students + lower_credit_students + 
                      pass_students + inactive_students + fail_students + withdrawn_students)

    # ===================================================================
    # STEP 5: WRITE STUDENT DATA WITH FORMATTING AND PROPER SERIAL NUMBERS
    # ===================================================================

    row = 7  # Start from row 7 (after header)
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Alternate row colors
    even_row_fill = PatternFill(
        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
    )
    withdrawn_fill = PatternFill(
        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
    )
    inactive_fill = PatternFill(
        start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"
    )

    # Color coding for withdrawn column
    withdrawn_yes_fill = PatternFill(
        start_color="FFCCCB", end_color="FFCCCB", fill_type="solid"
    )
    withdrawn_no_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )

    # CRITICAL FIX: Proper serial numbers from 1 to n
    serial_number = 1

    for idx, s in enumerate(sorted_students):
        # Determine row fill
        if s["withdrawn"]:
            row_fill = withdrawn_fill
        elif s["class_of_award"] == "INACTIVE":
            row_fill = inactive_fill
        elif idx % 2 == 0:
            row_fill = even_row_fill
        else:
            row_fill = PatternFill()  # White

        # Serial Number (PROPERLY SORTED from 1 to n)
        cell = cgpa_ws.cell(row, 1, value=serial_number)
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        serial_number += 1

        # Exam Number
        cell = cgpa_ws.cell(row, 2, value=s["exam_no"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Name
        cell = cgpa_ws.cell(row, 3, value=s["name"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Probation History (LEFT ALIGNED)
        cell = cgpa_ws.cell(row, 4, value=s["probation_history"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Semester GPAs - UPDATED MAPPING TO NEW HEADERS
        semester_mapping = {
            "ND-FIRST-YEAR-FIRST-SEMESTER": 5,  # Y1S1
            "ND-FIRST-YEAR-SECOND-SEMESTER": 6,  # Y1S2
            "ND-SECOND-YEAR-FIRST-SEMESTER": 7,  # Y2S1
            "ND-SECOND-YEAR-SECOND-SEMESTER": 8,  # Y2S2
        }

        for sem_key, col_idx in semester_mapping.items():
            gpa_value = s["gpas"].get(sem_key, "")
            cell = cgpa_ws.cell(row, col_idx, value=gpa_value)
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if gpa_value:
                cell.number_format = "0.00"

        # CGPA
        cell = cgpa_ws.cell(row, 9, value=s["cgpa"])
        cell.border = data_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "0.00"
        cell.font = Font(bold=True)
        
        # CLASS OF AWARD (with color coding)
        class_of_award = s["class_of_award"]
        cell = cgpa_ws.cell(row, 10, value=class_of_award)
        cell.border = data_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply color coding for CLASS OF AWARD
        if class_of_award == "Distinction":
            cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            cell.font = Font(bold=True, color="006400")  # Dark green
        elif class_of_award == "Upper Credit":
            cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            cell.font = Font(bold=True, color="000080")  # Navy blue
        elif class_of_award == "Lower Credit":
            cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            cell.font = Font(bold=True, color="8B4513")  # Saddle brown
        elif class_of_award == "Pass":
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.font = Font(bold=True, color="000000")  # Black
        elif class_of_award == "Fail":
            cell.fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
            cell.font = Font(bold=True, color="8B0000")  # Dark red
        elif class_of_award == "INACTIVE":
            cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
            cell.font = Font(bold=True, color="FF4500")  # Orange red
        elif class_of_award == "WITHDRAWN":
            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            cell.font = Font(bold=True, color="696969")  # Dim gray

        # Withdrawn status with COLOR CODING
        withdrawn_status = "Yes" if s["withdrawn"] else "No"
        cell = cgpa_ws.cell(row, 11, value=withdrawn_status)
        cell.border = data_border
        
        if s["withdrawn"]:
            cell.fill = withdrawn_yes_fill
            cell.font = Font(bold=True, color="FF0000")
        else:
            cell.fill = withdrawn_no_fill
            cell.font = Font(bold=True, color="006400")
            
        cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    # ===================================================================
    # STEP 6: ADD SUMMARY STATISTICS
    # ===================================================================

    summary_start_row = row + 2

    total_students = len(students)
    avg_cgpa = (
        round(sum(s["cgpa"] for s in students) / total_students, 2)
        if total_students > 0
        else 0
    )
    highest_cgpa = max(s["cgpa"] for s in students) if students else 0
    lowest_cgpa = min(s["cgpa"] for s in students if s["cgpa"] > 0) if students else 0
    withdrawn_count = len(withdrawn_students)
    inactive_count = len(inactive_students)

    # Summary header
    cgpa_ws.merge_cells(f"A{summary_start_row}:{last_column}{summary_start_row}")
    summary_header = cgpa_ws.cell(summary_start_row, 1)
    summary_header.value = "SUMMARY STATISTICS"
    summary_header.font = Font(bold=True, size=12, name="Calibri")
    summary_header.fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    summary_header.alignment = Alignment(horizontal="center", vertical="center")
    summary_header.border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="thin"),
    )

    # Summary data
    summary_data = [
        ("Total Students:", total_students),
        ("Average Cumulative CGPA:", avg_cgpa),
        ("Highest Cumulative CGPA:", highest_cgpa),
        ("Lowest Cumulative CGPA:", lowest_cgpa),
        ("Withdrawn Students:", withdrawn_count),
        ("Inactive Students:", inactive_count),
    ]

    summary_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    for i, (label, value) in enumerate(summary_data):
        current_row = summary_start_row + 1 + i

        # Label (columns A-E merged)
        cgpa_ws.merge_cells(f"A{current_row}:E{current_row}")
        label_cell = cgpa_ws.cell(current_row, 1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=11, name="Calibri")
        label_cell.fill = summary_fill
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.border = Border(
            left=Side(style="medium"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Value (columns F-K merged)
        cgpa_ws.merge_cells(f"F{current_row}:{last_column}{current_row}")
        value_cell = cgpa_ws.cell(current_row, 6)
        value_cell.value = value
        value_cell.font = Font(bold=True, size=11, name="Calibri")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="medium"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if isinstance(value, float):
            value_cell.number_format = "0.00"

    # Bottom border for summary section
    last_summary_row = summary_start_row + len(summary_data)
    for col in range(1, total_columns + 1):
        cell = cgpa_ws.cell(last_summary_row, col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=Side(style="medium"),
        )

    # ===================================================================
    # STEP 7: ADJUST COLUMN WIDTHS
    # ===================================================================

    cgpa_ws.column_dimensions["A"].width = 8   # S/N
    cgpa_ws.column_dimensions["B"].width = 18  # EXAM NUMBER
    cgpa_ws.column_dimensions["C"].width = 35  # NAME
    cgpa_ws.column_dimensions["D"].width = 20  # PROBATION HISTORY
    cgpa_ws.column_dimensions["E"].width = 10  # Y1S1
    cgpa_ws.column_dimensions["F"].width = 10  # Y1S2
    cgpa_ws.column_dimensions["G"].width = 10  # Y2S1
    cgpa_ws.column_dimensions["H"].width = 10  # Y2S2
    cgpa_ws.column_dimensions["I"].width = 12  # CGPA
    cgpa_ws.column_dimensions["J"].width = 18  # CLASS OF AWARD
    cgpa_ws.column_dimensions["K"].width = 12  # WITHDRAWN

    # Set specific row heights
    cgpa_ws.row_dimensions[1].height = 20  # Title
    cgpa_ws.row_dimensions[2].height = 18  # Department
    cgpa_ws.row_dimensions[3].height = 20  # Class/Title
    cgpa_ws.row_dimensions[4].height = 16  # Date

    print(
        f" ✅ CGPA_SUMMARY updated with {len(students)} students, {withdrawn_count} withdrawn, {inactive_count} inactive"
    )
    print(f" ✅ CLASS OF AWARD column included with proper color coding")
    print(f" ✅ FIXED: INACTIVE students are now below Pass and before Withdrawn in the sorting order")


def update_cgpa_summary_with_withdrawn(wb, withdrawn_students):
    """Update CGPA SUMMARY sheet with PERSISTENT withdrawn status and apply sorting"""
    if "CGPA SUMMARY" not in wb.sheetnames:
        return

    cgpa_ws = wb["CGPA SUMMARY"]
    header_row_found, headers_dict = find_sheet_structure(cgpa_ws)

    if not header_row_found:
        return

    exam_col = headers_dict.get("EXAM NUMBER")
    remarks_col = headers_dict.get("REMARKS")
    cgpa_col = headers_dict.get("CGPA")
    withdrawn_col = None

    # Find withdrawn column if it exists
    for header, col in headers_dict.items():
        if "WITHDRAW" in header.upper():
            withdrawn_col = col
            break

    if not all([exam_col, remarks_col, cgpa_col]):
        return

    # Update remarks AND withdrawn column for withdrawn students
    for row in range(header_row_found + 1, cgpa_ws.max_row + 1):
        exam_no = cgpa_ws.cell(row, exam_col).value
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break

        exam_no_clean = str(exam_no).strip()
        if exam_no_clean in withdrawn_students:
            # Update remarks to ensure withdrawn status
            cgpa_ws.cell(row, remarks_col).value = "WITHDRAWN"

            # Update withdrawn column if it exists
            if withdrawn_col:
                cgpa_ws.cell(row, withdrawn_col).value = "Yes"

            print(f" 🔒 Maintaining withdrawn status in CGPA: {exam_no_clean}")

    # Apply sorting to CGPA SUMMARY
    apply_student_sorting(cgpa_ws, header_row_found, headers_dict)

    print(f" ✅ CGPA SUMMARY updated with PERSISTENT withdrawn status")


def update_analysis_sheet_fixed(
    wb, semester_key, course_columns, headers, header_row, set_name
):
    """Update ANALYSIS sheet - ENHANCED VERSION with PERSISTENT withdrawn tracking and sorting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    print(f" 📊 Updating ANALYSIS with PERSISTENT withdrawn tracking...")

    if "ANALYSIS" not in wb.sheetnames:
        print(" ℹ️ No ANALYSIS sheet")
        return

    analysis_ws = wb["ANALYSIS"]

    # FIXED: Properly handle merged cells by unmerging first
    try:
        merged_ranges = list(analysis_ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            analysis_ws.unmerge_cells(str(merged_range))
        print(f" ✅ Unmerged {len(merged_ranges)} cell ranges in ANALYSIS sheet")
    except Exception as e:
        print(f" ⚠️ Could not unmerge cells in ANALYSIS sheet: {e}")

    # Clear ALL old data (including headers)
    for row in range(1, analysis_ws.max_row + 1):
        for col in range(1, analysis_ws.max_column + 1):
            try:
                cell = analysis_ws.cell(row, col)
                if cell.value is not None:
                    cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()
                cell.border = Border()
            except AttributeError:
                continue
    # ===================================================================
    # STEP 1: CREATE PROFESSIONAL HEADER SECTION
    # ===================================================================

    class_name = set_name

    # Get semester info
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )

    # Calculate total columns needed (7 columns for our data)
    total_columns = 7
    last_column = get_column_letter(total_columns)

    # Row 1: Institution Name
    analysis_ws.merge_cells(f"A1:{last_column}1")
    title_cell = analysis_ws["A1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Department
    analysis_ws.merge_cells(f"A2:{last_column}2")
    dept_cell = analysis_ws["A2"]
    dept_cell.value = "DEPARTMENT OF NURSING"
    dept_cell.font = Font(bold=True, size=12, name="Calibri")
    dept_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Class and Sheet Title
    analysis_ws.merge_cells(f"A3:{last_column}3")
    class_cell = analysis_ws["A3"]
    class_cell.value = f"{class_name} CLASS - PERFORMANCE ANALYSIS"
    class_cell.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
    class_cell.fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    class_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 4: Date and Academic Session
    analysis_ws.merge_cells(f"A4:{last_column}4")
    date_cell = analysis_ws["A4"]
    current_year = datetime.now().year
    date_cell.value = f"{current_year}/{current_year + 1} Academic Session - Generated on {datetime.now().strftime('%B %d, %Y')}"
    date_cell.font = Font(size=11, name="Calibri", italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5: Empty spacer
    analysis_ws.row_dimensions[5].height = 5

    # ===================================================================
    # STEP 2: CREATE COLUMN HEADERS
    # ===================================================================

    # Row 6: Column Headers
    headers_list = [
        "SEMESTER",
        "TOTAL",
        "PASSED",
        "CARRYOVER",
        "WITHDRAWN",
        "AVG GPA",
        "PASS RATE (%)",
    ]

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers_list, 1):
        cell = analysis_ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    analysis_ws.row_dimensions[6].height = 25
    # ===================================================================
    # STEP 3: ENHANCED DATA COLLECTION WITH PERSISTENT WITHDRAWN TRACKING
    # ===================================================================

    semester_keys = [
        "ND-FIRST-YEAR-FIRST-SEMESTER",
        "ND-FIRST-YEAR-SECOND-SEMESTER",
        "ND-SECOND-YEAR-FIRST-SEMESTER",
        "ND-SECOND-YEAR-SECOND-SEMESTER",
    ]

    semester_display_names = {
        "ND-FIRST-YEAR-FIRST-SEMESTER": "Year 1 - Semester 1",
        "ND-FIRST-YEAR-SECOND-SEMESTER": "Year 1 - Semester 2",
        "ND-SECOND-YEAR-FIRST-SEMESTER": "Year 2 - Semester 1",
        "ND-SECOND-YEAR-SECOND-SEMESTER": "Year 2 - Semester 2",
    }

    semester_stats = {}
    overall_total = 0
    overall_passed = 0
    overall_carryover = 0
    overall_withdrawn = 0
    overall_gpa_sum = 0
    overall_students_for_gpa = 0

    # Track withdrawn students across ALL semesters - CRITICAL FIX
    all_withdrawn_students = set()

    # FIRST PASS: Identify ALL historically withdrawn students from ALL semesters
    print(f" 🔍 FIRST PASS: Identifying ALL historically withdrawn students...")

    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue

            exam_col = headers_dict.get("EXAM NUMBER")
            remarks_col = headers_dict.get("REMARKS")

            if not all([exam_col, remarks_col]):
                continue

            # Collect withdrawn students from this semester
            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break

                exam_no_clean = str(exam_no).strip()
                remarks = ws.cell(row, remarks_col).value or ""
                remarks_upper = str(remarks).upper()

                # CRITICAL FIX: Track ALL students who were EVER withdrawn
                if "WITHDRAW" in remarks_upper:
                    all_withdrawn_students.add(exam_no_clean)
                    print(
                        f" 📝 Found historically withdrawn: {exam_no_clean} in {sheet_name}"
                    )

    print(
        f" ✅ Identified {len(all_withdrawn_students)} historically withdrawn students"
    )

    # SECOND PASS: Process each semester with PERSISTENT withdrawn status
    print(f" 🔍 SECOND PASS: Processing semesters with persistent withdrawn status...")

    for key in semester_keys:
        sheet_name = None
        for sheet in wb.sheetnames:
            if key.upper() in sheet.upper():
                sheet_name = sheet
                break

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_found, headers_dict = find_sheet_structure(ws)
            if not header_row_found:
                continue

            exam_col = headers_dict.get("EXAM NUMBER")
            gpa_col = headers_dict.get("GPA")
            remarks_col = headers_dict.get("REMARKS")

            if not all([exam_col, gpa_col, remarks_col]):
                continue

            total = 0
            passed = 0
            resit = 0
            probation = 0
            withdrawn = 0
            gpa_sum = 0

            # Track students for this semester
            semester_students = set()

            for row in range(header_row_found + 1, ws.max_row + 1):
                exam_no = ws.cell(row, exam_col).value
                if not exam_no or "SUMMARY" in str(exam_no).upper():
                    break

                exam_no_clean = str(exam_no).strip()
                semester_students.add(exam_no_clean)
                total += 1

                remarks = ws.cell(row, remarks_col).value or ""
                remarks_upper = str(remarks).upper()
                gpa_val = ws.cell(row, gpa_col).value

                # CRITICAL FIX: Check if student is historically withdrawn (PERSISTENT STATUS)
                is_withdrawn = exam_no_clean in all_withdrawn_students

                if is_withdrawn:
                    withdrawn += 1
                    # ENSURE withdrawn status is maintained in remarks
                    if "WITHDRAW" not in remarks_upper:
                        ws.cell(row, remarks_col).value = "WITHDRAWN"
                        print(
                            f" 🔒 Maintaining withdrawn status for: {exam_no_clean} in {sheet_name}"
                        )
                elif "PASSED" in remarks_upper:
                    passed += 1
                elif "RESIT" in remarks_upper or "CARRYOVER" in remarks_upper:
                    resit += 1
                elif "PROBATION" in remarks_upper:
                    probation += 1

                try:
                    gpa_sum += float(gpa_val) if gpa_val else 0
                except (ValueError, TypeError):
                    pass

            # Apply sorting to semester sheet (Passed -> Carryover -> Withdrawn)
            apply_student_sorting_with_serial_numbers(
                ws, header_row_found, headers_dict
            )

            avg_gpa = round(gpa_sum / total, 2) if total > 0 else 0
            pass_rate = round(passed / total * 100, 2) if total > 0 else 0
            carryover = resit + probation

            semester_stats[key] = {
                "total": total,
                "passed": passed,
                "carryover": carryover,
                "withdrawn": withdrawn,
                "avg_gpa": avg_gpa,
                "pass_rate": pass_rate,
            }

            overall_total += total
            overall_passed += passed
            overall_carryover += carryover
            overall_withdrawn += withdrawn
            overall_gpa_sum += gpa_sum
            overall_students_for_gpa += total
    # Update CGPA SUMMARY with PERSISTENT withdrawn status
    update_cgpa_summary_with_withdrawn(wb, all_withdrawn_students)
    # ===================================================================
    # STEP 4: WRITE SEMESTER DATA WITH FORMATTING
    # ===================================================================

    row = 7  # Start from row 7 (after header)
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Alternate row colors
    even_row_fill = PatternFill(
        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
    )

    for idx, key in enumerate(semester_keys):
        if key in semester_stats:
            stats = semester_stats[key]

            # Determine row fill
            row_fill = even_row_fill if idx % 2 == 0 else PatternFill()

            # Semester name
            cell = analysis_ws.cell(row, 1, value=semester_display_names[key])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(bold=True)

            # Total students
            cell = analysis_ws.cell(row, 2, value=stats["total"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Passed
            cell = analysis_ws.cell(row, 3, value=stats["passed"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Carryover
            cell = analysis_ws.cell(row, 4, value=stats["carryover"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Withdrawn
            cell = analysis_ws.cell(row, 5, value=stats["withdrawn"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Average GPA
            cell = analysis_ws.cell(row, 6, value=stats["avg_gpa"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"

            # Pass Rate
            cell = analysis_ws.cell(row, 7, value=stats["pass_rate"])
            cell.border = data_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00"

            row += 1
    # ===================================================================
    # STEP 5: ADD OVERALL SUMMARY ROW
    # ===================================================================

    row += 1  # Skip a row

    overall_avg_gpa = (
        round(overall_gpa_sum / overall_students_for_gpa, 2)
        if overall_students_for_gpa > 0
        else 0
    )
    overall_pass_rate = (
        round(overall_passed / overall_total * 100, 2) if overall_total > 0 else 0
    )

    overall_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    overall_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # OVERALL label
    cell = analysis_ws.cell(row, 1, value="OVERALL")
    cell.border = overall_border
    cell.fill = overall_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(bold=True, size=12)

    # Overall stats
    overall_data = [
        overall_total,
        overall_passed,
        overall_carryover,
        overall_withdrawn,
        overall_avg_gpa,
        overall_pass_rate,
    ]

    for col_idx, value in enumerate(overall_data, 2):
        cell = analysis_ws.cell(row, col_idx, value=value)
        cell.border = overall_border
        cell.fill = overall_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=11)

        if col_idx >= 6:  # GPA and Pass Rate
            cell.number_format = "0.00"
    # ===================================================================
    # STEP 6: ADJUST COLUMN WIDTHS
    # ===================================================================

    analysis_ws.column_dimensions["A"].width = 25  # SEMESTER
    analysis_ws.column_dimensions["B"].width = 12  # TOTAL
    analysis_ws.column_dimensions["C"].width = 12  # PASSED
    analysis_ws.column_dimensions["D"].width = 14  # CARRYOVER
    analysis_ws.column_dimensions["E"].width = 14  # WITHDRAWN
    analysis_ws.column_dimensions["F"].width = 12  # AVG GPA
    analysis_ws.column_dimensions["G"].width = 15  # PASS RATE (%)

    # Set specific row heights
    analysis_ws.row_dimensions[1].height = 20  # Title
    analysis_ws.row_dimensions[2].height = 18  # Department
    analysis_ws.row_dimensions[3].height = 20  # Class/Title
    analysis_ws.row_dimensions[4].height = 16  # Date

    print(
        f" ✅ ANALYSIS updated with PERSISTENT withdrawn tracking: {overall_withdrawn} withdrawn students"
    )
    print(f" ✅ Student sorting applied across all sheets")


def apply_complete_professional_formatting(wb, semester_key, header_row, set_name):
    """Apply complete professional formatting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    print(f" 🎨 Applying formatting...")

    # Find sheet
    current_sheet = None
    for sheet in wb.sheetnames:
        if semester_key.upper() in sheet.upper():
            current_sheet = sheet
            break

    if not current_sheet:
        return

    ws = wb[current_sheet]

    # Define styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Format header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Set specific widths
    if header_row > 0:
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                header_str = str(header_val).upper()
                if "EXAM NUMBER" in header_str:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20
                elif "NAME" in header_str:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 35

    print(f" ✅ Formatting applied")


# ----------------------------
# ENHANCED CUMULATIVE UPDATE FUNCTION
# ----------------------------

def identify_withdrawn_students(ws, headers, header_row):
    """Identify withdrawn students from worksheet."""
    withdrawn_students = []
    
    remarks_col = None
    exam_col = None
    
    for header, col_idx in headers.items():
        if "REMARKS" in header.upper():
            remarks_col = col_idx
        if "EXAM NUMBER" in header.upper():
            exam_col = col_idx
    
    if not remarks_col or not exam_col:
        return withdrawn_students
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        exam_no = ws.cell(row=row_idx, column=exam_col).value
        remarks = ws.cell(row=row_idx, column=remarks_col).value
        
        if not exam_no or "SUMMARY" in str(exam_no).upper():
            break
        
        if remarks and "WITHDRAW" in str(remarks).upper():
            withdrawn_students.append(str(exam_no).strip())
    
    return withdrawn_students


def recalculate_all_student_records(ws, headers, header_row, course_columns, course_units_dict):
    """Recalculate all student records with current scores."""
    exam_col = None
    for header, col_idx in headers.items():
        if "EXAM NUMBER" in header.upper():
            exam_col = col_idx
            break
    
    if not exam_col:
        return
    
    summary_columns = {}
    summary_keys = ["FAILED COURSES", "REMARKS", "CU Passed", "CU Failed", "TCPE", "GPA", "AVERAGE", "CGPA"]
    
    for key in summary_keys:
        for header, col_idx in headers.items():
            if key.upper() in header.upper():
                summary_columns[key] = col_idx
                break
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        try:
            exam_no_cell = ws.cell(row=row_idx, column=exam_col)
            exam_no = str(exam_no_cell.value).strip().upper() if exam_no_cell.value else None
            
            if not exam_no or exam_no in ["", "NAN", "NONE"]:
                continue
            
            if "SUMMARY" in str(exam_no).upper():
                break
            
            # Recalculate using current scores
            failed_courses = []
            cu_passed = 0
            cu_failed = 0
            total_credits = 0
            total_grade_points = 0.0
            total_score = 0.0
            valid_courses = 0
            
            for course_code, course_col in course_columns.items():
                score_cell = ws.cell(row=row_idx, column=course_col)
                score_value = score_cell.value
                
                if score_value is not None and score_value != "":
                    try:
                        score = float(score_value)
                        total_score += score
                        valid_courses += 1
                        
                        credit_unit = find_credit_unit_simple(course_code, course_units_dict)
                        total_credits += credit_unit
                        
                        grade_point = get_grade_point(score)
                        total_grade_points += grade_point * credit_unit
                        
                        if score >= DEFAULT_PASS_THRESHOLD:
                            cu_passed += credit_unit
                        else:
                            cu_failed += credit_unit
                            failed_courses.append(course_code)
                    except (ValueError, TypeError):
                        continue
            
            # Calculate metrics
            gpa = round(total_grade_points / total_credits, 2) if total_credits > 0 else 0.0
            average = round(total_score / valid_courses, 2) if valid_courses > 0 else 0.0
            
            # Determine remarks
            passed_percent = cu_passed / total_credits if total_credits > 0 else 0
            if passed_percent < 0.45:
                remarks = "WITHDRAW"
            elif cu_failed == 0:
                remarks = "PASSED"
            else:
                if gpa >= 2.0:
                    remarks = "RESIT"
                else:
                    remarks = "PROBATION"
            
            # Update summary columns
            if "FAILED COURSES" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["FAILED COURSES"]).value = ", ".join(failed_courses) if failed_courses else "NONE"
            
            if "REMARKS" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["REMARKS"]).value = remarks
            
            if "CU Passed" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["CU Passed"]).value = cu_passed
            
            if "CU Failed" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["CU Failed"]).value = cu_failed
            
            if "TCPE" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["TCPE"]).value = total_credits
            
            if "GPA" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["GPA"]).value = gpa
            
            if "AVERAGE" in summary_columns:
                ws.cell(row=row_idx, column=summary_columns["AVERAGE"]).value = average
            
        except Exception as e:
            print(f"⚠️ Error recalculating row {row_idx}: {e}")
            continue


def find_credit_unit_simple(course_code, units_dict):
    """Simple credit unit finder."""
    if course_code in units_dict:
        return units_dict[course_code]
    code_no_space = course_code.replace(" ", "")
    if code_no_space in units_dict:
        return units_dict[code_no_space]
    return 2


# ----------------------------
# UPDATED CGPA SUMMARY AND ANALYSIS SHEETS FOR CARRYOVER PROCESSOR
# ----------------------------

def create_cgpa_summary_sheet_carryover(mastersheet_path, timestamp):
    """
    Create CGPA summary sheet for carryover processor - adapted from regular processor
    FIXED: Proper header row detection based on actual worksheet structure
    UPDATED: Added S/N column and fixed creation logic
    UPDATED: Professional column headings (Y1S1, Y1S2, Y2S1, Y2S2) and color-coded withdrawn status
    UPDATED TITLE: Changed to match the requested format
    UPDATED: Added CLASS OF AWARD column with the specified criteria
    FIXED: Dynamic GPA column detection and proper data extraction
    UPDATED: Auto-fit column width for all columns and subtle color coding for CLASS OF AWARD
    UPDATED: INACTIVE class of award for students with less than 4 valid semester GPAs
    UPDATED: Recalculated all GPAs for 4.0 scale and proper sorting
    FIXED: Freeze headings (rows 1-5), proper serial numbering, and auto-fit columns
    UPDATED: NAME and PROBATION HISTORY left aligned, abbreviated probation history
    """
    try:
        print("📊 Creating CGPA Summary Sheet for Carryover...")
        
        # Load the mastersheet workbook
        wb = load_workbook(mastersheet_path)
        
        # Collect CGPA data from all semesters
        cgpa_data = {}
        
        # Map full semester names to abbreviated professional headings
        semester_abbreviation_map = {
            "ND-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
            "ND-FIRST-YEAR-SECOND-SEMESTER": "Y1S2",
            "ND-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
            "ND-SECOND-YEAR-SECOND-SEMESTER": "Y2S2"
        }

        # Track which semesters we actually found data for
        semesters_with_data = set()

        for sheet_name in wb.sheetnames:
            if sheet_name in SEMESTER_ORDER:
                try:
                    print(f"🔍 Processing sheet: {sheet_name}")
                    
                    # DYNAMIC APPROACH: Try multiple header rows to find the actual data
                    best_header_row = None
                    best_gpa_col = None
                    best_exam_col = None
                    best_df = None
                    
                    # Try header rows from 0 to 20 (covering all possible positions)
                    for header_row in range(0, 21):
                        try:
                            df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=header_row)
                            
                            # Skip if dataframe is empty or has very few columns
                            if df.empty or len(df.columns) < 3:
                                continue
                                
                            # Look for exam number column
                            exam_col = find_exam_number_column(df)
                            if not exam_col:
                                continue
                                
                            # Look for GPA column with multiple possible patterns
                            gpa_col = None
                            for col in df.columns:
                                col_str = str(col).upper().strip()
                                # More flexible GPA column detection
                                if any(pattern in col_str for pattern in ["GPA", "GRADE POINT", "POINT"]):
                                    gpa_col = col
                                    break
                            
                            if not gpa_col:
                                continue
                            
                            # Validate that we have actual data in these columns
                            valid_students = 0
                            sample_gpas = []
                            
                            for idx, row in df.iterrows():
                                exam_no = str(row[exam_col]).strip()
                                gpa_val = row[gpa_col]
                                
                                # Check for valid exam number and GPA
                                if (exam_no and exam_no != "nan" and exam_no != "" and 
                                    not exam_no.lower().startswith(("exam", "reg", "registration")) and
                                    len(exam_no) >= 5):  # Reasonable exam number length
                                    
                                    if pd.notna(gpa_val) and str(gpa_val).strip() != "":
                                        try:
                                            gpa_float = float(gpa_val)
                                            if 0 <= gpa_float <= 5.0:  # Valid GPA range
                                                valid_students += 1
                                                sample_gpas.append((exam_no, gpa_float))
                                                if valid_students >= 5:  # Found enough valid data
                                                    break
                                        except (ValueError, TypeError):
                                            continue
                            
                            if valid_students >= 3:  # Require at least 3 valid students
                                best_header_row = header_row
                                best_gpa_col = gpa_col
                                best_exam_col = exam_col
                                best_df = df
                                print(f"✅ Found valid data at header row {header_row}: {valid_students} students with GPA")
                                print(f"   Sample: {sample_gpas[:3]}")  # Show first 3 samples
                                break
                                    
                        except Exception as e:
                            # Continue to next header row if this one fails
                            continue
                    
                    # Process the data if we found a valid configuration
                    if best_header_row is not None and best_df is not None:
                        df = best_df
                        exam_col = best_exam_col
                        gpa_col = best_gpa_col
                        
                        print(f"📊 Processing {sheet_name} with header row {best_header_row}")
                        print(f"📝 Using columns - Exam: '{exam_col}', GPA: '{gpa_col}'")
                        
                        # Also look for name and remarks columns
                        name_col = None
                        remarks_col = None
                        for col in df.columns:
                            col_str = str(col).upper().strip()
                            if "NAME" in col_str and "COURSE" not in col_str:
                                name_col = col
                            elif "REMARKS" in col_str:
                                remarks_col = col
                        
                        students_found = 0
                        for idx, row in df.iterrows():
                            exam_no = str(row[exam_col]).strip()
                            
                            # Validate exam number format
                            if (exam_no and exam_no != "nan" and exam_no != "" and 
                                not exam_no.lower().startswith(("exam", "reg", "registration")) and
                                len(exam_no) >= 5):
                                
                                if exam_no not in cgpa_data:
                                    cgpa_data[exam_no] = {
                                        "name": (
                                            str(row[name_col]) if name_col and pd.notna(row.get(name_col))
                                            else ""
                                        ),
                                        "gpas": {},
                                        "status": "Active",
                                        "probation_semesters": [],
                                    }
                                
                                # Extract GPA value and convert to 4.0 scale if needed
                                gpa_value = row[gpa_col]
                                if pd.notna(gpa_value) and str(gpa_value).strip() != "":
                                    try:
                                        gpa_float = float(gpa_value)
                                        # Convert from 5.0 scale to 4.0 scale if necessary
                                        if gpa_float > 4.0:  # Assume it's in 5.0 scale
                                            gpa_float = (gpa_float / 5.0) * 4.0
                                        if 0 <= gpa_float <= 4.0:  # Valid GPA range for 4.0 scale
                                            cgpa_data[exam_no]["gpas"][sheet_name] = round(gpa_float, 2)
                                            students_found += 1
                                            semesters_with_data.add(sheet_name)
                                            
                                            # Update student tracker with CGPA data
                                            if exam_no in STUDENT_TRACKER:
                                                STUDENT_TRACKER[exam_no]["has_cgpa_data"] = True
                                                STUDENT_TRACKER[exam_no]["current_gpa"] = gpa_float
                                                STUDENT_TRACKER[exam_no]["cgpa_status"] = "Active in CGPA"
                                    except (ValueError, TypeError):
                                        continue
                                
                                # Track probation status - ABBREVIATE SEMESTER NAMES
                                if remarks_col and pd.notna(row.get(remarks_col)):
                                    remarks = str(row[remarks_col])
                                    if (remarks == "Probation" and 
                                        sheet_name not in cgpa_data[exam_no]["probation_semesters"]):
                                        # Use abbreviated semester name for probation history
                                        abbrev_semester = semester_abbreviation_map.get(sheet_name, sheet_name)
                                        cgpa_data[exam_no]["probation_semesters"].append(abbrev_semester)
                        
                        print(f"📊 Extracted GPA data for {students_found} students in {sheet_name}")
                        
                    else:
                        print(f"❌ Could not find valid GPA data in {sheet_name} after trying all header rows")
                        # Try to debug by showing available columns from first few rows
                        try:
                            for test_row in [0, 1, 2, 3, 4, 5]:
                                test_df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=test_row)
                                if not test_df.empty:
                                    print(f"   Row {test_row} columns: {[str(col) for col in test_df.columns[:8]]}")
                                    # Show sample data
                                    if len(test_df) > 0:
                                        sample = {}
                                        for col in test_df.columns[:5]:
                                            sample[col] = test_df.iloc[0][col]
                                        print(f"   Sample data: {sample}")
                                    break
                        except Exception as e:
                            print(f"   Debug failed: {e}")
                        
                except Exception as e:
                    print(f"⚠️ Warning: Could not process sheet {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue

        print(f"📊 Total students with CGPA data: {len(cgpa_data)}")
        print(f"📊 Semesters with data: {semesters_with_data}")

        # Create CGPA summary dataframe with probation tracking
        summary_data = []
        for exam_no, data in cgpa_data.items():
            row = {
                "EXAM NUMBER": exam_no,
                "NAME": data["name"],
                "PROBATION HISTORY": (
                    ", ".join(data["probation_semesters"])
                    if data["probation_semesters"]
                    else "None"
                ),
            }
            
            # Add GPA for each semester using abbreviated headings and calculate cumulative
            total_gpa = 0
            semester_count = 0
            for semester in SEMESTER_ORDER:
                if semester in data["gpas"]:
                    # Use abbreviated column name
                    abbrev_semester = semester_abbreviation_map.get(semester, semester)
                    row[abbrev_semester] = data["gpas"][semester]
                    if pd.notna(data["gpas"][semester]):
                        total_gpa += data["gpas"][semester]
                        semester_count += 1
                else:
                    abbrev_semester = semester_abbreviation_map.get(semester, semester)
                    row[abbrev_semester] = None
                    
            # Calculate Cumulative CGPA
            cumulative_cgpa = (
                round(total_gpa / semester_count, 2) if semester_count > 0 else 0.0
            )
            row["CUMULATIVE CGPA"] = cumulative_cgpa
            
            # FIX 1: Calculate CLASS OF AWARD with WITHDRAWN check FIRST
            is_withdrawn = is_student_withdrawn(exam_no)
            valid_semester_count = semester_count  # Count of semesters with valid GPA data

            # CRITICAL: Check withdrawal status FIRST
            if is_withdrawn:
                class_of_award = "WITHDRAWN"  # Withdrawn students should show WITHDRAWN, not Pass/Fail
            elif not is_withdrawn and valid_semester_count < 4:
                class_of_award = "INACTIVE"
            else:
                # Normal award calculation based on CGPA (4.0 scale)
                if cumulative_cgpa >= 3.5:
                    class_of_award = "Distinction"
                elif cumulative_cgpa >= 3.0:
                    class_of_award = "Upper Credit"
                elif cumulative_cgpa >= 2.0:
                    class_of_award = "Lower Credit"
                elif cumulative_cgpa >= 1.0:
                    class_of_award = "Pass"
                else:
                    class_of_award = "Fail"
                    
            row["CLASS OF AWARD"] = class_of_award
            row["WITHDRAWN"] = "Yes" if is_withdrawn else "No"
            summary_data.append(row)

        # Create summary dataframe
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            
            # UPDATED: Sort the data as requested
            def get_sort_key(row):
                award = row["CLASS OF AWARD"]
                if award in ["Distinction", "Upper Credit", "Lower Credit"]:
                    return (0, -row["CUMULATIVE CGPA"])  # Active students first
                elif award == "INACTIVE":
                    return (1, 0)
                elif award == "WITHDRAWN":
                    return (3, 0)  # Withdrawn students at the end
                elif award == "Fail":
                    return (2, 0)
                else:  # Pass and others
                    return (0, -row["CUMULATIVE CGPA"])
            
            # Create a temporary sort key column
            sort_keys = summary_df.apply(get_sort_key, axis=1)
            summary_df = summary_df.iloc[sort_keys.argsort()].reset_index(drop=True)
            
            # FIX 2: Add proper serial numbering after sorting
            summary_df.insert(0, "S/N", range(1, len(summary_df) + 1))
            
            print(f"✅ Successfully created CGPA summary with {len(summary_df)} students")
            # Show sample of the summary data
            print("📋 Sample of CGPA summary data:")
            for i in range(min(5, len(summary_df))):
                student = summary_df.iloc[i]
                print(f"  {i+1}. {student['S/N']}. {student['EXAM NUMBER']}: CGPA={student.get('CUMULATIVE CGPA', 'N/A')}, Award={student.get('CLASS OF AWARD', 'N/A')}, Semesters={sum(1 for sem in SEMESTER_ORDER if pd.notna(student.get(semester_abbreviation_map.get(sem, sem))))}")
        else:
            print("⚠️ No CGPA data found for any students")
            # Create empty dataframe with correct columns
            headers = (
                ["S/N", "EXAM NUMBER", "NAME", "PROBATION HISTORY"]
                + [semester_abbreviation_map.get(sem, sem) for sem in SEMESTER_ORDER]
                + ["CUMULATIVE CGPA", "WITHDRAWN", "CLASS OF AWARD"]
            )
            summary_df = pd.DataFrame(columns=headers)

        # Add the summary sheet to the workbook
        if "CGPA_SUMMARY" in wb.sheetnames:
            del wb["CGPA_SUMMARY"]
        ws = wb.create_sheet("CGPA_SUMMARY")

        # Define headers for column calculation
        headers = (
            ["S/N", "EXAM NUMBER", "NAME", "PROBATION HISTORY"]
            + [semester_abbreviation_map.get(sem, sem) for sem in SEMESTER_ORDER]
            + ["CUMULATIVE CGPA", "WITHDRAWN", "CLASS OF AWARD"]
        )
        num_columns = len(headers)
        end_col_letter = get_column_letter(num_columns)

        # ADD PROFESSIONAL HEADER WITH SCHOOL NAME - FIXED: Use exact requested format
        ws.merge_cells(f"A1:{end_col_letter}1")
        title_cell = ws["A1"]
        title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )

        ws.merge_cells(f"A2:{end_col_letter}2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = "DEPARTMENT OF NURSING"
        subtitle_cell.font = Font(bold=True, size=14, color="000000")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        subtitle_cell.fill = PatternFill(
            start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
        )

        # ADD DYNAMIC TITLE ROW - FIXED: Use exact requested format
        ws.merge_cells(f"A3:{end_col_letter}3")
        exam_title_cell = ws["A3"]
        exam_title_cell.value = "NDII CGPA SUMMARY REPORT"
        exam_title_cell.font = Font(bold=True, size=12, color="000000")
        exam_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        exam_title_cell.fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
        )

        ws.merge_cells(f"A4:{end_col_letter}4")
        date_cell = ws["A4"]
        date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add empty row for spacing
        ws.row_dimensions[5].height = 10

        # Write header starting from row 6
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Write data starting from row 7
        if not summary_df.empty:
            for row_idx, row_data in enumerate(summary_df.to_dict("records"), 7):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    if pd.isna(value):
                        value = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply basic styling to data cells
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
                    
                    # Alternate row coloring
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # FIX 1: LEFT ALIGN NAME AND PROBATION HISTORY COLUMNS
                    if col_idx == 1:  # S/N - center
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 2:  # EXAM NUMBER - center
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx in [3, 4]:  # NAME and PROBATION HISTORY - LEFT ALIGNED with wrap text
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:  # All other columns (numeric) - center
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                    # Apply subtle color coding for CLASS OF AWARD column
                    if header == "CLASS OF AWARD" and value:
                        if value == "Distinction":
                            cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Very light green
                        elif value == "Upper Credit":
                            cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")  # Very light blue
                        elif value == "Lower Credit":
                            cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # Very light yellow
                        elif value == "Pass":
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Very light gray
                        elif value == "Fail":
                            cell.fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")  # Very light red
                        elif value == "INACTIVE":  # FIX 1: Add color for INACTIVE
                            cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")  # Very light red/pink
                        elif value == "WITHDRAWN":  # FIX 1: Add color for WITHDRAWN
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light gray
            
            print(f"✅ Written {len(summary_df)} students to CGPA summary sheet")
        else:
            print("⚠️ No data to write to CGPA summary sheet")
            ws.cell(row=7, column=1, value="No CGPA data available")

        # FIX 1: FREEZE PANES AT ROW 7 (so rows 1-6 are frozen)
        ws.freeze_panes = "A7"
        print("✅ Frozen headings (rows 1-6)")

        # FIX 2: AUTO-FIT COLUMN WIDTHS FOR ALL COLUMNS WITH BETTER LOGIC
        print("📏 Auto-fitting column widths to fit content...")
        
        # Define minimum and maximum widths for better control
        min_width = 8
        max_width = 35
        
        # Enhanced auto-fit: Calculate maximum content length for each column
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Check header length first
            header_length = len(str(header))
            max_length = max(max_length, header_length)
            
            # Check data content lengths
            for row_idx in range(7, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        # For text columns, allow more width
                        if header in ["NAME", "PROBATION HISTORY"]:
                            cell_length = min(cell_length, 50)  # Cap very long text
                        max_length = max(max_length, cell_length)
                    except:
                        pass
            
            # Add padding and apply limits
            adjusted_width = min(max_length + 3, max_width)  # Increased padding to 3
            adjusted_width = max(adjusted_width, min_width)
            
            # Special handling for specific columns
            if header == "S/N":
                adjusted_width = 6
            elif header == "EXAM NUMBER":
                adjusted_width = min(max(adjusted_width, 15), 20)
            elif header == "NAME":
                adjusted_width = min(max(adjusted_width, 20), 35)
            elif header == "PROBATION HISTORY":
                adjusted_width = min(max(adjusted_width, 15), 25)  # Reduced width since we abbreviated
            elif header in ["Y1S1", "Y1S2", "Y2S1", "Y2S2"]:
                adjusted_width = min(max(adjusted_width, 8), 12)
            elif header == "CUMULATIVE CGPA":
                adjusted_width = min(max(adjusted_width, 12), 18)
            elif header == "WITHDRAWN":
                adjusted_width = min(max(adjusted_width, 10), 12)
            elif header == "CLASS OF AWARD":
                adjusted_width = min(max(adjusted_width, 12), 20)
            
            ws.column_dimensions[column_letter].width = adjusted_width
            print(f"   📐 Column {column_letter} ({header}): width {adjusted_width}")

        # Adjust row heights for better visibility
        for row in range(6, ws.max_row + 1):
            if row == 6:  # Header row
                ws.row_dimensions[row].height = 25
            else:
                ws.row_dimensions[row].height = 20

        # Add summary statistics if we have data
        if not summary_df.empty:
            stats_row = len(summary_df) + 8
            stats_cell = ws.cell(row=stats_row, column=1, value="SUMMARY STATISTICS")
            stats_cell.font = Font(bold=True, size=12)
            stats_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            if "CUMULATIVE CGPA" in summary_df.columns:
                # FIX 1: Count INACTIVE and WITHDRAWN students
                inactive_count = (summary_df['CLASS OF AWARD'] == 'INACTIVE').sum()
                withdrawn_count = (summary_df['CLASS OF AWARD'] == 'WITHDRAWN').sum()
                
                stats_data = [
                    f"Total Students: {len(summary_df)}",
                    f"Average Cumulative CGPA: {summary_df['CUMULATIVE CGPA'].mean():.2f}",
                    f"Highest Cumulative CGPA: {summary_df['CUMULATIVE CGPA'].max():.2f}",
                    f"Lowest Cumulative CGPA: {summary_df['CUMULATIVE CGPA'].min():.2f}",
                    f"Withdrawn Students: {withdrawn_count}",
                    f"INACTIVE Students (less than 4 semesters): {inactive_count}",
                    f"Students with Probation History: {(summary_df['PROBATION HISTORY'] != 'None').sum()}",
                ]
                
                # Add award distribution
                if "CLASS OF AWARD" in summary_df.columns:
                    award_counts = summary_df['CLASS OF AWARD'].value_counts()
                    stats_data.append("Award Distribution:")
                    for award, count in award_counts.items():
                        stats_data.append(f"  - {award}: {count} students")
                
                for i, stat in enumerate(stats_data, start=stats_row + 1):
                    cell = ws.cell(row=i, column=1, value=stat)
                    if stat.startswith("Award Distribution:"):
                        cell.font = Font(bold=True)
                    elif stat.startswith("  -"):
                        cell.font = Font(italic=True)

        # Save the workbook
        wb.save(mastersheet_path)
        print("✅ CGPA Summary sheet created successfully with:")
        print("   - Frozen headings (rows 1-6)")
        print("   - Proper serial numbering")
        print("   - Auto-fit column widths")
        print("   - NAME and PROBATION HISTORY left aligned")
        print("   - Abbreviated probation history (Y1S1, Y1S2, Y2S1, Y2S2)")
        return summary_df
        
    except Exception as e:
        print(f"❌ Error creating CGPA summary sheet: {e}")
        import traceback
        traceback.print_exc()
        return None


def create_analysis_sheet_carryover(mastersheet_path, timestamp):
    """
    Create an analysis sheet with comprehensive statistics for carryover processor.
    FIXED: Dynamic header detection, robust column finding, accurate student counting, and proper serial numbering
    ENHANCED: Professional formatting with proper cell fitting and visual appeal
    UPDATED: REMOVED INACTIVE STUDENTS column as requested
    """
    try:
        print("📈 Creating Analysis Sheet for Carryover...")
        wb = load_workbook(mastersheet_path)
        
        # Collect data from all semesters
        analysis_data = {
            "semester": [],
            "total_students": [],
            "passed_all": [],
            "resit_students": [],
            "probation_students": [],
            "withdrawn_students": [],
            "average_gpa": [],
            "pass_rate": [],
        }

        # Semester short name mapping
        semester_short_names = {
            "ND-FIRST-YEAR-FIRST-SEMESTER": "Y1-1STS",
            "ND-FIRST-YEAR-SECOND-SEMESTER": "Y1-2NDS", 
            "ND-SECOND-YEAR-FIRST-SEMESTER": "Y2-1STS",
            "ND-SECOND-YEAR-SECOND-SEMESTER": "Y2-2NDS"
        }

        for sheet_name in wb.sheetnames:
            if sheet_name not in SEMESTER_ORDER:
                continue
                
            try:
                print(f"\n🔍 Processing sheet: {sheet_name}")
                
                # ========================================
                # DYNAMIC HEADER ROW DETECTION
                # ========================================
                best_header_row = None
                best_df = None
                
                # Try different header rows (0 to 20)
                for header_row in range(0, 21):
                    try:
                        df = pd.read_excel(
                            mastersheet_path, 
                            sheet_name=sheet_name, 
                            header=header_row,
                            dtype=str
                        )
                        
                        if df.empty or len(df.columns) < 3:
                            continue
                        
                        # Check if we have the required columns
                        columns_upper = [str(col).upper().strip() for col in df.columns]
                        
                        # Look for exam number column
                        exam_col_found = any(
                            keyword in col_upper 
                            for col_upper in columns_upper 
                            for keyword in ["EXAM", "REG", "MATRIC", "STUDENT"]
                        )
                        
                        # Look for remarks column
                        remarks_col_found = any(
                            "REMARK" in col_upper 
                            for col_upper in columns_upper
                        )
                        
                        # Look for GPA column
                        gpa_col_found = any(
                            "GPA" in col_upper or "GRADE POINT" in col_upper
                            for col_upper in columns_upper
                        )
                        
                        if not (exam_col_found and remarks_col_found and gpa_col_found):
                            continue
                        
                        # Validate we have actual student data
                        exam_col = None
                        for col in df.columns:
                            col_str = str(col).upper().strip()
                            if any(keyword in col_str for keyword in ["EXAM", "REG", "MATRIC"]):
                                exam_col = col
                                break
                        
                        if not exam_col:
                            continue
                        
                        # Count valid exam numbers
                        valid_count = 0
                        for idx, row in df.iterrows():
                            exam_no = str(row[exam_col]).strip()
                            # Check if it looks like a real exam number
                            if (exam_no and 
                                exam_no != "nan" and 
                                exam_no != "" and
                                not exam_no.upper().startswith(("EXAM", "REG", "REGISTRATION", "STUDENT")) and
                                len(exam_no) >= 5):
                                valid_count += 1
                                if valid_count >= 3:  # Need at least 3 valid students
                                    break
                        
                        if valid_count >= 3:
                            best_header_row = header_row
                            best_df = df
                            print(f"✅ Found valid header at row {header_row} with {valid_count}+ students")
                            break
                            
                    except Exception as e:
                        continue
                
                if best_header_row is None or best_df is None:
                    print(f"❌ Could not find valid data structure in {sheet_name}")
                    continue
                
                df = best_df
                print(f"📊 Using header row {best_header_row}")
                
                # ========================================
                # DYNAMIC COLUMN DETECTION
                # ========================================
                
                # Find EXAM NUMBER column
                exam_col = None
                for col in df.columns:
                    col_upper = str(col).upper().strip()
                    if any(keyword in col_upper for keyword in [
                        "EXAM NUMBER", "EXAM NO", "REG NO", "REG. NO", 
                        "REGISTRATION", "MATRIC", "STUDENT ID"
                    ]):
                        exam_col = col
                        print(f"✅ Found EXAM column: '{col}'")
                        break
                
                # Find REMARKS column
                remarks_col = None
                for col in df.columns:
                    col_upper = str(col).upper().strip()
                    if "REMARK" in col_upper:
                        remarks_col = col
                        print(f"✅ Found REMARKS column: '{col}'")
                        break
                
                # Find GPA column
                gpa_col = None
                for col in df.columns:
                    col_upper = str(col).upper().strip()
                    if "GPA" in col_upper or "GRADE POINT" in col_upper:
                        gpa_col = col
                        print(f"✅ Found GPA column: '{col}'")
                        break
                
                # Validate we found all required columns
                if not exam_col:
                    print(f"❌ Could not find EXAM NUMBER column in {sheet_name}")
                    continue
                if not remarks_col:
                    print(f"❌ Could not find REMARKS column in {sheet_name}")
                    continue
                if not gpa_col:
                    print(f"❌ Could not find GPA column in {sheet_name}")
                    continue
                
                # ========================================
                # COUNT STUDENTS BY STATUS
                # ========================================
                
                # Use sets to track unique students
                all_students = set()
                passed_students = set()
                resit_students = set()
                probation_students = set()
                withdrawn_students_this_sem = set()
                gpa_values = []
                
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    
                    # Validate exam number
                    if not exam_no or exam_no == "nan" or exam_no == "":
                        continue
                    
                    # Skip header-like values
                    exam_no_upper = exam_no.upper()
                    if any(keyword in exam_no_upper for keyword in [
                        "EXAM", "REG", "REGISTRATION", "STUDENT", "MATRIC", "NUMBER"
                    ]):
                        continue
                    
                    # Must be reasonable length
                    if len(exam_no) < 5:
                        continue
                    
                    # Valid student found
                    all_students.add(exam_no)
                    
                    # Get remarks
                    remarks = str(row[remarks_col]).strip() if pd.notna(row.get(remarks_col)) else ""
                    
                    # Categorize by status
                    if remarks == "Passed":
                        passed_students.add(exam_no)
                    elif remarks == "Resit":
                        resit_students.add(exam_no)
                    elif remarks == "Probation":
                        probation_students.add(exam_no)
                    elif remarks == "Withdrawn":
                        withdrawn_students_this_sem.add(exam_no)
                    
                    # Collect GPA values
                    gpa_val = row[gpa_col]
                    if pd.notna(gpa_val):
                        try:
                            gpa_float = float(gpa_val)
                            if 0 <= gpa_float <= 5.0:
                                gpa_values.append(gpa_float)
                        except (ValueError, TypeError):
                            pass
                
                # Calculate statistics
                total_students = len(all_students)
                passed_all = len(passed_students)
                resit_count = len(resit_students)
                probation_count = len(probation_students)
                withdrawn_count = len(withdrawn_students_this_sem)
                
                avg_gpa = round(sum(gpa_values) / len(gpa_values), 2) if gpa_values else 0.0
                pass_rate = round((passed_all / total_students * 100), 2) if total_students > 0 else 0.0
                
                # Log results
                print(f"📊 {sheet_name} Statistics:")
                print(f"   Total Students: {total_students}")
                print(f"   Passed: {passed_all}")
                print(f"   Resit: {resit_count}")
                print(f"   Probation: {probation_count}")
                print(f"   Withdrawn: {withdrawn_count}")
                print(f"   Average GPA: {avg_gpa}")
                print(f"   Pass Rate: {pass_rate}%")
                
                # Add to analysis data
                short_semester = semester_short_names.get(sheet_name, sheet_name)
                analysis_data["semester"].append(short_semester)
                analysis_data["total_students"].append(total_students)
                analysis_data["passed_all"].append(passed_all)
                analysis_data["resit_students"].append(resit_count)
                analysis_data["probation_students"].append(probation_count)
                analysis_data["withdrawn_students"].append(withdrawn_count)
                analysis_data["average_gpa"].append(avg_gpa)
                analysis_data["pass_rate"].append(pass_rate)
                
            except Exception as e:
                print(f"⚠️ Error processing sheet {sheet_name}: {e}")
                import traceback
                traceback.print_exc()
                continue

        # ========================================
        # CREATE ANALYSIS DATAFRAME
        # ========================================
        
        if not analysis_data["semester"]:
            print("❌ No analysis data collected from any semester")
            # Create empty dataframe
            analysis_df = pd.DataFrame({
                "SEMESTER": [],
                "TOTAL STUDENTS": [],
                "PASSED ALL": [],
                "RESIT STUDENTS": [],
                "PROBATION STUDENTS": [],
                "WITHDRAWN STUDENTS": [],
                "AVERAGE GPA": [],
                "PASS RATE (%)": []
            })
        else:
            analysis_df = pd.DataFrame(analysis_data)
            
            # Rename columns to match expected format
            analysis_df.columns = [
                "SEMESTER",
                "TOTAL STUDENTS",
                "PASSED ALL",
                "RESIT STUDENTS",
                "PROBATION STUDENTS",
                "WITHDRAWN STUDENTS",
                "AVERAGE GPA",
                "PASS RATE (%)"
            ]
            
            # Add overall statistics
            overall_stats = {
                "SEMESTER": "OVERALL",
                "TOTAL STUDENTS": sum(analysis_data["total_students"]),
                "PASSED ALL": sum(analysis_data["passed_all"]),
                "RESIT STUDENTS": sum(analysis_data["resit_students"]),
                "PROBATION STUDENTS": sum(analysis_data["probation_students"]),
                "WITHDRAWN STUDENTS": sum(analysis_data["withdrawn_students"]),
                "AVERAGE GPA": round(sum(analysis_data["average_gpa"]) / len(analysis_data["average_gpa"]), 2) if analysis_data["average_gpa"] else 0.0,
                "PASS RATE (%)": round(sum(analysis_data["pass_rate"]) / len(analysis_data["pass_rate"]), 2) if analysis_data["pass_rate"] else 0.0,
            }
            analysis_df = pd.concat([analysis_df, pd.DataFrame([overall_stats])], ignore_index=True)
            
            print(f"\n✅ Analysis data collected for {len(analysis_data['semester'])} semesters")

        # Add serial number AFTER creating the dataframe
        analysis_df.insert(0, "S/N", range(1, len(analysis_df) + 1))

        # ========================================
        # CREATE ANALYSIS WORKSHEET
        # ========================================
        
        if "ANALYSIS" in wb.sheetnames:
            del wb["ANALYSIS"]
        ws = wb.create_sheet("ANALYSIS")

        headers = list(analysis_df.columns)  # Use actual dataframe columns
        num_columns = len(headers)
        end_col_letter = get_column_letter(num_columns)

        # Professional header with proper row heights
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A1:{end_col_letter}1")
        title_cell = ws["A1"]
        title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")

        ws.row_dimensions[2].height = 25
        ws.merge_cells(f"A2:{end_col_letter}2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = "DEPARTMENT OF NURSING"
        subtitle_cell.font = Font(bold=True, size=14, color="000000")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        subtitle_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        ws.row_dimensions[3].height = 22
        ws.merge_cells(f"A3:{end_col_letter}3")
        exam_title_cell = ws["A3"]
        exam_title_cell.value = "NDII SEMESTER ANALYSIS REPORT"
        exam_title_cell.font = Font(bold=True, size=12, color="000000")
        exam_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        exam_title_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

        ws.row_dimensions[4].height = 20
        ws.merge_cells(f"A4:{end_col_letter}4")
        date_cell = ws["A4"]
        date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add spacing row
        ws.row_dimensions[5].height = 10

        # Write headers (row 6) with proper height
        ws.row_dimensions[6].height = 35
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Write data (starting row 7)
        for row_idx, (_, row_data) in enumerate(analysis_df.iterrows(), start=7):
            ws.row_dimensions[row_idx].height = 25
            
            for col_idx, header in enumerate(headers, 1):
                value = row_data[header]
                if pd.isna(value):
                    value = ""
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Styling
                is_overall = (row_data["SEMESTER"] == "OVERALL")
                
                if is_overall:
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    cell.font = Font(bold=True, size=10)
                else:
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.font = Font(size=10)
                
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Professional column widths
        column_widths = {
            'A': 8,   # S/N
            'B': 12,  # SEMESTER
            'C': 16,  # TOTAL STUDENTS
            'D': 14,  # PASSED ALL
            'E': 16,  # RESIT STUDENTS
            'F': 18,  # PROBATION STUDENTS
            'G': 18,  # WITHDRAWN STUDENTS
            'H': 14,  # AVERAGE GPA
            'I': 16,  # PASS RATE (%)
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Analysis notes with proper spacing
        notes_row = ws.max_row + 2
        ws.row_dimensions[notes_row].height = 25
        ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=num_columns)
        notes_cell = ws.cell(row=notes_row, column=1, value="ANALYSIS NOTES:")
        notes_cell.font = Font(bold=True, size=12, color="1E90FF")
        notes_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        notes_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        notes = [
            "• Resit Students: Students with GPA ≥ 2.0 who failed ≤45% of credits",
            "• Probation Students: Students with GPA < 2.0 OR failed >45% with GPA ≥ 2.0",
            "• Withdrawn Students: Students who failed >45% of credits with GPA < 2.0",
            "• Pass Rate: Percentage of students who passed all courses in the semester",
        ]
        
        for i, note in enumerate(notes):
            note_row = notes_row + i + 1
            ws.row_dimensions[note_row].height = 22
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=num_columns)
            note_cell = ws.cell(row=note_row, column=1, value=note)
            note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            note_cell.font = Font(size=10, italic=True)

        wb.save(mastersheet_path)
        print("✅ Analysis sheet created successfully with proper serial numbering and professional formatting")
        return analysis_df
        
    except Exception as e:
        print(f"❌ Error creating analysis sheet: {e}")
        import traceback
        traceback.print_exc()
        return None


# ----------------------------
# ENHANCED CUMULATIVE UPDATE FUNCTION WITH CGPA AND ANALYSIS
# ----------------------------

def update_mastersheet_with_cumulative_updates_carryover(
    mastersheet_path,
    updates,
    semester_key,
    original_zip_path,
    course_titles_dict,
    course_units_dict,
    set_name,
    clean_dir
):
    """FIXED: Update mastersheet with cumulative updates and proper versioning INCLUDING CGPA AND ANALYSIS SHEETS."""
    print(f"\n{'='*80}")
    print(f"🔄 CUMULATIVE UPDATE WITH CGPA & ANALYSIS: {semester_key}")
    print(f"📁 Set: {set_name}")
    print(f"{'='*80}")
    
    # Create backup only if it doesn't exist
    backup_path = create_backup_if_not_exists(original_zip_path)
    
    wb = None
    try:
        # SINGLE WORKBOOK LOAD
        print(f"📖 Loading workbook...")
        wb = load_workbook(mastersheet_path)
        
        # Ensure required sheets exist
        ensure_required_sheets_exist(wb)
        
        # Find semester sheet
        sheet_name = None
        for sheet in wb.sheetnames:
            if semester_key.upper() in sheet.upper():
                sheet_name = sheet
                break
        
        if not sheet_name:
            print(f"❌ No sheet found for: {semester_key}")
            return False
        
        ws = wb[sheet_name]
        print(f"✅ Working on sheet: {sheet_name}")
        
        # Find header structure
        header_row, headers = find_sheet_structure(ws)
        if not header_row:
            print(f"❌ Could not find header row")
            return False
        
        print(f"✅ Found header row at: {header_row}")
        
        # Identify course columns
        course_columns = identify_course_columns_properly(headers)
        if not course_columns:
            print(f"❌ No course columns found!")
            return False
        
        print(f"✅ Found {len(course_columns)} course columns")
        
        # Track withdrawn students BEFORE updates
        withdrawn_students_before = identify_withdrawn_students(ws, headers, header_row)
        print(f"📊 Withdrawn students before update: {len(withdrawn_students_before)}")
        
        # =============================================================
        # STEP 1: Apply ALL score updates
        # =============================================================
        print(f"\n📝 STEP 1: APPLYING SCORE UPDATES...")
        exam_col = None
        for header, col_idx in headers.items():
            if "EXAM NUMBER" in header.upper():
                exam_col = col_idx
                break
        
        if not exam_col:
            print("❌ No exam column found")
            return False
        
        students_updated = 0
        courses_updated = 0
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            exam_no_cell = ws.cell(row=row_idx, column=exam_col)
            exam_no = str(exam_no_cell.value).strip().upper() if exam_no_cell.value else None
            
            if not exam_no or exam_no in ["", "NAN", "NONE"]:
                continue
            
            if "SUMMARY" in str(exam_no).upper():
                break
            
            if exam_no in updates:
                for course_code, new_score in updates[exam_no].items():
                    if course_code in course_columns:
                        course_col = course_columns[course_code]
                        old_score = ws.cell(row=row_idx, column=course_col).value
                        
                        # Update the score
                        ws.cell(row=row_idx, column=course_col).value = new_score
                        
                        # Apply color coding
                        if new_score >= DEFAULT_PASS_THRESHOLD:
                            ws.cell(row=row_idx, column=course_col).fill = PatternFill(
                                start_color="90EE90", end_color="90EE90", fill_type="solid"
                            )
                        else:
                            ws.cell(row=row_idx, column=course_col).fill = PatternFill(
                                start_color="FFD580", end_color="FFD580", fill_type="solid"
                            )
                        
                        ws.cell(row=row_idx, column=course_col).font = Font(bold=True)
                        courses_updated += 1
                
                students_updated += 1
        
        print(f"✅ Updated {courses_updated} scores for {students_updated} students")
        
        # =============================================================
        # STEP 2: Recalculate student records
        # =============================================================
        print(f"\n🧮 STEP 2: RECALCULATING RECORDS...")
        recalculate_all_student_records(ws, headers, header_row, course_columns, course_units_dict)
        
        # =============================================================
        # STEP 3: Update summary section
        # =============================================================
        print(f"\n📊 STEP 3: UPDATING SUMMARY SECTION...")
        update_summary_section_fixed(ws, headers, header_row, course_columns)
        
        # =============================================================
        # STEP 4: Handle withdrawn students
        # =============================================================
        print(f"\n👥 STEP 4: HANDLING WITHDRAWN STUDENTS...")
        
        # Remove from current semester sheet
        remove_withdrawn_from_semester_sheets(wb, semester_key)
        
        # Preserve in summary sheets
        withdrawn_students_after = identify_withdrawn_students(ws, headers, header_row)
        preserve_withdrawn_in_summary_sheets(wb, withdrawn_students_after)
        
        # =============================================================
        # STEP 5: Apply formatting and sorting
        # =============================================================
        print(f"\n🎨 STEP 5: APPLYING FORMATTING...")
        apply_complete_professional_formatting(wb, semester_key, header_row, set_name)
        apply_student_sorting_with_serial_numbers(ws, header_row, headers)
        
        # =============================================================
        # CRITICAL FIX: STEP 6 - UPDATE CGPA AND ANALYSIS IN-MEMORY
        # =============================================================
        print(f"\n📈 STEP 6: UPDATING CGPA_SUMMARY AND ANALYSIS SHEETS (IN-MEMORY)...")
        
        # Update CGPA_SUMMARY sheet using in-memory workbook
        if "CGPA_SUMMARY" in wb.sheetnames:
            print(f"🎯 Updating CGPA_SUMMARY sheet...")
            update_cgpa_summary_sheet_fixed(wb, semester_key, header_row, set_name)
            print(f"✅ CGPA_SUMMARY updated successfully")
        else:
            print(f"⚠️ CGPA_SUMMARY sheet not found - it should have been created")
        
        # Update ANALYSIS sheet using in-memory workbook
        if "ANALYSIS" in wb.sheetnames:
            print(f"🎯 Updating ANALYSIS sheet...")
            # Re-find header structure for current semester (in case it changed)
            ws_current = wb[sheet_name]
            header_row_current, headers_current = find_sheet_structure(ws_current)
            course_columns_current = identify_course_columns_properly(headers_current)
            
            # Call the CORRECT function that works with in-memory workbook
            update_analysis_sheet_fixed(
                wb, 
                semester_key, 
                course_columns_current, 
                headers_current, 
                header_row_current, 
                set_name
            )
            print(f"✅ ANALYSIS sheet updated successfully")
        else:
            print(f"⚠️ ANALYSIS sheet not found - it should have been created")
        
        # =============================================================
        # STEP 7 - SAVE WORKBOOK
        # =============================================================
        print(f"\n💾 STEP 7: SAVING WORKBOOK WITH ALL UPDATES...")
        
        try:
            wb.save(mastersheet_path)
            print(f"✅ Mastersheet saved successfully with all updates")
        except Exception as save_error:
            print(f"❌ Error saving mastersheet: {save_error}")
            return False
        finally:
            # Close workbook to release file lock
            if wb:
                wb.close()
                wb = None  # Set to None so finally block doesn't try to close again
        
        # =============================================================
        # STEP 8: Create updated ZIP with versioning
        # =============================================================
        print(f"\n📦 STEP 8: CREATING UPDATED ZIP...")
        
        # Determine next version number
        next_version = get_next_version_number(clean_dir)
        updated_zip_name = f"UPDATED_{next_version}_{os.path.basename(original_zip_path)}"
        updated_zip_path = os.path.join(clean_dir, updated_zip_name)
        
        # Create updated ZIP
        temp_extract_dir = tempfile.mkdtemp()
        try:
            # Extract original ZIP
            with zipfile.ZipFile(original_zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_extract_dir)
            
            # Replace mastersheet in extracted files
            mastersheet_found = False
            for root, dirs, files in os.walk(temp_extract_dir):
                for file in files:
                    if "mastersheet" in file.lower() and file.endswith(".xlsx"):
                        old_mastersheet_path = os.path.join(root, file)
                        shutil.copy2(mastersheet_path, old_mastersheet_path)
                        mastersheet_found = True
                        print(f"✅ Replaced mastersheet in: {old_mastersheet_path}")
                        break
                if mastersheet_found:
                    break
            
            # Create new ZIP
            with zipfile.ZipFile(updated_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(temp_extract_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_extract_dir)
                        zipf.write(file_path, arcname)
            
            # Verify new ZIP
            if os.path.exists(updated_zip_path) and os.path.getsize(updated_zip_path) > 0:
                print(f"✅ SUCCESS: Created {updated_zip_name}")
                print(f"📦 File size: {os.path.getsize(updated_zip_path)} bytes")
                
                # Test ZIP integrity
                try:
                    with zipfile.ZipFile(updated_zip_path, 'r') as test_zip:
                        test_zip.testzip()
                    print(f"✅ ZIP integrity verified")
                except Exception as e:
                    print(f"⚠️ ZIP integrity check warning: {e}")
                
                return True
            else:
                print(f"❌ Failed to create updated ZIP")
                return False
                
        except Exception as e:
            print(f"❌ Error during ZIP creation: {e}")
            traceback.print_exc()
            return False
        finally:
            if os.path.exists(temp_extract_dir):
                shutil.rmtree(temp_extract_dir)
        
    except Exception as e:
        print(f"❌ Error in cumulative update: {e}")
        traceback.print_exc()
        return False
    finally:
        # Only close if wb is still open
        if wb:
            try:
                wb.close()
            except:
                pass

# ----------------------------
# CARRYOVER PROCESSING FUNCTIONS
# ----------------------------

def process_carryover_core(resit_file_path, mastersheet_path, semester_key, set_name, 
                          course_titles_dict, credit_units_dict, pass_threshold,
                          course_code_to_title, course_code_to_unit):
    """Core carryover processing logic - FIXED GPA LOADING."""
    print(f"📖 Reading files with enhanced header detection...")
    
    # Read resit file with enhanced header detection
    resit_headers, resit_header_idx = detect_all_headers_robust(resit_file_path)
    
    # DEBUG: Print what we found
    print(f"🔍 Resit headers detected: {resit_headers}")
    
    if not resit_headers['registration_col']:
        print("❌ Could not detect registration column in resit file")
        print("💡 Trying fallback method...")
        
        # Fallback: Try to read the file and manually find the registration column
        resit_df = pd.read_excel(resit_file_path, header=0)
        print(f"📊 Resit file columns: {list(resit_df.columns)}")
        
        # Look for exam number column manually
        for col in resit_df.columns:
            col_str = str(col).upper()
            if any(pattern in col_str for pattern in ["EXAM", "REG", "MATRIC", "STUDENT"]):
                resit_headers['registration_col'] = col
                print(f"✅ Found registration column via fallback: {col}")
                break
        
        if not resit_headers['registration_col'] and len(resit_df.columns) > 0:
            # Use first column as last resort
            resit_headers['registration_col'] = resit_df.columns[0]
            print(f"⚠️ Using first column as registration column: {resit_df.columns[0]}")
    
    if not resit_headers['registration_col']:
        print("❌ Still could not detect registration column in resit file")
        return []
    
    # Read resit file with detected header row
    resit_df = pd.read_excel(resit_file_path, header=resit_header_idx)
    
    # Read mastersheet with enhanced header detection - FIXED VERSION
    print(f"🔍 Reading mastersheet with enhanced header detection...")
    
    xl = pd.ExcelFile(mastersheet_path)
    sheet_name = get_matching_sheet(xl, semester_key)
    
    if not sheet_name:
        print(f"❌ No matching sheet found for {semester_key}")
        return []
    
    # CRITICAL FIX: Find the header row in mastersheet (it's NOT at row 0)
    mastersheet_df = None
    mastersheet_header_idx = None
    
    for header_row in range(0, 15):  # Try first 15 rows
        try:
            temp_df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=header_row)
            
            # Check if this row has EXAM NUMBER column
            has_exam_col = any('EXAM NUMBER' in str(col).upper() for col in temp_df.columns)
            
            if has_exam_col:
                mastersheet_df = temp_df
                mastersheet_header_idx = header_row
                print(f"✅ Found mastersheet headers at row {header_row}")
                print(f"📊 Mastersheet columns at row {header_row}: {list(mastersheet_df.columns)}")
                break
            else:
                print(f"❌ Row {header_row} doesn't have EXAM NUMBER, columns: {[str(col).upper() for col in temp_df.columns]}")
        except Exception as e:
            print(f"⚠️ Error reading row {header_row}: {e}")
            continue
    
    if mastersheet_df is None:
        print("❌ Could not find valid headers in mastersheet")
        return []
    
    # Detect mastersheet headers from the found dataframe
    mastersheet_headers = {
        'registration_col': None,
        'name_col': None,
        'course_cols': []
    }
    
    # Find EXAM NUMBER in mastersheet
    for col in mastersheet_df.columns:
        col_upper = str(col).upper().strip()
        if 'EXAM NUMBER' in col_upper:
            mastersheet_headers['registration_col'] = col
            print(f"✅ Found EXAM NUMBER in mastersheet: '{col}'")
            break
    
    if not mastersheet_headers['registration_col']:
        print("❌ Could not detect registration column in mastersheet")
        return []
    
    # Find NAME column in mastersheet
    for col in mastersheet_df.columns:
        col_upper = str(col).upper().strip()
        if 'NAME' == col_upper:
            mastersheet_headers['name_col'] = col
            print(f"✅ Found NAME in mastersheet: '{col}'")
            break
    
    # Find course columns in mastersheet - FIXED PATTERN
    import re
    course_pattern = re.compile(r'^[A-Z]{3}\d{3}$', re.IGNORECASE)
    
    for col in mastersheet_df.columns:
        col_str = str(col).strip()
        
        # Skip EXAM NUMBER and NAME
        if col == mastersheet_headers['registration_col'] or col == mastersheet_headers['name_col']:
            continue
        
        # Check if matches course pattern
        if course_pattern.match(col_str):
            mastersheet_headers['course_cols'].append(col)
            print(f"✅ Found course column in mastersheet: '{col}'")
    
    print(f"✅ Total course columns in mastersheet: {len(mastersheet_headers['course_cols'])}")
    
    # Process carryover students
    carryover_data = []
    resit_exam_col = resit_headers['registration_col']
    mastersheet_exam_col = mastersheet_headers['registration_col']
    
    # CRITICAL FIX: Load previous GPAs with enhanced function
    print(f"📊 Loading previous GPA data for CGPA calculation...")
    cgpa_data = load_previous_gpas_enhanced(mastersheet_path, semester_key)
    print(f"✅ Loaded previous GPA data for {len(cgpa_data)} students")
    
    for idx, resit_row in resit_df.iterrows():
        exam_no = str(resit_row[resit_exam_col]).strip().upper()
        if not exam_no or exam_no in ["NAN", "NONE", ""]:
            continue
        
        # Find student in mastersheet
        student_mask = (
            mastersheet_df[mastersheet_exam_col]
            .astype(str)
            .str.strip()
            .str.upper()
            == exam_no
        )
        
        if not student_mask.any():
            print(f"⚠️ Student {exam_no} not found in mastersheet")
            continue
        
        student_data = mastersheet_df[student_mask].iloc[0]
        student_name = "Unknown"
        
        if resit_headers['name_col'] and resit_headers['name_col'] in resit_row:
            student_name = str(resit_row[resit_headers['name_col']]).strip()
        
        # CRITICAL FIX: Get current GPA and credits from mastersheet
        current_gpa = 0.0
        current_credits = 0
        
        # Try to find GPA column in mastersheet
        gpa_col = None
        credits_col = None
        
        for col in mastersheet_df.columns:
            col_str = str(col).upper()
            if 'GPA' in col_str and 'CGPA' not in col_str:
                gpa_col = col
            if 'TCPE' in col_str or 'TOTAL CREDIT' in col_str or 'TOTAL UNIT' in col_str:
                credits_col = col
        
        if gpa_col and gpa_col in student_data:
            try:
                current_gpa = float(student_data[gpa_col]) if pd.notna(student_data[gpa_col]) else 0.0
            except (ValueError, TypeError):
                current_gpa = 0.0
        
        if credits_col and credits_col in student_data:
            try:
                current_credits = int(float(student_data[credits_col])) if pd.notna(student_data[credits_col]) else 30
            except (ValueError, TypeError):
                current_credits = 30
        
        student_record = {
            "EXAM NUMBER": exam_no,
            "NAME": student_name,
            "RESIT_COURSES": {},
            "CURRENT_GPA": current_gpa,
            "CURRENT_CREDITS": current_credits,
            "CURRENT_CGPA": 0,
            "PREVIOUS_GPAS": {}  # NEW: Store previous semester GPAs
        }
        
        # CRITICAL FIX: Load previous semester GPAs for display
        if exam_no in cgpa_data:
            previous_data = cgpa_data[exam_no]
            # Store individual previous GPAs for display
            for i, (prev_gpa, prev_credits) in enumerate(zip(previous_data["gpas"], previous_data["credits"])):
                semester_name = f"GPA_Semester_{i+1}"
                student_record[semester_name] = prev_gpa
            
            # Calculate CGPA
            student_record["CURRENT_CGPA"] = calculate_cgpa(
                previous_data, 
                current_gpa, 
                current_credits
            )
            print(f"📊 Calculated CGPA for {exam_no}: {student_record['CURRENT_CGPA']} (Previous GPAs: {previous_data['gpas']})")
        else:
            print(f"⚠️ No previous GPA data found for {exam_no}, using current GPA as CGPA")
            student_record["CURRENT_CGPA"] = current_gpa
        
        # Process resit courses
        for course_col in resit_headers['course_cols']:
            resit_score = resit_row.get(course_col)
            
            if detect_not_registered_content(resit_score) or pd.isna(resit_score):
                continue
            
            try:
                resit_score_val = float(resit_score)
            except (ValueError, TypeError):
                continue
            
            # Get original score from mastersheet
            original_score = None
            if course_col in mastersheet_df.columns:
                original_score = student_data.get(course_col)
            
            if pd.isna(original_score):
                continue
            
            try:
                original_score_val = float(original_score)
            except (ValueError, TypeError):
                original_score_val = 0.0
            
            # Only process if original was failure
            if original_score_val < pass_threshold:
                course_title = find_course_title(course_col, course_titles_dict, course_code_to_title)
                credit_unit = find_credit_unit(course_col, credit_units_dict, course_code_to_unit)
                
                student_record["RESIT_COURSES"][course_col] = {
                    "original_score": original_score_val,
                    "resit_score": resit_score_val,
                    "updated": resit_score_val >= pass_threshold,
                    "course_title": course_title,
                    "credit_unit": credit_unit,
                }
        
        if student_record["RESIT_COURSES"]:
            carryover_data.append(student_record)
            print(f"✅ Processed {exam_no}: {len(student_record['RESIT_COURSES'])} resit courses, GPA: {current_gpa}, CGPA: {student_record['CURRENT_CGPA']}")
    
    print(f"📊 Total carryover students processed: {len(carryover_data)}")
    return carryover_data


def build_updates_dict(carryover_data):
    """Build updates dictionary from carryover data."""
    updates = {}
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        updates[exam_no] = {}
        
        for course_code, course_data in student["RESIT_COURSES"].items():
            updates[exam_no][course_code] = course_data["resit_score"]
    
    return updates


def generate_carryover_outputs(carryover_data, output_dir, semester_key, set_name, timestamp,
                              course_titles_dict, credit_units_dict, course_code_to_title, course_code_to_unit):
    """Generate all carryover outputs - FIXED VERSION."""
    print(f"📊 Generating carryover outputs...")
    
    # Load CGPA data for mastersheet generation
    cgpa_data = {}
    
    # Generate carryover mastersheet
    carryover_mastersheet_path = generate_carryover_mastersheet(
        carryover_data, output_dir, semester_key, set_name, timestamp,
        cgpa_data, course_titles_dict, credit_units_dict, course_code_to_title, course_code_to_unit
    )
    
    if carryover_mastersheet_path and os.path.exists(carryover_mastersheet_path):
        print(f"✅ Carryover mastersheet created: {carryover_mastersheet_path}")
    else:
        print(f"❌ Failed to create carryover mastersheet")
    
    # Generate individual reports
    generate_individual_reports(carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data)
    
    # Save JSON records
    json_filepath = save_carryover_json_records(carryover_data, output_dir, semester_key)
    
    if json_filepath:
        copy_json_to_centralized_location(json_filepath, set_name, semester_key)
    
    # Create carryover ZIP - FIXED: Use the main output directory
    zip_filename = f"CARRYOVER_{set_name}_{semester_key}_{timestamp}.zip"
    zip_path = os.path.join(output_dir, zip_filename)
    
    print(f"📦 Creating carryover ZIP: {zip_path}")
    
    # Create ZIP with all carryover files
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add all files from the carryover output directory
            for root, dirs, files in os.walk(output_dir):
                for file in files:
                    if file.startswith("CARRYOVER_") or "CARRYOVER_RECORDS" in root or "INDIVIDUAL_REPORTS" in root:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, output_dir)
                        zipf.write(file_path, arcname)
                        print(f"✅ Added to ZIP: {arcname}")
        
        print(f"✅ Carryover ZIP created successfully: {zip_path}")
        return zip_path
        
    except Exception as e:
        print(f"❌ Error creating carryover ZIP: {e}")
        return None


def process_carryover_results_enhanced_with_cgpa(
    resit_file_path,
    source_path,
    source_type,
    semester_key,
    set_name,
    pass_threshold,
    output_dir
):
    """ENHANCED: Process carryover results with cumulative updates INCLUDING CGPA AND ANALYSIS SHEETS."""
    print(f"\n🔄 ENHANCED CARRYOVER PROCESSING WITH CGPA & ANALYSIS FOR {semester_key}")
    print("=" * 80)
    
    # Load course data
    (
        semester_course_titles,
        semester_credit_units,
        course_code_to_title,
        course_code_to_unit,
    ) = load_course_data()
    
    # Get semester info
    year, sem_num, level, sem_display, set_code, sem_name = get_semester_display_info(semester_key)
    
    # Find course mappings
    possible_sheet_keys = [
        f"{set_code} {sem_display}",
        f"{set_code.replace('NDII', 'ND II').replace('NDI', 'ND I')} {sem_display}",
        semester_key,
        semester_key.replace("-", " ").upper(),
        f"{set_code} {sem_name}",
        f"{level} {sem_display}",
    ]
    
    course_titles_dict = {}
    credit_units_dict = {}
    
    for sheet_key in possible_sheet_keys:
        sheet_standard = standardize_semester_key(sheet_key)
        if sheet_standard in semester_course_titles:
            course_titles_dict = semester_course_titles[sheet_standard]
            credit_units_dict = semester_credit_units[sheet_standard]
            print(f"✅ Using sheet key: '{sheet_key}' with {len(course_titles_dict)} courses")
            break
    
    if not course_titles_dict:
        course_titles_dict = course_code_to_title
        credit_units_dict = course_code_to_unit
    
    # Setup output directory
    timestamp = datetime.now().strftime(TIMESTAMP_FMT)
    carryover_output_dir = os.path.join(output_dir, f"CARRYOVER_{set_name}_{semester_key}_{timestamp}")
    os.makedirs(carryover_output_dir, exist_ok=True)
    
    # Validate resit file
    if not os.path.exists(resit_file_path):
        print(f"❌ Resit file not found: {resit_file_path}")
        return False
    
    temp_mastersheet_path = None
    temp_dir = None
    
    try:
        # Get mastersheet path
        temp_mastersheet_path, temp_dir = get_mastersheet_path(source_path, source_type, semester_key)
        if not temp_mastersheet_path:
            print(f"❌ Failed to get mastersheet")
            return False
        
        # Process carryover data
        carryover_data = process_carryover_core(
            resit_file_path, temp_mastersheet_path, semester_key, set_name,
            course_titles_dict, credit_units_dict, pass_threshold,
            course_code_to_title, course_code_to_unit
        )
        
        if not carryover_data:
            print("❌ No carryover data processed")
            return False
        
        print(f"✅ Successfully processed {len(carryover_data)} carryover students")
        
        # =============================================================
        # CRITICAL FIX: Generate carryover outputs FIRST
        # =============================================================
        print(f"\n📊 GENERATING CARRYOVER OUTPUTS...")
        
        # Build updates dictionary BEFORE generating outputs
        updates = build_updates_dict(carryover_data)
        
        # Generate all carryover outputs
        generate_carryover_outputs(
            carryover_data, carryover_output_dir, semester_key, set_name, timestamp,
            course_titles_dict, credit_units_dict, course_code_to_title, course_code_to_unit
        )
        
        # =============================================================
        # Apply cumulative updates to mastersheet WITH CGPA AND ANALYSIS
        # =============================================================
        print(f"\n🔄 APPLYING CUMULATIVE UPDATES WITH CGPA & ANALYSIS...")
        
        update_success = update_mastersheet_with_cumulative_updates_carryover(
            mastersheet_path=temp_mastersheet_path,
            updates=updates,
            semester_key=semester_key,
            original_zip_path=source_path,
            course_titles_dict=course_titles_dict,
            course_units_dict=credit_units_dict,
            set_name=set_name,
            clean_dir=output_dir
        )
        
        if update_success:
            print(f"\n🎉 CUMULATIVE UPDATE WITH CGPA & ANALYSIS COMPLETE!")
            print(f"✅ Original preserved as backup")
            print(f"✅ New version created with all updates")
            print(f"✅ CGPA Summary and Analysis sheets updated")
            print(f"✅ Withdrawn students handled properly")
            
            # =============================================================
            # CRITICAL: Verify carryover outputs were created
            # =============================================================
            print(f"\n🔍 VERIFYING CARRYOVER OUTPUTS...")
            
            # Check for CARRYOVER mastersheet
            carryover_files = []
            for file in os.listdir(carryover_output_dir):
                if file.startswith("CARRYOVER_") and file.endswith(".xlsx"):
                    carryover_files.append(file)
                    print(f"✅ Found carryover mastersheet: {file}")
                elif file.startswith("CARRYOVER_") and file.endswith(".zip"):
                    carryover_files.append(file)
                    print(f"✅ Found carryover ZIP: {file}")
            
            # Check for individual reports
            reports_dir = os.path.join(carryover_output_dir, "INDIVIDUAL_REPORTS")
            if os.path.exists(reports_dir):
                report_files = os.listdir(reports_dir)
                print(f"✅ Found {len(report_files)} individual reports")
                carryover_files.extend([f"INDIVIDUAL_REPORTS/{f}" for f in report_files])
            
            # Check for JSON records
            json_dir = os.path.join(carryover_output_dir, "CARRYOVER_RECORDS")
            if os.path.exists(json_dir):
                json_files = os.listdir(json_dir)
                print(f"✅ Found {len(json_files)} JSON records")
                carryover_files.extend([f"CARRYOVER_RECORDS/{f}" for f in json_files])
            
            print(f"📁 Total carryover files generated: {len(carryover_files)}")
            
            # Copy the main CARRYOVER files to the main output directory for easy access
            main_carryover_zip = None
            for file in os.listdir(carryover_output_dir):
                if file.startswith("CARRYOVER_") and file.endswith(".zip"):
                    source_zip = os.path.join(carryover_output_dir, file)
                    dest_zip = os.path.join(output_dir, file)
                    shutil.copy2(source_zip, dest_zip)
                    main_carryover_zip = dest_zip
                    print(f"📦 Copied main carryover ZIP to: {dest_zip}")
                    break
            
            return True
        else:
            print(f"❌ Cumulative update failed")
            return False
            
    except Exception as e:
        print(f"❌ Error in enhanced carryover processing: {e}")
        traceback.print_exc()
        return False
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)


# ----------------------------
# Carryover Processing Functions
# ----------------------------
def load_carryover_files(carryover_dir, semester_key=None):
    """Load carryover files - FIXED to handle both JSON and Excel"""
    carryover_files = []
    for file in os.listdir(carryover_dir):
        # Look for both JSON and Excel carryover files
        if file.startswith("co_student_") and (
            file.endswith(".json") or file.endswith(".xlsx")
        ):
            file_semester = extract_semester_from_filename(file)
            file_semester_standardized = (
                standardize_semester_name(file_semester) if file_semester else None
            )

            # Skip if semester doesn't match filter
            if semester_key and file_semester_standardized != semester_key:
                continue

            file_path = os.path.join(carryover_dir, file)
            try:
                if file.endswith(".json"):
                    with open(file_path, "r") as f:
                        data = json.load(f)
                else:  # .xlsx
                    df = pd.read_excel(file_path)
                    data = df.to_dict("records")

                carryover_files.append(
                    {
                        "filename": file,
                        "semester": file_semester_standardized,
                        "data": data,
                        "count": len(data),
                        "file_path": file_path,
                    }
                )
                print(f"✅ Loaded {len(data)} records from {file}")

            except Exception as e:
                print(f"❌ Error loading {file}: {e}")
    if not carryover_files:
        print(f"❌ No carryover files found in {carryover_dir}")
        return []
    print(f"📚 Total carryover files loaded: {len(carryover_files)}")
    return carryover_files


def save_carryover_json_records(carryover_data, carryover_output_dir, semester_key):
    """
    Save carryover records as JSON files
    """
    json_dir = os.path.join(carryover_output_dir, "CARRYOVER_RECORDS")
    os.makedirs(json_dir, exist_ok=True)
    print(f"\n💾 SAVING CARRYOVER JSON RECORDS")
    print(f"📁 JSON directory: {json_dir}")
    # Convert carryover_data to JSON-friendly format
    json_records = {}
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]

        # Build the JSON record structure
        record = {
            "exam_number": exam_no,
            "name": student["NAME"],
            "carryover_courses": {},
            "passed_resit_courses": {},
            "failed_resit_courses": {},
        }

        # Process each resit course
        for course_code, course_data in student["RESIT_COURSES"].items():
            course_info = {
                "course_code": course_code,
                "course_title": course_data.get("course_title", course_code),
                "credit_unit": course_data.get("credit_unit", 2),
                "original_score": course_data["original_score"],
                "resit_score": course_data["resit_score"],
            }

            # Add to carryover courses (all courses that were taken as resit)
            record["carryover_courses"][course_code] = course_info

            # Separate into passed/failed based on resit score
            if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD:
                record["passed_resit_courses"][course_code] = course_info
            else:
                record["failed_resit_courses"][course_code] = course_info

        json_records[exam_no] = record
    # Save as JSON file with timestamp and semester
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f"co_student_{semester_key}_{timestamp}.json"
    json_filepath = os.path.join(json_dir, json_filename)
    try:
        with open(json_filepath, "w", encoding="utf-8") as f:
            json.dump(json_records, f, indent=2, ensure_ascii=False)

        print(f"✅ Saved JSON carryover records: {json_filepath}")
        print(f"📊 Records saved: {len(json_records)} students")

        return json_filepath

    except Exception as e:
        print(f"❌ Error saving JSON records: {e}")
        traceback.print_exc()
        return None


def copy_json_to_centralized_location(json_filepath, set_name, semester_key):
    """
    Copy JSON file to centralized CARRYOVER_RECORDS location
    """
    try:
        # Determine the centralized location
        base_dir = get_base_directory()
        centralized_dir = os.path.join(
            base_dir,
            "EXAMS_INTERNAL",
            "ND",
            set_name,
            "CLEAN_RESULTS",
            "CARRYOVER_RECORDS",
        )

        os.makedirs(centralized_dir, exist_ok=True)

        # Copy the JSON file
        filename = os.path.basename(json_filepath)
        dest_path = os.path.join(centralized_dir, filename)

        shutil.copy2(json_filepath, dest_path)

        print(f"\n📋 COPIED TO CENTRALIZED LOCATION")
        print(f"✅ From: {json_filepath}")
        print(f"✅ To: {dest_path}")

        return dest_path

    except Exception as e:
        print(f"❌ Error copying to centralized location: {e}")
        traceback.print_exc()
        return None


def generate_carryover_mastersheet(
    carryover_data,
    output_dir,
    semester_key,
    set_name,
    timestamp,
    cgpa_data,
    course_titles,
    course_units,
    course_code_to_title,
    course_code_to_unit,
):
    """Generate CARRYOVER_mastersheet - FIXED VERSION WITH PREVIOUS GPA DISPLAY."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ND_CARRYOVER_RESULTS"
    program_name = "NATIONAL DIPLOMA"
    program_abbr = "ND"
    if os.path.exists(DEFAULT_LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image

            img = Image(DEFAULT_LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"⚠️ Could not add logo: {e}")
    current_year = 2025
    next_year = 2026
    year, sem_num, level, sem_display, set_code, current_semester_name = (
        get_semester_display_info(semester_key)
    )
    all_courses = set()
    for student in carryover_data:
        all_courses.update(student["RESIT_COURSES"].keys())
    previous_semesters = get_previous_semesters_for_display(semester_key)
    headers = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        headers.append(f"GPA {prev_sem}")
    course_headers = []
    for course in sorted(all_courses):
        course_headers.extend([f"{course}", f"{course}_RESIT"])
    headers.extend(course_headers)
    headers.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    total_columns = len(headers)
    last_column = get_column_letter(total_columns)
    ws.merge_cells(f"A3:{last_column}3")
    title_cell = ws["A3"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A4:{last_column}4")
    subtitle_cell = ws["A4"]
    subtitle_cell.value = f"RESIT - {current_year}/{next_year} SESSION {program_name} {level} {sem_display} EXAMINATIONS RESULT — {datetime.now().strftime('%B %d, %Y')}"
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    print(f"🔍 Courses found in resit data: {sorted(all_courses)}")
    print(
        f"📊 GPA columns for {semester_key}: Previous={previous_semesters}, Current={current_semester_name}"
    )
    headers = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        headers.append(f"GPA {prev_sem}")
    course_headers = []
    course_title_mapping = {}
    course_unit_mapping = {}
    for course in sorted(all_courses):
        course_title = find_course_title(course, course_titles, course_code_to_title)
        course_title_mapping[course] = course_title

        credit_unit = find_credit_unit(course, course_units, course_code_to_unit)
        course_unit_mapping[course] = credit_unit

        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        course_headers.extend([f"{course}", f"{course}_RESIT"])
    headers.extend(course_headers)
    headers.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    title_row = [""] * 3
    for prev_sem in previous_semesters:
        title_row.extend([""])
    for course in sorted(all_courses):
        course_title = course_title_mapping[course]
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        title_row.extend([course_title, course_title])
    title_row.extend(["", "", ""])
    ws.append(title_row)
    credit_row = [""] * 3
    for prev_sem in previous_semesters:
        credit_row.extend([""])
    for course in sorted(all_courses):
        credit_unit = course_unit_mapping[course]
        credit_row.extend([f"CU: {credit_unit}", f"CU: {credit_unit}"])
    credit_row.extend(["", "", ""])
    ws.append(credit_row)
    code_row = ["S/N", "EXAM NUMBER", "NAME"]
    for prev_sem in previous_semesters:
        code_row.append(f"GPA {prev_sem}")
    for course in sorted(all_courses):
        code_row.append(f"{course}")
        code_row.append(f"{course}_RESIT")
    code_row.extend([f"GPA {current_semester_name}", "CGPA", "REMARKS"])
    ws.append(code_row)
    course_colors = [
        "E6F3FF",
        "FFF0E6",
        "E6FFE6",
        "FFF6E6",
        "F0E6FF",
        "E6FFFF",
        "FFE6F2",
        "F5F5DC",
        "E6F7FF",
        "FFF5E6",
    ]
    start_col = 4
    if previous_semesters:
        start_col += len(previous_semesters)
    color_index = 0
    for course in sorted(all_courses):
        for row in [5, 6, 7]:
            for offset in [0, 1]:
                cell = ws.cell(row=row, column=start_col + offset)
                cell.fill = PatternFill(
                    start_color=course_colors[color_index % len(course_colors)],
                    end_color=course_colors[color_index % len(course_colors)],
                    fill_type="solid",
                )
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        for offset in [0, 1]:
            cell = ws.cell(row=5, column=start_col + offset)
            cell.alignment = Alignment(
                text_rotation=90, horizontal="center", vertical="center"
            )
            cell.font = Font(bold=True, size=9)

        color_index += 1
        start_col += 2
    for row in [5, 6, 7]:
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        gpa_col = 4
        for prev_sem in previous_semesters:
            cell = ws.cell(row=row, column=gpa_col)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            gpa_col += 1

        for col in range(len(headers) - 2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
    row_idx = 8
    failed_counts = {course: 0 for course in all_courses}
    start_col = 4
    if previous_semesters:
        start_col += len(previous_semesters)
    # CRITICAL FIX: Proper serial numbers from 1 to n
    serial_number = 1

    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]

        # Serial Number (PROPERLY SORTED from 1 to n)
        ws.cell(row=row_idx, column=1, value=serial_number)
        serial_number += 1

        ws.cell(row=row_idx, column=2, value=student["EXAM NUMBER"])
        ws.cell(row=row_idx, column=3, value=student["NAME"])

        # CRITICAL FIX: Write previous GPAs using the stored semester GPAs
        gpa_col = 4
        for i, prev_sem in enumerate(previous_semesters):
            gpa_key = f"GPA_Semester_{i+1}"
            gpa_value = student.get(gpa_key, "")
            ws.cell(row=row_idx, column=gpa_col, value=gpa_value)
            
            # Format GPA cell
            if gpa_value:
                try:
                    gpa_val = float(gpa_value)
                    if gpa_val >= 3.5:
                        ws.cell(row=row_idx, column=gpa_col).font = Font(bold=True, color="006400")
                    elif gpa_val < 2.0:
                        ws.cell(row=row_idx, column=gpa_col).font = Font(bold=True, color="FF0000")
                except (ValueError, TypeError):
                    pass
            
            gpa_col += 1

        course_col = gpa_col
        color_index = 0
        for course in sorted(all_courses):
            for offset in [0, 1]:
                cell = ws.cell(row=row_idx, column=course_col + offset)
                cell.fill = PatternFill(
                    start_color=course_colors[color_index % len(course_colors)],
                    end_color=course_colors[color_index % len(course_colors)],
                    fill_type="solid",
                )

            if course in student["RESIT_COURSES"]:
                course_data = student["RESIT_COURSES"][course]

                orig_cell = ws.cell(
                    row=row_idx, column=course_col, value=course_data["original_score"]
                )
                if course_data["original_score"] < DEFAULT_PASS_THRESHOLD:
                    orig_cell.fill = PatternFill(
                        start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                    )

                resit_cell = ws.cell(
                    row=row_idx, column=course_col + 1, value=course_data["resit_score"]
                )
                if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD:
                    resit_cell.fill = PatternFill(
                        start_color="90EE90", end_color="90EE90", fill_type="solid"
                    )
                else:
                    resit_cell.fill = PatternFill(
                        start_color="FFD580", end_color="FFD580", fill_type="solid"
                    )
                    failed_counts[course] += 1
            else:
                ws.cell(row=row_idx, column=course_col, value="")
                ws.cell(row=row_idx, column=course_col + 1, value="")

            color_index += 1
            course_col += 2

        ws.cell(row=row_idx, column=course_col, value=student["CURRENT_GPA"])
        ws.cell(row=row_idx, column=course_col + 1, value=student["CURRENT_CGPA"])

        remarks = generate_remarks(student["RESIT_COURSES"])
        ws.cell(row=row_idx, column=course_col + 2, value=remarks)

        row_idx += 1
    failed_row_idx = row_idx
    ws.cell(row=failed_row_idx, column=1, value="FAILED COUNT BY COURSE:").font = Font(
        bold=True
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=failed_row_idx, column=col)
        cell.fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
    course_col = gpa_col
    for course in sorted(all_courses):
        count_cell = ws.cell(
            row=failed_row_idx, column=course_col + 1, value=failed_counts[course]
        )
        count_cell.font = Font(bold=True)
        count_cell.fill = PatternFill(
            start_color="FFFF99", end_color="FFFF99", fill_type="solid"
        )
        course_col += 2
    summary_start_row = failed_row_idx + 2
    total_students = len(carryover_data)
    passed_all = sum(
        1
        for student in carryover_data
        if all(
            course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
            for course_data in student["RESIT_COURSES"].values()
        )
    )
    carryover_count = total_students - passed_all
    total_failed_attempts = sum(failed_counts.values())
    summary_data = [
        ["CARRYOVER SUMMARY"],
        [
            f"A total of {total_students} students registered and sat for the Carryover Examination"
        ],
        [f"A total of {passed_all} students passed all carryover courses"],
        [
            f"A total of {carryover_count} students failed one or more carryover courses and must repeat them"
        ],
        [f"Total failed resit attempts: {total_failed_attempts} across all courses"],
        [
            f"Carryover processing completed on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"
        ],
        [""],
        [""],
        ["", ""],
        ["________________________", "________________________"],
        ["Mrs. Abini Hauwa", "Mrs. Olukemi Ogunleye"],
        ["Head of Exams", "Chairman, ND Program C'tee"],
    ]
    for i, row_data in enumerate(summary_data):
        row_num = summary_start_row + i
        if len(row_data) == 1:
            if row_data[0]:
                ws.merge_cells(
                    start_row=row_num, start_column=1, end_row=row_num, end_column=10
                )
                cell = ws.cell(row=row_num, column=1, value=row_data[0])
                if i == 0:
                    cell.font = Font(bold=True, size=12, underline="single")
                else:
                    cell.font = Font(bold=False, size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        elif len(row_data) == 2:
            left_cell = ws.cell(row=row_num, column=1, value=row_data[0])
            right_cell = ws.cell(row=row_num, column=4, value=row_data[1])

            if i >= len(summary_data) - 3:
                left_cell.alignment = Alignment(horizontal="left")
                right_cell.alignment = Alignment(horizontal="left")
                left_cell.font = Font(bold=True, size=11)
                right_cell.font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=7, max_row=row_idx - 1, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border = thin_border
    ws.freeze_panes = "D8"
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or not cell.font.bold:
                cell.font = Font(name="Calibri", size=11)
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in column:
            try:
                if cell.value is not None:
                    cell_value = str(cell.value)
                    cell_length = len(cell_value)

                    if cell.row == 5 and cell.alignment.text_rotation == 90:
                        cell_length = max(cell_length, 10)

                    if isinstance(cell.value, (int, float)) and not isinstance(
                        cell.value, bool
                    ):
                        cell_length = max(cell_length, 8)

                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, 50)

        if col_idx == 1:
            adjusted_width = 8
        elif col_idx == 2:
            adjusted_width = 18
        elif col_idx == 3:
            adjusted_width = 35
        elif col_idx >= 4 and col_idx <= (4 + len(previous_semesters) - 1):
            adjusted_width = 15
        elif col_idx >= len(headers) - 2:
            adjusted_width = 15
        else:
            adjusted_width = min(max(adjusted_width, 12), 25)

        ws.column_dimensions[column_letter].width = adjusted_width
    for row_idx in range(8, row_idx):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                if (
                    cell.fill.start_color.index == "00000000"
                    or cell.fill.start_color.index == "00FFFFFF"
                ):
                    cell.fill = PatternFill(
                        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
                    )
    gpa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    if previous_semesters:
        for row in range(8, row_idx):
            for col in range(4, 4 + len(previous_semesters)):
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.index == "00000000":
                    cell.fill = gpa_fill
    final_gpa_fill = PatternFill(
        start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"
    )
    for row in range(8, row_idx):
        for col in range(len(headers) - 2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.start_color.index == "00000000":
                cell.fill = final_gpa_fill
    filename = f"CARRYOVER_mastersheet_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    wb.close()  # ✅ CRITICAL: Properly close the workbook
    print(f"✅ CARRYOVER mastersheet generated: {filepath}")
    return filepath


def generate_individual_reports(
    carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data
):
    """Generate individual student reports."""
    reports_dir = os.path.join(output_dir, "INDIVIDUAL_REPORTS")
    os.makedirs(reports_dir, exist_ok=True)
    for student in carryover_data:
        exam_no = student["EXAM NUMBER"]
        safe_exam_no = sanitize_filename(exam_no)

        filename = f"carryover_report_{safe_exam_no}_{timestamp}.csv"
        filepath = os.path.join(reports_dir, filename)

        report_data = []
        report_data.append(["ND CARRYOVER RESULT REPORT"])
        report_data.append(["FCT COLLEGE OF NURSING SCIENCES"])
        report_data.append([f"ND Set: {set_name}"])
        report_data.append([f"ND Semester: {semester_key}"])
        report_data.append([])
        report_data.append(["ND STUDENT INFORMATION"])
        report_data.append(["Exam Number:", student["EXAM NUMBER"]])
        report_data.append(["Name:", student["NAME"]])
        report_data.append([])

        report_data.append(["ND PREVIOUS GPAs"])
        for key in sorted([k for k in student.keys() if k.startswith("GPA_")]):
            semester = key.replace("GPA_", "")
            report_data.append([f"{semester}:", student[key]])
        report_data.append([])

        report_data.append(["ND CURRENT ACADEMIC RECORD"])
        report_data.append(["Current GPA:", student["CURRENT_GPA"]])
        report_data.append(["Current CGPA:", student["CURRENT_CGPA"]])
        report_data.append([])

        report_data.append(["ND RESIT COURSES"])
        report_data.append(
            [
                "Course Code",
                "Course Title",
                "Credit Unit",
                "Original Score",
                "Resit Score",
                "Status",
            ]
        )

        for course_code, course_data in student["RESIT_COURSES"].items():
            status = (
                "PASSED"
                if course_data["resit_score"] >= DEFAULT_PASS_THRESHOLD
                else "FAILED"
            )
            course_title = course_data.get("course_title", course_code)
            credit_unit = course_data.get("credit_unit", 0)
            report_data.append(
                [
                    course_code,
                    course_title,
                    credit_unit,
                    course_data["original_score"],
                    course_data["resit_score"],
                    status,
                ]
            )

        try:
            df = pd.DataFrame(report_data)
            df.to_csv(filepath, index=False, header=False)
            print(f"✅ Generated report for: {exam_no}")
        except Exception as e:
            print(f"❌ Error generating report for {exam_no}: {e}")
    print(
        f"✅ Generated {len(carryover_data)} individual student reports in {reports_dir}"
    )


# ----------------------------
# MAIN FUNCTION - ENHANCED WITH CGPA AND ANALYSIS
# ----------------------------
def main_enhanced_with_cgpa():
    """Enhanced main function with cumulative updates INCLUDING CGPA AND ANALYSIS."""
    print("=" * 80)
    print("🎯 ENHANCED CARRYOVER PROCESSOR - CUMULATIVE UPDATES WITH CGPA & ANALYSIS")
    print("🔄 MAINTAINS SINGLE BACKUP & PERSISTENT UPDATES")
    print("📊 INCLUDES CGPA SUMMARY AND ANALYSIS SHEETS")
    print("=" * 80)
    
    # Get environment variables
    set_name = os.getenv("SELECTED_SET", "")
    semester_key = os.getenv("SELECTED_SEMESTERS", "")
    resit_file_path = os.getenv("RESIT_FILE_PATH", "")
    base_result_path = os.getenv("BASE_RESULT_PATH", "")
    output_dir_env = os.getenv("OUTPUT_DIR", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", str(DEFAULT_PASS_THRESHOLD)))
    
    # Validate inputs
    if not set_name:
        print("❌ ERROR: SELECTED_SET not provided")
        return
    
    if not semester_key:
        print("❌ ERROR: SELECTED_SEMESTERS not provided")
        return
    
    if not resit_file_path or not os.path.exists(resit_file_path):
        print(f"❌ ERROR: RESIT_FILE_PATH not provided or doesn't exist: {resit_file_path}")
        return
    
    if not set_name.startswith("ND-"):
        print(f"❌ ERROR: Invalid ND set name: {set_name}")
        return
    
    print(f"✅ Processing ND Set: {set_name}")
    print(f"✅ Processing Semester: {semester_key}")
    print(f"✅ Resit file: {resit_file_path}")
    
    # Find directories
    clean_dir = None
    output_dir = None
    possible_base_dirs = [
        BASE_DIR,
        os.path.join(BASE_DIR, "EXAMS_INTERNAL"),
        os.path.join(os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"),
    ]
    
    for base in possible_base_dirs:
        test_clean_dir = os.path.join(base, "ND", set_name, "CLEAN_RESULTS")
        if os.path.exists(test_clean_dir):
            clean_dir = test_clean_dir
            output_dir = output_dir_env if output_dir_env else test_clean_dir
            print(f"✅ Found clean directory: {clean_dir}")
            break
    
    if not clean_dir:
        print(f"❌ ERROR: Clean directory not found for {set_name}")
        return
    
    print(f"📁 Clean directory: {clean_dir}")
    print(f"📁 Output directory: {output_dir}")
    
    # Get mastersheet source (prefers UPDATED_ files)
    if base_result_path and os.path.exists(base_result_path):
        print(f"✅ Using provided base result path: {base_result_path}")
        source_path = base_result_path
        source_type = "zip" if base_result_path.endswith(".zip") else "folder"
    else:
        source_path, source_type = get_mastersheet_source(clean_dir, set_name)
        if not source_path:
            print(f"❌ ERROR: No suitable mastersheet source found")
            return
    
    print(f"✅ Using source: {source_path} (type: {source_type})")
    
    # Check if this is the first run
    latest_updated = find_latest_updated_file(clean_dir)
    if latest_updated:
        print(f"🔄 Continuing from: {os.path.basename(latest_updated)}")
    else:
        print("🔄 Starting fresh - will create UPDATED_1")
    
    # Process carryover results with enhanced cumulative updates INCLUDING CGPA AND ANALYSIS
    success = process_carryover_results_enhanced_with_cgpa(
        resit_file_path=resit_file_path,
        source_path=source_path,
        source_type=source_type,
        semester_key=semester_key,
        set_name=set_name,
        pass_threshold=pass_threshold,
        output_dir=output_dir
    )
    
    if success:
        print("\n" + "=" * 80)
        print("✅ ENHANCED CARRYOVER PROCESSING WITH CGPA & ANALYSIS COMPLETED SUCCESSFULLY!")
        print("=" * 80)
        print("📁 Check the CLEAN_RESULTS directory for:")
        print("   - UPDATED_* files (cumulative updates with CGPA & Analysis)")
        print("   - CARRYOVER_* files (carryover results)")
        print("   - Original_BACKUP.zip (single backup)")
        print("📊 New sheets included:")
        print("   - CGPA_SUMMARY (with professional formatting)")
        print("   - ANALYSIS (with semester statistics)")
        print(f"💡 Location: {output_dir}")
    else:
        print("\n" + "=" * 80)
        print("❌ ENHANCED CARRYOVER PROCESSING WITH CGPA & ANALYSIS FAILED")
        print("=" * 80)


# ----------------------------
# Script Execution
# ----------------------------
if __name__ == "__main__":
    try:
        main_enhanced_with_cgpa()
    except Exception as e:
        print(f"❌ Fatal error in enhanced execution: {e}")
        traceback.print_exc()
        sys.exit(1)