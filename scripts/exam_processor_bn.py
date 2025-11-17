#!/usr/bin/env python3
"""
exam_processor_bn.py - Enhanced BN Examination Processor with ND Logic
Complete script with flexible threshold upgrade rule for BN results.
Enhanced with transposed data transformation, carryover management,
CGPA tracking, analysis sheet, and comprehensive BN course matching.
Web-compatible version with file upload support.
UPDATED WITH FIXES:
- Issue 1: STATUS Column Now Properly Reflects REMARKS from individual semester sheets
- Issue 2: Previous CGPA Loading Issues Fixed with Proper Header Detection
- Issue 3: Status Determination Logic Enhanced with Better Debugging
- Header Section Adjusted: Logo placement fixed to not cover headings and student data
- All functions updated to use proper header detection and status capture
- FIXED: CGPA_SUMMARY sheet position now moves towards end after last semester
- NEW: NBTE-compliant CGPA calculation using individual course quality points
"""

from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import re
import pandas as pd
from datetime import datetime
import platform
import difflib
import math
import glob
import tempfile
import shutil
import json
import zipfile
import time
import traceback
import logging

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ----------------------------
# Logging Configuration
# ----------------------------
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# ----------------------------
# Configuration
# ----------------------------
def is_running_on_railway():
    """Check if we're running on Railway"""
    return any(
        key in os.environ
        for key in [
            "RAILWAY_ENVIRONMENT",
            "RAILWAY_STATIC_URL",
            "RAILWAY_PROJECT_ID",
            "RAILWAY_SERVICE_NAME",
        ]
    )

def get_base_directory():
    """Get base directory - compatible with both local and Railway environments"""
    # Check Railway environment first
    railway_base = os.getenv("BASE_DIR")
    if railway_base and os.path.exists(railway_base):
        return railway_base
    
    # Check if we're running on Railway but BASE_DIR doesn't exist
    if is_running_on_railway():
        # Create the directory structure on Railway
        railway_base = "/app/EXAMS_INTERNAL"
        os.makedirs(railway_base, exist_ok=True)
        os.makedirs(os.path.join(railway_base, "BN", "BN-COURSES"), exist_ok=True)
        return railway_base
    
    # Local development fallback - updated to match new structure
    local_path = os.path.join(
        os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"
    )
    if os.path.exists(local_path):
        return local_path
    
    # Final fallback - current directory
    return os.path.join(os.path.dirname(__file__), "EXAMS_INTERNAL")

BASE_DIR = get_base_directory()

# UPDATED: BN directories now under BN folder
BN_BASE_DIR = os.path.join(BASE_DIR, "BN")
BN_COURSES_DIR = os.path.join(BN_BASE_DIR, "BN-COURSES")

# Ensure directories exist
os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(BN_BASE_DIR, exist_ok=True)
os.makedirs(BN_COURSES_DIR, exist_ok=True)

# Define semester processing order for BN
BN_SEMESTER_ORDER = [
    "N-FIRST-YEAR-FIRST-SEMESTER",
    "N-FIRST-YEAR-SECOND-SEMESTER", 
    "N-SECOND-YEAR-FIRST-SEMESTER",
    "N-SECOND-YEAR-SECOND-SEMESTER",
    "N-THIRD-YEAR-FIRST-SEMESTER",
    "N-THIRD-YEAR-SECOND-SEMESTER",
]

# Short semester names mapping
SHORT_SEMESTER_MAP = {
    "N-FIRST-YEAR-FIRST-SEMESTER": "NY1S1",
    "N-FIRST-YEAR-SECOND-SEMESTER": "NY1S2",
    "N-SECOND-YEAR-FIRST-SEMESTER": "NY2S1",
    "N-SECOND-YEAR-SECOND-SEMESTER": "NY2S2",
    "N-THIRD-YEAR-FIRST-SEMESTER": "NY3S1",
    "N-THIRD-YEAR-SECOND-SEMESTER": "NY3S2",
}

# Global variables for threshold upgrade
THRESHOLD_UPGRADED = False
ORIGINAL_THRESHOLD = 50.0
UPGRADE_MIN = 0  # Defaults to 0 (disabled) - no upgrades unless explicitly set
UPGRADE_MAX = 49

# Global student tracker
STUDENT_TRACKER = {}
WITHDRAWN_STUDENTS = {}
CARRYOVER_STUDENTS = {}  # New global carryover tracker

def is_web_mode():
    """Check if running in web mode (file upload)"""
    return os.getenv("WEB_MODE") == "true"

def get_uploaded_file_path():
    """Get path of uploaded file in web mode"""
    return os.getenv("UPLOADED_FILE_PATH")

def should_use_interactive_mode():
    """Check if we should use interactive mode (CLI) or non-interactive mode (web)."""
    # If specific environment variables are set by web form, use non-interactive
    if os.getenv("SELECTED_SET") or os.getenv("PROCESSING_MODE") or is_web_mode():
        return False
    # If we're running in a terminal with stdin available, use interactive mode
    if sys.stdin.isatty():
        return True
    # Default to interactive for backward compatibility
    return True

def validate_raw_file(file_path):
    """Validate raw Excel file before processing."""
    logger.info(f"🔍 Validating file: {file_path}")
    if not os.path.exists(file_path):
        logger.error(f"❌ File does not exist: {file_path}")
        return False, "File does not exist"
    if os.path.getsize(file_path) == 0:
        logger.error(f"❌ File is empty: {file_path}")
        return False, "File is empty"
    try:
        xl = pd.ExcelFile(file_path)
        if not xl.sheet_names:
            logger.error(f"❌ No sheets in Excel file: {file_path}")
            return False, "No sheets in Excel file"
        expected_sheets = ["CA", "OBJ", "EXAM"]
        found_sheets = [s for s in expected_sheets if s in xl.sheet_names]
        if not found_sheets:
            logger.error(
                f"❌ No expected sheets found in {file_path}. Has: {xl.sheet_names}"
            )
            return False, f"No expected sheets found. Has: {xl.sheet_names}"
        logger.info(f"✅ Valid file with sheets: {found_sheets}")
        return True, f"Valid file with sheets: {found_sheets}"
    except Exception as e:
        logger.error(f"❌ Cannot open file {file_path}: {e}")
        return False, f"Cannot open file: {e}"

def process_uploaded_file(uploaded_file_path, base_dir_norm):
    """
    Process uploaded file in web mode.
    This function handles the single uploaded file for web processing.
    """
    logger.info("🔧 Processing uploaded file in web mode")
    # Validate the uploaded file first
    is_valid, validation_msg = validate_raw_file(uploaded_file_path)
    if not is_valid:
        logger.error(f"❌ Uploaded file validation failed: {validation_msg}")
        return False
    
    # Extract set name from filename or use default
    filename = os.path.basename(uploaded_file_path)
    set_name = "BN-UPLOADED"
    
    # Create temporary directory structure
    temp_dir = tempfile.mkdtemp()
    raw_dir = os.path.join(temp_dir, set_name, "RAW_RESULTS")
    clean_dir = os.path.join(temp_dir, set_name, "CLEAN_RESULTS")
    os.makedirs(raw_dir, exist_ok=True)
    os.makedirs(clean_dir, exist_ok=True)
    
    # Copy uploaded file to raw directory
    dest_path = os.path.join(raw_dir, filename)
    shutil.copy2(uploaded_file_path, dest_path)
    
    # Get parameters from environment
    params = get_form_parameters()
    ts = datetime.now().strftime(TIMESTAMP_FMT)
    
    try:
        # Load course data
        (
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
        ) = load_bn_course_data()
        
        # Process the single file
        raw_files = [filename]
        # Detect semester from filename
        semester_key = detect_bn_semester_from_filename(filename)
        logger.info(f"🎯 Detected semester: {semester_key}")
        logger.info(f"📁 Processing uploaded file: {filename}")
        
        # Process the file
        result = process_bn_single_file(
            dest_path,
            raw_dir,
            clean_dir,
            ts,
            params["pass_threshold"],
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
            DEFAULT_LOGO_PATH,
            semester_key,
            set_name,
            previous_gpas=None,
            upgrade_min_threshold=get_upgrade_threshold_from_env(),
        )
        
        if result is not None:
            logger.info("✅ Successfully processed uploaded file")
            return True
        else:
            logger.error("❌ Failed to process uploaded file")
            return False
    except Exception as e:
        logger.error(f"❌ Error processing uploaded file: {e}")
        traceback.print_exc()
        return False
    finally:
        # Clean up temporary directory
        shutil.rmtree(temp_dir, ignore_errors=True)

def get_upgrade_threshold_from_env():
    """Get upgrade threshold from environment variables"""
    upgrade_threshold_str = os.getenv(
        "UPGRADE_THRESHOLD", "0"
    ).strip()  # Default to '0'
    if upgrade_threshold_str and upgrade_threshold_str.isdigit():
        upgrade_value = int(upgrade_threshold_str)
        if upgrade_value == 0:
            return None  # 0 means disabled
        if 45 <= upgrade_value <= 49:
            return upgrade_value
    return None  # Default to disabled (no upgrade)

def check_bn_files_exist(raw_dir, semester_key):
    """Check if BN files actually exist for the given semester"""
    if not os.path.exists(raw_dir):
        logger.error(f"❌ Raw directory doesn't exist: {raw_dir}")
        return False
    
    raw_files = [
        f
        for f in os.listdir(raw_dir)
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
    ]
    if not raw_files:
        logger.error(f"❌ No Excel files found in: {raw_dir}")
        return False
    
    # Check if any files match the semester
    semester_files = []
    for rf in raw_files:
        detected_sem = detect_bn_semester_from_filename(rf)
        if detected_sem == semester_key:
            semester_files.append(rf)
    
    if not semester_files:
        logger.error(f"❌ No files found for semester {semester_key}")
        logger.info(f" Available files: {raw_files}")
        return False
    
    logger.info(
        f"✅ Found {len(semester_files)} files for {semester_key}: {semester_files}"
    )
    return True

def process_in_non_interactive_mode(params, base_dir_norm):
    """Process exams in non-interactive mode for web interface."""
    logger.info("🔧 Running in NON-INTERACTIVE mode (web interface)")
    
    # Use parameters from environment variables
    selected_set = params["selected_set"]
    processing_mode = params["processing_mode"]
    selected_semesters = params["selected_semesters"]
    
    # Get upgrade threshold from environment variable if provided
    upgrade_min_threshold = get_upgrade_threshold_from_env()
    
    # Get available sets
    available_sets = get_available_bn_sets(base_dir_norm)
    if not available_sets:
        logger.error("❌ No BN sets found")
        return False
    
    # Remove BN-COURSES from available sets if present
    available_sets = [s for s in available_sets if s != "BN-COURSES"]
    if not available_sets:
        logger.error("❌ No valid BN sets found (only BN-COURSES present)")
        return False
    
    # Determine which sets to process - FIXED: Only process the selected set
    if selected_set == "all":
        sets_to_process = available_sets
        logger.info(f"🎯 Processing ALL sets: {sets_to_process}")
    else:
        # CRITICAL FIX: Only process the specifically selected set
        if selected_set in available_sets:
            sets_to_process = [selected_set]
            logger.info(f"🎯 Processing ONLY selected set: {selected_set}")
        else:
            logger.error(
                f"❌ Selected set '{selected_set}' not found in available sets: {available_sets}"
            )
            return False  # Don't fall back to processing all sets
    
    # Determine which semesters to process
    if (
        processing_mode == "auto"
        or not selected_semesters
        or "all" in selected_semesters
    ):
        semesters_to_process = BN_SEMESTER_ORDER.copy()
        logger.info(f"🎯 Processing ALL semesters: {semesters_to_process}")
    else:
        semesters_to_process = selected_semesters
        logger.info(f"🎯 Processing selected semesters: {semesters_to_process}")
    
    # Load course data once
    try:
        (
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
        ) = load_bn_course_data()
    except Exception as e:
        logger.error(f"❌ Could not load course data: {e}")
        return False
    
    ts = datetime.now().strftime(TIMESTAMP_FMT)
    
    # Process each set and semester
    total_processed = 0
    for bn_set in sets_to_process:
        logger.info(f"\n{'='*60}")
        logger.info(f"PROCESSING SET: {bn_set}")
        logger.info(f"{'='*60}")
        
        # UPDATED: Raw and clean directories now under BN folder
        raw_dir = normalize_path(
            os.path.join(base_dir_norm, "BN", bn_set, "RAW_RESULTS")
        )
        clean_dir = normalize_path(
            os.path.join(base_dir_norm, "BN", bn_set, "CLEAN_RESULTS")
        )
        
        # Create directories if they don't exist
        os.makedirs(raw_dir, exist_ok=True)
        os.makedirs(clean_dir, exist_ok=True)
        
        if not os.path.exists(raw_dir):
            logger.warning(f"⚠️ RAW_RESULTS directory not found: {raw_dir}")
            continue
        
        raw_files = [
            f
            for f in os.listdir(raw_dir)
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
        ]
        if not raw_files:
            logger.warning(f"⚠️ No raw files in {raw_dir}; skipping {bn_set}")
            continue
        
        logger.info(f"📁 Found {len(raw_files)} raw files in {bn_set}: {raw_files}")
        
        # Create timestamped folder for this set
        set_output_dir = os.path.join(clean_dir, "{}_RESULT-{}".format(bn_set, ts))
        os.makedirs(set_output_dir, exist_ok=True)
        logger.info(f"📁 Created BN set output directory: {set_output_dir}")
        
        # Process selected semesters
        semester_processed = 0
        for semester_key in semesters_to_process:
            if semester_key not in BN_SEMESTER_ORDER:
                logger.warning(f"⚠️ Skipping unknown semester: {semester_key}")
                continue
            
            # Check if there are files for this semester - USING BM'S APPROACH
            semester_files_exist = False
            for rf in raw_files:
                try:
                    detected_sem = detect_bn_semester_from_filename(rf)
                    if detected_sem == semester_key:
                        semester_files_exist = True
                        break
                except ValueError as e:
                    logger.warning(f"⚠️ Could not detect semester for {rf}: {e}")
                    continue
            
            if semester_files_exist:
                logger.info(f"\n🎯 Processing BN {semester_key} in {bn_set}...")
                try:
                    # Add file existence check
                    if not check_bn_files_exist(raw_dir, semester_key):
                        logger.error(
                            f"❌ Skipping {semester_key} - no valid files found"
                        )
                        continue
                    
                    # Process the semester with the upgrade threshold
                    result = process_bn_semester_files(
                        semester_key,
                        raw_files,
                        raw_dir,
                        set_output_dir,
                        ts,
                        params["pass_threshold"],
                        semester_course_maps,
                        semester_credit_units,
                        semester_lookup,
                        semester_course_titles,
                        DEFAULT_LOGO_PATH,
                        bn_set,
                        previous_gpas=None,
                        upgrade_min_threshold=upgrade_min_threshold,
                    )
                    
                    if result and result.get("success", False):
                        logger.info(f"✅ Successfully processed {semester_key}")
                        total_processed += result.get("files_processed", 0)
                        semester_processed += result.get("files_processed", 0)
                    else:
                        logger.error(f"❌ Failed to process {semester_key}")
                except Exception as e:
                    logger.error(f"❌ Error processing {semester_key}: {e}")
                    traceback.print_exc()
            else:
                logger.warning(
                    f"⚠️ No files found for {semester_key} in {bn_set}, skipping..."
                )
        
        # Create ZIP of BN results ONLY if files were processed
        if semester_processed > 0:
            try:
                zip_success = create_bn_zip_for_set(
                    clean_dir, bn_set, ts, set_output_dir
                )
                if zip_success:
                    logger.info(f"✅ Successfully created ZIP for {bn_set}")
                else:
                    logger.warning(
                        f"⚠️ ZIP creation failed for {bn_set}, files remain in: {set_output_dir}"
                    )
            except Exception as e:
                logger.warning(f"⚠️ Failed to create BN ZIP for {bn_set}: {e}")
                traceback.print_exc()
        else:
            logger.warning(f"⚠️ No files processed for {bn_set}, skipping ZIP creation")
            if os.path.exists(set_output_dir):
                shutil.rmtree(set_output_dir)
    
    # Only return True if files were actually processed
    if total_processed == 0:
        logger.error("❌ No files were actually processed. Check if:")
        logger.error(" - Raw files exist in the correct directory")
        logger.error(" - File naming matches semester patterns")
        logger.error(" - Course data files are available")
        return False
    
    # Print BN-specific summaries
    logger.info("\n📊 BN STUDENT TRACKING SUMMARY:")
    logger.info(f"Total unique BN students tracked: {len(STUDENT_TRACKER)}")
    logger.info(f"Total BN withdrawn students: {len(WITHDRAWN_STUDENTS)}")
    if CARRYOVER_STUDENTS:
        logger.info("\n📋 BN CARRYOVER STUDENT SUMMARY:")
        logger.info(f"Total BN carryover students: {len(CARRYOVER_STUDENTS)}")
    
    return True

def get_form_parameters():
    """Get parameters from environment variables set by the web form."""
    logger.info("🎯 DEBUG - FORM PARAMETERS:")
    logger.info(f" SELECTED_SET: {os.getenv('SELECTED_SET')}")
    logger.info(f" PROCESSING_MODE: {os.getenv('PROCESSING_MODE')}")
    logger.info(f" SELECTED_SEMESTERS: {os.getenv('SELECTED_SEMESTERS')}")
    logger.info(f" UPGRADE_THRESHOLD: {os.getenv('UPGRADE_THRESHOLD')}")
    
    selected_set = os.getenv("SELECTED_SET", "all")
    processing_mode = os.getenv("PROCESSING_MODE", "auto")
    selected_semesters_str = os.getenv("SELECTED_SEMESTERS", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", "50.0"))
    generate_pdf = os.getenv("GENERATE_PDF", "True").lower() == "true"
    track_withdrawn = os.getenv("TRACK_WITHDRAWN", "True").lower() == "true"
    
    # Convert semester string to list
    selected_semesters = []
    if selected_semesters_str:
        selected_semesters = selected_semesters_str.split(",")
    
    logger.info("🎯 FINAL PARAMETERS:")
    logger.info(f" Selected Set: {selected_set}")
    logger.info(f" Processing Mode: {processing_mode}")
    logger.info(f" Selected Semesters: {selected_semesters}")
    logger.info(f" Pass Threshold: {pass_threshold}")
    
    return {
        "selected_set": selected_set,
        "processing_mode": processing_mode,
        "selected_semesters": selected_semesters,
        "pass_threshold": pass_threshold,
        "generate_pdf": generate_pdf,
        "track_withdrawn": track_withdrawn,
    }

def get_pass_threshold():
    """Get pass threshold - now handles upgrade logic interactively."""
    threshold_str = os.getenv("PASS_THRESHOLD", "50.0")
    try:
        threshold = float(threshold_str)
    except ValueError:
        threshold = 50.0
    return threshold

DEFAULT_PASS_THRESHOLD = get_pass_threshold()
TIMESTAMP_FMT = "%Y-%m-%d_%H%M%S"
DEFAULT_LOGO_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "launcher", "static", "logo.png")
)
NAME_WIDTH_CAP = 40

# ============================================================================
# GPA & CGPA CALCULATION FUNCTIONS
# ============================================================================

def calculate_cumulative_cgpa(student_data, current_gpa, current_credits):
    """
    Calculate Cumulative CGPA = Weighted average of all semester GPAs.
    Formula: (Σ(GPA × Credits)) / (Σ Credits)
    """
    if not student_data:
        return current_gpa
    total_grade_points = 0.0
    total_credits = 0
    # Add all previous semesters
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
    # Add current semester
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits
    return (
        round(total_grade_points / total_credits, 2)
        if total_credits > 0
        else current_gpa
    )

def calculate_cgpa(student_data, current_gpa, current_credits):
    """Calculate Cumulative GPA (CGPA) based on all previous semesters and current semester."""
    if not student_data:
        return current_gpa
    total_grade_points = 0.0
    total_credits = 0
    # Add previous semesters
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
    # Add current semester
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits
    if total_credits > 0:
        return round(total_grade_points / total_credits, 2)
    else:
        return current_gpa

# ============================================================================
# CRITICAL FUNCTION FIXES - ISSUE 1, 2, 3 RESOLUTIONS
# ============================================================================

def create_bn_cgpa_summary_sheet(mastersheet_path, ts, semester_credit_units, set_name="", logo_path=None):
    """Create a CGPA summary sheet that aggregates GPA across all BN semesters with proper weighted CGPA and correct status display.
    UPDATED: Sheet position now moves towards end after the last semester."""
    try:
        logger.info("📊 Creating BN CGPA Summary Sheet...")
        # Load the mastersheet workbook
        wb = load_workbook(mastersheet_path)
        
        # FIXED: Calculate the correct position for CGPA_SUMMARY sheet
        # We want it after the last semester sheet but before ANALYSIS
        sheet_names = wb.sheetnames
        max_semester_index = -1
        
        # Find the highest index of semester sheets
        for i, sheet_name in enumerate(sheet_names):
            if sheet_name in BN_SEMESTER_ORDER:
                max_semester_index = i
        
        # Calculate position: after last semester sheet
        if max_semester_index >= 0:
            cgpa_sheet_position = max_semester_index + 1
            logger.info(f"📋 CGPA_SUMMARY sheet will be inserted at position {cgpa_sheet_position} (after last semester)")
        else:
            # If no semester sheets found, insert before ANALYSIS or at the end
            cgpa_sheet_position = len(sheet_names)
            if "ANALYSIS" in sheet_names:
                analysis_index = sheet_names.index("ANALYSIS")
                cgpa_sheet_position = analysis_index  # Insert before ANALYSIS
            logger.info(f"📋 CGPA_SUMMARY sheet will be inserted at position {cgpa_sheet_position}")
        
        # Collect GPA data from all BN semesters
        cgpa_data = {}
        for sheet_name in wb.sheetnames:
            if sheet_name in BN_SEMESTER_ORDER:
                logger.info(f"📖 Processing sheet: {sheet_name}")
               
                try:
                    # FIXED: Use fixed row 6 as header and dynamic column lookup for remarks
                    df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5) # Row 6 is header (0-indexed 5)
                   
                    # Clean column names
                    df.columns = [str(col).strip() if col is not None else f"Unnamed_{i}" for i, col in enumerate(df.columns)]
                   
                    # Log available columns for debugging
                    logger.info(f"📋 Columns in {sheet_name}: {df.columns.tolist()}")
                   
                    # Find exam number column - FIXED: Use specific column names
                    exam_col = None
                    for col in df.columns:
                        col_str = str(col).upper().strip()
                        if "EXAM" in col_str and "NUMBER" in col_str:
                            exam_col = col
                            break
                   
                    if not exam_col:
                        # Try alternative column names
                        for col in df.columns:
                            col_str = str(col).upper().strip()
                            if "REG" in col_str or "NO" in col_str:
                                exam_col = col
                                break
                   
                    if not exam_col:
                        logger.warning(f"⚠️ No exam number column found in {sheet_name}")
                        continue
                   
                    # Filter valid rows - CRITICAL: Remove summary rows
                    df = df[df[exam_col].notna() & (df[exam_col] != "")]
                    df = df[~df[exam_col].astype(str).str.upper().str.contains("SUMMARY|TOTAL|FAIL", na=False)]
                   
                    # Find required columns - FIXED: Use dynamic lookup
                    gpa_col = None
                    name_col = None
                   
                    for col in df.columns:
                        col_str = str(col).upper().strip()
                        if "GPA" in col_str and not any(x in col_str for x in ["CGPA", "PREVIOUS", "OVERALL"]):
                            gpa_col = col
                            logger.info(f"✅ Found GPA column: '{col}'")
                        elif "NAME" in col_str:
                            name_col = col
                            logger.info(f"✅ Found NAME column: '{col}'")
                    if not exam_col or not gpa_col:
                        logger.warning(f"⚠️ Missing required columns in {sheet_name}")
                        logger.warning(f" exam_col={exam_col}, gpa_col={gpa_col}")
                        continue
                    # FIXED: Use dynamic lookup for REMARKS column instead of hardcoded index
                    remarks_cols = [c for c in df.columns if str(c).strip().upper() == "REMARKS"]
                    if remarks_cols:
                        remarks_col_name = remarks_cols[0]
                        logger.info(f"✅ Found REMARKS column: '{remarks_col_name}'")
                    else:
                        remarks_col_name = None
                        logger.warning(f"⚠️ No REMARKS column found in {sheet_name}")
                    # Process each student
                    students_processed = 0
                    for idx, row in df.iterrows():
                        exam_no = str(row[exam_col]).strip()
                       
                        # Skip invalid exam numbers
                        if not exam_no or exam_no == "nan" or exam_no == "" or len(exam_no) < 3:
                            continue
                       
                        # Clean exam number (remove .0 if present)
                        if "." in exam_no and exam_no.endswith(".0"):
                            exam_no = exam_no[:-2]
                       
                        # Initialize student data
                        if exam_no not in cgpa_data:
                            cgpa_data[exam_no] = {
                                "name": (
                                    str(row[name_col]).strip()
                                    if name_col and pd.notna(row.get(name_col))
                                    else ""
                                ),
                                "gpas": {},
                                "credits": {},
                                "status_history": {},
                                "last_semester": sheet_name,
                                "last_remarks": "Active",
                            }
                       
                        # Update GPA
                        gpa_value = row[gpa_col]
                        if pd.notna(gpa_value):
                            try:
                                cgpa_data[exam_no]["gpas"][sheet_name] = float(gpa_value)
                            except (ValueError, TypeError):
                                cgpa_data[exam_no]["gpas"][sheet_name] = 0.0
                        else:
                            cgpa_data[exam_no]["gpas"][sheet_name] = 0.0
                       
                        # Update credits
                        if sheet_name in semester_credit_units:
                            total_credits = sum(semester_credit_units[sheet_name].values())
                            cgpa_data[exam_no]["credits"][sheet_name] = total_credits
                       
                        # CRITICAL FIX: Capture status from REMARKS column using dynamic lookup
                        if remarks_col_name and remarks_col_name in row:
                            remarks_value = row[remarks_col_name]
                           
                            if pd.notna(remarks_value):
                                remarks = str(remarks_value).strip()
                               
                                # Clean up remarks (remove extra whitespace, normalize)
                                remarks = " ".join(remarks.split())
                               
                                # Store in status history
                                cgpa_data[exam_no]["status_history"][sheet_name] = remarks
                                cgpa_data[exam_no]["last_remarks"] = remarks
                                cgpa_data[exam_no]["last_semester"] = sheet_name
                               
                                students_processed += 1
                               
                                # Debug logging for first few students
                                if students_processed <= 3:
                                    logger.info(f"📝 Status captured from column '{remarks_col_name}': {exam_no} in {sheet_name} -> '{remarks}'")
                            else:
                                # If no remarks in REMARKS column, try to determine from other data
                                cu_passed = row.get("CU Passed", 0)
                                cu_failed = row.get("CU Failed", 0)
                                gpa_val = cgpa_data[exam_no]["gpas"][sheet_name]
                               
                                # Simple status determination if remarks not available
                                if cu_failed == 0:
                                    remarks = "Passed"
                                elif gpa_val >= 2.0:
                                    remarks = "Resit"
                                elif gpa_val < 2.0:
                                    remarks = "Probation"
                                else:
                                    remarks = "Active"
                                   
                                cgpa_data[exam_no]["status_history"][sheet_name] = remarks
                                cgpa_data[exam_no]["last_remarks"] = remarks
                                students_processed += 1
                        else:
                            logger.warning(f"⚠️ REMARKS column '{remarks_col_name}' not found or not accessible in {sheet_name}")
               
                    logger.info(f"✅ Processed {students_processed} students from {sheet_name}")
               
                except Exception as e:
                    logger.warning(f"⚠️ Error processing sheet {sheet_name}: {e}")
                    traceback.print_exc()
                    continue
        logger.info(f"📊 Collected data for {len(cgpa_data)} students")
        # Create CGPA summary dataframe with weighted average
        summary_data = []
        for exam_no, data in cgpa_data.items():
            # CRITICAL FIX: Improved status determination logic
            current_status = "Active"
            status_history = data.get("status_history", {})
           
            # Priority order for status selection (most recent semester first)
            for semester in reversed(BN_SEMESTER_ORDER):
                if semester in status_history:
                    status = status_history[semester]
                    # Normalize status values
                    if status in ["Passed", "Pass"]:
                        current_status = "Passed"
                    elif status in ["Resit", "Carry Over"]:
                        current_status = "Resit"
                    elif status in ["Probation"]:
                        current_status = "Probation"
                    elif status in ["Withdrawn", "Withdraw"]:
                        current_status = "Withdrawn"
                    else:
                        current_status = status
                    break
           
            # If no status in history but last_remarks exists
            if current_status == "Active" and data.get("last_remarks") != "Active":
                current_status = data["last_remarks"]
            row = {
                "EXAM NO": exam_no,
                "NAME": data["name"],
                "STATUS": current_status,
                "PROB HIST": (
                    ", ".join([sem for sem, status in status_history.items() if "Probation" in str(status)])
                    if status_history
                    else "None"
                ),
            }
            # Add GPA for each semester and calculate weighted CGPA
            total_grade_points = 0.0
            total_credits = 0
            for semester in BN_SEMESTER_ORDER:
                if semester in data["gpas"]:
                    row[SHORT_SEMESTER_MAP[semester]] = data["gpas"][semester]
                    # Calculate weighted contribution
                    semester_credits = data["credits"].get(semester, 0)
                    if semester_credits > 0:
                        total_grade_points += data["gpas"][semester] * semester_credits
                        total_credits += semester_credits
                else:
                    row[SHORT_SEMESTER_MAP[semester]] = None
            # Calculate Overall CGPA (weighted average)
            if total_credits > 0:
                row["CGPA"] = round(total_grade_points / total_credits, 2)
            else:
                row["CGPA"] = 0.0
               
            summary_data.append(row)
        # Create summary dataframe
        if not summary_data:
            logger.warning("⚠️ No summary data collected")
            return None
           
        summary_df = pd.DataFrame(summary_data)
        # Sort by STATUS (Active/Passed first) then Overall CGPA descending
        def status_sort_key(status):
            status_order = {"Passed": 0, "Resit": 1, "Probation": 2, "Withdrawn": 3, "Active": 4}
            return status_order.get(status, 5)
       
        summary_df["status_sort"] = summary_df["STATUS"].apply(status_sort_key)
        summary_df = summary_df.sort_values(
            by=["status_sort", "CGPA"], ascending=[True, False]
        ).drop(columns=["status_sort"]).reset_index(drop=True)
        # Add the summary sheet to the workbook
        if "CGPA_SUMMARY" in wb.sheetnames:
            del wb["CGPA_SUMMARY"]
        
        # FIXED: Insert CGPA_SUMMARY at the calculated position (after last semester)
        ws = wb.create_sheet("CGPA_SUMMARY", cgpa_sheet_position)
        logger.info(f"✅ CGPA_SUMMARY sheet created at position {cgpa_sheet_position}")
        
        # Add document headings
        headers = (
            ["EXAM NO", "NAME", "STATUS", "PROB HIST"]
            + [SHORT_SEMESTER_MAP[sem] for sem in BN_SEMESTER_ORDER]
            + ["CGPA"]
        )
        last_letter = get_column_letter(len(headers))
        
        # FIXED HEADER SECTION: Adjust starting row to avoid logo coverage
        start_row = 6  # Start from row 6 to leave space for logo
        
        # Add logo if provided
        if logo_path and os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width, img.height = 110, 80
                img.anchor = "A1"
                ws.add_image(img, "A1")
            except Exception as e:
                logger.warning(f"⚠ Could not place logo in CGPA_SUMMARY: {e}")
        # College name - FIXED: Start from row 1 but merge across columns
        ws.merge_cells(f"C1:{last_letter}1")
        title_cell = ws["C1"]
        title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )
        # Department
        ws.merge_cells(f"C2:{last_letter}2")
        dept_cell = ws["C2"]
        dept_cell.value = "Department of Nursing"
        dept_cell.font = Font(bold=True, size=14, color="000000")
        dept_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Set name and title
        ws.merge_cells(f"C3:{last_letter}3")
        set_title_cell = ws["C3"]
        set_title_cell.value = f"Set: {set_name} - CGPA Summary Sheet"
        set_title_cell.font = Font(bold=True, size=12, color="000000")
        set_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add spacing rows
        ws.row_dimensions[4].height = 5
        ws.row_dimensions[5].height = 5
        
        # Write header starting from row 6 (adjusted from row 4)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4A90E2", end_color="4A90E2", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        # Write data starting from row 7 (adjusted from row 5)
        for row_idx, row_data in enumerate(summary_data, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
               
                # Color code status cells
                if header == "STATUS":
                    if value == "Withdrawn":
                        cell.fill = PatternFill(
                            start_color="FFCCCB", end_color="FFCCCB", fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="CC0000")
                    elif value == "Probation":
                        cell.fill = PatternFill(
                            start_color="FFA500", end_color="FFA500", fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif value == "Resit":
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="9C6500")
                    elif value == "Passed":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="006100")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(header))
           
            for row_idx in range(start_row + 1, len(summary_data) + start_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))
           
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        # Freeze headings
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        wb.save(mastersheet_path)
        logger.info("✅ BN CGPA Summary sheet created successfully with CORRECT status display and POSITION")
       
        # Print detailed summary statistics
        status_counts = summary_df["STATUS"].value_counts()
        logger.info("\n📊 CGPA Summary Status Distribution:")
        for status, count in status_counts.items():
            logger.info(f" {status}: {count} students")
           
        logger.info(f"📊 Overall CGPA Range: {summary_df['CGPA'].min():.2f} - {summary_df['CGPA'].max():.2f}")
        return summary_df
    except Exception as e:
        logger.error(f"❌ Error creating BN CGPA summary sheet: {e}")
        traceback.print_exc()
        return None

def create_bn_analysis_sheet(mastersheet_path, ts, semester_credit_units, set_name="", logo_path=None):
    """Create an analysis sheet with comprehensive statistics for BN - UPDATED WITH PROBATION SEPARATION AND FIXED HEADER."""
    try:
        logger.info("📈 Creating BN Analysis Sheet...")
        wb = load_workbook(mastersheet_path)
        
        # Collect data from all semesters - UPDATED STRUCTURE
        analysis_data = {
            "SEMESTER": [],
            "TOTAL STUDENTS": [],
            "PASSED ALL": [],
            "RESIT STUDENTS": [], # NEW: Separate Resit
            "PROBATION STUDENTS": [], # NEW: Separate Probation
            "WITHDRAWN STUDENTS": [],
            "AVERAGE GPA": [],
            "PASS RATE": [],
            "TOTAL COURSES": [],
            "TOTAL CREDIT UNITS": [],
        }
        
        for sheet_name in wb.sheetnames:
            if sheet_name in BN_SEMESTER_ORDER:
                df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)
                
                # Filter valid rows: drop rows where 'EXAMS NUMBER' is NaN or empty
                exam_col = find_exam_number_column(df)
                if exam_col:
                    df = df[df[exam_col].notna() & (df[exam_col] != "")]
                
                # Basic statistics
                total_students = len(df)
                
                # Passed all: REMARKS == 'Passed' (strip whitespace) - UPDATED
                passed_all = 0
                if "REMARKS" in df.columns:
                    df["REMARKS"] = df["REMARKS"].astype(str).str.strip()
                    passed_all = len(df[df["REMARKS"] == "Passed"])
                
                # UPDATED: Separate Resit and Probation calculation
                resit_count = 0
                probation_count = 0
                if "REMARKS" in df.columns:
                    resit_count = len(df[df["REMARKS"] == "Resit"])
                    probation_count = len(df[df["REMARKS"] == "Probation"])
                
                # FIXED: Withdrawn calculation - look for "Withdrawn" in REMARKS column
                withdrawn_count = 0
                if "REMARKS" in df.columns:
                    withdrawn_count = len(df[df["REMARKS"] == "Withdrawn"])
                
                # Average GPA: mean of non-NaN GPA values
                avg_gpa = 0
                if "GPA" in df.columns:
                    gpa_series = pd.to_numeric(df["GPA"], errors="coerce")
                    avg_gpa = (
                        gpa_series.dropna().mean()
                        if not gpa_series.dropna().empty
                        else 0
                    )
                
                # Pass rate calculation
                pass_rate = (
                    (passed_all / total_students * 100) if total_students > 0 else 0
                )
                
                # Course and credit unit information
                total_courses = 0
                total_credit_units = 0
                if sheet_name in semester_credit_units:
                    total_courses = len(semester_credit_units[sheet_name])
                    total_credit_units = sum(semester_credit_units[sheet_name].values())
                
                analysis_data["SEMESTER"].append(SHORT_SEMESTER_MAP.get(sheet_name, sheet_name))
                analysis_data["TOTAL STUDENTS"].append(total_students)
                analysis_data["PASSED ALL"].append(passed_all)
                analysis_data["RESIT STUDENTS"].append(resit_count) # NEW
                analysis_data["PROBATION STUDENTS"].append(probation_count) # NEW
                analysis_data["WITHDRAWN STUDENTS"].append(withdrawn_count)
                analysis_data["AVERAGE GPA"].append(round(avg_gpa, 2))
                analysis_data["PASS RATE"].append(round(pass_rate, 2))
                analysis_data["TOTAL COURSES"].append(total_courses)
                analysis_data["TOTAL CREDIT UNITS"].append(total_credit_units)
        
        # Create analysis dataframe
        analysis_df = pd.DataFrame(analysis_data)
        
        # Add overall statistics
        if not analysis_df.empty:
            overall_stats = {
                "SEMESTER": "OVERALL",
                "TOTAL STUDENTS": analysis_df["TOTAL STUDENTS"].sum(),
                "PASSED ALL": analysis_df["PASSED ALL"].sum(),
                "RESIT STUDENTS": analysis_df["RESIT STUDENTS"].sum(), # NEW
                "PROBATION STUDENTS": analysis_df["PROBATION STUDENTS"].sum(), # NEW
                "WITHDRAWN STUDENTS": analysis_df["WITHDRAWN STUDENTS"].sum(),
                "AVERAGE GPA": round(analysis_df["AVERAGE GPA"].mean(), 2),
                "PASS RATE": round(analysis_df["PASS RATE"].mean(), 2),
                "TOTAL COURSES": analysis_df["TOTAL COURSES"].sum(),
                "TOTAL CREDIT UNITS": analysis_df["TOTAL CREDIT UNITS"].sum(),
            }
            analysis_df = pd.concat(
                [analysis_df, pd.DataFrame([overall_stats])], ignore_index=True
            )
        
        # Add the analysis sheet to the workbook - INSERT AT THE END
        if "ANALYSIS" in wb.sheetnames:
            del wb["ANALYSIS"]
        ws = wb.create_sheet("ANALYSIS")
        logger.info("✅ ANALYSIS sheet created at the end")
        
        # Add document headings
        headers = [
            "SEMESTER",
            "TOTAL STUDENTS",
            "PASSED ALL",
            "RESIT STUDENTS",
            "PROBATION STUDENTS",
            "WITHDRAWN STUDENTS",
            "AVERAGE GPA",
            "PASS RATE (%)",
            "TOTAL COURSES",
            "TOTAL CREDIT UNITS",
        ]
        last_letter = get_column_letter(len(headers))
        
        # FIXED HEADER SECTION: Adjust starting row to avoid logo coverage
        start_row = 6  # Start from row 6 to leave space for logo
        
        # Add logo if provided
        if logo_path and os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width, img.height = 110, 80
                img.anchor = "A1"
                ws.add_image(img, "A1")
            except Exception as e:
                logger.warning(f"⚠ Could not place logo in ANALYSIS: {e}")
        
        # College name - FIXED: Start from row 1 but merge across columns
        ws.merge_cells(f"C1:{last_letter}1")
        title_cell = ws["C1"]
        title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )
        
        # Department
        ws.merge_cells(f"C2:{last_letter}2")
        dept_cell = ws["C2"]
        dept_cell.value = "Department of Nursing"
        dept_cell.font = Font(bold=True, size=14, color="000000")
        dept_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set name and title
        ws.merge_cells(f"C3:{last_letter}3")
        set_title_cell = ws["C3"]
        set_title_cell.value = f"Set: {set_name} - Analysis Sheet"
        set_title_cell.font = Font(bold=True, size=12, color="000000")
        set_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add spacing rows
        ws.row_dimensions[4].height = 5
        ws.row_dimensions[5].height = 5
        
        # Write header starting from row 6 (adjusted from row 4)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col_idx, value=header)
        
        # Write data starting from row 7 (adjusted from row 5)
        for row_idx, row_data in analysis_df.iterrows():
            ws.cell(row=row_idx + start_row + 1, column=1, value=row_data["SEMESTER"])
            ws.cell(row=row_idx + start_row + 1, column=2, value=row_data["TOTAL STUDENTS"])
            ws.cell(row=row_idx + start_row + 1, column=3, value=row_data["PASSED ALL"])
            ws.cell(row=row_idx + start_row + 1, column=4, value=row_data["RESIT STUDENTS"])
            ws.cell(row=row_idx + start_row + 1, column=5, value=row_data["PROBATION STUDENTS"])
            ws.cell(row=row_idx + start_row + 1, column=6, value=row_data["WITHDRAWN STUDENTS"])
            ws.cell(row=row_idx + start_row + 1, column=7, value=row_data["AVERAGE GPA"])
            ws.cell(row=row_idx + start_row + 1, column=8, value=row_data["PASS RATE"])
            ws.cell(row=row_idx + start_row + 1, column=9, value=row_data["TOTAL COURSES"])
            ws.cell(row=row_idx + start_row + 1, column=10, value=row_data["TOTAL CREDIT UNITS"])
        
        # Style the header
        for cell in ws[start_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="27ae60", end_color="27ae60", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        
        # Style data rows and auto-adjust column widths
        max_lengths = [len(str(h)) for h in headers] # Initialize with header lengths
        for row in range(start_row + 1, len(analysis_df) + start_row + 1):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                if row == len(analysis_df) + start_row: # Overall row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                    )
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                # Update max length for auto-adjust
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_lengths[col - 1]:
                    max_lengths[col - 1] = len(cell_value)
        
        # Auto-adjust column widths
        for col_idx, max_len in enumerate(max_lengths, 1):
            col_letter = get_column_letter(col_idx)
            adjusted_width = min(max_len + 2, 50) # Cap at 50
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Freeze headings
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        wb.save(mastersheet_path)
        logger.info(
            "✅ BN Analysis sheet created successfully with auto-adjusted columns and FIXED withdrawn counting"
        )
        return analysis_df
        
    except Exception as e:
        logger.error(f"❌ Error creating BN analysis sheet: {e}")
        traceback.print_exc()
        return None

def load_previous_cgpas_from_processed_files(
    output_dir, current_semester_key, timestamp
):
    """
    Load CGPA from the immediately previous semester.
    FIXED: Properly reads Excel files with correct header detection.
    """
    previous_cgpas = {}
    
    # Determine which semester's data to load
    current_year, current_semester_num, _, _, _ = get_semester_display_info(
        current_semester_key
    )
    
    if current_semester_num == 1 and current_year == 1:
        logger.info("📊 First semester of first year - no previous CGPA available")
        return previous_cgpas
    elif current_semester_num == 2 and current_year == 1:
        prev_semester = "N-FIRST-YEAR-FIRST-SEMESTER"
    elif current_semester_num == 1 and current_year == 2:
        prev_semester = "N-FIRST-YEAR-SECOND-SEMESTER"
    elif current_semester_num == 2 and current_year == 2:
        prev_semester = "N-SECOND-YEAR-FIRST-SEMESTER"
    elif current_semester_num == 1 and current_year == 3:
        prev_semester = "N-SECOND-YEAR-SECOND-SEMESTER"
    elif current_semester_num == 2 and current_year == 3:
        prev_semester = "N-THIRD-YEAR-FIRST-SEMESTER"
    else:
        logger.warning(f"⚠️ Unknown semester progression for {current_semester_key}")
        return previous_cgpas
    
    # Load from mastersheet
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")
    if os.path.exists(mastersheet_path):
        try:
            logger.info(f"🔍 Loading previous CGPA data from: {prev_semester}")
            # CRITICAL FIX: Read Excel properly and find header row
            df_raw = pd.read_excel(
                mastersheet_path, sheet_name=prev_semester, header=None
            )
            
            # Find the actual header row
            header_row_idx = None
            for idx, row in df_raw.iterrows():
                row_values = [str(val).strip().upper() for val in row if pd.notna(val)]
                if any(
                    "EXAMS NUMBER" in val or ("EXAM" in val and "NUMBER" in val)
                    for val in row_values
                ):
                    header_row_idx = idx
                    break
            
            if header_row_idx is None:
                logger.warning(f"⚠️ Could not find header row in {prev_semester}")
                return previous_cgpas
            
            # Read again with correct header
            df = pd.read_excel(
                mastersheet_path, sheet_name=prev_semester, header=header_row_idx
            )
            
            # Clean column names
            df.columns = [str(col).strip() for col in df.columns]
            
            # Find exam number and GPA columns
            exam_col = find_exam_number_column(df)
            gpa_col = None
            for col in df.columns:
                col_upper = str(col).upper().strip()
                if "GPA" in col_upper and not any(
                    x in col_upper for x in ["CGPA", "PREVIOUS", "OVERALL"]
                ):
                    gpa_col = col
                    break
            
            if exam_col and gpa_col:
                # Filter out summary rows
                df = df[df[exam_col].notna() & (df[exam_col] != "")]
                df = df[
                    ~df[exam_col]
                    .astype(str)
                    .str.upper()
                    .str.contains("SUMMARY|TOTAL|FAIL", na=False)
                ]
                
                cgpas_loaded = 0
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    # Clean exam number
                    if "." in exam_no and exam_no.endswith(".0"):
                        exam_no = exam_no[:-2]
                    
                    # Skip invalid exam numbers
                    if (
                        not exam_no
                        or exam_no == "nan"
                        or exam_no == ""
                        or len(exam_no) < 3
                    ):
                        continue
                    
                    cgpa = row[gpa_col]
                    if pd.notna(cgpa):
                        try:
                            previous_cgpas[exam_no] = float(cgpa)
                            cgpas_loaded += 1
                            if cgpas_loaded <= 5:
                                logger.info(f"📝 Loaded CGPA: {exam_no} → {cgpa}")
                        except (ValueError, TypeError):
                            continue
                
                logger.info(
                    f"✅ Loaded {cgpas_loaded} previous CGPAs from {prev_semester}"
                )
            else:
                logger.warning(f"⚠️ Could not find required columns in {prev_semester}")
                logger.warning(f" exam_col={exam_col}, gpa_col={gpa_col}")
                
        except Exception as e:
            logger.warning(f"⚠️ Could not load previous CGPAs from {prev_semester}: {e}")
            traceback.print_exc()
    else:
        logger.warning(f"⚠️ Mastersheet not found: {mastersheet_path}")
    
    return previous_cgpas

def load_all_previous_cgpas_for_cumulative(output_dir, current_semester_key, timestamp):
    """
    Load GPA and credit data from ALL completed semesters.
    FIXED: Properly handles Excel file reading with header detection.
    """
    current_year, current_semester_num, _, _, _ = get_semester_display_info(
        current_semester_key
    )
    
    # Determine which semesters to include
    semesters_to_load = []
    if current_semester_num == 1 and current_year == 1:
        return {}
    elif current_semester_num == 2 and current_year == 1:
        semesters_to_load = ["N-FIRST-YEAR-FIRST-SEMESTER"]
    elif current_semester_num == 1 and current_year == 2:
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
        ]
    elif current_semester_num == 2 and current_year == 2:
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
        ]
    elif current_semester_num == 1 and current_year == 3:
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
            "N-SECOND-YEAR-SECOND-SEMESTER",
        ]
    elif current_semester_num == 2 and current_year == 3:
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
            "N-SECOND-YEAR-SECOND-SEMESTER",
            "N-THIRD-YEAR-FIRST-SEMESTER",
        ]
    
    all_student_data = {}
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")
    
    for semester in semesters_to_load:
        try:
            logger.info(f"📖 Loading cumulative data from: {semester}")
            # CRITICAL FIX: Read Excel properly
            df_raw = pd.read_excel(mastersheet_path, sheet_name=semester, header=None)
            
            # Find header row
            header_row_idx = None
            for idx, row in df_raw.iterrows():
                row_values = [str(val).strip().upper() for val in row if pd.notna(val)]
                if any(
                    "EXAMS NUMBER" in val or ("EXAM" in val and "NUMBER" in val)
                    for val in row_values
                ):
                    header_row_idx = idx
                    break
            
            if header_row_idx is None:
                logger.warning(f"⚠️ Could not find header in {semester}")
                continue
            
            df = pd.read_excel(
                mastersheet_path, sheet_name=semester, header=header_row_idx
            )
            df.columns = [str(col).strip() for col in df.columns]
            
            exam_col = find_exam_number_column(df)
            gpa_col = None
            credit_passed_col = None
            credit_failed_col = None
            
            for col in df.columns:
                col_str = str(col).upper()
                if "GPA" in col_str and not any(
                    x in col_str for x in ["CGPA", "PREVIOUS", "OVERALL"]
                ):
                    gpa_col = col
                elif "CU PASSED" in col_str:
                    credit_passed_col = col
                elif "CU FAILED" in col_str:
                    credit_failed_col = col
            
            if exam_col and gpa_col:
                # Filter out summary rows
                df = df[df[exam_col].notna() & (df[exam_col] != "")]
                df = df[
                    ~df[exam_col]
                    .astype(str)
                    .str.upper()
                    .str.contains("SUMMARY|TOTAL|FAIL", na=False)
                ]
                
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    # Clean exam number
                    if "." in exam_no and exam_no.endswith(".0"):
                        exam_no = exam_no[:-2]
                    
                    if (
                        not exam_no
                        or exam_no == "nan"
                        or exam_no == ""
                        or len(exam_no) < 3
                    ):
                        continue
                    
                    gpa = row[gpa_col]
                    if pd.notna(gpa):
                        if exam_no not in all_student_data:
                            all_student_data[exam_no] = {"gpas": [], "credits": []}
                        
                        credits_passed = (
                            int(row[credit_passed_col])
                            if credit_passed_col and pd.notna(row[credit_passed_col])
                            else 0
                        )
                        credits_failed = (
                            int(row[credit_failed_col])
                            if credit_failed_col and pd.notna(row[credit_failed_col])
                            else 0
                        )
                        credits = credits_passed + credits_failed
                        if credits == 0:
                            credits = 30
                        
                        all_student_data[exam_no]["gpas"].append(float(gpa))
                        all_student_data[exam_no]["credits"].append(credits)
                        
        except Exception as e:
            logger.warning(f"⚠️ Could not load data from {semester}: {e}")
            traceback.print_exc()
    
    logger.info(f"📊 Loaded cumulative data for {len(all_student_data)} students")
    return all_student_data

def determine_student_status(row, total_cu, pass_threshold):
    """
    ENFORCE the rule based on:
    | Category | GPA Condition | Credit Units Passed | Status |
    | -------- | ------------- | ------------------- | ------------------------------------------------------ |
    | 1 | GPA ≥ 2.00 | ≥ 45% | To **resit** failed courses next session |
    | 2 | GPA < 2.00 | ≥ 45% | **Placed on Probation**, to resit courses next session |
    | 3 | Any GPA | < 45% | **Advised to withdraw** |
    IMPROVED: Better logging and validation
    """
    exam_no = row.get("EXAMS NUMBER", "Unknown")
    
    # Safely get GPA
    try:
        gpa = float(row.get("GPA", 0))
    except (ValueError, TypeError):
        gpa = 0.0
        logger.warning(f"⚠️ Invalid GPA for {exam_no}, defaulting to 0.0")
    
    # Safely get credit units
    try:
        cu_passed = int(row.get("CU Passed", 0))
    except (ValueError, TypeError):
        cu_passed = 0
        logger.warning(f"⚠️ Invalid CU Passed for {exam_no}, defaulting to 0")
    
    try:
        cu_failed = int(row.get("CU Failed", 0))
    except (ValueError, TypeError):
        cu_failed = 0
        logger.warning(f"⚠️ Invalid CU Failed for {exam_no}, defaulting to 0")
    
    # Validate total_cu
    if total_cu == 0:
        logger.error(
            f"❌ CRITICAL: total_cu is 0 for {exam_no}! Cannot calculate percentage."
        )
        return "Error"
    
    # Calculate percentage of credit units passed
    passed_percentage = cu_passed / total_cu * 100
    
    # ENFORCED DECISION LOGIC
    # Rule 1: No failures → PASSED
    if cu_failed == 0:
        status = "Passed"
        reason = "No failed courses"
    # Rule 3: Passed < 45% of credits → WITHDRAWN (regardless of GPA)
    elif passed_percentage < 45:
        status = "Withdrawn"
        reason = f"Passed only {passed_percentage:.1f}% of credits (<45%) - Advised to withdraw"
    # Rule 1 & 2: Passed ≥ 45% of credits
    else:
        if gpa >= 2.00:
            status = "Resit"
            reason = f"GPA {gpa:.2f} ≥ 2.00, passed {passed_percentage:.1f}% of credits (≥45%) - To resit failed courses"
        else:
            status = "Probation"
            reason = f"GPA {gpa:.2f} < 2.00, passed {passed_percentage:.1f}% of credits (≥45%) - Placed on probation, to resit courses"
    
    # Enhanced debug logging
    if not hasattr(determine_student_status, "logged_count"):
        determine_student_status.logged_count = 0
    
    if determine_student_status.logged_count < 10:
        logger.info(f"\n 🔍 Status Debug for {exam_no}:")
        logger.info(f" Total CU: {total_cu}")
        logger.info(f" CU Passed: {cu_passed} ({passed_percentage:.1f}%)")
        logger.info(f" CU Failed: {cu_failed}")
        logger.info(f" GPA: {gpa:.2f}")
        logger.info(f" → Status: {status}")
        logger.info(f" → Reason: {reason}")
        determine_student_status.logged_count += 1
    
    return status

# Reset counter on module load
determine_student_status.logged_count = 0

def validate_probation_withdrawal_logic(mastersheet, total_cu):
    """
    Validate that probation and withdrawal statuses are correctly assigned.
    IMPROVED: Better validation and comprehensive reporting.
    """
    logger.info("\n" + "=" * 70)
    logger.info("🔍 VALIDATING PROBATION/WITHDRAWAL LOGIC - ENFORCED RULE")
    logger.info("=" * 70)
    
    # Validation check: Ensure total_cu is valid
    if total_cu == 0:
        logger.error("❌ CRITICAL ERROR: total_cu is 0! Cannot validate logic.")
        return
    
    # Check students who passed < 45% (should be Withdrawn regardless of GPA)
    low_pass_students = mastersheet[
        (mastersheet["CU Passed"] / total_cu < 0.45) & (mastersheet["CU Failed"] > 0)
    ]
    logger.info(f"\n📊 Students with <45% credits passed:")
    logger.info(f" Total: {len(low_pass_students)}")
    
    if len(low_pass_students) > 0:
        logger.info(f"\n Should ALL be 'Withdrawn' (regardless of GPA):")
        errors = 0
        for idx, row in low_pass_students.head(15).iterrows():
            exam_no = row["EXAMS NUMBER"]
            gpa = row["GPA"]
            cu_passed = row["CU Passed"]
            cu_failed = row["CU Failed"]
            passed_pct = cu_passed / total_cu * 100
            status = row["REMARKS"]
            
            if status == "Withdrawn":
                correct = "✅"
            else:
                correct = f"❌ WRONG! (got '{status}')"
                errors += 1
            
            logger.info(
                f" {exam_no}: GPA={gpa:.2f}, Passed={passed_pct:.1f}%, Failed={cu_failed}, Status={status} {correct}"
            )
        
        if errors > 0:
            logger.error(f"\n ❌ FOUND {errors} INCORRECT WITHDRAWAL ASSIGNMENTS!")
    
    # Check students who passed ≥ 45% with GPA >= 2.00 (should be Resit)
    high_gpa_adequate_pass = mastersheet[
        (mastersheet["CU Passed"] / total_cu >= 0.45)
        & (mastersheet["GPA"] >= 2.00)
        & (mastersheet["CU Failed"] > 0)
    ]
    logger.info(f"\n📊 Students with ≥45% credits passed AND GPA ≥ 2.00:")
    logger.info(f" Total: {len(high_gpa_adequate_pass)}")
    
    if len(high_gpa_adequate_pass) > 0:
        logger.info(f"\n Should ALL be 'Resit':")
        errors = 0
        for idx, row in high_gpa_adequate_pass.head(15).iterrows():
            exam_no = row["EXAMS NUMBER"]
            gpa = row["GPA"]
            cu_passed = row["CU Passed"]
            passed_pct = cu_passed / total_cu * 100
            status = row["REMARKS"]
            
            if status == "Resit":
                correct = "✅"
            else:
                correct = f"❌ WRONG! (got '{status}')"
                errors += 1
            
            logger.info(
                f" {exam_no}: GPA={gpa:.2f}, Passed={passed_pct:.1f}%, Status={status} {correct}"
            )
        
        if errors > 0:
            logger.error(f"\n ❌ FOUND {errors} INCORRECT RESIT ASSIGNMENTS!")
    
    # Check students who passed ≥ 45% with GPA < 2.00 (should be Probation)
    low_gpa_adequate_pass = mastersheet[
        (mastersheet["CU Passed"] / total_cu >= 0.45)
        & (mastersheet["GPA"] < 2.00)
        & (mastersheet["CU Failed"] > 0)
    ]
    logger.info(f"\n📊 Students with ≥45% credits passed AND GPA < 2.00:")
    logger.info(f" Total: {len(low_gpa_adequate_pass)}")
    
    if len(low_gpa_adequate_pass) > 0:
        logger.info(f"\n Should ALL be 'Probation':")
        errors = 0
        for idx, row in low_gpa_adequate_pass.head(15).iterrows():
            exam_no = row["EXAMS NUMBER"]
            gpa = row["GPA"]
            cu_passed = row["CU Passed"]
            passed_pct = cu_passed / total_cu * 100
            status = row["REMARKS"]
            
            if status == "Probation":
                correct = "✅"
            else:
                correct = f"❌ WRONG! (got '{status}')"
                errors += 1
            
            logger.info(
                f" {exam_no}: GPA={gpa:.2f}, Passed={passed_pct:.1f}%, Status={status} {correct}"
            )
        
        if errors > 0:
            logger.error(f"\n ❌ FOUND {errors} INCORRECT PROBATION ASSIGNMENTS!")
    
    # Status distribution
    logger.info(f"\n📊 Overall Status Distribution:")
    status_counts = mastersheet["REMARKS"].value_counts()
    for status in ["Passed", "Resit", "Probation", "Withdrawn"]:
        count = status_counts.get(status, 0)
        pct = (count / len(mastersheet) * 100) if len(mastersheet) > 0 else 0
        logger.info(f" {status:12s}: {count:3d} ({pct:5.1f}%)")
    
    logger.info("=" * 70)

# ============================================================================
# ORIGINAL CODE CONTINUES WITH THE FIXED FUNCTIONS ABOVE
# ============================================================================

def identify_carryover_students(
    mastersheet_df, semester_key, set_name, pass_threshold=50.0
):
    """
    Identify BN students with carryover courses from current semester processing.
    UPDATED: Includes both Resit and Probation students as carryover students
    """
    carryover_students = []
    
    # Get course columns (excluding student info columns)
    course_columns = [
        col
        for col in mastersheet_df.columns
        if col
        not in [
            "S/N",
            "EXAMS NUMBER",
            "NAME",
            "FAILED COURSES",
            "REMARKS",
            "CU Passed",
            "CU Failed",
            "TCPE",
            "GPA",
            "AVERAGE",
        ]
    ]
    
    for idx, student in mastersheet_df.iterrows():
        failed_courses = []
        exam_no = str(student["EXAMS NUMBER"])
        student_name = student["NAME"]
        remarks = str(student["REMARKS"])
        
        # Include both Resit and Probation students in carryover
        if remarks in ["Resit", "Probation"]:
            for course in course_columns:
                score = student.get(course, 0)
                try:
                    score_val = float(score) if pd.notna(score) else 0
                    if score_val < pass_threshold:
                        failed_courses.append(
                            {
                                "course_code": course,
                                "original_score": score_val,
                                "semester": semester_key,
                                "set": set_name,
                                "resit_attempts": 0,
                                "best_score": score_val,
                                "status": "Failed",
                            }
                        )
                except (ValueError, TypeError):
                    continue
            
            if failed_courses:
                carryover_data = {
                    "exam_number": exam_no,
                    "name": student_name,
                    "failed_courses": failed_courses,
                    "semester": semester_key,
                    "set": set_name,
                    "identified_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_resit_attempts": 0,
                    "status": "Active",
                    "probation_status": remarks == "Probation", # Track if on probation
                }
                carryover_students.append(carryover_data)
                # Update global tracker
                student_key = f"{exam_no}_{semester_key}"
                CARRYOVER_STUDENTS[student_key] = carryover_data
    
    logger.info(
        f"📊 Identified {len(carryover_students)} carryover students ({len([s for s in carryover_students if s.get('probation_status', False)])} on probation)"
    )
    return carryover_students

def update_student_tracker(
    semester_key, exam_numbers, withdrawn_students=None, probation_students=None
):
    """
    Update the student tracker with current semester's students.
    UPDATED: Tracks probation status separately
    """
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    logger.info(f"📊 Updating student tracker for {semester_key}")
    logger.info(f"📝 Current students in this semester: {len(exam_numbers)}")
    
    # Track withdrawn students
    if withdrawn_students:
        for exam_no in withdrawn_students:
            if exam_no not in WITHDRAWN_STUDENTS:
                WITHDRAWN_STUDENTS[exam_no] = {
                    "withdrawn_semester": semester_key,
                    "withdrawn_date": datetime.now().strftime(TIMESTAMP_FMT),
                    "reappeared_semesters": [],
                }
                logger.info(f"🚫 Marked as withdrawn: {exam_no} in {semester_key}")
    
    # Track probation students
    probation_count = 0
    for exam_no in exam_numbers:
        if exam_no not in STUDENT_TRACKER:
            STUDENT_TRACKER[exam_no] = {
                "first_seen": semester_key,
                "last_seen": semester_key,
                "semesters_present": [semester_key],
                "status": "Active",
                "withdrawn": False,
                "withdrawn_semester": None,
                "probation_history": [],
                "current_probation": False,
            }
        else:
            STUDENT_TRACKER[exam_no]["last_seen"] = semester_key
            if semester_key not in STUDENT_TRACKER[exam_no]["semesters_present"]:
                STUDENT_TRACKER[exam_no]["semesters_present"].append(semester_key)
            
            # Check if student was previously withdrawn and has reappeared
            if STUDENT_TRACKER[exam_no]["withdrawn"]:
                logger.warning(f"⚠️ PREVIOUSLY WITHDRAWN STUDENT REAPPEARED: {exam_no}")
                if exam_no in WITHDRAWN_STUDENTS:
                    if (
                        semester_key
                        not in WITHDRAWN_STUDENTS[exam_no]["reappeared_semesters"]
                    ):
                        WITHDRAWN_STUDENTS[exam_no]["reappeared_semesters"].append(
                            semester_key
                        )
        
        # Update probation status if this student is on probation
        if probation_students and exam_no in probation_students:
            if semester_key not in STUDENT_TRACKER[exam_no]["probation_history"]:
                STUDENT_TRACKER[exam_no]["probation_history"].append(semester_key)
            STUDENT_TRACKER[exam_no]["current_probation"] = True
            probation_count += 1
    
    logger.info(f"📈 Total unique students tracked: {len(STUDENT_TRACKER)}")
    logger.info(f"🚫 Total withdrawn students: {len(WITHDRAWN_STUDENTS)}")
    logger.info(f"⚠️ Total probation students: {probation_count}")

# ----------------------------
# Enhanced Course Data Loading for BN
# ----------------------------
def load_bn_course_data():
    """
    Reads N-course-code-creditUnit.xlsx and returns BN course data
    """
    course_file = os.path.join(BN_COURSES_DIR, "N-course-code-creditUnit.xlsx")
    logger.info(f"Loading BN course data from: {course_file}")
    
    if not os.path.exists(course_file):
        raise FileNotFoundError(f"BN course file not found: {course_file}")
    
    xl = pd.ExcelFile(course_file)
    semester_course_maps = {}
    semester_credit_units = {}
    semester_lookup = {}
    semester_course_titles = {}
    
    for sheet in xl.sheet_names:
        df = pd.read_excel(course_file, sheet_name=sheet, engine="openpyxl", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        
        expected = ["COURSE CODE", "COURSE TITLE", "CU"]
        if not all(col in df.columns for col in expected):
            logger.warning(
                f"Warning: sheet '{sheet}' missing expected columns {expected} — skipped"
            )
            continue
        
        # Enhanced data cleaning for BN
        dfx = df.dropna(subset=["COURSE CODE", "COURSE TITLE"])
        dfx = dfx[
            ~dfx["COURSE CODE"].astype(str).str.contains("TOTAL", case=False, na=False)
        ]
        
        valid_mask = (
            dfx["CU"].astype(str).str.replace(".", "", regex=False).str.isdigit()
        )
        dfx = dfx[valid_mask]
        
        if dfx.empty:
            logger.warning(
                f"Warning: sheet '{sheet}' has no valid rows after cleaning — skipped"
            )
            continue
        
        codes = dfx["COURSE CODE"].astype(str).str.strip().tolist()
        titles = dfx["COURSE TITLE"].astype(str).str.strip().tolist()
        cus = dfx["CU"].astype(float).astype(int).tolist()
        
        # Create enhanced course map for BN
        enhanced_course_map = {}
        for title, code in zip(titles, codes):
            normalized_title = normalize_course_name(title)
            enhanced_course_map[normalized_title] = {
                "original_name": title,
                "code": code,
                "normalized": normalized_title,
            }
        
        semester_course_maps[sheet] = enhanced_course_map
        semester_credit_units[sheet] = dict(zip(codes, cus))
        semester_course_titles[sheet] = dict(zip(codes, titles))
        
        # Create BN-specific lookup variations
        norm = normalize_for_matching(sheet)
        semester_lookup[norm] = sheet
        
        # Add BN-specific variations
        norm_no_bn = norm.replace("bn-", "").replace("bn ", "")
        semester_lookup[norm_no_bn] = sheet
    
    return (
        semester_course_maps,
        semester_credit_units,
        semester_lookup,
        semester_course_titles,
    )

# ----------------------------
# Data Transformation Functions
# ----------------------------
def transform_transposed_data(df, sheet_type):
    """
    Transform transposed data format to wide format for BN.
    Input: Each student appears multiple times with different courses
    Output: Each student appears once with all courses as columns
    """
    logger.info(f"🔄 Transforming {sheet_type} sheet from transposed to wide format...")
    
    # Find the registration and name columns for BN
    reg_col = find_column_by_names(
        df, ["REG. No", "Reg No", "Registration Number", "EXAM NUMBER"]
    )
    name_col = find_column_by_names(df, ["NAME", "Full Name", "Candidate Name"])
    
    if not reg_col:
        logger.error("❌ Could not find registration column for transformation")
        return df
    
    # Get all course columns (columns that contain course codes)
    course_columns = [
        col
        for col in df.columns
        if col not in [reg_col, name_col] and col not in ["", None]
    ]
    
    logger.info(f"📊 Found {len(course_columns)} course columns: {course_columns}")
    
    # Create a new dataframe to store transformed data
    transformed_data = []
    student_dict = {}
    
    # Process each row
    for idx, row in df.iterrows():
        exam_no = str(row[reg_col]).strip()
        student_name = (
            str(row[name_col]).strip()
            if name_col and pd.notna(row.get(name_col))
            else ""
        )
        
        if exam_no not in student_dict:
            student_dict[exam_no] = {"REG. No": exam_no, "NAME": student_name}
        
        # Add course scores for this student
        for course_col in course_columns:
            score = row.get(course_col)
            if pd.notna(score) and score != "" and score != " ":
                # Create column name with sheet type suffix
                column_name = "{}_{}".format(course_col, sheet_type)
                student_dict[exam_no][column_name] = score
    
    # Convert dictionary to list
    transformed_data = list(student_dict.values())
    
    # Create new DataFrame
    if transformed_data:
        transformed_df = pd.DataFrame(transformed_data)
        logger.info(
            f"✅ Transformed data: {len(transformed_df)} students, {len(transformed_df.columns)} columns"
        )
        return transformed_df
    else:
        logger.error("❌ No data after transformation")
        return df

def detect_data_format(df, sheet_type):
    """
    Detect if data is in transposed format (students appear multiple times)
    Returns True if transposed format is detected - IMPROVED for BN
    """
    reg_col = find_column_by_names(df, ["REG. No", "Reg No", "EXAM NUMBER"])
    if not reg_col or df.empty:
        return False
    
    # Clean data first
    valid_exam_nos = df[reg_col].dropna().astype(str).str.strip()
    valid_exam_nos = valid_exam_nos[valid_exam_nos != ""]
    
    if len(valid_exam_nos) == 0:
        return False
    
    # Count occurrences
    student_counts = valid_exam_nos.value_counts()
    duplicates = (student_counts > 1).sum()
    duplicate_percentage = (duplicates / len(student_counts)) * 100
    
    # More than 30% students appearing multiple times = transposed
    is_transposed = duplicate_percentage > 30
    
    if is_transposed:
        logger.info(f"📊 Transposed data detected in {sheet_type}:")
        logger.info(
            f" {duplicates} students with duplicates ({duplicate_percentage:.1f}%)"
        )
    
    return is_transposed

# ----------------------------
# Enhanced Course Matching for BN
# ----------------------------
def normalize_course_name(name):
    """Enhanced normalization for BN course title matching"""
    if not isinstance(name, str):
        return ""
    
    # Convert to lowercase and remove extra spaces
    normalized = name.lower().strip()
    # Replace multiple spaces with single space
    normalized = re.sub(r"\s+", " ", normalized)
    # Remove special characters and extra words
    normalized = re.sub(r"[^\w\s]", "", normalized)
    
    # BN-specific substitutions for variations
    bn_substitutions = {
        "coomunication": "communication",
        "nsg": "nursing",
        "foundation": "foundations",
        "of of": "of",
        "emergency care": "emergency",
        "nursing/ emergency": "nursing emergency",
        "care i": "care",
        "foundations of nursing": "foundations nursing",
        "foundation of nsg": "foundations nursing",
        "foundation of nursing": "foundations nursing",
        # BN-specific courses
        "maternal": "maternal health",
        "child health": "child health nursing",
        "community health": "community health nursing",
        "psychiatric": "psychiatric nursing",
    }
    
    for old, new in bn_substitutions.items():
        normalized = normalized.replace(old, new)
    
    return normalized.strip()

def find_best_course_match(column_name, course_map):
    """Find the best matching BN course using enhanced matching algorithm."""
    if not isinstance(column_name, str):
        return None
    
    normalized_column = normalize_course_name(column_name)
    
    # First try exact match
    if normalized_column in course_map:
        return course_map[normalized_column]
    
    # Try partial matches with higher priority
    for course_norm, course_info in course_map.items():
        # Check if one is contained in the other
        if course_norm in normalized_column or normalized_column in course_norm:
            return course_info
    
    # Try word-based matching
    column_words = set(normalized_column.split())
    best_match = None
    best_score = 0
    
    for course_norm, course_info in course_map.items():
        course_words = set(course_norm.split())
        common_words = column_words.intersection(course_words)
        
        if common_words:
            score = len(common_words)
            # Bonus for matching BN key words
            bn_key_words = [
                "nursing",
                "health",
                "care",
                "maternal",
                "child",
                "community",
                "psychiatric",
            ]
            for word in bn_key_words:
                if word in column_words and word in course_words:
                    score += 2
            
            if score > best_score:
                best_score = score
                best_match = course_info
    
    # Require at least 2 common words or 1 key word with good match
    if best_match and best_score >= 2:
        return best_match
    
    # Final fallback: use difflib for fuzzy matching
    best_match = None
    best_ratio = 0
    for course_norm, course_info in course_map.items():
        ratio = difflib.SequenceMatcher(None, normalized_column, course_norm).ratio()
        if ratio > best_ratio and ratio > 0.6:
            best_ratio = ratio
            best_match = course_info
    
    return best_match

# ----------------------------
# Carryover Management for BN - FIXED VERSION
# ----------------------------
def initialize_carryover_tracker():
    """Initialize the global carryover tracker for BN."""
    global CARRYOVER_STUDENTS
    CARRYOVER_STUDENTS = {}
    
    # Load previous carryover records from all JSON files
    carryover_jsons = glob.glob(
        os.path.join(BN_BASE_DIR, "**/co_student*.json"), recursive=True
    )
    
    for jf in sorted(
        carryover_jsons, key=os.path.getmtime
    ):  # Load in chronological order, later files override
        try:
            with open(jf, "r") as f:
                data = json.load(f)
                for student in data:
                    student_key = f"{student['exam_number']}_{student['semester']}"
                    CARRYOVER_STUDENTS[student_key] = student
        except Exception as e:
            logger.warning(f"⚠️ Failed to load carryover from {jf}: {e}")
    
    logger.info(f"📂 Loaded {len(CARRYOVER_STUDENTS)} previous carryover records")

def save_carryover_records(carryover_students, output_dir, set_name, semester_key):
    """
    Save BN carryover student records to the clean results folder.
    USING ND SCRIPT SIMPLIFIED APPROACH
    """
    if not carryover_students:
        logger.info("ℹ️ No carryover students to save")
        return None
    
    # Use the output_dir directly (simplified from ND script)
    carryover_dir = os.path.join(output_dir, "CARRYOVER_RECORDS")
    os.makedirs(carryover_dir, exist_ok=True)
    logger.info(f"📁 Saving carryover records to: {carryover_dir}")
    
    # Generate filename with set and semester tags
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"co_student_{set_name}_{semester_key}_{timestamp}"
    
    # Save as Excel - SIMPLE STRUCTURE LIKE ND SCRIPT
    excel_file = os.path.join(carryover_dir, f"{filename}.xlsx")
    
    # Prepare data for Excel - simple structure without enhanced formatting
    records_data = []
    for student in carryover_students:
        for course in student["failed_courses"]:
            record = {
                "EXAMS NUMBER": student["exam_number"],
                "NAME": student["name"],
                "COURSE CODE": course["course_code"],
                "ORIGINAL SCORE": course["original_score"],
                "SEMESTER": student["semester"],
                "SET": student["set"],
                "RESIT ATTEMPTS": course["resit_attempts"],
                "BEST SCORE": course["best_score"],
                "STATUS": course["status"],
                "IDENTIFIED DATE": student["identified_date"],
            }
            records_data.append(record)
    
    if records_data:
        df = pd.DataFrame(records_data)
        df.to_excel(excel_file, index=False)
        logger.info(f"✅ BN Carryover records saved: {excel_file}")
        
        # Add basic formatting only (like ND script)
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Simple header formatting only
            header_row = ws[1]
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            logger.info("✅ Added basic formatting to BN carryover Excel file")
        except Exception as e:
            logger.warning(
                f"⚠️ Could not add basic formatting to BN carryover file: {e}"
            )
    
    # Save as JSON for easy processing
    json_file = os.path.join(carryover_dir, f"{filename}.json")
    with open(json_file, "w") as f:
        json.dump(carryover_students, f, indent=2)
    
    logger.info(f"📁 BN Carryover records saved in: {carryover_dir}")
    
    # Print summary
    total_failed_courses = sum(len(s["failed_courses"]) for s in carryover_students)
    logger.info(
        f"📊 Carryover Summary: {len(carryover_students)} students, {total_failed_courses} failed courses"
    )
    
    return carryover_dir

# ----------------------------
# GPA & CGPA CALCULATION FUNCTIONS - UPDATED TERMINOLOGY
# ----------------------------
def calculate_semester_gpa(
    student_row, ordered_codes, filtered_credit_units, pass_threshold
):
    """
    Calculate GPA for current semester only.
    GPA = Total Grade Points / Total Credit Units
    Uses Nigerian 5.0 grading scale.
    """
    total_grade_points = 0.0
    total_units = 0
    total_units_passed = 0
    total_units_failed = 0
    
    for code in ordered_codes:
        score = student_row.get(code, 0)
        if pd.isna(score) or score == "":
            continue
        
        score_val = float(score)
        cu = filtered_credit_units.get(code, 0)
        
        # Get grade point (5.0 scale: A=5.0, B=4.0, C=3.0, D=2.0, E=1.0, F=0.0)
        grade_point = get_grade_point(score_val)
        
        # Accumulate grade points
        total_grade_points += grade_point * cu
        total_units += cu
        
        # Track passed/failed units
        if score_val >= pass_threshold:
            total_units_passed += cu
        else:
            total_units_failed += cu
    
    # Calculate GPA for current semester
    current_semester_gpa = (
        round(total_grade_points / total_units, 2) if total_units > 0 else 0.0
    )
    
    # Calculate TCPE (Total Credit Points Earned)
    tcpe = round(total_grade_points, 1)
    
    return {
        "gpa": current_semester_gpa,
        "tcpe": tcpe,
        "cu_passed": total_units_passed,
        "cu_failed": total_units_failed,
        "total_units": total_units,
    }

def get_grade_point(score):
    """Convert numeric score to grade point (5.0 scale)."""
    try:
        score = float(score)
        if score >= 70:
            return 5.0  # A
        elif score >= 60:
            return 4.0  # B
        elif score >= 50:
            return 3.0  # C
        elif score >= 45:
            return 2.0  # D
        elif score >= 40:
            return 1.0  # E
        else:
            return 0.0  # F
    except:
        return 0.0

# ----------------------------
# ZIP File Creation for BN - FIXED VERSION
# ----------------------------
def create_bn_zip_for_set(clean_dir, set_name, ts, set_output_dir):
    """Create ZIP file for BN set with verification."""
    try:
        # Wait for file operations
        time.sleep(2)
        
        # Verify source directory
        if not os.path.exists(set_output_dir):
            logger.error(f"❌ Output directory doesn't exist: {set_output_dir}")
            return False
        
        # Count files
        file_count = sum(len(files) for _, _, files in os.walk(set_output_dir))
        if file_count == 0:
            logger.error(f"❌ No files to zip in: {set_output_dir}")
            return False
        
        # Create ZIP
        zip_filename = f"{set_name}_RESULT-{ts}.zip"
        zip_path = os.path.join(clean_dir, zip_filename)
        
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(set_output_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, set_output_dir)
                    zipf.write(file_path, arcname)
        
        # Verify ZIP
        if os.path.exists(zip_path) and os.path.getsize(zip_path) > 1000:
            with zipfile.ZipFile(zip_path, "r") as test_zip:
                if len(test_zip.namelist()) == file_count:
                    # Safe to remove original
                    shutil.rmtree(set_output_dir)
                    logger.info(f"✅ ZIP created and verified: {zip_path}")
                    return True
        
        return False
    except Exception as e:
        logger.error(f"❌ ZIP creation failed: {e}")
        return False

# ----------------------------
# Enhanced Semester Detection for BN - FIXED VERSION
# ----------------------------
def detect_bn_semester_from_filename(filename):
    """
    Detect semester from filename for BN program - HANDLES ALL FORMATS INCLUDING YOURS.
    Returns semester_key
    """
    filename_upper = filename.upper()
    logger.info(f"🔍 Detecting semester from: '{filename}'")
    
    # ✅ COMPREHENSIVE patterns matching YOUR file format: THIRD-YEAR-SECOND-SEMESTER.xlsx
    semester_patterns = {
        # Pattern 1: Direct word matching (handles hyphens, underscores, spaces)
        # Matches: THIRD-YEAR-SECOND-SEMESTER, THIRD_YEAR_SECOND_SEMESTER, THIRD YEAR SECOND SEMESTER
        r"(?:N[-_\s])?(?:THIRD|3RD)[-_\s]+YEAR[-_\s]+(?:SECOND|2ND)[-_\s]+SEMESTER": "N-THIRD-YEAR-SECOND-SEMESTER",
        r"(?:N[-_\s])?(?:THIRD|3RD)[-_\s]+YEAR[-_\s]+(?:FIRST|1ST)[-_\s]+SEMESTER": "N-THIRD-YEAR-FIRST-SEMESTER",
        r"(?:N[-_\s])?(?:SECOND|2ND)[-_\s]+YEAR[-_\s]+(?:SECOND|2ND)[-_\s]+SEMESTER": "N-SECOND-YEAR-SECOND-SEMESTER",
        r"(?:N[-_\s])?(?:SECOND|2ND)[-_\s]+YEAR[-_\s]+(?:FIRST|1ST)[-_\s]+SEMESTER": "N-SECOND-YEAR-FIRST-SEMESTER",
        r"(?:N[-_\s])?(?:FIRST|1ST)[-_\s]+YEAR[-_\s]+(?:SECOND|2ND)[-_\s]+SEMESTER": "N-FIRST-YEAR-SECOND-SEMESTER",
        r"(?:N[-_\s])?(?:FIRST|1ST)[-_\s]+YEAR[-_\s]+(?:FIRST|1ST)[-_\s]+SEMESTER": "N-FIRST-YEAR-FIRST-SEMESTER",
        
        # Pattern 2: Flexible matching with wildcards
        # Matches variations with different word orders or separators
        r"(?:THIRD|3RD).*YEAR.*(?:SECOND|2ND).*SEMESTER": "N-THIRD-YEAR-SECOND-SEMESTER",
        r"(?:THIRD|3RD).*YEAR.*(?:FIRST|1ST).*SEMESTER": "N-THIRD-YEAR-FIRST-SEMESTER",
        r"(?:SECOND|2ND).*YEAR.*(?:SECOND|2ND).*SEMESTER": "N-SECOND-YEAR-SECOND-SEMESTER",
        r"(?:SECOND|2ND).*YEAR.*(?:FIRST|1ST).*SEMESTER": "N-SECOND-YEAR-FIRST-SEMESTER",
        r"(?:FIRST|1ST).*YEAR.*(?:SECOND|2ND).*SEMESTER": "N-FIRST-YEAR-SECOND-SEMESTER",
        r"(?:FIRST|1ST).*YEAR.*(?:FIRST|1ST).*SEMESTER": "N-FIRST-YEAR-FIRST-SEMESTER",
        
        # Pattern 3: Year X Semester Y format
        # Matches: YEAR 3 SEMESTER 2, YR3SEM2, etc.
        r"(?:YEAR|YR)[-_\s]*3[-_\s]*(?:SEMESTER|SEM)[-_\s]*2": "N-THIRD-YEAR-SECOND-SEMESTER",
        r"(?:YEAR|YR)[-_\s]*3[-_\s]*(?:SEMESTER|SEM)[-_\s]*1": "N-THIRD-YEAR-FIRST-SEMESTER",
        r"(?:YEAR|YR)[-_\s]*2[-_\s]*(?:SEMESTER|SEM)[-_\s]*2": "N-SECOND-YEAR-SECOND-SEMESTER",
        r"(?:YEAR|YR)[-_\s]*2[-_\s]*(?:SEMESTER|SEM)[-_\s]*1": "N-SECOND-YEAR-FIRST-SEMESTER",
        r"(?:YEAR|YR)[-_\s]*1[-_\s]*(?:SEMESTER|SEM)[-_\s]*2": "N-FIRST-YEAR-SECOND-SEMESTER",
        r"(?:YEAR|YR)[-_\s]*1[-_\s]*(?:SEMESTER|SEM)[-_\s]*1": "N-FIRST-YEAR-FIRST-SEMESTER",
    }
    
    for pattern, semester_key in semester_patterns.items():
        match = re.search(pattern, filename_upper)
        if match:
            logger.info(f"✅ MATCHED: '{filename}' → '{semester_key}'")
            logger.info(f" Pattern: {pattern}")
            logger.info(f" Matched text: '{match.group()}'")
            return semester_key
    
    # If no match, log and raise error
    logger.error(f"❌ NO MATCH for filename: '{filename}'")
    logger.error(f" Uppercase: '{filename_upper}'")
    raise ValueError(f"Could not detect semester from filename: {filename}")

# ----------------------------
# Utilities
# ----------------------------
def normalize_path(path: str) -> str:
    """Normalize user paths for Windows/WSL/Linux."""
    path = os.path.expanduser(path)
    path = os.path.normpath(path)
    
    if platform.system().lower() == "linux" and path.startswith("C:\\"):
        path = "/mnt/" + path[0].lower() + path[2:].replace("\\", "/")
    if platform.system().lower() == "linux" and path.startswith("c:\\"):
        path = "/mnt/" + path[0].lower() + path[2:].replace("\\", "/")
    if path.startswith("/c/") and os.path.exists("/mnt/c"):
        path = path.replace("/c/", "/mnt/c/", 1)
    
    return os.path.abspath(path)

def normalize_for_matching(s):
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"\b1st\b", "first", s)
    s = re.sub(r"\b2nd\b", "second", s)
    s = re.sub(r"\b3rd\b", "third", s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def find_column_by_names(df, candidate_names):
    norm_map = {
        col: re.sub(r"\s+", " ", str(col).strip().lower()) for col in df.columns
    }
    candidates = [re.sub(r"\s+", " ", c.strip().lower()) for c in candidate_names]
    
    for cand in candidates:
        for col, ncol in norm_map.items():
            if ncol == cand:
                return col
    
    return None

def find_exam_number_column(df):
    """Find exam number column in dataframe"""
    return find_column_by_names(
        df, ["REG. No", "Reg No", "Registration Number", "EXAM NUMBER", "EXAMS NUMBER"]
    )

# ----------------------------
# Student Tracking Functions
# ----------------------------
def initialize_student_tracker():
    """Initialize the global student tracker."""
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    STUDENT_TRACKER = {}
    WITHDRAWN_STUDENTS = {}

def mark_student_withdrawn(exam_no, semester_key):
    """Mark a student as withdrawn in a specific semester."""
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    if exam_no in STUDENT_TRACKER:
        STUDENT_TRACKER[exam_no]["withdrawn"] = True
        STUDENT_TRACKER[exam_no]["withdrawn_semester"] = semester_key
        STUDENT_TRACKER[exam_no]["status"] = "Withdrawn"
    
    if exam_no not in WITHDRAWN_STUDENTS:
        WITHDRAWN_STUDENTS[exam_no] = {
            "withdrawn_semester": semester_key,
            "withdrawn_date": datetime.now().strftime(TIMESTAMP_FMT),
            "reappeared_semesters": [],
        }

def is_student_withdrawn(exam_no):
    """Check if a student has been withdrawn in any previous semester."""
    return exam_no in WITHDRAWN_STUDENTS

def get_withdrawal_history(exam_no):
    """Get withdrawal history for a student."""
    if exam_no in WITHDRAWN_STUDENTS:
        return WITHDRAWN_STUDENTS[exam_no]
    return None

def filter_out_withdrawn_students(mastersheet, semester_key):
    """
    Filter out students who were withdrawn in previous semesters.
    Returns filtered mastersheet and list of removed students.
    """
    removed_students = []
    filtered_mastersheet = mastersheet.copy()
    
    for idx, row in mastersheet.iterrows():
        exam_no = str(row["EXAMS NUMBER"]).strip()
        if is_student_withdrawn(exam_no):
            withdrawal_history = get_withdrawal_history(exam_no)
            # Only remove if student was withdrawn in a PREVIOUS semester
            if (
                withdrawal_history
                and withdrawal_history["withdrawn_semester"] != semester_key
            ):
                removed_students.append(exam_no)
                filtered_mastersheet = filtered_mastersheet[
                    filtered_mastersheet["EXAMS NUMBER"] != exam_no
                ]
    
    if removed_students:
        logger.info(
            f"🚫 Removed {len(removed_students)} previously withdrawn students from {semester_key}:"
        )
        for exam_no in removed_students:
            withdrawal_history = get_withdrawal_history(exam_no)
            logger.info(
                f" - {exam_no} (withdrawn in {withdrawal_history['withdrawn_semester']})"
            )
    
    return filtered_mastersheet, removed_students

# ----------------------------
# Set Selection Functions
# ----------------------------
def get_available_bn_sets(base_dir):
    """Get all available BN sets (SET47, SET48, etc.)"""
    # UPDATED: Sets are now under BN folder
    bn_dir = os.path.join(base_dir, "BN")
    if not os.path.exists(bn_dir):
        logger.error(f"❌ BN directory not found: {bn_dir}")
        return []
    
    sets = []
    for item in os.listdir(bn_dir):
        item_path = os.path.join(bn_dir, item)
        if os.path.isdir(item_path) and item.upper().startswith("SET"):
            sets.append(item)
    
    return sorted(sets)

def get_user_set_choice(available_sets):
    """
    Prompt user to choose which set to process.
    Returns the selected set directory name.
    """
    print("\n🎯 AVAILABLE SETS:")
    for i, set_name in enumerate(available_sets, 1):
        print("{}. {}".format(i, set_name))
    
    print("{}. Process ALL sets".format(len(available_sets) + 1))
    
    while True:
        try:
            choice = input(
                "\nEnter your choice (1-{}): ".format(len(available_sets) + 1)
            ).strip()
            
            if not choice:
                print("❌ Please enter a choice.")
                continue
            
            if choice.isdigit():
                choice_num = int(choice)
                if 1 <= choice_num <= len(available_sets):
                    selected_set = available_sets[choice_num - 1]
                    print("✅ Selected set: {}".format(selected_set))
                    return [selected_set]
                elif choice_num == len(available_sets) + 1:
                    print("✅ Selected: ALL sets")
                    return available_sets
                else:
                    print(
                        "❌ Invalid choice. Please enter a number between 1-{}.".format(
                            len(available_sets) + 1
                        )
                    )
            else:
                print("❌ Please enter a valid number.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print("❌ Error: {}. Please try again.".format(e))

# ----------------------------
# Grade and GPA calculation
# ----------------------------
def get_grade(score):
    """Convert numeric score to letter grade - single letter only."""
    try:
        score = float(score)
        if score >= 70:
            return "A"
        elif score >= 60:
            return "B"
        elif score >= 50:
            return "C"
        elif score >= 45:
            return "D"
        elif score >= 40:
            return "E"
        else:
            return "F"
    except BaseException:
        return "F"

# ----------------------------
# Upgrade Rule Functions
# ----------------------------
def get_upgrade_threshold_from_user(semester_key, set_name):
    """
    Prompt user to choose upgrade threshold for BN results.
    Returns: (min_threshold, upgraded_count) or (None, 0) if skipped
    """
    print("\n🎯 MANAGEMENT THRESHOLD UPGRADE RULE DETECTED")
    print("📚 Semester: {}".format(semester_key))
    print("📁 Set: {}".format(set_name))
    print(
        "\nSelect minimum score to upgrade (45-49). All scores >= selected value up to 49 will be upgraded to 50."
    )
    print("Enter 0 to skip upgrade.")
    
    while True:
        try:
            choice = input("\nEnter your choice (0, 45, 46, 47, 48, 49): ").strip()
            if not choice:
                print("❌ Please enter a value.")
                continue
            
            if choice == "0":
                print("⏭️ Skipping upgrade for this semester.")
                return None, 0
            
            if choice in ["45", "46", "47", "48", "49"]:
                min_threshold = int(choice)
                print("✅ Upgrade rule selected: {}–49 → 50".format(min_threshold))
                return min_threshold, 0
            else:
                print("❌ Invalid choice. Please enter 0, 45, 46, 47, 48, or 49.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print("❌ Error: {}. Please try again.".format(e))

def apply_upgrade_rule(mastersheet, ordered_codes, min_threshold):
    """
    Apply upgrade rule to mastersheet scores.
    Returns: (updated_mastersheet, upgraded_count)
    """
    if min_threshold is None:
        return mastersheet, 0
    
    upgraded_count = 0
    upgraded_students = set()
    logger.info(f"🔄 Applying upgrade rule: {min_threshold}–49 → 50")
    
    for code in ordered_codes:
        for idx in mastersheet.index:
            score = mastersheet.at[idx, code]
            if isinstance(score, (int, float)) and min_threshold <= score <= 49:
                exam_no = mastersheet.at[idx, "EXAMS NUMBER"]
                original_score = score
                mastersheet.at[idx, code] = 50
                upgraded_count += 1
                upgraded_students.add(exam_no)
                
                # Log first few upgrades for visibility
                if upgraded_count <= 5:
                    logger.info(f"🔼 {exam_no} - {code}: {original_score} → 50")
    
    if upgraded_count > 0:
        logger.info(
            f"✅ Upgraded {upgraded_count} scores from {min_threshold}–49 to 50"
        )
        logger.info(f"📊 Affected {len(upgraded_students)} students")
    else:
        logger.info(f"ℹ️ No scores found in range {min_threshold}–49 to upgrade")
    
    return mastersheet, upgraded_count

# ----------------------------
# Semester Display Info
# ----------------------------
def get_semester_display_info(semester_key):
    """Get display information for a given semester key.
    Returns: (year, semester_num, level_display, semester_display, set_code)"""
    semester_lower = semester_key.lower()
    
    if "first-year-first-semester" in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BN1"
    elif "first-year-second-semester" in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "BN1"
    elif "second-year-first-semester" in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "BN2"
    elif "second-year-second-semester" in semester_lower:
        return 2, 2, "YEAR TWO", "SECOND SEMESTER", "BN2"
    elif "third-year-first-semester" in semester_lower:
        return 3, 1, "YEAR THREE", "FIRST SEMESTER", "BN3"
    elif "third-year-second-semester" in semester_lower:
        return 3, 2, "YEAR THREE", "SECOND SEMESTER", "BN3"
    elif "first" in semester_lower and "second" not in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BN1"
    elif "second" in semester_lower and "third" not in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "BN1"
    elif "third" in semester_lower:
        return 3, 1, "YEAR THREE", "FIRST SEMESTER", "BN3"
    else:
        # Default to first semester, first year
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BN1"

# ----------------------------
# GPA Loading Functions - UPDATED WITH CORRECT TERMINOLOGY
# ----------------------------
def load_previous_gpas_from_processed_files(
    output_dir, current_semester_key, timestamp
):
    """
    Load previous GPA data from previously processed mastersheets - UPDATED TERMINOLOGY.
    Returns dict: {exam_number: previous_cgpa}
    """
    previous_cgpas = {}
    logger.info(f"\n🔍 LOADING PREVIOUS CGPA for: {current_semester_key}")
    
    # Determine previous semester based on current
    current_year, current_semester_num, _, _, _ = get_semester_display_info(
        current_semester_key
    )
    
    # Map current semester to previous semester
    semester_sequence = {
        (1, 1): None,  # First semester of first year - no previous CGPA
        (1, 2): "N-FIRST-YEAR-FIRST-SEMESTER",
        (2, 1): "N-FIRST-YEAR-SECOND-SEMESTER",
        (2, 2): "N-SECOND-YEAR-FIRST-SEMESTER",
        (3, 1): "N-SECOND-YEAR-SECOND-SEMESTER",
        (3, 2): "N-THIRD-YEAR-FIRST-SEMESTER",
    }
    
    prev_semester = semester_sequence.get((current_year, current_semester_num))
    
    if not prev_semester:
        logger.info("📊 First semester of first year - no previous CGPA available")
        return previous_cgpas
    
    logger.info(f"🔍 Looking for previous CGPA data from: {prev_semester}")
    
    # CRITICAL FIX: Look in the output directory directly for mastersheet
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")
    if not os.path.exists(mastersheet_path):
        logger.error(f"❌ Mastersheet not found: {mastersheet_path}")
        return previous_cgpas
    
    try:
        # Read the Excel file properly, skipping the header rows
        df = pd.read_excel(mastersheet_path, sheet_name=prev_semester, header=5)
        logger.info(f"📋 Columns in {prev_semester}: {df.columns.tolist()}")
        
        # Find the actual column names
        exam_col = None
        gpa_col = None
        for col in df.columns:
            col_str = str(col).upper().strip()
            if "EXAM" in col_str or "REG" in col_str or "NUMBER" in col_str:
                exam_col = col
            elif "GPA" in col_str:
                gpa_col = col
                break
        
        if exam_col and gpa_col:
            logger.info(f"✅ Found exam column: '{exam_col}', GPA column: '{gpa_col}'")
            cgpas_loaded = 0
            for idx, row in df.iterrows():
                exam_no = str(row[exam_col]).strip()
                # Clean exam number
                if "." in exam_no and exam_no.endswith(".0"):
                    exam_no = exam_no[:-2]
                
                gpa = row[gpa_col]
                if (
                    pd.notna(gpa)
                    and pd.notna(exam_no)
                    and exam_no != "nan"
                    and exam_no != ""
                ):
                    try:
                        previous_cgpas[exam_no] = float(gpa)
                        cgpas_loaded += 1
                        if cgpas_loaded <= 5:
                            logger.info(f"📝 Loaded CGPA: {exam_no} → {gpa}")
                    except (ValueError, TypeError):
                        continue
            
            logger.info(
                f"✅ Loaded previous CGPAs for {cgpas_loaded} students from {prev_semester}"
            )
        else:
            logger.error(f"❌ Could not find required columns in {prev_semester}")
    except Exception as e:
        logger.warning(f"⚠️ Could not read mastersheet: {str(e)}")
        traceback.print_exc()
    
    logger.info(f"📊 FINAL: Loaded {len(previous_cgpas)} previous CGPAs")
    return previous_cgpas

def load_all_previous_gpas_for_cgpa(output_dir, current_semester_key, timestamp):
    """
    Load ALL previous GPAs from all completed semesters for CGPA calculation.
    Returns dict: {exam_number: {'gpas': [gpa1, gpa2, ...], 'credits': [credits1, credits2, ...]}}
    """
    logger.info(
        f"\n🔍 LOADING ALL PREVIOUS GPAs for CGPA calculation: {current_semester_key}"
    )
    
    current_year, current_semester_num, _, _, _ = get_semester_display_info(
        current_semester_key
    )
    
    # Determine which semesters to load based on current semester
    semesters_to_load = []
    
    if current_semester_num == 1 and current_year == 1:
        # First semester - no previous data
        return {}
    elif current_semester_num == 2 and current_year == 1:
        # Second semester of first year - load first semester
        semesters_to_load = ["N-FIRST-YEAR-FIRST-SEMESTER"]
    elif current_semester_num == 1 and current_year == 2:
        # First semester of second year - load both first year semesters
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
        ]
    elif current_semester_num == 2 and current_year == 2:
        # Second semester of second year - load all previous semesters
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
        ]
    elif current_semester_num == 1 and current_year == 3:
        # First semester of third year - load all previous semesters
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
            "N-SECOND-YEAR-SECOND-SEMESTER",
        ]
    elif current_semester_num == 2 and current_year == 3:
        # Second semester of third year - load all previous semesters
        semesters_to_load = [
            "N-FIRST-YEAR-FIRST-SEMESTER",
            "N-FIRST-YEAR-SECOND-SEMESTER",
            "N-SECOND-YEAR-FIRST-SEMESTER",
            "N-SECOND-YEAR-SECOND-SEMESTER",
            "N-THIRD-YEAR-FIRST-SEMESTER",
        ]
    
    logger.info(f"📚 Semesters to load for CGPA: {semesters_to_load}")
    
    all_student_data = {}
    mastersheet_path = os.path.join(output_dir, "mastersheet_{}.xlsx".format(timestamp))
    
    if not os.path.exists(mastersheet_path):
        logger.error(f"❌ Mastersheet not found: {mastersheet_path}")
        return {}
    
    for semester in semesters_to_load:
        logger.info(f"📖 Loading data from: {semester}")
        try:
            # Read the Excel file properly, skipping the header rows
            df = pd.read_excel(mastersheet_path, sheet_name=semester, header=5)
            
            # Filter valid rows
            exam_col = find_exam_number_column(df)
            if exam_col:
                df = df[df[exam_col].notna() & (df[exam_col] != "")]
            
            # Find columns
            exam_col = find_exam_number_column(df)
            gpa_col = None
            credit_passed_col = None
            credit_failed_col = None
            
            for col in df.columns:
                col_str = str(col).upper().strip()
                if "GPA" in col_str:
                    gpa_col = col
                elif "CU PASSED" in col_str or "CREDIT" in col_str:
                    credit_passed_col = col
                elif "CU FAILED" in col_str:
                    credit_failed_col = col
            
            if exam_col and gpa_col:
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    # Clean exam number
                    if "." in exam_no and exam_no.endswith(".0"):
                        exam_no = exam_no[:-2]
                    
                    gpa = row[gpa_col]
                    if (
                        pd.notna(gpa)
                        and pd.notna(exam_no)
                        and exam_no != "nan"
                        and exam_no != ""
                    ):
                        try:
                            # Get credits completed (use CU Passed if available, otherwise estimate)
                            credits_passed = (
                                int(row[credit_passed_col])
                                if credit_passed_col
                                and pd.notna(row[credit_passed_col])
                                else 0
                            )
                            credits_failed = (
                                int(row[credit_failed_col])
                                if credit_failed_col
                                and pd.notna(row[credit_failed_col])
                                else 0
                            )
                            credits_completed = credits_passed + credits_failed
                            
                            if credits_completed == 0:
                                credits_completed = 30  # Typical estimate
                            
                            if exam_no not in all_student_data:
                                all_student_data[exam_no] = {"gpas": [], "credits": []}
                            
                            all_student_data[exam_no]["gpas"].append(float(gpa))
                            all_student_data[exam_no]["credits"].append(
                                credits_completed
                            )
                        except (ValueError, TypeError):
                            continue
        except Exception as e:
            logger.warning(f"⚠️ Could not load data from {semester}: {str(e)}")
    
    logger.info(f"📊 Loaded cumulative data for {len(all_student_data)} students")
    return all_student_data

def format_failed_courses_remark(failed_courses, max_line_length=60):
    """Format failed courses remark with line breaks for long lists.
    Returns list of formatted lines."""
    if not failed_courses:
        return [""]
    
    failed_str = ", ".join(sorted(failed_courses))
    
    # If the string is short enough, return as single line
    if len(failed_str) <= max_line_length:
        return [failed_str]
    
    # Split into multiple lines
    lines = []
    current_line = ""
    
    for course in sorted(failed_courses):
        if not current_line:
            current_line = course
        elif len(current_line) + len(course) + 2 <= max_line_length:  # +2 for ", "
            current_line += ", " + course
        else:
            lines.append(current_line)
            current_line = course
    
    if current_line:
        lines.append(current_line)
    
    return lines

def get_user_semester_choice():
    """Prompt user to choose which semesters to process.
    Returns list of semester keys to process."""
    print("\n🎯 SEMESTER PROCESSING OPTIONS:")
    print("1. Process ALL semesters in order")
    
    for i, semester in enumerate(BN_SEMESTER_ORDER, 2):
        year, sem_num, level, sem_display, set_code = get_semester_display_info(
            semester
        )
        print("{}. Process {} - {} only".format(i, level, sem_display))
    
    print("{}. Custom selection".format(len(BN_SEMESTER_ORDER) + 2))
    
    while True:
        try:
            choice = input(
                "\nEnter your choice (1-{}): ".format(len(BN_SEMESTER_ORDER) + 2)
            ).strip()
            
            if choice == "1":
                return BN_SEMESTER_ORDER.copy()
            elif choice.isdigit():
                choice_num = int(choice)
                if 2 <= choice_num <= len(BN_SEMESTER_ORDER) + 1:
                    return [BN_SEMESTER_ORDER[choice_num - 2]]
                elif choice_num == len(BN_SEMESTER_ORDER) + 2:
                    return get_custom_semester_selection()
                else:
                    print(
                        "❌ Invalid choice. Please enter a number between 1-{}.".format(
                            len(BN_SEMESTER_ORDER) + 2
                        )
                    )
            else:
                print("❌ Please enter a valid number.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print("❌ Error: {}. Please try again.".format(e))

def get_custom_semester_selection():
    """Allow user to select multiple semesters for processing."""
    print("\n📚 AVAILABLE SEMESTERS:")
    for i, semester in enumerate(BN_SEMESTER_ORDER, 1):
        year, sem_num, level, sem_display, set_code = get_semester_display_info(
            semester
        )
        print("{}. {} - {}".format(i, level, sem_display))
    
    print("{}. Select all".format(len(BN_SEMESTER_ORDER) + 1))
    
    selected = []
    while True:
        try:
            choices = input(
                "\nEnter semester numbers separated by commas (1-{}): ".format(
                    len(BN_SEMESTER_ORDER) + 1
                )
            ).strip()
            
            if not choices:
                print("❌ Please enter at least one semester number.")
                continue
            
            choice_list = [c.strip() for c in choices.split(",")]
            
            # Check for "select all" option
            if str(len(BN_SEMESTER_ORDER) + 1) in choice_list:
                return BN_SEMESTER_ORDER.copy()
            
            # Validate and convert choices
            valid_choices = []
            for choice in choice_list:
                if not choice.isdigit():
                    print("❌ '{}' is not a valid number.".format(choice))
                    continue
                
                choice_num = int(choice)
                if 1 <= choice_num <= len(BN_SEMESTER_ORDER):
                    valid_choices.append(choice_num)
                else:
                    print("❌ '{}' is not a valid semester number.".format(choice))
            
            if valid_choices:
                selected_semesters = [BN_SEMESTER_ORDER[i - 1] for i in valid_choices]
                print(
                    "✅ Selected semesters: {}".format(
                        [
                            get_semester_display_info(sem)[3]
                            for sem in selected_semesters
                        ]
                    )
                )
                return selected_semesters
            else:
                print("❌ No valid semesters selected. Please try again.")
        except KeyboardInterrupt:
            print("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            print("❌ Error: {}. Please try again.".format(e))

# ----------------------------
# PDF Generation - Individual Student Report - UPDATED TERMINOLOGY WITH PROBATION
# ----------------------------
def generate_individual_student_pdf(
    mastersheet_df,
    out_pdf_path,
    semester_key,
    logo_path=None,
    prev_mastersheet_df=None,
    filtered_credit_units=None,
    ordered_codes=None,
    course_titles_map=None,
    previous_gpas=None,
    cgpa_data=None,
    total_cu=None,
    pass_threshold=None,
    upgrade_min_threshold=None,
):
    """Create a PDF with one page per student matching the sample format exactly with UPDATED TERMINOLOGY AND PROBATION."""
    doc = SimpleDocTemplate(
        out_pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=20,
        bottomMargin=20,
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        "CustomHeader",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=2,
    )
    
    main_header_style = ParagraphStyle(
        "MainHeader",
        parent=styles["Normal"],
        fontSize=16,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceAfter=6,
        textColor=colors.HexColor("#800080"),
    )
    
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Normal"],
        fontSize=12,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceAfter=4,
    )
    
    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=10,
        textColor=colors.red,
    )
    
    # Left alignment style for course code and title
    left_align_style = ParagraphStyle(
        "LeftAlign",
        parent=styles["Normal"],
        fontSize=9,
        alignment=TA_LEFT,
        leftIndent=4,
    )
    
    center_align_style = ParagraphStyle(
        "CenterAlign", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER
    )
    
    # Style for remarks with smaller font
    remarks_style = ParagraphStyle(
        "RemarksStyle", parent=styles["Normal"], fontSize=8, alignment=TA_LEFT
    )
    
    elems = []
    
    for idx, r in mastersheet_df.iterrows():
        # Logo and header
        logo_img = None
        if logo_path and os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path, width=0.8 * inch, height=0.8 * inch)
            except Exception as e:
                logger.warning(f"Could not load logo: {e}")
        
        # Header table with logo and title
        if logo_img:
            header_data = [
                [
                    logo_img,
                    Paragraph("FCT COLLEGE OF NURSING SCIENCES", main_header_style),
                ]
            ]
            header_table = Table(header_data, colWidths=[1.0 * inch, 5.0 * inch])
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ]
                )
            )
            elems.append(header_table)
        else:
            elems.append(
                Paragraph("FCT COLLEGE OF NURSING SCIENCES", main_header_style)
            )
        
        # Address and contact info
        elems.append(Paragraph("P.O.Box 507, Gwagwalada-Abuja, Nigeria", header_style))
        elems.append(Paragraph("<b>EXAMINATIONS OFFICE</b>", header_style))
        elems.append(Paragraph("fctsonexamsoffice@gmail.com", header_style))
        elems.append(Spacer(1, 8))
        
        elems.append(Paragraph("STUDENT'S ACADEMIC PROGRESS REPORT", title_style))
        elems.append(Paragraph("(THIS IS NOT A TRANSCRIPT)", subtitle_style))
        elems.append(Spacer(1, 8))
        
        # Student particulars - SEPARATE FROM PASSPORT PHOTO
        exam_no = str(r.get("EXAMS NUMBER", r.get("REG. No", "")))
        student_name = str(r.get("NAME", ""))
        
        # Determine level and semester using the new function
        year, semester_num, level_display, semester_display, set_code = (
            get_semester_display_info(semester_key)
        )
        
        # Create two tables: one for student particulars, one for passport photo
        particulars_data = [
            [Paragraph("<b>STUDENT'S PARTICULARS</b>", styles["Normal"])],
            [Paragraph("<b>NAME:</b>", styles["Normal"]), student_name],
            [
                Paragraph("<b>LEVEL OF<br/>STUDY:</b>", styles["Normal"]),
                level_display,
                Paragraph("<b>SEMESTER:</b>", styles["Normal"]),
                semester_display,
            ],
            [
                Paragraph("<b>REG NO.</b>", styles["Normal"]),
                exam_no,
                Paragraph("<b>SET:</b>", styles["Normal"]),
                set_code,
            ],
        ]
        
        particulars_table = Table(
            particulars_data, colWidths=[1.2 * inch, 2.3 * inch, 0.8 * inch, 1.5 * inch]
        )
        particulars_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (3, 0)),
                    ("SPAN", (1, 1), (3, 1)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        
        # Passport photo table (separate box)
        passport_data = [
            [Paragraph("Affix Recent<br/>Passport<br/>Photograph", styles["Normal"])]
        ]
        passport_table = Table(
            passport_data, colWidths=[1.5 * inch], rowHeights=[1.2 * inch]
        )
        passport_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        
        # Create a combined table with particulars and passport side by side
        combined_data = [[particulars_table, passport_table]]
        combined_table = Table(combined_data, colWidths=[5.8 * inch, 1.5 * inch])
        combined_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        
        elems.append(combined_table)
        elems.append(Spacer(1, 12))
        
        # Semester result header
        elems.append(Paragraph("<b>SEMESTER RESULT</b>", title_style))
        elems.append(Spacer(1, 6))
        
        # Course results table - LEFT-ALIGNED CODE AND TITLE
        course_data = [
            [
                Paragraph("<b>S/N</b>", styles["Normal"]),
                Paragraph("<b>CODE</b>", styles["Normal"]),
                Paragraph("<b>COURSE TITLE</b>", styles["Normal"]),
                Paragraph("<b>UNITS</b>", styles["Normal"]),
                Paragraph("<b>SCORE</b>", styles["Normal"]),
                Paragraph("<b>GRADE</b>", styles["Normal"]),
            ]
        ]
        
        sn = 1
        total_grade_points = 0.0
        total_units = 0
        total_units_passed = 0
        total_units_failed = 0
        failed_courses_list = []
        
        for code in ordered_codes if ordered_codes else []:
            score = r.get(code)
            if pd.isna(score) or score == "":
                continue
            
            try:
                score_val = float(score)
                # Apply upgrade rule for PDF display consistency
                if (
                    upgrade_min_threshold is not None
                    and upgrade_min_threshold <= score_val <= 49
                ):
                    score_val = 50.0
                    score_display = "50"
                else:
                    score_display = str(int(round(score_val)))
                
                grade = get_grade(score_val)
                grade_point = get_grade_point(score_val)
            except Exception:
                score_display = str(score)
                grade = "F"
                grade_point = 0.0
            
            cu = filtered_credit_units.get(code, 0) if filtered_credit_units else 0
            course_title = (
                course_titles_map.get(code, code) if course_titles_map else code
            )
            
            total_grade_points += grade_point * cu
            total_units += cu
            
            if score_val >= pass_threshold:
                total_units_passed += cu
            else:
                total_units_failed += cu
                failed_courses_list.append(code)
            
            course_data.append(
                [
                    Paragraph(str(sn), center_align_style),
                    Paragraph(code, left_align_style),
                    Paragraph(course_title, left_align_style),
                    Paragraph(str(cu), center_align_style),
                    Paragraph(score_display, center_align_style),
                    Paragraph(grade, center_align_style),
                ]
            )
            sn += 1
        
        course_table = Table(
            course_data,
            colWidths=[
                0.4 * inch,
                0.7 * inch,
                2.8 * inch,
                0.6 * inch,
                0.6 * inch,
                0.6 * inch,
            ],
        )
        course_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("ALIGN", (1, 1), (2, -1), "LEFT"),
                ]
            )
        )
        
        elems.append(course_table)
        elems.append(Spacer(1, 14))
        
        # Calculate current semester GPA
        current_gpa = (
            round(total_grade_points / total_units, 2) if total_units > 0 else 0.0
        )
        
        # Calculate CGPA using the new helper functions
        exam_no_clean = str(r.get("EXAMS NUMBER", "")).strip()
        # Clean exam number
        if "." in exam_no_clean and exam_no_clean.endswith(".0"):
            exam_no_clean = exam_no_clean[:-2]
        
        # Calculate CGPA using the cumulative method
        student_cgpa_data = None
        if cgpa_data and exam_no_clean in cgpa_data:
            student_cgpa_data = cgpa_data[exam_no_clean]
            display_overall_cgpa = calculate_cgpa(
                student_cgpa_data, current_gpa, total_units
            )
        else:
            display_overall_cgpa = current_gpa
        
        # Get values from dataframe
        tcpe = round(total_grade_points, 1)
        tcup = total_units_passed
        tcuf = total_units_failed
        
        # Determine student status based on performance - USING UPDATED FUNCTION
        student_status = determine_student_status(r, total_cu, pass_threshold)
        
        # Check if student was previously withdrawn
        withdrawal_history = get_withdrawal_history(exam_no_clean)
        previously_withdrawn = withdrawal_history is not None
        
        # Format failed courses with line breaks if needed
        failed_courses_formatted = format_failed_courses_remark(failed_courses_list)
        
        # Get remarks from the dataframe (new column structure)
        remarks_status = r.get("REMARKS", "")
        failed_courses_col = r.get("FAILED COURSES", "")
        
        # Combine course-specific remarks with overall status - UPDATED WITH PROBATION
        final_remarks_lines = []
        
        if (
            previously_withdrawn
            and withdrawal_history["withdrawn_semester"] == semester_key
        ):
            if failed_courses_formatted:
                final_remarks_lines.append(
                    "Failed: {}".format(failed_courses_formatted[0])
                )
                if len(failed_courses_formatted) > 1:
                    final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("Withdrawn")
            else:
                final_remarks_lines.append("Withdrawn")
        elif previously_withdrawn:
            withdrawn_semester = withdrawal_history["withdrawn_semester"]
            year, sem_num, level, sem_display, set_code = get_semester_display_info(
                withdrawn_semester
            )
            final_remarks_lines.append(
                "STUDENT WAS WITHDRAWN FROM {} - {}".format(level, sem_display)
            )
            final_remarks_lines.append(
                "This result should not be processed as student was previously withdrawn"
            )
        elif remarks_status == "Passed":  # UPDATED: Changed from "Pass"
            final_remarks_lines.append("Passed")
        elif remarks_status == "Resit":
            if failed_courses_col:
                final_remarks_lines.append("Failed: {}".format(failed_courses_col))
                final_remarks_lines.append("To Resit Courses")
            else:
                final_remarks_lines.append("To Resit Courses")
        elif remarks_status == "Probation":  # NEW CASE
            if failed_courses_col:
                final_remarks_lines.append("Failed: {}".format(failed_courses_col))
            # Add specific probation reason based on ENFORCED rule
            passed_percentage = (tcup / total_cu * 100) if total_cu > 0 else 0
            if passed_percentage >= 45 and current_gpa < 2.00:
                final_remarks_lines.append(
                    "Placed on Probation (Passed ≥45% but GPA < 2.00)"
                )
            final_remarks_lines.append("To Resit Failed Courses")
        elif remarks_status == "Withdrawn":
            if failed_courses_col:
                final_remarks_lines.append("Failed: {}".format(failed_courses_col))
                final_remarks_lines.append("Withdrawn")
            else:
                final_remarks_lines.append("Withdrawn")
        else:
            final_remarks_lines.append(str(remarks_status))
        
        final_remarks = "<br/>".join(final_remarks_lines)
        
        display_gpa = current_gpa
        
        # Summary section - UPDATED TERMINOLOGY
        summary_data = [
            [Paragraph("<b>SUMMARY</b>", styles["Normal"]), "", "", ""],
            [
                Paragraph("<b>TCPE:</b>", styles["Normal"]),
                str(tcpe),
                Paragraph("<b>CURRENT GPA:</b>", styles["Normal"]),
                str(display_gpa),
            ],
            [
                Paragraph("<b>TCUP:</b>", styles["Normal"]),
                str(tcup),
                Paragraph("<b>PREVIOUS CGPA:</b>", styles["Normal"]),
                str(previous_gpas.get(exam_no_clean, "N/A") if previous_gpas else "N/A"),
            ],
            [
                Paragraph("<b>TCUF:</b>", styles["Normal"]),
                str(tcuf),
                Paragraph("<b>OVERALL CGPA:</b>", styles["Normal"]),
                str(display_overall_cgpa),
            ],
        ]
        
        # Add remarks with multiple lines if needed
        remarks_paragraph = Paragraph(final_remarks, remarks_style)
        summary_data.append(
            [Paragraph("<b>REMARKS:</b>", styles["Normal"]), remarks_paragraph, "", ""]
        )
        
        # Calculate row heights based on content
        row_heights = [0.3 * inch] * len(summary_data)
        # Adjust height for remarks row based on number of lines
        total_remark_lines = len(final_remarks_lines)
        if total_remark_lines > 1:
            row_heights[-1] = max(0.4 * inch, 0.2 * inch * (total_remark_lines + 1))
        
        summary_table = Table(
            summary_data,
            colWidths=[1.5 * inch, 1.0 * inch, 1.5 * inch, 1.0 * inch],
            rowHeights=row_heights,
        )
        summary_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (3, 0)),
                    ("SPAN", (1, len(summary_data) - 1), (3, len(summary_data) - 1)),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#E0E0E0")),
                    ("ALIGN", (0, 0), (3, 0), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        
        elems.append(summary_table)
        elems.append(Spacer(1, 25))
        
        # Signature section
        sig_data = [
            ["", ""],
            ["____________________", "____________________"],
            [
                Paragraph(
                    "<b>EXAMS SECRETARY</b>",
                    ParagraphStyle(
                        "SigStyle",
                        parent=styles["Normal"],
                        fontSize=10,
                        alignment=TA_CENTER,
                    ),
                ),
                Paragraph(
                    "<b>V.P. ACADEMICS</b>",
                    ParagraphStyle(
                        "SigStyle",
                        parent=styles["Normal"],
                        fontSize=10,
                        alignment=TA_CENTER,
                    ),
                ),
            ],
        ]
        
        sig_table = Table(sig_data, colWidths=[3.0 * inch, 3.0 * inch])
        sig_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        
        elems.append(sig_table)
        
        # Page break for next student
        if idx < len(mastersheet_df) - 1:
            elems.append(PageBreak())
    
    doc.build(elems)
    logger.info(f"✅ Individual student PDF written: {out_pdf_path}")

# ----------------------------
# Main BN Processing Functions - UPDATED WITH ND LOGIC AND COLUMN STRUCTURE
# ----------------------------
def process_bn_semester_files(
    semester_key,
    raw_files,
    raw_dir,
    output_dir,
    ts,
    pass_threshold,
    semester_course_maps,
    semester_credit_units,
    semester_lookup,
    semester_course_titles,
    logo_path,
    set_name,
    previous_gpas=None,
    upgrade_min_threshold=None,
):
    """Process all files for a specific BN semester."""
    logger.info(f"\n{'='*60}")
    logger.info(f"PROCESSING BN SEMESTER: {semester_key}")
    logger.info(f"{'='*60}")
    
    # Filter files for this semester
    semester_files = []
    for rf in raw_files:
        try:
            detected_sem = detect_bn_semester_from_filename(rf)
            if detected_sem == semester_key:
                semester_files.append(rf)
        except ValueError as e:
            logger.warning(f"⚠️ Could not detect semester for {rf}: {e}")
            continue
    
    if not semester_files:
        logger.warning(f"⚠️ No files found for semester {semester_key}")
        return {
            "success": False,
            "files_processed": 0,
            "error": "No files found for semester",
        }
    
    logger.info(
        f"📁 Found {len(semester_files)} files for {semester_key}: {semester_files}"
    )
    
    files_processed = 0
    
    # Process each file for this semester
    for rf in semester_files:
        raw_path = os.path.join(raw_dir, rf)
        logger.info(f"\n📄 Processing: {rf}")
        
        try:
            # Enhanced file validation
            is_valid, validation_msg = validate_raw_file(raw_path)
            if not is_valid:
                logger.error(f"❌ File validation failed for {rf}: {validation_msg}")
                continue
            
            # Load previous CGPAs for this specific semester - UPDATED TERMINOLOGY
            current_previous_cgpas = load_previous_cgpas_from_processed_files(
                output_dir, semester_key, ts
            )
            
            # Load CGPA data (all previous semesters)
            cgpa_data = load_all_previous_cgpas_for_cumulative(
                output_dir, semester_key, ts
            )
            
            # Process the file - UPDATED: Pass the loaded CGPAs correctly
            result = process_bn_single_file(
                raw_path,
                raw_dir,
                output_dir,
                ts,
                pass_threshold,
                semester_course_maps,
                semester_credit_units,
                semester_lookup,
                semester_course_titles,
                logo_path,
                semester_key,
                set_name,
                previous_gpas=current_previous_cgpas,  # UPDATED: Pass the loaded previous CGPAs
                cgpa_data=cgpa_data,
                upgrade_min_threshold=upgrade_min_threshold,
            )
            
            if result is not None:
                logger.info(f"✅ Successfully processed {rf}")
                files_processed += 1
            else:
                logger.error(f"❌ Failed to process {rf}")
        except Exception as e:
            logger.error(f"❌ Error processing {rf}: {e}")
            logger.info("🔍 Detailed error traceback:")
            traceback.print_exc()
            continue  # Continue with next file instead of stopping entirely
    
    # Return result with proper structure
    if files_processed > 0:
        # Create CGPA summary after processing all files
        mastersheet_path = os.path.join(output_dir, "mastersheet_{}.xlsx".format(ts))
        if os.path.exists(mastersheet_path):
            try:
                create_bn_cgpa_summary_sheet(
                    mastersheet_path, ts, semester_credit_units, set_name, logo_path
                )
                create_bn_analysis_sheet(mastersheet_path, ts, semester_credit_units, set_name, logo_path)
            except Exception as e:
                logger.warning(f"⚠️ Could not create summary sheets: {e}")
        
        return {
            "success": True,
            "files_processed": files_processed,
            "semester": semester_key,
        }
    else:
        return {"success": False, "files_processed": 0, "error": "No files processed"}

def process_bn_single_file(
    path,
    raw_dir,  # Added
    output_dir,
    ts,
    pass_threshold,
    semester_course_maps,
    semester_credit_units,
    semester_lookup,
    semester_course_titles,
    logo_path,
    semester_key,
    set_name,
    previous_gpas=None,
    cgpa_data=None,
    upgrade_min_threshold=None,
    is_resit=False,
):
    """Process a single BN raw file with all enhanced features and UPDATED ND LOGIC."""
    fname = os.path.basename(path)
    
    # Enhanced file validation
    is_valid, validation_msg = validate_raw_file(path)
    if not is_valid:
        logger.error(f"❌ File validation failed: {validation_msg}")
        return None
    
    logger.info(f"🔍 Processing BN file: {fname} for semester: {semester_key}")
    
    try:
        xl = pd.ExcelFile(path)
        logger.info(f"✅ Successfully opened BN Excel file: {fname}")
        logger.info(f"📋 Sheets found: {xl.sheet_names}")
        
        # Check if file has any sheets
        if not xl.sheet_names:
            logger.error(f"❌ Excel file has no sheets: {path}")
            return None
    except Exception as e:
        logger.error(f"❌ Error opening BN excel {path}: {e}")
        return None
    
    expected_sheets = ["CA", "OBJ", "EXAM"]
    dfs = {}
    
    for s in expected_sheets:
        if s in xl.sheet_names:
            try:
                dfs[s] = pd.read_excel(path, sheet_name=s, dtype=str, header=0)
                logger.info(f"✅ Loaded BN sheet {s} with shape: {dfs[s].shape}")
                
                # Check if data is in transposed format and transform if needed
                if detect_data_format(dfs[s], s):
                    logger.info(
                        f"🔄 BN Data in {s} sheet is in transposed format, transforming..."
                    )
                    dfs[s] = transform_transposed_data(dfs[s], s)
                    logger.info(f"✅ Transformed BN {s} sheet to wide format")
            except Exception as e:
                logger.error(f"❌ Error reading BN sheet {s}: {e}")
                dfs[s] = pd.DataFrame()
        else:
            logger.warning(f"⚠️ BN Sheet {s} not found in {fname}")
            dfs[s] = pd.DataFrame()
    
    if not dfs:
        logger.error("No CA/OBJ/EXAM sheets detected — skipping file.")
        return None
    
    # Use the provided semester key
    sem = semester_key
    year, semester_num, level_display, semester_display, set_code = (
        get_semester_display_info(sem)
    )
    
    logger.info(
        f"📁 Processing: {level_display} - {semester_display} - Set: {set_code}"
    )
    logger.info(f"📊 Using course sheet: {sem}")
    logger.info(
        f"📊 Previous CGPAs provided: {len(previous_gpas) if previous_gpas else 0} students"
    )
    logger.info(
        f"📊 CGPA data available for: {len(cgpa_data) if cgpa_data else 0} students"
    )
    
    # Check if semester exists in course maps
    if sem not in semester_course_maps:
        logger.error(
            f"❌ Semester '{sem}' not found in course data. Available semesters: {list(semester_course_maps.keys())}"
        )
        return None
    
    course_map = semester_course_maps[sem]
    credit_units = semester_credit_units[sem]
    course_titles = semester_course_titles[sem]
    
    # Extract ordered codes from enhanced course map
    ordered_titles = list(course_map.keys())
    ordered_codes = [course_map[t]["code"] for t in ordered_titles if course_map.get(t)]
    ordered_codes = [c for c in ordered_codes if credit_units.get(c, 0) > 0]
    filtered_credit_units = {c: credit_units[c] for c in ordered_codes}
    
    # FIXED: Calculate total_cu early so it's available for status determination
    total_cu = sum(filtered_credit_units.values()) if filtered_credit_units else 0
    logger.info(f"📊 Total credit units for {semester_key}: {total_cu}")
    
    reg_no_cols = {
        s: find_column_by_names(
            df, ["REG. No", "Reg No", "Registration Number", "EXAM NUMBER"]
        )
        for s, df in dfs.items()
    }
    
    name_cols = {
        s: find_column_by_names(df, ["NAME", "Full Name", "Candidate Name"])
        for s, df in dfs.items()
    }
    
    merged = None
    
    for s, df in dfs.items():
        df = df.copy()
        regcol = reg_no_cols.get(s)
        namecol = name_cols.get(s)
        
        if not regcol:
            regcol = df.columns[0] if len(df.columns) > 0 else None
        
        if not namecol and len(df.columns) > 1:
            namecol = df.columns[1]
        
        if regcol is None:
            logger.warning(f"Skipping sheet {s}: no reg column found")
            continue
        
        df["REG. No"] = df[regcol].astype(str).str.strip()
        if namecol:
            df["NAME"] = df[namecol].astype(str).str.strip()
        else:
            df["NAME"] = pd.NA
        
        to_drop = [c for c in [regcol, namecol] if c and c not in ["REG. No", "NAME"]]
        df.drop(columns=to_drop, errors="ignore", inplace=True)
        
        # Enhanced course matching using the new algorithm
        for col in [c for c in df.columns if c not in ["REG. No", "NAME"]]:
            best_match = find_best_course_match(col, course_map)
            if best_match:
                matched_code = best_match["code"]
                newcol = "{}_{}".format(matched_code, s.upper())
                df.rename(columns={col: newcol}, inplace=True)
                logger.info(f"✅ Matched '{col}' -> '{matched_code}'")
        
        cur_cols = ["REG. No", "NAME"] + [
            c for c in df.columns if c.endswith("_{}".format(s.upper()))
        ]
        cur = df[cur_cols].copy()
        
        if merged is None:
            merged = cur
        else:
            merged = merged.merge(cur, on="REG. No", how="outer", suffixes=("", "_dup"))
            if "NAME_dup" in merged.columns:
                merged["NAME"] = merged["NAME"].combine_first(merged["NAME_dup"])
                merged.drop(columns=["NAME_dup"], inplace=True)
    
    if merged is None or merged.empty:
        logger.error("No data merged from sheets — skipping file.")
        return None
    
    mastersheet = merged[["REG. No", "NAME"]].copy()
    mastersheet.rename(columns={"REG. No": "EXAMS NUMBER"}, inplace=True)
    mastersheet["EXAMS NUMBER"] = mastersheet["EXAMS NUMBER"].apply(
        lambda x: str(int(float(x))) if "." in str(x) else str(x)
    )
    
    for code in ordered_codes:
        ca_col = "{}_CA".format(code)
        obj_col = "{}_OBJ".format(code)
        exam_col = "{}_EXAM".format(code)
        
        ca_series = (
            pd.to_numeric(merged[ca_col], errors="coerce")
            if ca_col in merged.columns
            else pd.Series([0] * len(merged), index=merged.index)
        )
        obj_series = (
            pd.to_numeric(merged[obj_col], errors="coerce")
            if obj_col in merged.columns
            else pd.Series([0] * len(merged), index=merged.index)
        )
        exam_series = (
            pd.to_numeric(merged[exam_col], errors="coerce")
            if exam_col in merged.columns
            else pd.Series([0] * len(merged), index=merged.index)
        )
        
        ca_norm = (ca_series / 20) * 100
        obj_norm = (obj_series / 20) * 100
        exam_norm = (exam_series / 80) * 100
        
        ca_norm = ca_norm.fillna(0).clip(upper=100)
        obj_norm = obj_norm.fillna(0).clip(upper=100)
        exam_norm = exam_norm.fillna(0).clip(upper=100)
        
        total = (ca_norm * 0.2) + (((obj_norm + exam_norm) / 2) * 0.8)
        mastersheet[code] = total.round(0).clip(upper=100).values
    
    # APPLY FLEXIBLE UPGRADE RULE
    if should_use_interactive_mode():
        upgrade_min_threshold, upgraded_scores_count = get_upgrade_threshold_from_user(
            semester_key, set_name
        )
    else:
        # In non-interactive mode, use the provided threshold or None
        upgraded_scores_count = 0
        if upgrade_min_threshold is not None:
            logger.info(
                f"🔄 Applying upgrade rule from parameters: {upgrade_min_threshold}–49 → 50"
            )
    
    if upgrade_min_threshold is not None:
        mastersheet, upgraded_scores_count = apply_upgrade_rule(
            mastersheet, ordered_codes, upgrade_min_threshold
        )
    
    for c in ordered_codes:
        if c not in mastersheet.columns:
            mastersheet[c] = 0
    
    # TEMPORARY DEBUG: Check first few students
    logger.info(
        f"🔍 DEBUG: Checking status calculation for first 5 students (Total CU: {total_cu})"
    )
    for idx, row in mastersheet.head().iterrows():
        exam_no = row["EXAMS NUMBER"]
        cu_passed = sum(
            1
            for code in ordered_codes
            if float(row.get(code, 0) or 0) >= pass_threshold
        )
        cu_failed = sum(
            1 for code in ordered_codes if float(row.get(code, 0) or 0) < pass_threshold
        )
        passed_pct = (cu_passed / total_cu * 100) if total_cu > 0 else 0
        logger.info(
            f" {exam_no}: Passed={cu_passed}({passed_pct:.1f}%), Failed={cu_failed}, Total CU={total_cu}"
        )
    
    # (RE)CALCULATE REMARKS AND METRICS AFTER POSSIBLE UPDATES
    def compute_failed_courses(row):
        """Compute failed courses list for the new column structure."""
        fails = [c for c in ordered_codes if float(row.get(c, 0) or 0) < pass_threshold]
        if not fails:
            return ""
        return ", ".join(sorted(fails))
    
    def compute_student_remarks(row):
        """Compute student status for the new REMARKS column - USING UPDATED ND LOGIC"""
        # FIXED: Pass total_cu to determine_student_status
        status = determine_student_status(row, total_cu, pass_threshold)
        return status
    
    # Calculate TCPE, TCUP, TCUF correctly
    # (RE)CALCULATE METRICS AFTER POSSIBLE UPGRADES
    def calc_tcpe_tcup_tcuf(row):
        tcpe = 0.0
        tcup = 0
        tcuf = 0
        for code in ordered_codes:
            score = float(row.get(code, 0) or 0)
            cu = filtered_credit_units.get(code, 0)
            gp = get_grade_point(score)
            tcpe += gp * cu
            if score >= pass_threshold:
                tcup += cu
            else:
                tcuf += cu
        return tcpe, tcup, tcuf
    
    results = mastersheet.apply(calc_tcpe_tcup_tcuf, axis=1, result_type="expand")
    mastersheet["TCPE"] = results[0].round(1)
    mastersheet["CU Passed"] = results[1].astype(int)
    mastersheet["CU Failed"] = results[2].astype(int)
    
    # <<< GPA MUST COME FIRST >>>
    mastersheet["GPA"] = mastersheet.apply(
        lambda row: calculate_semester_gpa(
            row, ordered_codes, filtered_credit_units, pass_threshold
        )["gpa"],
        axis=1,
    )
    
    mastersheet["AVERAGE"] = mastersheet[ordered_codes].mean(axis=1).round(0)
    
    # Now REMARKS sees the real GPA
    mastersheet["FAILED COURSES"] = mastersheet.apply(compute_failed_courses, axis=1)
    mastersheet["REMARKS"] = mastersheet.apply(compute_student_remarks, axis=1)
    
    # Calculate GPA using the new function
    def calculate_gpa(row):
        gpa_data = calculate_semester_gpa(
            row, ordered_codes, filtered_credit_units, pass_threshold
        )
        return gpa_data["gpa"]
    
    mastersheet["GPA"] = mastersheet.apply(calculate_gpa, axis=1)
    mastersheet["AVERAGE"] = (
        mastersheet[[c for c in ordered_codes]].mean(axis=1).round(0)
    )
    
    # VALIDATE PROBATION/WITHDRAWAL LOGIC (like ND script)
    validate_probation_withdrawal_logic(mastersheet, total_cu)
    
    # FILTER OUT PREVIOUSLY WITHDRAWN STUDENTS
    mastersheet, removed_students = filter_out_withdrawn_students(
        mastersheet, semester_key
    )
    
    # Identify withdrawn students in this semester (after filtering)
    withdrawn_students = []
    for idx, row in mastersheet.iterrows():
        student_status = determine_student_status(row, total_cu, pass_threshold)
        if student_status == "Withdrawn":
            exam_no = str(row["EXAMS NUMBER"]).strip()
            withdrawn_students.append(exam_no)
            mark_student_withdrawn(exam_no, semester_key)
            logger.info(f"🚫 Student {exam_no} marked as withdrawn in {semester_key}")
    
    # UPDATED: Identify probation students for tracking
    probation_students = []
    for idx, row in mastersheet.iterrows():
        if row["REMARKS"] == "Probation":
            exam_no = str(row["EXAMS NUMBER"]).strip()
            probation_students.append(exam_no)
    
    # Update student tracker with current semester's students (after filtering)
    exam_numbers = mastersheet["EXAMS NUMBER"].astype(str).str.strip().tolist()
    update_student_tracker(
        semester_key, exam_numbers, withdrawn_students, probation_students
    )
    
    # IDENTIFY CARRYOVER STUDENTS - ENSURE THIS IS CALLED
    carryover_students = identify_carryover_students(
        mastersheet, semester_key, set_name, pass_threshold
    )
    
    if carryover_students:
        carryover_dir = save_carryover_records(
            carryover_students, output_dir, set_name, semester_key
        )
        logger.info(
            f"✅ Saved {len(carryover_students)} BN carryover records to: {carryover_dir}"
        )
    else:
        logger.info("✅ No carryover students identified for this semester")
    
    # NEW SORTING LOGIC: Passed, Resit, Probation, Withdrawn with secondary GPA sort
    def status_key(s):
        return {"Passed": 0, "Resit": 1, "Probation": 2, "Withdrawn": 3}.get(s, 4)
    
    mastersheet["status_key"] = mastersheet["REMARKS"].apply(status_key)
    mastersheet = (
        mastersheet.sort_values(by=["status_key", "GPA"], ascending=[True, False])
        .drop(columns=["status_key"])
        .reset_index(drop=True)
    )
    
    if "S/N" not in mastersheet.columns:
        mastersheet.insert(0, "S/N", range(1, len(mastersheet) + 1))
    else:
        mastersheet["S/N"] = range(1, len(mastersheet) + 1)
        cols = list(mastersheet.columns)
        if cols[0] != "S/N":
            cols.remove("S/N")
            mastersheet = mastersheet[["S/N"] + cols]
    
    course_cols = ordered_codes
    
    # UPDATED COLUMN ORDER
    out_cols = (
        ["S/N", "EXAMS NUMBER", "NAME"]
        + course_cols
        + [
            "FAILED COURSES",
            "REMARKS",
            "CU Passed",
            "CU Failed",
            "TCPE",
            "GPA",
            "AVERAGE",
        ]
    )
    
    for c in out_cols:
        if c not in mastersheet.columns:
            mastersheet[c] = pd.NA
    
    mastersheet = mastersheet[out_cols]
    
    if is_resit:
        # For resit processing, skip saving and PDF generation
        logger.info("ℹ️ Skipping save and PDF for resit data")
        return mastersheet
    
    # Create proper output directory structure
    output_subdir = output_dir  # Use the provided output_dir directly
    os.makedirs(output_subdir, exist_ok=True)
    
    out_xlsx = os.path.join(output_subdir, "mastersheet_{}.xlsx".format(ts))
    
    if not os.path.exists(out_xlsx):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
    else:
        wb = load_workbook(out_xlsx)
    
    if sem not in wb.sheetnames:
        ws = wb.create_sheet(title=sem)
    else:
        ws = wb[sem]
    
    try:
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
    except Exception:
        pass
    
    ws.insert_rows(1, 2)
    
    logo_path_norm = os.path.normpath(logo_path) if logo_path else None
    if logo_path_norm and os.path.exists(logo_path_norm):
        try:
            img = XLImage(logo_path_norm)
            img.width, img.height = 110, 80
            img.anchor = "A1"
            ws.add_image(img, "A1")
        except Exception as e:
            logger.warning(f"⚠ Could not place logo: {e}")
    
    # FIXED MERGE RANGE
    last_letter = get_column_letter(len(out_cols))
    ws.merge_cells(f"C1:{last_letter}1")
    title_cell = ws["C1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="1E90FF", end_color="1E90FF", fill_type="solid"
    )
    
    border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    title_cell.border = border
    
    # Use expanded semester name in the subtitle
    expanded_semester_name = "{} {}".format(level_display, semester_display)
    ws.merge_cells(f"C2:{last_letter}2")
    subtitle_cell = ws["C2"]
    subtitle_cell.value = (
        "{}/{} SESSION BASIC NURSING {} EXAMINATIONS RESULT — {}".format(
            datetime.now().year,
            datetime.now().year + 1,
            expanded_semester_name,
            datetime.now().strftime("%B %d, %Y"),
        )
    )
    subtitle_cell.font = Font(bold=True, size=12, color="000000")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    start_row = 3
    display_course_titles = []
    for t, c in zip(ordered_titles, [course_map[t]["code"] for t in ordered_titles]):
        if c in ordered_codes:
            display_course_titles.append(course_map[t]["original_name"])
    
    ws.append([""] * 3 + display_course_titles + [""] * 7)  # Updated for new columns
    
    for i, cell in enumerate(
        ws[start_row][3 : 3 + len(display_course_titles)], start=3
    ):
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=45
        )
        cell.font = Font(bold=True, size=9)
    
    ws.row_dimensions[start_row].height = 18
    
    cu_list = [filtered_credit_units.get(c, "") for c in ordered_codes]
    ws.append([""] * 3 + cu_list + [""] * 7)  # Updated for new columns
    
    for cell in ws[start_row + 1][3 : 3 + len(cu_list)]:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=135
        )
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
    
    headers = out_cols
    ws.append(headers)
    
    for cell in ws[start_row + 2]:
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="4A90E2", end_color="4A90E2", fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    
    for _, r in mastersheet.iterrows():
        rowvals = [r[col] for col in headers]
        ws.append(rowvals)
    
    # Freeze the column headers
    ws.freeze_panes = ws.cell(row=start_row + 3, column=1)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for row in ws.iter_rows(
        min_row=start_row + 3, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
    
    # Colorize course columns - SPECIAL COLOR FOR UPGRADED SCORES
    upgraded_fill = PatternFill(
        start_color="E6FFCC", end_color="E6FFCC", fill_type="solid"
    )
    passed_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    failed_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    
    for idx, code in enumerate(ordered_codes, start=4):
        col_letter = get_column_letter(idx)
        for r_idx in range(start_row + 3, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=idx)
            try:
                val = float(cell.value) if cell.value not in (None, "") else 0
                if (
                    upgrade_min_threshold is not None
                    and upgrade_min_threshold <= val <= 49
                ):
                    # This score was upgraded - use special color
                    cell.fill = upgraded_fill
                    cell.font = Font(color="006600", bold=True)
                elif val >= pass_threshold:
                    cell.fill = passed_fill
                    cell.font = Font(color="006100")
                else:
                    cell.fill = failed_fill
                    cell.font = Font(color="FF0000", bold=True)
            except Exception:
                continue
    
    # Apply specific column alignments
    left_align_columns = ["CU Passed", "CU Failed", "TCPE", "GPA", "AVERAGE"]
    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in left_align_columns:
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row + 3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        # Center align S/N column
        elif col_name == "S/N":
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row + 3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Calculate optimal column widths
    longest_name_len = (
        max([len(str(x)) for x in mastersheet["NAME"].fillna("")])
        if "NAME" in mastersheet.columns
        else 10
    )
    name_col_width = min(max(longest_name_len + 2, 10), NAME_WIDTH_CAP)
    
    # Enhanced REMARKS column width calculation
    longest_remark_len = 0
    for remark in mastersheet["REMARKS"].fillna(""):
        remark_str = str(remark)
        total_length = len(remark_str)
        if total_length > longest_remark_len:
            longest_remark_len = total_length
    
    # Enhanced FAILED COURSES column width calculation
    longest_failed_len = 0
    for failed in mastersheet["FAILED COURSES"].fillna(""):
        failed_str = str(failed)
        if failed_str:
            total_length = len(failed_str)
            if total_length > longest_failed_len:
                longest_failed_len = total_length
    
    failed_courses_col_width = min(max(longest_failed_len + 4, 40), 80)
    remarks_col_width = min(max(longest_remark_len + 4, 15), 30)
    
    # Apply text wrapping and left alignment to FAILED COURSES and REMARKS columns
    failed_courses_col_idx = headers.index("FAILED COURSES") + 1
    remarks_col_idx = headers.index("REMARKS") + 1
    
    for row_idx in range(start_row + 3, ws.max_row + 1):
        # FAILED COURSES column
        cell = ws.cell(row=row_idx, column=failed_courses_col_idx)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # REMARKS column
        cell = ws.cell(row=row_idx, column=remarks_col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color code remarks - UPDATED WITH PROBATION
        remarks_value = str(cell.value) if cell.value else ""
        if remarks_value == "Passed":  # UPDATED: Changed from "Pass"
            cell.fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            cell.font = Font(color="006100", bold=True)
        elif remarks_value == "Resit":
            cell.fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
            cell.font = Font(color="9C6500", bold=True)
        elif remarks_value == "Probation":  # NEW
            cell.fill = PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
        elif remarks_value == "Withdrawn":
            cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            cell.font = Font(color="9C0006", bold=True)
    
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        if column_letter == "A":  # S/N
            ws.column_dimensions[column_letter].width = 6
        elif column_letter == "B" or headers[col_idx - 1] in [
            "EXAMS NUMBER",
            "EXAM NO",
        ]:
            ws.column_dimensions[column_letter].width = 18
        elif headers[col_idx - 1] == "NAME":
            ws.column_dimensions[column_letter].width = name_col_width
        elif 4 <= col_idx < 4 + len(ordered_codes):  # course columns
            ws.column_dimensions[column_letter].width = 8
        elif headers[col_idx - 1] in ["FAILED COURSES"]:
            ws.column_dimensions[column_letter].width = failed_courses_col_width
        elif headers[col_idx - 1] in ["REMARKS"]:
            ws.column_dimensions[column_letter].width = remarks_col_width
        else:
            ws.column_dimensions[column_letter].width = 12
    
    # Fails per course row
    fails_per_course = (
        mastersheet[ordered_codes].apply(lambda x: (x < pass_threshold).sum()).tolist()
    )
    footer_vals = (
        [""] * 2
        + ["FAILS PER COURSE:"]
        + fails_per_course
        + [""] * (len(headers) - 3 - len(ordered_codes))
    )
    ws.append(footer_vals)
    
    for cell in ws[ws.max_row]:
        if 4 <= cell.column < 4 + len(ordered_codes):
            cell.fill = PatternFill(
                start_color="F0E68C", end_color="F0E68C", fill_type="solid"
            )
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        elif cell.column == 3:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    
    # COMPREHENSIVE SUMMARY BLOCK - ENFORCED RULE
    total_students = len(mastersheet)
    passed_all = len(mastersheet[mastersheet["REMARKS"] == "Passed"])
    
    # Count students by status with ENFORCED rule
    resit_students = len(mastersheet[mastersheet["REMARKS"] == "Resit"])
    probation_students = len(mastersheet[mastersheet["REMARKS"] == "Probation"])
    withdrawn_students_count = len(mastersheet[mastersheet["REMARKS"] == "Withdrawn"])
    
    # ENFORCED RULE: Break down students by the new criteria
    resit_rule_students = len(
        mastersheet[
            (mastersheet["REMARKS"] == "Resit")
            & (mastersheet["CU Passed"] / total_cu >= 0.45)
            & (mastersheet["GPA"] >= 2.00)
        ]
    )
    probation_rule_students = len(
        mastersheet[
            (mastersheet["REMARKS"] == "Probation")
            & (mastersheet["CU Passed"] / total_cu >= 0.45)
            & (mastersheet["GPA"] < 2.00)
        ]
    )
    withdrawn_rule_students = len(
        mastersheet[
            (mastersheet["REMARKS"] == "Withdrawn")
            & (mastersheet["CU Passed"] / total_cu < 0.45)
        ]
    )
    
    # Add withdrawn student tracking to summary - UPDATED WITH ENFORCED RULE
    ws.append([])
    ws.append(["SUMMARY"])
    ws.append(
        [f"A total of {total_students} students registered and sat for the Examination"]
    )
    ws.append(
        [
            f"A total of {passed_all} students passed in all courses registered and are to proceed to the next semester"
        ]
    )
    ws.append(
        [
            f"A total of {resit_rule_students} students with Grade Point Average (GPA) of 2.00 and above who passed ≥45% of credit units failed various courses, and are to resit these courses in the next session."
        ]
    )
    ws.append(
        [
            f"A total of {probation_rule_students} students with Grade Point Average (GPA) below 2.00 who passed ≥45% of credit units failed various courses, and are placed on Probation, to resit these courses in the next session."
        ]
    )
    ws.append(
        [
            f"A total of {withdrawn_rule_students} students who passed less than 45% of their registered credit units have been advised to withdraw"
        ]
    )
    
    # Add upgrade notice in summary section
    if upgrade_min_threshold is not None:
        ws.append(
            [
                f"✅ Upgraded all scores between {upgrade_min_threshold}–49 to 50 as per management decision ({upgraded_scores_count} scores upgraded)"
            ]
        )
    
    # Add removed withdrawn students info
    if removed_students:
        ws.append(
            [
                f"NOTE: {len(removed_students)} previously withdrawn students were removed from this semester's results as they should not be processed."
            ]
        )
    
    ws.append(
        [
            "The above decisions are in line with the provisions of the General Information Section of the NMCN/NBTE Examinations Regulations (Pg 4) adopted by the College."
        ]
    )
    
    ws.append([])
    ws.append(
        [
            "________________________",
            "",
            "",
            "________________________",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Mrs. Abini Hauwa",
            "",
            "",
            "Mrs. Olukemi Ogunleye",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Head of Exams",
            "",
            "",
            "Chairman, ND/HND Program C'tee",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    
    wb.save(out_xlsx)
    logger.info(f"✅ Mastersheet saved: {out_xlsx}")
    
    # Generate individual student PDF
    safe_sem = re.sub(r"[^\w\-]", "_", sem)
    student_pdf_path = os.path.join(
        output_subdir, "mastersheet_students_{}_{}.pdf".format(ts, safe_sem)
    )
    
    logger.info("📊 FINAL CHECK before PDF generation:")
    logger.info(f" Previous CGPAs loaded: {len(previous_gpas) if previous_gpas else 0}")
    logger.info(
        f" CGPA data available for: {len(cgpa_data) if cgpa_data else 0} students"
    )
    
    try:
        generate_individual_student_pdf(
            mastersheet,
            student_pdf_path,
            sem,
            logo_path=logo_path_norm,
            prev_mastersheet_df=None,
            filtered_credit_units=filtered_credit_units,
            ordered_codes=ordered_codes,
            course_titles_map=course_titles,
            previous_gpas=previous_gpas,  # UPDATED: Pass the actual previous_gpas parameter
            cgpa_data=cgpa_data,
            total_cu=total_cu,
            pass_threshold=pass_threshold,
            upgrade_min_threshold=upgrade_min_threshold,
        )
        logger.info(f"✅ PDF generated successfully for {sem}")
    except Exception as e:
        logger.error(f"❌ Failed to generate student PDF for {sem}: {e}")
        traceback.print_exc()
    
    return mastersheet

# ----------------------------
# Main runner
# ----------------------------
def main():
    """Main entry point with comprehensive error handling."""
    try:
        # Initialize
        initialize_student_tracker()
        initialize_carryover_tracker()
        
        logger.info(
            "Starting BN Examination Results Processing with Enhanced Features..."
        )
        ts = datetime.now().strftime(TIMESTAMP_FMT)
        
        # Check if running in web mode
        if is_web_mode():
            uploaded_file_path = get_uploaded_file_path()
            if uploaded_file_path and os.path.exists(uploaded_file_path):
                logger.info("🔧 Running in WEB MODE with uploaded file")
                success = process_uploaded_file(
                    uploaded_file_path, normalize_path(BASE_DIR)
                )
                if success:
                    logger.info("✅ Uploaded file processing completed successfully")
                else:
                    logger.error("❌ Uploaded file processing failed")
                return
            else:
                logger.error("❌ No uploaded file found in web mode")
                return
        
        # Get parameters from form
        params = get_form_parameters()
        
        # Use the parameters
        global DEFAULT_PASS_THRESHOLD
        DEFAULT_PASS_THRESHOLD = params["pass_threshold"]
        
        base_dir_norm = normalize_path(BASE_DIR)
        logger.info(f"Using base directory: {base_dir_norm}")
        
        # Check if we should use interactive or non-interactive mode
        if should_use_interactive_mode():
            logger.info("🔧 Running in INTERACTIVE mode (CLI)")
            try:
                (
                    semester_course_maps,
                    semester_credit_units,
                    semester_lookup,
                    semester_course_titles,
                ) = load_bn_course_data()
            except Exception as e:
                logger.error(f"❌ Could not load BN course data: {e}")
                return
            
            # Get available sets and let user choose
            available_sets = get_available_bn_sets(base_dir_norm)
            if not available_sets:
                logger.error(
                    f"No BN SET* directories found in {base_dir_norm}. Nothing to process."
                )
                return
            
            logger.info(
                f"📚 Found {len(available_sets)} available BN sets: {available_sets}"
            )
            
            # Let user choose which set(s) to process
            sets_to_process = get_user_set_choice(available_sets)
            logger.info(f"\n🎯 PROCESSING SELECTED SETS: {sets_to_process}")
            
            for bn_set in sets_to_process:
                logger.info(f"\n{'='*60}")
                logger.info(f"PROCESSING BN SET: {bn_set}")
                logger.info(f"{'='*60}")
                
                raw_dir = normalize_path(
                    os.path.join(base_dir_norm, "BN", bn_set, "RAW_RESULTS")
                )
                clean_dir = normalize_path(
                    os.path.join(base_dir_norm, "BN", bn_set, "CLEAN_RESULTS")
                )
                
                # Create directories if they don't exist
                os.makedirs(raw_dir, exist_ok=True)
                os.makedirs(clean_dir, exist_ok=True)
                
                if not os.path.exists(raw_dir):
                    logger.warning(f"⚠️ BN RAW_RESULTS directory not found: {raw_dir}")
                    continue
                
                raw_files = [
                    f
                    for f in os.listdir(raw_dir)
                    if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
                ]
                if not raw_files:
                    logger.warning(f"⚠️ No raw files in {raw_dir}; skipping {bn_set}")
                    continue
                
                logger.info(
                    f"📁 Found {len(raw_files)} raw files in {bn_set}: {raw_files}"
                )
                
                # Create timestamped folder for this set
                set_output_dir = os.path.join(
                    clean_dir, "{}_RESULT-{}".format(bn_set, ts)
                )
                os.makedirs(set_output_dir, exist_ok=True)
                logger.info(f"📁 Created BN set output directory: {set_output_dir}")
                
                # Get user choice for which semesters to process
                semesters_to_process = get_user_semester_choice()
                logger.info(
                    f"\n🎯 PROCESSING SELECTED SEMESTERS for {bn_set}: {[get_semester_display_info(sem)[3] for sem in semesters_to_process]}"
                )
                
                # Process selected semesters in the correct order
                semester_processed = 0
                for semester_key in semesters_to_process:
                    if semester_key not in BN_SEMESTER_ORDER:
                        logger.warning(f"⚠️ Skipping unknown semester: {semester_key}")
                        continue
                    
                    # Check if there are files for this semester - USING BM'S APPROACH
                    semester_files_exist = False
                    for rf in raw_files:
                        try:
                            detected_sem = detect_bn_semester_from_filename(rf)
                            if detected_sem == semester_key:
                                semester_files_exist = True
                                break
                        except ValueError as e:
                            logger.warning(f"⚠️ Could not detect semester for {rf}: {e}")
                            continue
                    
                    if semester_files_exist:
                        logger.info(f"\n🎯 Processing BN {semester_key} in {bn_set}...")
                        result = process_bn_semester_files(
                            semester_key,
                            raw_files,
                            raw_dir,
                            set_output_dir,
                            ts,
                            DEFAULT_PASS_THRESHOLD,
                            semester_course_maps,
                            semester_credit_units,
                            semester_lookup,
                            semester_course_titles,
                            DEFAULT_LOGO_PATH,
                            bn_set,
                        )
                        if result and result.get("success", False):
                            semester_processed += result.get("files_processed", 0)
                    else:
                        logger.warning(
                            f"⚠️ No files found for BN {semester_key} in {bn_set}, skipping..."
                        )
                
                # Create ZIP of BN results ONLY if files were processed
                if semester_processed > 0:
                    try:
                        zip_success = create_bn_zip_for_set(
                            clean_dir, bn_set, ts, set_output_dir
                        )
                        if zip_success:
                            logger.info(f"✅ Successfully created ZIP for {bn_set}")
                        else:
                            logger.warning(
                                f"⚠️ ZIP creation failed for {bn_set}, files remain in: {set_output_dir}"
                            )
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to create BN ZIP for {bn_set}: {e}")
                        traceback.print_exc()
                else:
                    logger.warning(
                        f"⚠️ No files processed for {bn_set}, skipping ZIP creation"
                    )
                    if os.path.exists(set_output_dir):
                        shutil.rmtree(set_output_dir)
            
            # Print BN-specific summaries
            logger.info("\n📊 BN STUDENT TRACKING SUMMARY:")
            logger.info(f"Total unique BN students tracked: {len(STUDENT_TRACKER)}")
            logger.info(f"Total BN withdrawn students: {len(WITHDRAWN_STUDENTS)}")
            if CARRYOVER_STUDENTS:
                logger.info("\n📋 BN CARRYOVER STUDENT SUMMARY:")
                logger.info(f"Total BN carryover students: {len(CARRYOVER_STUDENTS)}")
            
            # Analyze student progression
            sem_counts = {}
            for student_data in STUDENT_TRACKER.values():
                sem_count = len(student_data["semesters_present"])
                if sem_count not in sem_counts:
                    sem_counts[sem_count] = 0
                sem_counts[sem_count] += 1
            
            for sem_count, student_count in sorted(sem_counts.items()):
                logger.info(
                    f"Students present in {sem_count} semester(s): {student_count}"
                )
            
            logger.info(
                "\n✅ BN Examination Results Processing completed successfully."
            )
        else:
            logger.info("🔧 Running in NON-INTERACTIVE mode (Web)")
            success = process_in_non_interactive_mode(params, base_dir_norm)
            if success:
                logger.info(
                    "✅ BN Examination Results Processing completed successfully"
                )
            else:
                logger.error("❌ BN Examination Results Processing failed")
            return
            
    except KeyboardInterrupt:
        logger.info("\n👋 Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        traceback.print_exc()
        sys.exit(1)
    finally:
        # Cleanup
        logger.info("\n📊 Final Summary:")
        logger.info(f" Students tracked: {len(STUDENT_TRACKER)}")
        logger.info(f" Carryover students: {len(CARRYOVER_STUDENTS)}")
        logger.info(f" Withdrawn students: {len(WITHDRAWN_STUDENTS)}")

if __name__ == "__main__":
    main()