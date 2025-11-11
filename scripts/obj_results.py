import pandas as pd
import glob
import os
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------
# Step 1: Define directories (Railway + Local compatible)
# -----------------------------
IS_RAILWAY = os.getenv("RAILWAY_ENVIRONMENT") is not None

if IS_RAILWAY:
    BASE_DIR = os.getenv("BASE_DIR", "/app/EXAMS_INTERNAL")
    print("🚂 Running on Railway")
else:
    BASE_DIR = os.path.join(
        os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"
    )
    print("💻 Running locally")

input_dir = os.path.join(BASE_DIR, "OBJ_RESULT", "RAW_OBJ")
output_base_dir = os.path.join(BASE_DIR, "OBJ_RESULT", "CLEAN_OBJ")

# -----------------------------
# Step 2: Create timestamped output folder
# -----------------------------
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
folder_name = f"obj_result_{timestamp}"
output_dir = os.path.join(output_base_dir, folder_name)
os.makedirs(output_dir, exist_ok=True)

print(f"📁 Input Directory: {input_dir}")
print(f"📁 Output Directory: {output_dir}\n")


# -----------------------------
# Step 3: Helper Functions
# -----------------------------
def sanitize_filename(name):
    """Remove invalid filename characters"""
    return re.sub(r'[<>:"/\\|?*]', "_", name)


def extract_set_identifier(filename):
    """
    Extracts the session pattern like ND2023-SET1, ND2024-SET2 from filename.
    Returns 'UNKNOWN' if not matched.
    """
    match = re.search(r"(ND\d{4}-SET\d+)", filename.upper())
    return match.group(1) if match else "UNKNOWN"


def apply_professional_formatting(excel_file):
    """
    Apply professional formatting to Excel file:
    - Bold headers with background color
    - Auto-fit column widths
    - Center alignment for headers
    - Format "Overall average" rows
    - Add borders
    """
    wb = load_workbook(excel_file)
    ws = wb.active

    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    overall_avg_font = Font(bold=True, size=10, color="C00000")
    overall_avg_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Auto-fit column widths and apply formatting
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)

        # Calculate max length for auto-fit
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set column width with padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
        ws.column_dimensions[column_letter].width = adjusted_width

    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Check if this is an "Overall average" row
        mat_no_cell = row[1] if len(row) > 1 else None  # MAT NO. column
        is_overall_avg = False

        if mat_no_cell and mat_no_cell.value:
            if "overall average" in str(mat_no_cell.value).lower():
                is_overall_avg = True

        for col_idx, cell in enumerate(row):
            cell.border = thin_border

            if is_overall_avg:
                # Format "Overall average" rows
                cell.font = overall_avg_font
                cell.fill = overall_avg_fill
                cell.alignment = center_alignment if col_idx > 1 else left_alignment
            else:
                # Format regular data rows
                if col_idx == 0:  # SN column
                    cell.alignment = center_alignment
                elif col_idx == 1:  # MAT NO. column
                    cell.alignment = center_alignment
                elif col_idx == 2:  # FULL NAME column
                    cell.alignment = left_alignment
                else:  # Grade columns
                    cell.alignment = center_alignment

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save formatted workbook
    wb.save(excel_file)
    print(f"      🎨 Applied professional formatting")


# -----------------------------
# Step 4: Clean file names (remove extra spaces)
# -----------------------------
for filename in os.listdir(input_dir):
    old_path = os.path.join(input_dir, filename)
    clean_name = filename.strip()
    new_path = os.path.join(input_dir, clean_name)
    if filename != clean_name:
        os.rename(old_path, new_path)
        print(f"🧹 Renamed: '{filename}' → '{clean_name}'")

# -----------------------------
# Step 5: Collect all valid files
# -----------------------------
csv_files = glob.glob(os.path.join(input_dir, "*.csv"))
xls_files = glob.glob(os.path.join(input_dir, "*.xls")) + glob.glob(
    os.path.join(input_dir, "*.xlsx")
)
all_files = [
    f for f in (csv_files + xls_files) if not os.path.basename(f).startswith("~$")
]

if not all_files:
    print("❌ No valid input files found for Internal Examination Results.")
    print(f"   Checked directory: {input_dir}")
    print("\n🔍 Files actually present:")
    for f in os.listdir(input_dir):
        print("   -", f)
    print(
        "\n💡 Hint: Ensure your files are .csv, .xls, or .xlsx and not temporary (~$)"
    )
    exit()

print(f"📊 Found {len(all_files)} file(s) to process\n")

# -----------------------------
# Step 6: Group files by ND-SET pattern
# -----------------------------
grouped_files = {}
for file in all_files:
    set_id = extract_set_identifier(os.path.basename(file))
    grouped_files.setdefault(set_id, []).append(file)

# -----------------------------
# Step 7: Process each group
# -----------------------------
for set_id, files in grouped_files.items():
    print(f"\n📘 Processing group: {set_id} ({len(files)} file(s))\n")
    group_cleaned_dfs = []

    for file in files:
        file_name = os.path.basename(file)
        print(f"   ➤ Processing: {file_name}")

        try:
            if file.lower().endswith(".csv"):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file, engine="openpyxl")
        except Exception as e:
            print(f"⚠️  Error reading {file_name}: {e}")
            continue

        # Detect Grade column dynamically
        grade_cols = [col for col in df.columns if str(col).startswith("Grade/")]
        if not grade_cols:
            print(f"⚠️  Skipping {file_name}: No Grade column found")
            continue
        grade_col = grade_cols[0]

        required_cols = ["Surname", "First name", grade_col]
        if not all(col in df.columns for col in required_cols):
            print(f"⚠️  Skipping {file_name}: Missing one or more required columns")
            continue

        # Clean dataframe
        cleaned_df = df[required_cols].copy()
        cleaned_df.rename(
            columns={
                "Surname": "MAT NO.",
                "First name": "FULL NAME",
            },
            inplace=True,
        )

        # Normalize text format
        cleaned_df["MAT NO."] = (
            cleaned_df["MAT NO."].astype(str).str.strip().str.upper()
        )
        cleaned_df["FULL NAME"] = (
            cleaned_df["FULL NAME"].astype(str).str.strip().str.title()
        )

        # Sort by MAT NO.
        cleaned_df.sort_values(by="MAT NO.", key=lambda x: x.str.upper(), inplace=True)

        # Add serial number
        cleaned_df.reset_index(drop=True, inplace=True)
        cleaned_df.insert(0, "SN", range(1, len(cleaned_df) + 1))

        # Handle "Overall average" rows
        mask = cleaned_df["MAT NO."].str.contains(
            "Overall average", case=False, na=False
        )
        cleaned_df.loc[mask, "SN"] = ""

        # Uniform uppercase headers
        cleaned_df.columns = [col.upper().strip() for col in cleaned_df.columns]

        # Save as Excel with formatting (instead of CSV)
        safe_base_name = sanitize_filename(os.path.splitext(file_name)[0])
        output_file = os.path.join(
            output_dir, f"{set_id}_cleaned_{safe_base_name}.xlsx"
        )
        cleaned_df.to_excel(output_file, index=False, engine="openpyxl")

        # Apply professional formatting
        apply_professional_formatting(output_file)

        print(f"      ✅ Saved: {output_file}\n")
        group_cleaned_dfs.append(cleaned_df)

    # Combine group files into a master sheet for that SET
    if group_cleaned_dfs:
        master_df = pd.concat(group_cleaned_dfs, ignore_index=True)
        mask_master = master_df["MAT NO."].str.contains(
            "Overall average", case=False, na=False
        )
        master_df.loc[mask_master, "SN"] = ""

        master_output = os.path.join(output_dir, f"{set_id}_MASTER_CLEANED.xlsx")
        master_df.to_excel(master_output, index=False, engine="openpyxl")

        # Apply professional formatting to master file
        apply_professional_formatting(master_output)

        print(f"🎓 Group Master Excel saved: {master_output}")
        print(f"   Total records: {len(master_df)}")

print("\n✅ All processing completed successfully!")
print(f"📂 All results are saved in: {output_dir}")
