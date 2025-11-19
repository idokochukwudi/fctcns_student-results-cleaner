"""
exam_processor_bm.py

Complete script for Basic Midwifery (BM) exam processing.
FIXED VERSION with:
1. Corrected NBTE standard logic for Probation/Withdrawn determination
2. Fixed color application for Passed, Resit, Probation, Withdrawn remarks
3. Fixed logic for student status determination
4. Auto-fit column widths
5. Proper remarks column formatting
"""

from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import re
import pandas as pd
from datetime import datetime
import platform
import difflib
import math
import glob
import json
import tempfile
import shutil
import zipfile
import time
import logging
import traceback

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ----------------------------
# BM-Specific Configuration
# ----------------------------

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def is_running_on_railway():
    """Check if we're running on Railway"""
    return any(
        key in os.environ
        for key in [
            "RAILWAY_ENVIRONMENT",
            "RAILWAY_STATIC_URL",
            "RAILWAY_PROJECT_ID",
            "RAILWAY_SERVICE_NAME",
        ]
    )


def get_base_directory():
    """Get base directory - compatible with both local and Railway environments"""
    # Check Railway environment first
    railway_base = os.getenv("BASE_DIR")
    if railway_base and os.path.exists(railway_base):
        return railway_base

    # Check if we're running on Railway but BASE_DIR doesn't exist
    if is_running_on_railway():
        # Create the directory structure on Railway
        railway_base = "/app/EXAMS_INTERNAL"
        os.makedirs(railway_base, exist_ok=True)
        os.makedirs(os.path.join(railway_base, "BM", "BM-COURSES"), exist_ok=True)
        return railway_base

    # Local development fallback - updated to match BM structure
    local_path = os.path.join(
        os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL"
    )
    if os.path.exists(local_path):
        return local_path

    # Final fallback - current directory
    return os.path.join(os.path.dirname(__file__), "EXAMS_INTERNAL")


BASE_DIR = get_base_directory()
# BM directories under BM folder
BM_BASE_DIR = os.path.join(BASE_DIR, "BM")
BM_COURSES_DIR = os.path.join(BM_BASE_DIR, "BM-COURSES")

# Ensure directories exist
os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(BM_BASE_DIR, exist_ok=True)
os.makedirs(BM_COURSES_DIR, exist_ok=True)

# Global variables for threshold upgrade
THRESHOLD_UPGRADED = False
ORIGINAL_THRESHOLD = 50.0
UPGRADE_MIN = None
UPGRADE_MAX = 49


def is_web_mode():
    """Check if running in web mode (file upload)"""
    return os.getenv("WEB_MODE") == "true"


def get_uploaded_file_path():
    """Get path of uploaded file in web mode"""
    return os.getenv("UPLOADED_FILE_PATH")


def should_use_interactive_mode():
    """Check if we should use interactive mode (CLI) or non-interactive mode (web)."""
    # If specific environment variables are set by web form, use non-interactive
    if os.getenv("SELECTED_SET") or os.getenv("PROCESSING_MODE") or is_web_mode():
        return False
    # If we're running in a terminal with stdin available, use interactive mode
    if sys.stdin.isatty():
        return True
    # Default to interactive for backward compatibility
    return True


def process_uploaded_file(uploaded_file_path, base_dir_norm):
    """
    Process uploaded file in web mode.
    This function handles the single uploaded file for web processing.
    """
    logger.info("🔧 Processing uploaded file in web mode")

    # Extract set name from filename or use default
    filename = os.path.basename(uploaded_file_path)
    set_name = "BM-UPLOADED"

    # Create temporary directory structure
    temp_dir = tempfile.mkdtemp()
    raw_dir = os.path.join(temp_dir, set_name, "RAW_RESULTS")
    clean_dir = os.path.join(temp_dir, set_name, "CLEAN_RESULTS")
    os.makedirs(raw_dir, exist_ok=True)
    os.makedirs(clean_dir, exist_ok=True)

    # Copy uploaded file to raw directory
    dest_path = os.path.join(raw_dir, filename)
    shutil.copy2(uploaded_file_path, dest_path)

    # Get parameters from environment
    params = get_form_parameters()
    ts = datetime.now().strftime(TIMESTAMP_FMT)

    try:
        # Load course data
        (
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
        ) = load_course_data()

        # Process the single file
        raw_files = [filename]

        # Detect semester from filename
        semester_key, _, _, _, _, _ = detect_semester_from_filename(filename)

        logger.info(f"🎯 Detected semester: {semester_key}")
        logger.info(f"📁 Processing uploaded file: {filename}")

        # Process the file
        result = process_single_file(
            dest_path,
            clean_dir,
            ts,
            params["pass_threshold"],
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
            DEFAULT_LOGO_PATH,
            semester_key,
            set_name,
            previous_gpas=None,
            upgrade_min_threshold=get_upgrade_threshold_from_env(),
        )

        if result is not None:
            logger.info(f"✅ Successfully processed uploaded file")

            # Zip the results
            result_folder = os.path.join(clean_dir, f"{set_name}_RESULT-{ts}")
            if os.path.exists(result_folder):
                zip_path = os.path.join(clean_dir, f"{set_name}_RESULT-{ts}.zip")
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                    for root, _, files in os.walk(result_folder):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, result_folder)
                            zipf.write(file_path, arcname)
                logger.info(f"✅ Zipped results: {zip_path}")
                # Remove the folder after zipping to clean up
                shutil.rmtree(result_folder)

            return True
        else:
            logger.error(f"❌ Failed to process uploaded file")
            return False

    except Exception as e:
        logger.error(f"❌ Error processing uploaded file: {e}")
        traceback.print_exc()
        return False
    finally:
        # Clean up temporary directory
        shutil.rmtree(temp_dir, ignore_errors=True)


def get_upgrade_threshold_from_env():
    """Get upgrade threshold from environment variables"""
    upgrade_threshold_str = os.getenv("UPGRADE_THRESHOLD", "0").strip()
    if upgrade_threshold_str and upgrade_threshold_str.isdigit():
        upgrade_value = int(upgrade_threshold_str)
        if 45 <= upgrade_value <= 49:
            return upgrade_value
    return None


def create_bm_zip_for_set(clean_dir, set_name, ts, set_output_dir):
    """Create ZIP file for BM set with verification."""
    try:
        # Wait for file operations
        time.sleep(2)

        # Verify source directory
        if not os.path.exists(set_output_dir):
            logger.error(f"❌ Output directory doesn't exist: {set_output_dir}")
            return False

        # Count files
        file_count = sum(len(files) for _, _, files in os.walk(set_output_dir))
        if file_count == 0:
            logger.error(f"❌ No files to zip in: {set_output_dir}")
            return False

        # Create ZIP
        zip_filename = f"{set_name}_RESULT-{ts}.zip"
        zip_path = os.path.join(clean_dir, zip_filename)

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(set_output_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, set_output_dir)
                    zipf.write(file_path, arcname)

        # Verify ZIP
        if os.path.exists(zip_path) and os.path.getsize(zip_path) > 1000:
            with zipfile.ZipFile(zip_path, "r") as test_zip:
                if len(test_zip.namelist()) == file_count:
                    # Safe to remove original
                    shutil.rmtree(set_output_dir)
                    logger.info(f"✅ ZIP created and verified: {zip_path}")
                    return True

        return False

    except Exception as e:
        logger.error(f"❌ ZIP creation failed: {e}")
        return False


def process_in_non_interactive_mode(params, base_dir_norm):
    """Process exams in non-interactive mode for web interface."""
    logger.info("🔧 Running in NON-INTERACTIVE mode (web interface)")

    # Use parameters from environment variables
    selected_set = params["selected_set"]
    processing_mode = params["processing_mode"]
    selected_semesters = params["selected_semesters"]

    # Get upgrade threshold from environment variable if provided
    upgrade_min_threshold = get_upgrade_threshold_from_env()

    # Get available sets
    available_sets = get_available_sets(base_dir_norm)

    if not available_sets:
        logger.error("❌ No BM sets found")
        return False

    # Remove BM-COURSES from available sets if present
    available_sets = [s for s in available_sets if s != "BM-COURSES"]

    if not available_sets:
        logger.error("❌ No valid BM sets found (only BM-COURSES present)")
        return False

    # Determine which sets to process
    if selected_set == "all":
        sets_to_process = available_sets
        logger.info(f"🎯 Processing ALL sets: {sets_to_process}")
    else:
        if selected_set in available_sets:
            sets_to_process = [selected_set]
            logger.info(f"🎯 Processing selected set: {selected_set}")
        else:
            logger.warning(
                f"⚠️ Selected set '{selected_set}' not found, processing all sets"
            )
            sets_to_process = available_sets

    # Determine which semesters to process
    if (
        processing_mode == "auto"
        or not selected_semesters
        or "all" in selected_semesters
    ):
        semesters_to_process = SEMESTER_ORDER.copy()
        logger.info(f"🎯 Processing ALL semesters: {semesters_to_process}")
    else:
        semesters_to_process = selected_semesters
        logger.info(f"🎯 Processing selected semesters: {semesters_to_process}")

    # Load course data once
    try:
        (
            semester_course_maps,
            semester_credit_units,
            semester_lookup,
            semester_course_titles,
        ) = load_course_data()
    except Exception as e:
        logger.error(f"❌ Could not load course data: {e}")
        return False

    ts = datetime.now().strftime(TIMESTAMP_FMT)

    # Process each set and semester
    total_processed = 0
    for bm_set in sets_to_process:
        logger.info(f"\n{'='*60}")
        logger.info(f"PROCESSING SET: {bm_set}")
        logger.info(f"{'='*60}")

        # BM: Raw and clean directories under BM folder
        raw_dir = os.path.join(base_dir_norm, "BM", bm_set, "RAW_RESULTS")
        clean_dir = os.path.join(base_dir_norm, "BM", bm_set, "CLEAN_RESULTS")

        # Create directories if they don't exist
        os.makedirs(raw_dir, exist_ok=True)
        os.makedirs(clean_dir, exist_ok=True)

        if not os.path.exists(raw_dir):
            logger.warning(f"⚠️ RAW_RESULTS directory not found: {raw_dir}")
            continue

        raw_files = [
            f
            for f in os.listdir(raw_dir)
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
        ]
        if not raw_files:
            logger.warning(f"⚠️ No raw files in {raw_dir}; skipping {bm_set}")
            continue

        logger.info(f"📁 Found {len(raw_files)} raw files in {bm_set}: {raw_files}")

        # Create timestamped output directory
        set_output_dir = os.path.join(clean_dir, f"{bm_set}_RESULT-{ts}")
        os.makedirs(set_output_dir, exist_ok=True)

        # Track semesters processed for this set
        semesters_processed = []

        # Process selected semesters
        for semester_key in semesters_to_process:
            if semester_key not in SEMESTER_ORDER:
                logger.warning(f"⚠️ Skipping unknown semester: {semester_key}")
                continue

            # Check if there are files for this semester
            semester_files_exist = False
            for rf in raw_files:
                detected_sem, _, _, _, _, _ = detect_semester_from_filename(rf)
                if detected_sem == semester_key:
                    semester_files_exist = True
                    break

            if semester_files_exist:
                logger.info(f"\n🎯 Processing {semester_key} in {bm_set}...")
                try:
                    # Process the semester with the upgrade threshold
                    result = process_semester_files(
                        semester_key,
                        raw_files,
                        raw_dir,
                        clean_dir,
                        ts,
                        params["pass_threshold"],
                        semester_course_maps,
                        semester_credit_units,
                        semester_lookup,
                        semester_course_titles,
                        DEFAULT_LOGO_PATH,
                        bm_set,
                        previous_gpas=None,
                        upgrade_min_threshold=upgrade_min_threshold,
                    )

                    if result is not None and result.get("success", False):
                        logger.info(f"✅ Successfully processed {semester_key}")
                        total_processed += 1
                        semesters_processed.append(semester_key)
                    else:
                        logger.error(f"❌ Failed to process {semester_key}")

                except Exception as e:
                    logger.error(f"❌ Error processing {semester_key}: {e}")
                    traceback.print_exc()
            else:
                logger.warning(
                    f"⚠️ No files found for {semester_key} in {bm_set}, skipping..."
                )

        # Only create ZIP if semesters were actually processed
        if semesters_processed:
            logger.info(
                f"📦 Creating ZIP for {bm_set} ({len(semesters_processed)} semesters)"
            )
            create_bm_zip_for_set(clean_dir, bm_set, ts, set_output_dir)
        else:
            logger.warning(f"⚠️ No semesters processed for {bm_set}, skipping ZIP")
            if os.path.exists(set_output_dir):
                shutil.rmtree(set_output_dir)

    logger.info(f"\n📊 PROCESSING SUMMARY: {total_processed} semester(s) processed")

    # NEW: Print BM-specific summaries
    logger.info("\n📊 BM STUDENT TRACKING SUMMARY:")
    logger.info("Total unique BM students tracked: {}".format(len(STUDENT_TRACKER)))
    logger.info("Total BM withdrawn students: {}".format(len(WITHDRAWN_STUDENTS)))

    if CARRYOVER_STUDENTS:
        logger.info("\n📋 BM CARRYOVER STUDENT SUMMARY:")
        logger.info("Total BM carryover students: {}".format(len(CARRYOVER_STUDENTS)))

    return total_processed > 0


def get_form_parameters():
    """Get parameters from environment variables set by the web form."""
    selected_set = os.getenv("SELECTED_SET", "all")
    processing_mode = os.getenv("PROCESSING_MODE", "auto")
    selected_semesters_str = os.getenv("SELECTED_SEMESTERS", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", "50.0"))
    generate_pdf = os.getenv("GENERATE_PDF", "True").lower() == "true"
    track_withdrawn = os.getenv("TRACK_WITHDRAWN", "True").lower() == "true"

    # Convert semester string to list
    selected_semesters = []
    if selected_semesters_str:
        selected_semesters = selected_semesters_str.split(",")

    logger.info(f"🎯 FORM PARAMETERS:")
    logger.info(f"   Selected Set: {selected_set}")
    logger.info(f"   Processing Mode: {processing_mode}")
    logger.info(f"   Selected Semesters: {selected_semesters}")
    logger.info(f"   Pass Threshold: {pass_threshold}")
    logger.info(f"   Generate PDF: {generate_pdf}")
    logger.info(f"   Track Withdrawn: {track_withdrawn}")

    return {
        "selected_set": selected_set,
        "processing_mode": processing_mode,
        "selected_semesters": selected_semesters,
        "pass_threshold": pass_threshold,
        "generate_pdf": generate_pdf,
        "track_withdrawn": track_withdrawn,
    }


def get_pass_threshold():
    """Get pass threshold - now handles upgrade logic interactively."""
    threshold_str = os.getenv("PASS_THRESHOLD", "50.0")
    try:
        threshold = float(threshold_str)
    except ValueError:
        threshold = 50.0

    return threshold


DEFAULT_PASS_THRESHOLD = get_pass_threshold()

TIMESTAMP_FMT = "%Y-%m-%d_%H%M%S"

DEFAULT_LOGO_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "launcher", "static", "logo.png")
)

NAME_WIDTH_CAP = 40

# Define BM semester processing order (3-year program) - FIXED: Changed BM- prefix to M- prefix consistently
SEMESTER_ORDER = [
    "M-FIRST-YEAR-FIRST-SEMESTER",
    "M-FIRST-YEAR-SECOND-SEMESTER",
    "M-SECOND-YEAR-FIRST-SEMESTER",
    "M-SECOND-YEAR-SECOND-SEMESTER",
    "M-THIRD-YEAR-FIRST-SEMESTER",
    "M-THIRD-YEAR-SECOND-SEMESTER",
]

# Global student tracker
STUDENT_TRACKER = {}
WITHDRAWN_STUDENTS = {}
CARRYOVER_STUDENTS = {}  # New global carryover tracker for BM

# ----------------------------
# NEW: CGPA SUMMARY SHEET FUNCTION
# ----------------------------

def create_bm_cgpa_summary_sheet(mastersheet_path, timestamp, set_name):
    """Create a CGPA summary sheet that aggregates GPA across all BM semesters with professional title and short headings."""
    try:
        logger.info("📊 Creating BM CGPA Summary Sheet...")

        # Load the mastersheet workbook
        wb = load_workbook(mastersheet_path)

        # Collect GPA data from all BM semesters
        cgpa_data = {}

        # Map semester names to short codes
        semester_short_codes = {
            "M-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
            "M-FIRST-YEAR-SECOND-SEMESTER": "Y1S2", 
            "M-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
            "M-SECOND-YEAR-SECOND-SEMESTER": "Y2S2",
            "M-THIRD-YEAR-FIRST-SEMESTER": "Y3S1",
            "M-THIRD-YEAR-SECOND-SEMESTER": "Y3S2"
        }

        for sheet_name in wb.sheetnames:
            if sheet_name in SEMESTER_ORDER:
                df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)

                # Find exam number and GPA columns
                exam_col = find_column_by_names(
                    df, ["EXAMS NUMBER", "REG. No", "Reg No"]
                )
                gpa_col = None
                name_col = None

                for col in df.columns:
                    col_str = str(col).upper()
                    if "GPA" in col_str:
                        gpa_col = col
                    elif "NAME" in col_str:
                        name_col = col

                if exam_col and gpa_col:
                    for idx, row in df.iterrows():
                        exam_no = str(row[exam_col]).strip()
                        # Remove .0 from exam numbers
                        if exam_no.endswith('.0'):
                            exam_no = exam_no[:-2]
                        if exam_no and exam_no != "nan":
                            if exam_no not in cgpa_data:
                                cgpa_data[exam_no] = {
                                    "name": (
                                        row[name_col]
                                        if name_col and pd.notna(row.get(name_col))
                                        else ""
                                    ),
                                    "gpas": {},
                                    "remarks": {},  # Store remarks for status determination
                                }
                            cgpa_data[exam_no]["gpas"][sheet_name] = row[gpa_col]
                            # Store remarks if available
                            if "REMARKS" in df.columns:
                                cgpa_data[exam_no]["remarks"][sheet_name] = row.get("REMARKS", "")

        # Create CGPA summary dataframe
        summary_data = []
        for exam_no, data in cgpa_data.items():
            # Ensure exam number doesn't have .0
            clean_exam_no = exam_no[:-2] if exam_no.endswith('.0') else exam_no
            row = {"EXAMS NUMBER": clean_exam_no, "NAME": data["name"]}

            # Add GPA for each semester
            total_gpa = 0
            semester_count = 0
            semesters_completed = 0

            for semester in SEMESTER_ORDER:
                if semester in data["gpas"]:
                    # Use short code for semester column
                    short_code = semester_short_codes.get(semester, semester)
                    row[short_code] = data["gpas"][semester]
                    if pd.notna(data["gpas"][semester]):
                        total_gpa += data["gpas"][semester]
                        semester_count += 1
                        semesters_completed += 1
                else:
                    short_code = semester_short_codes.get(semester, semester)
                    row[short_code] = None

            # Calculate CGPA
            row["CGPA"] = (
                round(total_gpa / semester_count, 2) if semester_count > 0 else 0.0
            )

            # Determine STATUS and GRADUATED columns
            status = "Active"
            graduated = "In Progress"

            # Check if student is withdrawn in any semester
            is_withdrawn = False
            for semester, remarks in data.get("remarks", {}).items():
                if remarks == "Withdrawn":
                    is_withdrawn = True
                    break

            if is_withdrawn:
                status = "Withdrawn"
                graduated = "Withdrawn"  # Better wording for withdrawn students
            elif semesters_completed == len(SEMESTER_ORDER):
                # Completed all semesters - consider graduated
                status = "Active"
                graduated = "Graduated"
            else:
                # Still completing semesters
                status = "Active"
                graduated = "In Progress"

            # Add the two new columns
            row["STATUS"] = status
            row["GRADUATED"] = graduated

            summary_data.append(row)

        # Create summary dataframe
        summary_df = pd.DataFrame(summary_data)

        # Add the summary sheet to the workbook
        if "CGPA_SUMMARY" in wb.sheetnames:
            del wb["CGPA_SUMMARY"]

        ws = wb.create_sheet("CGPA_SUMMARY")

        # Determine the number of columns needed to span the student data
        short_semesters = [semester_short_codes.get(sem, sem) for sem in SEMESTER_ORDER]
        headers = ["EXAMS NUMBER", "NAME"] + short_semesters + ["CGPA", "STATUS", "GRADUATED"]
        num_columns = len(headers)

        # Create professional title - EXPANDED TO MATCH STUDENT HEADING WIDTH
        title_row1 = 1
        ws.merge_cells(f'A{title_row1}:{get_column_letter(num_columns)}{title_row1}')
        title_cell1 = ws[f'A{title_row1}']
        title_cell1.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA"
        title_cell1.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell1.alignment = Alignment(horizontal="center", vertical="center")
        title_cell1.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )

        title_row2 = 2
        ws.merge_cells(f'A{title_row2}:{get_column_letter(num_columns)}{title_row2}')
        title_cell2 = ws[f'A{title_row2}']
        title_cell2.value = "DEPARTMENT OF MIDWIFERY"
        title_cell2.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell2.alignment = Alignment(horizontal="center", vertical="center")
        title_cell2.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )

        # Add set record title - EXPANDED TO MATCH STUDENT HEADING WIDTH
        set_row = title_row2 + 1
        ws.merge_cells(f'A{set_row}:{get_column_letter(num_columns)}{set_row}')
        set_cell = ws[f'A{set_row}']
        set_cell.value = f"{set_name.upper()} - CGPA SUMMARY"
        set_cell.font = Font(bold=True, size=12, color="000000")
        set_cell.alignment = Alignment(horizontal="center", vertical="center")
        set_cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        # Write header (start from row 4)
        header_row = set_row + 1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4A90E2", end_color="4A90E2", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Write data - ensure exam numbers don't have .0
        for row_idx, row_data in enumerate(summary_data, header_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                # Remove .0 from exam numbers if they appear
                if header == "EXAMS NUMBER" and isinstance(value, str) and value.endswith('.0'):
                    value = value[:-2]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # Apply color coding for STATUS column
                if header == "STATUS":
                    if value == "Withdrawn":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:  # Active
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                        cell.font = Font(bold=True, color="006400")

                # Apply color coding for GRADUATED column
                elif header == "GRADUATED":
                    if value == "Graduated":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                        cell.font = Font(bold=True, color="006400")
                    elif value == "Withdrawn":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:  # In Progress
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                        cell.font = Font(bold=True, color="000000")

        # AUTO-FIT COLUMN WIDTHS for all columns
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            column_letter = get_column_letter(col_idx)
            
            # Check all rows in this column
            for row_idx in range(header_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    # Remove .0 from exam numbers for width calculation
                    cell_value_str = str(cell_value)
                    if col_idx == 1 and cell_value_str.endswith('.0'):  # EXAMS NUMBER column
                        cell_value_str = cell_value_str[:-2]
                    cell_length = len(cell_value_str)
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Set width with some padding, but reasonable limits
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            if adjusted_width < 8:  # Minimum width
                adjusted_width = 8
            ws.column_dimensions[column_letter].width = adjusted_width

        # Specific column width adjustments for better readability
        ws.column_dimensions['A'].width = 15  # EXAMS NUMBER
        ws.column_dimensions['B'].width = 25  # NAME
        # Semester columns (Y1S1, Y1S2, etc.) will auto-fit
        ws.column_dimensions[get_column_letter(len(headers) - 2)].width = 10  # CGPA column
        ws.column_dimensions[get_column_letter(len(headers) - 1)].width = 12  # STATUS column
        ws.column_dimensions[get_column_letter(len(headers))].width = 12     # GRADUATED column

        wb.save(mastersheet_path)
        logger.info("✅ BM CGPA Summary sheet created successfully with professional title and two status columns")

        return summary_df

    except Exception as e:
        logger.error(f"❌ Error creating BM CGPA summary sheet: {e}")
        return None

# ----------------------------
# NEW: ANALYSIS SHEET FUNCTION
# ----------------------------

def create_bm_analysis_sheet(mastersheet_path, timestamp, set_name):
    """Create an analysis sheet with comprehensive statistics for BM with professional title and short headings."""
    try:
        logger.info("📈 Creating BM Analysis Sheet...")

        wb = load_workbook(mastersheet_path)

        # Map semester names to short codes
        semester_short_codes = {
            "M-FIRST-YEAR-FIRST-SEMESTER": "Y1S1",
            "M-FIRST-YEAR-SECOND-SEMESTER": "Y1S2", 
            "M-SECOND-YEAR-FIRST-SEMESTER": "Y2S1",
            "M-SECOND-YEAR-SECOND-SEMESTER": "Y2S2",
            "M-THIRD-YEAR-FIRST-SEMESTER": "Y3S1",
            "M-THIRD-YEAR-SECOND-SEMESTER": "Y3S2"
        }

        # Collect data from all semesters
        analysis_data = {
            "semester": [],
            "total_students": [],
            "passed_all": [],
            "carryover_students": [],
            "withdrawn_students": [],
            "average_gpa": [],
            "pass_rate": [],
        }

        for sheet_name in wb.sheetnames:
            if sheet_name in SEMESTER_ORDER:
                df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)

                # Basic statistics
                total_students = len(df)
                passed_all = (
                    len(df[df["REMARKS"] == "Passed"]) if "REMARKS" in df.columns else 0
                )

                # Calculate carryover students (Resit + Probation)
                carryover_count = 0
                if "REMARKS" in df.columns:
                    carryover_count = len(df[df["REMARKS"].isin(["Resit", "Probation"])])

                # Calculate withdrawn students
                withdrawn_count = len(df[df["REMARKS"] == "Withdrawn"]) if "REMARKS" in df.columns else 0

                # Calculate average GPA
                avg_gpa = (
                    df["GPA"].mean()
                    if "GPA" in df.columns and not df["GPA"].isna().all()
                    else 0
                )

                # Calculate pass rate
                pass_rate = (
                    (passed_all / total_students * 100) if total_students > 0 else 0
                )

                # Use short code for semester name
                short_semester = semester_short_codes.get(sheet_name, sheet_name)
                analysis_data["semester"].append(short_semester)
                analysis_data["total_students"].append(total_students)
                analysis_data["passed_all"].append(passed_all)
                analysis_data["carryover_students"].append(carryover_count)
                analysis_data["withdrawn_students"].append(withdrawn_count)
                analysis_data["average_gpa"].append(round(avg_gpa, 2))
                analysis_data["pass_rate"].append(round(pass_rate, 2))

        # Create analysis dataframe
        analysis_df = pd.DataFrame(analysis_data)

        # Add overall statistics
        overall_stats = {
            "semester": "OVERALL",
            "total_students": analysis_df["total_students"].sum(),
            "passed_all": analysis_df["passed_all"].sum(),
            "carryover_students": analysis_df["carryover_students"].sum(),
            "withdrawn_students": analysis_df["withdrawn_students"].sum(),
            "average_gpa": round(analysis_df["average_gpa"].mean(), 2),
            "pass_rate": round(analysis_df["pass_rate"].mean(), 2),
        }
        analysis_df = pd.concat(
            [analysis_df, pd.DataFrame([overall_stats])], ignore_index=True
        )

        # Add the analysis sheet to the workbook
        if "ANALYSIS" in wb.sheetnames:
            del wb["ANALYSIS"]

        ws = wb.create_sheet("ANALYSIS")

        # Determine the number of columns
        headers = [
            "SEMESTER",
            "TOTAL STUDENTS",
            "PASSED ALL",
            "CARRYOVER",
            "WITHDRAWN",
            "AVG GPA",
            "PASS RATE %",
        ]
        num_columns = len(headers)

        # Create professional title - EXPANDED TO MATCH STUDENT HEADING WIDTH
        title_row1 = 1
        ws.merge_cells(f'A{title_row1}:{get_column_letter(num_columns)}{title_row1}')
        title_cell1 = ws[f'A{title_row1}']
        title_cell1.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA"
        title_cell1.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell1.alignment = Alignment(horizontal="center", vertical="center")
        title_cell1.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )

        title_row2 = 2
        ws.merge_cells(f'A{title_row2}:{get_column_letter(num_columns)}{title_row2}')
        title_cell2 = ws[f'A{title_row2}']
        title_cell2.value = "DEPARTMENT OF MIDWIFERY"
        title_cell2.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell2.alignment = Alignment(horizontal="center", vertical="center")
        title_cell2.fill = PatternFill(
            start_color="1E90FF", end_color="1E90FF", fill_type="solid"
        )

        # Add set record title - EXPANDED TO MATCH STUDENT HEADING WIDTH
        set_row = title_row2 + 1
        ws.merge_cells(f'A{set_row}:{get_column_letter(num_columns)}{set_row}')
        set_cell = ws[f'A{set_row}']
        set_cell.value = f"{set_name.upper()} - ACADEMIC PERFORMANCE ANALYSIS"
        set_cell.font = Font(bold=True, size=12, color="000000")
        set_cell.alignment = Alignment(horizontal="center", vertical="center")
        set_cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        # Write header (start from row 4)
        header_row = set_row + 1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="27ae60", end_color="27ae60", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Write data
        for row_idx, row_data in analysis_df.iterrows():
            ws.cell(row=row_idx + header_row + 1, column=1, value=row_data["semester"])
            ws.cell(row=row_idx + header_row + 1, column=2, value=row_data["total_students"])
            ws.cell(row=row_idx + header_row + 1, column=3, value=row_data["passed_all"])
            ws.cell(row=row_idx + header_row + 1, column=4, value=row_data["carryover_students"])
            ws.cell(row=row_idx + header_row + 1, column=5, value=row_data["withdrawn_students"])
            ws.cell(row=row_idx + header_row + 1, column=6, value=row_data["average_gpa"])
            ws.cell(row=row_idx + header_row + 1, column=7, value=row_data["pass_rate"])

            # Apply borders to data rows
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row_idx + header_row + 1, column=col)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # IMPROVED AUTO-FIT COLUMN WIDTHS for all columns
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            column_letter = get_column_letter(col_idx)
            
            # Check all rows in this column including data rows
            for row_idx in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Set width with some padding
            adjusted_width = max_length + 3
            # Set reasonable limits
            if adjusted_width > 25:
                adjusted_width = 25
            if adjusted_width < 10:
                adjusted_width = 10
                
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(mastersheet_path)
        logger.info("✅ BM Analysis sheet created successfully with professional title")

        return analysis_df

    except Exception as e:
        logger.error(f"❌ Error creating BM analysis sheet: {e}")
        return None

# ----------------------------
# Upgrade Rule Functions
# ----------------------------


def get_upgrade_threshold_from_user(semester_key, set_name):
    """
    Prompt user to choose upgrade threshold for BM results.
    Returns: (min_threshold, upgraded_count) or (None, 0) if skipped
    """
    logger.info(f"\n🎯 MANAGEMENT THRESHOLD UPGRADE RULE DETECTED")
    logger.info(f"📚 Semester: {semester_key}")
    logger.info(f"📁 Set: {set_name}")
    logger.info(
        "\nSelect minimum score to upgrade (45-49). All scores >= selected value up to 49 will be upgraded to 50."
    )
    logger.info("Enter 0 to skip upgrade.")

    while True:
        try:
            choice = input("\nEnter your choice (0, 45, 46, 47, 48, 49): ").strip()

            if not choice:
                logger.error("❌ Please enter a value.")
                continue

            if choice == "0":
                logger.info("⏭️ Skipping upgrade for this semester.")
                return None, 0

            if choice in ["45", "46", "47", "48", "49"]:
                min_threshold = int(choice)
                logger.info(f"✅ Upgrade rule selected: {min_threshold}–49 → 50")
                return min_threshold, 0
            else:
                logger.error(
                    "❌ Invalid choice. Please enter 0, 45, 46, 47, 48, or 49."
                )

        except KeyboardInterrupt:
            logger.info("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            logger.error(f"❌ Error: {e}. Please try again.")


def apply_upgrade_rule(mastersheet, ordered_codes, min_threshold):
    """
    Apply upgrade rule to mastersheet scores.
    Returns: (updated_mastersheet, upgraded_count)
    """
    if min_threshold is None:
        return mastersheet, 0

    upgraded_count = 0
    upgraded_students = set()

    logger.info(f"🔄 Applying upgrade rule: {min_threshold}–49 → 50")

    for code in ordered_codes:
        for idx in mastersheet.index:
            score = mastersheet.at[idx, code]
            if isinstance(score, (int, float)) and min_threshold <= score <= 49:
                exam_no = mastersheet.at[idx, "EXAMS NUMBER"]
                original_score = score
                mastersheet.at[idx, code] = 50
                upgraded_count += 1
                upgraded_students.add(exam_no)

                # Log first few upgrades for visibility
                if upgraded_count <= 5:
                    logger.info(f"🔼 {exam_no} - {code}: {original_score} → 50")

    if upgraded_count > 0:
        logger.info(
            f"✅ Upgraded {upgraded_count} scores from {min_threshold}–49 to 50"
        )
        logger.info(f"📊 Affected {len(upgraded_students)} students")
    else:
        logger.info(f"ℹ️ No scores found in range {min_threshold}–49 to upgrade")

    return mastersheet, upgraded_count


# ----------------------------
# Utilities
# ----------------------------


def normalize_path(path: str) -> str:
    """Normalize user paths for Windows/WSL/Linux."""
    path = os.path.expanduser(path)
    path = os.path.normpath(path)
    if platform.system().lower() == "linux" and path.startswith("C:\\"):
        path = "/mnt/" + path[0].lower() + path[2:].replace("\\", "/")
    if platform.system().lower() == "linux" and path.startswith("c:\\"):
        path = "/mnt/" + path[0].lower() + path[2:].replace("\\", "/")
    if path.startswith("/c/") and os.path.exists("/mnt/c"):
        path = path.replace("/c/", "/mnt/c/", 1)
    return os.path.abspath(path)


def normalize_course_name(name):
    """Simple normalization for course title matching."""
    return re.sub(r"\s+", " ", str(name).strip().lower()).replace(
        "coomunication", "communication"
    )


def normalize_for_matching(s):
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"\b1st\b", "first", s)
    s = re.sub(r"\b2nd\b", "second", s)
    s = re.sub(r"\b3rd\b", "third", s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ----------------------------
# Student Tracking Functions
# ----------------------------


def initialize_student_tracker():
    """Initialize the global student tracker."""
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS
    STUDENT_TRACKER = {}
    WITHDRAWN_STUDENTS = {}


def update_student_tracker(semester_key, exam_numbers, withdrawn_students=None):
    """
    Update the student tracker with current semester's students.
    This helps track which students are present in each semester.
    """
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS

    logger.info(f"📊 Updating student tracker for {semester_key}")
    logger.info(f"📝 Current students in this semester: {len(exam_numbers)}")

    # Track withdrawn students
    if withdrawn_students:
        for exam_no in withdrawn_students:
            if exam_no not in WITHDRAWN_STUDENTS:
                WITHDRAWN_STUDENTS[exam_no] = {
                    "withdrawn_semester": semester_key,
                    "withdrawn_date": datetime.now().strftime(TIMESTAMP_FMT),
                    "reappeared_semesters": [],
                }
                logger.info(f"🚫 Marked as withdrawn: {exam_no} in {semester_key}")

    for exam_no in exam_numbers:
        if exam_no not in STUDENT_TRACKER:
            STUDENT_TRACKER[exam_no] = {
                "first_seen": semester_key,
                "last_seen": semester_key,
                "semesters_present": [semester_key],
                "status": "Active",
                "withdrawn": False,
                "withdrawn_semester": None,
            }
        else:
            STUDENT_TRACKER[exam_no]["last_seen"] = semester_key
            if semester_key not in STUDENT_TRACKER[exam_no]["semesters_present"]:
                STUDENT_TRACKER[exam_no]["semesters_present"].append(semester_key)

            # Check if student was previously withdrawn and has reappeared
            if STUDENT_TRACKER[exam_no]["withdrawn"]:
                logger.warning(f"⚠️ PREVIOUSLY WITHDRAWN STUDENT REAPPEARED: {exam_no}")
                if exam_no in WITHDRAWN_STUDENTS:
                    if (
                        semester_key
                        not in WITHDRAWN_STUDENTS[exam_no]["reappeared_semesters"]
                    ):
                        WITHDRAWN_STUDENTS[exam_no]["reappeared_semesters"].append(
                            semester_key
                        )

    logger.info(f"📈 Total unique students tracked: {len(STUDENT_TRACKER)}")
    logger.info(f"🚫 Total withdrawn students: {len(WITHDRAWN_STUDENTS)}")


def mark_student_withdrawn(exam_no, semester_key):
    """Mark a student as withdrawn in a specific semester."""
    global STUDENT_TRACKER, WITHDRAWN_STUDENTS

    if exam_no in STUDENT_TRACKER:
        STUDENT_TRACKER[exam_no]["withdrawn"] = True
        STUDENT_TRACKER[exam_no]["withdrawn_semester"] = semester_key
        STUDENT_TRACKER[exam_no]["status"] = "Withdrawn"

    if exam_no not in WITHDRAWN_STUDENTS:
        WITHDRAWN_STUDENTS[exam_no] = {
            "withdrawn_semester": semester_key,
            "withdrawn_date": datetime.now().strftime(TIMESTAMP_FMT),
            "reappeared_semesters": [],
        }


def is_student_withdrawn(exam_no):
    """Check if a student has been withdrawn in any previous semester."""
    return exam_no in WITHDRAWN_STUDENTS


def get_withdrawal_history(exam_no):
    """Get withdrawal history for a student."""
    if exam_no in WITHDRAWN_STUDENTS:
        return WITHDRAWN_STUDENTS[exam_no]
    return None


def filter_out_withdrawn_students(mastersheet, semester_key):
    """
    Filter out students who were withdrawn in previous semesters.
    Returns filtered mastersheet and list of removed students.
    """
    removed_students = []
    filtered_mastersheet = mastersheet.copy()

    for idx, row in mastersheet.iterrows():
        exam_no = str(row["EXAMS NUMBER"]).strip()
        if is_student_withdrawn(exam_no):
            withdrawal_history = get_withdrawal_history(exam_no)
            # Only remove if student was withdrawn in a PREVIOUS semester
            if (
                withdrawal_history
                and withdrawal_history["withdrawn_semester"] != semester_key
            ):
                removed_students.append(exam_no)
                filtered_mastersheet = filtered_mastersheet[
                    filtered_mastersheet["EXAMS NUMBER"] != exam_no
                ]

    if removed_students:
        logger.info(
            f"🚫 Removed {len(removed_students)} previously withdrawn students from {semester_key}:"
        )
        for exam_no in removed_students:
            withdrawal_history = get_withdrawal_history(exam_no)
            logger.info(
                f"   - {exam_no} (withdrawn in {withdrawal_history['withdrawn_semester']})"
            )

    return filtered_mastersheet, removed_students


# ----------------------------
# Carryover Management for BM
# ----------------------------
def initialize_carryover_tracker():
    """Initialize the global carryover tracker for BM."""
    global CARRYOVER_STUDENTS
    CARRYOVER_STUDENTS = {}

    # Load previous carryover records from all JSON files
    carryover_jsons = glob.glob(
        os.path.join(BM_BASE_DIR, "**/co_student*.json"), recursive=True
    )
    for jf in sorted(
        carryover_jsons, key=os.path.getmtime
    ):  # Load in chronological order, later files override
        try:
            with open(jf, "r") as f:
                data = json.load(f)
                for student in data:
                    student_key = f"{student['exam_number']}_{student['semester']}"
                    CARRYOVER_STUDENTS[student_key] = student
        except Exception as e:
            logger.warning(f"⚠️ Failed to load carryover from {jf}: {e}")

    logger.info(f"📂 Loaded {len(CARRYOVER_STUDENTS)} previous carryover records")


def identify_carryover_students(
    mastersheet_df, semester_key, set_name, pass_threshold=50.0
):
    """
    Identify BM students with carryover courses from current semester processing.
    """
    carryover_students = []

    # Get course columns (excluding student info columns)
    course_columns = [
        col
        for col in mastersheet_df.columns
        if col
        not in [
            "S/N",
            "EXAMS NUMBER",
            "NAME",
            "FAILED COURSES",
            "REMARKS",
            "CU Passed",
            "CU Failed",
            "TCPE",
            "GPA",
            "AVERAGE",
        ]
    ]

    for idx, student in mastersheet_df.iterrows():
        failed_courses = []
        exam_no = str(student["EXAMS NUMBER"])
        student_name = student["NAME"]

        for course in course_columns:
            score = student.get(course, 0)
            try:
                score_val = float(score) if pd.notna(score) else 0
                if score_val < pass_threshold:
                    failed_courses.append(
                        {
                            "course_code": course,
                            "original_score": score_val,
                            "semester": semester_key,
                            "set": set_name,
                            "resit_attempts": 0,
                            "best_score": score_val,
                            "status": "Failed",
                        }
                    )
            except (ValueError, TypeError):
                continue

        if failed_courses:
            carryover_data = {
                "exam_number": exam_no,
                "name": student_name,
                "failed_courses": failed_courses,
                "semester": semester_key,
                "set": set_name,
                "identified_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "total_resit_attempts": 0,
                "status": "Active",
            }
            carryover_students.append(carryover_data)

            # Update global tracker
            student_key = "{}_{}".format(exam_no, semester_key)
            CARRYOVER_STUDENTS[student_key] = carryover_data

    return carryover_students


def save_carryover_records(carryover_students, output_dir, set_name, semester_key):
    """
    Save BM carryover student records to the clean results folder.
    UPDATED: Use centralized CARRYOVER_RECORDS folder like BN
    """
    if not carryover_students:
        logger.info("ℹ️ No carryover students to save")
        return None

    # FIXED: Always navigate to CLEAN_RESULTS parent
    # Find CLEAN_RESULTS directory by walking up the path
    current_dir = output_dir
    while current_dir and not current_dir.endswith("CLEAN_RESULTS"):
        parent = os.path.dirname(current_dir)
        if parent == current_dir:  # Reached root
            current_dir = output_dir
            break
        current_dir = parent

    # If we're in a timestamped folder, go up one level
    if "_RESULT-" in os.path.basename(current_dir):
        clean_results_dir = os.path.dirname(current_dir)
    else:
        clean_results_dir = current_dir

    carryover_dir = os.path.join(clean_results_dir, "CARRYOVER_RECORDS")
    os.makedirs(carryover_dir, exist_ok=True)

    logger.info(f"📁 Saving carryover records to centralized location: {carryover_dir}")

    # Generate filename with set and semester tags
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "co_student_{}_{}_{}".format(set_name, semester_key, timestamp)

    # Save as Excel
    excel_file = os.path.join(carryover_dir, "{}.xlsx".format(filename))

    # Prepare data for Excel
    records_data = []
    for student in carryover_students:
        for course in student["failed_courses"]:
            record = {
                "EXAMS NUMBER": student["exam_number"],
                "NAME": student["name"],
                "COURSE CODE": course["course_code"],
                "ORIGINAL SCORE": course["original_score"],
                "SEMESTER": student["semester"],
                "SET": student["set"],
                "RESIT ATTEMPTS": course["resit_attempts"],
                "BEST SCORE": course["best_score"],
                "STATUS": course["status"],
                "IDENTIFIED DATE": student["identified_date"],
                "OVERALL_STATUS": student["status"],
            }
            records_data.append(record)

    if records_data:
        df = pd.DataFrame(records_data)
        df.to_excel(excel_file, index=False)
        logger.info("✅ BM Carryover records saved: {}".format(excel_file))

        # Add basic formatting
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            # Simple header formatting only
            header_row = ws[1]
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(excel_file)
            logger.info("✅ Added basic formatting to BM carryover Excel file")

        except Exception as e:
            logger.warning(
                "⚠️ Could not add basic formatting to BM carryover file: {}".format(e)
            )

    # Save as JSON for easy processing
    json_file = os.path.join(carryover_dir, "{}.json".format(filename))
    with open(json_file, "w") as f:
        json.dump(carryover_students, f, indent=2)

    # Save individual CSV reports
    individual_dir = os.path.join(carryover_dir, "INDIVIDUAL_REPORTS")
    os.makedirs(individual_dir, exist_ok=True)

    for student in carryover_students:
        student_records = [
            r for r in records_data if r["EXAMS NUMBER"] == student["exam_number"]
        ]
        if student_records:
            student_df = pd.DataFrame(student_records)
            student_filename = (
                f"carryover_report_{student['exam_number']}_{timestamp}.csv"
            )
            student_path = os.path.join(individual_dir, student_filename)
            student_df.to_csv(student_path, index=False)
            logger.info(f"✅ Saved individual carryover report: {student_path}")

    logger.info("📁 BM Carryover records saved in: {}".format(carryover_dir))
    return carryover_dir


# ----------------------------
# Set Selection Functions
# ----------------------------


def get_available_sets(base_dir):
    """Get all available BM sets (SET2023, SET2024, SET2025, etc.)"""
    # BM: Sets are under BM folder
    bm_dir = os.path.join(base_dir, "BM")
    if not os.path.exists(bm_dir):
        logger.error(f"❌ BM directory not found: {bm_dir}")
        return []

    sets = []
    for item in os.listdir(bm_dir):
        item_path = os.path.join(bm_dir, item)
        if os.path.isdir(item_path) and item.upper().startswith("SET"):
            sets.append(item)
    return sorted(sets)


def get_user_set_choice(available_sets):
    """
    Prompt user to choose which set to process.
    Returns the selected set directory name.
    """
    logger.info("\n🎯 AVAILABLE SETS:")
    for i, set_name in enumerate(available_sets, 1):
        logger.info(f"{i}. {set_name}")
    logger.info(f"{len(available_sets) + 1}. Process ALL sets")

    while True:
        try:
            choice = input(
                f"\nEnter your choice (1-{len(available_sets) + 1}): "
            ).strip()
            if not choice:
                logger.error("❌ Please enter a choice.")
                continue

            if choice.isdigit():
                choice_num = int(choice)
                if 1 <= choice_num <= len(available_sets):
                    selected_set = available_sets[choice_num - 1]
                    logger.info(f"✅ Selected set: {selected_set}")
                    return [selected_set]
                elif choice_num == len(available_sets) + 1:
                    logger.info("✅ Selected: ALL sets")
                    return available_sets
                else:
                    logger.error(
                        f"❌ Invalid choice. Please enter a number between 1-{len(available_sets) + 1}."
                    )
            else:
                logger.error("❌ Please enter a valid number.")

        except KeyboardInterrupt:
            logger.info("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            logger.error(f"❌ Error: {e}. Please try again.")


# ----------------------------
# Grade and GPA calculation
# ----------------------------


def get_grade(score):
    """Convert numeric score to letter grade - single letter only."""
    try:
        score = float(score)
        if score >= 70:
            return "A"
        elif score >= 60:
            return "B"
        elif score >= 50:
            return "C"
        elif score >= 45:
            return "D"
        elif score >= 40:
            return "E"
        else:
            return "F"
    except BaseException:
        return "F"


def get_grade_point(score):
    """Convert score to grade point for GPA calculation."""
    try:
        score = float(score)
        if score >= 70:
            return 5.0  # A
        elif score >= 60:
            return 4.0  # B
        elif score >= 50:
            return 3.0  # C
        elif score >= 45:
            return 2.0  # D
        elif score >= 40:
            return 1.0  # E
        else:
            return 0.0  # F
    except BaseException:
        return 0.0


# ----------------------------
# Load Course Data
# ----------------------------


def load_course_data():
    """
    Reads M-course-code-creditUnit.xlsx and returns:
      (semester_course_maps, semester_credit_units,
       semester_lookup, semester_course_titles)
    """
    # FIXED: Flexible course file detection
    possible_filenames = [
        "M-course-code-creditUnit.xlsx",
        "BM-course-code-creditUnit.xlsx",
        "m-course-code-creditUnit.xlsx",
        "bm-course-code-creditUnit.xlsx",
    ]

    course_file = None
    for filename in possible_filenames:
        test_path = os.path.join(BM_COURSES_DIR, filename)
        if os.path.exists(test_path):
            course_file = test_path
            break

    if not course_file:
        raise FileNotFoundError(
            f"BM course file not found. Tried: {possible_filenames}\n"
            f"In directory: {BM_COURSES_DIR}"
        )

    logger.info(f"Loading BM course data from: {course_file}")

    xl = pd.ExcelFile(course_file)
    semester_course_maps = {}
    semester_credit_units = {}
    semester_lookup = {}
    semester_course_titles = {}  # code -> title mapping

    for sheet in xl.sheet_names:
        df = pd.read_excel(course_file, sheet_name=sheet, engine="openpyxl", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        expected = ["COURSE CODE", "COURSE TITLE", "CU"]
        if not all(col in df.columns for col in expected):
            logger.warning(
                f"Warning: sheet '{sheet}' missing expected columns {expected} — skipped"
            )
            continue
        dfx = df.dropna(subset=["COURSE CODE", "COURSE TITLE"])
        dfx = dfx[
            ~dfx["COURSE CODE"].astype(str).str.contains("TOTAL", case=False, na=False)
        ]
        valid_mask = (
            dfx["CU"].astype(str).str.replace(".", "", regex=False).str.isdigit()
        )
        dfx = dfx[valid_mask]
        if dfx.empty:
            logger.warning(
                f"Warning: sheet '{sheet}' has no valid rows after cleaning — skipped"
            )
            continue
        codes = dfx["COURSE CODE"].astype(str).str.strip().tolist()
        titles = dfx["COURSE TITLE"].astype(str).str.strip().tolist()
        cus = dfx["CU"].astype(float).astype(int).tolist()

        semester_course_maps[sheet] = dict(zip(titles, codes))
        semester_credit_units[sheet] = dict(zip(codes, cus))
        semester_course_titles[sheet] = dict(zip(codes, titles))

        # Create multiple lookup variations for flexible matching
        norm = normalize_for_matching(sheet)
        semester_lookup[norm] = sheet

        # Add variations without "BM-" prefix
        norm_no_bm = norm.replace("bm-", "").replace("bm ", "")
        semester_lookup[norm_no_bm] = sheet

        # Add variations with different separators
        norm_hyphen = norm.replace("-", " ")
        semester_lookup[norm_hyphen] = sheet

        norm_space = norm.replace(" ", "-")
        semester_lookup[norm_space] = sheet

    if not semester_course_maps:
        raise ValueError("No course data loaded from BM course workbook")
    logger.info(f"Loaded BM course sheets: {list(semester_course_maps.keys())}")
    return (
        semester_course_maps,
        semester_credit_units,
        semester_lookup,
        semester_course_titles,
    )


# ----------------------------
# Helper functions
# ----------------------------


def detect_semester_from_filename(filename):
    """
    Detect semester from filename for BM program.
    Returns: (
    semester_key,
    year,
    semester_num,
    level_display,
    semester_display,
     set_code)
    """
    filename_upper = filename.upper()

    # Map filename patterns to actual BM course sheet names - FIXED: Changed BM- prefix to M- prefix consistently
    semester_patterns = {
        r"FIRST[-_\s]YEAR[-_\s]FIRST[-_\s]SEMESTER": "M-FIRST-YEAR-FIRST-SEMESTER",
        r"FIRST[-_\s]YEAR[-_\s]SECOND[-_\s]SEMESTER": "M-FIRST-YEAR-SECOND-SEMESTER",
        r"SECOND[-_\s]YEAR[-_\s]FIRST[-_\s]SEMESTER": "M-SECOND-YEAR-FIRST-SEMESTER",
        r"SECOND[-_\s]YEAR[-_\s]SECOND[-_\s]SEMESTER": "M-SECOND-YEAR-SECOND-SEMESTER",
        r"THIRD[-_\s]YEAR[-_\s]FIRST[-_\s]SEMESTER": "M-THIRD-YEAR-FIRST-SEMESTER",
        r"THIRD[-_\s]YEAR[-_\s]SECOND[-_\s]SEMESTER": "M-THIRD-YEAR-SECOND-SEMESTER",
    }

    for pattern, semester_key in semester_patterns.items():
        if re.search(pattern, filename_upper):
            logger.info(f"✅ Detected {semester_key} from: {filename}")
            # Get display info using the fixed function
            year, semester_num, level_display, semester_display, set_code = (
                get_semester_display_info(semester_key)
            )
            return (
                semester_key,
                year,
                semester_num,
                level_display,
                semester_display,
                set_code,
            )

    # If no match, raise error instead of defaulting
    raise ValueError(f"❌ Could not detect semester from filename: {filename}")


def get_semester_display_info(semester_key):
    """
    Get display information for a given BM semester key.
    Returns: (year, semester_num, level_display, semester_display, set_code)
    """
    semester_lower = semester_key.lower()

    # FIXED: Changed BM- prefix to M- prefix with correct set codes
    if "first-year-first-semester" in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BMI"
    elif "first-year-second-semester" in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "BMI"
    elif "second-year-first-semester" in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "BMII"
    elif "second-year-second-semester" in semester_lower:
        return 2, 2, "YEAR TWO", "SECOND SEMESTER", "BMII"
    elif "third-year-first-semester" in semester_lower:
        return 3, 1, "YEAR THREE", "FIRST SEMESTER", "BMIII"
    elif "third-year-second-semester" in semester_lower:
        return 3, 2, "YEAR THREE", "SECOND SEMESTER", "BMIII"
    elif "first" in semester_lower and "second" not in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "BMI"
    elif "second" in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "BMI"
    elif "third" in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "BMII"
    else:
        # Default with warning
        logger.warning(f"⚠️ Unknown semester: {semester_key}, using default")
        return (1, 1, "YEAR ONE", "FIRST SEMESTER", "BMI")


def match_semester_from_filename(fname, semester_lookup):
    """Match semester using the lookup table with flexible matching."""
    fn = normalize_for_matching(fname)

    # Try exact matches first
    for norm, sheet in semester_lookup.items():
        if norm in fn:
            return sheet

    # Try close matches
    keys = list(semester_lookup.keys())
    best = difflib.get_close_matches(fn, keys, n=1, cutoff=0.55)
    if best:
        return semester_lookup[best[0]]

    # Fallback to filename-based detection
    sem, _, _, _, _, _ = detect_semester_from_filename(fname)
    return sem


def find_column_by_names(df, candidate_names):
    norm_map = {
        col: re.sub(r"\s+", " ", str(col).strip().lower()) for col in df.columns
    }
    candidates = [re.sub(r"\s+", " ", c.strip().lower()) for c in candidate_names]
    for cand in candidates:
        for col, ncol in norm_map.items():
            if ncol == cand:
                return col
    return None


# ----------------------------
# FIXED: Previous GPA Loading Function for BM
# ----------------------------


def load_previous_gpas_from_processed_files(
    output_dir, current_semester_key, timestamp
):
    """
    Load previous GPA data - FIXED VERSION for BM.
    Returns dict: {exam_number: previous_gpa}
    """
    previous_gpas = {}

    logger.info(f"\n🔍 LOADING PREVIOUS GPA for: {current_semester_key}")

    current_year, current_semester_num, _, _, _ = get_semester_display_info(
        current_semester_key
    )

    # Map for BM semesters
    semester_sequence = {
        (1, 1): None,
        (1, 2): "M-FIRST-YEAR-FIRST-SEMESTER",
        (2, 1): "M-FIRST-YEAR-SECOND-SEMESTER",
        (2, 2): "M-SECOND-YEAR-FIRST-SEMESTER",
        (3, 1): "M-SECOND-YEAR-SECOND-SEMESTER",
        (3, 2): "M-THIRD-YEAR-FIRST-SEMESTER",
    }

    prev_semester = semester_sequence.get((current_year, current_semester_num))

    if not prev_semester:
        logger.info("📊 First semester - no previous GPA available")
        return previous_gpas

    logger.info(f"🔍 Looking for previous GPA data from: {prev_semester}")

    # FIXED: Direct path to mastersheet
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")

    if not os.path.exists(mastersheet_path):
        logger.error(f"❌ Mastersheet not found: {mastersheet_path}")
        return previous_gpas

    try:
        df = pd.read_excel(mastersheet_path, sheet_name=prev_semester, header=5)
        logger.info(f"📋 Columns in {prev_semester}: {df.columns.tolist()}")

        exam_col = None
        gpa_col = None

        for col in df.columns:
            col_str = str(col).upper().strip()
            if "EXAM" in col_str or "REG" in col_str or "NUMBER" in col_str:
                exam_col = col
            elif "GPA" in col_str:
                gpa_col = col
                break

        if exam_col and gpa_col:
            logger.info(f"✅ Found columns: '{exam_col}', '{gpa_col}'")

            for idx, row in df.iterrows():
                exam_no = str(row[exam_col]).strip()
                gpa = row[gpa_col]

                if (
                    pd.notna(gpa)
                    and pd.notna(exam_no)
                    and exam_no != "nan"
                    and exam_no != ""
                ):
                    try:
                        previous_gpas[exam_no] = float(gpa)
                    except (ValueError, TypeError):
                        continue

            logger.info(f"✅ Loaded {len(previous_gpas)} previous GPAs")
        else:
            logger.error(f"❌ Could not find required columns")

    except Exception as e:
        logger.warning(f"⚠️ Could not read mastersheet: {str(e)}")

    return previous_gpas


def load_all_previous_gpas_for_cgpa(output_dir, current_semester_key, timestamp):
    """
    Load ALL previous GPAs from all completed semesters for CGPA calculation.
    Returns dict: {exam_number: {'gpas': [gpa1, gpa2, ...], 'credits': [credits1, credits2, ...]}}
    """
    logger.info(
        f"\n🔍 LOADING ALL PREVIOUS GPAs for CGPA calculation: {current_semester_key}"
    )

    current_year, current_semester_num, _, _, _ = get_semester_display_info(
        current_semester_key
    )

    # Determine which semesters to load based on current semester
    semesters_to_load = []

    if current_semester_num == 1 and current_year == 1:
        # First semester - no previous data
        return {}
    elif current_semester_num == 2 and current_year == 1:
        # Second semester of first year - load first semester
        semesters_to_load = [
            "M-FIRST-YEAR-FIRST-SEMESTER"
        ]  # FIXED: Changed BM- prefix to M- prefix
    elif current_semester_num == 1 and current_year == 2:
        # First semester of second year - load both first year semesters
        semesters_to_load = [
            "M-FIRST-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-FIRST-YEAR-SECOND-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
        ]
    elif current_semester_num == 2 and current_year == 2:
        # Second semester of second year - load all previous semesters
        semesters_to_load = [
            "M-FIRST-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-FIRST-YEAR-SECOND-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-SECOND-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
        ]
    elif current_semester_num == 1 and current_year == 3:
        # First semester of third year - load all first and second year semesters
        semesters_to_load = [
            "M-FIRST-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-FIRST-YEAR-SECOND-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-SECOND-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-SECOND-YEAR-SECOND-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
        ]
    elif current_semester_num == 2 and current_year == 3:
        # Second semester of third year - load all previous semesters
        semesters_to_load = [
            "M-FIRST-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-FIRST-YEAR-SECOND-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-SECOND-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-SECOND-YEAR-SECOND-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
            "M-THIRD-YEAR-FIRST-SEMESTER",  # FIXED: Changed BM- prefix to M- prefix
        ]

    logger.info(f"📚 Semesters to load for CGPA: {semesters_to_load}")

    all_student_data = {}
    mastersheet_path = os.path.join(output_dir, f"mastersheet_{timestamp}.xlsx")

    if not os.path.exists(mastersheet_path):
        logger.error(f"❌ Mastersheet not found: {mastersheet_path}")
        return {}

    for semester in semesters_to_load:
        logger.info(f"📖 Loading data from: {semester}")
        try:
            # Load the semester data, skipping header rows
            df = pd.read_excel(mastersheet_path, sheet_name=semester, header=5)

            # Find columns
            exam_col = None
            gpa_col = None
            credit_col = None

            for col in df.columns:
                col_str = str(col).upper().strip()
                if "EXAM" in col_str or "REG" in col_str or "NUMBER" in col_str:
                    exam_col = col
                elif "GPA" in col_str:
                    gpa_col = col
                elif "CU PASSED" in col_str or "CREDIT" in col_str:
                    credit_col = col

            if exam_col and gpa_col:
                for idx, row in df.iterrows():
                    exam_no = str(row[exam_col]).strip()
                    gpa = row[gpa_col]

                    if (
                        pd.notna(gpa)
                        and pd.notna(exam_no)
                        and exam_no != "nan"
                        and exam_no != ""
                    ):
                        try:
                            # Get credits completed (use CU Passed if
                            # available, otherwise estimate)
                            credits_completed = 0
                            if credit_col and pd.notna(row[credit_col]):
                                credits_completed = int(row[credit_col])
                            else:
                                # Estimate credits based on typical semester
                                # load for BM
                                if "FIRST-YEAR-FIRST-SEMESTER" in semester:
                                    credits_completed = (
                                        30  # Typical first semester credits
                                    )
                                elif "FIRST-YEAR-SECOND-SEMESTER" in semester:
                                    credits_completed = (
                                        30  # Typical second semester credits
                                    )
                                elif "SECOND-YEAR-FIRST-SEMESTER" in semester:
                                    credits_completed = (
                                        30  # Typical third semester credits
                                    )
                                elif "SECOND-YEAR-SECOND-SEMESTER" in semester:
                                    credits_completed = (
                                        30  # Typical fourth semester credits
                                    )
                                elif "THIRD-YEAR-FIRST-SEMESTER" in semester:
                                    credits_completed = (
                                        30  # Typical fifth semester credits
                                    )

                            if exam_no not in all_student_data:
                                all_student_data[exam_no] = {"gpas": [], "credits": []}

                            all_student_data[exam_no]["gpas"].append(float(gpa))
                            all_student_data[exam_no]["credits"].append(
                                credits_completed
                            )

                        except (ValueError, TypeError):
                            continue

        except Exception as e:
            logger.warning(f"⚠️ Could not load data from {semester}: {str(e)}")

    logger.info(f"📊 Loaded cumulative data for {len(all_student_data)} students")
    return all_student_data


def calculate_cgpa(student_data, current_gpa, current_credits):
    """
    Calculate Cumulative GPA (CGPA) based on all previous semesters and current semester.
    """
    if not student_data:
        return current_gpa

    total_grade_points = 0.0
    total_credits = 0

    # Add previous semesters
    for prev_gpa, prev_credits in zip(student_data["gpas"], student_data["credits"]):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits

    # Add current semester
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits

    if total_credits > 0:
        return round(total_grade_points / total_credits, 2)
    else:
        return current_gpa


def get_cumulative_gpa(current_gpa, previous_gpa, current_credits, previous_credits):
    """
    Calculate cumulative GPA based on current and previous semester performance.
    """
    if previous_gpa is None:
        return current_gpa

    # For simplicity, we'll assume equal credit weights if not provided
    if current_credits is None or previous_credits is None:
        return round((current_gpa + previous_gpa) / 2, 2)

    total_points = (current_gpa * current_credits) + (previous_gpa * previous_credits)
    total_credits = current_credits + previous_credits
    return round(total_points / total_credits, 2) if total_credits > 0 else 0.0


def determine_student_status(row, total_cu, pass_threshold):
    """
    Determine student status based on performance metrics following NBTE standard.
    FIXED: Correct NBTE logic for Probation/Withdrawn determination
    """
    gpa = row.get("GPA", 0)
    cu_passed = row.get("CU Passed", 0)
    cu_failed = row.get("CU Failed", 0)

    # Calculate percentage of failed credit units
    failed_percentage = (cu_failed / total_cu) * 100 if total_cu > 0 else 0

    # FIXED: CORRECT NBTE STANDARD LOGIC
    # 1. No failed courses = Pass
    if cu_failed == 0:
        return "Pass"
    
    # 2. GPA >= 2.0 AND failed_percentage <= 45% = Carry Over (Resit)
    elif gpa >= 2.0 and failed_percentage <= 45:
        return "Carry Over"
    
    # 3. GPA < 2.0 AND failed_percentage <= 45% = Probation  
    elif gpa < 2.0 and failed_percentage <= 45:
        return "Probation"
    
    # 4. failed_percentage > 45% = Withdrawn (regardless of GPA)
    elif failed_percentage > 45:
        return "Withdrawn"
    
    # Fallback
    else:
        return "Carry Over"


def format_failed_courses_remark(failed_courses, max_line_length=60):
    """
    Format failed courses remark with line breaks for long lists.
    Returns list of formatted lines.
    """
    if not failed_courses:
        return [""]

    failed_str = ", ".join(sorted(failed_courses))

    # If the string is short enough, return as single line
    if len(failed_str) <= max_line_length:
        return [failed_str]

    # Split into multiple lines
    lines = []
    current_line = ""

    for course in sorted(failed_courses):
        if not current_line:
            current_line = course
        elif len(current_line) + len(course) + 2 <= max_line_length:  # +2 for ", "
            current_line += ", " + course
        else:
            lines.append(current_line)
            current_line = course

    if current_line:
        lines.append(current_line)

    return lines


def get_user_semester_choice():
    """
    Prompt user to choose which semesters to process.
    Returns list of semester keys to process.
    """
    logger.info("\n🎯 SEMESTER PROCESSING OPTIONS:")
    logger.info("1. Process ALL semesters in order")
    logger.info("2. Process FIRST YEAR - FIRST SEMESTER only")
    logger.info("3. Process FIRST YEAR - SECOND SEMESTER only")
    logger.info("4. Process SECOND YEAR - FIRST SEMESTER only")
    logger.info("5. Process SECOND YEAR - SECOND SEMESTER only")
    logger.info("6. Process THIRD YEAR - FIRST SEMESTER only")
    logger.info("7. Process THIRD YEAR - SECOND SEMESTER only")
    logger.info("8. Custom selection")

    while True:
        try:
            choice = input("\nEnter your choice (1-8): ").strip()
            if choice == "1":
                return SEMESTER_ORDER.copy()
            elif choice == "2":
                return [
                    "M-FIRST-YEAR-FIRST-SEMESTER"
                ]  # FIXED: Changed BM- prefix to M- prefix
            elif choice == "3":
                return [
                    "M-FIRST-YEAR-SECOND-SEMESTER"
                ]  # FIXED: Changed BM- prefix to M- prefix
            elif choice == "4":
                return [
                    "M-SECOND-YEAR-FIRST-SEMESTER"
                ]  # FIXED: Changed BM- prefix to M- prefix
            elif choice == "5":
                return [
                    "M-SECOND-YEAR-SECOND-SEMESTER"
                ]  # FIXED: Changed BM- prefix to M- prefix
            elif choice == "6":
                return [
                    "M-THIRD-YEAR-FIRST-SEMESTER"
                ]  # FIXED: Changed BM- prefix to M- prefix
            elif choice == "7":
                return [
                    "M-THIRD-YEAR-SECOND-SEMESTER"
                ]  # FIXED: Changed BM- prefix to M- prefix
            elif choice == "8":
                return get_custom_semester_selection()
            else:
                logger.error("❌ Invalid choice. Please enter a number between 1-8.")
        except KeyboardInterrupt:
            logger.info("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            logger.error(f"❌ Error: {e}. Please try again.")


def get_custom_semester_selection():
    """
    Allow user to select multiple semesters for processing.
    """
    logger.info("\n📚 AVAILABLE SEMESTERS:")
    for i, semester in enumerate(SEMESTER_ORDER, 1):
        year, sem_num, level, sem_display, set_code = get_semester_display_info(
            semester
        )
        logger.info(f"{i}. {level} - {sem_display}")

    logger.info(f"{len(SEMESTER_ORDER) + 1}. Select all")

    selected = []
    while True:
        try:
            choices = input(
                f"\nEnter semester numbers separated by commas (1-{len(SEMESTER_ORDER) + 1}): "
            ).strip()
            if not choices:
                logger.error("❌ Please enter at least one semester number.")
                continue

            choice_list = [c.strip() for c in choices.split(",")]

            # Check for "select all" option
            if str(len(SEMESTER_ORDER) + 1) in choice_list:
                return SEMESTER_ORDER.copy()

            # Validate and convert choices
            valid_choices = []
            for choice in choice_list:
                if not choice.isdigit():
                    logger.error(f"❌ '{choice}' is not a valid number.")
                    continue

                choice_num = int(choice)
                if 1 <= choice_num <= len(SEMESTER_ORDER):
                    valid_choices.append(choice_num)
                else:
                    logger.error(f"❌ '{choice}' is not a valid semester number.")

            if valid_choices:
                selected_semesters = [SEMESTER_ORDER[i - 1] for i in valid_choices]
                logger.info(
                    f"✅ Selected semesters: {[get_semester_display_info(sem)[3] for sem in selected_semesters]}"
                )
                return selected_semesters
            else:
                logger.error("❌ No valid semesters selected. Please try again.")

        except KeyboardInterrupt:
            logger.info("\n👋 Operation cancelled by user.")
            sys.exit(0)
        except Exception as e:
            logger.error(f"❌ Error: {e}. Please try again.")


# ----------------------------
# File Validation
# ----------------------------


def validate_raw_file(file_path):
    """Validate raw Excel file before processing."""
    if not os.path.exists(file_path):
        return False, "File does not exist"

    if os.path.getsize(file_path) == 0:
        return False, "File is empty"

    try:
        xl = pd.ExcelFile(file_path)
        if not xl.sheet_names:
            return False, "No sheets in Excel file"

        expected_sheets = ["CA", "OBJ", "EXAM"]
        found_sheets = [s for s in expected_sheets if s in xl.sheet_names]

        if not found_sheets:
            return False, f"No expected sheets found. Has: {xl.sheet_names}"

        return True, f"Valid file with sheets: {found_sheets}"

    except Exception as e:
        return False, f"Cannot open file: {e}"


# ----------------------------
# PDF Generation - Individual Student Report
# ----------------------------


def generate_individual_student_pdf(
    mastersheet_df,
    out_pdf_path,
    semester_key,
    logo_path=None,
    prev_mastersheet_df=None,
    filtered_credit_units=None,
    ordered_codes=None,
    course_titles_map=None,
    previous_gpas=None,
    cgpa_data=None,
    total_cu=None,
    pass_threshold=None,
    upgrade_min_threshold=None,
):
    """
    Create a PDF with one page per student matching the sample format exactly.
    """
    doc = SimpleDocTemplate(
        out_pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=20,
        bottomMargin=20,
    )
    styles = getSampleStyleSheet()

    # Custom styles
    header_style = ParagraphStyle(
        "CustomHeader",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=2,
    )

    main_header_style = ParagraphStyle(
        "MainHeader",
        parent=styles["Normal"],
        fontSize=16,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceAfter=6,
        textColor=colors.HexColor("#800080"),
    )

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Normal"],
        fontSize=12,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceAfter=4,
    )

    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=10,
        textColor=colors.red,
    )

    # Left alignment style for course code and title
    left_align_style = ParagraphStyle(
        "LeftAlign",
        parent=styles["Normal"],
        fontSize=9,
        alignment=TA_LEFT,
        leftIndent=4,
    )

    center_align_style = ParagraphStyle(
        "CenterAlign", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER
    )

    # Style for remarks with smaller font
    remarks_style = ParagraphStyle(
        "RemarksStyle", parent=styles["Normal"], fontSize=8, alignment=TA_LEFT
    )

    elems = []

    for idx, r in mastersheet_df.iterrows():
        # Logo and header
        logo_img = None
        if logo_path and os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path, width=0.8 * inch, height=0.8 * inch)
            except Exception as e:
                logger.warning(f"Warning: Could not load logo: {e}")

        # Header table with logo and title
        if logo_img:
            header_data = [
                [
                    logo_img,
                    Paragraph("FCT COLLEGE OF NURSING SCIENCES", main_header_style),
                ]
            ]
            header_table = Table(header_data, colWidths=[1.0 * inch, 5.0 * inch])
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ]
                )
            )
            elems.append(header_table)
        else:
            elems.append(
                Paragraph("FCT COLLEGE OF NURSING SCIENCES", main_header_style)
            )

        # Address and contact info
        elems.append(Paragraph("P.O.Box 507, Gwagwalada-Abuja, Nigeria", header_style))
        elems.append(Paragraph("<b>EXAMINATIONS OFFICE</b>", header_style))
        elems.append(Paragraph("fctsonexamsoffice@gmail.com", header_style))

        elems.append(Spacer(1, 8))
        elems.append(Paragraph("STUDENT'S ACADEMIC PROGRESS REPORT", title_style))
        elems.append(Paragraph("(THIS IS NOT A TRANSCRIPT)", subtitle_style))

        elems.append(Spacer(1, 8))

        # Student particulars - SEPARATE FROM PASSPORT PHOTO
        exam_no = str(r.get("EXAMS NUMBER", r.get("REG. No", "")))
        student_name = str(r.get("NAME", ""))

        # Determine level and semester using the new function
        year, semester_num, level_display, semester_display, set_code = (
            get_semester_display_info(semester_key)
        )

        # Create two tables: one for student particulars, one for passport
        # photo
        particulars_data = [
            [Paragraph("<b>STUDENT'S PARTICULARS</b>", styles["Normal"])],
            [Paragraph("<b>NAME:</b>", styles["Normal"]), student_name],
            [
                Paragraph("<b>LEVEL OF<br/>STUDY:</b>", styles["Normal"]),
                level_display,
                Paragraph("<b>SEMESTER:</b>", styles["Normal"]),
                semester_display,
            ],
            [
                Paragraph("<b>REG NO.</b>", styles["Normal"]),
                exam_no,
                Paragraph("<b>SET:</b>", styles["Normal"]),
                set_code,
            ],
        ]

        particulars_table = Table(
            particulars_data, colWidths=[1.2 * inch, 2.3 * inch, 0.8 * inch, 1.5 * inch]
        )
        particulars_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (3, 0)),
                    ("SPAN", (1, 1), (3, 1)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )

        # Passport photo table (separate box)
        passport_data = [
            [Paragraph("Affix Recent<br/>Passport<br/>Photograph", styles["Normal"])]
        ]

        passport_table = Table(
            passport_data, colWidths=[1.5 * inch], rowHeights=[1.2 * inch]
        )
        passport_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )

        # Create a combined table with particulars and passport side by side
        combined_data = [[particulars_table, passport_table]]

        combined_table = Table(combined_data, colWidths=[5.8 * inch, 1.5 * inch])
        combined_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        elems.append(combined_table)
        elems.append(Spacer(1, 12))

        # Semester result header
        elems.append(Paragraph("<b>SEMESTER RESULT</b>", title_style))
        elems.append(Spacer(1, 6))

        # Course results table - LEFT-ALIGNED CODE AND TITLE
        course_data = [
            [
                Paragraph("<b>S/N</b>", styles["Normal"]),
                Paragraph("<b>CODE</b>", styles["Normal"]),
                Paragraph("<b>COURSE TITLE</b>", styles["Normal"]),
                Paragraph("<b>UNITS</b>", styles["Normal"]),
                Paragraph("<b>SCORE</b>", styles["Normal"]),
                Paragraph("<b>GRADE</b>", styles["Normal"]),
            ]
        ]

        sn = 1
        total_grade_points = 0.0
        total_units = 0
        total_units_passed = 0
        total_units_failed = 0
        failed_courses_list = []

        for code in ordered_codes if ordered_codes else []:
            score = r.get(code)
            if pd.isna(score) or score == "":
                continue

            try:
                score_val = float(score)

                # -----------------------
                # FIX: Auto-upgrade borderline scores when threshold upgrade applies
                # This ensures PDF shows the same upgraded scores as Excel
                if (
                    upgrade_min_threshold is not None
                    and upgrade_min_threshold <= score_val <= 49
                ):
                    # Use the upgraded score for PDF display
                    score_val = 50.0
                    score_display = "50"
                    logger.info(f"🔼 PDF: Upgraded score for {exam_no} - {code}: → 50")
                else:
                    score_display = str(int(round(score_val)))
                # -----------------------

                grade = get_grade(score_val)
                grade_point = get_grade_point(score_val)
            except Exception:
                score_display = str(score)
                grade = "F"
                grade_point = 0.0

            cu = filtered_credit_units.get(code, 0) if filtered_credit_units else 0
            course_title = (
                course_titles_map.get(code, code) if course_titles_map else code
            )

            total_grade_points += grade_point * cu
            total_units += cu

            if score_val >= pass_threshold:
                total_units_passed += cu
            else:
                total_units_failed += cu
                failed_courses_list.append(code)

            course_data.append(
                [
                    Paragraph(str(sn), center_align_style),
                    Paragraph(code, left_align_style),
                    Paragraph(course_title, left_align_style),
                    Paragraph(str(cu), center_align_style),
                    Paragraph(score_display, center_align_style),
                    Paragraph(grade, center_align_style),
                ]
            )
            sn += 1

        course_table = Table(
            course_data,
            colWidths=[
                0.4 * inch,
                0.7 * inch,
                2.8 * inch,
                0.6 * inch,
                0.6 * inch,
                0.6 * inch,
            ],
        )
        course_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("ALIGN", (1, 1), (2, -1), "LEFT"),
                ]
            )
        )
        elems.append(course_table)
        elems.append(Spacer(1, 14))

        # Calculate current semester GPA
        current_gpa = (
            round(total_grade_points / total_units, 2) if total_units > 0 else 0.0
        )

        # Get previous GPA if available
        exam_no = str(r.get("EXAMS NUMBER", "")).strip()
        previous_gpa = previous_gpas.get(exam_no, None) if previous_gpas else None

        # Calculate CGPA if available
        cgpa = None
        if cgpa_data and exam_no in cgpa_data:
            cgpa = calculate_cgpa(cgpa_data[exam_no], current_gpa, total_units_passed)

        logger.info(f"📊 PDF GENERATION for {exam_no}:")
        logger.info(f"   Current GPA: {current_gpa}")
        logger.info(f"   Previous GPA available: {previous_gpa is not None}")
        logger.info(f"   CGPA available: {cgpa is not None}")
        if previous_gpa is not None:
            logger.info(f"   Previous GPA value: {previous_gpa}")
        if cgpa is not None:
            logger.info(f"   CGPA value: {cgpa}")

        # Get values from dataframe
        tcpe = round(total_grade_points, 1)
        tcup = total_units_passed
        tcuf = total_units_failed

        # Determine student status based on performance using FIXED NBTE logic
        student_status = determine_student_status(r, total_cu, pass_threshold)

        # Check if student was previously withdrawn
        withdrawal_history = get_withdrawal_history(exam_no)
        previously_withdrawn = withdrawal_history is not None

        # Format failed courses with line breaks if needed
        failed_courses_formatted = format_failed_courses_remark(failed_courses_list)

        # Combine course-specific remarks with overall status
        final_remarks_lines = []

        # For withdrawn students in their withdrawal semester, show appropriate
        # remarks
        if (
            previously_withdrawn
            and withdrawal_history["withdrawn_semester"] == semester_key
        ):
            # This is the actual withdrawal semester - show normal withdrawal
            # remarks
            if failed_courses_formatted:
                final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                if len(failed_courses_formatted) > 1:
                    final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("Advised to Withdraw")
            else:
                final_remarks_lines.append("Advised to Withdraw")
        elif previously_withdrawn:
            # Student was withdrawn in a previous semester but appears here -
            # this shouldn't happen due to filtering
            withdrawn_semester = withdrawal_history["withdrawn_semester"]
            year, sem_num, level, sem_display, set_code = get_semester_display_info(
                withdrawn_semester
            )
            final_remarks_lines.append(
                f"STUDENT WAS WITHDRAWN FROM {level} - {sem_display}"
            )
            final_remarks_lines.append(
                "This result should not be processed as student was previously withdrawn"
            )
        elif student_status == "Pass":
            final_remarks_lines.append("Passed")
        elif student_status == "Carry Over":
            if failed_courses_formatted:
                final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                if len(failed_courses_formatted) > 1:
                    final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("To Carry Over Courses")
            else:
                final_remarks_lines.append("To Carry Over Courses")
        elif student_status == "Probation":
            if failed_courses_formatted:
                final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                if len(failed_courses_formatted) > 1:
                    final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("Placed on Probation")
            else:
                final_remarks_lines.append("Placed on Probation")
        elif student_status == "Withdrawn":
            if failed_courses_formatted:
                final_remarks_lines.append(f"Failed: {failed_courses_formatted[0]}")
                if len(failed_courses_formatted) > 1:
                    final_remarks_lines.extend(failed_courses_formatted[1:])
                final_remarks_lines.append("Advised to Withdraw")
            else:
                final_remarks_lines.append("Advised to Withdraw")
        else:
            final_remarks_lines.append(str(r.get("REMARKS", "")))

        final_remarks = "<br/>".join(final_remarks_lines)
        display_gpa = current_gpa
        display_cgpa = cgpa if cgpa is not None else current_gpa

        # Summary section - EXPANDED TO ACCOMMODATE LONG REMARKS
        summary_data = [
            [Paragraph("<b>SUMMARY</b>", styles["Normal"]), "", "", ""],
            [
                Paragraph("<b>TCPE:</b>", styles["Normal"]),
                str(tcpe),
                Paragraph("<b>CURRENT GPA:</b>", styles["Normal"]),
                str(display_gpa),
            ],
        ]

        # Add previous GPA if available (from first year second semester
        # upward)
        if previous_gpa is not None:
            logger.info(f"✅ ADDING PREVIOUS GPA to PDF: {previous_gpa}")
            summary_data.append(
                [
                    Paragraph("<b>TCUP:</b>", styles["Normal"]),
                    str(tcup),
                    Paragraph("<b>PREVIOUS GPA:</b>", styles["Normal"]),
                    str(previous_gpa),
                ]
            )
        else:
            summary_data.append(
                [Paragraph("<b>TCUP:</b>", styles["Normal"]), str(tcup), "", ""]
            )

        # Add CGPA if available (from second semester onward)
        if cgpa is not None:
            logger.info(f"✅ ADDING CGPA to PDF: {cgpa}")
            summary_data.append(
                [
                    Paragraph("<b>TCUF:</b>", styles["Normal"]),
                    str(tcuf),
                    Paragraph("<b>OVERALL GPA:</b>", styles["Normal"]),
                    str(display_cgpa),
                ]
            )
        else:
            summary_data.append(
                [Paragraph("<b>TCUF:</b>", styles["Normal"]), str(tcuf), "", ""]
            )

        # Add remarks with multiple lines if needed
        remarks_paragraph = Paragraph(final_remarks, remarks_style)
        summary_data.append(
            [Paragraph("<b>REMARKS:</b>", styles["Normal"]), remarks_paragraph, "", ""]
        )

        # Calculate row heights based on content
        row_heights = [0.3 * inch] * len(summary_data)  # Default height

        # Adjust height for remarks row based on number of lines
        total_remark_lines = len(final_remarks_lines)
        if total_remark_lines > 1:
            # Add extra height for multiple lines
            row_heights[-1] = max(0.4 * inch, 0.2 * inch * (total_remark_lines + 1))

        summary_table = Table(
            summary_data,
            colWidths=[1.5 * inch, 1.0 * inch, 1.5 * inch, 1.0 * inch],
            rowHeights=row_heights,
        )
        summary_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (3, 0)),
                    ("SPAN", (1, len(summary_data) - 1), (3, len(summary_data) - 1)),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#E0E0E0")),
                    ("ALIGN", (0, 0), (3, 0), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        elems.append(summary_table)
        elems.append(Spacer(1, 25))

        # Signature section
        sig_data = [
            ["", ""],
            ["____________________", "____________________"],
            [
                Paragraph(
                    "<b>EXAMS SECRETARY</b>",
                    ParagraphStyle(
                        "SigStyle",
                        parent=styles["Normal"],
                        fontSize=10,
                        alignment=TA_CENTER,
                    ),
                ),
                Paragraph(
                    "<b>V.P. ACADEMICS</b>",
                    ParagraphStyle(
                        "SigStyle",
                        parent=styles["Normal"],
                        fontSize=10,
                        alignment=TA_CENTER,
                    ),
                ),
            ],
        ]

        sig_table = Table(sig_data, colWidths=[3.0 * inch, 3.0 * inch])
        sig_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elems.append(sig_table)

        # Page break for next student
        if idx < len(mastersheet_df) - 1:
            elems.append(PageBreak())

    doc.build(elems)
    logger.info(f"✅ Individual student PDF written: {out_pdf_path}")


# ----------------------------
# Main file processing
# ----------------------------


def process_semester_files(
    semester_key,
    raw_files,
    raw_dir,
    clean_dir,
    ts,
    pass_threshold,
    semester_course_maps,
    semester_credit_units,
    semester_lookup,
    semester_course_titles,
    logo_path,
    set_name,
    previous_gpas=None,
    upgrade_min_threshold=None,
):
    """
    Process all files for a specific semester.
    """
    logger.info(f"\n{'='*60}")
    logger.info(f"PROCESSING SEMESTER: {semester_key}")
    logger.info(f"{'='*60}")

    # Filter files for this semester
    semester_files = []
    for rf in raw_files:
        detected_sem, _, _, _, _, _ = detect_semester_from_filename(rf)
        if detected_sem == semester_key:
            semester_files.append(rf)

    if not semester_files:
        logger.warning(f"⚠️ No files found for semester {semester_key}")
        return None

    logger.info(
        f"📁 Found {len(semester_files)} files for {semester_key}: {semester_files}"
    )

    # Process each file for this semester
    files_processed = 0
    for rf in semester_files:
        raw_path = os.path.join(raw_dir, rf)
        logger.info(f"\n📄 Processing: {rf}")

        # Validate file before processing
        is_valid, validation_msg = validate_raw_file(raw_path)
        if not is_valid:
            logger.error(f"❌ File validation failed: {validation_msg}")
            continue

        try:
            # Load previous GPAs for this specific semester
            current_previous_gpas = (
                load_previous_gpas_from_processed_files(clean_dir, semester_key, ts)
                if previous_gpas is None
                else previous_gpas
            )

            # Load CGPA data (all previous semesters)
            cgpa_data = load_all_previous_gpas_for_cgpa(clean_dir, semester_key, ts)

            # Process the file
            result = process_single_file(
                raw_path,
                clean_dir,
                ts,
                pass_threshold,
                semester_course_maps,
                semester_credit_units,
                semester_lookup,
                semester_course_titles,
                logo_path,
                semester_key,
                set_name,
                current_previous_gpas,
                cgpa_data,
                upgrade_min_threshold,
            )

            if result is not None:
                logger.info(f"✅ Successfully processed {rf}")
                files_processed += 1
            else:
                logger.error(f"❌ Failed to process {rf}")

        except Exception as e:
            logger.error(f"❌ Error processing {rf}: {e}")
            traceback.print_exc()

    # NEW: Create CGPA and Analysis sheets after processing files
    if files_processed > 0:
        try:
            mastersheet_path = os.path.join(
                clean_dir, f"{set_name}_RESULT-{ts}", f"mastersheet_{ts}.xlsx"
            )
            if os.path.exists(mastersheet_path):
                create_bm_cgpa_summary_sheet(mastersheet_path, ts, set_name)
                create_bm_analysis_sheet(mastersheet_path, ts, set_name)
                logger.info("✅ Created CGPA and Analysis sheets for BM")
        except Exception as e:
            logger.warning(f"⚠️ Could not create summary sheets: {e}")

    # Return proper result dictionary
    if files_processed > 0:
        return {
            "success": True,
            "files_processed": files_processed,
            "semester": semester_key,
        }
    else:
        return {"success": False, "files_processed": 0, "error": "No files processed"}


def compute_remarks(row, total_cu, pass_threshold):
    """
    Compute remarks based on student status - FIXED NBTE STANDARD LOGIC.
    This function MUST be called AFTER GPA calculation.
    """
    # Get all course codes from the row (excluding non-course columns)
    course_columns = [col for col in row.index if col not in [
        "S/N", "EXAMS NUMBER", "NAME", "FAILED COURSES", "REMARKS", 
        "CU Passed", "CU Failed", "TCPE", "GPA", "AVERAGE"
    ]]
    
    # Count failed courses
    fails = [c for c in course_columns if float(row.get(c, 0) or 0) < pass_threshold]
    if not fails:
        return "Passed"
    
    # Get already calculated values
    cu_failed = row.get("CU Failed", 0)
    gpa = row.get("GPA", 0)
    
    failed_percentage = (cu_failed / total_cu) * 100 if total_cu > 0 else 0
    
    # FIXED: CORRECT NBTE STANDARD LOGIC
    # 1. GPA >= 2.0 AND failed_percentage <= 45% = Resit
    if gpa >= 2.0 and failed_percentage <= 45:
        return "Resit"
    
    # 2. GPA < 2.0 AND failed_percentage <= 45% = Probation  
    elif gpa < 2.0 and failed_percentage <= 45:
        return "Probation"
    
    # 3. failed_percentage > 45% = Withdrawn (regardless of GPA)
    elif failed_percentage > 45:
        return "Withdrawn"
    
    # Fallback
    else:
        return "Resit"


def process_single_file(
    path,
    output_dir,
    ts,
    pass_threshold,
    semester_course_maps,
    semester_credit_units,
    semester_lookup,
    semester_course_titles,
    logo_path,
    semester_key,
    set_name,
    previous_gpas,
    cgpa_data=None,
    upgrade_min_threshold=None,
):
    """
    Process a single raw file and produce mastersheet Excel and PDFs.
    FIXED: Correct NBTE standard logic for Probation/Withdrawn determination
    """
    fname = os.path.basename(path)

    # FIX: Initialize output_subdir EARLY to avoid UnboundLocalError
    output_subdir = os.path.join(output_dir, f"{set_name}_RESULT-{ts}")
    os.makedirs(output_subdir, exist_ok=True)
    logger.info(f"📁 Created output directory: {output_subdir}")

    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        logger.error(f"Error opening excel {path}: {e}")
        return None

    expected_sheets = ["CA", "OBJ", "EXAM"]
    dfs = {}
    for s in expected_sheets:
        if s in xl.sheet_names:
            dfs[s] = pd.read_excel(path, sheet_name=s, dtype=str)
    if not dfs:
        logger.error("No CA/OBJ/EXAM sheets detected — skipping file.")
        return None

    # Use the provided semester key
    sem = semester_key
    year, semester_num, level_display, semester_display, set_code = (
        get_semester_display_info(sem)
    )
    logger.info(
        f"📁 Processing: {level_display} - {semester_display} - Set: {set_code}"
    )
    logger.info(f"📊 Using course sheet: {sem}")

    logger.info(f"📊 Previous GPAs provided: {len(previous_gpas)} students")
    logger.info(
        f"📊 CGPA data available for: {len(cgpa_data) if cgpa_data else 0} students"
    )

    # Check if semester exists in course maps
    if sem not in semester_course_maps:
        logger.error(
            f"❌ Semester '{sem}' not found in course data. Available semesters: {list(semester_course_maps.keys())}"
        )
        return None

    course_map = semester_course_maps[sem]
    credit_units = semester_credit_units[sem]
    course_titles = semester_course_titles[sem]

    ordered_titles = list(course_map.keys())
    ordered_codes = [course_map[t] for t in ordered_titles if course_map.get(t)]
    ordered_codes = [c for c in ordered_codes if credit_units.get(c, 0) > 0]
    filtered_credit_units = {c: credit_units[c] for c in ordered_codes}
    total_cu = sum(filtered_credit_units.values())

    reg_no_cols = {
        s: find_column_by_names(
            df,
            [
                "REG. No",
                "Reg No",
                "Registration Number",
                "Mat No",
                "Exam No",
                "Student ID",
            ],
        )
        for s, df in dfs.items()
    }
    name_cols = {
        s: find_column_by_names(df, ["NAME", "Full Name", "Candidate Name"])
        for s, df in dfs.items()
    }

    merged = None
    for s, df in dfs.items():
        df = df.copy()
        regcol = reg_no_cols.get(s)
        namecol = name_cols.get(s)
        if not regcol:
            regcol = df.columns[0] if len(df.columns) > 0 else None
        if not namecol and len(df.columns) > 1:
            namecol = df.columns[1]

        if regcol is None:
            logger.warning(f"Skipping sheet {s}: no reg column found")
            continue

        df["REG. No"] = df[regcol].astype(str).str.strip()
        if namecol:
            df["NAME"] = df[namecol].astype(str).str.strip()
        else:
            df["NAME"] = pd.NA

        to_drop = [c for c in [regcol, namecol] if c and c not in ["REG. No", "NAME"]]
        df.drop(columns=to_drop, errors="ignore", inplace=True)

        for col in [c for c in df.columns if c not in ["REG. No", "NAME"]]:
            norm = normalize_course_name(col)
            matched_code = None
            for title, code in zip(
                ordered_titles, [course_map[t] for t in ordered_titles]
            ):
                if normalize_course_name(title) == norm:
                    matched_code = code
                    break
            if matched_code:
                newcol = f"{matched_code}_{s.upper()}"
                df.rename(columns={col: newcol}, inplace=True)

        cur_cols = ["REG. No", "NAME"] + [
            c for c in df.columns if c.endswith(f"_{s.upper()}")
        ]
        cur = df[cur_cols].copy()
        if merged is None:
            merged = cur
        else:
            merged = merged.merge(cur, on="REG. No", how="outer", suffixes=("", "_dup"))
            if "NAME_dup" in merged.columns:
                merged["NAME"] = merged["NAME"].combine_first(merged["NAME_dup"])
                merged.drop(columns=["NAME_dup"], inplace=True)

    if merged is None or merged.empty:
        logger.error("No data merged from sheets — skipping file.")
        return None

    mastersheet = merged[["REG. No", "NAME"]].copy()
    mastersheet.rename(columns={"REG. No": "EXAMS NUMBER"}, inplace=True)

    for code in ordered_codes:
        ca_col = f"{code}_CA"
        obj_col = f"{code}_OBJ"
        exam_col = f"{code}_EXAM"

        ca_series = (
            pd.to_numeric(merged[ca_col], errors="coerce")
            if ca_col in merged.columns
            else pd.Series([0] * len(merged), index=merged.index)
        )
        obj_series = (
            pd.to_numeric(merged[obj_col], errors="coerce")
            if obj_col in merged.columns
            else pd.Series([0] * len(merged), index=merged.index)
        )
        exam_series = (
            pd.to_numeric(merged[exam_col], errors="coerce")
            if exam_col in merged.columns
            else pd.Series([0] * len(merged), index=merged.index)
        )

        ca_norm = (ca_series / 20) * 100
        obj_norm = (obj_series / 20) * 100
        exam_norm = (exam_series / 80) * 100
        ca_norm = ca_norm.fillna(0).clip(upper=100)
        obj_norm = obj_norm.fillna(0).clip(upper=100)
        exam_norm = exam_norm.fillna(0).clip(upper=100)
        total = (ca_norm * 0.2) + (((obj_norm + exam_norm) / 2) * 0.8)
        mastersheet[code] = total.round(0).clip(upper=100).values

    # Apply upgrade rule if specified
    if should_use_interactive_mode():
        upgrade_min_threshold, upgraded_scores_count = get_upgrade_threshold_from_user(
            semester_key, set_name
        )
    else:
        upgraded_scores_count = 0
        if upgrade_min_threshold is not None:
            logger.info(
                f"🔄 Applying upgrade rule from parameters: {upgrade_min_threshold}–49 → 50"
            )

    if upgrade_min_threshold is not None:
        mastersheet, upgraded_scores_count = apply_upgrade_rule(
            mastersheet, ordered_codes, upgrade_min_threshold
        )

    for c in ordered_codes:
        if c not in mastersheet.columns:
            mastersheet[c] = 0

    # Calculate TCPE, TCUP, TCUF correctly FIRST
    def calc_tcpe_tcup_tcuf(row):
        tcpe = 0.0
        tcup = 0
        tcuf = 0

        for code in ordered_codes:
            score = float(row.get(code, 0) or 0)
            cu = filtered_credit_units.get(code, 0)
            gp = get_grade_point(score)

            # TCPE: Grade Point × Credit Units
            tcpe += gp * cu

            # TCUP/TCUF: Count credit units based on pass/fail
            if score >= pass_threshold:
                tcup += cu
            else:
                tcuf += cu

        return tcpe, tcup, tcuf

    # Apply calculations to each row
    results = mastersheet.apply(calc_tcpe_tcup_tcuf, axis=1, result_type="expand")
    mastersheet["TCPE"] = results[0].round(1)
    mastersheet["CU Passed"] = results[1]
    mastersheet["CU Failed"] = results[2]
    
    # Calculate GPA BEFORE remarks
    def calculate_gpa(row):
        tcpe = row["TCPE"]
        return round((tcpe / total_cu), 2) if total_cu > 0 else 0.0

    mastersheet["GPA"] = mastersheet.apply(calculate_gpa, axis=1)

    # NOW calculate FAILED COURSES and REMARKS after GPA is available
    def get_failed_courses(row):
        """Get list of failed courses for a student."""
        fails = [c for c in ordered_codes if float(row.get(c, 0) or 0) < pass_threshold]
        return ", ".join(sorted(fails)) if fails else ""

    # Add FAILED COURSES column
    mastersheet["FAILED COURSES"] = mastersheet.apply(get_failed_courses, axis=1)
    
    # FIXED: Use the corrected compute_remarks function with proper NBTE logic
    mastersheet["REMARKS"] = mastersheet.apply(
        lambda row: compute_remarks(row, total_cu, pass_threshold), 
        axis=1
    )

    mastersheet["AVERAGE"] = (
        mastersheet[[c for c in ordered_codes]].mean(axis=1).round(0)
    )

    # Filter out previously withdrawn students
    mastersheet, removed_students = filter_out_withdrawn_students(
        mastersheet, semester_key
    )

    # Identify withdrawn students in this semester
    withdrawn_students = []
    for idx, row in mastersheet.iterrows():
        student_status = determine_student_status(row, total_cu, pass_threshold)
        if student_status == "Withdrawn":
            exam_no = str(row["EXAMS NUMBER"]).strip()
            withdrawn_students.append(exam_no)
            mark_student_withdrawn(exam_no, semester_key)
            logger.info(f"🚫 Student {exam_no} marked as withdrawn in {semester_key}")

    # Update student tracker
    exam_numbers = mastersheet["EXAMS NUMBER"].astype(str).str.strip().tolist()
    update_student_tracker(semester_key, exam_numbers, withdrawn_students)

    # Identify carryover students
    carryover_students = identify_carryover_students(
        mastersheet, semester_key, set_name, pass_threshold
    )

    if carryover_students:
        carryover_dir = save_carryover_records(
            carryover_students, output_subdir, set_name, semester_key
        )
        logger.info("✅ Saved {} BM carryover records".format(len(carryover_students)))

    # Sort by remarks
    def sort_key(remark):
        if remark == "Passed":
            return (0, "")
        elif remark == "Resit":
            return (1, "")
        elif remark == "Probation":
            return (2, "")
        else:  # Withdrawn
            return (3, "")

    mastersheet = mastersheet.sort_values(
        by="REMARKS", key=lambda x: x.map(sort_key)
    ).reset_index(drop=True)

    if "S/N" not in mastersheet.columns:
        mastersheet.insert(0, "S/N", range(1, len(mastersheet) + 1))
    else:
        mastersheet["S/N"] = range(1, len(mastersheet) + 1)
        cols = list(mastersheet.columns)
        if cols[0] != "S/N":
            cols.remove("S/N")
            mastersheet = mastersheet[["S/N"] + cols]

    course_cols = ordered_codes
    # UPDATED: New column order with FAILED COURSES before REMARKS
    out_cols = (
        ["S/N", "EXAMS NUMBER", "NAME"]
        + course_cols
        + ["FAILED COURSES", "REMARKS", "CU Passed", "CU Failed", "TCPE", "GPA", "AVERAGE"]
    )
    for c in out_cols:
        if c not in mastersheet.columns:
            mastersheet[c] = pd.NA
    mastersheet = mastersheet[out_cols]

    # Save to Excel
    out_xlsx = os.path.join(output_subdir, f"mastersheet_{ts}.xlsx")

    if not os.path.exists(out_xlsx):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
    else:
        wb = load_workbook(out_xlsx)

    if sem not in wb.sheetnames:
        ws = wb.create_sheet(title=sem)
    else:
        ws = wb[sem]

    try:
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
    except Exception:
        pass

    ws.insert_rows(1, 2)
    logo_path_norm = os.path.normpath(logo_path) if logo_path else None
    if logo_path_norm and os.path.exists(logo_path_norm):
        try:
            img = XLImage(logo_path_norm)
            img.width, img.height = 110, 110
            ws.add_image(img, "A1")
        except Exception as e:
            logger.warning(f"⚠ Could not place logo: {e}")

    ws.merge_cells("C1:Q1")
    title_cell = ws["C1"]
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="1E90FF", end_color="1E90FF", fill_type="solid"
    )
    border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    title_cell.border = border

    # Use expanded semester name in the subtitle
    expanded_semester_name = f"{level_display} {semester_display}"

    ws.merge_cells("C2:Q2")
    subtitle_cell = ws["C2"]
    subtitle_cell.value = f"{datetime.now().year}/{datetime.now().year + 1} SESSION  BASIC MIDWIFERY {expanded_semester_name} EXAMINATIONS RESULT — {datetime.now().strftime('%B %d, %Y')}"
    subtitle_cell.font = Font(bold=True, size=12, color="000000")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # FIXED: Remove the upgrade notice from the header to prevent covering course titles
    # The upgrade notice will only appear in the summary section
    start_row = 3

    display_course_titles = []
    for t, c in zip(ordered_titles, [course_map[t] for t in ordered_titles]):
        if c in ordered_codes:
            display_course_titles.append(t)

    ws.append([""] * 3 + display_course_titles + [""] * 5)
    for i, cell in enumerate(
        ws[start_row][3 : 3 + len(display_course_titles)], start=3
    ):
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=45
        )
        cell.font = Font(bold=True, size=9)
    ws.row_dimensions[start_row].height = 18

    cu_list = [filtered_credit_units.get(c, "") for c in ordered_codes]
    ws.append([""] * 3 + cu_list + [""] * 5)
    for cell in ws[start_row + 1][3 : 3 + len(cu_list)]:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=135
        )
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

    headers = out_cols
    ws.append(headers)
    for cell in ws[start_row + 2]:
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="4A90E2", end_color="4A90E2", fill_type="solid"
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    for _, r in mastersheet.iterrows():
        rowvals = [r[col] for col in headers]
        ws.append(rowvals)

    # FIXED: Freeze the column headers (S/N, EXAMS NUMBER, NAME, etc.) at row start_row + 3
    # This ensures all column headers remain visible when scrolling
    ws.freeze_panes = ws.cell(row=start_row + 3, column=1)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=start_row + 3, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border

    # Colorize course columns - SPECIAL COLOR FOR UPGRADED SCORES
    upgraded_fill = PatternFill(
        start_color="E6FFCC", end_color="E6FFCC", fill_type="solid"
    )  # Light green for upgraded scores
    passed_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )  # Normal green for passed
    failed_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )  # White for failed

    for idx, code in enumerate(ordered_codes, start=4):
        col_letter = get_column_letter(idx)
        for r_idx in range(start_row + 3, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=idx)
            try:
                val = float(cell.value) if cell.value not in (None, "") else 0
                if (
                    upgrade_min_threshold is not None
                    and upgrade_min_threshold <= val <= 49
                ):
                    # This score was upgraded - use special color
                    cell.fill = upgraded_fill
                    # Dark green for upgraded scores
                    cell.font = Font(color="006600", bold=True)
                elif val >= pass_threshold:
                    cell.fill = passed_fill
                    cell.font = Font(color="006100")
                else:
                    cell.fill = failed_fill
                    cell.font = Font(color="FF0000", bold=True)
            except Exception:
                continue

    # FIXED: Apply specific column alignments
    left_align_columns = ["CU Passed", "CU Failed", "TCPE", "GPA", "AVERAGE"]

    for col_idx, col_name in enumerate(headers, start=1):
        if col_name in left_align_columns:
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row + 3, ws.max_row + 1):  # Start from data rows
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Center align S/N column
        elif col_name == "S/N":
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row + 3, ws.max_row + 1):  # Start from data rows
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # FIXED: Colorize REMARKS column based on status
    remarks_col_index = None
    for col_idx, col_name in enumerate(headers, 1):
        if col_name == "REMARKS":
            remarks_col_index = col_idx
            break

    if remarks_col_index:
        # Define colors for different remarks
        passed_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        resit_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   # Yellow
        probation_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        withdrawn_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

        for row_idx in range(start_row + 3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=remarks_col_index)
            remark_value = str(cell.value).strip() if cell.value else ""
            
            if remark_value == "Passed":
                cell.fill = passed_fill
                cell.font = Font(bold=True, color="006400")  # Dark green text
            elif remark_value == "Resit":
                cell.fill = resit_fill
                cell.font = Font(bold=True, color="8B8000")  # Dark yellow text
            elif remark_value == "Probation":
                cell.fill = probation_fill
                cell.font = Font(bold=True, color="8B4500")  # Dark orange text
            elif remark_value == "Withdrawn":
                cell.fill = withdrawn_fill
                cell.font = Font(bold=True, color="8B0000")  # Dark red text

    # FIXED: Auto-fit column widths for ALL columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Skip the first few header rows when calculating max length
        for cell in column:
            if cell.row >= start_row + 2:  # Start from header row
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
        
        # Add some padding and set a reasonable maximum width
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters wide
        if adjusted_width < 8:  # Minimum width
            adjusted_width = 8
        ws.column_dimensions[column_letter].width = adjusted_width

    # Special handling for NAME column - make it wider
    name_col_index = None
    for col_idx, col_name in enumerate(headers, 1):
        if col_name == "NAME":
            name_col_index = col_idx
            break
    
    if name_col_index:
        name_col_letter = get_column_letter(name_col_index)
        ws.column_dimensions[name_col_letter].width = 30  # Fixed width for names

    # Special handling for FAILED COURSES column - make it wider
    failed_courses_col_index = None
    for col_idx, col_name in enumerate(headers, 1):
        if col_name == "FAILED COURSES":
            failed_courses_col_index = col_idx
            break
    
    if failed_courses_col_index:
        failed_courses_col_letter = get_column_letter(failed_courses_col_index)
        ws.column_dimensions[failed_courses_col_letter].width = 40  # Wider for course lists

    # Fails per course row
    fails_per_course = (
        mastersheet[ordered_codes].apply(lambda x: (x < pass_threshold).sum()).tolist()
    )
    footer_vals = (
        [""] * 2
        + ["FAILS PER COURSE:"]
        + fails_per_course
        + [""] * (len(headers) - 3 - len(ordered_codes))
    )
    ws.append(footer_vals)
    for cell in ws[ws.max_row]:
        if 4 <= cell.column < 4 + len(ordered_codes):
            cell.fill = PatternFill(
                start_color="F0E68C", end_color="F0E68C", fill_type="solid"
            )
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        elif cell.column == 3:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # COMPREHENSIVE SUMMARY BLOCK
    total_students = len(mastersheet)
    passed_all = len(mastersheet[mastersheet["REMARKS"] == "Passed"])
    resit_count = len(mastersheet[mastersheet["REMARKS"] == "Resit"])
    probation_count = len(mastersheet[mastersheet["REMARKS"] == "Probation"])
    withdrawn_count = len(mastersheet[mastersheet["REMARKS"] == "Withdrawn"])

    # Add withdrawn student tracking to summary
    ws.append([])
    ws.append(["SUMMARY"])
    ws.append(
        [f"A total of {total_students} students registered and sat for the Examination"]
    )
    ws.append(
        [
            f"A total of {passed_all} students passed in all courses registered and are to proceed to Second Semester, BM I"
        ]
    )
    ws.append(
        [
            f"A total of {resit_count} students with Grade Point Average (GPA) of 2.00 and above failed various courses, but passed at least 45% of the total registered credit units, and are to carry these courses over to the next session."
        ]
    )
    ws.append(
        [
            f"A total of {probation_count} students with Grade Point Average (GPA) below 2.00 failed various courses, but passed at least 45% of the total registered credit units, and are placed on Probation, to carry these courses over to the next session."
        ]
    )
    ws.append(
        [
            f"A total of {withdrawn_count} students failed in more than 45% of their registered credit units in various courses and have been advised to withdraw"
        ]
    )

    # FIXED: Keep the upgrade notice only in the summary section, not in the header
    if upgrade_min_threshold is not None:
        ws.append(
            [
                f"✅ Upgraded all scores between {upgrade_min_threshold}–49 to 50 as per management decision ({upgraded_scores_count} scores upgraded)"
            ]
        )

    # Add removed withdrawn students info
    if removed_students:
        ws.append(
            [
                f"NOTE: {len(removed_students)} previously withdrawn students were removed from this semester's results as they should not be processed."
            ]
        )

    ws.append(
        [
            "The above decisions are in line with the provisions of the General Information Section of the General NMCN/NBTE Examinations Regulations (Pg 4) adopted by the College."
        ]
    )
    ws.append([])
    ws.append(
        [
            "________________________",
            "",
            "",
            "________________________",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Mrs. Abini Hauwa",
            "",
            "",
            "Mrs. Olukemi Ogunleye",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Head of Exams",
            "",
            "",
            "Chairman, BM Program C'tee",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    wb.save(out_xlsx)
    logger.info(f"✅ Mastersheet saved: {out_xlsx}")

    # Generate individual student PDF with previous GPAs and CGPA
    safe_sem = re.sub(r"[^\w\-]", "_", sem)
    student_pdf_path = os.path.join(
        output_subdir, f"mastersheet_students_{ts}_{safe_sem}.pdf"
    )

    logger.info(f"📊 FINAL CHECK before PDF generation:")
    logger.info(f"   Previous GPAs loaded: {len(previous_gpas)}")
    logger.info(
        f"   CGPA data available for: {len(cgpa_data) if cgpa_data else 0} students"
    )
    if previous_gpas:
        sample = list(previous_gpas.items())[:3]
        logger.info(f"   Sample GPAs: {sample}")

    try:
        generate_individual_student_pdf(
            mastersheet,
            student_pdf_path,
            sem,
            logo_path=logo_path_norm,
            prev_mastersheet_df=None,
            filtered_credit_units=filtered_credit_units,
            ordered_codes=ordered_codes,
            course_titles_map=course_titles,
            previous_gpas=previous_gpas,
            cgpa_data=cgpa_data,
            total_cu=total_cu,
            pass_threshold=pass_threshold,
            upgrade_min_threshold=upgrade_min_threshold,
        )  # PASS THE UPGRADE THRESHOLD TO PDF
        logger.info(f"✅ PDF generated successfully for {sem}")
    except Exception as e:
        logger.error(f"❌ Failed to generate student PDF for {sem}: {e}")
        traceback.print_exc()

    return mastersheet


# ----------------------------
# Main runner
# ----------------------------


def main():
    """Main entry point with comprehensive error handling."""
    try:
        logger.info("Starting BM Examination Results Processing...")
        ts = datetime.now().strftime(TIMESTAMP_FMT)

        # Initialize trackers
        initialize_student_tracker()
        initialize_carryover_tracker()  # NEW: Initialize carryover tracker

        # Check if running in web mode
        if is_web_mode():
            uploaded_file_path = get_uploaded_file_path()
            if uploaded_file_path and os.path.exists(uploaded_file_path):
                logger.info("🔧 Running in WEB MODE with uploaded file")
                success = process_uploaded_file(
                    uploaded_file_path, normalize_path(BASE_DIR)
                )
                if success:
                    logger.info("✅ Uploaded file processing completed successfully")
                else:
                    logger.error("❌ Uploaded file processing failed")
                return
            else:
                logger.error("❌ No uploaded file found in web mode")
                return

        # Get parameters from form
        params = get_form_parameters()

        # Use the parameters
        global DEFAULT_PASS_THRESHOLD
        DEFAULT_PASS_THRESHOLD = params["pass_threshold"]

        base_dir_norm = normalize_path(BASE_DIR)
        logger.info(f"Using base directory: {base_dir_norm}")

        # Check if we should use interactive or non-interactive mode
        if should_use_interactive_mode():
            logger.info("🔧 Running in INTERACTIVE mode (CLI)")

            try:
                (
                    semester_course_maps,
                    semester_credit_units,
                    semester_lookup,
                    semester_course_titles,
                ) = load_course_data()
            except Exception as e:
                logger.error(f"❌ Could not load course data: {e}")
                return

            # Get available sets and let user choose
            available_sets = get_available_sets(base_dir_norm)

            if not available_sets:
                logger.error(
                    f"No BM-* directories found in {base_dir_norm}. Nothing to process."
                )
                logger.error(f"Available directories: {os.listdir(base_dir_norm)}")
                return

            logger.info(
                f"📚 Found {len(available_sets)} available sets: {available_sets}"
            )

            # Let user choose which set(s) to process
            sets_to_process = get_user_set_choice(available_sets)

            logger.info(f"\n🎯 PROCESSING SELECTED SETS: {sets_to_process}")

            for bm_set in sets_to_process:
                logger.info(f"\n{'='*60}")
                logger.info(f"PROCESSING SET: {bm_set}")
                logger.info(f"{'='*60}")

                # BM: Raw and clean directories under BM folder
                raw_dir = normalize_path(
                    os.path.join(base_dir_norm, "BM", bm_set, "RAW_RESULTS")
                )
                clean_dir = normalize_path(
                    os.path.join(base_dir_norm, "BM", bm_set, "CLEAN_RESULTS")
                )

                # Create directories if they don't exist
                os.makedirs(raw_dir, exist_ok=True)
                os.makedirs(clean_dir, exist_ok=True)

                # Check if raw directory exists and has files
                if not os.path.exists(raw_dir):
                    logger.warning(f"⚠️ RAW_RESULTS directory not found: {raw_dir}")
                    continue

                raw_files = [
                    f
                    for f in os.listdir(raw_dir)
                    if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
                ]
                if not raw_files:
                    logger.warning(f"⚠️ No raw files in {raw_dir}; skipping {bm_set}")
                    logger.info(f"   Available files: {os.listdir(raw_dir)}")
                    continue

                logger.info(
                    f"📁 Found {len(raw_files)} raw files in {bm_set}: {raw_files}"
                )

                # Get user choice for which semesters to process
                semesters_to_process = get_user_semester_choice()

                logger.info(
                    f"\n🎯 PROCESSING SELECTED SEMESTERS for {bm_set}: {[get_semester_display_info(sem)[3] for sem in semesters_to_process]}"
                )

                # Process selected semesters in the correct order
                for semester_key in semesters_to_process:
                    if semester_key not in SEMESTER_ORDER:
                        logger.warning(f"⚠️ Skipping unknown semester: {semester_key}")
                        continue

                    # Check if there are files for this semester
                    semester_files_exist = False
                    for rf in raw_files:
                        detected_sem, _, _, _, _, _ = detect_semester_from_filename(rf)
                        if detected_sem == semester_key:
                            semester_files_exist = True
                            break

                    if semester_files_exist:
                        logger.info(f"\n🎯 Processing {semester_key} in {bm_set}...")
                        result = process_semester_files(
                            semester_key,
                            raw_files,
                            raw_dir,
                            clean_dir,
                            ts,
                            DEFAULT_PASS_THRESHOLD,
                            semester_course_maps,
                            semester_credit_units,
                            semester_lookup,
                            semester_course_titles,
                            DEFAULT_LOGO_PATH,
                            bm_set,
                        )

                        if result is not None and result.get("success", False):
                            logger.info(f"✅ Successfully processed {semester_key}")
                        else:
                            logger.error(f"❌ Failed to process {semester_key}")
                    else:
                        logger.warning(
                            f"⚠️ No files found for {semester_key} in {bm_set}, skipping..."
                        )

                # Zip the results after processing all semesters for this set
                result_folder = os.path.join(clean_dir, f"{bm_set}_RESULT-{ts}")
                if os.path.exists(result_folder):
                    zip_path = os.path.join(clean_dir, f"{bm_set}_RESULT-{ts}.zip")
                    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                        for root, _, files in os.walk(result_folder):
                            for file in files:
                                file_path = os.path.join(root, file)
                                arcname = os.path.relpath(file_path, result_folder)
                                zipf.write(file_path, arcname)
                    logger.info(f"✅ Zipped results: {zip_path}")
                    # Remove the folder after zipping to clean up
                    shutil.rmtree(result_folder)

            # Print student tracking summary
            logger.info(f"\n📊 STUDENT TRACKING SUMMARY:")
            logger.info(f"Total unique students tracked: {len(STUDENT_TRACKER)}")
            logger.info(f"Total withdrawn students: {len(WITHDRAWN_STUDENTS)}")

            # NEW: Print carryover summary
            if CARRYOVER_STUDENTS:
                logger.info(f"\n📋 BM CARRYOVER STUDENT SUMMARY:")
                logger.info(f"Total BM carryover students: {len(CARRYOVER_STUDENTS)}")

            # Print withdrawn students who reappeared
            reappeared_count = 0
            for exam_no, data in WITHDRAWN_STUDENTS.items():
                if data["reappeared_semesters"]:
                    reappeared_count += 1
                    logger.warning(
                        f"🚨 {exam_no}: Withdrawn in {data['withdrawn_semester']}, reappeared in {data['reappeared_semesters']}"
                    )

            if reappeared_count > 0:
                logger.warning(
                    f"🚨 ALERT: {reappeared_count} previously withdrawn students have reappeared in later semesters!"
                )

            # Analyze student progression
            sem_counts = {}
            for student_data in STUDENT_TRACKER.values():
                sem_count = len(student_data["semesters_present"])
                if sem_count not in sem_counts:
                    sem_counts[sem_count] = 0
                sem_counts[sem_count] += 1

            for sem_count, student_count in sorted(sem_counts.items()):
                logger.info(
                    f"Students present in {sem_count} semester(s): {student_count}"
                )

            logger.info(
                "\n✅ BM Examination Results Processing completed successfully."
            )
        else:
            logger.info("🔧 Running in NON-INTERACTIVE mode (Web)")
            success = process_in_non_interactive_mode(params, base_dir_norm)
            if success:
                logger.info(
                    "✅ BM Examination Results Processing completed successfully"
                )
            else:
                logger.error("❌ BM Examination Results Processing failed")
            return

    except KeyboardInterrupt:
        logger.info("\n👋 Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        traceback.print_exc()
        sys.exit(1)
    finally:
        # Cleanup
        logger.info("\n📊 Final Summary:")
        logger.info(f"   Students tracked: {len(STUDENT_TRACKER)}")
        logger.info(f"   Carryover students: {len(CARRYOVER_STUDENTS)}")
        logger.info(f"   Withdrawn students: {len(WITHDRAWN_STUDENTS)}")


if __name__ == "__main__":
    main()
