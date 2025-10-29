#!/usr/bin/env python3
"""
integrated_carryover_processor.py - FIXED VERSION WITH ALL-SEMESTER GPA TRACKING
Enhanced to track GPA for all semesters, not just second year second semester.
"""

import os
import sys
import re
import pandas as pd
import numpy as np
from datetime import datetime
import glob
import json
import traceback
import shutil
import zipfile
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def standardize_semester_key(semester_key):
    """Standardize semester key to canonical format."""
    if not semester_key:
        return None
    
    key_upper = semester_key.upper()
    
    # Define canonical mappings
    canonical_mappings = {
        # First Year First Semester variants
        ("FIRST", "YEAR", "FIRST", "SEMESTER"): "ND-First-YEAR-First-SEMESTER",
        ("1ST", "YEAR", "1ST", "SEMESTER"): "ND-First-YEAR-First-SEMESTER",
        ("YEAR", "1", "SEMESTER", "1"): "ND-First-YEAR-First-SEMESTER",
        
        # First Year Second Semester variants
        ("FIRST", "YEAR", "SECOND", "SEMESTER"): "ND-First-YEAR-SECOND-SEMESTER",
        ("1ST", "YEAR", "2ND", "SEMESTER"): "ND-First-YEAR-SECOND-SEMESTER",
        ("YEAR", "1", "SEMESTER", "2"): "ND-First-YEAR-SECOND-SEMESTER",
        
        # Second Year First Semester variants
        ("SECOND", "YEAR", "FIRST", "SEMESTER"): "ND-SECOND-YEAR-First-SEMESTER",
        ("2ND", "YEAR", "1ST", "SEMESTER"): "ND-SECOND-YEAR-First-SEMESTER",
        ("YEAR", "2", "SEMESTER", "1"): "ND-SECOND-YEAR-First-SEMESTER",
        
        # Second Year Second Semester variants
        ("SECOND", "YEAR", "SECOND", "SEMESTER"): "ND-SECOND-YEAR-SECOND-SEMESTER",
        ("2ND", "YEAR", "2ND", "SEMESTER"): "ND-SECOND-YEAR-SECOND-SEMESTER",
        ("YEAR", "2", "SEMESTER", "2"): "ND-SECOND-YEAR-SECOND-SEMESTER",
    }
    
    # Extract key components using regex
    patterns = [
        r'(FIRST|1ST|YEAR.?1).*?(FIRST|1ST|SEMESTER.?1)',
        r'(FIRST|1ST|YEAR.?1).*?(SECOND|2ND|SEMESTER.?2)',
        r'(SECOND|2ND|YEAR.?2).*?(FIRST|1ST|SEMESTER.?1)',
        r'(SECOND|2ND|YEAR.?2).*?(SECOND|2ND|SEMESTER.?2)',
    ]
    
    for pattern_idx, pattern in enumerate(patterns):
        if re.search(pattern, key_upper):
            if pattern_idx == 0:
                return "ND-First-YEAR-First-SEMESTER"
            elif pattern_idx == 1:
                return "ND-First-YEAR-SECOND-SEMESTER"
            elif pattern_idx == 2:
                return "ND-SECOND-YEAR-First-SEMESTER"
            elif pattern_idx == 3:
                return "ND-SECOND-YEAR-SECOND-SEMESTER"
    
    # If no match, return original
    print(f"Could not standardize semester key: {semester_key}")
    return semester_key

def get_previous_semester(semester_key):
    """Get the previous semester key for carryover."""
    standardized = standardize_semester_key(semester_key)
    if standardized == "ND-First-YEAR-SECOND-SEMESTER":
        return "ND-First-YEAR-First-SEMESTER"
    elif standardized == "ND-SECOND-YEAR-First-SEMESTER":
        return "ND-First-YEAR-SECOND-SEMESTER"
    elif standardized == "ND-SECOND-YEAR-SECOND-SEMESTER":
        return "ND-SECOND-YEAR-First-SEMESTER"
    else:
        return None  # No previous for first semester

# ----------------------------
# Configuration
# ----------------------------
def get_base_directory():
    """Get base directory - ENHANCED VERSION."""
    # First try environment variable
    if os.getenv('BASE_DIR'):
        base_dir = os.getenv('BASE_DIR')
        if os.path.exists(base_dir):
            return base_dir
    
    # Try the user's home directory with student_result_cleaner
    home_dir = os.path.expanduser('~')
    default_dir = os.path.join(home_dir, 'student_result_cleaner')
    
    # Check if EXAMS_INTERNAL exists in the default directory
    if os.path.exists(os.path.join(default_dir, "EXAMS_INTERNAL")):
        return default_dir
    
    # If not, check if we're already in a directory that contains EXAMS_INTERNAL
    current_script_dir = os.path.dirname(os.path.abspath(__file__))
    if os.path.exists(os.path.join(current_script_dir, "EXAMS_INTERNAL")):
        return current_script_dir
    
    # Check parent directory
    parent_dir = os.path.dirname(current_script_dir)
    if os.path.exists(os.path.join(parent_dir, "EXAMS_INTERNAL")):
        return parent_dir
    
    # Final fallback
    return default_dir

BASE_DIR = get_base_directory()
TIMESTAMP_FMT = "%d-%m-%Y_%H%M%S"
DEFAULT_PASS_THRESHOLD = 50.0
DEFAULT_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")

def sanitize_filename(filename):
    """Remove or replace characters that are not safe for filenames."""
    return re.sub(r'[^\w\-_.]', '_', filename)

def find_exam_number_column(df):
    """Find the exam number column in a DataFrame."""
    possible_names = ['EXAM NUMBER', 'REG. No', 'REG NO', 'REGISTRATION NUMBER', 'MAT NO', 'STUDENT ID']
    for col in df.columns:
        col_upper = str(col).upper()
        for possible_name in possible_names:
            if possible_name in col_upper:
                return col
    return None

# NEW: Enhanced debugging function for course matching
def debug_course_matching(resit_file_path, course_code_to_title, course_code_to_unit):
    """Debug function to specifically check why course codes aren't matching."""
    print(f"\n🔍 DEBUGGING COURSE MATCHING")
    print("=" * 50)
    
    # Read resit file to see what course codes we have
    resit_df = pd.read_excel(resit_file_path, header=0)
    resit_exam_col = find_exam_number_column(resit_df)
    
    # Get all course codes from resit file (excluding exam number and name columns)
    resit_courses = []
    for col in resit_df.columns:
        if col != resit_exam_col and col != 'NAME' and not 'Unnamed' in str(col):
            resit_courses.append(col)
    
    print(f"📋 Course codes from resit file: {resit_courses}")
    print(f"📊 Total courses in resit file: {len(resit_courses)}")
    
    # Check each resit course against course file
    for course in resit_courses:
        print(f"\n🔍 Checking course: '{course}'")
        original_code = str(course).strip()
        
        # Generate the same variants as in find_course_title
        variants = [
            original_code.upper().strip(),
            original_code.strip(),
            original_code.upper(),
            original_code,
            original_code.upper().replace(' ', ''),
            original_code.replace(' ', ''),
            re.sub(r'\s+', '', original_code.upper()),
            re.sub(r'\s+', '', original_code),
            re.sub(r'[^a-zA-Z0-9]', '', original_code.upper()),
            re.sub(r'[^a-zA-Z0-9]', '', original_code),
            original_code.upper().replace('-', ''),
            original_code.upper().replace('_', ''),
            original_code.replace('-', '').replace('_', ''),
            original_code.lower(),
            original_code.title(),
        ]
        
        # Remove duplicates
        variants = list(dict.fromkeys([v for v in variants if v and v != 'NAN']))
        
        print(f"   Trying variants: {variants}")
        
        found = False
        for variant in variants:
            if variant in course_code_to_title:
                title = course_code_to_title[variant]
                unit = course_code_to_unit.get(variant, 0)
                print(f"   ✅ FOUND: '{variant}' -> '{title}' (CU: {unit})")
                found = True
                break
        
        if not found:
            print(f"   ❌ NOT FOUND: No match for '{course}'")
            # Show some similar keys that exist
            similar_keys = [k for k in course_code_to_title.keys() if course.upper()[:3] in k.upper()]
            if similar_keys:
                print(f"   💡 Similar keys in course file: {similar_keys[:5]}")
            else:
                print(f"   💡 Available keys sample: {list(course_code_to_title.keys())[:10]}")

# NEW: Function to find alternative course files
def find_alternative_course_files():
    """Look for alternative course files in case the main one is wrong."""
    base_dirs = [
        os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", "ND-COURSES"),
        os.path.join(BASE_DIR, "ND", "ND-COURSES"),
        os.path.join(BASE_DIR, "COURSES"),
        os.path.join(BASE_DIR, "EXAMS_INTERNAL"),
    ]
    
    course_files = []
    for base_dir in base_dirs:
        if os.path.exists(base_dir):
            for file in os.listdir(base_dir):
                if 'course' in file.lower() and file.endswith(('.xlsx', '.xls')):
                    full_path = os.path.join(base_dir, file)
                    course_files.append(full_path)
                    print(f"📁 Found course file: {full_path}")
    
    return course_files

# ENHANCED: Course Data Loading with multiple file support
def load_course_data():
    """Load course data from course-code-creditUnit.xlsx with enhanced matching and multiple file support."""
    # Try multiple possible course file locations
    possible_course_files = [
        os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", "ND-COURSES", "course-code-creditUnit.xlsx"),
        os.path.join(BASE_DIR, "ND", "ND-COURSES", "course-code-creditUnit.xlsx"),
        os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND-COURSES", "course-code-creditUnit.xlsx"),
        os.path.join(BASE_DIR, "course-code-creditUnit.xlsx"),
    ]
    
    course_file = None
    for possible_file in possible_course_files:
        if os.path.exists(possible_file):
            course_file = possible_file
            print(f"✅ Found course file: {course_file}")
            break
    
    # If main file doesn't exist, try alternative files
    if not course_file:
        print(f"❌ Main course file not found in standard locations")
        alternative_files = find_alternative_course_files()
        if alternative_files:
            course_file = alternative_files[0]
            print(f"🔄 Using alternative course file: {course_file}")
        else:
            print("❌ No course files found anywhere!")
            return {}, {}, {}, {}
    
    print(f"📚 Loading course data from: {course_file}")
    
    try:
        xl = pd.ExcelFile(course_file)
        semester_course_titles = {}
        semester_credit_units = {}
        course_code_to_title = {}
        course_code_to_unit = {}
        
        print(f"📖 Available sheets: {xl.sheet_names}")
        
        for sheet in xl.sheet_names:
            sheet_standard = standardize_semester_key(sheet)
            print(f"📖 Reading sheet: {sheet} (standardized: {sheet_standard})")
            try:
                df = pd.read_excel(course_file, sheet_name=sheet, engine='openpyxl', header=0)
                
                # Print raw columns for debugging
                print(f"🔍 Raw columns in sheet '{sheet}': {list(df.columns)}")
                
                # Convert columns to string and clean
                df.columns = [str(c).strip().upper() for c in df.columns]
                
                print(f"🔍 Cleaned columns in sheet '{sheet}': {list(df.columns)}")
                
                # Look for course code, title, and credit unit columns with flexible matching
                code_col = None
                title_col = None
                unit_col = None
                
                for col in df.columns:
                    col_clean = str(col).upper()
                    if any(keyword in col_clean for keyword in ['COURSE CODE', 'CODE', 'COURSECODE']):
                        code_col = col
                    elif any(keyword in col_clean for keyword in ['COURSE TITLE', 'TITLE', 'COURSENAME']):
                        title_col = col
                    elif any(keyword in col_clean for keyword in ['CU', 'CREDIT', 'UNIT', 'CREDIT UNIT']):
                        unit_col = col
                
                print(f"🔍 Detected columns - Code: {code_col}, Title: {title_col}, Unit: {unit_col}")
                
                if not all([code_col, title_col, unit_col]):
                    print(f"⚠️ Sheet '{sheet}' missing required columns - found: code={code_col}, title={title_col}, unit={unit_col}")
                    # Try to use first three columns as fallback
                    if len(df.columns) >= 3:
                        code_col, title_col, unit_col = df.columns[0], df.columns[1], df.columns[2]
                        print(f"🔄 Using fallback columns: {code_col}, {title_col}, {unit_col}")
                    else:
                        print(f"❌ Sheet '{sheet}' doesn't have enough columns - skipped")
                        continue
                
                # Clean the data
                df_clean = df.dropna(subset=[code_col]).copy()
                if df_clean.empty:
                    print(f"⚠️ Sheet '{sheet}' has no data after cleaning - skipped")
                    continue
                
                # Convert credit units to numeric, handling errors
                df_clean[unit_col] = pd.to_numeric(df_clean[unit_col], errors='coerce')
                df_clean = df_clean.dropna(subset=[unit_col])
                
                # Remove rows with "TOTAL" in course code
                df_clean = df_clean[~df_clean[code_col].astype(str).str.contains('TOTAL', case=False, na=False)]
                
                if df_clean.empty:
                    print(f"⚠️ Sheet '{sheet}' has no valid rows after cleaning - skipped")
                    continue
                
                codes = df_clean[code_col].astype(str).str.strip().tolist()
                titles = df_clean[title_col].astype(str).str.strip().tolist()
                units = df_clean[unit_col].astype(float).tolist()

                print(f"📋 Found {len(codes)} courses in {sheet}:")
                for i, (code, title, unit) in enumerate(zip(codes[:5], titles[:5], units[:5])):
                    print(f"   - '{code}': '{title}' (CU: {unit})")

                # Create mapping dictionaries with multiple normalization strategies
                sheet_titles = {}
                sheet_units = {}
                
                for code, title, unit in zip(codes, titles, units):
                    if not code or code.upper() in ['NAN', 'NONE', '']:
                        continue
                    
                    # Create multiple normalization variants for robust matching
                    variants = [
                        # Original variations
                        code.upper().strip(),
                        code.strip(),
                        # Remove all spaces and special characters
                        re.sub(r'[^a-zA-Z0-9]', '', code.upper()),
                        re.sub(r'[^a-zA-Z0-9]', '', code),
                        # Common formatting variations
                        code.upper().replace(' ', ''),
                        code.replace(' ', ''),
                        # Handle case variations
                        code.lower(),
                        code.upper(),
                        # Additional variants for edge cases
                        code.upper().replace('-', '').replace('_', ''),
                        code.replace('-', '').replace('_', ''),
                    ]
                    
                    # Remove duplicates while preserving order
                    variants = list(dict.fromkeys([v for v in variants if v]))
                    
                    # Add all variants to mappings
                    for variant in variants:
                        sheet_titles[variant] = title
                        sheet_units[variant] = unit
                        course_code_to_title[variant] = title
                        course_code_to_unit[variant] = unit
                
                semester_course_titles[sheet_standard] = sheet_titles
                semester_credit_units[sheet_standard] = sheet_units
                
            except Exception as e:
                print(f"❌ Error processing sheet '{sheet}': {e}")
                traceback.print_exc()
                continue
        
        print(f"✅ Loaded course data for sheets: {list(semester_course_titles.keys())}")
        print(f"📊 Total course mappings: {len(course_code_to_title)}")
        
        # Debug: Show some course mappings
        print("🔍 Sample course mappings:")
        sample_items = list(course_code_to_title.items())[:15]
        for code, title in sample_items:
            unit = course_code_to_unit.get(code, 0)
            print(f"   '{code}' -> '{title}' (CU: {unit})")
            
        return semester_course_titles, semester_credit_units, course_code_to_title, course_code_to_unit
        
    except Exception as e:
        print(f"❌ Error loading course data: {e}")
        traceback.print_exc()
        return {}, {}, {}, {}

# FIX 2: Enhanced Course Matching Functions
def find_course_title(course_code, course_titles_dict, course_code_to_title):
    """Robust function to find course title with comprehensive matching strategies."""
    if not course_code or str(course_code).upper() in ['NAN', 'NONE', '']:
        return str(course_code) if course_code else "Unknown Course"
    
    original_code = str(course_code).strip()
    
    # Generate comprehensive matching variants
    variants = [
        # Basic normalizations
        original_code.upper().strip(),
        original_code.strip(),
        original_code.upper(),
        original_code,
        
        # Space handling variations
        original_code.upper().replace(' ', ''),
        original_code.replace(' ', ''),
        re.sub(r'\s+', '', original_code.upper()),
        re.sub(r'\s+', '', original_code),
        
        # Special character handling
        re.sub(r'[^a-zA-Z0-9]', '', original_code.upper()),
        re.sub(r'[^a-zA-Z0-9]', '', original_code),
        
        # Common formatting issues
        original_code.upper().replace('-', ''),
        original_code.upper().replace('_', ''),
        original_code.replace('-', '').replace('_', ''),
        
        # Case variations
        original_code.lower(),
        original_code.title(),
    ]
    
    # Remove duplicates
    variants = list(dict.fromkeys([v for v in variants if v and v != 'NAN']))
    
    # Try each strategy in order
    for variant in variants:
        # Try course_titles_dict first (semester-specific)
        if variant in course_titles_dict:
            title = course_titles_dict[variant]
            return title
        
        # Try global course_code_to_title
        if variant in course_code_to_title:
            title = course_code_to_title[variant]
            return title
    
    # If no match found, log and return descriptive original code
    print(f"⚠️ Could not find course title for: '{original_code}'")
    return f"{original_code} (Title Not Found)"

def find_credit_unit(course_code, credit_units_dict, course_code_to_unit):
    """Robust function to find credit unit with comprehensive matching strategies."""
    if not course_code or str(course_code).upper() in ['NAN', 'NONE', '']:
        return 0
    
    original_code = str(course_code).strip()
    
    # Generate the same variants as title matching
    variants = [
        original_code.upper().strip(),
        original_code.strip(),
        original_code.upper(),
        original_code,
        original_code.upper().replace(' ', ''),
        original_code.replace(' ', ''),
        re.sub(r'\s+', '', original_code.upper()),
        re.sub(r'\s+', '', original_code),
        re.sub(r'[^a-zA-Z0-9]', '', original_code.upper()),
        re.sub(r'[^a-zA-Z0-9]', '', original_code),
        original_code.upper().replace('-', ''),
        original_code.upper().replace('_', ''),
        original_code.replace('-', '').replace('_', ''),
        original_code.lower(),
        original_code.title(),
    ]
    
    # Remove duplicates
    variants = list(dict.fromkeys([v for v in variants if v and v != 'NAN']))
    
    # Try each strategy
    for variant in variants:
        if variant in credit_units_dict:
            unit = credit_units_dict[variant]
            return unit
        
        if variant in course_code_to_unit:
            unit = course_code_to_unit[variant]
            return unit
    
    print(f"⚠️ Could not find credit unit for: '{original_code}'")
    return 0

def debug_course_file_structure():
    """Debug function to check the actual structure of the course file."""
    course_file = os.path.join(BASE_DIR, "EXAMS_INTERNAL", "ND", "ND-COURSES", "course-code-creditUnit.xlsx")
    print(f"\n🔍 DEBUGGING COURSE FILE: {course_file}")
    
    if not os.path.exists(course_file):
        print("❌ Course file not found!")
        # Try alternative files
        alternative_files = find_alternative_course_files()
        if alternative_files:
            course_file = alternative_files[0]
            print(f"🔄 Using alternative course file: {course_file}")
        else:
            print("❌ No course files found anywhere!")
            return
    
    try:
        xl = pd.ExcelFile(course_file)
        print(f"📖 Sheets in course file: {xl.sheet_names}")
        
        for sheet_name in xl.sheet_names:
            print(f"\n📄 Sheet: {sheet_name}")
            df = pd.read_excel(course_file, sheet_name=sheet_name)
            print(f"   Shape: {df.shape}")
            print(f"   Columns: {list(df.columns)}")
            
            # Show first few rows
            for i in range(min(3, len(df))):
                row = df.iloc[i]
                print(f"   Row {i}: {dict(row)}")
                
    except Exception as e:
        print(f"❌ Error reading course file: {e}")

def extract_mastersheet_from_zip(zip_path, semester_key):
    """Extract mastersheet from ZIP file and return temporary file path."""
    try:
        print(f"📦 Looking for mastersheet in ZIP: {zip_path}")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # List all files in ZIP for debugging
            all_files = zip_ref.namelist()
            print(f"📁 Files in ZIP: {all_files}")
            
            # Look for mastersheet files in the ZIP
            mastersheet_files = [f for f in all_files if 'mastersheet' in f.lower() and f.endswith('.xlsx')]
            
            if not mastersheet_files:
                print(f"❌ No mastersheet found in ZIP")
                return None, None
            
            # Use the first mastersheet found
            mastersheet_name = mastersheet_files[0]
            print(f"✅ Found mastersheet: {mastersheet_name}")
            
            # Extract to temporary file
            temp_dir = tempfile.mkdtemp()
            temp_mastersheet_path = os.path.join(temp_dir, f"mastersheet_{semester_key}.xlsx")
            
            with open(temp_mastersheet_path, 'wb') as f:
                f.write(zip_ref.read(mastersheet_name))
            
            print(f"✅ Extracted mastersheet to: {temp_mastersheet_path}")
            return temp_mastersheet_path, temp_dir
            
    except Exception as e:
        print(f"❌ Error extracting mastersheet from ZIP: {e}")
        traceback.print_exc()
        return None, None

def find_latest_zip_file(clean_dir):
    """Find the latest ZIP file in clean results directory - IMPROVED VERSION."""
    print(f"🔍 Looking for ZIP files in: {clean_dir}")
    
    if not os.path.exists(clean_dir):
        print(f"❌ Clean directory doesn't exist: {clean_dir}")
        return None
    
    # List all files in directory
    all_files = os.listdir(clean_dir)
    print(f"📁 Files in clean directory: {all_files}")
    
    # Look for ZIP files but EXCLUDE carryover ZIPs and include only regular result ZIPs
    zip_files = []
    for f in all_files:
        if f.lower().endswith('.zip'):
            # EXCLUDE carryover files
            if 'carryover' in f.lower():
                print(f"⚠️ Skipping carryover ZIP: {f}")
                continue
            
            # INCLUDE regular result files (like ND-2024_RESULT-*.zip)
            if any(pattern in f for pattern in ['_RESULT-', 'RESULT_', 'RESULT-']):
                zip_files.append(f)
                print(f"✅ Found regular results ZIP: {f}")
            else:
                print(f"ℹ️ Found other ZIP (not a result file): {f}")
    
    if not zip_files:
        print(f"❌ No regular results ZIP files found (excluding carryover files)")
        print(f"💡 Looking for any ZIP file that might contain mastersheet...")
        
        # Fallback: look for any ZIP that might contain mastersheet
        fallback_zips = [f for f in all_files if f.lower().endswith('.zip') and 'carryover' not in f.lower()]
        if fallback_zips:
            print(f"⚠️ Using fallback ZIP files: {fallback_zips}")
            zip_files = fallback_zips
        else:
            print(f"❌ No ZIP files found at all in {clean_dir}")
            return None
    
    print(f"✅ Final ZIP files to consider: {zip_files}")
    
    # Sort by modification time and return the latest
    zip_files_with_path = [os.path.join(clean_dir, f) for f in zip_files]
    latest_zip = sorted(zip_files_with_path, key=os.path.getmtime, reverse=True)[0]
    
    print(f"🎯 Using latest ZIP: {latest_zip}")
    return latest_zip

def find_latest_result_folder(clean_dir, set_name):
    """Find the latest result folder in clean results directory."""
    print(f"🔍 Looking for result folders in: {clean_dir}")
    
    if not os.path.exists(clean_dir):
        print(f"❌ Clean directory doesn't exist: {clean_dir}")
        return None
    
    # List all items in directory
    all_items = os.listdir(clean_dir)
    print(f"📁 Items in clean directory: {all_items}")
    
    # Look for result folders (assuming pattern like "{set_name}_RESULT-{timestamp}")
    result_folders = [f for f in all_items if os.path.isdir(os.path.join(clean_dir, f)) and f.startswith(f"{set_name}_RESULT-")]
    
    if not result_folders:
        print(f"❌ No result folders found")
        return None
    
    print(f"✅ Found result folders: {result_folders}")
    
    # Sort by modification time and return the latest
    folders_with_path = [os.path.join(clean_dir, f) for f in result_folders]
    latest_folder = sorted(folders_with_path, key=os.path.getmtime, reverse=True)[0]
    
    print(f"🎯 Using latest result folder: {latest_folder}")
    return latest_folder

def find_latest_mastersheet_source(clean_dir, set_name):
    """Find the latest source for mastersheet: prefer ZIP, fallback to folder - IMPROVED."""
    print(f"🔍 Looking for mastersheet source in: {clean_dir}")
    
    if not os.path.exists(clean_dir):
        print(f"❌ Clean directory doesn't exist: {clean_dir}")
        return None, None
    
    # First try to find ZIP file
    zip_path = find_latest_zip_file(clean_dir)
    if zip_path:
        print(f"✅ Using ZIP source: {zip_path}")
        # Verify ZIP contains mastersheet
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_files = zip_ref.namelist()
                mastersheet_files = [f for f in zip_files if 'mastersheet' in f.lower() and f.endswith('.xlsx')]
                if mastersheet_files:
                    print(f"✅ ZIP contains mastersheet files: {mastersheet_files}")
                    return zip_path, 'zip'
                else:
                    print(f"⚠️ ZIP found but no mastersheet inside: {zip_path}")
        except Exception as e:
            print(f"⚠️ Error checking ZIP contents: {e}")
    
    # Fallback to folder
    folder_path = find_latest_result_folder(clean_dir, set_name)
    if folder_path:
        print(f"✅ Using folder source: {folder_path}")
        return folder_path, 'folder'
    
    print(f"❌ No valid ZIP files or result folders found in {clean_dir}")
    print(f"📁 Available files: {os.listdir(clean_dir) if os.path.exists(clean_dir) else 'Directory not found'}")
    return None, None

def get_mastersheet_path(source_path, source_type, semester_key):
    """Get mastersheet path based on source type (zip or folder)."""
    temp_dir = None
    if source_type == 'zip':
        temp_mastersheet_path, temp_dir = extract_mastersheet_from_zip(source_path, semester_key)
        if not temp_mastersheet_path:
            print("❌ Failed to extract mastersheet from ZIP")
            return None, None
    elif source_type == 'folder':
        # Look for mastersheet in the folder
        all_files = os.listdir(source_path)
        mastersheet_files = [f for f in all_files if 'mastersheet' in f.lower() and f.endswith('.xlsx')]
        
        if not mastersheet_files:
            print(f"❌ No mastersheet found in folder {source_path}")
            return None, None
        
        mastersheet_name = mastersheet_files[0]
        temp_mastersheet_path = os.path.join(source_path, mastersheet_name)
        print(f"✅ Found mastersheet in folder: {temp_mastersheet_path}")
    else:
        return None, None
    
    return temp_mastersheet_path, temp_dir

def get_semester_display_info(semester_key):
    """Get display information for a semester key."""
    semester_lower = semester_key.lower()
    if 'first-year-first-semester' in semester_lower:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI", "Semester 1"
    elif 'first-year-second-semester' in semester_lower:
        return 1, 2, "YEAR ONE", "SECOND SEMESTER", "NDI", "Semester 2"
    elif 'second-year-first-semester' in semester_lower:
        return 2, 1, "YEAR TWO", "FIRST SEMESTER", "NDII", "Semester 3"
    elif 'second-year-second-semester' in semester_lower:
        return 2, 2, "YEAR TWO", "SECOND SEMESTER", "NDII", "Semester 4"
    else:
        return 1, 1, "YEAR ONE", "FIRST SEMESTER", "NDI", "Semester 1"

def get_grade_point(score):
    """Determine grade point based on score - NIGERIAN 5.0 SCALE."""
    try:
        score = float(score)
        if score >= 70: return 5.0
        elif score >= 60: return 4.0
        elif score >= 50: return 3.0
        elif score >= 45: return 2.0
        elif score >= 40: return 1.0
        else: return 0.0
    except (ValueError, TypeError):
        return 0.0

def get_matching_sheet(xl, target_key):
    """Find matching sheet name with variants - FIXED."""
    target_upper = target_key.upper().replace('-', ' ').replace('_', ' ').replace('.', ' ')
    target_upper = ' '.join(target_upper.split())  # Normalize spaces
    
    possible_keys = [
        target_key,
        target_key.upper(),
        target_key.lower(),
        target_key.title(),
        target_key.replace('-', ' ').upper(),
        target_key.replace('-', ' ').lower(),
        target_key.replace('-', ' ').title(),
        target_key.replace('-', '_').upper(),
        target_key.replace('-', '_').lower(),
        target_key.replace('-', '_').title(),
        target_key.replace('First', '1st'),
        target_key.replace('Second', '2nd'),
        target_key.replace('YEAR', 'YR'),
        target_key.replace('SEMESTER', 'SEM'),
        target_upper,
        target_upper.replace('FIRST', '1ST'),
        target_upper.replace('SECOND', '2ND'),
        target_upper.replace('YEAR', 'YR'),
        target_upper.replace('SEMESTER', 'SEM'),
    ]
    
    # Remove duplicates
    possible_keys = list(set([k for k in possible_keys if k]))
    
    print(f"🔍 Trying sheet variants for '{target_key}': {possible_keys}")
    
    for sheet in xl.sheet_names:
        sheet_normalized = sheet.upper().replace('-', ' ').replace('_', ' ').replace('.', ' ')
        sheet_normalized = ' '.join(sheet_normalized.split())
        
        if any(p == sheet or p in sheet or p == sheet_normalized or p in sheet_normalized for p in possible_keys):
            print(f"✅ Found matching sheet: '{sheet}' for '{target_key}'")
            return sheet
    
    print(f"❌ No matching sheet found for '{target_key}'")
    print(f"📖 Available sheets: {xl.sheet_names}")
    return None

# ENHANCED: GPA Loading for ALL semesters
def load_previous_gpas(mastersheet_path, current_semester_key):
    """Load previous GPA data from mastersheet for CGPA calculation - ENHANCED FOR ALL SEMESTERS."""
    all_student_data = {}
    current_standard = standardize_semester_key(current_semester_key)
    
    # Define all semesters based on current semester
    all_semesters = {
        "ND-First-YEAR-First-SEMESTER": ["ND-FIRST-YEAR-FIRST-SEMESTER"],
        "ND-First-YEAR-SECOND-SEMESTER": ["ND-FIRST-YEAR-FIRST-SEMESTER"],
        "ND-SECOND-YEAR-First-SEMESTER": ["ND-FIRST-YEAR-FIRST-SEMESTER", "ND-FIRST-YEAR-SECOND-SEMESTER"],
        "ND-SECOND-YEAR-SECOND-SEMESTER": ["ND-FIRST-YEAR-FIRST-SEMESTER", "ND-FIRST-YEAR-SECOND-SEMESTER", "ND-SECOND-YEAR-FIRST-SEMESTER"]
    }
    
    semesters_to_load = all_semesters.get(current_standard, [])
    print(f"📊 Loading previous GPAs for {current_standard}: {semesters_to_load}")

    if not os.path.exists(mastersheet_path):
        print(f"❌ Mastersheet not found: {mastersheet_path}")
        return {}

    try:
        xl = pd.ExcelFile(mastersheet_path)
        print(f"📖 Available sheets in mastersheet: {xl.sheet_names}")
    except Exception as e:
        print(f"❌ Error opening mastersheet: {e}")
        return {}

    for semester in semesters_to_load:
        try:
            sheet_name = get_matching_sheet(xl, semester)
            if not sheet_name:
                print(f"⚠️ Skipping semester {semester} - no matching sheet found")
                continue
            
            print(f"📖 Reading sheet '{sheet_name}' for semester {semester}")
            df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=5)
            
            # If header row 5 doesn't work, try row 0
            if df.empty or len(df.columns) < 3:
                df = pd.read_excel(mastersheet_path, sheet_name=sheet_name, header=0)
                print(f"🔄 Using header row 0 for sheet '{sheet_name}'")
            
            exam_col = find_exam_number_column(df)
            gpa_col = None
            credit_col = None
            
            # Find GPA and credit columns with flexible matching
            for col in df.columns:
                col_str = str(col).upper()
                if 'GPA' in col_str and 'CGPA' not in col_str:
                    gpa_col = col
                if 'CU PASSED' in col_str or 'CREDIT' in col_str or 'UNIT' in col_str:
                    credit_col = col
            
            print(f"🔍 Columns found - Exam: {exam_col}, GPA: {gpa_col}, Credits: {credit_col}")
            
            if exam_col and gpa_col:
                for idx, row in df.iterrows():
                    try:
                        exam_no = str(row[exam_col]).strip()
                        if pd.isna(exam_no) or exam_no in ['', 'NAN', 'NONE']:
                            continue
                            
                        gpa_value = row[gpa_col]
                        if pd.isna(gpa_value):
                            continue
                            
                        # Get credits passed - default to 30 if not found
                        credits = 30
                        if credit_col and credit_col in row and pd.notna(row[credit_col]):
                            try:
                                credits = int(float(row[credit_col]))
                            except (ValueError, TypeError):
                                credits = 30
                        
                        if exam_no not in all_student_data:
                            all_student_data[exam_no] = {'gpas': [], 'credits': [], 'semesters': []}
                        
                        all_student_data[exam_no]['gpas'].append(float(gpa_value))
                        all_student_data[exam_no]['credits'].append(credits)
                        all_student_data[exam_no]['semesters'].append(semester)
                        
                        if idx < 3:  # Log first few for debugging
                            print(f"📊 Loaded GPA for {exam_no}: {gpa_value} with {credits} credits from {semester}")
                            
                    except (ValueError, TypeError) as e:
                        print(f"⚠️ Error processing row {idx} for {semester}: {e}")
                        continue
            else:
                print(f"⚠️ Missing required columns in {sheet_name}: exam_col={exam_col}, gpa_col={gpa_col}")
                
        except Exception as e:
            print(f"⚠️ Could not load data from {semester}: {e}")
            traceback.print_exc()
    
    print(f"📊 Loaded cumulative data for {len(all_student_data)} students")
    return all_student_data

def calculate_cgpa(student_data, current_gpa, current_credits):
    """Calculate Cumulative GPA - ENHANCED FOR ALL SEMESTERS."""
    if not student_data or not student_data.get('gpas'):
        print(f"⚠️ No previous GPA data, using current GPA: {current_gpa}")
        return current_gpa

    total_grade_points = 0.0
    total_credits = 0

    print(f"🔢 Calculating CGPA from {len(student_data['gpas'])} previous semesters")
    
    for prev_gpa, prev_credits in zip(student_data['gpas'], student_data['credits']):
        total_grade_points += prev_gpa * prev_credits
        total_credits += prev_credits
        print(f"   - GPA: {prev_gpa}, Credits: {prev_credits}, Running Total: {total_grade_points}/{total_credits}")

    # Add current semester
    total_grade_points += current_gpa * current_credits
    total_credits += current_credits

    print(f"📊 Final calculation: {total_grade_points} / {total_credits}")

    if total_credits > 0:
        cgpa = round(total_grade_points / total_credits, 2)
        print(f"✅ Calculated CGPA: {cgpa}")
        return cgpa
    else:
        print(f"⚠️ No credits, returning current GPA: {current_gpa}")
        return current_gpa

def get_previous_semesters_for_display(current_semester_key):
    """Get list of previous semesters for GPA display in mastersheet."""
    current_standard = standardize_semester_key(current_semester_key)
    
    semester_mapping = {
        "ND-First-YEAR-First-SEMESTER": [],
        "ND-First-YEAR-SECOND-SEMESTER": ["Semester 1"],
        "ND-SECOND-YEAR-First-SEMESTER": ["Semester 1", "Semester 2"], 
        "ND-SECOND-YEAR-SECOND-SEMESTER": ["Semester 1", "Semester 2", "Semester 3"]
    }
    
    return semester_mapping.get(current_standard, [])

def extract_semester_from_filename(filename):
    """Extract semester from filename using comprehensive pattern matching - FIXED."""
    filename_upper = filename.upper()
    
    # First, try to extract any semester-like pattern from filename
    semester_pattern = r'(ND[-_]?(?:FIRST|SECOND|1ST|2ND)[-_]?YEAR[-_]?(?:FIRST|SECOND|1ST|2ND)[-_]?SEMESTER)'
    match = re.search(semester_pattern, filename_upper)
    
    if match:
        extracted = match.group(1)
        # Standardize the extracted pattern
        standardized = standardize_semester_key(extracted)
        print(f"✅ Extracted and standardized: '{filename}' → '{standardized}'")
        return standardized
    
    # Fallback to comprehensive pattern mapping (existing code)
    semester_patterns = {
        "ND-First-YEAR-First-SEMESTER": [
            "FIRST.YEAR.FIRST.SEMESTER", "FIRST-YEAR-FIRST-SEMESTER",
        ],
        "ND-First-YEAR-SECOND-SEMESTER": [
            "FIRST.YEAR.SECOND.SEMESTER", "FIRST-YEAR-SECOND-SEMESTER",
        ],
        "ND-SECOND-YEAR-First-SEMESTER": [
            "SECOND.YEAR.FIRST.SEMESTER", "SECOND-YEAR-FIRST-SEMESTER",
        ],
        "ND-SECOND-YEAR-SECOND-SEMESTER": [
            "SECOND.YEAR.SECOND.SEMESTER", "SECOND-YEAR-SECOND-SEMESTER",
        ]
    }
    
    for semester_key, patterns in semester_patterns.items():
        for pattern in patterns:
            flexible_pattern = pattern.replace('.', '[._\\- ]?')
            if re.search(flexible_pattern, filename_upper, re.IGNORECASE):
                print(f"✅ Matched semester '{semester_key}' for filename: {filename}")
                return semester_key
    
    print(f"❌ Could not determine semester for filename: {filename}")
    return "UNKNOWN_SEMESTER"

def load_carryover_json_files(carryover_dir, semester_key=None):
    """Load carryover JSON files from directory - FIXED."""
    carryover_files = []
    
    # Standardize the target semester key
    if semester_key:
        semester_key = standardize_semester_key(semester_key)
    
    previous_semester = get_previous_semester(semester_key)
    print(f"🔑 Target semester: {semester_key}")
    print(f"🔑 Previous semester for carryover: {previous_semester}")
    
    for file in os.listdir(carryover_dir):
        if file.startswith("co_student_") and file.endswith(".json"):
            # Extract semester from filename and standardize it
            file_semester = extract_semester_from_filename(file)
            file_semester_standardized = standardize_semester_key(file_semester)
            
            print(f"📄 Found carryover file: {file}")
            print(f"   Original semester: {file_semester}")
            print(f"   Standardized: {file_semester_standardized}")
            print(f"   Target previous: {previous_semester}")
            
            # If previous_semester is specified, only load matching files (carryover from previous)
            if previous_semester and file_semester_standardized != previous_semester:
                print(f"   ⏭️ Skipping (doesn't match previous semester)")
                continue
            
            file_path = os.path.join(carryover_dir, file)
            try:
                with open(file_path, 'r') as f:
                    data = json.load(f)
                    carryover_files.append({
                        'filename': file,
                        'semester': file_semester_standardized,  # Use standardized key
                        'data': data,
                        'count': len(data),
                        'file_path': file_path
                    })
                    print(f"   ✅ Loaded: {len(data)} records")
            except Exception as e:
                print(f"Error loading {file}: {e}")
    
    print(f"📊 Total carryover files loaded: {len(carryover_files)}")
    return carryover_files

def get_carryover_records_from_zip(zip_path, set_name, semester_key):
    """Get carryover records from ZIP file."""
    # This function is not defined in the provided code, assuming it extracts the ZIP and calls load_carryover_json_files
    # For completeness, implement it here
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            temp_dir = tempfile.mkdtemp()
            carryover_dir = os.path.join(temp_dir, "CARRYOVER_RECORDS")
            os.makedirs(carryover_dir, exist_ok=True)
            for member in zip_ref.namelist():
                if member.startswith("CARRYOVER_RECORDS/"):
                    zip_ref.extract(member, temp_dir)
            records = load_carryover_json_files(carryover_dir, semester_key)
            shutil.rmtree(temp_dir)
            print(f"✅ Loaded {len(records)} carryover records from ZIP")
            return records
    except Exception as e:
        print(f"❌ Error loading from ZIP: {e}")
        return []

def get_carryover_records(program, set_name, semester_key=None):
    """Get carryover records for a specific program, set, and semester - FIXED."""
    try:
        # Standardize semester key first
        if semester_key:
            semester_key = standardize_semester_key(semester_key)
            print(f"🔑 Using standardized semester key: {semester_key}")
        
        clean_dir = os.path.join(BASE_DIR, program, set_name, "CLEAN_RESULTS")
        if not os.path.exists(clean_dir):
            print(f"❌ Clean directory not found: {clean_dir}")
            return []
        
        # Look for both folders and ZIP files (REGULAR results, not carryover)
        timestamp_items = []
        
        for item in os.listdir(clean_dir):
            item_path = os.path.join(clean_dir, item)
            
            # ONLY include regular result files (exclude carryover files)
            if item.startswith(f"{set_name}_RESULT-") and "CARRYOVER" not in item.upper():
                if os.path.isdir(item_path) or item.endswith('.zip'):
                    timestamp_items.append(item)
                    print(f"Found regular result: {item}")
        
        if not timestamp_items:
            print(f"❌ No regular result files found in: {clean_dir}")
            return []
        
        latest_item = sorted(timestamp_items)[-1]
        latest_path = os.path.join(clean_dir, latest_item)
        print(f"✅ Using latest result: {latest_item}")
        
        # Extract from ZIP or use folder
        if latest_item.endswith('.zip'):
            return get_carryover_records_from_zip(latest_path, set_name, semester_key)
        else:
            carryover_dir = os.path.join(latest_path, "CARRYOVER_RECORDS")
            if not os.path.exists(carryover_dir):
                print(f"❌ No CARRYOVER_RECORDS folder in: {latest_path}")
                return []
            return load_carryover_json_files(carryover_dir, semester_key)
            
    except Exception as e:
        print(f"Error getting carryover records: {e}")
        return []

# FIX 3: Enhanced Semester Key Matching with DEBUGGING
def process_carryover_results(resit_file_path, source_path, source_type, semester_key, set_name, pass_threshold, output_dir):
    """
    Process carryover results and generate CARRYOVER_mastersheet.
    """
    print(f"\n🔄 PROCESSING CARRYOVER RESULTS FOR {semester_key}")
    print("=" * 60)
    
    # Load course data
    semester_course_titles, semester_credit_units, course_code_to_title, course_code_to_unit = load_course_data()
    
    # DEBUG: Check course matching specifically
    debug_course_matching(resit_file_path, course_code_to_title, course_code_to_unit)
    
    # Get display info and try multiple sheet key formats
    year, sem_num, level, sem_display, set_code, sem_name = get_semester_display_info(semester_key)
    
    # Try multiple sheet key formats for robust matching
    possible_sheet_keys = [
        f"{set_code} {sem_display}",                    # "NDII SECOND SEMESTER"
        f"{set_code.replace('NDII', 'ND II').replace('NDI', 'ND I')} {sem_display}",  # "ND II SECOND SEMESTER"
        semester_key,                                   # Original key
        semester_key.replace('-', ' ').upper(),         # "ND SECOND YEAR SECOND SEMESTER"
        f"{set_code} {sem_name}",                       # "NDII Semester 4"
        f"{level} {sem_display}",                       # "YEAR TWO SECOND SEMESTER"
    ]
    
    course_titles_dict = {}
    credit_units_dict = {}
    
    for sheet_key in possible_sheet_keys:
        sheet_standard = standardize_semester_key(sheet_key)
        if sheet_standard in semester_course_titles:
            course_titles_dict = semester_course_titles[sheet_standard]
            credit_units_dict = semester_credit_units[sheet_standard]
            print(f"✅ Using sheet key: '{sheet_key}' with {len(course_titles_dict)} courses")
            break
        else:
            print(f"❌ Sheet key not found: '{sheet_key}'")
    
    # If no semester-specific dict found, use global dictionaries
    if not course_titles_dict:
        print("⚠️ No semester-specific course data found, using global course mappings")
        course_titles_dict = course_code_to_title
        credit_units_dict = course_code_to_unit
    
    print(f"📊 Final course mappings: {len(course_titles_dict)} titles, {len(credit_units_dict)} units")
    
    # Create output directory
    timestamp = datetime.now().strftime(TIMESTAMP_FMT)
    carryover_output_dir = os.path.join(output_dir, f"CARRYOVER_{set_name}_{semester_key}_{timestamp}")
    os.makedirs(carryover_output_dir, exist_ok=True)
    
    if not os.path.exists(resit_file_path):
        print(f"❌ Resit file not found: {resit_file_path}")
        return False
    
    temp_mastersheet_path = None
    temp_dir = None
    
    try:
        # Get mastersheet path based on source type
        temp_mastersheet_path, temp_dir = get_mastersheet_path(source_path, source_type, semester_key)
        
        if not temp_mastersheet_path:
            print("❌ Failed to get mastersheet")
            return False
        
        # Read files
        print("📖 Reading files...")
        resit_df = pd.read_excel(resit_file_path, header=0)
        
        # FIXED: Try multiple header positions for mastersheet
        xl = pd.ExcelFile(temp_mastersheet_path)
        sheet_name = get_matching_sheet(xl, semester_key)
        if not sheet_name:
            print(f"❌ No matching sheet found for {semester_key}")
            return False
        
        print(f"📖 Using sheet '{sheet_name}' for current semester {semester_key}")
        
        try:
            mastersheet_df = pd.read_excel(temp_mastersheet_path, sheet_name=sheet_name, header=5)
        except:
            try:
                mastersheet_df = pd.read_excel(temp_mastersheet_path, sheet_name=sheet_name, header=0)
                print("⚠️ Using header row 0 for mastersheet")
            except Exception as e:
                print(f"❌ Error reading mastersheet: {e}")
                return False
        
        print(f"✅ Files loaded - Resit: {len(resit_df)} rows, Mastersheet: {len(mastersheet_df)} students")
        
        # Find exam number columns
        resit_exam_col = find_exam_number_column(resit_df)
        mastersheet_exam_col = find_exam_number_column(mastersheet_df) or 'EXAM NUMBER'
        
        if not resit_exam_col:
            print("❌ Cannot find exam number column in resit file")
            return False
        
        print(f"📝 Exam columns - Resit: '{resit_exam_col}', Mastersheet: '{mastersheet_exam_col}'")
        
        # Load previous GPAs for CGPA calculation - ENHANCED FOR ALL SEMESTERS
        cgpa_data = load_previous_gpas(temp_mastersheet_path, semester_key)
        
        # Create carryover mastersheet data structure
        carryover_data = []
        updated_students = set()
        
        print(f"\n🎯 PROCESSING RESIT SCORES...")
        
        for idx, resit_row in resit_df.iterrows():
            exam_no = str(resit_row[resit_exam_col]).strip().upper()
            if not exam_no or exam_no in ['NAN', 'NONE', '']:
                continue
            
            # Find student in mastersheet
            student_mask = mastersheet_df[mastersheet_exam_col].astype(str).str.strip().str.upper() == exam_no
            if not student_mask.any():
                print(f"⚠️ Student {exam_no} not found in mastersheet - skipping")
                continue
            
            student_data = mastersheet_df[student_mask].iloc[0]
            student_name = student_data.get('NAME', 'Unknown')
            
            # Get current credits passed for CGPA calculation
            current_credits = 0
            for col in mastersheet_df.columns:
                if 'CU PASSED' in str(col).upper():
                    current_credits = student_data.get(col, 0)
                    break
            
            # Initialize student record for carryover mastersheet
            student_record = {
                'EXAM NUMBER': exam_no,
                'NAME': student_name,
                'RESIT_COURSES': {},
                'CURRENT_GPA': student_data.get('GPA', 0),
                'CURRENT_CREDITS': current_credits
            }
            
            # Calculate CGPA properly - ENHANCED FOR ALL SEMESTERS
            if exam_no in cgpa_data:
                student_record['CURRENT_CGPA'] = calculate_cgpa(
                    cgpa_data[exam_no], 
                    student_record['CURRENT_GPA'], 
                    current_credits
                )
            else:
                student_record['CURRENT_CGPA'] = student_record['CURRENT_GPA']
            
            # Add previous GPAs for ALL semesters (not just Semester 4)
            if exam_no in cgpa_data:
                student_gpa_data = cgpa_data[exam_no]
                for i, prev_semester in enumerate(student_gpa_data['semesters']):
                    # Get semester display name (Semester 1, Semester 2, etc.)
                    sem_display_name = get_semester_display_info(prev_semester)[5]
                    student_record[f'GPA_{sem_display_name}'] = student_gpa_data['gpas'][i]
                    print(f"📊 Stored GPA for {exam_no}: {sem_display_name} = {student_gpa_data['gpas'][i]}")
            
            # Process each course in resit file
            for col in resit_df.columns:
                if col == resit_exam_col or col == 'NAME' or 'Unnamed' in str(col):
                    continue
                    
                resit_score = resit_row.get(col)
                if pd.isna(resit_score) or resit_score == '':
                    continue
                
                try:
                    resit_score_val = float(resit_score)
                except (ValueError, TypeError):
                    continue
                
                # Check if this course column exists in mastersheet
                if col in mastersheet_df.columns:
                    original_score = student_data.get(col)
                    if pd.isna(original_score):
                        continue
                    
                    try:
                        original_score_val = float(original_score) if not pd.isna(original_score) else 0.0
                    except (ValueError, TypeError):
                        original_score_val = 0.0
                    
                    # Only include courses that were re-sat (failed originally and now resat)
                    if original_score_val < pass_threshold:
                        # Get course title using robust matching
                        course_title = find_course_title(col, course_titles_dict, course_code_to_title)
                        # Get credit unit using robust matching
                        credit_unit = find_credit_unit(col, credit_units_dict, course_code_to_unit)
                        
                        student_record['RESIT_COURSES'][col] = {
                            'original_score': original_score_val,
                            'resit_score': resit_score_val,
                            'updated': resit_score_val >= pass_threshold,
                            'course_title': course_title,
                            'credit_unit': credit_unit
                        }
            
            # Only add student to carryover mastersheet if they have resit courses
            if student_record['RESIT_COURSES']:
                carryover_data.append(student_record)
                updated_students.add(exam_no)
                print(f"✅ {exam_no}: {len(student_record['RESIT_COURSES'])} resit courses, CGPA: {student_record['CURRENT_CGPA']}")
        
        # Generate CARRYOVER_mastersheet
        if carryover_data:
            print(f"\n📊 GENERATING CARRYOVER MASTERSHEET...")
            carryover_mastersheet_path = generate_carryover_mastersheet(
                carryover_data, carryover_output_dir, semester_key, set_name, timestamp, 
                cgpa_data, course_titles_dict, credit_units_dict, course_code_to_title, course_code_to_unit
            )
            
            # Generate individual student reports
            print(f"\n📄 GENERATING INDIVIDUAL STUDENT REPORTS...")
            generate_individual_reports(
                carryover_data, carryover_output_dir, semester_key, set_name, timestamp, cgpa_data
            )
            
            # Create final ZIP
            zip_path = os.path.join(output_dir, f"CARRYOVER_{set_name}_{semester_key}_{timestamp}.zip")
            if create_carryover_zip(carryover_output_dir, zip_path):
                print(f"✅ Final carryover ZIP created: {zip_path}")
            
            print(f"\n🎉 CARRYOVER PROCESSING COMPLETED!")
            print(f"📁 Output directory: {carryover_output_dir}")
            print(f"📦 ZIP file: {zip_path}")
            print(f"👨‍🎓 Students processed: {len(carryover_data)}")
            
            return True
        else:
            print("❌ No carryover data found to process")
            return False
            
    except Exception as e:
        print(f"❌ Error processing carryover results: {e}")
        traceback.print_exc()
        return False
    finally:
        # Clean up temporary files
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print("✅ Cleaned up temporary files")

def generate_carryover_mastersheet(carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data, course_titles, course_units, course_code_to_title, course_code_to_unit):
    """Generate the CARRYOVER_mastersheet with enhanced GPA tracking for ALL semesters."""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CARRYOVER_RESULTS"
    
    # Add logo if available
    if os.path.exists(DEFAULT_LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image
            img = Image(DEFAULT_LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"⚠️ Could not add logo: {e}")
    
    # Title and headers - UPDATED with dynamic GPA columns
    current_year = 2025
    next_year = 2026
    year, sem_num, level, sem_display, set_code, current_semester_name = get_semester_display_info(semester_key)
    
    # Calculate total columns needed for merging
    all_courses = set()
    for student in carryover_data:
        all_courses.update(student['RESIT_COURSES'].keys())
    
    # Get previous semesters for GPA display - DYNAMIC FOR ALL SEMESTERS
    previous_semesters = get_previous_semesters_for_display(semester_key)
    
    # Build headers structure with dynamic GPA columns
    headers = ['S/N', 'EXAM NUMBER', 'NAME']
    
    # Add previous GPA columns dynamically
    for prev_sem in previous_semesters:
        headers.append(f'GPA {prev_sem}')
    
    course_headers = []
    for course in sorted(all_courses):
        course_headers.extend([f'{course}', f'{course}_RESIT'])
    
    headers.extend(course_headers)
    headers.extend([f'GPA {current_semester_name}', 'CGPA', 'REMARKS'])
    
    total_columns = len(headers)
    last_column = get_column_letter(total_columns)
    
    # CENTRALIZED TITLE ROWS
    ws.merge_cells(f'A3:{last_column}3')
    title_cell = ws['A3']
    title_cell.value = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA-ABUJA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(f'A4:{last_column}4')
    subtitle_cell = ws['A4']
    subtitle_cell.value = f"RESIT - {current_year}/{next_year} SESSION NATIONAL DIPLOMA {level} {sem_display} EXAMINATIONS RESULT — October 28, 2025"
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"🔍 Courses found in resit data: {sorted(all_courses)}")
    print(f"📊 GPA columns for {semester_key}: Previous={previous_semesters}, Current={current_semester_name}")
    
    # Build headers structure with course titles, codes, and credit units
    print(f"🔍 Courses found in resit data: {sorted(all_courses)}")
    
    # Build headers structure with course titles, codes, and credit units
    headers = ['S/N', 'EXAM NUMBER', 'NAME']
    
    # Add previous GPA columns dynamically
    for prev_sem in previous_semesters:
        headers.append(f'GPA {prev_sem}')
    
    # Add course columns with titles, codes, and credit units
    course_headers = []
    course_title_mapping = {}  # Store the actual titles we find
    course_unit_mapping = {}   # Store the credit units
    
    for course in sorted(all_courses):
        # Use robust title lookup
        course_title = find_course_title(course, course_titles, course_code_to_title)
        course_title_mapping[course] = course_title
        
        # Get credit unit using robust matching
        credit_unit = find_credit_unit(course, course_units, course_code_to_unit)
        course_unit_mapping[course] = credit_unit
        
        # Truncate long course titles for display
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        course_headers.extend([f'{course}', f'{course}_RESIT'])
    
    headers.extend(course_headers)
    headers.extend([f'GPA {current_semester_name}', 'CGPA', 'REMARKS'])
    
    # Write course titles row (row 5) with counterclockwise orientation
    title_row = [''] * 3  # S/N, EXAM NUMBER, NAME
    
    # Add previous GPA placeholders
    for prev_sem in previous_semesters:
        title_row.extend([''])  # GPA placeholders
    
    # Add course titles with counterclockwise orientation
    for course in sorted(all_courses):
        course_title = course_title_mapping[course]
        if len(course_title) > 30:
            course_title = course_title[:27] + "..."
        title_row.extend([course_title, course_title])  # Use title for both original and resit columns
    
    title_row.extend(['', '', ''])  # GPA Current, CGPA, REMARKS
    
    ws.append(title_row)  # This is row 5
    
    # Write credit units row (row 6)
    credit_row = [''] * 3  # S/N, EXAM NUMBER, NAME
    
    # Add previous GPA placeholders
    for prev_sem in previous_semesters:
        credit_row.extend([''])  # GPA placeholders
    
    # Add credit units for each course
    for course in sorted(all_courses):
        credit_unit = course_unit_mapping[course]
        credit_row.extend([f'CU: {credit_unit}', f'CU: {credit_unit}'])  # Credit unit for both original and resit columns
    
    credit_row.extend(['', '', ''])  # GPA Current, CGPA, REMARKS
    
    ws.append(credit_row)  # This is row 6
    
    # Write course codes row (row 7)
    code_row = ['S/N', 'EXAM NUMBER', 'NAME']
    
    # Add previous GPA headers
    for prev_sem in previous_semesters:
        code_row.append(f'GPA {prev_sem}')
    
    # Add course codes
    for course in sorted(all_courses):
        code_row.extend([f'{course}', f'{course}_RESIT'])
    
    code_row.extend([f'GPA {current_semester_name}', 'CGPA', 'REMARKS'])
    
    ws.append(code_row)  # This is row 7
    
    # Define print-friendly colors for course title columns (light pastel colors)
    course_colors = [
        "E6F3FF",  # Light blue
        "FFF0E6",  # Light orange
        "E6FFE6",  # Light green
        "FFF6E6",  # Light peach
        "F0E6FF",  # Light purple
        "E6FFFF",  # Light cyan
        "FFE6F2",  # Light pink
        "F5F5DC",  # Light beige
        "E6F7FF",  # Light sky blue
        "FFF5E6",  # Light apricot
    ]
    
    # Apply colors to course columns in all header rows (5, 6, 7)
    start_col = 4  # Start after S/N, EXAM NUMBER, NAME
    if previous_semesters:  # Skip GPA columns if they exist
        start_col += len(previous_semesters)
    
    color_index = 0
    for course in sorted(all_courses):
        # Apply colors to all three header rows for this course pair
        for row in [5, 6, 7]:  # CHANGED: Now rows 5, 6, 7
            for offset in [0, 1]:  # Original and resit columns
                cell = ws.cell(row=row, column=start_col + offset)
                cell.fill = PatternFill(start_color=course_colors[color_index % len(course_colors)], 
                                      end_color=course_colors[color_index % len(course_colors)], 
                                      fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        # Apply rotation only to course titles row (row 5)
        for offset in [0, 1]:
            cell = ws.cell(row=5, column=start_col + offset)  # CHANGED: Row 5 for course titles
            cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=9)
        
        color_index += 1
        start_col += 2  # Move to next course pair
    
    # Style the non-course header columns (S/N, EXAM NUMBER, NAME, GPA columns)
    for row in [5, 6, 7]:  # CHANGED: Rows 5, 6, 7
        for col in range(1, 4):  # S/N, EXAM NUMBER, NAME
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Style GPA columns if they exist
        gpa_col = 4
        for prev_sem in previous_semesters:
            cell = ws.cell(row=row, column=gpa_col)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            gpa_col += 1
        
        # Style final GPA columns
        for col in range(len(headers)-2, len(headers)+1):  # GPA Current, CGPA, REMARKS
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Write data starting from row 8 (after headers)
    row_idx = 8  # CHANGED: Data starts at row 8 now
    failed_counts = {course: 0 for course in all_courses}
    
    # Apply colors to data rows for course columns
    start_col = 4  # Reset start column
    if previous_semesters:  # Skip GPA columns if they exist
        start_col += len(previous_semesters)
    
    for student in carryover_data:
        exam_no = student['EXAM NUMBER']
        
        # Basic info
        ws.cell(row=row_idx, column=1, value=row_idx-7)  # S/N (adjusted for new row)
        ws.cell(row=row_idx, column=2, value=student['EXAM NUMBER'])
        ws.cell(row=row_idx, column=3, value=student['NAME'])
        
        # Previous GPAs - DYNAMIC FOR ALL SEMESTERS
        gpa_col = 4
        for prev_sem in previous_semesters:
            gpa_value = student.get(f'GPA_{prev_sem}', '')
            ws.cell(row=row_idx, column=gpa_col, value=gpa_value)
            gpa_col += 1
        
        # Course scores - APPLY COLORS TO DATA ROWS
        course_col = gpa_col
        color_index = 0
        for course in sorted(all_courses):
            # Apply the same alternating colors to data cells
            for offset in [0, 1]:
                cell = ws.cell(row=row_idx, column=course_col + offset)
                cell.fill = PatternFill(start_color=course_colors[color_index % len(course_colors)], 
                                      end_color=course_colors[color_index % len(course_colors)], 
                                      fill_type="solid")
            
            if course in student['RESIT_COURSES']:
                course_data = student['RESIT_COURSES'][course]
                
                # Original score (color red if failed)
                orig_cell = ws.cell(row=row_idx, column=course_col, value=course_data['original_score'])
                if course_data['original_score'] < DEFAULT_PASS_THRESHOLD:
                    orig_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                # Resit score (color green if passed, red if failed)
                resit_cell = ws.cell(row=row_idx, column=course_col+1, value=course_data['resit_score'])
                if course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD:
                    resit_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    resit_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    failed_counts[course] += 1
            else:
                ws.cell(row=row_idx, column=course_col, value='')
                ws.cell(row=row_idx, column=course_col+1, value='')
            
            color_index += 1
            course_col += 2
        
        # Current GPA and CGPA
        ws.cell(row=row_idx, column=course_col, value=student['CURRENT_GPA'])
        ws.cell(row=row_idx, column=course_col+1, value=student['CURRENT_CGPA'])
        
        # Remarks
        remarks = generate_remarks(student['RESIT_COURSES'])
        ws.cell(row=row_idx, column=course_col+2, value=remarks)
        
        row_idx += 1
    
    # Add failed count summary - MOVED TO THE EMPTY ROW IMMEDIATELY AFTER DATA
    failed_row_idx = row_idx  # Use the current row (empty row after data)
    ws.cell(row=failed_row_idx, column=1, value="FAILED COUNT BY COURSE:").font = Font(bold=True)
    
    # Apply color to failed count row - LIGHT YELLOW BACKGROUND
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=failed_row_idx, column=col)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Write failed counts under each course's RESIT column
    course_col = gpa_col  # Start at first course column
    for course in sorted(all_courses):
        # Write the failed count in the RESIT column (course_col + 1)
        count_cell = ws.cell(row=failed_row_idx, column=course_col+1, value=failed_counts[course])
        count_cell.font = Font(bold=True)
        count_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        course_col += 2
    
    # Add main summary section - LEFT ALIGNED
    summary_start_row = failed_row_idx + 2  # One empty row after failed count
    
    # Calculate summary statistics
    total_students = len(carryover_data)
    passed_all = sum(1 for student in carryover_data 
                    if all(course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD 
                          for course_data in student['RESIT_COURSES'].values()))
    
    carryover_count = total_students - passed_all
    total_failed_attempts = sum(failed_counts.values())
    
    # LEFT-ALIGNED SUMMARY DATA
    summary_data = [
        ["SUMMARY"],
        [f"A total of {total_students} students registered and sat for the Carryover Examination"],
        [f"A total of {passed_all} students passed all carryover courses"],
        [f"A total of {carryover_count} students failed one or more carryover courses and must repeat them"],
        [f"Total failed resit attempts: {total_failed_attempts} across all courses"],
        [f"Carryover processing completed on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"],
        [""],  # Empty row for spacing
        [""],  # Another empty row
        ["", ""],  # Signatories will be placed in separate columns
        ["________________________", "________________________"],
        ["Mrs. Abini Hauwa", "Mrs. Olukemi Ogunleye"],
        ["Head of Exams", "Chairman, ND/HND Program C'tee"]
    ]
    
    for i, row_data in enumerate(summary_data):
        row_num = summary_start_row + i
        if len(row_data) == 1:
            if row_data[0]:  # Only merge if there's actual content
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
                cell = ws.cell(row=row_num, column=1, value=row_data[0])
                if i == 0:  # "SUMMARY" header
                    cell.font = Font(bold=True, size=12, underline='single')
                else:
                    cell.font = Font(bold=False, size=11)
                # LEFT ALIGNMENT for summary text
                cell.alignment = Alignment(horizontal='left', vertical='center')
        elif len(row_data) == 2:
            # MOVED SIGNATORIES FURTHER LEFT - aligned with summary
            left_cell = ws.cell(row=row_num, column=1, value=row_data[0])
            right_cell = ws.cell(row=row_num, column=4, value=row_data[1])
            
            # Style signatory rows
            if i >= len(summary_data) - 3:  # Last 3 rows are signatories
                left_cell.alignment = Alignment(horizontal='left')
                right_cell.alignment = Alignment(horizontal='left')
                left_cell.font = Font(bold=True, size=11)
                right_cell.font = Font(bold=True, size=11)
    
    # Apply borders to data area
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=7, max_row=row_idx-1, min_col=1, max_col=len(headers)):  # CHANGED: min_row=7
        for cell in row:
            cell.border = thin_border
    
    # Professional formatting
    ws.freeze_panes = 'D8'  # CHANGED: Freeze at row 8 (data start)
    
    # Set professional font for entire worksheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or not cell.font.bold:
                cell.font = Font(name='Calibri', size=11)
    
    # IMPROVED AUTO-ADJUST COLUMN WIDTHS - PROPERLY FITS LONGEST TEXT
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in the column to find the longest content
        for cell in column:
            try:
                if cell.value is not None:
                    # Convert to string and calculate length
                    cell_value = str(cell.value)
                    cell_length = len(cell_value)
                    
                    # For rotated text in row 5, we need to handle differently
                    if cell.row == 5 and cell.alignment.text_rotation == 90:
                        # For rotated text, we want wider columns to accommodate the text
                        cell_length = max(cell_length, 10)  # Minimum width for rotated text
                    
                    # Adjust for numeric values (scores, GPAs)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell_length = max(cell_length, 8)  # Ensure enough space for numbers
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width based on content with reasonable limits
        adjusted_width = min(max_length + 2, 50)  # Add padding, cap at 50
        
        # Apply specific adjustments for different column types
        if col_idx == 1:  # S/N
            adjusted_width = 8
        elif col_idx == 2:  # EXAM NUMBER
            adjusted_width = 18
        elif col_idx == 3:  # NAME
            adjusted_width = 35  # Generous space for full names
        elif col_idx >= 4 and col_idx <= (4 + len(previous_semesters) - 1):
            # GPA columns
            adjusted_width = 15
        elif col_idx >= len(headers) - 2:  # GPA Current, CGPA, REMARKS
            adjusted_width = 15
        else:
            # Course columns - ensure they're wide enough for content
            adjusted_width = min(max(adjusted_width, 12), 25)  # Course columns between 12-25 width
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply alternating row colors for better readability
    for row_idx in range(8, row_idx):  # Data rows only (starting from row 8)
        if row_idx % 2 == 0:  # Even rows
            for cell in ws[row_idx]:
                # Only apply if no special fill (like course colors or pass/fail colors)
                if (cell.fill.start_color.index == '00000000' or 
                    cell.fill.start_color.index == '00FFFFFF'):
                    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    
    # Color code GPA columns in data area for better distinction
    gpa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light purple
    if previous_semesters:
        for row in range(8, row_idx):  # CHANGED: Starting from row 8
            for col in range(4, 4 + len(previous_semesters)):  # Previous GPA columns
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.index == '00000000':
                    cell.fill = gpa_fill
    
    # Color code final GPA columns in data area
    final_gpa_fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light cyan
    for row in range(8, row_idx):  # CHANGED: Starting from row 8
        for col in range(len(headers)-2, len(headers)+1):  # GPA Current, CGPA, REMARKS
            cell = ws.cell(row=row, column=col)
            if cell.fill.start_color.index == '00000000':
                cell.fill = final_gpa_fill
    
    # Save file
    filename = f"CARRYOVER_mastersheet_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    print(f"✅ CARRYOVER mastersheet generated: {filepath}")
    print(f"📊 Course title mapping used: {course_title_mapping}")
    print(f"📊 Credit units used: {course_unit_mapping}")
    print(f"🎨 Applied color coding: Course title row (row 5) with pastel colors")
    return filepath

def generate_remarks(resit_courses):
    """Generate remarks based on resit performance."""
    passed_count = sum(1 for course_data in resit_courses.values() 
                      if course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD)
    total_count = len(resit_courses)
    
    if passed_count == total_count:
        return "All courses passed in resit"
    elif passed_count > 0:
        return f"{passed_count}/{total_count} courses passed in resit"
    else:
        return "No improvement in resit"

def generate_individual_reports(carryover_data, output_dir, semester_key, set_name, timestamp, cgpa_data):
    """Generate individual student reports in CSV format."""
    reports_dir = os.path.join(output_dir, "INDIVIDUAL_REPORTS")
    os.makedirs(reports_dir, exist_ok=True)
    
    for student in carryover_data:
        exam_no = student['EXAM NUMBER']
        # Sanitize the exam number for filename safety
        safe_exam_no = sanitize_filename(exam_no)
        filename = f"carryover_report_{safe_exam_no}_{timestamp}.csv"
        filepath = os.path.join(reports_dir, filename)
        
        report_data = []
        report_data.append(["CARRYOVER RESULT REPORT"])
        report_data.append(["FCT COLLEGE OF NURSING SCIENCES"])
        report_data.append([f"Set: {set_name}"])
        report_data.append([f"Semester: {semester_key}"])
        report_data.append([])
        report_data.append(["STUDENT INFORMATION"])
        report_data.append(["Exam Number:", student['EXAM NUMBER']])
        report_data.append(["Name:", student['NAME']])
        report_data.append([])
        
        # Previous GPAs - ENHANCED FOR ALL SEMESTERS
        report_data.append(["PREVIOUS GPAs"])
        for key in sorted([k for k in student.keys() if k.startswith('GPA_')]):
            semester = key.replace('GPA_', '')
            report_data.append([f"{semester}:", student[key]])
        report_data.append([])
        
        # Current GPA and CGPA
        report_data.append(["CURRENT ACADEMIC RECORD"])
        report_data.append(["Current GPA:", student['CURRENT_GPA']])
        report_data.append(["Current CGPA:", student['CURRENT_CGPA']])
        report_data.append([])
        
        # Resit courses
        report_data.append(["RESIT COURSES"])
        report_data.append(["Course Code", "Course Title", "Credit Unit", "Original Score", "Resit Score", "Status"])
        
        for course_code, course_data in student['RESIT_COURSES'].items():
            status = "PASSED" if course_data['resit_score'] >= DEFAULT_PASS_THRESHOLD else "FAILED"
            course_title = course_data.get('course_title', course_code)
            credit_unit = course_data.get('credit_unit', 0)
            report_data.append([
                course_code, 
                course_title,
                credit_unit,
                course_data['original_score'], 
                course_data['resit_score'], 
                status
            ])
        
        # Save CSV
        try:
            df = pd.DataFrame(report_data)
            df.to_csv(filepath, index=False, header=False)
            print(f"✅ Generated report for: {exam_no}")
        except Exception as e:
            print(f"❌ Error generating report for {exam_no}: {e}")
    
    print(f"✅ Generated {len(carryover_data)} individual student reports in {reports_dir}")

def create_carryover_zip(source_dir, zip_path):
    """Create ZIP file of carryover results."""
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        print(f"✅ ZIP file created: {zip_path}")
        return True
    except Exception as e:
        print(f"❌ Error creating ZIP: {e}")
        return False

def main():
    """Main function to process carryover results - FIXED TO ONLY SCAN SELECTED PROGRAM."""
    print("🎯 CARRYOVER RESULT PROCESSOR - PROGRAM-SPECIFIC VERSION")
    print("=" * 50)
    
    # Check for alternative course files
    print("🔍 Looking for course files...")
    alternative_files = find_alternative_course_files()
    
    # Debug course file structure first
    debug_course_file_structure()
    
    # Configuration
    set_name = os.getenv("SELECTED_SET", "ND-2025")  # Changed to ND-2025 as requested
    semester_key = os.getenv("SELECTED_SEMESTERS", "ND-FIRST-YEAR-SECOND-SEMESTER")  # Updated default
    resit_file_path = os.getenv("RESIT_FILE_PATH", "")
    pass_threshold = float(os.getenv("PASS_THRESHOLD", str(DEFAULT_PASS_THRESHOLD)))
    
    print(f"🎯 Processing: Set={set_name}, Semester={semester_key}")
    
    # Determine program from set - FIXED LOGIC
    if set_name.startswith("ND-"):
        program = "ND"
    elif set_name.startswith("SET4"):
        program = "BN" 
    elif set_name.startswith("BM") or set_name.startswith("SET"):
        program = "BM"
    else:
        # Default to ND if we can't determine
        program = "ND"
        print(f"⚠️ Could not determine program from set name '{set_name}', defaulting to ND")
    
    # FIXED: Use correct base directory structure based on your file structure
    # Your files are in EXAMS_INTERNAL/ND/ND-2025/
    possible_base_dirs = [
        BASE_DIR,  # Original base (~/student_result_cleaner)
        os.path.join(BASE_DIR, "EXAMS_INTERNAL"),  # This is where your files actually are
        os.path.join(os.path.expanduser('~'), 'student_result_cleaner', 'EXAMS_INTERNAL'),
    ]
    
    raw_dir = None
    clean_dir = None
    output_dir = None
    
    for base in possible_base_dirs:
        # Try the standard structure first - ONLY FOR THE SELECTED PROGRAM
        test_raw_dir = os.path.join(base, program, set_name, "RAW_RESULTS")
        test_clean_dir = os.path.join(base, program, set_name, "CLEAN_RESULTS")
        
        print(f"🔍 Checking directory: {test_clean_dir}")
        
        if os.path.exists(test_clean_dir):
            raw_dir = test_raw_dir
            clean_dir = test_clean_dir
            output_dir = test_clean_dir
            print(f"✅ Found clean directory: {clean_dir}")
            break
    
    # If not found in standard locations, try alternative structures
    if not clean_dir:
        print("🔍 Trying alternative directory structures...")
        # Try direct path to EXAMS_INTERNAL/ND/ND-2025
        alt_clean_dir = os.path.join(BASE_DIR, "EXAMS_INTERNAL", program, set_name, "CLEAN_RESULTS")
        print(f"🔍 Checking alternative directory: {alt_clean_dir}")
        
        if os.path.exists(alt_clean_dir):
            raw_dir = os.path.join(BASE_DIR, "EXAMS_INTERNAL", program, set_name, "RAW_RESULTS")
            clean_dir = alt_clean_dir
            output_dir = alt_clean_dir
            print(f"✅ Found alternative clean directory: {clean_dir}")
    
    if not clean_dir:
        print(f"❌ Clean directory not found for {program}/{set_name}")
        print("💡 Please check:")
        print(f"   - Set name: {set_name}")
        print(f"   - Program: {program}") 
        print(f"   - Base directory: {BASE_DIR}")
        print(f"   - Expected directory: .../EXAMS_INTERNAL/{program}/{set_name}/CLEAN_RESULTS/")
        
        # Show what directories actually exist
        exams_internal_path = os.path.join(BASE_DIR, "EXAMS_INTERNAL")
        if os.path.exists(exams_internal_path):
            print(f"📁 Contents of EXAMS_INTERNAL: {os.listdir(exams_internal_path)}")
            program_path = os.path.join(exams_internal_path, program)
            if os.path.exists(program_path):
                print(f"📁 Contents of {program}: {os.listdir(program_path)}")
        
        print("💡 Please run the regular result processor first to generate clean results")
        return
    
    print(f"📁 Base directory: {BASE_DIR}")
    print(f"📁 Raw directory: {raw_dir}")
    print(f"📁 Clean directory: {clean_dir}")
    print(f"📁 Output directory: {output_dir}")
    print(f"📁 Resit file path: {resit_file_path}")
    
    # Validate resit file path
    if not resit_file_path or not os.path.exists(resit_file_path):
        print(f"❌ Resit file not provided or doesn't exist: {resit_file_path}")
        print("💡 Please set the RESIT_FILE_PATH environment variable to a valid resit file")
        return

    # Check if raw directory exists (as indicator of set existence)
    if not os.path.exists(raw_dir):
        print(f"⚠️ Raw directory doesn't exist: {raw_dir}")
        print("💡 The set might not be properly set up")
    
    # For carryover, use previous set's clean_dir if available
    previous_set = None
    if program == "ND" and set_name.startswith("ND-"):
        try:
            year = int(set_name.split('-')[1])
            previous_set = f"ND-{year - 1}"
            print(f"✅ Using previous set for carryover: {previous_set}")
        except:
            print(f"⚠️ Could not determine previous set")
    
    # Find latest mastersheet source (ZIP or folder) - ONLY IN THE SELECTED PROGRAM'S CLEAN_DIR
    print(f"🔍 Looking for mastersheet in: {clean_dir}")
    source_path, source_type = find_latest_mastersheet_source(clean_dir, set_name)
    if not source_path:
        print(f"❌ No ZIP files or result folders found in {clean_dir}")
        print(f"📁 Available files in {clean_dir}:")
        if os.path.exists(clean_dir):
            try:
                files = os.listdir(clean_dir)
                for f in files:
                    print(f"   - {f}")
            except Exception as e:
                print(f"   Error listing directory: {e}")
        else:
            print("   Directory not exist")
        print("💡 Please run the regular result processor first to generate clean results")
        return
    
    # Process carryover results
    success = process_carryover_results(
        resit_file_path, source_path, source_type, semester_key, set_name, pass_threshold, output_dir
    )
    
    if success:
        print(f"\n✅ Carryover processing completed successfully!")
        print(f"📁 Check the CLEAN_RESULTS directory for the CARRYOVER output")
    else:
        print(f"\n❌ Carryover processing failed!")

if __name__ == "__main__":
    main()