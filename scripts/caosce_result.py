#!/usr/bin/env python3
"""
caosce_result_fixed_all_papers.py
ENHANCED CAOSCE cleaning script with multi-college support and multi-paper processing
   FIXED: Overall average calculation (divide by actual papers taken, not always 5)
   FIXED: Failed papers logic (only fail if score > 0 and < 50)
   FIXED: Added upgrade count tracking and display
   FIXED: Proper handling of Paper I, II, III and CAOSCE
"""

import os
import re
import copy
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# ---------------------------
# College Configuration
# ---------------------------
COLLEGE_CONFIGS = {
    "YAGONGWO": {
        "name": "YAGONGWO COLLEGE OF NURSING SCIENCE, KUJE, ABUJA",
        "exam_patterns": [r'BN/A\d{2}/\d{3}', r'BN/\w+/\d+'],
        "mat_no_label": "MAT NO.",
        "logo": "LOGO_YAN.png",
        "output_prefix": "YAGONGWO"
    },
    "FCT": {
        "name": "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA", 
        "exam_patterns": [r'^\d{4}$', r'FCTCONS/ND\d{2}/\d{3}', r'FCTCONS/\w+/\d+'],
        "mat_no_label": "EXAM NO.",
        "logo": "logo.png",
        "output_prefix": "FCT"
    }
}

# ---------------------------
# Configuration
# ---------------------------
IS_RAILWAY = os.getenv("RAILWAY_ENVIRONMENT") is not None

if IS_RAILWAY:
    BASE_DIR = os.getenv("BASE_DIR", "/app/EXAMS_INTERNAL")
else:
    BASE_DIR = os.path.join(os.path.expanduser("~"), "student_result_cleaner", "EXAMS_INTERNAL")

DEFAULT_BASE_DIR = os.path.join(BASE_DIR, "CAOSCE_RESULT")
DEFAULT_RAW_DIR = os.path.join(DEFAULT_BASE_DIR, "RAW_CAOSCE_RESULT")
DEFAULT_CLEAN_DIR = os.path.join(DEFAULT_BASE_DIR, "CLEAN_CAOSCE_RESULT")

LOGO_BASE_PATH = os.path.join(os.path.expanduser("~"), "student_result_cleaner", "launcher", "static")

TIMESTAMP_FMT = "%Y-%m-d_%H%M%S"
CURRENT_YEAR = datetime.now().year

# Get upgrade threshold from environment variable
UPGRADE_THRESHOLD = int(os.getenv("UPGRADE_THRESHOLD", "0"))

# UPDATED: OSCE station weights - stations sum to 90% (6 stations × 15% each), viva is 10%
OSCE_STATION_WEIGHTS = {
    "procedure_station_one": 15.0,
    "procedure_station_three": 15.0,
    "procedure_station_five": 15.0,
    "question_station_two": 15.0,
    "question_station_four": 15.0,
    "question_station_six": 15.0,
    "viva": 10.0,
}

# Paper patterns
PAPER_I_PATTERNS = [
    r"PAPERI_PAPERII-PAPER I-grades",
    r"PAPER I", 
    r"PAPER 1",
    r"PAPERI",
    r"CLASS-PAPER I",
    r"CLASS-PAPER 1"
]

PAPER_II_PATTERNS = [
    r"PAPERI_PAPERII-PAPER II-grades",
    r"PAPER II",
    r"PAPER 2", 
    r"PAPERII",
    r"CLASS-PAPER II",
    r"CLASS-PAPER 2"
]

PAPER_III_PATTERNS = [
    r"PAPER III",
    r"PAPER 3",
    r"PAPERIII",
    r"MIDWIFERY",
    r"PAPER 3.*MIDWIFERY",
    r"PAPER III.*MIDWIFERY",
    r"MIDWIFERY.*PAPER",
    r"PAPER.*MIDWIFERY",
    r"CLASS-PAPER III",
    r"CLASS-PAPER 3"
]

COMBINED_PAPER_PATTERNS = [
    r"paper[_\s]*i[_\s]*paper[_\s]*ii",
    r"paper[_\s]*1[_\s]*paper[_\s]*2",
    r"paper[_\s]*i[_\s]*paper[_\s]*ii[_\s]*paper[_\s]*iii",
    r"paper[_\s]*1[_\s]*paper[_\s]*2[_\s]*paper[_\s]*3",
    r"combined[_\s]*papers",
    r"paper[_\s]*i[_\s]*&\s*paper[_\s]*ii",
    r"paper[_\s]*i[_\s]*and[_\s]*paper[_\s]*ii",
    r"all[_\s]*papers",
    r"papers[_\s]*i[_\s]*ii[_\s]*iii"
]

# Station mapping
STATION_COLUMN_MAP = {
    "procedure_station_one": "PS1_Score",
    "procedure_station_three": "PS3_Score",
    "procedure_station_five": "PS5_Score",
    "question_station_two": "QS2_Score",
    "question_station_four": "QS4_Score",
    "question_station_six": "QS6_Score",
    "viva": "VIVA",
}

# Station display names
STATION_DISPLAY_NAMES = {
    "procedure_station_one": "PS ONE (/10) (15%)",
    "procedure_station_three": "PS THREE (/10) (15%)", 
    "procedure_station_five": "PS FIVE (/10) (15%)",
    "question_station_two": "QS TWO (/10) (15%)",
    "question_station_four": "QS FOUR (/10) (15%)",
    "question_station_six": "QS SIX (/10) (15%)",
    "viva": "VIVA (/10) (10%)"
}

# Pass mark configuration
PASS_MARK = 50.0
OSCE_TOTAL_WEIGHT = 100.0

# Styling
NO_SCORE_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
NO_SCORE_FONT = Font(bold=True, color="9C0006", size=10, name="Calibri")
HEADER_FONT = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, name="Calibri", color="1F4E78")
SUBTITLE_FONT = Font(bold=True, size=14, name="Calibri", color="1F4E78")
DATE_FONT = Font(bold=True, size=11, name="Calibri", color="1F4E78")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
AVERAGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
AVERAGE_FONT = Font(bold=True, size=10, name="Calibri", color="7F6000")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(bold=True, size=10, name="Calibri", color="006100")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(bold=True, size=10, name="Calibri", color="9C0006")
FAILED_PAPERS_FONT = Font(size=10, name="Calibri", color="9C0006")
SUMMARY_HEADER_FONT = Font(bold=True, size=12, name="Calibri", color="1F4E78", underline="single")
SUMMARY_BODY_FONT = Font(size=11, name="Calibri")
ANALYSIS_HEADER_FONT = Font(bold=True, size=12, name="Calibri", color="1F4E78", underline="single")
ANALYSIS_BODY_FONT = Font(bold=True, size=11, name="Calibri")
SIGNATURE_FONT = Font(bold=True, size=11, name="Calibri", color="1F4E78")
SIGNATURE_LINE_FONT = Font(size=11, name="Calibri")

UNWANTED_COL_PATTERNS = [
    r"phone", r"department", r"city", r"town", r"state",
    r"started on", r"started", r"completed", r"time taken", r"duration",
    r"q\.?\s*\d+", r"email", r"status", r"date", r"groups?",
]

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------------------------
# Helper Functions
# ---------------------------

def apply_score_upgrade(score):
    """
    Apply upgrade to a score if UPGRADE_THRESHOLD is set and score is within upgrade range
    Returns: (upgraded_score, was_upgraded)
    """
    if UPGRADE_THRESHOLD > 0 and UPGRADE_THRESHOLD <= score < 50:
        upgraded_score = 50.0
        logger.debug(f"⬆️ Score upgrade: {score:.2f}% → 50.00%")
        return upgraded_score, True
    return score, False

def determine_remark_and_failed_papers(paper_i_score, paper_ii_score, paper_iii_score, caosce_score):
    """
    Determine REMARK and Failed Papers based on pass mark of 50
    FIXED: Only fail if score > 0 and < 50 (0 means paper not taken)
    """
    failed_papers = []
    failed_count = 0
    
    # Convert scores to float for comparison
    try:
        paper_i_score = float(paper_i_score) if paper_i_score is not None else 0.0
    except (ValueError, TypeError):
        paper_i_score = 0.0
    
    try:
        paper_ii_score = float(paper_ii_score) if paper_ii_score is not None else 0.0
    except (ValueError, TypeError):
        paper_ii_score = 0.0
    
    try:
        paper_iii_score = float(paper_iii_score) if paper_iii_score is not None else 0.0
    except (ValueError, TypeError):
        paper_iii_score = 0.0
    
    try:
        caosce_score = float(caosce_score) if caosce_score is not None else 0.0
    except (ValueError, TypeError):
        caosce_score = 0.0
    
    # Check each paper - only fail if score > 0 and < 50
    if paper_i_score > 0 and paper_i_score < PASS_MARK:
        failed_papers.append("Failed Paper 1")
        failed_count += 1
    if paper_ii_score > 0 and paper_ii_score < PASS_MARK:
        failed_papers.append("Failed Paper 2")
        failed_count += 1
    if paper_iii_score > 0 and paper_iii_score < PASS_MARK:
        failed_papers.append("Failed Paper 3 (Midwifery)")
        failed_count += 1
    if caosce_score > 0 and caosce_score < PASS_MARK:
        failed_papers.append("Failed CAOSCE")
        failed_count += 1
    
    # Determine overall remark
    if failed_count == 0:
        remark = "Passed"
    else:
        remark = "Failed"
    
    # Format failed papers string
    if failed_papers:
        failed_papers_str = ", ".join(failed_papers)
    else:
        failed_papers_str = ""
    
    return remark, failed_papers_str, failed_count

def detect_college_from_exam_numbers(exam_numbers):
    """Detect college based on exam number patterns"""
    if not exam_numbers:
        return "YAGONGWO", COLLEGE_CONFIGS["YAGONGWO"]
    
    yagongwo_count = 0
    fct_count = 0
    
    for exam_no in exam_numbers:
        exam_no_str = str(exam_no).strip().upper()
        
        # Check Yagongwo patterns
        yagongwo_matched = False
        for pattern in COLLEGE_CONFIGS["YAGONGWO"]["exam_patterns"]:
            if re.match(pattern, exam_no_str):
                yagongwo_count += 1
                yagongwo_matched = True
                break
        
        # Check FCT patterns only if not matched by Yagongwo
        if not yagongwo_matched:
            for pattern in COLLEGE_CONFIGS["FCT"]["exam_patterns"]:
                if re.match(pattern, exam_no_str):
                    fct_count += 1
                    break
    
    logger.info(f"College detection - Yagongwo: {yagongwo_count}, FCT: {fct_count}")
    
    if fct_count > yagongwo_count:
        return "FCT", COLLEGE_CONFIGS["FCT"]
    else:
        return "YAGONGWO", COLLEGE_CONFIGS["YAGONGWO"]

def find_first_col(df, candidates):
    cols = {c.strip().lower(): c for c in df.columns}
    for cand in candidates:
        cand_norm = cand.strip().lower()
        if cand_norm in cols:
            return cols[cand_norm]
        for c in df.columns:
            if cand_norm in str(c).strip().lower():
                return c
    return None

def find_username_col(df):
    return find_first_col(df, [
        "last name", "surname", "username", "user name", "exam no", 
        "exam number", "registration no", "reg no", "mat no", "matno", 
        "regnum", "groups", "group", "id"
    ])

def find_fullname_col(df):
    return find_first_col(df, [
        "user full name", "full name", "name", "candidate name", "student name",
        "first name", "given name", "first names"
    ])

def find_viva_score_col(df):
    return find_first_col(df, [
        "enter student score below", "enter student's score", "score", "grade"
    ])

def find_grade_column(df, filename="", station_key=None):
    """Dynamically find the grade/score column"""
    # First, look for any column containing / followed by a number
    for c in df.columns:
        cn = str(c).strip()
        match = re.search(r'/([\d.]+)', cn)
        if match:
            try:
                max_score = float(match.group(1))
                return (c, max_score)
            except:
                pass
    
    # For "Class-PAPER X-grades" format
    if "CLASS-PAPER" in filename.upper():
        logger.info(f"  Detected 'Class-PAPER X-grades' format for {filename}")
        
        # Look for columns with "grade", "total", or "score"
        for c in df.columns:
            cn = str(c).strip().lower()
            if "grade" in cn or "score" in cn or "total" in cn or "mark" in cn:
                logger.info(f"  Found likely grade column for Class-PAPER format: '{c}'")
                return (c, 100.0)
    
    # Special handling for VIVA station
    if station_key == "viva":
        for c in df.columns:
            cn = str(c).strip().lower()
            if "viva" in cn or "oral" in cn:
                match = re.search(r'/([\d.]+)', cn)
                if match:
                    try:
                        max_score = float(match.group(1))
                        return (c, max_score)
                    except:
                        pass
                logger.info(f"  Found VIVA column without denominator '{c}', assuming /10")
                return (c, 10.0)
    
    # Fallback: look for columns with 'grade', 'total', or 'score'
    for c in df.columns:
        cn = str(c).strip().lower()
        if "grade" in cn or "total" in cn or "score" in cn:
            if station_key and station_key != "viva":
                return (c, 10.0)
            else:
                return (c, 100.0)
    
    # Ultimate fallback
    logger.warning(f"  Could not find grade column by name. Using fallback.")
    for c in reversed(df.columns):
        cn = str(c).strip().lower()
        if not any(keyword in cn for keyword in ["name", "id", "no", "number", "mat", "reg"]):
            logger.info(f"  Using fallback grade column: '{c}'")
            return (c, 100.0)
    
    return (None, 100.0)

def extract_exam_number_from_fullname(text):
    if pd.isna(text):
        return None
    s = str(text).strip().upper()
    
    # Try Yagongwo pattern
    match = re.search(r'\b(BN/A\d{2}/\d{3})\b', s)
    if match:
        return match.group(1)
    
    # Try FCT pattern
    match = re.search(r'\b(FCTCONS/ND\d{2}/\d{3})\b', s)
    if match:
        return match.group(1)
    
    # Try 4-digit pattern
    match = re.search(r'\b(\d{4})\b', s)
    if match:
        return match.group(1)
    
    return None

def extract_fullname_from_text(text, exam_no):
    if pd.isna(text):
        return None
    s = str(text).strip()
    if exam_no:
        s = s.replace(exam_no, "")
    s = re.sub(r'\s*-\s*', ' ', s)
    s = " ".join(s.split())
    if s and re.search(r"[A-Za-z]{3,}", s):
        return s
    return None

def sanitize_exam_no(v):
    if pd.isna(v):
        return ""
    s = str(v).strip().upper()
    s = re.sub(r"\.0+$", "", s)
    if " - " in s:
        s = s.split(" - ")[0].strip()
    return s

def numeric_safe(v):
    try:
        if pd.isna(v) or str(v).strip() == "":
            return None
        v2 = str(v).strip().replace(",", "")
        return float(v2)
    except:
        return None

def is_overall_average_row(row, username_col, fullname_col):
    """Check if a row represents an overall average row"""
    if username_col and pd.notna(row.get(username_col)):
        if 'overall' in str(row[username_col]).lower() and 'average' in str(row[username_col]).lower():
            return True
    if fullname_col and pd.notna(row.get(fullname_col)):
        if 'overall' in str(row[fullname_col]).lower() and 'average' in str(row[fullname_col]).lower():
            return True
    return False

def apply_autofit_columns(ws, header_row, data_end_row):
    """Apply optimal column widths based on content"""
    max_lengths = {}
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        
        # Check header
        header_cell = ws.cell(row=header_row, column=col_idx)
        header_value = str(header_cell.value or "")
        max_length = len(header_value)
        
        # Check data rows
        for row_idx in range(header_row + 1, data_end_row + 1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_value = str(cell.value)
                    if isinstance(cell.value, (int, float)):
                        if cell.number_format == '0.00':
                            cell_value = f"{cell.value:.2f}"
                        elif cell.number_format == '0':
                            cell_value = f"{int(cell.value)}"
                    
                    cell_length = len(cell_value)
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        max_lengths[col_idx] = max_length
    
    # Set optimal widths
    for col_idx, max_length in max_lengths.items():
        col_letter = get_column_letter(col_idx)
        header_cell = ws.cell(row=header_row, column=col_idx)
        header_value = str(header_cell.value or "")
        
        optimal_width = min(80, max(8, max_length + 4))
        
        # Apply intelligent sizing
        if header_value == "S/N":
            optimal_width = max(6, min(optimal_width, 8))
        elif header_value in ["MAT NO.", "EXAM NO."]:
            optimal_width = max(12, min(optimal_width, 20))
        elif header_value == "FULL NAME":
            optimal_width = max(25, min(optimal_width, 40))
        elif "Score/" in header_value or "VIVA/" in header_value or "(15%)" in header_value or "(10%)" in header_value:
            optimal_width = max(15, min(optimal_width, 20))
        elif "Total Raw Score" in header_value:
            optimal_width = max(18, min(optimal_width, 25))
        elif "Percentage" in header_value:
            optimal_width = max(15, min(optimal_width, 18))
        elif "PAPER I" in header_value or "PAPER II" in header_value or "PAPER III" in header_value or "CAOSCE" in header_value:
            optimal_width = max(15, min(optimal_width, 20))
        elif "OVERALL AVERAGE" in header_value:
            optimal_width = max(18, min(optimal_width, 22))
        elif "REMARK" in header_value:
            optimal_width = max(10, min(optimal_width, 15))
        elif "FAILED PAPERS" in header_value:
            optimal_width = max(40, min(optimal_width, 60))
            if max_length > 50:
                optimal_width = min(80, max_length + 10)
        else:
            optimal_width = max(12, min(optimal_width, 25))
        
        # Ensure header fits
        header_length = len(header_value)
        if header_length + 2 > optimal_width:
            optimal_width = header_length + 2
        
        ws.column_dimensions[col_letter].width = optimal_width
        
        # Set wrap text for long content columns
        if header_value == "FAILED PAPERS":
            for row_idx in range(header_row + 1, data_end_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

def create_document_sections(ws, total_students, avg_percentage, highest_percentage, lowest_percentage, 
                           total_max_score, data_end_row, college_config, sheet_type="CAOSCE", 
                           upgraded_count=0, paper_upgrade_counts=None):
    """Create well-structured summary, analysis and signatories sections"""
    if paper_upgrade_counts is None:
        paper_upgrade_counts = {}
    
    doc_start_row = data_end_row + 3
    
    # ====================== SUMMARY SECTION ======================
    summary_header_row = doc_start_row
    last_col_letter = get_column_letter(ws.max_column)
    ws.merge_cells(f"A{summary_header_row}:{last_col_letter}{summary_header_row}")
    
    if sheet_type == "COMBINED":
        # Determine number of papers from column headers
        paper_columns = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=7, column=col).value
            if cell_value and ("PAPER" in str(cell_value).upper() or "MIDWIFERY" in str(cell_value).upper() or "CAOSCE" in str(cell_value).upper()):
                paper_columns.append(cell_value)
        
        paper_count = len(paper_columns)
        
        if paper_count >= 4:
            ws.cell(row=summary_header_row, column=1, value=f"EXAMINATION SUMMARY ({paper_count} Papers)")
            summary_rows = [
                "",
                f"Total Possible Score: {paper_count * 100} marks (100 per paper)",
                "",
                "Scoring Methodology:",
            ]
            
            # Add specific methodology for each paper
            if any("PAPER I" in str(col).upper() for col in paper_columns):
                summary_rows.append("- Paper I Score: 0-100 marks")
            if any("PAPER II" in str(col).upper() for col in paper_columns):
                summary_rows.append("- Paper II Score: 0-100 marks")
            if any("PAPER III" in str(col).upper() or "MIDWIFERY" in str(col).upper() for col in paper_columns):
                summary_rows.append("- Paper III (Midwifery) Score: 0-100 marks")
            if any("CAOSCE" in str(col).upper() for col in paper_columns):
                summary_rows.append("- CAOSCE Score: 0-100 marks (percentage from station performance)")
            
            summary_rows.extend([
                f"- Overall Average = Sum of all papers ÷ {paper_count}",
                "",
                "Pass/Fail Criteria:",
                f"- Pass Mark: {PASS_MARK}% for each paper",
                "- REMARK: 'Passed' if all papers ≥ 50%, 'Failed' otherwise",
                "- Failed Papers: Lists specific failed papers",
            ])
            
            # Add upgrade information if any upgrades were applied
            if UPGRADE_THRESHOLD > 0 and upgraded_count > 0:
                summary_rows.extend([
                    "",
                    "Score Upgrade Information:",
                    f"- Upgrade Threshold: {UPGRADE_THRESHOLD}%",
                    "- Scores ≥ threshold and < 50% are upgraded to 50%",
                    f"- Total Upgrades Applied: {upgraded_count}",
                ])
                
                # Add paper-specific upgrade counts
                if paper_upgrade_counts:
                    summary_rows.append("- Paper-specific upgrades:")
                    for paper, count in paper_upgrade_counts.items():
                        if count > 0:
                            summary_rows.append(f"  • {paper}: {count} upgrade(s)")
        else:
            ws.cell(row=summary_header_row, column=1, value="EXAMINATION SUMMARY")
            summary_rows = [
                "",
                f"Total Possible Score: {paper_count * 100} marks (100 per paper)",
                "",
                "Scoring Methodology:",
                "- Each Paper Score: 0-100 marks",
                f"- Overall Average = Sum of all papers ÷ {paper_count}",
                "",
                "Pass/Fail Criteria:",
                f"- Pass Mark: {PASS_MARK}% for each paper",
                "- REMARK: 'Passed' if all papers ≥ 50%, 'Failed' otherwise",
                "- Failed Papers: Lists specific failed papers",
            ]
            
            if UPGRADE_THRESHOLD > 0 and upgraded_count > 0:
                summary_rows.extend([
                    "",
                    "Score Upgrade Information:",
                    f"- Upgrade Threshold: {UPGRADE_THRESHOLD}%",
                    "- Scores ≥ threshold and < 50% are upgraded to 50%",
                    f"- Total Upgrades Applied: {upgraded_count}",
                ])
    else:
        # CAOSCE sheet
        ws.cell(row=summary_header_row, column=1, value="OSCE EXAMINATION SUMMARY (Updated Grading)")
        summary_rows = [
            "",
            "Updated OSCE Station Configuration:",
            "- 6 Stations (Procedure Stations 1, 3, 5 and Question Stations 2, 4, 6)",
            "- Each station scored out of 10 marks",
            "- Each station weighted at 15% of total OSCE (6 × 15% = 90%)",
            "- VIVA scored out of 10 marks, weighted at 10% of total OSCE",
            "",
            "Scoring Methodology:",
            "1. Sum of all 6 stations (maximum 60 marks)",
            "2. Convert to 90% weight: (sum ÷ 60) × 90",
            "3. Add VIVA score converted to 10%: (VIVA ÷ 10) × 10",
            "4. Total OSCE Score = Step 2 + Step 3",
            "",
            "Pass/Fail Criteria:",
            f"- Pass Mark: {PASS_MARK}%",
            "- REMARK: 'Passed' if ≥ 50%, 'Failed' otherwise",
        ]
        
        if UPGRADE_THRESHOLD > 0 and upgraded_count > 0:
            summary_rows.extend([
                "",
                "Score Upgrade Information:",
                f"- Upgrade Threshold: {UPGRADE_THRESHOLD}%",
                "- Scores ≥ threshold and < 50% are upgraded to 50%",
                f"- Total Upgrades Applied: {upgraded_count}",
            ])
    
    ws.cell(row=summary_header_row, column=1).font = SUMMARY_HEADER_FONT
    ws.cell(row=summary_header_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    for i, line in enumerate(summary_rows, 1):
        row_num = summary_header_row + i
        ws.merge_cells(f"A{row_num}:{last_col_letter}{row_num}")
        cell = ws.cell(row=row_num, column=1, value=line)
        if "Methodology:" in line or line.startswith("-") or "Upgrade" in line:
            cell.font = Font(bold=True, size=11, name="Calibri")
        else:
            cell.font = SUMMARY_BODY_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # ====================== ANALYSIS SECTION ======================
    analysis_start_row = summary_header_row + len(summary_rows) + 2
    
    ws.merge_cells(f"A{analysis_start_row}:{last_col_letter}{analysis_start_row}")
    ws.cell(row=analysis_start_row, column=1, value="PERFORMANCE ANALYSIS")
    ws.cell(row=analysis_start_row, column=1).font = ANALYSIS_HEADER_FONT
    ws.cell(row=analysis_start_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    if sheet_type == "COMBINED":
        # Get paper-wise averages from the overall average row
        paper_averages = {}
        paper_failed_counts = {}
        total_passed = 0
        total_failed = 0
        
        # Find paper columns and their indices
        paper_columns_indices = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=7, column=col_idx).value
            if cell_value and ("PAPER" in str(cell_value).upper() or "MIDWIFERY" in str(cell_value).upper() or "CAOSCE" in str(cell_value).upper()):
                paper_columns_indices[col_idx] = cell_value
        
        # Find the overall average row
        calculated_header_row = 7
        overall_avg_row_idx = None
        for row_idx in range(calculated_header_row + 1, data_end_row + 1):
            if ws.cell(row=row_idx, column=2).value == "OVERALL AVERAGE":
                overall_avg_row_idx = row_idx
                for col_idx, paper_name in paper_columns_indices.items():
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        try:
                            paper_averages[paper_name] = float(cell_value)
                        except (ValueError, TypeError):
                            paper_averages[paper_name] = 0.0
                    else:
                        paper_averages[paper_name] = 0.0
                break
        
        # Calculate pass/fail statistics
        student_overall_scores = []
        
        for row_idx in range(calculated_header_row + 1, data_end_row + 1):
            mat_no = ws.cell(row=row_idx, column=2).value
            if mat_no == "OVERALL AVERAGE":
                continue
                
            # Get overall score
            overall_score = 0
            for col_idx in range(ws.max_column, 0, -1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and ("OVERALL AVERAGE" in str(cell_value) or col_idx == 7):
                    try:
                        overall_score = float(cell_value) if cell_value is not None else 0
                    except (ValueError, TypeError):
                        overall_score = 0
                    break
            
            student_overall_scores.append(overall_score)
            
            # Get remark
            remark = ""
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value == "Passed" or cell_value == "Failed":
                    remark = cell_value
                    break
            
            if remark == "Passed":
                total_passed += 1
            elif remark == "Failed":
                total_failed += 1
            
            # Count failures per paper
            for col_idx, paper_name in paper_columns_indices.items():
                score_cell = ws.cell(row=row_idx, column=col_idx).value
                score = 0
                if score_cell is not None:
                    try:
                        score = float(score_cell)
                    except (ValueError, TypeError):
                        score = 0
                
                if score > 0 and score < PASS_MARK:
                    if paper_name not in paper_failed_counts:
                        paper_failed_counts[paper_name] = 0
                    paper_failed_counts[paper_name] += 1
        
        # Calculate highest and lowest scores
        if student_overall_scores:
            valid_scores = []
            for score in student_overall_scores:
                if score is not None:
                    try:
                        valid_scores.append(float(score))
                    except (ValueError, TypeError):
                        pass
            
            if valid_scores:
                highest_individual = round(max(valid_scores), 2)
                lowest_individual = round(min(valid_scores), 2)
            else:
                highest_individual = 0
                lowest_individual = 0
        else:
            highest_individual = 0
            lowest_individual = 0
        
        # Calculate percentages
        passed_percentage = (total_passed / total_students * 100) if total_students > 0 else 0
        failed_percentage = (total_failed / total_students * 100) if total_students > 0 else 0
        
        # Calculate class performance rating
        performance_rating = ""
        if avg_percentage >= 70:
            performance_rating = "EXCELLENT"
        elif avg_percentage >= 60:
            performance_rating = "GOOD"
        elif avg_percentage >= 50:
            performance_rating = "AVERAGE"
        elif avg_percentage >= 40:
            performance_rating = "BELOW AVERAGE"
        else:
            performance_rating = "POOR"
        
        # Build analysis rows
        analysis_rows = [
            "",
            f"Total Candidates: {total_students}",
            f"Overall Class Average: {avg_percentage}%",
            f"Highest Individual Score: {highest_individual}%",
            f"Lowest Individual Score: {lowest_individual}%",
            f"Class Performance Rating: {performance_rating}",
            "",
            "════════════════════════════════════════════════════════════════",
            "PASS/FAIL STATISTICS:",
            f"• Candidates Passed: {total_passed} ({passed_percentage:.1f}%)",
            f"• Candidates Failed: {total_failed} ({failed_percentage:.1f}%)",
            "",
            "════════════════════════════════════════════════════════════════",
            "PAPER-WISE PERFORMANCE ANALYSIS:",
        ]
        
        # Add each paper's average
        for paper_name, avg_score in paper_averages.items():
            analysis_rows.append(f"• {paper_name} Average: {avg_score:.1f}%")
        
        analysis_rows.append("")
        analysis_rows.append("════════════════════════════════════════════════════════════════")
        analysis_rows.append("PAPER-WISE FAILURE ANALYSIS:")
        
        # Add each paper's failure count
        for paper_name, failed_count in paper_failed_counts.items():
            failed_pct = (failed_count / total_students * 100) if total_students > 0 else 0
            analysis_rows.append(f"• Failed {paper_name}: {failed_count} ({failed_pct:.1f}%)")
    else:
        # For CAOSCE sheet
        analysis_rows = [
            "",
            f"Total Candidates: {total_students}",
            f"Overall OSCE Average: {avg_percentage}%",
            f"Highest OSCE Score: {highest_percentage}%",
            f"Lowest OSCE Score: {lowest_percentage}%",
            "",
            "Performance Rating:",
        ]
        
        # Add performance rating for OSCE
        if avg_percentage >= 70:
            analysis_rows.append("• EXCELLENT: Strong OSCE performance")
        elif avg_percentage >= 60:
            analysis_rows.append("• GOOD: Satisfactory OSCE performance")
        elif avg_percentage >= 50:
            analysis_rows.append("• AVERAGE: Acceptable OSCE performance")
        elif avg_percentage >= 40:
            analysis_rows.append("• BELOW AVERAGE: Needs improvement in OSCE")
        else:
            analysis_rows.append("• POOR: Significant OSCE deficiencies")
    
    for i, line in enumerate(analysis_rows, 1):
        row_num = analysis_start_row + i
        ws.merge_cells(f"A{row_num}:{last_col_letter}{row_num}")
        cell = ws.cell(row=row_num, column=1, value=line)
        if line.startswith("════") or line.endswith(":"):
            cell.font = Font(bold=True, size=11, name="Calibri", color="1F4E78")
        elif line.startswith("•"):
            cell.font = ANALYSIS_BODY_FONT
        else:
            cell.font = ANALYSIS_BODY_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # ====================== SIGNATORIES SECTION ======================
    signatories_start_row = analysis_start_row + len(analysis_rows) + 3
    
    # Get total number of columns to determine layout
    total_columns = ws.max_column
    
    # Adjust signatories layout based on available columns
    if total_columns >= 8:
        # Wide layout
        examiners_col_end = total_columns // 2
        approved_col_start = examiners_col_end + 1
        approved_col_end = total_columns
        
        # Examiners section
        ws.merge_cells(f"A{signatories_start_row}:{get_column_letter(examiners_col_end)}{signatories_start_row}")
        ws.cell(row=signatories_start_row, column=1, value="EXAMINERS' NAME & SIGNATURE:")
        ws.cell(row=signatories_start_row, column=1).font = SIGNATURE_FONT
        ws.cell(row=signatories_start_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        # Add 4 numbered lines for examiners
        for i in range(1, 5):
            row_num = signatories_start_row + i
            ws.merge_cells(f"A{row_num}:{get_column_letter(examiners_col_end)}{row_num}")
            signature_line = f"{i}. _______________________________________________________________"
            ws.cell(row=row_num, column=1, value=signature_line)
            ws.cell(row=row_num, column=1).font = SIGNATURE_LINE_FONT
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 25
        
        # Approved by section
        ws.merge_cells(f"{get_column_letter(approved_col_start)}{signatories_start_row}:{get_column_letter(approved_col_end)}{signatories_start_row}")
        approved_cell = ws.cell(row=signatories_start_row, column=approved_col_start, value="APPROVED BY:")
        approved_cell.font = SIGNATURE_FONT
        approved_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Signature line
        ws.merge_cells(f"{get_column_letter(approved_col_start)}{signatories_start_row + 2}:{get_column_letter(approved_col_end)}{signatories_start_row + 2}")
        signature_cell = ws.cell(row=signatories_start_row + 2, column=approved_col_start, 
                                 value="________________________________________________________________")
        signature_cell.font = SIGNATURE_LINE_FONT
        signature_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Provost title
        ws.merge_cells(f"{get_column_letter(approved_col_start)}{signatories_start_row + 4}:{get_column_letter(approved_col_end)}{signatories_start_row + 4}")
        provost_cell = ws.cell(row=signatories_start_row + 4, column=approved_col_start, value="PROVOST'S SIGNATURE")
        provost_cell.font = SIGNATURE_FONT
        provost_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Name line
        ws.merge_cells(f"{get_column_letter(approved_col_start)}{signatories_start_row + 6}:{get_column_letter(approved_col_end)}{signatories_start_row + 6}")
        name_cell = ws.cell(row=signatories_start_row + 6, column=approved_col_start, 
                           value="NAME: ____________________________________________________________")
        name_cell.font = SIGNATURE_LINE_FONT
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Date line
        ws.merge_cells(f"{get_column_letter(approved_col_start)}{signatories_start_row + 8}:{get_column_letter(approved_col_end)}{signatories_start_row + 8}")
        date_cell = ws.cell(row=signatories_start_row + 8, column=approved_col_start, 
                           value="DATE: _____________________________________________________________")
        date_cell.font = SIGNATURE_LINE_FONT
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        # Narrow layout
        examiners_col_end = total_columns
        
        # Examiners section
        ws.merge_cells(f"A{signatories_start_row}:{get_column_letter(examiners_col_end)}{signatories_start_row}")
        ws.cell(row=signatories_start_row, column=1, value="EXAMINERS' NAME & SIGNATURE:")
        ws.cell(row=signatories_start_row, column=1).font = SIGNATURE_FONT
        ws.cell(row=signatories_start_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        # Add 4 numbered lines for examiners
        for i in range(1, 5):
            row_num = signatories_start_row + i
            ws.merge_cells(f"A{row_num}:{get_column_letter(examiners_col_end)}{row_num}")
            signature_line = f"{i}. _______________________________________________________________"
            ws.cell(row=row_num, column=1, value=signature_line)
            ws.cell(row=row_num, column=1).font = SIGNATURE_LINE_FONT
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 25
        
        # Approved by section
        approved_start_row = signatories_start_row + 6
        
        ws.merge_cells(f"A{approved_start_row}:{get_column_letter(examiners_col_end)}{approved_start_row}")
        approved_cell = ws.cell(row=approved_start_row, column=1, value="APPROVED BY:")
        approved_cell.font = SIGNATURE_FONT
        approved_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Signature line
        ws.merge_cells(f"A{approved_start_row + 2}:{get_column_letter(examiners_col_end)}{approved_start_row + 2}")
        signature_cell = ws.cell(row=approved_start_row + 2, column=1, 
                                 value="________________________________________________________________")
        signature_cell.font = SIGNATURE_LINE_FONT
        signature_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Provost title
        ws.merge_cells(f"A{approved_start_row + 4}:{get_column_letter(examiners_col_end)}{approved_start_row + 4}")
        provost_cell = ws.cell(row=approved_start_row + 4, column=1, value="PROVOST'S SIGNATURE")
        provost_cell.font = SIGNATURE_FONT
        provost_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Name line
        ws.merge_cells(f"A{approved_start_row + 6}:{get_column_letter(examiners_col_end)}{approved_start_row + 6}")
        name_cell = ws.cell(row=approved_start_row + 6, column=1, 
                           value="NAME: ____________________________________________________________")
        name_cell.font = SIGNATURE_LINE_FONT
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Date line
        ws.merge_cells(f"A{approved_start_row + 8}:{get_column_letter(examiners_col_end)}{approved_start_row + 8}")
        date_cell = ws.cell(row=approved_start_row + 8, column=1, 
                           value="DATE: _____________________________________________________________")
        date_cell.font = SIGNATURE_LINE_FONT
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    return signatories_start_row + 10

def detect_paper_type(filename):
    """Detect if file is Paper I, Paper II, Paper III, CAOSCE station, or Combined Papers"""
    fname_upper = filename.upper()
    fname_lower = filename.lower()
    
    # Check for combined paper patterns first
    for pattern in COMBINED_PAPER_PATTERNS:
        if re.search(pattern, fname_lower, re.IGNORECASE):
            return "COMBINED_PAPERS"
    
    # Check for files that contain multiple paper markers
    has_paper_i = bool(re.search(r'\bPAPER[\s-]+I\b(?!I)', fname_upper))
    has_paper_ii = bool(re.search(r'\bPAPER[\s-]+II\b', fname_upper))
    has_paper_iii = bool(re.search(r'\bPAPER[\s-]+III\b', fname_upper))
    
    if (has_paper_i and has_paper_ii) or (has_paper_i and has_paper_iii) or (has_paper_ii and has_paper_iii):
        return "COMBINED_PAPERS"
    
    # Check for MIDWIFERY with other papers
    if "MIDWIFERY" in fname_upper and ("PAPER I" in fname_upper or "PAPER II" in fname_upper):
        return "COMBINED_PAPERS"
    
    # Check for Paper III
    paper_iii_patterns = [
        r"\bCLASS-PAPER[\s-]+III\b",
        r"\bCLASS-PAPER[\s-]+3\b",
        r"\bPAPER[\s-]+III\b",
        r"\bPAPER[\s-]+3\b",
        r"\bPAPERIII\b",
        r"\bMIDWIFERY\b",
        r"\bPAPER[\s-]+3.*MIDWIFERY\b",
        r"\bPAPER[\s-]+III.*MIDWIFERY\b",
    ]
    
    for pattern in paper_iii_patterns:
        if re.search(pattern, fname_upper, re.IGNORECASE):
            return "PAPER_III"
    
    # Check for Paper II
    paper_ii_patterns = [
        r"\bCLASS-PAPER[\s-]+II\b",
        r"\bCLASS-PAPER[\s-]+2\b",
        r"\bPAPER[\s-]+II\b",
        r"\bPAPER[\s-]+2\b",
        r"\bPAPERII\b",
        r"PAPERI_PAPERII-PAPER[\s-]+II",
        r"PAPER[\s-]+II-GRADES"
    ]
    
    for pattern in paper_ii_patterns:
        if re.search(pattern, fname_upper, re.IGNORECASE):
            return "PAPER_II"
    
    # Check for Paper I
    paper_i_patterns = [
        r"\bCLASS-PAPER[\s-]+I\b(?!\s*I)",
        r"\bCLASS-PAPER[\s-]+1\b",
        r"\bPAPER[\s-]+I\b(?!\s*I)",
        r"\bPAPER[\s-]+1\b",
        r"\bPAPERI\b(?!I)",
        r"PAPERI_PAPERII-PAPER[\s-]+I\b(?!\s*I)",
        r"PAPER[\s-]+I-GRADES"
    ]
    
    for pattern in paper_i_patterns:
        if re.search(pattern, fname_upper, re.IGNORECASE):
            return "PAPER_I"
    
    # Check for CAOSCE station patterns
    if any(station in fname_lower for station in ["procedure", "question", "viva", "ps-", "qs-", "ps1", "ps3", "ps5", "qs2", "qs4", "qs6"]):
        return "CAOSCE_STATION"
    
    return "UNKNOWN"

def process_paper_files(files, raw_dir):
    """Process Paper I, II, and III files including combined format"""
    paper_results = {}
    paper_upgrade_counts = {"PAPER I": 0, "PAPER II": 0, "PAPER III": 0}
    
    logger.info("=" * 60)
    logger.info("STARTING PAPER FILES PROCESSING (I, II, and III)")
    logger.info("=" * 60)
    
    # First, identify and process COMBINED_PAPERS files
    combined_files = []
    separate_paper_files = []
    
    for fname in files:
        paper_type = detect_paper_type(fname)
        if paper_type == "COMBINED_PAPERS":
            combined_files.append(fname)
        elif paper_type in ["PAPER_I", "PAPER_II", "PAPER_III"]:
            separate_paper_files.append(fname)
    
    logger.info(f"Found {len(combined_files)} combined paper file(s)")
    logger.info(f"Found {len(separate_paper_files)} separate paper file(s): {separate_paper_files}")
    
    # Process COMBINED files first
    for fname in combined_files:
        logger.info(f"\n{'='*60}")
        logger.info(f"Processing COMBINED file: {fname}")
        logger.info(f"{'='*60}")
        
        path = os.path.join(raw_dir, fname)
        
        try:
            if fname.lower().endswith(".csv"):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str)
        except Exception as e:
            logger.error(f"❌ Error reading {fname}: {e}")
            continue
            
        df.rename(columns=lambda c: str(c).strip(), inplace=True)
        
        # Process combined papers
        processed_combined = process_combined_papers(df, fname)
        
        if not processed_combined:
            continue
        
        # Merge results from combined file into paper_results
        for exam_no, data in processed_combined.items():
            # Initialize if new student
            if exam_no not in paper_results:
                paper_results[exam_no] = {
                    "PAPER I": 0.00,
                    "PAPER II": 0.00,
                    "PAPER III": 0.00,
                    "FULL NAME": data.get("FULL NAME")
                }
            
            # Update Paper scores
            for paper in ["PAPER I", "PAPER II", "PAPER III"]:
                paper_score = data.get(paper)
                if paper_score is not None:
                    try:
                        numeric_score = float(paper_score)
                    except (ValueError, TypeError):
                        numeric_score = 0.00
                    
                    # Apply upgrade
                    upgraded_score, was_upgraded = apply_score_upgrade(numeric_score)
                    if was_upgraded:
                        paper_upgrade_counts[paper] += 1
                    
                    paper_results[exam_no][paper] = upgraded_score
            
            # Update full name if not set
            if data.get("FULL NAME") and not paper_results[exam_no]["FULL NAME"]:
                paper_results[exam_no]["FULL NAME"] = data.get("FULL NAME")
    
    # Process separate PAPER_I, PAPER_II, and PAPER_III files
    for fname in separate_paper_files:
        paper_type = detect_paper_type(fname)
        path = os.path.join(raw_dir, fname)
        
        logger.info(f"\n{'='*40}")
        logger.info(f"Processing {paper_type} file: {fname}")
        logger.info(f"{'='*40}")
        
        try:
            if fname.lower().endswith(".csv"):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str)
        except Exception as e:
            logger.error(f"Error reading {fname}: {e}")
            continue
            
        df.rename(columns=lambda c: str(c).strip(), inplace=True)
        
        # Find columns
        username_col = find_username_col(df)
        fullname_col = find_fullname_col(df)
        grade_col, max_score = find_grade_column(df, fname)
        
        if not grade_col:
            logger.error(f"  CRITICAL: Could not find grade column in {fname}")
            continue
            
        # Remove unwanted columns
        for pattern in UNWANTED_COL_PATTERNS:
            df.drop(columns=[c for c in df.columns if re.search(pattern, str(c), flags=re.I)], 
                   inplace=True, errors="ignore")
        
        paper_label = paper_type.replace("_", " ")
        
        for idx, row in df.iterrows():
            # Skip overall average rows
            if is_overall_average_row(row, username_col, fullname_col):
                continue
                
            exam_no = None
            full_name = None
            
            # Extract exam number
            if username_col and pd.notna(row.get(username_col)):
                exam_no = sanitize_exam_no(row.get(username_col))
            
            # Try fullname column if username didn't work
            if (not exam_no or exam_no == "") and fullname_col and pd.notna(row.get(fullname_col)):
                fullname_value = str(row.get(fullname_col, "")).strip()
                exam_no = extract_exam_number_from_fullname(fullname_value)
                if exam_no:
                    full_name = extract_fullname_from_text(fullname_value, exam_no)
            
            # Try scanning all columns for exam number patterns
            if not exam_no or exam_no == "":
                for col in df.columns:
                    if col == username_col or col == fullname_col:
                        continue
                    val = str(row.get(col, "")).strip()
                    if re.search(r'BN/A\d{2}/\d{3}|FCTCONS/ND\d{2}/\d{3}|\b\d{4}\b', val):
                        exam_no = sanitize_exam_no(val)
                        break
            
            if not exam_no or exam_no == "":
                continue
            
            # Extract full name
            if not full_name and fullname_col and pd.notna(row.get(fullname_col)):
                full_name = str(row.get(fullname_col, "")).strip()
                if full_name and not re.search(r'[A-Za-z]{3,}', full_name):
                    full_name = None
                
            # Initialize student record if not exists
            if exam_no not in paper_results:
                paper_results[exam_no] = {
                    "PAPER I": 0.00,
                    "PAPER II": 0.00,
                    "PAPER III": 0.00,
                    "FULL NAME": full_name
                }
            
            # Update full name if not set
            if full_name and not paper_results[exam_no]["FULL NAME"]:
                paper_results[exam_no]["FULL NAME"] = full_name
                
            # Extract and normalize score
            score_val = numeric_safe(row.get(grade_col))
            
            if score_val is not None:
                # Normalize if needed
                if score_val > 100:
                    normalized_score = (score_val / max_score) * 100
                else:
                    normalized_score = score_val
                
                # Apply upgrade if enabled
                upgraded_score, was_upgraded = apply_score_upgrade(normalized_score)
                if was_upgraded:
                    paper_upgrade_counts[paper_label] += 1
                
                rounded_score = round(upgraded_score, 2)
                
                # Only fill if this score is currently 0.00
                current_score = paper_results[exam_no][paper_label]
                if current_score == 0.00:
                    paper_results[exam_no][paper_label] = rounded_score
    
    # Final summary
    logger.info(f"\n{'='*60}")
    logger.info(f"PAPER FILES PROCESSING COMPLETE")
    logger.info(f"{'='*60}")
    logger.info(f"Total students with paper data: {len(paper_results)}")
    
    return paper_results, paper_upgrade_counts

def process_combined_papers(df, filename):
    """Process combined papers format including PAPER III"""
    logger.info(f"=== PROCESSING COMBINED PAPERS FILE: {filename} ===")
    
    results = {}
    
    # Find columns by flexible matching
    mat_no_col = None
    full_name_col = None
    paper_cols = {}
    
    # Find MAT NO column
    mat_no_patterns = ["MAT NO", "MATNO", "MAT.NO", "EXAM NO", "REG NO", "REGISTRATION", "ID", "STUDENT ID"]
    for col in df.columns:
        col_upper = str(col).strip().upper()
        for pattern in mat_no_patterns:
            if pattern in col_upper:
                mat_no_col = col
                logger.info(f"  Found MAT NO column: '{col}'")
                break
        if mat_no_col:
            break
    
    # Find FULL NAME column
    name_patterns = ["FULL NAME", "FULLNAME", "NAME", "STUDENT NAME", "CANDIDATE NAME", "FIRST NAME", "LAST NAME"]
    for col in df.columns:
        col_upper = str(col).strip().upper()
        for pattern in name_patterns:
            if pattern in col_upper:
                full_name_col = col
                logger.info(f"  Found FULL NAME column: '{col}'")
                break
        if full_name_col:
            break
    
    # Find Paper columns
    for col in df.columns:
        col_str = str(col).strip()
        col_upper = col_str.upper()
        
        # Check for PAPER I
        if re.search(r'PAPER\s*I|PAPER\s*1|PAPERI', col_upper) and not re.search(r'PAPER\s*II|PAPER\s*2|PAPERIII', col_upper):
            paper_cols["PAPER I"] = col
            logger.info(f"  Found Paper I column: '{col}'")
        
        # Check for PAPER II
        if re.search(r'PAPER\s*II|PAPER\s*2|PAPERII', col_upper) and not re.search(r'PAPER\s*III|PAPER\s*3', col_upper):
            paper_cols["PAPER II"] = col
            logger.info(f"  Found Paper II column: '{col}'")
        
        # Check for PAPER III (Midwifery)
        if re.search(r'PAPER\s*III|PAPER\s*3|PAPERIII|MIDWIFERY', col_upper, re.IGNORECASE):
            paper_cols["PAPER III"] = col
            logger.info(f"  Found Paper III column: '{col}'")
    
    # Validate required columns
    if not mat_no_col:
        logger.error("  ❌ ERROR: Could not find MAT NO column!")
        return results
    
    rows_processed = 0
    
    for idx, row in df.iterrows():
        # Get exam number
        exam_no = sanitize_exam_no(row.get(mat_no_col))
        
        # Skip "Overall average" rows
        if not exam_no or exam_no == "":
            continue
            
        exam_no_str = str(exam_no).lower()
        fullname_str = str(row.get(full_name_col, "")).lower() if full_name_col else ""
        if ("overall" in exam_no_str and "average" in exam_no_str) or \
           ("overall" in fullname_str and "average" in fullname_str):
            continue
        
        # Get full name
        full_name = None
        if full_name_col:
            full_name = str(row.get(full_name_col, "")).strip()
            if not full_name or not re.search(r'[A-Za-z]{3,}', full_name):
                full_name = None
        
        # Initialize result entry
        results[exam_no] = {
            "FULL NAME": full_name,
        }
        
        # Get scores for each paper found
        for paper_name, col_name in paper_cols.items():
            paper_score = 0.00
            paper_val = numeric_safe(row.get(col_name))
            if paper_val is not None:
                paper_score = round(float(paper_val), 2)
                
                # Apply upgrade if enabled
                upgraded_score, was_upgraded = apply_score_upgrade(paper_score)
                if was_upgraded:
                    paper_score = upgraded_score
            
            results[exam_no][paper_name] = paper_score
        
        rows_processed += 1
    
    logger.info(f"  Processed {rows_processed} rows")
    logger.info(f"  Unique students: {len(results)}")
    
    return results

def process_caosce_station_files(files, raw_dir):
    """Process CAOSCE station files"""
    caosce_results = {}
    station_max_scores = {}
    station_overall_averages = {}
    all_exam_numbers = set()
    caosce_upgrade_count = 0
    
    # Initialize all station keys
    station_keys = list(STATION_COLUMN_MAP.keys())
    
    for fname in sorted(files):
        path = os.path.join(raw_dir, fname)
        paper_type = detect_paper_type(fname)
        
        # Only process CAOSCE station files
        if paper_type != "CAOSCE_STATION":
            continue
            
        lower = fname.lower()
        station_key = None
        
        # Determine station type
        if "procedure" in lower or "ps-" in lower or "ps1" in lower or "ps3" in lower or "ps5" in lower:
            if "one" in lower or "ps1" in lower or "_1" in lower:
                station_key = "procedure_station_one"
            elif "three" in lower or "ps3" in lower or "_3" in lower:
                station_key = "procedure_station_three"
            elif "five" in lower or "ps5" in lower or "_5" in lower:
                station_key = "procedure_station_five"
        elif "question" in lower or "qs-" in lower or "qs2" in lower or "qs4" in lower or "qs6" in lower:
            if "two" in lower or "qs2" in lower or "_2" in lower:
                station_key = "question_station_two"
            elif "four" in lower or "qs4" in lower or "_4" in lower:
                station_key = "question_station_four"
            elif "six" in lower or "qs6" in lower or "_6" in lower:
                station_key = "question_station_six"
        elif "viva" in lower:
            station_key = "viva"

        if not station_key:
            logger.warning(f"Could not determine station for {fname} – skipping")
            continue

        try:
            if fname.lower().endswith(".csv"):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str)
        except Exception as e:
            logger.error(f"Error reading {fname}: {e}")
            continue

        df.rename(columns=lambda c: str(c).strip(), inplace=True)

        username_col = find_username_col(df)
        fullname_col = find_fullname_col(df)
        grade_col, max_score = find_grade_column(df, fname, station_key)
        viva_score_col = find_viva_score_col(df) if station_key == "viva" else None

        if grade_col:
            # Store the actual max score for this station
            station_max_scores[station_key] = max_score
        else:
            # Use standard weights if not found
            if station_key == "viva":
                station_max_scores[station_key] = 10.0
            else:
                station_max_scores[station_key] = 10.0

        # Remove unwanted columns
        for pattern in UNWANTED_COL_PATTERNS:
            df.drop(columns=[c for c in df.columns if re.search(pattern, str(c), flags=re.I)], 
                   inplace=True, errors="ignore")

        rows_added = 0
        station_scores = []
        
        for _, row in df.iterrows():
            # Skip overall average rows
            if is_overall_average_row(row, username_col, fullname_col):
                continue
                
            exam_no = None
            full_name = None
            
            if station_key == "viva":
                if fullname_col and pd.notna(row.get(fullname_col)):
                    fullname_value = str(row[fullname_col]).strip()
                    exam_no = extract_exam_number_from_fullname(fullname_value)
                    if exam_no:
                        full_name = extract_fullname_from_text(fullname_value, exam_no)
                if not exam_no and username_col:
                    exam_no = sanitize_exam_no(row.get(username_col))
                if not exam_no:
                    continue
            else:
                if username_col:
                    raw_value = row.get(username_col)
                    exam_no = sanitize_exam_no(raw_value)
                    if exam_no and pd.notna(raw_value):
                        full_name = extract_fullname_from_text(str(raw_value), exam_no)
                if not full_name and fullname_col and pd.notna(row.get(fullname_col)):
                    fullname_value = str(row[fullname_col]).strip()
                    if not exam_no:
                        exam_no = extract_exam_number_from_fullname(fullname_value)
                    full_name = extract_fullname_from_text(fullname_value, exam_no)
                if not exam_no:
                    for c in df.columns:
                        val = sanitize_exam_no(row.get(c))
                        if val and len(val) > 2 and re.search(r'\d', val):
                            exam_no = val
                            break
                if not full_name:
                    for c in df.columns:
                        if c == username_col or c == fullname_col:
                            continue
                        val = str(row.get(c, "")).strip()
                        if val and re.search(r"[A-Za-z]{3,}", val) and not re.search(r'\d{2,}', val):
                            full_name = val
                            break
                if not exam_no:
                    continue

            if exam_no:
                all_exam_numbers.add(exam_no)

            # Initialize student with all station keys if not exists
            if exam_no not in caosce_results:
                caosce_results[exam_no] = {
                    "MAT NO.": exam_no,
                    "FULL NAME": None,
                }
                # Initialize all station scores to 0.00
                for sk in station_keys:
                    caosce_results[exam_no][STATION_COLUMN_MAP[sk]] = 0.00

            if full_name and not caosce_results[exam_no]["FULL NAME"]:
                caosce_results[exam_no]["FULL NAME"] = full_name

            score_val = None
            if station_key == "viva" and viva_score_col:
                score_val = numeric_safe(row.get(viva_score_col))
            elif grade_col:
                score_val = numeric_safe(row.get(grade_col))

            if score_val is not None:
                out_col = STATION_COLUMN_MAP[station_key]
                if station_key == "viva":
                    caosce_results[exam_no][out_col] = round(score_val, 2)
                else:
                    caosce_results[exam_no][out_col] = round(score_val, 2)
                
                station_scores.append(caosce_results[exam_no][out_col])

            rows_added += 1

        # Calculate average for this station
        if station_scores:
            station_avg = sum(station_scores) / len(station_scores)
            station_overall_averages[station_key] = round(station_avg, 2)

        logger.info(f"Processed {fname} → {rows_added} rows (Station: {station_key})")
    
    return caosce_results, station_max_scores, station_overall_averages, all_exam_numbers, caosce_upgrade_count

def calculate_osce_percentage(student_scores, station_max_scores):
    """Calculate OSCE percentage"""
    # Sum all 6 stations
    procedure_question_stations = [
        "procedure_station_one",
        "procedure_station_three", 
        "procedure_station_five",
        "question_station_two",
        "question_station_four",
        "question_station_six"
    ]
    
    station_scores_sum = 0
    for station_key in procedure_question_stations:
        score = student_scores.get(STATION_COLUMN_MAP[station_key], 0) or 0
        station_scores_sum += float(score)
    
    # Convert to 90%: (sum ÷ 60) × 90
    if station_scores_sum > 0:
        stations_percentage = (station_scores_sum / 60.0) * 90.0
    else:
        stations_percentage = 0
    
    # Add VIVA/10 (already 10%)
    viva_score = student_scores.get(STATION_COLUMN_MAP["viva"], 0) or 0
    viva_percentage = float(viva_score)
    
    # Total OSCE percentage
    total_percentage = round(stations_percentage + viva_percentage, 2)
    
    return total_percentage

def merge_results(caosce_results, paper_results, station_max_scores, paper_upgrade_counts):
    """
    Merge CAOSCE and paper results into combined structure
    FIXED: Properly calculate average based on actual papers taken
    """
    combined_results = {}
    all_exam_numbers = set(caosce_results.keys()) | set(paper_results.keys())
    total_upgrades = sum(paper_upgrade_counts.values())
    
    for exam_no in all_exam_numbers:
        combined_results[exam_no] = {
            "MAT NO.": exam_no,
            "FULL NAME": None,
            "PAPER I": 0.00,
            "PAPER II": 0.00,
            "PAPER III": 0.00,
            "CAOSCE": 0.00,
            "OVERALL AVERAGE": 0.00,
            "REMARK": "",
            "FAILED PAPERS": "",
            "FAILED_COUNT": 0
        }
        
        # Add CAOSCE data
        if exam_no in caosce_results:
            combined_results[exam_no]["FULL NAME"] = caosce_results[exam_no]["FULL NAME"]
            
            # Calculate CAOSCE percentage
            caosce_percentage = calculate_osce_percentage(caosce_results[exam_no], station_max_scores)
            
            # Apply upgrade to CAOSCE score
            upgraded_caosce, was_upgraded = apply_score_upgrade(caosce_percentage)
            if was_upgraded:
                total_upgrades += 1
                
            combined_results[exam_no]["CAOSCE"] = upgraded_caosce
        
        # Add paper data  
        if exam_no in paper_results:
            paper_data = paper_results[exam_no]
            if not combined_results[exam_no]["FULL NAME"] and paper_data.get("FULL NAME"):
                combined_results[exam_no]["FULL NAME"] = paper_data["FULL NAME"]
                
            # Get already upgraded paper scores
            for paper in ["PAPER I", "PAPER II", "PAPER III"]:
                paper_score = float(paper_data.get(paper, 0.00) or 0.00)
                combined_results[exam_no][paper] = paper_score
        
        # CRITICAL FIX: Calculate overall average based on ACTUAL papers (4 papers total)
        paper_i = float(combined_results[exam_no]["PAPER I"] or 0)
        paper_ii = float(combined_results[exam_no]["PAPER II"] or 0)  
        paper_iii = float(combined_results[exam_no]["PAPER III"] or 0)
        caosce_score = float(combined_results[exam_no]["CAOSCE"] or 0)
        
        # For PRE-COUNCIL exams, we always have 4 papers: Paper I, II, III, and CAOSCE
        # Even if score is 0, it's still a paper that was taken
        total_score = paper_i + paper_ii + paper_iii + caosce_score
        overall_avg = total_score / 4.0  # Always divide by 4 for 4 papers
            
        combined_results[exam_no]["OVERALL AVERAGE"] = round(overall_avg, 2)
        
        # Determine REMARK and FAILED PAPERS
        remark, failed_papers, failed_count = determine_remark_and_failed_papers(
            paper_i, paper_ii, paper_iii, caosce_score
        )
        combined_results[exam_no]["REMARK"] = remark
        combined_results[exam_no]["FAILED PAPERS"] = failed_papers
        combined_results[exam_no]["FAILED_COUNT"] = failed_count
    
    return combined_results, total_upgrades

def sort_combined_results(combined_results):
    """Sort combined results"""
    results_list = [(exam_no, data) for exam_no, data in combined_results.items()]
    
    def sort_key(item):
        exam_no, data = item
        remark_order = 0 if data.get("REMARK") == "Passed" else 1
        failed_count = int(data.get("FAILED_COUNT", 0) or 0)
        
        # Extract numeric part from exam number for better sorting
        exam_str = str(exam_no).upper() if exam_no else ""
        numeric_part = 0
        
        if exam_str:
            match = re.search(r'(\d{2,})', exam_str)
            if match:
                try:
                    numeric_part = int(match.group(1))
                except (ValueError, TypeError):
                    numeric_part = 0
        
        return (remark_order, failed_count, numeric_part, exam_str)
    
    try:
        sorted_results = sorted(results_list, key=sort_key)
    except TypeError as e:
        logger.error(f"Sorting error: {e}")
        # Fallback: sort by exam number only
        sorted_results = sorted(results_list, key=lambda x: str(x[0]))
    
    return sorted_results

def create_caosce_sheet(wb, df_caosce, college_config, station_max_scores, station_overall_averages):
    """Create the CAOSCE Results sheet"""
    ws = wb.create_sheet("CAOSCE Results", 0)
    
    # Write data to worksheet
    for r in dataframe_to_rows(df_caosce, index=False, header=True):
        ws.append(r)
    
    data_end_row = apply_excel_formatting(ws, df_caosce, college_config, "CAOSCE", station_max_scores)
    return ws, data_end_row

def create_combined_sheet(wb, df_combined, college_config):
    """Create the Combined Results sheet"""
    ws = wb.create_sheet("Combined Results")
    
    # Write data to worksheet
    for r in dataframe_to_rows(df_combined, index=False, header=True):
        ws.append(r)
    
    data_end_row = apply_excel_formatting(ws, df_combined, college_config, "COMBINED")
    return ws, data_end_row

def apply_excel_formatting(ws, df, college_config, sheet_type, station_max_scores=None):
    """Apply consistent Excel formatting to worksheets"""
    TITLE_ROWS = 6
    ws.insert_rows(1, TITLE_ROWS)
    header_row = TITLE_ROWS + 1

    last_col_letter = get_column_letter(ws.max_column)

    # Add college-specific logo
    logo_path = None
    base_logo_name = college_config["logo"]
    
    for ext in ['.png', '.jpg', '.jpeg']:
        potential_path = os.path.join(LOGO_BASE_PATH, base_logo_name.replace('.png', ext))
        if os.path.exists(potential_path):
            logo_path = potential_path
            break
    
    if not logo_path:
        exact_path = os.path.join(LOGO_BASE_PATH, base_logo_name)
        if os.path.exists(exact_path):
            logo_path = exact_path
    
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 120
            img.height = 120
            ws.add_image(img, "A1")
            logger.info(f"✓ Added logo: {os.path.basename(logo_path)}")
        except Exception as e:
            logger.warning(f"Could not add logo {logo_path}: {e}")

    # College-specific titles
    ws.merge_cells(f"B1:{last_col_letter}1")
    ws["B1"] = college_config["name"]
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"B2:{last_col_letter}2")
    ws["B2"] = "PRE-COUNCIL EXAMINATION RESULT"
    ws["B2"].font = SUBTITLE_FONT
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f"B3:{last_col_letter}3")
    exam_type = "COMBINED" if sheet_type == "COMBINED" else "CAOSCE"
    ws["B3"] = f"{exam_type}_{CURRENT_YEAR}"
    ws["B3"].font = Font(bold=True, size=12, name="Calibri", color="1F4E78")
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # DATE
    ws.merge_cells(f"B4:{last_col_letter}4")
    ws.cell(row=4, column=2, value=f"Date: {datetime.now().strftime('%d %B %Y')}")
    ws.cell(row=4, column=2).font = DATE_FONT
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 25

    # CLASS
    ws.merge_cells(f"A5:{last_col_letter}5")
    class_cell = ws.cell(row=5, column=1, value="CLASS: _________________________________________________________________________________________")
    class_cell.font = Font(bold=True, size=11, name="Calibri", color="1F4E78")
    class_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 28

    # Empty row for spacing
    ws.row_dimensions[6].height = 18

    # Header styling
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.fill = HEADER_FILL
    ws.row_dimensions[header_row].height = 20

    # Update MAT NO. header to college-specific label
    mat_no_idx = 2
    full_name_idx = 3
    ws.cell(row=header_row, column=mat_no_idx).value = college_config["mat_no_label"]
    ws.cell(row=header_row, column=mat_no_idx).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=header_row, column=full_name_idx).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    ws.freeze_panes = f"A{header_row + 1}"

    # Determine column indices based on sheet type
    if sheet_type == "COMBINED":
        # Find paper columns dynamically
        paper_columns = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value and ("PAPER" in str(cell_value).upper() or "MIDWIFERY" in str(cell_value).upper() or "CAOSCE" in str(cell_value).upper()):
                paper_columns[col_idx] = cell_value
        
        # Find other column indices
        overall_idx = None
        remark_idx = None
        failed_papers_idx = None
        
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).upper()
                if "OVERALL" in cell_str:
                    overall_idx = col_idx
                elif "REMARK" in cell_str:
                    remark_idx = col_idx
                elif "FAILED" in cell_str:
                    failed_papers_idx = col_idx
        
        # Format data rows for combined sheet
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            is_avg_row = (row[1].value == "OVERALL AVERAGE")
            
            for cell in row:
                if cell.column == 1:  # S/N
                    if is_avg_row:
                        cell.value = ""
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=10, name="Calibri")
                elif cell.column == mat_no_idx:  # MAT NO./EXAM NO.
                    if is_avg_row:
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif cell.column == full_name_idx:  # FULL NAME
                    if is_avg_row:
                        cell.value = ""
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif cell.column in paper_columns:  # Paper scores
                    if is_avg_row:
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    elif cell.value == 0 or cell.value == 0.0 or cell.value is None:
                        cell.value = 0.00
                        cell.fill = NO_SCORE_FILL
                        cell.font = NO_SCORE_FONT
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column == overall_idx:  # OVERALL AVERAGE
                    cell.number_format = "0.00"
                    if is_avg_row:
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.font = Font(bold=True, size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column == remark_idx:  # REMARK
                    if is_avg_row:
                        cell.value = ""
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    elif cell.value == "Passed":
                        cell.font = PASS_FONT
                        cell.fill = PASS_FILL
                    elif cell.value == "Failed":
                        cell.font = FAIL_FONT
                        cell.fill = FAIL_FILL
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column == failed_papers_idx:  # FAILED PAPERS
                    if is_avg_row:
                        cell.value = ""
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    elif cell.value:
                        cell.font = FAILED_PAPERS_FONT
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=10, name="Calibri")
    else:
        # CAOSCE sheet formatting
        score_cols_count = len([col for col in df.columns if "Score/" in col or "VIVA/" in col or "(15%)" in str(col) or "(10%)" in str(col)])
        station_col_indices = list(range(4, 4 + score_cols_count))
        total_col_idx = 4 + score_cols_count
        percent_col_idx = total_col_idx + 1

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            is_avg_row = (row[1].value == "OVERALL AVERAGE")
            
            for cell in row:
                if cell.column == 1:  # S/N
                    if is_avg_row:
                        cell.value = ""
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=10, name="Calibri")
                elif cell.column == mat_no_idx:  # MAT NO.
                    if is_avg_row:
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif cell.column == full_name_idx:  # FULL NAME
                    if is_avg_row:
                        cell.value = ""
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif cell.column in station_col_indices:  # Station scores
                    if is_avg_row:
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    elif cell.value == 0 or cell.value == 0.0 or cell.value is None:
                        cell.value = 0.00
                        cell.fill = NO_SCORE_FILL
                        cell.font = NO_SCORE_FONT
                    else:
                        cell.font = Font(size=10, name="Calibri")
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column in [total_col_idx, percent_col_idx]:  # Total and Percentage
                    if cell.column == percent_col_idx:
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.00"
                    if is_avg_row:
                        cell.font = AVERAGE_FONT
                        cell.fill = AVERAGE_FILL
                    else:
                        cell.font = Font(bold=True, size=10, name="Calibri")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(size=10, name="Calibri")

    # Apply autofit columns with improved sizing
    data_end_row = ws.max_row
    apply_autofit_columns(ws, header_row, data_end_row)

    return data_end_row

def generate_caosce_dataframe(caosce_results, station_max_scores, station_overall_averages):
    """Generate DataFrame for CAOSCE results"""
    caosce_results = copy.deepcopy(caosce_results)
    caosce_upgrade_count = 0
    
    score_cols = []
    for station_key in ["procedure_station_one", "procedure_station_three", "procedure_station_five",
                        "question_station_two", "question_station_four", "question_station_six", "viva"]:
        display_name = STATION_DISPLAY_NAMES[station_key]
        score_cols.append(display_name)

    # Create a new dictionary with the display column names
    processed_results = {}
    for exam_no, student_data in caosce_results.items():
        processed_results[exam_no] = {
            "MAT NO.": student_data["MAT NO."],
            "FULL NAME": student_data["FULL NAME"]
        }
        
        # Add station scores with display names
        for i, station_key in enumerate(["procedure_station_one", "procedure_station_three", "procedure_station_five",
                                         "question_station_two", "question_station_four", "question_station_six", "viva"]):
            base_col = STATION_COLUMN_MAP[station_key]
            score_value = student_data.get(base_col, 0.00)
            try:
                processed_results[exam_no][score_cols[i]] = float(score_value) if score_value is not None else 0.00
            except (ValueError, TypeError):
                processed_results[exam_no][score_cols[i]] = 0.00

    base_cols = ["MAT NO.", "FULL NAME"] + score_cols
    df_out = pd.DataFrame.from_dict(processed_results, orient="index")
    
    # Ensure all required columns exist
    for col in base_cols:
        if col not in df_out.columns:
            df_out[col] = 0.00

    df_out = df_out[base_cols]

    df_out["FULL NAME"] = df_out.groupby("MAT NO.")["FULL NAME"].transform(
        lambda x: x.fillna(method='ffill').fillna(method='bfill')
    )

    # Sort by exam number
    try:
        def extract_numeric(x):
            if pd.isna(x):
                return 0
            x_str = str(x)
            match = re.search(r'(\d+)', x_str)
            if match:
                try:
                    return int(match.group(1))
                except (ValueError, TypeError):
                    return 0
            return 0
        
        df_out["__sort"] = df_out["MAT NO."].apply(extract_numeric)
        df_out["__sort_str"] = df_out["MAT NO."].astype(str)
        df_out.sort_values(["__sort", "__sort_str"], inplace=True)
        df_out.drop(columns=["__sort", "__sort_str"], inplace=True)
    except Exception as e:
        logger.error(f"Sorting error in generate_caosce_dataframe: {e}")
        df_out.sort_values("MAT NO.", inplace=True)
    
    df_out.reset_index(drop=True, inplace=True)

    df_out.insert(0, "S/N", range(1, len(df_out) + 1))

    df_out[score_cols] = df_out[score_cols].apply(pd.to_numeric, errors="coerce")
    for col in score_cols:
        df_out[col] = df_out[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else 0.00)

    # Calculate OSCE score
    df_out["OSCE Total Score"] = 0.00
    df_out["OSCE Percentage (%)"] = 0
    
    for idx, row in df_out.iterrows():
        # Sum all 6 stations
        stations_sum = 0
        station_columns = ["PS ONE (/10) (15%)", "PS THREE (/10) (15%)", "PS FIVE (/10) (15%)",
                          "QS TWO (/10) (15%)", "QS FOUR (/10) (15%)", "QS SIX (/10) (15%)"]
        for col in station_columns:
            stations_sum += float(row[col]) if pd.notna(row[col]) else 0
        
        # Convert to 90%: (sum ÷ 60) × 90
        stations_percentage = (stations_sum / 60.0) * 90.0 if stations_sum > 0 else 0
        
        # Add VIVA/10
        viva_score = float(row["VIVA (/10) (10%)"]) if pd.notna(row["VIVA (/10) (10%)"]) else 0
        
        # Total OSCE percentage
        total_osce = stations_percentage + viva_score
        
        # Apply upgrade if enabled
        upgraded_score, was_upgraded = apply_score_upgrade(total_osce)
        if was_upgraded:
            caosce_upgrade_count += 1
            total_osce = upgraded_score
        
        df_out.at[idx, "OSCE Total Score"] = round(total_osce, 2)
        df_out.at[idx, "OSCE Percentage (%)"] = int(round(total_osce, 0))

    final_display_cols = ["S/N", "MAT NO.", "FULL NAME"] + score_cols + ["OSCE Total Score", "OSCE Percentage (%)"]
    df_out = df_out[final_display_cols]

    # Add overall average row
    avg_row = {
        "S/N": "",
        "MAT NO.": "OVERALL AVERAGE",
        "FULL NAME": "",
    }
    
    for i, station_key in enumerate(["procedure_station_one", "procedure_station_three", "procedure_station_five",
                                    "question_station_two", "question_station_four", "question_station_six", "viva"]):
        col_name = score_cols[i]
        avg_row[col_name] = float(station_overall_averages.get(station_key, 0.00))
    
    # Calculate average weighted total
    avg_stations_sum = 0
    station_avg_columns = ["procedure_station_one", "procedure_station_three", "procedure_station_five",
                          "question_station_two", "question_station_four", "question_station_six"]
    for station_key in station_avg_columns:
        station_avg = station_overall_averages.get(station_key, 0.00)
        avg_stations_sum += station_avg
    
    avg_stations_percentage = (avg_stations_sum / 60.0) * 90.0 if avg_stations_sum > 0 else 0
    avg_viva_score = station_overall_averages.get("viva", 0.00)
    avg_total_osce = avg_stations_percentage + avg_viva_score
    
    avg_row["OSCE Total Score"] = round(avg_total_osce, 2)
    avg_row["OSCE Percentage (%)"] = int(round(avg_total_osce, 0))

    df_out = pd.concat([df_out, pd.DataFrame([avg_row])], ignore_index=True)

    return df_out, caosce_upgrade_count

# ---------------------------
# Main Processing Function
# ---------------------------

def process_files():
    """Main function to process all files and generate ONE workbook with TWO sheets"""
    logger.info("Starting Enhanced CAOSCE Pre-Council Results Cleaning...")
    logger.info(f"Processing year: CAOSCE_{CURRENT_YEAR}")
    
    if UPGRADE_THRESHOLD > 0:
        logger.info(f"✅ UPGRADE ENABLED: {UPGRADE_THRESHOLD}-49 → 50")
    else:
        logger.info("ℹ️ No upgrades - strict grading mode")

    RAW_DIR = DEFAULT_RAW_DIR
    BASE_CLEAN_DIR = DEFAULT_CLEAN_DIR

    ts = datetime.now().strftime(TIMESTAMP_FMT)
    
    # Get all files in raw directory
    files = [f for f in os.listdir(RAW_DIR) if f.lower().endswith((".xlsx", ".xls", ".csv"))]

    if not files:
        logger.error(f"No raw files found in {RAW_DIR}")
        return

    logger.info(f"Found {len(files)} files to process")
    
    # Process CAOSCE station files
    caosce_results, station_max_scores, station_overall_averages, caosce_exam_numbers, caosce_upgrade_count = process_caosce_station_files(files, RAW_DIR)
    
    # Process Paper I, II, and III files
    paper_results, paper_upgrade_counts = process_paper_files(files, RAW_DIR)
    
    # Check for valid data
    if not caosce_results and not paper_results:
        logger.error("No valid data found in any files")
        return
    elif not caosce_results and paper_results:
        logger.warning("No CAOSCE files found, but paper files exist. Processing paper files only.")
    elif caosce_results and not paper_results:
        logger.warning("No paper files found, but CAOSCE files exist. Processing CAOSCE files only.")

    # Detect college using all exam numbers
    all_exam_numbers = caosce_exam_numbers | set(paper_results.keys())
    college_key, college_config = detect_college_from_exam_numbers(all_exam_numbers)
    
    logger.info(f"Detected college: {college_config['name']}")
    logger.info(f"Using logo: {college_config['logo']}")
    logger.info(f"Exam number label: {college_config['mat_no_label']}")

    # Create college-specific output directory
    output_dir = os.path.join(BASE_CLEAN_DIR, f"{college_config['output_prefix']}_COMBINED_{ts}")
    os.makedirs(output_dir, exist_ok=True)

    # Generate only ONE workbook with TWO sheets
    combined_output = generate_combined_output(caosce_results, paper_results, station_max_scores, station_overall_averages,
                                             paper_upgrade_counts, caosce_upgrade_count, college_config, output_dir, ts)

    # Print summary
    logger.info("\n" + "="*50)
    logger.info("PROCESSING COMPLETE")
    logger.info("="*50)
    
    if combined_output:
        logger.info(f"✓ Combined Results (2 sheets): {os.path.basename(combined_output)}")
    
    logger.info(f"📁 Output directory: {output_dir}")
    
    caosce_count = len(caosce_results) if caosce_results else 0
    paper_count = len(paper_results) if paper_results else 0
    combined_count = len(set(caosce_results.keys()) | set(paper_results.keys())) if caosce_results or paper_results else 0
    
    logger.info(f"📊 Students processed: CAOSCE={caosce_count}, Papers={paper_count}, Combined={combined_count}")
    
    # Print upgrade summary
    total_upgrades = caosce_upgrade_count + sum(paper_upgrade_counts.values())
    if total_upgrades > 0:
        logger.info(f"⬆️ Total upgrades applied: {total_upgrades}")
        if paper_upgrade_counts:
            for paper, count in paper_upgrade_counts.items():
                if count > 0:
                    logger.info(f"  • {paper}: {count} upgrade(s)")
        if caosce_upgrade_count > 0:
            logger.info(f"  • CAOSCE: {caosce_upgrade_count} upgrade(s)")

def generate_combined_output(caosce_results, paper_results, station_max_scores, station_overall_averages,
                           paper_upgrade_counts, caosce_upgrade_count, college_config, output_dir, timestamp):
    """Generate ONE workbook with TWO sheets: CAOSCE Results and Combined Results"""
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Process CAOSCE sheet if we have CAOSCE data
    total_upgrades = caosce_upgrade_count + sum(paper_upgrade_counts.values())
    
    if caosce_results:
        caosce_df, caosce_upgrade_count_from_df = generate_caosce_dataframe(caosce_results, station_max_scores, station_overall_averages)
        caosce_upgrade_count = caosce_upgrade_count_from_df
        total_upgrades = caosce_upgrade_count + sum(paper_upgrade_counts.values())
        
        ws_caosce, data_end_row_caosce = create_caosce_sheet(wb, caosce_df, college_config, station_max_scores, station_overall_averages)
        
        # Calculate statistics for CAOSCE sheet
        student_percentages = caosce_df["OSCE Percentage (%)"].iloc[:-1]
        total_students_caosce = len(student_percentages)
        avg_percentage_caosce = round(student_percentages.mean(), 1) if total_students_caosce > 0 else 0
        highest_percentage_caosce = round(student_percentages.max(), 1) if total_students_caosce > 0 else 0
        lowest_percentage_caosce = round(student_percentages.min(), 1) if total_students_caosce > 0 else 0
        
        # Total max score for OSCE is always 100%
        total_max_score_caosce = 100.0
        
        # Add documentation to CAOSCE sheet
        create_document_sections(
            ws_caosce, total_students_caosce, avg_percentage_caosce, highest_percentage_caosce, lowest_percentage_caosce,
            total_max_score_caosce, data_end_row_caosce, college_config, "CAOSCE", caosce_upgrade_count
        )
    else:
        logger.warning("No CAOSCE data available for CAOSCE sheet")
    
    # Process Combined sheet if we have any data
    if caosce_results or paper_results:
        # Merge results for combined sheet
        if caosce_results and paper_results:
            combined_results, combined_upgrades = merge_results(caosce_results, paper_results, station_max_scores, paper_upgrade_counts)
            # Update total upgrades
            if combined_upgrades > 0:
                total_upgrades = combined_upgrades
        else:
            # Handle case where we only have one type of data
            combined_results = {}
            combined_upgrades = 0
            
        # Sort the combined results
        sorted_results = sort_combined_results(combined_results)
        
        # Create DataFrame from sorted results
        if sorted_results:
            sorted_data = []
            for exam_no, data in sorted_results:
                row_data = {k: v for k, v in data.items() if k != "FAILED_COUNT"}
                sorted_data.append(row_data)
            
            df_combined = pd.DataFrame(sorted_data)
        else:
            df_combined = pd.DataFrame()
            
        if not df_combined.empty:
            # Define column order
            column_order = ["MAT NO.", "FULL NAME", "PAPER I", "PAPER II", "PAPER III", "CAOSCE", "OVERALL AVERAGE", "REMARK", "FAILED PAPERS"]
            existing_columns = [col for col in column_order if col in df_combined.columns]
            df_combined = df_combined[existing_columns]
            
            # Rename columns for display with /100 notation
            rename_dict = {}
            for col in ["PAPER I", "PAPER II", "PAPER III", "CAOSCE", "OVERALL AVERAGE"]:
                if col in df_combined.columns:
                    rename_dict[col] = f"{col}/100"
            df_combined.rename(columns=rename_dict, inplace=True)
            
        # Add S/N column at the beginning
        if not df_combined.empty:
            df_combined.insert(0, "S/N", range(1, len(df_combined) + 1))
        else:
            df_combined = pd.DataFrame()

        # Calculate overall averages for the average row
        if not df_combined.empty:
            # Ensure numeric conversion
            for col in ["PAPER I/100", "PAPER II/100", "PAPER III/100", "CAOSCE/100", "OVERALL AVERAGE/100"]:
                if col in df_combined.columns:
                    df_combined[col] = pd.to_numeric(df_combined[col], errors='coerce').fillna(0)
            
            # Calculate averages for each paper
            paper_i_avg = df_combined["PAPER I/100"].mean() if "PAPER I/100" in df_combined.columns else 0
            paper_ii_avg = df_combined["PAPER II/100"].mean() if "PAPER II/100" in df_combined.columns else 0
            paper_iii_avg = df_combined["PAPER III/100"].mean() if "PAPER III/100" in df_combined.columns else 0
            caosce_avg = df_combined["CAOSCE/100"].mean() if "CAOSCE/100" in df_combined.columns else 0
            
            # CRITICAL FIX: Calculate overall average of ALL papers (4 papers)
            # Calculate average of column averages
            paper_averages = []
            if paper_i_avg > 0:
                paper_averages.append(paper_i_avg)
            if paper_ii_avg > 0:
                paper_averages.append(paper_ii_avg)
            if paper_iii_avg > 0:
                paper_averages.append(paper_iii_avg)
            if caosce_avg > 0:
                paper_averages.append(caosce_avg)
            
            if paper_averages:
                total_avg = sum(paper_averages) / len(paper_averages)
            else:
                total_avg = 0

            # Add overall average row
            avg_row = {
                "S/N": "",
                "MAT NO.": "OVERALL AVERAGE",
                "FULL NAME": "",
            }
            
            # Add paper averages if columns exist
            if "PAPER I/100" in df_combined.columns:
                avg_row["PAPER I/100"] = round(paper_i_avg, 2)
            if "PAPER II/100" in df_combined.columns:
                avg_row["PAPER II/100"] = round(paper_ii_avg, 2)
            if "PAPER III/100" in df_combined.columns:
                avg_row["PAPER III/100"] = round(paper_iii_avg, 2)
            if "CAOSCE/100" in df_combined.columns:
                avg_row["CAOSCE/100"] = round(caosce_avg, 2)
            if "OVERALL AVERAGE/100" in df_combined.columns:
                avg_row["OVERALL AVERAGE/100"] = round(total_avg, 2)
            
            # Empty values for REMARK and FAILED PAPERS in average row
            avg_row["REMARK"] = ""
            avg_row["FAILED PAPERS"] = ""
            
            df_combined = pd.concat([df_combined, pd.DataFrame([avg_row])], ignore_index=True)

            # Calculate statistics for documentation
            total_students = len(df_combined) - 1  # Exclude average row
            
            # Calculate statistics correctly
            student_data = df_combined[df_combined["MAT NO."] != "OVERALL AVERAGE"]
            if not student_data.empty and "OVERALL AVERAGE/100" in student_data.columns:
                student_overall = pd.to_numeric(student_data["OVERALL AVERAGE/100"], errors='coerce').fillna(0)
                avg_percentage = round(student_overall.mean(), 2) if not student_overall.empty else 0
                highest_percentage = round(student_overall.max(), 2) if not student_overall.empty else 0
                lowest_percentage = round(student_overall.min(), 2) if not student_overall.empty else 0
            else:
                avg_percentage = 0
                highest_percentage = 0
                lowest_percentage = 0

            # Calculate total possible score based on number of papers
            paper_count = 0
            for col in df_combined.columns:
                if "PAPER" in str(col).upper() or "CAOSCE" in str(col).upper() or "MIDWIFERY" in str(col).upper():
                    paper_count += 1
            
            total_possible_score = paper_count * 100

            # Add Combined sheet
            ws_combined, data_end_row_combined = create_combined_sheet(wb, df_combined, college_config)
            
            # Add documentation to Combined sheet with upgrade information
            create_document_sections(
                ws_combined, total_students, avg_percentage, highest_percentage, lowest_percentage,
                total_possible_score, data_end_row_combined, college_config, "COMBINED", 
                total_upgrades, paper_upgrade_counts
            )
        else:
            logger.error("No data available for Combined sheet")
            return None
    else:
        logger.error("No data available for any sheet")
        return None

    # Save the workbook
    output_basename = f"{college_config['output_prefix']}_PRE_COUNCIL_CLEANED"
    out_xlsx = os.path.join(output_dir, f"{output_basename}_{timestamp}.xlsx")
    wb.save(out_xlsx)
    
    logger.info(f"✓ Saved combined results: {os.path.basename(out_xlsx)}")
    return out_xlsx

if __name__ == "__main__":
    process_files()