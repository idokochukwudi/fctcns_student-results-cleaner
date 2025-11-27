#!/usr/bin/env python3
"""
utme_result.py

Robust UTME/PUTME cleaning and reporting script (no matplotlib required).

Place raw files (CSV/XLS/XLSX) into:
  PROCESS_RESULT/PUTME_RESULT/RAW_PUTME_RESULT (or specify via --input-dir)

Put raw candidate batches (optional) into:
  PROCESS_RESULT/PUTME_RESULT/RAW_CANDIDATE_BATCHES (or specify via --candidate-dir)

Outputs cleaned CSV + formatted XLSX are saved to:
  PROCESS_RESULT/PUTME_RESULT/CLEAN_PUTME_RESULT/UTME_RESULT-<timestamp> (or specify via --output-dir)

Combined output for all batches is saved as:
  PROCESS_RESULT/PUTME_RESULT/CLEAN_PUTME_RESULT/UTME_RESULT-<timestamp>/PUTME_COMBINE_RESULT_<timestamp>.xlsx

Unsorted combined output is saved as:
  PROCESS_RESULT/PUTME_RESULT/CLEAN_PUTME_RESULT/UTME_RESULT-<timestamp>/PUTME_COMBINE_RESULT_UNSORTED_<timestamp>.xlsx

Features:
 - Maps common header variations to canonical names
 - Drops unwanted columns (Username, Department, State, Started on, Completed, Time taken)
 - Detects Grade/... column and renames to Score/...
 - Adds Score/100 (0 decimals) in individual and sorted combined results, and Score/100.00 (2 decimals) in unsorted result
 - Removes 'overall average' rows and invalid rows
 - Sorts combined result by STATE (A->Z) then Score (Z->A highest first)
 - Unsorted result preserves raw input order from RAW_PUTME_RESULT with columns: S/N, FULL NAME, APPLICATION ID, PHONE NUMBER, STATE, Score/100.00 (2 decimals), without state-based coloring or green highlighting for scores >= pass threshold
 - Adds batch-specific worksheets (e.g., Batch1A, Batch1B) in unsorted result, each with a bold title row (e.g., BATCH1A UNSORTED), unsorted data, same columns as Unsorted Results, and an overall average row for Score/100.00
 - Adds an overall average row for Score/100.00 in the Unsorted Results sheet
 - Adds S/N (global) and STATE_SN (restarts per state, excluded in unsorted result)
 - Cleans phone numbers (remove .0, non-digits, ensure leading 0)
 - Highlights passes (>= PASS_THRESHOLD) in green for individual and sorted combined results
 - Applies a soft, pastel row color per STATE (except in unsorted result)
 - Produces Analysis sheet and clear, management-friendly Charts with a legend table
 - Lists registered-but-absent candidates in Absent sheet with Batch column, batch-specific for individual files
 - Detects and reports candidates appearing in results of a different batch than their registered batch
 - Combines all batches into a single PUTME_COMBINE_RESULT file with Rebatched sheet
 - Generates unsorted combined result with specified columns, batch-specific worksheets, and overall averages
 - Enhanced logging for invalid APPLICATION IDs, batch-specific absent candidates, and batch mismatches
 - Supports batch IDs with or without spaces (e.g., Batch1A or Batch 1A)
 - Ensures absent counts reflects the exact difference between registered candidates and those who sat
"""

import os
import sys
import re
import math
import hashlib
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import argparse

# ---------------------------
# Directory configuration (defaults)
# ---------------------------
DEFAULT_BASE_DIR = "/mnt/c/Users/MTECH COMPUTERS/Documents/PROCESS_RESULT/PUTME_RESULT"
DEFAULT_RAW_DIR = os.path.join(DEFAULT_BASE_DIR, "RAW_PUTME_RESULT")
DEFAULT_CANDIDATE_DIR = os.path.join(DEFAULT_BASE_DIR, "RAW_CANDIDATE_BATCHES")
DEFAULT_CLEAN_DIR = os.path.join(DEFAULT_BASE_DIR, "CLEAN_PUTME_RESULT")
DEFAULT_PASS_THRESHOLD = 50.0
TIMESTAMP_FMT = "%Y-%m-%d_%H:%M:%S"


def parse_args():
    """Parse command-line arguments with defaults."""
    parser = argparse.ArgumentParser(description="UTME/PUTME Results Cleaning Script")
    parser.add_argument(
        "--input-dir",
        default=DEFAULT_RAW_DIR,
        help=f"Input directory for raw files (default: {DEFAULT_RAW_DIR})",
    )
    parser.add_argument(
        "--candidate-dir",
        default=DEFAULT_CANDIDATE_DIR,
        help=f"Directory for candidate batch files (default: {DEFAULT_CANDIDATE_DIR})",
    )
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_CLEAN_DIR,
        help=f"Output directory for cleaned files (default: {DEFAULT_CLEAN_DIR})",
    )
    parser.add_argument(
        "--pass-threshold",
        type=float,
        default=DEFAULT_PASS_THRESHOLD,
        help=f"Pass threshold for highlighting (default: {DEFAULT_PASS_THRESHOLD})",
    )
    parser.add_argument(
        "--batch-id",
        help="Specific batch ID to process (e.g., 'Batch1A' or 'Batch 1A')",
    )
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="Skip interactive prompts (e.g., for converted score)",
    )
    parser.add_argument(
        "--converted-score-max",
        type=int,
        help="Target maximum for converted score (e.g., 60 for Score/60%) when non-interactive",
    )
    return parser.parse_args()


# ---------------------------
# Soft pastel color palette for states and charts
# ---------------------------
PALETTE = [
    "A8E6CF",  # Soft Green
    "D7B9D5",  # Soft Purple
    "FFCCCB",  # Soft Red
    "B3CDE0",  # Soft Blue
    "F4E1A4",  # Soft Yellow
    "D1E8E2",  # Soft Mint
    "E4C1F9",  # Soft Lavender
    "F8E9A1",  # Soft Cream
    "A2D2FF",  # Soft Sky
    "D9BF77",  # Soft Beige
    "C6D4E1",  # Soft Grey-Blue
    "F7D6E0",  # Soft Pink
    "B5EAD7",  # Soft Seafoam
    "E2D1F9",  # Soft Lilac
    "F9AFA4",  # Soft Coral
    "B8E1FF",  # Soft Azure
    "D8E2DC",  # Soft Neutral
    "FFE5D9",  # Soft Peach
    "A1C7E0",  # Soft Slate
    "D4A5A5",  # Soft Rose
]


def state_color_for(state_name):
    """Deterministic pastel fill for a state label."""
    if state_name is None:
        state_name = ""
    key = state_name.strip().lower().encode("utf8")
    h = int(hashlib.md5(key).hexdigest(), 16)
    color = PALETTE[h % len(PALETTE)]
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# ---------------------------
# Helpers
# ---------------------------
def find_column_by_names(df, candidate_names):
    norm_map = {
        col: re.sub(r"\s+", " ", str(col).strip().lower()) for col in df.columns
    }
    candidates = [re.sub(r"\s+", " ", c.strip().lower()) for c in candidate_names]
    for cand in candidates:
        for col, ncol in norm_map.items():
            if ncol == cand:
                return col
    for cand in candidates:
        candc = cand.replace(" ", "")
        for col, ncol in norm_map.items():
            if ncol.replace(" ", "") == candc:
                return col
    return None


def find_grade_column(df):
    for col in df.columns:
        if str(col).strip().lower().startswith("grade/"):
            return col
    for col in df.columns:
        if re.search(r"grade|score", str(col), flags=re.I):
            return col
    return None


def to_numeric_safe(series):
    return pd.to_numeric(
        series.astype(str).str.replace(",", "").str.strip(), errors="coerce"
    )


def clean_phone_value(s):
    if pd.isna(s):
        return ""
    st = str(s).strip()
    if st.lower() in ("nan", "none", ""):
        return ""
    st = re.sub(r"\.0+$", "", st)
    digits = re.sub(r"\D", "", st)
    if digits == "":
        return ""
    if digits.startswith("234") and len(digits) > 3:
        digits = "0" + digits[3:]
    if not digits.startswith("0"):
        digits = "0" + digits
    return digits


def drop_overall_average_rows(df):
    mask = df.apply(
        lambda row: row.astype(str)
        .str.contains("overall average", case=False, na=False)
        .any(),
        axis=1,
    )
    if mask.any():
        return df[~mask].copy()
    return df


def auto_column_width(ws, min_width=8, max_width=60):
    for i, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            if cell.value is not None:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[get_column_letter(i)].width = min(
            max_width, max(min_width, max_len + 2)
        )


def normalize_id(s):
    if s is None:
        return None
    return str(s).strip().lower()


# ---------------------------
# Candidate batches helpers
# ---------------------------
def load_candidate_batches(folder, batch_id=None):
    """Load candidate-batch files. If batch_id is provided, load only the matching batch file; otherwise, load all with batch IDs."""
    if not os.path.isdir(folder):
        print(f"Warning: Candidate batch directory {folder} does not exist.")
        return (
            pd.DataFrame(
                columns=["EXAM_NO", "FULL_NAME", "PHONE NUMBER", "STATE", "BATCH_ID"]
            ),
            {},
            {},
        )

    files = [
        f
        for f in os.listdir(folder)
        if f.lower().endswith((".csv", ".xlsx", ".xls")) and not f.startswith("~$")
    ]
    if not files:
        print(f"Warning: No candidate batch files found in {folder}.")
        return (
            pd.DataFrame(
                columns=["EXAM_NO", "FULL_NAME", "PHONE NUMBER", "STATE", "BATCH_ID"]
            ),
            {},
            {},
        )

    batch_id_map = {}  # Maps normalized ID to registered BATCH_ID
    numeric_to_full = {}  # Maps numeric token to original EXAM_NO
    rows = []
    for fname in sorted(files):
        path = os.path.join(folder, fname)
        batch_id_match = re.search(r"(Batch\s*\d+[A-Za-z]?)", fname, re.IGNORECASE)
        current_batch_id = batch_id_match.group(1) if batch_id_match else fname
        if batch_id:
            batch_id_normalized = re.sub(r"\s+", "", batch_id.lower())
            if batch_id_normalized not in re.sub(r"\s+", "", fname.lower()):
                continue
        print(f"Loading candidate batch file: {fname} (Batch ID: {current_batch_id})")
        try:
            if fname.lower().endswith(".csv"):
                cdf = pd.read_csv(path, dtype=str)
            else:
                cdf = pd.read_excel(path, dtype=str)
        except Exception as e:
            print(f"Error reading candidate batch {fname}: {e}")
            continue
        exam_col = find_column_by_names(
            cdf, ["username", "exam no", "reg no", "mat no", "regnum", "reg number"]
        )
        name_col = find_column_by_names(
            cdf,
            [
                "firstname",
                "full name",
                "name",
                "candidate name",
                "user full name",
                "RG_CANDNAME",
            ],
        )
        phone_col = find_column_by_names(
            cdf, ["phone1", "phone", "phone number", "PHONE", "Mobile", "PhoneNumber", "Phone No"]
        )
        state_col = find_column_by_names(
            cdf,
            [
                "city",
                "city/town",
                "City /town",
                "City",
                "Town",
                "State of Origin",
                "STATE",
            ],
        )
        for _, r in cdf.iterrows():
            ex = None
            name = None
            phone = None
            state = None
            if exam_col and pd.notna(r.get(exam_col)):
                ex = str(r.get(exam_col)).strip()
            if name_col and pd.notna(r.get(name_col)):
                name = str(r.get(name_col)).strip()
            if phone_col and pd.notna(r.get(phone_col)):
                phone = str(r.get(phone_col)).strip()
            if state_col and pd.notna(r.get(state_col)):
                state = str(r.get(state_col)).strip()
            if (not name) and ex and re.search(r"[A-Za-z]{2,}", ex):
                m = re.search(r"(.+?)\s+(\d{3,})$", ex)
                if m:
                    name = m.group(1).strip()
                    ex = m.group(2).strip()
            if not ex or not re.search(r"\d", ex):
                for c in cdf.columns:
                    v = r.get(c)
                    if pd.isna(v):
                        continue
                    s = str(v).strip()
                    m = re.search(r"\b(\d{3,})\b", s)
                    if m and not ex:
                        ex = m.group(1)
                    if (
                        not name
                        and re.search(r"[A-Za-z]{2,}", s)
                        and not re.fullmatch(r"\d+", s)
                    ):
                        name = s
                    if not phone and re.search(r"\d{10,}", s):
                        phone = s
                    if (
                        not state
                        and re.search(r"[A-Za-z]{2,}", s)
                        and c.lower() in ["city", "state"]
                    ):
                        state = s
            if ex:
                rows.append(
                    {
                        "EXAM_NO": ex,
                        "FULL_NAME": name or "",
                        "PHONE NUMBER": phone or "",
                        "STATE": state or "",
                        "BATCH_ID": current_batch_id,
                    }
                )
                norm_id = normalize_id(ex)
                batch_id_map[norm_id] = current_batch_id
                tok = extract_numeric_token(ex)
                if tok:
                    numeric_to_full[tok] = ex
    if not rows:
        print(
            f"No valid candidate records found in batch files{' for ' + batch_id if batch_id else ''}."
        )
        return (
            pd.DataFrame(
                columns=["EXAM_NO", "FULL_NAME", "PHONE NUMBER", "STATE", "BATCH_ID"]
            ),
            {},
            {},
        )

    cdf_all = pd.DataFrame(rows)
    duplicates = cdf_all[cdf_all.duplicated(subset=["EXAM_NO"], keep=False)]
    if not duplicates.empty:
        print(
            f"Found {len(duplicates)} duplicate EXAM_NO entries in candidate batches{' for ' + batch_id if batch_id else ''}:"
        )
        print(duplicates[["EXAM_NO", "FULL_NAME", "BATCH_ID"]].head().to_string())
    cdf_all = cdf_all.drop_duplicates(subset=["EXAM_NO"], keep="first")
    print(
        f"Loaded {len(cdf_all)} unique registered candidates from batch files{' for ' + batch_id if batch_id else ''}."
    )
    if "PHONE NUMBER" in cdf_all.columns:
        cdf_all["PHONE NUMBER"] = cdf_all["PHONE NUMBER"].apply(clean_phone_value)
    return cdf_all, batch_id_map, numeric_to_full


# ---------------------------
# Extract numeric tokens helper
# ---------------------------
def extract_numeric_token(s):
    """Return first sequence of 3+ digits found in string, else None."""
    if s is None:
        return None
    s = re.sub(r"\W+", "", str(s).strip().lower())
    m = re.search(r"(\d{3,})", s)
    return m.group(1) if m else None


# ---------------------------
# Format Excel sheet helper
# ---------------------------
def format_excel_sheet(
    wb,
    ws_name,
    cleaned,
    state_colors,
    score_header,
    pass_threshold,
    apply_state_colors=True,
    highlight_passing_scores=True,
    title_row=None,
    document_heading=None,
):
    ws = wb[ws_name]
    header_row = 1 if title_row is None and document_heading is None else (2 if document_heading is None else 3)
    
    # Apply document heading if provided
    if document_heading is not None:
        # Insert rows for heading and date/time
        ws.insert_rows(1)
        ws.insert_rows(1)
        
        # Add document heading
        ws.cell(row=1, column=1, value=document_heading)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cleaned.columns))
        heading_cell = ws.cell(row=1, column=1)
        heading_cell.font = Font(bold=True, size=16)
        heading_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add date and time
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=2, column=1, value=f"Date and Time: {current_datetime}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cleaned.columns))
        datetime_cell = ws.cell(row=2, column=1)
        datetime_cell.font = Font(bold=True, size=12)
        datetime_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply title row formatting if provided
    if title_row is not None:
        title_row_num = 3 if document_heading is not None else 1
        if document_heading is None:
            ws.insert_rows(1)
        ws.cell(row=title_row_num, column=1, value=title_row)
        ws.merge_cells(start_row=title_row_num, start_column=1, end_row=title_row_num, end_column=len(cleaned.columns))
        title_cell = ws.cell(row=title_row_num, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_font = Font(bold=True, size=12)
    for cell in ws[header_row]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if "APPLICATION ID" in cleaned.columns:
        try:
            col_idx = list(cleaned.columns).index("APPLICATION ID") + 1
            for r in ws.iter_rows(
                min_row=header_row + 1,
                min_col=col_idx,
                max_col=col_idx,
                max_row=ws.max_row,
            ):
                for cell in r:
                    cell.alignment = Alignment(horizontal="left")
        except Exception:
            pass

    ws.freeze_panes = f"A{header_row + 1}"
    auto_column_width(ws)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Apply borders to all cells including heading and title rows
    start_border_row = 1
    for r in ws.iter_rows(
        min_row=start_border_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in r:
            cell.border = border
            # Apply bold formatting to "Overall Average" row
            if (
                cell.row > header_row
                and ws.cell(row=cell.row, column=1).value == "Overall Average"
            ):
                cell.font = Font(bold=True)

    if apply_state_colors:
        try:
            state_col_index = list(cleaned.columns).index("STATE") + 1
            for row_idx in range(header_row + 1, ws.max_row + 1):
                state_val = ws.cell(row=row_idx, column=state_col_index).value
                # Skip coloring for "Overall Average" row
                if ws.cell(row=row_idx, column=1).value == "Overall Average":
                    continue
                fill = state_color_for(state_val)
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
        except Exception:
            pass

    if highlight_passing_scores:
        score_cols_indices = [
            i + 1 for i, c in enumerate(cleaned.columns) if str(c).startswith("Score/")
        ]
        pass_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        pass_font = Font(color="006100")
        for col_idx in score_cols_indices:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                # Skip highlighting for "Overall Average" row
                if ws.cell(row=row_idx, column=1).value == "Overall Average":
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is None or str(cell.value).strip() == "":
                        continue
                    val = float(cell.value)
                    if val >= pass_threshold:
                        cell.fill = pass_fill
                        cell.font = pass_font
                except Exception:
                    continue


# ---------------------------
# Create Analysis and Charts sheets
# ---------------------------
def create_analysis_and_charts(
    wb,
    cleaned,
    df,
    candidates_df,
    full_candidates_df,
    batch_id_map,
    numeric_to_full,
    score_header,
    output_dir,
    ts,
    fname="Combined",
    pass_threshold=DEFAULT_PASS_THRESHOLD,
    current_batch_id=None,
    skip_batch_mismatches=False,
    rebatched_count=0,
):
    total_candidates = len(cleaned)
    score_series = (
        df["_ScoreNum"].dropna().astype(float)
        if "_ScoreNum" in df.columns
        else pd.Series([], dtype=float)
    )
    highest_score = score_series.max() if not score_series.empty else None
    lowest_score = score_series.min() if not score_series.empty else None
    avg_score = round(score_series.mean(), 2) if not score_series.empty else None

    state_counts = (
        df.groupby("STATE")["_ScoreNum"].count().sort_index()
        if "_ScoreNum" in df.columns
        else pd.Series(dtype=int)
    )
    state_avg = (
        df.groupby("STATE")["_ScoreNum"].mean().round(2).sort_index()
        if "_ScoreNum" in df.columns
        else pd.Series(dtype=float)
    )
    state_pass_count = (
        df[df["_ScoreNum"].astype(float) >= pass_threshold]
        .groupby("STATE")["_ScoreNum"]
        .count()
        if "_ScoreNum" in df.columns
        else pd.Series(dtype=int)
    )

    if "Analysis" in wb.sheetnames:
        wb.remove(wb["Analysis"])
    anal = wb.create_sheet("Analysis")
    anal.append(["Metric", "Value"])
    anal.append(["Total Candidates", int(total_candidates)])
    anal.append(
        [
            "Highest Score (raw)",
            float(highest_score) if highest_score is not None else None,
        ]
    )
    anal.append(
        [
            "Lowest Score (raw)",
            float(lowest_score) if lowest_score is not None else None,
        ]
    )
    anal.append(["Average Score (raw)", avg_score])
    anal.append([])

    anal.append(
        [
            "State",
            "Candidates",
            "Average Score",
            f"Pass Count (≥{int(pass_threshold)})",
            "Pass Rate (%)",
        ]
    )
    per_state_rows = []
    for st in sorted(set(df["STATE"].tolist())):
        cnt = int(state_counts.get(st, 0))
        avg = float(state_avg.get(st, float("nan"))) if st in state_avg.index else None
        pcnt = int(state_pass_count.get(st, 0)) if st in state_pass_count.index else 0
        prate = round((pcnt / cnt * 100), 2) if cnt > 0 else 0.0
        per_state_rows.append(
            [
                st,
                cnt,
                avg if not (isinstance(avg, float) and math.isnan(avg)) else None,
                pcnt,
                prate,
            ]
        )
    for row in per_state_rows:
        anal.append(row)

    anal["A1"].font = Font(bold=True, size=12)
    anal["B1"].font = Font(bold=True, size=12)
    for i, r in enumerate(anal.iter_rows(values_only=True), start=1):
        if r and r[0] == "State" and r[1] == "Candidates":
            for cell in anal[i]:
                cell.font = Font(bold=True, size=12)
            break

    if per_state_rows:
        sorted_by_count = sorted(per_state_rows, key=lambda r: r[1], reverse=True)
        most_state, most_cnt = sorted_by_count[0][0], sorted_by_count[0][1]
        least_state, least_cnt = sorted_by_count[-1][0], sorted_by_count[-1][1]
        sorted_by_avg = sorted(
            [r for r in per_state_rows if r[2] is not None],
            key=lambda r: r[2],
            reverse=True,
        )
        best_avg_state = sorted_by_avg[0][0] if sorted_by_avg else None
        sorted_by_pr = sorted(per_state_rows, key=lambda r: r[4], reverse=True)
        best_pass_state = sorted_by_pr[0][0] if sorted_by_pr else None

        overall_pass = sum(r[3] for r in per_state_rows)
        overall_fail = total_candidates - overall_pass
        overall_pass_rate = (
            round(overall_pass / total_candidates * 100, 2)
            if total_candidates > 0
            else 0.0
        )

        anal.append([])
        anal.append(["Narrative Summary"])
        anal.append([f"Most candidates: {most_state} ({most_cnt})"])
        anal.append([f"Fewest candidates: {least_state} ({least_cnt})"])
        if best_avg_state:
            anal.append([f"Highest average score: {best_avg_state}"])
        if best_pass_state:
            anal.append([f"Highest pass rate: {best_pass_state}"])
        anal.append([f"Overall pass rate: {overall_pass_rate}%"])

        print(f"\nSummary for {fname}:")
        print(f"  Total candidates processed: {total_candidates}")
        print(f"  Highest raw score: {highest_score}")
        print(f"  Lowest raw score: {lowest_score}")
        print(f"  Average raw score: {avg_score}")
        print(f"  State with most candidates: {most_state} ({most_cnt})")
        print(f"  State with fewest candidates: {least_state} ({least_cnt})")

    if "Charts" in wb.sheetnames:
        wb.remove(wb["Charts"])
    chs = wb.create_sheet("Charts")

    chs["A1"] = "UTME Results Charts Overview"
    chs["A1"].font = Font(bold=True, size=16)
    chs["A2"] = (
        "1. Candidates per State: Number of candidates from each state.\n"
        "2. Average Score per State: Average score (raw) per state.\n"
        f"3. Pass Rate per State: Percentage of candidates scoring ≥{int(pass_threshold)} per state.\n"
        "4. Score Distribution: Number of candidates in 10-point score ranges.\n"
        "5. Pass/Fail Distribution: Proportion of all candidates who passed vs. failed."
    )
    chs["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    chs.column_dimensions["A"].width = 40

    chs["C2"] = "State Color Legend"
    chs["C2"].font = Font(bold=True, size=12)
    chs["C3"] = "State"
    chs["D3"] = "Color"
    chs["C3"].font = Font(bold=True)
    chs["D3"].font = Font(bold=True)
    row = 4
    state_colors = {}
    for st in sorted(set(df["STATE"].tolist())):
        fill = state_color_for(st)
        state_colors[st] = fill.start_color.rgb[2:]
        chs[f"C{row}"] = st
        chs[f"D{row}"].fill = fill
        row += 1
    auto_column_width(chs, min_width=10, max_width=30)

    start_row = row + 2
    chs.cell(row=start_row, column=1, value="State")
    chs.cell(row=start_row, column=2, value="Candidates")
    chs.cell(row=start_row, column=3, value="Average Score")
    chs.cell(row=start_row, column=4, value="Pass Rate (%)")
    counts_sorted = sorted(per_state_rows, key=lambda r: r[1], reverse=True)
    for i, r in enumerate(counts_sorted):
        chs.cell(row=start_row + i + 1, column=1, value=r[0])
        chs.cell(row=start_row + i + 1, column=2, value=r[1])
        chs.cell(row=start_row + i + 1, column=3, value=r[2] if r[2] is not None else 0)
        chs.cell(row=start_row + i + 1, column=4, value=r[4])

    n = len(counts_sorted)
    if n > 0:
        try:
            c1 = BarChart()
            c1.title = "Candidates per State"
            c1.x_axis.title = "State"
            c1.y_axis.title = "Number of Candidates"
            c1.width = 32
            c1.height = 16
            c1.gapWidth = 50
            data_ref = Reference(
                chs, min_col=2, min_row=start_row + 1, max_row=start_row + n
            )
            cats_ref = Reference(
                chs, min_col=1, min_row=start_row + 1, max_row=start_row + n
            )
            c1.add_data(data_ref, titles_from_data=True)
            c1.set_categories(cats_ref)
            c1.dLbls = DataLabelList()
            c1.dLbls.showVal = True
            c1.dLbls.showCatName = False
            c1.style = 10
            for i, series in enumerate(c1.series):
                state = counts_sorted[i][0]
                series.graphicalProperties.solidFill = state_colors[state]
            c1.title.font = Font(size=16, bold=True)
            c1.x_axis.titleFont = Font(size=14)
            c1.y_axis.titleFont = Font(size=14)
            chs.add_chart(c1, f"F{start_row}")
        except Exception as e:
            print(f"Error creating Candidates chart: {e}")

        try:
            c2 = BarChart()
            c2.title = "Average Score per State"
            c2.x_axis.title = "State"
            c2.y_axis.title = "Average Score"
            c2.width = 32
            c2.height = 16
            c2.gapWidth = 50
            data_ref2 = Reference(
                chs, min_col=3, min_row=start_row + 1, max_row=start_row + n
            )
            cats_ref2 = Reference(
                chs, min_col=1, min_row=start_row + 1, max_row=start_row + n
            )
            c2.add_data(data_ref2, titles_from_data=True)
            c2.set_categories(cats_ref2)
            c2.dLbls = DataLabelList()
            c2.dLbls.showVal = True
            c2.dLbls.showCatName = False
            c2.style = 10
            for i, series in enumerate(c2.series):
                state = counts_sorted[i][0]
                series.graphicalProperties.solidFill = state_colors[state]
            c2.title.font = Font(size=16, bold=True)
            c2.x_axis.titleFont = Font(size=14)
            c2.y_axis.titleFont = Font(size=14)
            chs.add_chart(c2, f"F{start_row + 20}")
        except Exception as e:
            print(f"Error creating Average Score chart: {e}")

        try:
            c3 = BarChart()
            c3.title = f"Pass Rate per State (≥ {int(pass_threshold)})"
            c3.x_axis.title = "State"
            c3.y_axis.title = "Pass Rate (%)"
            c3.width = 32
            c3.height = 16
            c3.gapWidth = 50
            data_ref3 = Reference(
                chs, min_col=4, min_row=start_row + 1, max_row=start_row + n
            )
            cats_ref3 = Reference(
                chs, min_col=1, min_row=start_row + 1, max_row=start_row + n
            )
            c3.add_data(data_ref3, titles_from_data=True)
            c3.set_categories(cats_ref3)
            c3.dLbls = DataLabelList()
            c3.dLbls.showVal = True
            c3.dLbls.showCatName = False
            c3.style = 10
            for i, series in enumerate(c3.series):
                state = counts_sorted[i][0]
                series.graphicalProperties.solidFill = state_colors[state]
            c3.title.font = Font(size=16, bold=True)
            c3.x_axis.titleFont = Font(size=14)
            c3.y_axis.titleFont = Font(size=14)
            chs.add_chart(c3, f"F{start_row + 40}")
        except Exception as e:
            print(f"Error creating Pass Rate chart: {e}")

    try:
        score_vals = score_series.dropna().astype(float)
        if len(score_vals) > 0:
            bins = list(range(0, 101, 10))
            dist = pd.cut(score_vals, bins=bins, right=False)
            freq = dist.value_counts().sort_index()
            insert_row = start_row + n + 3
            chs.cell(row=insert_row, column=1, value="Score Range")
            chs.cell(row=insert_row, column=2, value="Candidates")
            r = insert_row + 1
            for rng, cnt in freq.items():
                chs.cell(row=r, column=1, value=str(rng))
                chs.cell(row=r, column=2, value=int(cnt))
                r += 1
            n_bins = len(freq)
            try:
                c4 = BarChart()
                c4.title = "Score Distribution (10-point Ranges)"
                c4.x_axis.title = "Score Range"
                c4.y_axis.title = "Number of Candidates"
                c4.width = 32
                c4.height = 12
                c4.gapWidth = 50
                data_ref4 = Reference(
                    chs, min_col=2, min_row=insert_row + 1, max_row=insert_row + n_bins
                )
                cats_ref4 = Reference(
                    chs, min_col=1, min_row=insert_row + 1, max_row=insert_row + n_bins
                )
                c4.add_data(data_ref4, titles_from_data=True)
                c4.set_categories(cats_ref4)
                c4.dLbls = DataLabelList()
                c4.dLbls.showVal = True
                c4.dLbls.showCatName = False
                c4.style = 10
                for i, series in enumerate(c4.series):
                    series.graphicalProperties.solidFill = PALETTE[i % len(PALETTE)]
                c4.title.font = Font(size=16, bold=True)
                c4.x_axis.titleFont = Font(size=14)
                c4.y_axis.titleFont = Font(size=14)
                chs.add_chart(c4, f"F{insert_row}")
            except Exception as e:
                print(f"Error creating Score Distribution chart: {e}")
    except Exception as e:
        print(f"Error processing score distribution: {e}")

    try:
        if total_candidates > 0:
            insert_row_pie = chs.max_row + 3
            chs.cell(row=insert_row_pie, column=1, value="Category")
            chs.cell(row=insert_row_pie, column=2, value="Count")
            chs.cell(row=insert_row_pie + 1, column=1, value="Pass")
            chs.cell(row=insert_row_pie + 1, column=2, value=overall_pass)
            chs.cell(row=insert_row_pie + 2, column=1, value="Fail")
            chs.cell(row=insert_row_pie + 2, column=2, value=overall_fail)
            pie = PieChart()
            pie.title = f"Overall Pass/Fail Distribution (≥ {int(pass_threshold)})"
            pie.width = 15
            pie.height = 10
            data_ref_pie = Reference(
                chs, min_col=2, min_row=insert_row_pie + 1, max_row=insert_row_pie + 2
            )
            cats_ref_pie = Reference(
                chs, min_col=1, min_row=insert_row_pie + 1, max_row=insert_row_pie + 2
            )
            pie.add_data(data_ref_pie, titles_from_data=True)
            pie.set_categories(cats_ref_pie)
            pie.dLbls = DataLabelList()
            pie.dLbls.showVal = True
            pie.dLbls.showPercent = True
            pie.dLbls.showCatName = True
            pie.style = 10
            for i, series in enumerate(pie.series):
                series.graphicalProperties.solidFill = "A8E6CF" if i == 0 else "FFCCCB"
            pie.title.font = Font(size=16, bold=True)
            chs.add_chart(pie, f"F{insert_row_pie}")
    except Exception as e:
        print(f"Error creating Pass/Fail Pie chart: {e}")

    # Calculate absent candidates and batch mismatches
    absent_count = 0
    absent_df = pd.DataFrame(
        columns=["APPLICATION ID", "FULL NAME", "PHONE NO.", "STATE", "Batch"]
    )
    mismatch_df = pd.DataFrame(
        columns=[
            "APPLICATION ID",
            "FULL NAME",
            "PHONE NO.",
            "STATE",
            "REGISTERED BATCH",
        ]
    )
    invalid_ids = []
    total_registered = len(candidates_df)

    if not candidates_df.empty:
        present_ids = set()
        mismatch_rows = []
        if "APPLICATION ID" in cleaned.columns:
            duplicates = cleaned[cleaned["APPLICATION ID"].duplicated(keep=False)]
            if not duplicates.empty:
                print(
                    f"Found {len(duplicates)} duplicate APPLICATION ID entries in {fname}:"
                )
                print(duplicates[["APPLICATION ID", "FULL NAME"]].head().to_string())
            for index, v in cleaned["APPLICATION ID"].astype(str).items():
                tok = extract_numeric_token(v)
                norm_id = normalize_id(v)
                registered_batch = None
                full_ex = None
                if norm_id in batch_id_map:
                    registered_batch = batch_id_map[norm_id]
                    full_ex = v
                elif tok and tok in numeric_to_full:
                    full_ex = numeric_to_full[tok]
                    norm_id = normalize_id(full_ex)
                    registered_batch = batch_id_map.get(norm_id, "")
                if (
                    registered_batch
                    and current_batch_id
                    and re.sub(r"\s+", "", registered_batch.lower())
                    != re.sub(r"\s+", "", current_batch_id.lower())
                ):
                    mismatch_row = full_candidates_df[
                        full_candidates_df["EXAM_NO"].apply(normalize_id) == norm_id
                    ]
                    if not mismatch_row.empty:
                        mismatch_rows.append(
                            {
                                "APPLICATION ID": v,
                                "FULL NAME": cleaned.at[index, "FULL NAME"],
                                "PHONE NO.": cleaned.at[index, "PHONE NUMBER"],
                                "STATE": cleaned.at[index, "STATE"],
                                "REGISTERED BATCH": registered_batch,
                            }
                        )
                if tok:
                    present_ids.add(tok)
                else:
                    invalid_ids.append((v, cleaned.at[index, "FULL NAME"]))
        for index, v in cleaned["FULL NAME"].astype(str).items():
            tok = extract_numeric_token(v)
            if tok and tok not in present_ids:
                registered_batch = ""
                full_ex = None
                if tok in numeric_to_full:
                    full_ex = numeric_to_full[tok]
                    norm_id = normalize_id(full_ex)
                    registered_batch = batch_id_map.get(norm_id, "")
                    if current_batch_id and re.sub(
                        r"\s+", "", registered_batch.lower()
                    ) != re.sub(r"\s+", "", current_batch_id.lower()):
                        mismatch_row = full_candidates_df[
                            full_candidates_df["EXAM_NO"] == full_ex
                        ]
                        if not mismatch_row.empty:
                            mismatch_rows.append(
                                {
                                    "APPLICATION ID": full_ex,
                                    "FULL NAME": v,
                                    "PHONE NO.": cleaned.at[index, "PHONE NUMBER"],
                                    "STATE": cleaned.at[index, "STATE"],
                                    "REGISTERED BATCH": registered_batch,
                                }
                            )
                    present_ids.add(tok)

        # Calculate absent candidates
        candidates_df["EXAM_NO_NORMAL"] = candidates_df["EXAM_NO"].apply(
            lambda x: extract_numeric_token(x) or normalize_id(x)
        )
        candidate_ids = set(candidates_df["EXAM_NO_NORMAL"])
        absent_ids = sorted([cid for cid in candidate_ids if cid not in present_ids])
        absent_df = candidates_df[candidates_df["EXAM_NO_NORMAL"].isin(absent_ids)][
            ["EXAM_NO", "FULL_NAME", "PHONE NUMBER", "STATE", "BATCH_ID"]
        ].drop_duplicates()
        absent_df = absent_df.rename(
            columns={
                "EXAM_NO": "APPLICATION ID",
                "FULL_NAME": "FULL NAME",
                "PHONE NUMBER": "PHONE NO.",
                "STATE": "STATE",
                "BATCH_ID": "Batch",
            }
        )
        absent_count = len(absent_df)

        # Handle batch mismatches
        if mismatch_rows:
            mismatch_df = pd.DataFrame(mismatch_rows)
            print(
                f"Found {len(mismatch_df)} candidates in {fname} results who were registered in a different batch:"
            )
            print(
                mismatch_df[
                    ["APPLICATION ID", "FULL NAME", "REGISTERED BATCH"]
                ].to_string(index=False)
            )

        print(f"Absent candidates check for {fname}:")
        print(
            f"  Total registered candidates (from RAW_CANDIDATE_BATCHES): {total_registered}"
        )
        print(f"  Total present candidates (sat for exam): {len(present_ids)}")
        print(f"  Total absent candidates (registered but did not sit): {absent_count}")
        if absent_count != len(absent_df):
            print(
                f"Warning: Discrepancy detected: Expected {absent_count} absent candidates, but found {len(absent_df)} in absent_df"
            )
        if absent_ids:
            print(f"  Sample absent IDs (first 5): {absent_ids[:5]}")
        if invalid_ids:
            print(
                f"  Note: {len(invalid_ids)} invalid APPLICATION IDs may affect absent count."
            )

    if "Absent" in wb.sheetnames:
        wb.remove(wb["Absent"])
    abs_ws = wb.create_sheet("Absent")
    abs_ws.append(["APPLICATION ID", "FULL NAME", "PHONE NO.", "STATE", "Batch"])
    if absent_df.empty:
        abs_ws.append(["No absent candidates found.", "", "", "", ""])
    else:
        for _, r in absent_df.iterrows():
            abs_ws.append(
                [
                    r.get("APPLICATION ID"),
                    r.get("FULL NAME"),
                    r.get("PHONE NO."),
                    r.get("STATE"),
                    r.get("Batch"),
                ]
            )
    auto_column_width(abs_ws)

    if not skip_batch_mismatches:
        if "BatchMismatches" in wb.sheetnames:
            wb.remove(wb["BatchMismatches"])
        mismatch_ws = wb.create_sheet("BatchMismatches")
        mismatch_ws.append(
            ["APPLICATION ID", "FULL NAME", "PHONE NO.", "STATE", "REGISTERED BATCH"]
        )
        if mismatch_df.empty:
            mismatch_ws.append(["No candidates found in wrong batch.", "", "", "", ""])
        else:
            for _, r in mismatch_df.iterrows():
                mismatch_ws.append(
                    [
                        r.get("APPLICATION ID"),
                        r.get("FULL NAME"),
                        r.get("PHONE NO."),
                        r.get("STATE"),
                        r.get("REGISTERED BATCH"),
                    ]
                )
        auto_column_width(mismatch_ws)

    anal.append([])
    anal.append(["Registered (candidate batches) total", int(total_registered)])
    anal.append(["Registered but absent (did not sit)", int(absent_count)])
    if skip_batch_mismatches:
        anal.append(["Rebatched candidates", rebatched_count])
    else:
        anal.append(["Candidates in wrong batch", len(mismatch_df)])
    anal.append([])

    print(f"  Registered (candidate batches): {total_registered}")
    print(f"  Registered but absent: {absent_count}")
    if skip_batch_mismatches:
        print(f"  Rebatched candidates: {rebatched_count}")
    else:
        print(f"  Candidates in wrong batch: {len(mismatch_df)}")

    auto_column_width(chs)
    auto_column_width(anal)

    return absent_df, mismatch_df


# ---------------------------
# Core processing for a single file
# ---------------------------
def process_file(
    path,
    output_dir,
    ts,
    candidate_dir,
    full_candidates_df,
    full_batch_id_map,
    full_numeric_to_full,
    pass_threshold,
):
    fname = os.path.basename(path)
    print(f"\nProcessing: {fname}")

    batch_id_match = re.search(r"(Batch\s*\d+[A-Za-z]?)", fname, re.IGNORECASE)
    batch_id = batch_id_match.group(1) if batch_id_match else None
    if batch_id:
        print(f"Detected batch ID: {batch_id}")
        candidates_df, _, _ = load_candidate_batches(candidate_dir, batch_id=batch_id)
    else:
        print(
            f"Warning: Could not detect batch ID in filename {fname}. Using empty candidates list for batch-specific absent calculation."
        )
        candidates_df = pd.DataFrame(
            columns=["EXAM_NO", "FULL_NAME", "PHONE NUMBER", "STATE", "BATCH_ID"]
        )

    try:
        if fname.lower().endswith(".csv"):
            df = pd.read_csv(path, dtype=str)
        else:
            try:
                df = pd.read_excel(path, dtype=str)
            except Exception:
                try:
                    df = pd.read_excel(path, dtype=str, engine="openpyxl")
                except Exception:
                    df = pd.read_excel(path, dtype=str, engine="xlrd")
    except Exception as e:
        print(f"Error reading {fname}: {e}")
        return None, None, None, None

    df.rename(columns=lambda c: str(c).strip(), inplace=True)

    fullname_col = find_column_by_names(
        df, ["First name", "Firstname", "Full name", "Name", "Surname"]
    )
    if fullname_col:
        df.rename(columns={fullname_col: "FULL NAME"}, inplace=True)

    appid_col = find_column_by_names(
        df, ["Surname", "Mat no", "Mat No", "MAT NO.", "APPLICATION ID", "Username"]
    )
    if appid_col and appid_col != fullname_col:
        df.rename(columns={appid_col: "APPLICATION ID"}, inplace=True)

    # Dynamically look for phone number column with more variations
    phone_col = find_column_by_names(
        df, ["Phone", "Phone number", "PHONE", "Mobile", "PhoneNumber", "Phone No", "PHONE NUMBER", "Phone No."]
    )
    if phone_col:
        df.rename(columns={phone_col: "PHONE NUMBER"}, inplace=True)
        # Clean phone numbers immediately after finding the column
        df["PHONE NUMBER"] = df["PHONE NUMBER"].apply(clean_phone_value)

    city_col = find_column_by_names(
        df, ["City/town", "City /town", "City", "Town", "State of Origin", "STATE"]
    )
    if city_col:
        df.rename(columns={city_col: "STATE"}, inplace=True)

    for drop_col in [
        "Username",
        "Department",
        "State",
        "Started on",
        "Completed",
        "Time taken",
        "USERNAME",
        "DEPARTMENT",
    ]:
        if drop_col in df.columns:
            df.drop(columns=[drop_col], inplace=True)

    if "FULL NAME" not in df.columns:
        df["FULL NAME"] = pd.NA
    if "PHONE NUMBER" not in df.columns:
        df["PHONE NUMBER"] = pd.NA
    if "STATE" not in df.columns:
        df["STATE"] = pd.NA

    df["FULL NAME"] = df["FULL NAME"].astype(str).str.strip()
    df["STATE"] = df["STATE"].astype(str).str.strip()
    df["STATE"] = df["STATE"].replace(["", "nan", "None"], "NO STATE OF ORIGIN")
    df["STATE"] = df["STATE"].fillna("NO STATE OF ORIGIN")

    grade_col = find_grade_column(df)
    if not grade_col:
        print(
            f"Skipping {fname}: Missing required column: a 'Grade/...' or 'Grade' column was not found."
        )
        return None, None, None, None

    if str(grade_col).strip().lower().startswith("grade/"):
        score_header = re.sub(r"(?i)^grade", "Score", grade_col, flags=re.I)
    else:
        found = re.search(r"(\d+(?:\.\d+)?)", str(grade_col))
        suffix = found.group(1) if found else ""
        score_header = f"Score/{suffix}" if suffix else "Score"

    df.rename(columns={grade_col: score_header}, inplace=True)

    df["_ScoreNum"] = to_numeric_safe(df[score_header])

    df = drop_overall_average_rows(df)
    df["FULL NAME"] = df["FULL NAME"].astype(str).str.strip()
    df = df[~df["FULL NAME"].isin(["", "nan", "None"])].copy()
    df = df[df["_ScoreNum"].notna()].copy()

    if df.empty:
        print(f"No valid rows remain after cleaning {fname}; skipping file.")
        return None, None, None, None

    # Ensure phone numbers are cleaned (in case they weren't cleaned earlier)
    if "PHONE NUMBER" in df.columns:
        df["PHONE NUMBER"] = df["PHONE NUMBER"].apply(clean_phone_value)

    df.sort_values(
        by=["STATE", "_ScoreNum"],
        ascending=[True, False],
        inplace=True,
        na_position="last",
    )
    df.reset_index(drop=True, inplace=True)
    df.insert(0, "S/N", range(1, len(df) + 1))
    df["STATE_SN"] = df.groupby("STATE").cumcount() + 1

    out_cols = ["S/N", "STATE_SN", "FULL NAME"]
    if "APPLICATION ID" in df.columns:
        out_cols.append("APPLICATION ID")
    out_cols += ["PHONE NUMBER", "STATE", score_header]

    for c in out_cols:
        if c not in df.columns:
            df[c] = pd.NA

    cleaned = df[out_cols].copy()

    found_max = re.search(r"(\d+(?:\.\d+)?)", str(score_header))
    original_max = float(found_max.group(1)) if found_max else 100.0
    cleaned["Score/100"] = (
        ((df["_ScoreNum"].astype(float) / original_max) * 100).round(0).astype("Int64")
    )

    cleaned = cleaned[
        ~(
            (cleaned["FULL NAME"].astype(str).str.strip() == "")
            & (cleaned["PHONE NUMBER"].astype(str).str.strip() == "")
        )
    ].copy()

    base_name = f"UTME_RESULT_{os.path.splitext(fname)[0]}_{ts}"
    out_csv = os.path.join(output_dir, base_name + ".csv")
    out_xlsx = os.path.join(output_dir, base_name + ".xlsx")

    cleaned.to_csv(out_csv, index=False)
    try:
        cleaned.to_excel(out_xlsx, index=False, engine="openpyxl")
    except Exception:
        cleaned.to_excel(out_xlsx, index=False)
    print(f"Saved processed file: {os.path.basename(out_csv)}")
    print(f"Saved processed file: {os.path.basename(out_xlsx)}")

    wb = load_workbook(out_xlsx)
    ws = wb.active
    ws.title = "Results"
    
    # Add document heading
    document_heading = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA\nPOST UTME RESULT"
    format_excel_sheet(wb, "Results", cleaned, {}, score_header, pass_threshold, document_heading=document_heading)

    absent_df, mismatch_df = create_analysis_and_charts(
        wb,
        cleaned,
        df,
        candidates_df,
        full_candidates_df,
        full_batch_id_map,
        full_numeric_to_full,
        score_header,
        output_dir,
        ts,
        fname=fname,
        pass_threshold=pass_threshold,
        current_batch_id=batch_id,
    )

    wb.save(out_xlsx)
    print(
        f"Saved processed file with Analysis, Charts, Absent, and BatchMismatches: {os.path.basename(out_xlsx)}"
    )

    return cleaned, df, absent_df, mismatch_df


# ---------------------------
# Process file for unsorted result (minimal processing, no sorting)
# ---------------------------
def process_file_for_unsorted(path):
    fname = os.path.basename(path)
    print(f"Processing for unsorted result: {fname}")

    batch_id_match = re.search(r"(Batch\s*\d+[A-Za-z]?)", fname, re.IGNORECASE)
    batch_id = batch_id_match.group(1) if batch_id_match else os.path.splitext(fname)[0]

    try:
        if fname.lower().endswith(".csv"):
            df = pd.read_csv(path, dtype=str)
        else:
            try:
                df = pd.read_excel(path, dtype=str)
            except Exception:
                try:
                    df = pd.read_excel(path, dtype=str, engine="openpyxl")
                except Exception:
                    df = pd.read_excel(path, dtype=str, engine="xlrd")
    except Exception as e:
        print(f"Error reading {fname} for unsorted result: {e}")
        return None, None, None

    df.rename(columns=lambda c: str(c).strip(), inplace=True)

    fullname_col = find_column_by_names(
        df, ["First name", "Firstname", "Full name", "Name", "Surname"]
    )
    if fullname_col:
        df.rename(columns={fullname_col: "FULL NAME"}, inplace=True)

    appid_col = find_column_by_names(
        df, ["Surname", "Mat no", "Mat No", "MAT NO.", "APPLICATION ID", "Username"]
    )
    if appid_col and appid_col != fullname_col:
        df.rename(columns={appid_col: "APPLICATION ID"}, inplace=True)

    # Dynamically look for phone number column with more variations
    phone_col = find_column_by_names(
        df, ["Phone", "Phone number", "PHONE", "Mobile", "PhoneNumber", "Phone No", "PHONE NUMBER", "Phone No."]
    )
    if phone_col:
        df.rename(columns={phone_col: "PHONE NUMBER"}, inplace=True)
        # Clean phone numbers immediately after finding the column
        df["PHONE NUMBER"] = df["PHONE NUMBER"].apply(clean_phone_value)

    city_col = find_column_by_names(
        df, ["City/town", "City /town", "City", "Town", "State of Origin", "STATE"]
    )
    if city_col:
        df.rename(columns={city_col: "STATE"}, inplace=True)

    for drop_col in [
        "Username",
        "Department",
        "State",
        "Started on",
        "Completed",
        "Time taken",
        "USERNAME",
        "DEPARTMENT",
    ]:
        if drop_col in df.columns:
            df.drop(columns=[drop_col], inplace=True)

    if "FULL NAME" not in df.columns:
        df["FULL NAME"] = pd.NA
    if "PHONE NUMBER" not in df.columns:
        df["PHONE NUMBER"] = pd.NA
    if "STATE" not in df.columns:
        df["STATE"] = pd.NA

    df["FULL NAME"] = df["FULL NAME"].astype(str).str.strip()
    df["STATE"] = df["STATE"].astype(str).str.strip()
    df["STATE"] = df["STATE"].replace(["", "nan", "None"], "NO STATE OF ORIGIN")
    df["STATE"] = df["STATE"].fillna("NO STATE OF ORIGIN")

    grade_col = find_grade_column(df)
    if not grade_col:
        print(
            f"Skipping {fname} for unsorted result: Missing required column: a 'Grade/...' or 'Grade' column was not found."
        )
        return None, None, None

    if str(grade_col).strip().lower().startswith("grade/"):
        score_header = re.sub(r"(?i)^grade", "Score", grade_col, flags=re.I)
    else:
        found = re.search(r"(\d+(?:\.\d+)?)", str(grade_col))
        suffix = found.group(1) if found else ""
        score_header = f"Score/{suffix}" if suffix else "Score"

    df.rename(columns={grade_col: score_header}, inplace=True)

    df["_ScoreNum"] = to_numeric_safe(df[score_header])

    df = drop_overall_average_rows(df)
    df["FULL NAME"] = df["FULL NAME"].astype(str).str.strip()
    df = df[~df["FULL NAME"].isin(["", "nan", "None"])].copy()
    df = df[df["_ScoreNum"].notna()].copy()

    if df.empty:
        print(
            f"No valid rows remain after cleaning {fname} for unsorted result; skipping file."
        )
        return None, None, None

    # Ensure phone numbers are cleaned (in case they weren't cleaned earlier)
    if "PHONE NUMBER" in df.columns:
        df["PHONE NUMBER"] = df["PHONE NUMBER"].apply(clean_phone_value)

    out_cols = ["FULL NAME"]
    if "APPLICATION ID" in df.columns:
        out_cols.append("APPLICATION ID")
    out_cols += ["PHONE NUMBER", "STATE", score_header]

    for c in out_cols:
        if c not in df.columns:
            df[c] = pd.NA

    cleaned = df[out_cols].copy()

    found_max = re.search(r"(\d+(?:\.\d+)?)", str(score_header))
    original_max = float(found_max.group(1)) if found_max else 100.0
    cleaned["Score/100.00"] = (
        (df["_ScoreNum"].astype(float) / original_max) * 100
    ).round(2)

    cleaned = cleaned[
        ~(
            (cleaned["FULL NAME"].astype(str).str.strip() == "")
            & (cleaned["PHONE NUMBER"].astype(str).str.strip() == "")
        )
    ].copy()

    cleaned.insert(0, "S/N", range(1, len(cleaned) + 1))

    out_cols = [
        "S/N",
        "FULL NAME",
        "APPLICATION ID",
        "PHONE NUMBER",
        "STATE",
        "Score/100.00",
    ]
    cleaned = cleaned[out_cols].copy()

    # Calculate overall average for Score/100.00
    avg_score = cleaned["Score/100.00"].mean()
    avg_score = round(float(avg_score), 2) if pd.notna(avg_score) else None

    return cleaned, batch_id, avg_score


# ---------------------------
# Combine all batches
# ---------------------------
def combine_batches(
    cleaned_dfs,
    dfs,
    absent_dfs,
    mismatch_dfs,
    output_dir,
    ts,
    pass_threshold,
    candidate_dir,
    raw_dir,
    non_interactive=False,
    converted_score_max=None,
):
    if not cleaned_dfs:
        print("No valid batches to combine.")
        return

    # Create unsorted combined result with batch-specific worksheets and overall averages
    raw_files = [
        f
        for f in os.listdir(raw_dir)
        if f.lower().endswith((".csv", ".xlsx", ".xls")) and not f.startswith("~$")
    ]
    unsorted_dfs = []
    batch_sheets = {}
    batch_averages = {}
    for f in sorted(raw_files):  # Sort files to ensure consistent order
        cleaned, batch_id, avg_score = process_file_for_unsorted(
            os.path.join(raw_dir, f)
        )
        if cleaned is not None:
            unsorted_dfs.append(cleaned)
            batch_sheets[batch_id] = cleaned
            batch_averages[batch_id] = avg_score
            print(
                f"Prepared unsorted data for batch: {batch_id}, Overall Average: {avg_score}"
            )

    if unsorted_dfs:
        unsorted_cleaned = pd.concat(unsorted_dfs, ignore_index=True)
        unsorted_cols = [
            "S/N",
            "FULL NAME",
            "APPLICATION ID",
            "PHONE NUMBER",
            "STATE",
            "Score/100.00",
        ]
        missing_cols = [c for c in unsorted_cols if c not in unsorted_cleaned.columns]
        if missing_cols:
            print(
                f"Error: Missing required columns for unsorted result: {missing_cols}"
            )
        else:
            unsorted_cleaned = unsorted_cleaned[unsorted_cols].copy()
            if "STATE_SN" in unsorted_cleaned.columns:
                unsorted_cleaned = unsorted_cleaned.drop(columns=["STATE_SN"])
            # Calculate overall average for combined unsorted data
            combined_avg = (
                round(float(unsorted_cleaned["Score/100.00"].mean()), 2)
                if pd.notna(unsorted_cleaned["Score/100.00"].mean())
                else None
            )
            out_unsorted_xlsx = os.path.join(
                output_dir, f"PUTME_COMBINE_RESULT_UNSORTED_{ts}.xlsx"
            )
            try:
                with pd.ExcelWriter(out_unsorted_xlsx, engine="openpyxl") as writer:
                    # Write Unsorted Results sheet
                    unsorted_cleaned.to_excel(
                        writer, sheet_name="Unsorted Results", index=False
                    )
                    # Append overall average row
                    wb = writer.book
                    ws = wb["Unsorted Results"]
                    avg_row = (
                        ["Overall Average"]
                        + [""] * (len(unsorted_cols) - 2)
                        + [combined_avg]
                    )
                    ws.append(avg_row)
                    # Write batch-specific sheets with title row
                    for batch_id, batch_df in batch_sheets.items():
                        safe_batch_id = re.sub(r"[^\w\s]", "_", batch_id)[:31]
                        title = f"{batch_id.upper()} UNSORTED"
                        batch_df.to_excel(
                            writer, sheet_name=safe_batch_id, index=False, startrow=1
                        )
                        ws_batch = wb[safe_batch_id]
                        ws_batch.insert_rows(1)
                        ws_batch.cell(row=1, column=1).value = title
                        avg_row_batch = (
                            ["Overall Average"]
                            + [""] * (len(unsorted_cols) - 2)
                            + [batch_averages.get(batch_id)]
                        )
                        ws_batch.append(avg_row_batch)
                print(
                    f"Saved unsorted combined result with batch sheets: {os.path.basename(out_unsorted_xlsx)}"
                )
            except Exception as e:
                print(f"Error saving unsorted combined result: {e}")
                return

            wb_unsorted = load_workbook(out_unsorted_xlsx)
            # Add document heading to unsorted results
            document_heading = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA\nPOST UTME RESULT"
            format_excel_sheet(
                wb_unsorted,
                "Unsorted Results",
                unsorted_cleaned,
                {},
                "Score/100.00",
                pass_threshold,
                apply_state_colors=False,
                highlight_passing_scores=False,
                document_heading=document_heading,
            )
            for batch_id, batch_df in batch_sheets.items():
                safe_batch_id = re.sub(r"[^\w\s]", "_", batch_id)[:31]
                title = f"{batch_id.upper()} UNSORTED"
                format_excel_sheet(
                    wb_unsorted,
                    safe_batch_id,
                    batch_df,
                    {},
                    "Score/100.00",
                    pass_threshold,
                    apply_state_colors=False,
                    highlight_passing_scores=False,
                    title_row=title,
                    document_heading=document_heading,
                )
            wb_unsorted.save(out_unsorted_xlsx)
            print(
                f"Saved formatted unsorted combined result with batch sheets and overall averages: {os.path.basename(out_unsorted_xlsx)}"
            )
    else:
        print("No valid data for unsorted combined result.")

    # Create sorted combined result
    combined_cleaned = pd.concat(cleaned_dfs, ignore_index=True)
    duplicates = combined_cleaned[
        combined_cleaned["APPLICATION ID"].duplicated(keep=False)
    ]
    if not duplicates.empty:
        print(
            f"Found {len(duplicates)} duplicate APPLICATION ID entries in combined results:"
        )
        print(duplicates[["APPLICATION ID", "FULL NAME"]].head().to_string())
    combined_cleaned = combined_cleaned.drop_duplicates(
        subset=["APPLICATION ID"], keep="first"
    )
    combined_cleaned.sort_values(
        by=["STATE", "Score/100"], ascending=[True, False], inplace=True
    )
    combined_cleaned.reset_index(drop=True, inplace=True)
    combined_cleaned["S/N"] = range(1, len(combined_cleaned) + 1)
    combined_cleaned["STATE_SN"] = combined_cleaned.groupby("STATE").cumcount() + 1

    score_header = [
        col
        for col in combined_cleaned.columns
        if col.startswith("Score/") and col != "Score/100"
    ][0]
    found_max = re.search(r"(\d+(?:\.\d+)?)", str(score_header))
    original_max = float(found_max.group(1)) if found_max else 100.0

    if non_interactive and converted_score_max is not None:
        try:
            tgt = float(converted_score_max)
            new_col = f"Score/{int(tgt)}%"
            combined_cleaned[new_col] = (
                ((combined_cleaned["Score/100"].astype(float) / 100.0) * tgt)
                .round(0)
                .astype("Int64")
            )
            print(f"Added converted column '{new_col}' to combined result.")
        except Exception as e:
            print(f"Invalid conversion value; skipped converted column: {e}")
    elif not non_interactive:
        if sys.stdin.isatty():
            add_conv = (
                input(
                    "Add converted score column for combined result (e.g., convert Score/100 to Score/60)? (y/n): "
                )
                .strip()
                .lower()
            )
            if add_conv in ("y", "yes"):
                tgt_raw = input("Enter target maximum (integer), e.g., 60: ").strip()
                try:
                    tgt = float(tgt_raw)
                    new_col = f"Score/{int(tgt)}%"
                    combined_cleaned[new_col] = (
                        ((combined_cleaned["Score/100"].astype(float) / 100.0) * tgt)
                        .round(0)
                        .astype("Int64")
                    )
                    print(f"Added converted column '{new_col}' to combined result.")
                except Exception as e:
                    print(f"Invalid conversion value; skipped converted column: {e}")
        else:
            print("Non-interactive mode detected; skipping converted score column.")

    combined_df = pd.concat(dfs, ignore_index=True)
    combined_df = combined_df.drop_duplicates(subset=["APPLICATION ID"], keep="first")

    out_xlsx = os.path.join(output_dir, f"PUTME_COMBINE_RESULT_{ts}.xlsx")
    combined_cleaned.to_excel(out_xlsx, index=False, engine="openpyxl")
    print(f"Saved processed file: {os.path.basename(out_xlsx)}")

    wb = load_workbook(out_xlsx)
    ws = wb.active
    ws.title = "Results"
    
    # Add document heading
    document_heading = "FCT COLLEGE OF NURSING SCIENCES, GWAGWALADA, FCT-ABUJA\nPOST UTME RESULT"
    format_excel_sheet(
        wb, "Results", combined_cleaned, {}, score_header, pass_threshold, document_heading=document_heading
    )

    full_candidates_df, full_batch_id_map, full_numeric_to_full = (
        load_candidate_batches(candidate_dir)
    )

    # Detect batch mismatches for combined results
    mismatch_rows = []
    present_ids = set()
    invalid_ids = []
    id_to_source_batch = {}
    raw_files = [
        f
        for f in os.listdir(raw_dir)
        if f.lower().endswith((".csv", ".xlsx", ".xls")) and not f.startswith("~$")
    ]
    for df, fname in zip(cleaned_dfs, raw_files):
        batch_id_match = re.search(r"(Batch\s*\d+[A-Za-z]?)", fname, re.IGNORECASE)
        source_batch = batch_id_match.group(1) if batch_id_match else ""
        if source_batch and "APPLICATION ID" in df.columns:
            for app_id in df["APPLICATION ID"].astype(str):
                id_to_source_batch[normalize_id(app_id)] = source_batch

    if "APPLICATION ID" in combined_cleaned.columns:
        for index, v in combined_cleaned["APPLICATION ID"].astype(str).items():
            tok = extract_numeric_token(v)
            norm_id = normalize_id(v)
            registered_batch = None
            full_ex = None
            if norm_id in full_batch_id_map:
                registered_batch = full_batch_id_map[norm_id]
                full_ex = v
            elif tok and tok in full_numeric_to_full:
                full_ex = full_numeric_to_full[tok]
                norm_id = normalize_id(full_ex)
                registered_batch = full_batch_id_map.get(norm_id, "")
            if registered_batch:
                source_batch = id_to_source_batch.get(norm_id, "")
                if source_batch and re.sub(
                    r"\s+", "", registered_batch.lower()
                ) != re.sub(r"\s+", "", source_batch.lower()):
                    mismatch_row = full_candidates_df[
                        full_candidates_df["EXAM_NO"].apply(normalize_id) == norm_id
                    ]
                    if not mismatch_row.empty:
                        mismatch_rows.append(
                            {
                                "APPLICATION ID": v,
                                "FULL NAME": combined_cleaned.at[index, "FULL NAME"],
                                "PHONE NO.": combined_cleaned.at[index, "PHONE NUMBER"],
                                "STATE": combined_cleaned.at[index, "STATE"],
                                "REGISTERED BATCH": registered_batch,
                                "REBATCHED TO": source_batch,
                            }
                        )
            if tok:
                present_ids.add(tok)
            else:
                invalid_ids.append((v, combined_cleaned.at[index, "FULL NAME"]))
        if invalid_ids:
            print(
                f"Found {len(invalid_ids)} invalid APPLICATION ID entries in combined results (non-numeric):"
            )
            for invalid_id, full_name in invalid_ids[:5]:
                print(f"  APPLICATION ID: {invalid_id}, FULL NAME: {full_name}")

    # Create Rebatched sheet for combined results
    combined_mismatch_df = pd.DataFrame(mismatch_rows)
    if "Rebatched" in wb.sheetnames:
        wb.remove(wb["Rebatched"])
    rebatched_ws = wb.create_sheet("Rebatched")
    rebatched_ws.append(
        [
            "APPLICATION ID",
            "FULL NAME",
            "PHONE NO.",
            "STATE",
            "REGISTERED BATCH",
            "REBATCHED TO",
        ]
    )
    if combined_mismatch_df.empty:
        rebatched_ws.append(["No candidates found rebatched.", "", "", "", "", ""])
    else:
        print(
            f"Found {len(combined_mismatch_df)} candidates in combined results who were rebatched:"
        )
        print(
            combined_mismatch_df[
                ["APPLICATION ID", "FULL NAME", "REGISTERED BATCH", "REBATCHED TO"]
            ].to_string(index=False)
        )
        for _, r in combined_mismatch_df.iterrows():
            rebatched_ws.append(
                [
                    r.get("APPLICATION ID"),
                    r.get("FULL NAME"),
                    r.get("PHONE NO."),
                    r.get("STATE"),
                    r.get("REGISTERED BATCH"),
                    r.get("REBATCHED TO"),
                ]
            )
    auto_column_width(rebatched_ws)

    absent_df, _ = create_analysis_and_charts(
        wb,
        combined_cleaned,
        combined_df,
        full_candidates_df,
        full_candidates_df,
        full_batch_id_map,
        full_numeric_to_full,
        score_header,
        output_dir,
        ts,
        fname="Combined",
        pass_threshold=pass_threshold,
        skip_batch_mismatches=True,
        rebatched_count=len(combined_mismatch_df),
    )

    wb.save(out_xlsx)
    print(
        f"Saved processed file with Analysis, Charts, Absent, and Rebatched: {os.path.basename(out_xlsx)}"
    )


# ---------------------------
# Entrypoint
# ---------------------------
def main():
    args = parse_args()
    print("Starting UTME Results Cleaning...")

    # Use command-line arguments or defaults
    RAW_DIR = args.input_dir
    CANDIDATE_DIR = args.candidate_dir
    CLEAN_DIR = args.output_dir
    PASS_THRESHOLD = args.pass_threshold
    BATCH_ID = args.batch_id
    NON_INTERACTIVE = args.non_interactive
    CONVERTED_SCORE_MAX = args.converted_score_max

    # Ensure directories exist
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(CLEAN_DIR, exist_ok=True)
    os.makedirs(CANDIDATE_DIR, exist_ok=True)

    # Filter files based on batch_id if provided
    files = [
        f for f in os.listdir(RAW_DIR) if f.lower().endswith((".csv", ".xlsx", ".xls"))
    ]
    if BATCH_ID:
        batch_id_normalized = re.sub(r"\s+", "", BATCH_ID.lower())
        files = [
            f for f in files if batch_id_normalized in re.sub(r"\s+", "", f.lower())
        ]
        if not files:
            print(f"No raw files found matching batch ID '{BATCH_ID}' in {RAW_DIR}")
            return

    if not files:
        print(
            f"No raw files found in {RAW_DIR}\nPut your raw file(s) there and re-run."
        )
        return

    ts = datetime.now().strftime(TIMESTAMP_FMT)
    output_dir = os.path.join(CLEAN_DIR, f"UTME_RESULT-{ts}")
    os.makedirs(output_dir, exist_ok=True)

    # Load full candidates_df, batch_id_map, numeric_to_full from all candidates
    full_candidates_df, full_batch_id_map, full_numeric_to_full = (
        load_candidate_batches(CANDIDATE_DIR)
    )

    outputs = []
    cleaned_dfs = []
    dfs = []
    absent_dfs = []
    mismatch_dfs = []

    for f in files:
        try:
            cleaned, df, absent_df, mismatch_df = process_file(
                os.path.join(RAW_DIR, f),
                output_dir,
                ts,
                CANDIDATE_DIR,
                full_candidates_df,
                full_batch_id_map,
                full_numeric_to_full,
                PASS_THRESHOLD,
            )
            if cleaned is not None:
                outputs.append(
                    (
                        os.path.join(
                            output_dir, f"UTME_RESULT_{os.path.splitext(f)[0]}_{ts}.csv"
                        ),
                        os.path.join(
                            output_dir,
                            f"UTME_RESULT_{os.path.splitext(f)[0]}_{ts}.xlsx",
                        ),
                    )
                )
                cleaned_dfs.append(cleaned)
                dfs.append(df)
                absent_dfs.append(absent_df)
                mismatch_dfs.append(mismatch_df)
        except Exception as e:
            print(f"Error processing {f}: {e}", file=sys.stderr)

    combine_batches(
        cleaned_dfs,
        dfs,
        absent_dfs,
        mismatch_dfs,
        output_dir,
        ts,
        PASS_THRESHOLD,
        CANDIDATE_DIR,
        RAW_DIR,
        NON_INTERACTIVE,
        CONVERTED_SCORE_MAX,
    )

    print("\nProcessing completed successfully.")
    if outputs:
        for csvp, xl in outputs:
            print(f" - Saved processed file: {os.path.basename(csvp)}")
            print(f" - Saved processed file: {os.path.basename(xl)}")
    print(
        f" - Saved processed file: {os.path.basename(os.path.join(output_dir, f'PUTME_COMBINE_RESULT_{ts}.xlsx'))}"
    )
    print(
        f" - Saved processed file: {os.path.basename(os.path.join(output_dir, f'PUTME_COMBINE_RESULT_UNSORTED_{ts}.xlsx'))}"
    )


if __name__ == "__main__":
    main()